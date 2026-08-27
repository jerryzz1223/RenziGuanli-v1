import hashlib
import json
import re
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import cint, flt, getdate, now_datetime

from hrms.payroll.payroll_formula import (
	FIELD_BY_NAME,
	FIELD_DEFINITIONS,
	FORMULA_TEMPLATES,
	FUNCTIONS,
	FormulaError,
	compile_formula,
	evaluate_formula,
	evaluate_formula_set,
)


PAYROLL_VARIABLE_SHEETS = [
	"奖惩提报单（提交财务）",
	"证书、多能工津贴名单",
	"全勤奖",
	"学历补贴",
	"宿舍费",
	"社保名单",
	"继续服务奖",
	"提案改善表",
	"苹果树",
	"苹果树（修改后）",
	"离职人员薪资结算",
	"每月员工住宿费用明细表",
	"人员住宿登记表",
]
VARIABLE_BATCH_DOCTYPE = "HRMS Payroll Variable Import Batch"
VARIABLE_RECORD_DOCTYPE = "HRMS Payroll Variable Record"
VARIABLE_SOURCE_TYPE_DOCTYPE = "HRMS Payroll Variable Source Type"
PAYROLL_INPUT_DOCTYPE = "HRMS Payroll Input Record"
PAYROLL_SETTLEMENT_DOCTYPE = "HRMS Payroll Settlement Record"
LOCAL_PAYROLL_TEST_COMPANY = "TEST-HRMS"
SALARY_STRUCTURE_VERSION_DOCTYPE = "HRMS Salary Structure Version"
SALARY_GRADE_DOCTYPE = "HRMS Salary Grade"
EMPLOYEE_SALARY_CHANGE_DOCTYPE = "HRMS Employee Salary Change"
MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE = "HRMS Monthly Payroll Participation"
FORM_IMPORT_BATCH_DOCTYPE = "HRMS Form Import Batch"
WELFARE_SOURCE_DOCTYPE = "HRMS Payroll Welfare Source Record"
PAYROLL_RULE_DOCTYPE = "HRMS Payroll Rule"
PAYROLL_FIELD_MAPPING_DOCTYPE = "HRMS Payroll Field Mapping"
MONTHLY_ATTENDANCE_DOCTYPE = "HRMS Monthly Attendance Summary"
PAYROLL_STEP_LOCK_DOCTYPE = "HRMS Payroll Step Lock"
PAYROLL_STANDARD_HOURS_DIVISOR = 174
WELFARE_SOURCE_SYNC_SHEET = "福利扣款来源中心"
PAYROLL_SETTLEMENT_IMPORT_SHEET = "薪资结算表"
PAYROLL_SETTLEMENT_IMPORT_SOURCE = "完整薪资结算表导入"
PROCESSING_ATTENDANCE_LOCK_PREFIX = "处理终稿:"
MONTHLY_VARIABLE_SCOPE_PREFIX = "月度增减项:"
SOCIAL_INSURANCE_MANUAL_EXCLUSION_STATUSES = {"不参保（已确认）", "停缴"}
SOCIAL_INSURANCE_VARIABLE_TYPES = ("社保个人", "社保公司")
SINGLETON_MONTHLY_VARIABLE_TYPES = {
	"住房补贴", "学历补贴", "宿舍扣款", "水电费及扣款",
	"社保个人", "社保公司", "公积金个人", "公积金公司",
}
PAYROLL_PARTICIPATION_DECISIONS = {"正常计薪", "离职结算", "不参与计算", "异常待审核"}
PAYROLL_PARTICIPATION_APPROVED_STATUS = "审核通过"
# ``待审核`` is retained for batches created before the simplified entry flow
# was released.  New data only has one human action left: confirm it for payroll.
PENDING_VARIABLE_BATCH_STATUSES = {"待确认", "待审核"}
SIGNABLE_PAYROLL_SOURCE_CODES = {
	"certificate_skill",
	"continuing_service",
	"dormitory",
	"reward_punishment",
	"education",
	"social_insurance",
	"housing_fund",
}
TEST_MONTHLY_RESET_AREAS = {
	"attendance": "考勤派生数据",
	"payroll": "薪酬派生数据",
}

DEFAULT_PAYROLL_VARIABLE_SOURCE_TYPES = [
	{"source_code": "attendance_final", "source_name": "考勤终稿 / 全勤奖 / 住房补贴", "purpose": "锁定后的出勤、工时、全勤奖与住房补贴计算基础", "required_fields": "由考勤假期模块锁定终稿自动提供", "template_notes": "本页不上传；全勤奖和住房补贴均在考勤补充来源中导入、校验并随考勤终稿锁定后继承。", "target_area": "考勤继承"},
	{"source_code": "salary_change", "source_name": "员工定薪", "purpose": "底薪、职能/职务津贴与总薪资", "required_for_payroll": 1, "required_fields": "工号、姓名、生效日期、调整后薪资", "template_notes": "在员工定薪区域保存并提交后立即生效并参与试算。", "target_area": "员工定薪"},
	{"source_code": "certificate_skill", "source_name": "证书/多能工津贴", "purpose": "证书奖励与多能工奖励", "required_fields": "工号或姓名、证书津贴/多能工津贴", "template_notes": "一人可生成多条月度增项。"},
	{"source_code": "proposal", "source_name": "提案改善", "purpose": "改善提案奖金", "required_fields": "奖励人、奖金金额", "template_notes": "多个奖励人时必须在预览后分配到每人。"},
	{"source_code": "continuing_service", "source_name": "继续服务奖", "purpose": "按制度发放继续服务奖", "required_fields": "工号或姓名、金额", "template_notes": "支持原继续服务奖工作表。"},
	{"source_code": "dormitory", "source_name": "宿舍水电", "purpose": "宿舍费、水费、电费与当月扣款", "required_fields": "工号或姓名、实收/当月扣款", "template_notes": "金额作为扣项进入当月计算。"},
	{"source_code": "reward_punishment", "source_name": "奖惩", "purpose": "奖励或惩处金额", "required_fields": "受奖惩人、金额、奖惩性质", "template_notes": "审核时确认奖励为增项、惩处为减项。"},
	{"source_code": "education", "source_name": "学历补贴", "purpose": "按学历资格发放月度补贴", "required_fields": "工号或姓名、学历类别、补贴金额", "template_notes": "需审核学历资格与金额。"},
	{"source_code": "social_insurance", "source_name": "社保", "purpose": "个人与公司社保承担", "required_for_payroll": 1, "required_fields": "工号或姓名、个人承担、公司承担", "template_notes": "未参保员工应明确标记停缴/不参保。"},
	{"source_code": "housing_fund", "source_name": "公积金", "purpose": "个人与公司公积金承担", "required_fields": "工号或姓名、个人承担、公司承担", "template_notes": "未缴纳员工应明确保留零申报说明。"},
]

PAYROLL_WORKFLOW_STEPS = [
	("master", "人员基础"),
	("salary", "员工定薪"),
	("rules", "核算规则"),
	("attendance", "考勤计薪规则"),
	("sources", "月度数据封板"),
	("calculation", "试算复核"),
	("delivery", "报表发放"),
]
PAYROLL_WORKFLOW_STEP_LABELS = dict(PAYROLL_WORKFLOW_STEPS)
PAYROLL_ATTENDANCE_RULE_CODES = [
	"ATTENDANCE_FULL_ATTENDANCE_BONUS",
	"PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION",
	"ATTENDANCE_MISSED_PUNCH",
	"PAYROLL_SETTLEMENT_OVERTIME_PAY",
	"PAYROLL_SETTLEMENT_NIGHT_SHIFT",
]

PAYROLL_IMPORT_TEMPLATES = [
	{
		"template_key": "employee_salary_change",
		"sheet_name": "员工薪资异动导入",
		"target_doctype": EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		"description": "导入员工底薪、职能津贴、证书津贴、多能工津贴和薪资小计，是薪资结算表 E:H 列的主来源。",
		"required_columns": ["工号", "姓名", "生效日期"],
		"columns": [
			("薪资月份", "payroll_month", "可选；未填时使用页面月份"),
			("工号", "employee_code", "必填；用于匹配员工花名册"),
			("姓名", "employee_name", "必填；工号无法匹配时用姓名辅助匹配"),
			("部门", "department", "可选；未填时取员工花名册部门"),
			("岗位", "designation", "可选；未填时取员工花名册岗位"),
			("生效日期", "effective_date", "必填；格式 yyyy-mm-dd"),
			("薪资档位", "salary_grade", "可选；已维护薪资档位时填写"),
			("底薪", "base_salary", "对应薪资结算表 E 列"),
			("职能津贴", "function_allowance", "对应薪资结算表 F 列"),
			("证书津贴", "certificate_allowance", "证书津贴"),
			("多能工津贴", "multi_skill_allowance", "多能工津贴"),
			("薪资小计", "full_salary", "对应薪资结算表 H 列；未填时自动 E+F+证书+多能工"),
			("社保", "social_insurance_enabled", "1/0 或 是/否"),
			("公积金", "housing_fund_enabled", "1/0 或 是/否"),
			("备注", "remarks", "可选"),
		],
	},
	{
		"template_key": "welfare_source",
		"sheet_name": "福利扣款来源导入",
		"target_doctype": WELFARE_SOURCE_DOCTYPE,
		"description": "导入学历补贴、租房补贴、宿舍费、水电费、社保公积金、提案改善奖、继续服务奖、所得税等月度变量。",
		"required_columns": ["来源类型", "姓名", "金额"],
		"columns": [
			("薪资月份", "payroll_month", "可选；未填时使用页面月份"),
			("来源类型", "source_type", "如学历补贴、租房补贴、宿舍住宿费、宿舍水电费、社保个人、公积金个人、社保公司、公积金公司、提案改善奖、继续服务奖、所得税、年终奖所得税、水电费及扣款、已发福利、生产奖"),
			("工号", "employee_code", "建议填写"),
			("姓名", "employee_name", "必填"),
			("部门", "department", "可选"),
			("金额", "amount", "必填"),
			("资格状态", "eligibility_status", "默认符合"),
			("确认状态", "confirmation_status", "默认已确认"),
			("来源单据/说明", "source_reference", "如宿舍明细、社保名单、财务确认表"),
			("备注", "remarks", "可选"),
		],
	},
	{
		"template_key": "monthly_attendance_summary",
		"sheet_name": "月度考勤终稿导入",
		"target_doctype": MONTHLY_ATTENDANCE_DOCTYPE,
		"description": "导入考勤终稿，作为标准工时、基本出勤、加班、旷工、夜班和苹果树金额来源。",
		"required_columns": ["工号", "姓名", "标准工时", "基本出勤工时"],
		"columns": [
			("考勤月份", "attendance_month", "可选；未填时使用页面月份"),
			("工号", "employee_code", "必填；用于匹配员工花名册"),
			("姓名", "employee_name", "必填；工号无法匹配时用姓名辅助匹配"),
			("部门", "department", "可选"),
			("入职日期", "date_of_joining", "可选"),
			("标准工时", "standard_hours", "对应薪资结算表 I 列"),
			("基本出勤工时", "actual_attendance_hours", "对应薪资结算表 J 列"),
			("调整后工时", "adjusted_working_hours", "考勤终稿调整后工时"),
			("1.5倍加班", "overtime_1_5_hours", "对应平日加班时数"),
			("2倍加班", "overtime_2_hours", "对应调整前周末加班"),
			("3倍加班", "overtime_3_hours", "对应节假日加班"),
			("请假工时", "leave_hours", "月度请假合计"),
			("旷工工时", "absent_hours", "对应薪资结算表 AE 列"),
			("深夜班次数", "deep_night_shift_count", "按锁定考勤明细的深夜班时段自动匹配"),
			("大夜班次数", "large_night_shift_count", "扣除深夜班后的大夜班次数，对应薪资结算表 V 列"),
			("小夜班次数", "small_night_shift_count", "对应薪资结算表 W 列"),
			("红绿苹果金额", "apple_reward_amount", "对应薪资结算表 AA 列"),
			("全勤扣款", "full_attendance_deduction", "全勤奖扣款来源"),
			("状态", "status", "默认已确认"),
		],
	},
]

SHEET_VARIABLE_TYPES = {
	"奖惩提报单（提交财务）": "其他奖金",
	"证书、多能工津贴名单": "证书及多能工津贴",
	"全勤奖": "全勤奖",
	"学历补贴": "学历补贴",
	"宿舍费": "宿舍扣款",
	"社保名单": "社保个人",
	"继续服务奖": "继续服务奖",
	"提案改善表": "提案改善奖",
	"苹果树": "苹果树",
	"苹果树（修改后）": "苹果树",
	"离职人员薪资结算": "其他扣款",
	"每月员工住宿费用明细表": "宿舍扣款",
	"人员住宿登记表": "宿舍扣款",
}

VARIABLE_FIELD_MAP = {
	"全勤奖": "full_attendance_bonus",
	"住房补贴": "housing_subsidy",
	"学历补贴": "education_subsidy",
	"宿舍扣款": "dormitory_deduction",
	"社保个人": "social_security_personal",
	"公积金个人": "housing_fund_personal",
	"其他奖金": "other_bonus",
	"其他扣款": "other_deduction",
	"底薪": "base_salary",
	"职能津贴": "function_allowance",
	"职务津贴": "function_allowance",
	"证书津贴": "certificate_skill_allowance",
	"多能工津贴": "certificate_skill_allowance",
	"证书及多能工津贴": "certificate_skill_allowance",
	"全薪": "salary_subtotal",
	"薪资小计": "salary_subtotal",
	"生产奖": "production_bonus",
	"提案改善奖": "proposal_improvement_bonus",
	"继续服务奖": "continuing_service_bonus",
	"所得税": "income_tax",
	"年终奖所得税": "year_end_bonus_tax",
	"水电费及扣款": "utilities_deduction",
	"社保公司": "social_security_company",
	"公积金公司": "housing_fund_company",
	"已发福利": "paid_proposal_birthday_welfare",
	"夜班津贴": "night_shift_allowance",
	"迟到金额+全勤奖扣款": "late_full_attendance_deduction",
	"苹果树": "apple_reward_amount",
	"奖惩提报": "other_bonus",
	"离职薪资结算": "other_deduction",
}

WELFARE_SOURCE_VARIABLE_TYPE_MAP = {
	"薪资构成": "薪资小计",
	"学历补贴": "学历补贴",
	"租房补贴": "住房补贴",
	"宿舍住宿费": "宿舍扣款",
	"宿舍水电费": "水电费及扣款",
	"社保个人": "社保个人",
	"社保公司": "社保公司",
	"公积金个人": "公积金个人",
	"公积金公司": "公积金公司",
	"提案改善奖": "提案改善奖",
	"继续服务奖": "继续服务奖",
	"所得税": "所得税",
	"年终奖所得税": "年终奖所得税",
	"水电费及扣款": "水电费及扣款",
	"已发福利": "已发福利",
	"生产奖": "生产奖",
	"奖惩提报": "其他奖金",
	"证书多能工津贴": "证书及多能工津贴",
	"苹果树": "其他奖金",
	"离职薪资结算": "其他扣款",
	"高温补贴": "其他奖金",
	"手机话费补贴": "其他奖金",
	"油费补贴": "其他奖金",
	"其他奖金": "其他奖金",
	"其他扣款": "其他扣款",
}

WELFARE_SOURCE_RULES = [
	{
		"source_type": "薪资构成",
		"title": "薪资构成异动来源",
		"direction": "参考",
		"variable_type": "薪资小计",
		"rule": "异动前后岗位、职级、底薪、职能/职务津贴、证书津贴、多能工津贴、全薪只作为薪资主数据来源；结算以同公司已批准且生效的员工薪资异动为准。",
	},
	{
		"source_type": "证书多能工津贴",
		"title": "证书多能工津贴确认",
		"direction": "应发",
		"variable_type": "证书及多能工津贴",
		"rule": "证书津贴、多能工津贴需保留来源单据与确认状态；进入薪资主数据或同版本变量后参与试算。",
	},
	{
		"source_type": "奖惩提报",
		"title": "奖惩提报",
		"direction": "应发",
		"variable_type": "其他奖金",
		"rule": "奖惩提报默认作为奖励导入；如为扣款，可在导入表或来源记录中选择“应扣”。",
	},
	{
		"source_type": "苹果树",
		"title": "绩效与苹果树变量",
		"direction": "应发",
		"variable_type": "苹果树",
		"rule": "只读取绩效与试用报表线程输出的已确认苹果树变量，必须与薪资公司、月份、考勤锁定版本一致。",
	},
	{
		"source_type": "离职薪资结算",
		"title": "离职薪资结算",
		"direction": "应扣",
		"variable_type": "其他扣款",
		"rule": "离职薪资结算默认作为扣款导入；如为补发，可在导入表或来源记录中选择“应发”。",
	},
	{
		"source_type": "学历补贴",
		"title": "学历补贴资格与月报",
		"direction": "应发",
		"variable_type": "学历补贴",
		"rule": "非全日制大专100元；全日制大专、非全日制本科200元；全日制本科及以上300元；入职当月开始提交月报，共24个月。",
	},
	{
		"source_type": "租房补贴",
		"title": "租房补贴申请/登记/月度明细",
		"direction": "应发",
		"variable_type": "住房补贴",
		"rule": "外地户籍在苏州租房人员享受；苏州户籍或苏州买房不享受；入职10号前200元，11-20号100元，21号及以后0元；离职当月满勤200元，未做满0元。",
	},
	{
		"source_type": "宿舍住宿费",
		"title": "宿舍入住/退宿/住宿费",
		"direction": "应扣",
		"variable_type": "宿舍扣款",
		"rule": "干部宿舍300元/月，阁楼干部宿舍200元/月，线长单身宿舍400元/月，集体宿舍400元/月；每月10号前确认给财务。",
	},
	{
		"source_type": "宿舍水电费",
		"title": "宿舍入住/退宿/水电住宿费",
		"direction": "应扣",
		"variable_type": "水电费及扣款",
		"rule": "宿舍水费4.25元/吨，电费0.9元/度；按每月员工住宿费用明细表确认后进入薪资扣款。",
	},
	{
		"source_type": "社保个人",
		"title": "社保公积金个人/公司承担",
		"direction": "应扣",
		"variable_type": "社保个人",
		"rule": "社保对象以社保名单为准；员工档案明确为不参保、停缴或起缴日期未到时，当月社保个人与公司承担不计入薪资。",
	},
	{
		"source_type": "公积金个人",
		"title": "社保公积金个人/公司承担",
		"direction": "应扣",
		"variable_type": "公积金个人",
		"rule": "住房公积金对象为转正后员工；个人承担额来自公积金名单或财务确认表。",
	},
	{
		"source_type": "社保公司",
		"title": "社保公积金个人/公司承担",
		"direction": "公司承担",
		"variable_type": "社保公司",
		"rule": "公司承担额可手工确认；未提供时薪资结算按薪资结算表区间公式由个人社保反推。",
	},
	{
		"source_type": "公积金公司",
		"title": "社保公积金个人/公司承担",
		"direction": "公司承担",
		"variable_type": "公积金公司",
		"rule": "公司公积金承担额可手工确认；未提供时薪资结算默认等于个人公积金。",
	},
	{
		"source_type": "提案改善奖",
		"title": "提案改善奖、继续服务奖、所得税、水电扣款等月度变量",
		"direction": "应发",
		"variable_type": "提案改善奖",
		"rule": "按月度确认金额录入，进入奖金小计。",
	},
	{
		"source_type": "继续服务奖",
		"title": "提案改善奖、继续服务奖、所得税、水电扣款等月度变量",
		"direction": "应发",
		"variable_type": "继续服务奖",
		"rule": "按年度公告和缺勤区间确认金额；进入实发工资与公司实际负担。",
	},
	{
		"source_type": "所得税",
		"title": "提案改善奖、继续服务奖、所得税、水电扣款等月度变量",
		"direction": "应扣",
		"variable_type": "所得税",
		"rule": "按国家法令代扣代缴；可由财务确认后按月导入。",
	},
	{
		"source_type": "水电费及扣款",
		"title": "提案改善奖、继续服务奖、所得税、水电扣款等月度变量",
		"direction": "应扣",
		"variable_type": "水电费及扣款",
		"rule": "除宿舍水电外的月度水电及其他已确认扣款。",
	},
]

DEFAULT_PAYROLL_RULES = [
	{
		"rule_code": "SALARY_STRUCTURE_SUBTOTAL",
		"rule_name": "薪资小计",
		"rule_category": "薪资架构",
		"rule_scope": "所有员工",
		"formula_expression": "薪资小计 = 底薪 + 职能津贴",
		"parameters_json": {"base_salary": "底薪", "function_allowance": "职能津贴"},
		"rule_text": "员工薪资以《人事组薪资异动表》为准。证书津贴、多能工津贴自 2026 年 5 月起不参与全薪构成，作为月度补贴进入奖金小计。",
		"source_file": "5.1薪资福利.xlsx",
		"source_sheet": "薪资架构 / 人事组薪资异动表",
	},
	{
		"rule_code": "ATTENDANCE_FULL_ATTENDANCE_BONUS",
		"rule_name": "全勤奖",
		"rule_category": "考勤",
		"rule_scope": "月度考勤终稿",
		"formula_expression": "缺勤=标准工时-钉钉导出实际出勤-排休+病假/2；0=200，0~0.5=190，0.5~16=150，16~32=100，32~48=50，>48=0",
		"parameters_json": {"base_bonus": 200, "late_deduction": 10, "thresholds": [[0, 200], [0.5, 190], [16, 150], [32, 100], [48, 50]]},
		"rule_text": "特休、公休、公司原因调整上班时间不扣全勤；其他假别计入请假工时。新进与离职员工当月未上班时间计入缺勤。",
		"source_file": "5.1薪资福利.xlsx / 5.2人资考勤.xlsx",
		"source_sheet": "薪资福利作业规范 / 1.12考勤终稿",
		"source_cell": "薪资福利作业规范 rows 33-43；1.12考勤终稿 BF列",
	},
	{
		"rule_code": "ATTENDANCE_OVERTIME_HOURS",
		"rule_name": "考勤终稿工时换算",
		"rule_category": "考勤",
		"rule_scope": "月度考勤终稿",
		"formula_expression": "1倍结算工时、1.5倍结算工时、2倍结算工时、3倍节假日加班工时按1.12考勤终稿 AH:AV 列公式生成",
		"parameters_json": {"workday_multiplier": 1.5, "weekend_multiplier": 2, "holiday_multiplier": 3},
		"rule_text": "请假、病假补/扣工时、特休、工伤、丧假、婚假、排休等先在考勤终稿中换算，再进入薪资结算表。",
		"source_file": "5.2人资考勤.xlsx",
		"source_sheet": "1.12考勤终稿",
		"source_cell": "AH:AV",
	},
	{
		"rule_code": "PERFORMANCE_APPLE_REWARD",
		"rule_name": "苹果树奖惩金额",
		"rule_category": "奖金福利",
		"rule_scope": "审批通过并纳入考勤锁定版本的苹果树记录",
		"formula_expression": "苹果树金额 = (绿苹果颗数 - 红苹果颗数) * 5",
		"parameters_json": {"amount_per_apple": 5, "locked_attendance_required": True},
		"rule_text": "部门主管审批通过后，月初随考勤给员工签字确认；薪资不读取草稿或未锁定的苹果树记录。",
		"source_file": "4.2苹果树.xlsx / 5.2人资考勤.xlsx",
		"source_sheet": "苹果树管理办法 / 1.12考勤终稿",
		"source_cell": "苹果树管理办法 row 6；考勤终稿苹果树金额列",
	},
	{
		"rule_code": "PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION",
		"rule_name": "缺勤扣除金额",
		"rule_category": "薪资结算",
		"rule_scope": "薪资结算表",
		"formula_expression": "缺勤扣除金额 = ROUND(薪资小计 / 标准计薪工时 * 调整后缺勤工时, 2)；旷工扣款 = ROUND(薪资小计 / 标准计薪工时 * 旷工工时 * 旷工倍率, 2)",
		"parameters_json": {"standard_hours_divisor": 174, "absenteeism_multiplier": 3},
		"rule_text": "调整后缺勤工时 = IF(缺勤工时 - 调整前周末加班 > 0, 缺勤工时 - 调整前周末加班, 0)。",
		"source_file": "5.2人资考勤.xlsx",
		"source_sheet": "薪资结算表",
		"source_cell": "M:N",
	},
	{
		"rule_code": "ATTENDANCE_MISSED_PUNCH",
		"rule_name": "忘打卡红苹果",
		"rule_category": "考勤",
		"rule_scope": "忘打卡来源；仅影响后续重新加工的批次",
		"formula_expression": "每条纳入的忘打卡记录 = 红苹果颗数/次；金额 = 红苹果颗数/次 × 每颗金额",
		"parameters_json": {"red_apples_per_record": 2, "amount_per_apple": 5},
		"rule_text": "仅由“忘打卡”来源汇总，不从原始考勤缺卡标记重复生成红苹果；审批未通过、撤销及不计入记录不参与。",
		"source_file": "忘打卡.xlsx",
		"source_sheet": "忘打卡合计",
	},
	{
		"rule_code": "PAYROLL_SETTLEMENT_OVERTIME_PAY",
		"rule_name": "加班费",
		"rule_category": "薪资结算",
		"rule_scope": "薪资结算表",
		"formula_expression": "平日=ROUND(底薪/174*平日加班*1.5,2)；周末=ROUND(底薪/174*调整后周末加班*2,2)；节假日=ROUND(底薪/174*节假日加班*3,2)",
		"parameters_json": {"standard_hours_divisor": 174, "weekday": 1.5, "weekend": 2, "holiday": 3},
		"rule_text": "加班费小计为平日、周末、节假日加班费合计。",
		"source_file": "5.2人资考勤.xlsx",
		"source_sheet": "薪资结算表",
		"source_cell": "R:U",
	},
	{
		"rule_code": "PAYROLL_SETTLEMENT_NIGHT_SHIFT",
		"rule_name": "夜班津贴",
		"rule_category": "薪资结算",
		"rule_scope": "薪资结算表",
		"formula_expression": "夜班津贴 = 深夜班次数 * 55 + 大夜班次数 * 45 + 小夜班次数 * 24",
		"parameters_json": {"deep_night_shift": 55, "large_night_shift": 45, "small_night_shift": 24, "deep_night_shift_start": "20:00", "deep_night_shift_end": "08:00", "large_night_shift_start": "", "large_night_shift_end": "", "small_night_shift_start": "", "small_night_shift_end": ""},
		"rule_text": "深夜班、大夜班、小夜班均按各自设置的上下班时间匹配锁定考勤明细；同一条完整打卡记录只匹配一个档位。",
		"source_file": "5.2人资考勤.xlsx",
		"source_sheet": "薪资结算表",
		"source_cell": "V:X",
	},
	{
		"rule_code": "PAYROLL_SETTLEMENT_GROSS_PAY",
		"rule_name": "应付工资",
		"rule_category": "薪资结算",
		"rule_scope": "薪资结算表",
		"formula_expression": "应付工资 = 薪资小计 - 缺勤扣除金额 + 加班费小计 + 夜班津贴 + 奖金小计 - 惩处小计",
		"parameters_json": {"bonus_total": "提案改善奖+红绿苹果+全勤奖住房学历补贴+生产奖", "punishment_total": "旷工扣款+迟到金额+全勤奖扣款"},
		"rule_text": "奖金小计和惩处小计必须保留来源记录。",
		"source_file": "5.2人资考勤.xlsx",
		"source_sheet": "薪资结算表",
		"source_cell": "AD:AI",
	},
	{
		"rule_code": "PAYROLL_SETTLEMENT_NET_PAY",
		"rule_name": "实发工资",
		"rule_category": "薪资结算",
		"rule_scope": "薪资结算表",
		"formula_expression": "计税工资 = 应付工资 - 社保个人 - 公积金个人 + 已发福利；实发工资 = 计税工资 - 所得税 - 年终奖所得税 - 水电费及扣款 + 继续服务奖 - 已发福利",
		"parameters_json": {"paid_welfare": "提案改善奖&生日福利金（已发）"},
		"rule_text": "所得税和年终奖所得税由财务/个税规则确认后进入变量。",
		"source_file": "5.2人资考勤.xlsx",
		"source_sheet": "薪资结算表",
		"source_cell": "AM:AS",
	},
	{
		"rule_code": "WELFARE_EDUCATION_SUBSIDY",
		"rule_name": "学历补贴",
		"rule_category": "福利补贴",
		"rule_scope": "符合学历补贴人员",
		"formula_expression": "非全日制大专=100；全日制大专/非全日制本科=200；全日制本科及以上=300；最多24个月",
		"parameters_json": {"non_full_time_college": 100, "full_time_college_or_non_full_time_bachelor": 200, "full_time_bachelor_or_above": 300, "months": 24},
		"rule_text": "入职当月即可享受并提交月报；离职当月享受但离职前需提交报告。",
		"source_file": "5.1薪资福利.xlsx",
		"source_sheet": "学历补贴工作月报管理办法",
	},
	{
		"rule_code": "WELFARE_RENTAL_SUBSIDY",
		"rule_name": "租房补贴",
		"rule_category": "福利补贴",
		"rule_scope": "外地户籍在苏州租房人员",
		"formula_expression": "入职当月10号前=200；11-20号=100；21号及以后=0；离职当月满勤=200，未做满=0",
		"parameters_json": {"before_or_on_day_10": 200, "day_11_to_20": 100, "after_or_on_day_21": 0, "resignation_full_attendance": 200},
		"rule_text": "苏州户籍以及在苏州买房者不享受；住公司宿舍人员按宿舍规则处理。",
		"source_file": "5.5租房补贴.xlsx",
		"source_sheet": "租房补贴作业规范",
		"source_cell": "rows 17-31",
	},
	{
		"rule_code": "WELFARE_DORMITORY_FEE",
		"rule_name": "宿舍住宿费",
		"rule_category": "宿舍",
		"rule_scope": "住宿员工",
		"formula_expression": "干部宿舍=300/月；阁楼干部宿舍=200/月；线长单身宿舍=400/月；集体宿舍=400/月；水费=4.25/吨；电费=0.9/度",
		"parameters_json": {"manager_dorm": 300, "attic_manager_dorm": 200, "line_leader_single_dorm": 400, "group_dorm": 400, "water_per_ton": 4.25, "electricity_per_kwh": 0.9},
		"rule_text": "每月10号前住宿员工签字确认后给财务结算薪资用。",
		"source_file": "5.6员工宿舍.xlsx",
		"source_sheet": "员工宿舍作业规范",
		"source_cell": "rows 90-100",
	},
	{
		"rule_code": "WELFARE_SOCIAL_SECURITY_COMPANY",
		"rule_name": "公司社保承担",
		"rule_category": "社保公积金",
		"rule_scope": "转正后员工",
		"formula_expression": "若社保公司未手工确认，则按薪资结算表区间由社保个人反推：<524.96=0；=524.96=1256.82；520~531=1269；531~636=1522.8；>636=1649.7",
		"parameters_json": {"manual_override": True, "ranges": [[0, 524.96, 0], [524.96, 524.96, 1256.82], [520, 531, 1269], [531, 636, 1522.8], [636, None, 1649.7]]},
		"rule_text": "社保对象以社保名单为准；员工档案明确为不参保、停缴或起缴日期未到时，当月不计入社保。增减员、基数、比例、上下限仍需以社保名单或财务确认表为准。",
		"source_file": "5.1薪资福利.xlsx / 5.2人资考勤.xlsx",
		"source_sheet": "薪资福利作业规范 / 薪资结算表",
		"missing_rule_note": "社保个人基数、比例、增减员当月规则未在当前表中完整结构化。",
	},
	{
		"rule_code": "WELFARE_HOUSING_FUND_COMPANY",
		"rule_name": "公司公积金承担",
		"rule_category": "社保公积金",
		"rule_scope": "转正后员工",
		"formula_expression": "公司公积金 = 公积金个人，除非福利扣款来源中心手工确认公司承担额",
		"parameters_json": {"default_company_equals_personal": True, "manual_override": True},
		"rule_text": "住房公积金对象为转正后员工。",
		"source_file": "5.1薪资福利.xlsx / 5.2人资考勤.xlsx",
		"source_sheet": "薪资福利作业规范 / 薪资结算表",
		"missing_rule_note": "公积金基数、比例、增减员当月规则未在当前表中完整结构化。",
	},
	{
		"rule_code": "TAX_INCOME_TAX",
		"rule_name": "所得税代扣",
		"rule_category": "税费扣款",
		"rule_scope": "所有应税员工",
		"formula_expression": "所得税、年终奖所得税目前作为财务确认月度变量导入",
		"parameters_json": {"manual_import": True},
		"rule_text": "凡薪资、奖金核发，一律依照国家法令代扣缴个人所得税。",
		"source_file": "5.1薪资福利.xlsx",
		"source_sheet": "薪资福利作业规范",
		"source_cell": "row 27",
		"missing_rule_note": "专项扣除、累计预扣、年终奖税等完整个税计算参数未在公司资料中结构化提供。",
	},
	{
		"rule_code": "WELFARE_CONTINUING_SERVICE_BONUS",
		"rule_name": "继续服务奖",
		"rule_category": "奖金福利",
		"rule_scope": "符合年度公告且仍在职的员工",
		"formula_expression": "已确认继续服务奖金额作为月度变量进入实发工资和公司实际负担",
		"parameters_json": {"confirmed_source_required": True},
		"rule_text": "缺勤区间和年度公告决定是否享受；系统只读取同公司、同月份、同考勤锁定版本的已确认来源记录。",
		"source_file": "5.1薪资福利.xlsx",
		"source_sheet": "薪资福利作业规范",
		"missing_rule_note": "各年度基准金额、发放月份及特殊假别是否计入缺勤尚未形成稳定参数表，暂不自动计算。",
	},
	{
		"rule_code": "WELFARE_HIGH_TEMPERATURE_SUBSIDY",
		"rule_name": "高温补贴",
		"rule_category": "福利补贴",
		"rule_scope": "符合当年公告和出勤条件的员工",
		"formula_expression": "当月高温补贴 = 当年有效公告单价或金额 × 已确认资格/天数",
		"parameters_json": {"confirmed_source_required": True},
		"rule_text": "离职人员不享受；实际发放月份、岗位范围和标准以当年公司公告及确认明细为准。",
		"source_file": "5.1薪资福利.xlsx",
		"source_sheet": "薪资福利作业规范",
		"missing_rule_note": "当前资料未提供长期有效的年度单价、月份和岗位资格参数，必须按年度来源记录导入。",
	},
	{
		"rule_code": "PAYROLL_PART_MONTH_PRORATION",
		"rule_name": "入离职及未满月工资折算",
		"rule_category": "薪资结算",
		"rule_scope": "当月入职、离职、转正或调薪员工",
		"formula_expression": "按薪资异动生效日期、锁定考勤工时和经确认的未满月口径计算",
		"parameters_json": {"confirmed_policy_required": True},
		"rule_text": "薪资异动按生效日期取值；入离职未上班时间会影响全勤和缺勤，但固定薪资是否按工时、天数或分段薪资折算需单独确认。",
		"source_file": "5.1薪资福利.xlsx / 5.2人资考勤.xlsx",
		"source_sheet": "人事组薪资异动表 / 薪资结算表",
		"missing_rule_note": "未满月固定薪资折算、月中调薪分段、离职结算截止日尚未形成唯一可执行规则，当前系统阻止将说明文本当作公式执行。",
	},
	{
		"rule_code": "PAYROLL_TERMINATION_SETTLEMENT",
		"rule_name": "离职薪资结算",
		"rule_category": "薪资结算",
		"rule_scope": "当月离职、异常离职或开除员工",
		"formula_expression": "试用期未满7天按符合8小时的工作日100元/天，工作1天离职无薪；其他扣款按已批准离职申请单确认",
		"parameters_json": {"trial_under_seven_days_daily_amount": 100, "one_day_departure_amount": 0, "approved_source_required": True},
		"rule_text": "试用期未提前3个工作日申请、正式员工未提前30日申请、旷工或开除涉及特殊扣款，必须由离职申请单和考勤终稿共同确认。",
		"source_file": "6.9离职.xlsx",
		"source_sheet": "离职作业规范 / 人事组员工辞职申请单",
		"source_cell": "离职作业规范 rows 29-31；辞职申请单 rows 9-16",
		"missing_rule_note": "此规则涉及劳动关系和审批结论，当前仅作为离职薪资结算来源记录，不自动从员工状态推导扣款。",
	},
	{
		"rule_code": "PAYROLL_MINIMUM_WAGE_CHECK",
		"rule_name": "最低工资差异检查",
		"rule_category": "薪资结算",
		"rule_scope": "按员工工作地和规则生效区间",
		"formula_expression": "将适用口径工资与员工工作地、生效期间的已确认最低工资标准比较",
		"parameters_json": {"confirmed_legal_source_required": True, "enforcement": "warning"},
		"rule_text": "成熟系统页面提供最低工资配置入口；本项目只把它作为差异检查，不用未经确认的金额自动改写工资。",
		"source_file": "成熟系统参考截图",
		"source_sheet": "薪酬设置 / 最低工资",
		"missing_rule_note": "公司资料未提供工作地、标准金额、生效区间和比较口径，需人事/法务按现行政策确认后才能启用阻断。",
	},
]

# Formula text is deliberately not evaluated as Python/JavaScript.  Payroll is
# a high-risk financial domain: only registered calculation drivers may run.
# The editable JSON below is the executable part of a rule and is validated
# before a payroll month can be generated.
EXECUTABLE_PAYROLL_RULES = {
	"SALARY_STRUCTURE_SUBTOTAL": {"parameters": ()},
	"ATTENDANCE_FULL_ATTENDANCE_BONUS": {"parameters": ("thresholds",)},
	"PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION": {"parameters": ("standard_hours_divisor", "absenteeism_multiplier")},
	"ATTENDANCE_MISSED_PUNCH": {"parameters": ("red_apples_per_record", "amount_per_apple")},
	"PAYROLL_SETTLEMENT_OVERTIME_PAY": {"parameters": ("standard_hours_divisor", "weekday", "weekend", "holiday")},
	"PAYROLL_SETTLEMENT_NIGHT_SHIFT": {"parameters": ("deep_night_shift", "large_night_shift", "small_night_shift")},
	"WELFARE_SOCIAL_SECURITY_COMPANY": {"parameters": ("ranges",)},
	"WELFARE_HOUSING_FUND_COMPANY": {"parameters": ("default_company_equals_personal",)},
}

# These formulas are calculated by the settlement pipeline itself.  They are
# intentionally visible in the rule centre for audit, but their expression is
# not editable JSON: changing their structure requires a reviewed code change,
# rather than evaluating arbitrary text in a payroll run.
FIXED_PAYROLL_CALCULATION_RULES = {
	"ATTENDANCE_OVERTIME_HOURS",
	"PAYROLL_SETTLEMENT_GROSS_PAY",
	"PAYROLL_SETTLEMENT_NET_PAY",
}

PAYROLL_SETTLEMENT_FIELD_MAPPINGS = [
	{"mapping_code": "EXCEL_B_DEPARTMENT", "display_order": 2, "excel_column": "B", "excel_label": "部门", "system_field": "department", "source_module": "员工档案", "source_detail": "Employee.department 或考勤/变量来源部门"},
	{"mapping_code": "EXCEL_C_EMPLOYEE_CODE", "display_order": 3, "excel_column": "C", "excel_label": "工号", "system_field": "employee_code", "source_module": "员工档案", "source_detail": "Employee 工号"},
	{"mapping_code": "EXCEL_D_EMPLOYEE_NAME", "display_order": 4, "excel_column": "D", "excel_label": "姓名", "system_field": "employee_name", "source_module": "员工档案", "source_detail": "Employee.employee_name"},
	{"mapping_code": "EXCEL_E_BASE_SALARY", "display_order": 5, "excel_column": "E", "excel_label": "底薪", "system_field": "base_salary", "source_module": "薪资主数据", "source_detail": "员工薪资异动优先，其次薪资变量", "rule_code": "SALARY_STRUCTURE_SUBTOTAL"},
	{"mapping_code": "EXCEL_F_FUNCTION_ALLOWANCE", "display_order": 6, "excel_column": "F", "excel_label": "职能津贴", "system_field": "function_allowance", "source_module": "薪资主数据", "source_detail": "员工薪资异动优先，其次薪资变量", "rule_code": "SALARY_STRUCTURE_SUBTOTAL"},
	{"mapping_code": "EXCEL_G_CERTIFICATE_SKILL_ALLOWANCE", "display_order": 7, "excel_column": "G", "excel_label": "证书及多能工津贴", "system_field": "certificate_skill_allowance", "source_module": "薪资主数据", "source_detail": "证书津贴 + 多能工津贴", "rule_code": "SALARY_STRUCTURE_SUBTOTAL"},
	{"mapping_code": "EXCEL_H_SALARY_SUBTOTAL", "display_order": 8, "excel_column": "H", "excel_label": "薪资小计", "system_field": "salary_subtotal", "source_module": "公式计算", "formula_expression": "SUM(E:G)", "rule_code": "SALARY_STRUCTURE_SUBTOTAL"},
	{"mapping_code": "EXCEL_I_STANDARD_HOURS", "display_order": 9, "excel_column": "I", "excel_label": "标准工时", "system_field": "standard_hours", "source_module": "考勤终稿", "source_detail": "HRMS Monthly Attendance Summary.standard_hours"},
	{"mapping_code": "EXCEL_J_BASIC_ATTENDANCE_HOURS", "display_order": 10, "excel_column": "J", "excel_label": "调整前/基本出勤工时", "system_field": "basic_attendance_hours", "source_module": "考勤终稿", "source_detail": "实际出勤/基本出勤工时"},
	{"mapping_code": "EXCEL_K_ABSENCE_HOURS", "display_order": 11, "excel_column": "K", "excel_label": "缺勤工时", "system_field": "missing_hours", "source_module": "公式计算", "formula_expression": "I-J", "rule_code": "PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION"},
	{"mapping_code": "EXCEL_L_RAW_WEEKEND_OVERTIME", "display_order": 12, "excel_column": "L", "excel_label": "调整前周末加班", "system_field": "raw_weekend_overtime_hours", "source_module": "考勤终稿", "source_detail": "考勤终稿2倍加班工时"},
	{"mapping_code": "EXCEL_M_ADJUSTED_ABSENCE_HOURS", "display_order": 13, "excel_column": "M", "excel_label": "调整后缺勤工时", "system_field": "adjusted_absence_hours", "source_module": "公式计算", "formula_expression": "IF(K-L>0,K-L,0)", "rule_code": "PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION"},
	{"mapping_code": "EXCEL_N_ABSENCE_DEDUCTION", "display_order": 14, "excel_column": "N", "excel_label": "缺勤工时对应的扣除金额", "system_field": "absence_deduction_amount", "source_module": "公式计算", "formula_expression": "ROUND(H/174*M,2)", "rule_code": "PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION"},
	{"mapping_code": "EXCEL_O_ADJUSTED_WEEKEND_OVERTIME", "display_order": 15, "excel_column": "O", "excel_label": "调整后周末加班", "system_field": "weekend_overtime_hours", "source_module": "公式计算", "formula_expression": "L-K+M", "rule_code": "ATTENDANCE_OVERTIME_HOURS"},
	{"mapping_code": "EXCEL_P_WEEKDAY_OVERTIME_HOURS", "display_order": 16, "excel_column": "P", "excel_label": "平日加班时数", "system_field": "weekday_overtime_hours", "source_module": "考勤终稿", "rule_code": "ATTENDANCE_OVERTIME_HOURS"},
	{"mapping_code": "EXCEL_Q_HOLIDAY_OVERTIME_HOURS", "display_order": 17, "excel_column": "Q", "excel_label": "节假日加班时数", "system_field": "holiday_overtime_hours", "source_module": "考勤终稿", "rule_code": "ATTENDANCE_OVERTIME_HOURS"},
	{"mapping_code": "EXCEL_R_WEEKDAY_OVERTIME_PAY", "display_order": 18, "excel_column": "R", "excel_label": "加班费/平日", "system_field": "weekday_overtime_pay", "source_module": "公式计算", "formula_expression": "ROUND(E/174*P*1.5,2)", "rule_code": "PAYROLL_SETTLEMENT_OVERTIME_PAY"},
	{"mapping_code": "EXCEL_S_WEEKEND_OVERTIME_PAY", "display_order": 19, "excel_column": "S", "excel_label": "加班费/周末", "system_field": "weekend_overtime_pay", "source_module": "公式计算", "formula_expression": "ROUND(E/174*2*O,2)", "rule_code": "PAYROLL_SETTLEMENT_OVERTIME_PAY"},
	{"mapping_code": "EXCEL_T_HOLIDAY_OVERTIME_PAY", "display_order": 20, "excel_column": "T", "excel_label": "加班费/节假日", "system_field": "holiday_overtime_pay", "source_module": "公式计算", "formula_expression": "ROUND(E/174*3*Q,2)", "rule_code": "PAYROLL_SETTLEMENT_OVERTIME_PAY"},
	{"mapping_code": "EXCEL_U_OVERTIME_PAY_TOTAL", "display_order": 21, "excel_column": "U", "excel_label": "加班费小计", "system_field": "overtime_pay_total", "source_module": "公式计算", "formula_expression": "SUM(R:T)", "rule_code": "PAYROLL_SETTLEMENT_OVERTIME_PAY"},
	{"mapping_code": "EXCEL_V_LARGE_NIGHT_SHIFT_COUNT", "display_order": 22, "excel_column": "V", "excel_label": "大夜班次数", "system_field": "large_night_shift_count", "source_module": "考勤终稿", "source_detail": "已扣除符合深夜班时段的次数"},
	{"mapping_code": "EXCEL_W_SMALL_NIGHT_SHIFT_COUNT", "display_order": 23, "excel_column": "W", "excel_label": "小夜班次数", "system_field": "small_night_shift_count", "source_module": "考勤终稿"},
	{"mapping_code": "EXCEL_X_NIGHT_SHIFT_ALLOWANCE", "display_order": 24, "excel_column": "X", "excel_label": "夜班津贴", "system_field": "night_shift_allowance", "source_module": "公式计算", "formula_expression": "深夜班、大夜班、小夜班次数分别乘以当前规则设置的津贴标准", "rule_code": "PAYROLL_SETTLEMENT_NIGHT_SHIFT"},
	{"mapping_code": "EXCEL_Y_ATTENDANCE_WAGE", "display_order": 25, "excel_column": "Y", "excel_label": "出勤工资", "system_field": "attendance_wage", "source_module": "公式计算", "formula_expression": "H-N+U+X-AH", "rule_code": "PAYROLL_SETTLEMENT_GROSS_PAY"},
	{"mapping_code": "EXCEL_Z_PROPOSAL_IMPROVEMENT_BONUS", "display_order": 26, "excel_column": "Z", "excel_label": "提案改善奖", "system_field": "proposal_improvement_bonus", "source_module": "福利扣款", "source_detail": "福利扣款来源中心/薪资变量"},
	{"mapping_code": "EXCEL_AA_APPLE_REWARD_AMOUNT", "display_order": 27, "excel_column": "AA", "excel_label": "红绿苹果", "system_field": "apple_reward_amount", "source_module": "考勤终稿", "source_detail": "苹果树奖惩"},
	{"mapping_code": "EXCEL_AB_SUBSIDY_BONUS_TOTAL", "display_order": 28, "excel_column": "AB", "excel_label": "全勤奖,住房学历补贴", "system_field": "subsidy_bonus_total", "source_module": "考勤终稿/福利扣款", "formula_expression": "全勤奖+住房补贴+学历补贴", "rule_code": "ATTENDANCE_FULL_ATTENDANCE_BONUS"},
	{"mapping_code": "EXCEL_AC_PRODUCTION_BONUS", "display_order": 29, "excel_column": "AC", "excel_label": "生产奖", "system_field": "production_bonus", "source_module": "福利扣款"},
	{"mapping_code": "EXCEL_AD_BONUS_TOTAL", "display_order": 30, "excel_column": "AD", "excel_label": "奖金小计", "system_field": "bonus_total", "source_module": "公式计算", "formula_expression": "SUM(Z:AC)", "rule_code": "PAYROLL_SETTLEMENT_GROSS_PAY"},
	{"mapping_code": "EXCEL_AE_ABSENTEEISM_HOURS", "display_order": 31, "excel_column": "AE", "excel_label": "旷工工时", "system_field": "absenteeism_hours", "source_module": "考勤终稿"},
	{"mapping_code": "EXCEL_AF_ABSENTEEISM_DEDUCTION", "display_order": 32, "excel_column": "AF", "excel_label": "旷工扣款", "system_field": "absenteeism_deduction", "source_module": "公式计算", "formula_expression": "ROUND(H/174*AE*3,2)", "rule_code": "PAYROLL_SETTLEMENT_GROSS_PAY"},
	{"mapping_code": "EXCEL_AG_LATE_FULL_ATTENDANCE_DEDUCTION", "display_order": 33, "excel_column": "AG", "excel_label": "迟到金额+全勤奖扣款", "system_field": "late_full_attendance_deduction", "source_module": "薪资变量", "rule_code": "ATTENDANCE_FULL_ATTENDANCE_BONUS"},
	{"mapping_code": "EXCEL_AH_PUNISHMENT_TOTAL", "display_order": 34, "excel_column": "AH", "excel_label": "惩处小计", "system_field": "punishment_total", "source_module": "公式计算", "formula_expression": "SUM(AF:AG)", "rule_code": "PAYROLL_SETTLEMENT_GROSS_PAY"},
	{"mapping_code": "EXCEL_AI_GROSS_PAY", "display_order": 35, "excel_column": "AI", "excel_label": "应付工资", "system_field": "gross_pay", "source_module": "公式计算", "formula_expression": "H-N+U+X+AD-AH", "rule_code": "PAYROLL_SETTLEMENT_GROSS_PAY"},
	{"mapping_code": "EXCEL_AJ_SOCIAL_SECURITY_PERSONAL", "display_order": 36, "excel_column": "AJ", "excel_label": "保险基金员工负担额", "system_field": "social_security_personal", "source_module": "福利扣款", "source_detail": "社保个人"},
	{"mapping_code": "EXCEL_AK_HOUSING_FUND_PERSONAL", "display_order": 37, "excel_column": "AK", "excel_label": "住房公积金", "system_field": "housing_fund_personal", "source_module": "福利扣款", "source_detail": "公积金个人"},
	{"mapping_code": "EXCEL_AL_PAID_WELFARE", "display_order": 38, "excel_column": "AL", "excel_label": "提案改善奖&生日福利金（已发）", "system_field": "paid_proposal_birthday_welfare", "source_module": "福利扣款", "source_detail": "已发福利"},
	{"mapping_code": "EXCEL_AM_TAXABLE_SALARY", "display_order": 39, "excel_column": "AM", "excel_label": "计税工资", "system_field": "taxable_salary", "source_module": "公式计算", "formula_expression": "AI-AJ-AK+AL", "rule_code": "PAYROLL_SETTLEMENT_NET_PAY"},
	{"mapping_code": "EXCEL_AN_CONTINUING_SERVICE_BONUS", "display_order": 40, "excel_column": "AN", "excel_label": "继续服务奖", "system_field": "continuing_service_bonus", "source_module": "福利扣款"},
	{"mapping_code": "EXCEL_AP_INCOME_TAX", "display_order": 42, "excel_column": "AP", "excel_label": "所得税代扣款", "system_field": "income_tax", "source_module": "福利扣款", "rule_code": "TAX_INCOME_TAX"},
	{"mapping_code": "EXCEL_AQ_YEAR_END_BONUS_TAX", "display_order": 43, "excel_column": "AQ", "excel_label": "年终奖所得税", "system_field": "year_end_bonus_tax", "source_module": "福利扣款", "rule_code": "TAX_INCOME_TAX"},
	{"mapping_code": "EXCEL_AR_UTILITIES_DEDUCTION", "display_order": 44, "excel_column": "AR", "excel_label": "水电费及扣款", "system_field": "utilities_deduction", "source_module": "福利扣款"},
	{"mapping_code": "EXCEL_AS_NET_PAY", "display_order": 45, "excel_column": "AS", "excel_label": "实发工资", "system_field": "net_pay", "source_module": "公式计算", "formula_expression": "AM-AP-AQ-AR+AN-AL", "rule_code": "PAYROLL_SETTLEMENT_NET_PAY"},
	{"mapping_code": "EXCEL_AT_SOCIAL_SECURITY_COMPANY", "display_order": 46, "excel_column": "AT", "excel_label": "保险基金公司负担额", "system_field": "social_security_company", "source_module": "公式计算", "formula_expression": "社保公司手工确认优先；否则按AJ区间反推", "rule_code": "WELFARE_SOCIAL_SECURITY_COMPANY"},
	{"mapping_code": "EXCEL_AU_HOUSING_FUND_COMPANY", "display_order": 47, "excel_column": "AU", "excel_label": "住房公积金公司负担", "system_field": "housing_fund_company", "source_module": "公式计算", "formula_expression": "AK", "rule_code": "WELFARE_HOUSING_FUND_COMPANY"},
	{"mapping_code": "EXCEL_AV_COMPANY_COST_TOTAL", "display_order": 48, "excel_column": "AV", "excel_label": "公司实际负担总计", "system_field": "company_cost_total", "source_module": "公式计算", "formula_expression": "AI+AT+AU+AN+AL", "rule_code": "PAYROLL_SETTLEMENT_NET_PAY"},
	{"mapping_code": "EXCEL_AW_EXPORT_OVERTIME_PAY", "display_order": 49, "excel_column": "AW", "excel_label": "加班工资", "system_field": "overtime_pay_total", "source_module": "导出辅助", "formula_expression": "U"},
	{"mapping_code": "EXCEL_AX_EXPORT_NIGHT_SHIFT", "display_order": 50, "excel_column": "AX", "excel_label": "夜班津贴", "system_field": "night_shift_allowance", "source_module": "导出辅助", "formula_expression": "X"},
	{"mapping_code": "EXCEL_AY_EXPORT_GROSS_PAY", "display_order": 51, "excel_column": "AY", "excel_label": "应付工资", "system_field": "gross_pay", "source_module": "导出辅助", "formula_expression": "AI"},
	{"mapping_code": "EXCEL_AZ_EXPORT_PAID_WELFARE", "display_order": 52, "excel_column": "AZ", "excel_label": "各类奖金及福利", "system_field": "paid_proposal_birthday_welfare", "source_module": "导出辅助", "formula_expression": "AL"},
	{"mapping_code": "EXCEL_BA_EXPORT_CONTINUING_SERVICE", "display_order": 53, "excel_column": "BA", "excel_label": "继续服务奖", "system_field": "continuing_service_bonus", "source_module": "导出辅助", "formula_expression": "AN"},
	{"mapping_code": "EXCEL_BB_EXPORT_SOCIAL_SECURITY_COMPANY", "display_order": 54, "excel_column": "BB", "excel_label": "公司承担社保", "system_field": "social_security_company", "source_module": "导出辅助", "formula_expression": "AT"},
	{"mapping_code": "EXCEL_BC_EXPORT_HOUSING_FUND_COMPANY", "display_order": 55, "excel_column": "BC", "excel_label": "公司承担公积金", "system_field": "housing_fund_company", "source_module": "导出辅助", "formula_expression": "AU"},
	{"mapping_code": "EXCEL_BD_EXPORT_NET_PAY", "display_order": 56, "excel_column": "BD", "excel_label": "实发工资", "system_field": "net_pay", "source_module": "导出辅助", "formula_expression": "AS"},
	{"mapping_code": "EXCEL_BE_EXPORT_TAX_ADJUSTED_NET", "display_order": 57, "excel_column": "BE", "excel_label": "导出校验工资", "system_field": "export_tax_adjusted_net_pay", "source_module": "导出辅助", "formula_expression": "AM+AN-AR-AP-AL-AQ", "remarks": "Excel 原表 BE 列无表头，保留为导出校验口径。"},
]

# These are the calculated values explicitly copied into
# ``HRMS Payroll Settlement Record`` by ``generate_payroll_settlement_records``.
# Keeping the list beside the import mapping makes a configuration audit able to
# detect a formula that is valid syntactically but no longer reaches settlement.
PAYROLL_SETTLEMENT_FORMULA_OUTPUT_FIELDS = {
	"salary_subtotal", "missing_hours", "adjusted_absence_hours", "weekend_overtime_hours",
	"full_salary_hourly_rate", "base_salary_hourly_rate", "absence_deduction_amount",
	"weekday_overtime_pay", "weekend_overtime_pay", "holiday_overtime_pay", "overtime_pay_total",
	"night_shift_allowance", "subsidy_bonus_total", "bonus_total", "absenteeism_deduction",
	"punishment_total", "attendance_wage", "gross_pay", "taxable_salary", "net_pay",
	"social_security_company", "housing_fund_company", "company_cost_total", "export_tax_adjusted_net_pay",
}

# These values are assembled from approved salary changes, a locked attendance
# version and confirmed monthly variables before the formula engine runs.
PAYROLL_FORMULA_CONTEXT_FIELDS = {
	"base_salary", "function_allowance", "certificate_skill_allowance", "standard_hours",
	"basic_attendance_hours", "raw_weekend_overtime_hours", "weekday_overtime_hours",
	"holiday_overtime_hours", "deep_night_shift_count", "large_night_shift_count", "small_night_shift_count",
	"absenteeism_hours", "proposal_improvement_bonus", "apple_reward_amount",
	"full_attendance_bonus", "housing_subsidy", "education_subsidy", "other_bonus",
	"production_bonus", "late_full_attendance_deduction", "other_deduction",
	"social_security_personal", "housing_fund_personal", "paid_proposal_birthday_welfare",
	"continuing_service_bonus", "income_tax", "year_end_bonus_tax", "utilities_deduction",
	"manual_social_security_company", "manual_housing_fund_company",
}

# Atomic options that are combined into one historical Excel column.  They stay
# separate in the configuration centre and are aggregated only when the legacy
# settlement layout is produced.
PAYROLL_ATOMIC_CONFIGURATION_ITEMS = [
	{"item_code": "SALARY_DUTY_ALLOWANCE", "item_name": "职务津贴", "category": "固定薪资", "data_type": "金额", "direction": "应发", "source_module": "薪资主数据", "result_field": "function_allowance", "aggregate_target": "职能津贴"},
	{"item_code": "SALARY_CERTIFICATE_ALLOWANCE", "item_name": "证书津贴", "category": "固定薪资", "data_type": "金额", "direction": "应发", "source_module": "薪资主数据", "result_field": "certificate_skill_allowance", "aggregate_target": "证书及多能工津贴"},
	{"item_code": "SALARY_MULTI_SKILL_ALLOWANCE", "item_name": "多能工津贴", "category": "固定薪资", "data_type": "金额", "direction": "应发", "source_module": "薪资主数据", "result_field": "certificate_skill_allowance", "aggregate_target": "证书及多能工津贴"},
	{"item_code": "BONUS_FULL_ATTENDANCE", "item_name": "全勤奖", "category": "奖金补贴", "data_type": "金额", "direction": "应发", "source_module": "考勤终稿/福利扣款", "result_field": "subsidy_bonus_total", "aggregate_target": "全勤奖、住房及学历补贴", "rule_code": "ATTENDANCE_FULL_ATTENDANCE_BONUS"},
	{"item_code": "BONUS_HOUSING_SUBSIDY", "item_name": "住房补贴", "category": "奖金补贴", "data_type": "金额", "direction": "应发", "source_module": "考勤终稿", "result_field": "subsidy_bonus_total", "aggregate_target": "全勤奖、住房及学历补贴", "rule_code": "WELFARE_RENTAL_SUBSIDY"},
	{"item_code": "BONUS_EDUCATION_SUBSIDY", "item_name": "学历补贴", "category": "奖金补贴", "data_type": "金额", "direction": "应发", "source_module": "福利扣款", "result_field": "subsidy_bonus_total", "aggregate_target": "全勤奖、住房及学历补贴", "rule_code": "WELFARE_EDUCATION_SUBSIDY"},
	{"item_code": "BONUS_HIGH_TEMPERATURE", "item_name": "高温补贴", "category": "奖金补贴", "data_type": "金额", "direction": "应发", "source_module": "福利扣款", "result_field": "subsidy_bonus_total", "aggregate_target": "其他奖金"},
	{"item_code": "BONUS_PHONE", "item_name": "手机话费补贴", "category": "奖金补贴", "data_type": "金额", "direction": "应发", "source_module": "福利扣款", "result_field": "subsidy_bonus_total", "aggregate_target": "其他奖金"},
	{"item_code": "BONUS_FUEL", "item_name": "油费补贴", "category": "奖金补贴", "data_type": "金额", "direction": "应发", "source_module": "福利扣款", "result_field": "subsidy_bonus_total", "aggregate_target": "其他奖金"},
	{"item_code": "DEDUCTION_LATE", "item_name": "迟到扣款", "category": "扣款税费", "data_type": "金额", "direction": "应扣", "source_module": "考勤终稿/薪资变量", "result_field": "late_full_attendance_deduction", "aggregate_target": "迟到金额及全勤奖扣款"},
	{"item_code": "DEDUCTION_FULL_ATTENDANCE", "item_name": "全勤奖扣款", "category": "扣款税费", "data_type": "金额", "direction": "应扣", "source_module": "考勤终稿", "result_field": "late_full_attendance_deduction", "aggregate_target": "迟到金额及全勤奖扣款", "rule_code": "ATTENDANCE_FULL_ATTENDANCE_BONUS"},
	{"item_code": "DEDUCTION_DORMITORY", "item_name": "宿舍住宿费", "category": "扣款税费", "data_type": "金额", "direction": "应扣", "source_module": "福利扣款", "result_field": "utilities_deduction", "aggregate_target": "水电费及扣款", "rule_code": "WELFARE_DORMITORY_FEE"},
	{"item_code": "DEDUCTION_DORM_WATER", "item_name": "宿舍水费", "category": "扣款税费", "data_type": "金额", "direction": "应扣", "source_module": "福利扣款", "result_field": "utilities_deduction", "aggregate_target": "水电费及扣款", "rule_code": "WELFARE_DORMITORY_FEE"},
	{"item_code": "DEDUCTION_DORM_ELECTRICITY", "item_name": "宿舍电费", "category": "扣款税费", "data_type": "金额", "direction": "应扣", "source_module": "福利扣款", "result_field": "utilities_deduction", "aggregate_target": "水电费及扣款", "rule_code": "WELFARE_DORMITORY_FEE"},
	{"item_code": "BONUS_OTHER", "item_name": "其他奖金", "category": "奖金补贴", "data_type": "金额", "direction": "应发", "source_module": "已确认月度变量", "result_field": "subsidy_bonus_total", "aggregate_target": "奖金小计"},
	{"item_code": "DEDUCTION_OTHER", "item_name": "其他扣款", "category": "扣款税费", "data_type": "金额", "direction": "应扣", "source_module": "已确认月度变量", "result_field": "punishment_total", "aggregate_target": "惩处小计"},
]


def _get_file_content(file_url):
	name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not name:
		frappe.throw(_("未找到上传文件"))
	file_doc = frappe.get_doc("File", name)
	content = file_doc.get_content()
	if isinstance(content, str):
		content = content.encode()
	return content


def _load_workbook(file_url):
	from openpyxl import load_workbook

	return load_workbook(BytesIO(_get_file_content(file_url)), data_only=True, read_only=True)


def _text(value):
	if value is None:
		return ""
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d %H:%M:%S")
	if isinstance(value, date):
		return value.isoformat()
	return str(value).strip()


def _normalise(value):
	return re.sub(r"\s+", "", _text(value).replace("\n", ""))


def _doctype_exists(doctype):
	return frappe.db.exists("DocType", doctype)


def _doctype_has_field(doctype, fieldname):
	if fieldname == "name":
		return True
	if not _doctype_exists(doctype):
		return False
	return frappe.get_meta(doctype).has_field(fieldname)


def _safe_fields(doctype, fields):
	return [field for field in fields if _doctype_has_field(doctype, field)]


def _safe_count(doctype, filters=None):
	if not _doctype_exists(doctype):
		return 0
	try:
		return frappe.db.count(doctype, filters or {})
	except Exception:
		return 0


def _safe_get_all(doctype, fields=None, filters=None, order_by=None, limit_page_length=50):
	if not _doctype_exists(doctype):
		return []
	fields = _safe_fields(doctype, fields or ["name"])
	return frappe.get_all(
		doctype,
		fields=fields,
		filters=filters or {},
		order_by=order_by,
		limit_page_length=int(limit_page_length or 50),
	)


def _employee_code(row):
	return row.get("custom_employee_code") or ""


def _read_rows(sheet, max_rows=None):
	rows = []
	for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
		values = [_text(value) for value in row]
		if any(values):
			rows.append(values)
		if max_rows and index >= max_rows:
			break
	return rows


def _find_header_index(rows):
	keywords = {"工号", "姓名", "部门", "全勤奖", "住房补贴", "学历补贴", "个人承担", "当月扣款", "金额"}
	best_index = 0
	best_score = -1
	for index, row in enumerate(rows[:20]):
		headers = {_normalise(value) for value in row if value}
		score = len(keywords & headers)
		if score > best_score:
			best_index = index
			best_score = score
	return best_index


def _rows_as_dicts(sheet):
	rows = _read_rows(sheet)
	if not rows:
		return []
	header_index = _find_header_index(rows)
	seen = defaultdict(int)
	headers = []
	for value in rows[header_index]:
		header = _normalise(value)
		if not header:
			headers.append("")
			continue
		seen[header] += 1
		headers.append(f"{header}_{seen[header]}" if seen[header] > 1 else header)

	items = []
	for row in rows[header_index + 1 :]:
		item = {}
		for index, header in enumerate(headers):
			if header:
				item[header] = row[index] if index < len(row) else ""
		if any(item.values()):
			items.append(item)
	return items


def _first(row, *headers):
	for header in headers:
		value = row.get(_normalise(header))
		if value not in (None, ""):
			return value
	return ""


def _date_or_none(value):
	if not value:
		return None
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()
	text = _text(value)
	if not text:
		return None
	try:
		return getdate(text).isoformat()
	except Exception:
		return None


def _bool_value(value):
	text = _text(value).lower()
	if text in ("1", "true", "yes", "y", "是", "启用", "已启用"):
		return 1
	if text in ("0", "false", "no", "n", "否", "停用", "未启用"):
		return 0
	return 0


def _template_by_sheet(sheet_name):
	for template in PAYROLL_IMPORT_TEMPLATES:
		if template["sheet_name"] == sheet_name:
			return template
	return None


def _find_template_header_index(rows, template):
	required = {_normalise(column) for column in template.get("required_columns", [])}
	best_index = 0
	best_score = -1
	for index, row in enumerate(rows[:20]):
		headers = {_normalise(value) for value in row if value}
		score = len(required & headers)
		if score > best_score:
			best_index = index
			best_score = score
	return best_index


def _template_rows_as_dicts(sheet, template):
	rows = _read_rows(sheet)
	if not rows:
		return []
	header_index = _find_template_header_index(rows, template)
	seen = defaultdict(int)
	headers = []
	for value in rows[header_index]:
		header = _normalise(value)
		if not header:
			headers.append("")
			continue
		seen[header] += 1
		headers.append(f"{header}_{seen[header]}" if seen[header] > 1 else header)

	items = []
	for row in rows[header_index + 1 :]:
		item = {}
		for index, header in enumerate(headers):
			if header:
				item[header] = row[index] if index < len(row) else ""
		if any(item.values()):
			items.append(item)
	return items


def _amount_for_type(row, variable_type):
	if variable_type == "全勤奖":
		return flt(_first(row, "全勤奖", "金额"))
	if variable_type == "住房补贴":
		return flt(_first(row, "住房补贴", "金额"))
	if variable_type == "学历补贴":
		return flt(_first(row, "学历补贴", "补贴金额(元)", "补贴金额", "金额"))
	if variable_type == "宿舍扣款":
		return flt(_first(row, "当月扣款", "扣款", "金额"))
	if variable_type == "证书及多能工津贴":
		combined = _first(row, "证书及多能工津贴")
		return flt(combined) or flt(_first(row, "证书津贴")) + flt(_first(row, "多能工津贴")) or flt(_first(row, "金额"))
	if variable_type == "继续服务奖":
		return flt(_first(row, "继续服务奖", "金额"))
	if variable_type == "提案改善奖":
		return flt(_first(row, "提案改善奖", "金额"))
	if variable_type == "苹果树":
		return flt(_first(row, "苹果树", "绿苹果", "红苹果", "金额"))
	if variable_type == "其他奖金":
		return flt(_first(row, "金额（元）", "金额", "奖惩金额"))
	if variable_type == "其他扣款":
		return flt(_first(row, "扣款", "金额", "水电费及扣款"))
	if variable_type == "社保个人":
		pension = flt(_first(row, "企业养老8%", "个人养老", "养老个人"))
		unemployment = flt(_first(row, "失业保险约0.5%", "个人失业", "失业个人"))
		medical = flt(_first(row, "基本医疗2%", "个人医疗", "医疗个人"))
		mutual_medical = flt(_first(row, "互助医疗"))
		total = flt(_first(row, "个人合计承担10.5%+5", "个人合计", "个人承担合计", "金额"))
		return total or pension + unemployment + medical + mutual_medical
	return flt(_first(row, "金额"))


def _employee_lookup(employee_code=None, employee_name=None):
	if employee_code:
		for fieldname in ("custom_employee_code",):
			try:
				name = frappe.db.get_value("Employee", {fieldname: employee_code}, "name")
			except Exception:
				name = None
			if name:
				return name
	if employee_name:
		return frappe.db.get_value("Employee", {"employee_name": employee_name}, "name")
	return None


def _payroll_employee_lookup(employee_code=None, employee_name=None):
	"""Resolve a payroll employee using only the source employee code and name.

	Departments and job information are descriptive payroll fields; they must
	not influence identity matching.  When a source provides both identifiers,
	they have to resolve to the same employee rather than allowing one of them
	to silently override the other.
	"""
	employee_code = _text(employee_code)
	employee_name = _text(employee_name)
	by_code = _employee_lookup(employee_code, "") if employee_code else None
	by_name = _employee_lookup("", employee_name) if employee_name else None
	if employee_code and employee_name:
		return by_code if by_code and by_code == by_name else None
	return by_code or by_name


def _department_lookup(department, company=None):
	"""Resolve a Department link without leaking legacy company suffixes.

	The HR workspace now keeps the business department name as its document ID.
	Older imported records can still contain values such as ``行政科 - 1D``.
	Accept that historical representation only when it resolves inside the same
	company; never fall through to a department belonging to another company.
	"""
	department = (department or "").strip()
	if not department:
		return None
	filters = {"name": department}
	if company:
		filters["company"] = company
	if frappe.db.get_value("Department", filters, "name"):
		return department
	filters = {"department_name": department}
	if company:
		filters["company"] = company
	resolved = frappe.db.get_value("Department", filters, "name")
	if resolved:
		return resolved
	# Pre-normalisation identifiers were stored as “显示名 - 公司缩写”.  Do
	# not remove arbitrary punctuation: only retry the final suffix pattern.
	legacy_display_name = re.sub(r"\s+-\s+[^-]+$", "", department).strip()
	if legacy_display_name and legacy_display_name != department:
		filters = {"department_name": legacy_display_name}
		if company:
			filters["company"] = company
		return frappe.db.get_value("Department", filters, "name")
	return None


def _default_company():
	return frappe.defaults.get_user_default("Company") or frappe.db.get_single_value("Global Defaults", "default_company") or frappe.db.get_value("Company", {}, "name")


def _ensure_department(department):
	department = (department or "").strip()
	if not department:
		return None
	existing = _department_lookup(department)
	if existing:
		return existing
	company = _default_company()
	if not company:
		return department
	doc = frappe.get_doc({
		"doctype": "Department",
		"department_name": department,
		"company": company,
		"is_group": 0,
	})
	doc.insert(ignore_permissions=True)
	return doc.name


def _matching_sheet(workbook, expected_name):
	expected = (expected_name or "").strip()
	for sheet_name in workbook.sheetnames:
		actual = sheet_name.strip()
		if actual == expected or (expected and expected in actual):
			return workbook[sheet_name]
	return None


# The monthly source files issued by Yongxin are business forms, not system
# templates.  Their sheet names change with the month (for example ``2607社保
# 名单`` and ``6月继续服务奖``), so recognize them from their business headers.
# This keeps the original form usable while still importing through one audited
# payroll-variable batch.
RAW_PAYROLL_SOURCE_LABELS = {
	"salary_change": "薪资异动单",
	"certificate_skill": "证书及多能工津贴",
	"proposal": "提案改善",
	"continuing_service": "继续服务奖",
	"dormitory": "宿舍水电费",
	"reward_punishment": "奖惩提报单",
	"education": "学历补贴",
	"housing_allowance": "住房补贴",
	"housing_allowance_base": "住房补贴一阶基础数据",
	"attendance_bonus": "全勤奖（考勤终稿自动继承）",
	"housing_fund": "公积金",
	"social_insurance": "社保",
}


def _sheet_headers(sheet):
	return {_normalise(value) for row in _read_rows(sheet, max_rows=8) for value in row if value}


def _raw_payroll_source_kind(sheet):
	headers = _sheet_headers(sheet)
	title = _normalise(sheet.title)
	if {"调整月份", "调整后职能薪资", "工号", "姓名"}.issubset(headers):
		return "salary_change"
	if "证书津贴" in headers and "多能工津贴" in headers:
		return "certificate_skill"
	if "奖励人1" in headers and ({"奖金(元)", "奖金（元）"} & headers):
		return "proposal"
	if "继续服务奖" in title:
		return "continuing_service"
	if "当月扣款" in headers and "租房补贴" in headers and "实收金额" in headers:
		return "dormitory"
	if "奖惩条例" in headers and ({"金额(元)", "金额（元）"} & headers):
		return "reward_punishment"
	if "补贴金额(元)" in headers and "学历类别" in headers:
		return "education"
	if "全勤奖" in headers and "工号" in headers and "姓名" in headers:
		return "attendance_bonus"
	if "住房补贴" in headers and "工号" in headers and "姓名" in headers:
		return "housing_allowance"
	if {"工号", "姓名"}.issubset(headers) and headers & {"是否非苏州户籍", "是否在苏州租房", "是否在苏州购房", "是否住公司宿舍"}:
		return "housing_allowance_base"
	if "公积金账号" in headers and "公司承担" in headers and "个人承担" in headers:
		return "housing_fund"
	if "社保基数" in headers and "个人合计承担10.5%+5" in headers:
		return "social_insurance"
	return None


def _repeated_employee_amount_rows(sheet, amount_header):
	"""Read side-by-side employee blocks used by the July allowance forms.

	The source workbook repeats ``工号/姓名/单位/金额`` twice on one sheet.  A
	flat dict reader only keeps the first employee and can accidentally pair that
	employee with the second block's amount.  Resolve each amount column back to
	the nearest identity columns so every employee/amount pair stays intact.
	"""
	rows = _read_rows(sheet)
	if not rows:
		return []
	normalized_amount = _normalise(amount_header)
	header_index = next(
		(
			index
			for index, row in enumerate(rows[:12])
			if any(_normalise(value) == normalized_amount for value in row)
			and "工号" in {_normalise(value) for value in row}
			and "姓名" in {_normalise(value) for value in row}
		),
		None,
	)
	if header_index is None:
		return []
	header = rows[header_index]
	amount_columns = [index for index, value in enumerate(header) if _normalise(value) == normalized_amount]
	groups = []
	for amount_index in amount_columns:
		window_start = max(0, amount_index - 5)
		code_index = next((index for index in range(amount_index - 1, window_start - 1, -1) if _normalise(header[index]) == "工号"), None)
		name_index = next((index for index in range(amount_index - 1, window_start - 1, -1) if _normalise(header[index]) == "姓名"), None)
		department_index = next((index for index in range(amount_index - 1, window_start - 1, -1) if _normalise(header[index]) in {"部门", "单位"}), None)
		if code_index is not None and name_index is not None:
			groups.append((code_index, name_index, department_index, amount_index))
	items = []
	for row in rows[header_index + 1 :]:
		for code_index, name_index, department_index, amount_index in groups:
			code = row[code_index] if code_index < len(row) else ""
			name = row[name_index] if name_index < len(row) else ""
			if not code and not name:
				continue
			amount = row[amount_index] if amount_index < len(row) else ""
			items.append(
				{
					"工号": code,
					"姓名": name,
					"部门": row[department_index] if department_index is not None and department_index < len(row) else "",
					amount_header: amount,
					"金额": amount,
				}
			)
	return [item for item in items if _is_employee_source_row(item)]


def _is_employee_source_row(row):
	"""Exclude totals, signatures and rule tables without hiding real people."""
	code = _text(_first(row, "工号", "受奖惩人工号"))
	name = _text(_first(row, "姓名", "受奖/惩人", "受奖惩人姓名"))
	if not name or re.fullmatch(r"[\d.]+", name):
		return False
	if re.fullmatch(r"共\s*\d+\s*人", name) or any(marker in name for marker in ("合计", "共计", "小计", "总计", "审核", "制表", "批准", "备注", "标准", "月份", "人数", "补贴金额")):
		return False
	return not code or bool(re.fullmatch(r"\d+(?:\.0+)?", code))


def _raw_payroll_source_rows(sheet, source_kind):
	if source_kind == "salary_change":
		return _salary_change_form_rows(sheet)
	if source_kind == "housing_allowance":
		return _repeated_employee_amount_rows(sheet, "住房补贴")
	if source_kind == "housing_allowance_base":
		return [row for row in _rows_as_dicts(sheet) if _is_employee_source_row(row)]
	if source_kind == "attendance_bonus":
		return _repeated_employee_amount_rows(sheet, "全勤奖")
	if source_kind != "continuing_service":
		rows = _rows_as_dicts(sheet)
		if source_kind == "proposal":
			return [row for row in rows if _first(row, "提案人", "奖励人1", "奖励人2", "奖励人3")]
		return [row for row in rows if _is_employee_source_row(row)]
	rows = _read_rows(sheet)
	if not rows:
		return []
	header_index = next(
		(index for index, row in enumerate(rows[:10]) if sum(1 for value in row if _normalise(value) in {"工号", "姓名", "金额"}) >= 3),
		None,
	)
	if header_index is None:
		return []
	header = rows[header_index]
	groups = []
	for start in range(len(header)):
		if _normalise(header[start]) != "工号":
			continue
		name_index = next((index for index in range(start + 1, len(header)) if _normalise(header[index]) == "姓名"), None)
		amount_index = next((index for index in range((name_index or start) + 1, len(header)) if _normalise(header[index]) == "金额"), None)
		department_index = next((index for index in range(max(0, start - 2), start) if _normalise(header[index]) == "部门"), None)
		if name_index is not None and amount_index is not None:
			groups.append((department_index, start, name_index, amount_index))
	items = []
	for row in rows[header_index + 1 :]:
		for department_index, code_index, name_index, amount_index in groups:
			code = row[code_index] if code_index < len(row) else ""
			name = row[name_index] if name_index < len(row) else ""
			amount = row[amount_index] if amount_index < len(row) else ""
			if code or name:
				item = {
					"部门": row[department_index] if department_index is not None and department_index < len(row) else "",
					"工号": code,
					"姓名": name,
					"金额": amount,
				}
				if _is_employee_source_row(item):
					items.append(item)
	return items


HOUSING_ALLOWANCE_BASE_FIELDS = (
	"是否非苏州户籍", "是否在苏州租房", "是否在苏州购房", "是否住公司宿舍",
)


def _housing_yes_no(row, field):
	"""Read an explicit yes/no value without treating a blank as ``否``."""
	value = _normalise(_first(row, field))
	if value in {"是", "1", "yes", "y", "true"}:
		return True
	if value in {"否", "0", "no", "n", "false"}:
		return False
	return None


def _housing_allowance_calculation(row, payroll_month, company):
	"""Turn a housing *eligibility* row into an auditable second-level amount.

	The source workbook records facts, not a payroll amount.  This intentionally
	keeps the policy in ``WELFARE_RENTAL_SUBSIDY`` so a rule change takes effect
	through the rule centre and every generated row retains its source facts.
	"""
	parameters = (
		_effective_rule_config("WELFARE_RENTAL_SUBSIDY", payroll_month, company)["parameters"]
		if company
		else _rule_parameters(_default_rule("WELFARE_RENTAL_SUBSIDY").get("parameters_json"))
	)
	missing = [field for field in HOUSING_ALLOWANCE_BASE_FIELDS if _housing_yes_no(row, field) is None]
	if missing:
		return {
			"amount": 0,
			"excluded": False,
			"error": "一阶数据缺少或无法识别：{0}（请填写是/否）".format("、".join(missing)),
			"reason": "无法计算住房补贴",
			"mode": "一阶数据系统计算",
		}

	if not _housing_yes_no(row, "是否非苏州户籍"):
		reason = "苏州户籍，不符合租房补贴资格"
	elif not _housing_yes_no(row, "是否在苏州租房"):
		reason = "未标记为在苏州租房，不符合租房补贴资格"
	elif _housing_yes_no(row, "是否在苏州购房"):
		reason = "已在苏州购房，不符合租房补贴资格"
	elif _housing_yes_no(row, "是否住公司宿舍"):
		reason = "住公司宿舍，按宿舍规则处理，不计住房补贴"
	else:
		reason = ""
	if reason:
		return {"amount": 0, "excluded": True, "error": "", "reason": reason, "mode": "一阶数据系统计算"}

	join_date = _date_or_none(_first(row, "入职日期"))
	leave_date = _date_or_none(_first(row, "离职日期"))
	try:
		year, month = [int(value) for value in payroll_month.split("-")]
	except (AttributeError, ValueError):
		return {"amount": 0, "excluded": False, "error": "薪资月份无效，无法计算住房补贴", "reason": "无法计算住房补贴", "mode": "一阶数据系统计算"}

	if leave_date and getdate(leave_date).year == year and getdate(leave_date).month == month:
		full_attendance = _housing_yes_no(row, "离职当月是否满勤")
		if full_attendance is None:
			return {"amount": 0, "excluded": False, "error": "离职当月需填写“离职当月是否满勤”（是/否）", "reason": "无法计算住房补贴", "mode": "一阶数据系统计算"}
		amount = flt(parameters.get("resignation_full_attendance")) if full_attendance else 0
		reason = "离职当月满勤，按离职规则计算" if full_attendance else "离职当月未满勤，不计住房补贴"
		return {"amount": amount, "excluded": not bool(amount), "error": "", "reason": reason, "mode": "一阶数据系统计算"}

	if join_date and getdate(join_date).year == year and getdate(join_date).month == month:
		join_day = getdate(join_date).day
		if join_day <= 10:
			amount = flt(parameters.get("before_or_on_day_10"))
		elif join_day <= 20:
			amount = flt(parameters.get("day_11_to_20"))
		else:
			amount = flt(parameters.get("after_or_on_day_21"))
		return {"amount": amount, "excluded": not bool(amount), "error": "", "reason": "入职当月 {0} 日，按入职日期规则计算".format(join_day), "mode": "一阶数据系统计算"}

	return {"amount": flt(parameters.get("before_or_on_day_10")), "excluded": False, "error": "", "reason": "符合住房补贴资格，按正常月度规则计算", "mode": "一阶数据系统计算"}


def _salary_change_form_rows(sheet):
	"""Read Yongxin's two-row ``人员薪资调整模板（月）`` header.

	The form keeps the new salary fields below the merged heading ``调整后职能
	薪资``.  A normal one-row header reader loses those labels and would read the
	*before-adjustment* columns by mistake.  Keep both groups explicitly named so
	the import always uses the adjusted salary values.
	"""
	rows = _read_rows(sheet)
	header_index = next(
		(index for index, row in enumerate(rows[:-1]) if {"调整月份", "工号", "姓名", "调整后职能薪资"}.issubset({_normalise(value) for value in row})),
		None,
	)
	if header_index is None:
		return []
	top_headers = rows[header_index]
	sub_headers = rows[header_index + 1]
	title_month = next((value for row in rows[:header_index] for value in row if _payroll_month_from_source(value, "")), "")
	sections, current_section = [], ""
	for value in top_headers:
		if _normalise(value):
			current_section = _normalise(value)
		sections.append(current_section)

	items = []
	for row in rows[header_index + 2 :]:
		item = {}
		for index, section in enumerate(sections):
			value = row[index] if index < len(row) else ""
			sub_header = _normalise(sub_headers[index] if index < len(sub_headers) else "")
			if section in {"调整前职能薪资", "调整后职能薪资"} and sub_header:
				key = f"{'调整前' if section == '调整前职能薪资' else '调整后'}{sub_header}"
			elif sub_header and not section:
				key = sub_header
			else:
				key = section
			if key:
				item[key] = value
		if not _first(item, "工号", "姓名"):
			continue
		item["调整月份"] = _first(item, "调整月份") or title_month
		items.append(item)
	return items


def _payroll_month_from_source(value, fallback):
	text = _text(value)
	match = re.search(r"(20\d{2})\D?(\d{1,2})", text)
	if match:
		return f"{match.group(1)}-{int(match.group(2)):02d}"
	return fallback


def _salary_change_row_from_source(row, payroll_month):
	month = _payroll_month_from_source(_first(row, "调整月份", "月份"), payroll_month)
	return {
		"工号": _text(_first(row, "工号")),
		"姓名": _first(row, "姓名"),
		"部门": _first(row, "部门"),
		"岗位": _first(row, "调整后岗位", "岗位_2", "岗位"),
		"生效日期": f"{month}-01",
		"异动原因": _first(row, "异动原因"),
		"薪资架构版本": _first(row, "调整后版本", "版本"),
		"薪资序号": _first(row, "调整后薪资序号", "薪资序号"),
		"底薪": _first(row, "调整后底薪", "底薪_2", "底薪"),
		"职能津贴": _first(row, "调整后职能津贴", "职能津贴_2", "职能津贴"),
		"薪资小计": _first(row, "调整后全薪", "全薪_2", "全薪"),
		"社保": _first(row, "社保费用"),
		"公积金": _first(row, "住房公积金"),
		"公司总承担": _first(row, "公司总承担"),
		"状态": "已批准",
		"备注": _first(row, "备注"),
	}


def _salary_grade_from_structure(version_code, salary_level):
	"""Resolve Yongxin's ``版本 + 薪资序号`` pair to one salary-grade record."""
	version_code = _text(version_code)
	level = cint(salary_level)
	if not version_code or level <= 0:
		return None
	version_name = frappe.db.get_value(SALARY_STRUCTURE_VERSION_DOCTYPE, {"structure_version": version_code}, "name")
	if not version_name:
		return None
	row = frappe.db.get_value(
		SALARY_GRADE_DOCTYPE,
		{"salary_structure_version": version_name, "salary_level": level},
		["name", "base_salary", "function_allowance", "full_salary"],
		as_dict=True,
	)
	if row:
		return row
	# Compatibility for records imported before the numeric salary-level field.
	return frappe.db.get_value(
		SALARY_GRADE_DOCTYPE,
		{"salary_structure_version": version_name, "job_grade": str(level)},
		["name", "base_salary", "function_allowance", "full_salary"],
		as_dict=True,
	)


def _salary_grade_from_unique_amounts(base_salary, function_allowance, payroll_month=""):
	"""Resolve a grade only when the imported fixed-pay pair is unambiguous.

	Some historical salary-change rows intentionally omit ``版本`` and ``薪资序号``.
	They must remain manual when their figures are non-standard or match multiple
	levels, but a unique active grade is a reliable existing binding and should
	not be discarded merely because the two identifier columns were blank.
	"""
	base_salary = flt(base_salary)
	function_allowance = flt(function_allowance)
	if not (base_salary or function_allowance):
		return None
	versions = _active_salary_structure_versions(payroll_month)
	if not versions:
		versions = frappe.get_all(
			SALARY_STRUCTURE_VERSION_DOCTYPE,
			filters={"status": "已启用"},
			fields=["name"],
			limit_page_length=1000,
		)
	version_names = [version.name for version in versions]
	if not version_names:
		return None
	candidates = [
		row
		for row in frappe.get_all(
			SALARY_GRADE_DOCTYPE,
			filters={"salary_structure_version": ["in", version_names]},
			fields=["name", "base_salary", "function_allowance", "full_salary"],
			limit_page_length=1000,
		)
		if flt(row.base_salary) == base_salary and flt(row.function_allowance) == function_allowance
	]
	return candidates[0] if len(candidates) == 1 else None


def _salary_grade_from_import_row(row, payroll_month=""):
	"""Resolve an explicit grade first, then a safe unique amount-based match."""
	structure_version = _text(_first(row, "薪资架构版本", "版本"))
	salary_level = cint(_first(row, "薪资序号"))
	grade = _salary_grade_from_structure(structure_version, salary_level)
	if grade:
		return grade, "薪资架构"
	explicit_grade = _text(_first(row, "薪资档位"))
	if explicit_grade:
		grade = frappe.db.get_value(
			SALARY_GRADE_DOCTYPE,
			explicit_grade,
			["name", "base_salary", "function_allowance", "full_salary"],
			as_dict=True,
		)
		if grade:
			return grade, "薪资档位"
	# A fully supplied, invalid pair is an import error and must not silently
	# turn into an amount match.  The fallback only fills genuinely blank data.
	if structure_version and salary_level:
		return None, ""
	grade = _salary_grade_from_unique_amounts(
		_first(row, "底薪"), _first(row, "职能津贴"), payroll_month
	)
	return (grade, "表内金额唯一匹配") if grade else (None, "")


def _salary_grade_from_matching_history(employee, base_salary, function_allowance, effective_date):
	"""Keep an employee's existing grade only when the imported pay still matches it."""
	if not employee or not effective_date:
		return None
	for change in frappe.get_all(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		filters={
			"employee": employee,
			"effective_date": ["<=", effective_date],
			"salary_grade": ["is", "set"],
			"status": ["!=", "已作废"],
		},
		fields=["salary_grade", "base_salary", "function_allowance"],
		order_by="effective_date desc, modified desc",
		limit_page_length=100,
	):
		if flt(change.base_salary) != flt(base_salary) or flt(change.function_allowance) != flt(function_allowance):
			continue
		grade = frappe.db.get_value(
			SALARY_GRADE_DOCTYPE,
			change.salary_grade,
			["name", "base_salary", "function_allowance", "full_salary"],
			as_dict=True,
		)
		if grade:
			return grade
	return None


def _reward_punishment_variable_type(row):
	classification = " ".join(_text(_first(row, header)) for header in ("奖惩类型", "奖惩条例", "标准", "主旨"))
	return "其他扣款" if any(token in classification for token in ("惩", "罚", "扣", "警告")) else "其他奖金"


def _raw_variable_entries(source_kind, row):
	if source_kind == "certificate_skill":
		return [("证书及多能工津贴", _amount_for_type(row, "证书及多能工津贴"))]
	if source_kind == "education":
		return [("学历补贴", _amount_for_type(row, "学历补贴"))]
	if source_kind == "housing_allowance":
		return [("住房补贴", _amount_for_type(row, "住房补贴"))]
	if source_kind == "attendance_bonus":
		return [("全勤奖", _amount_for_type(row, "全勤奖"))]
	if source_kind == "dormitory":
		# Housing subsidy has one dedicated monthly source.  The legacy dormitory
		# form contains a convenience "租房补贴" column, but importing it here as
		# well would make the same subsidy payable twice.  Keep this source limited
		# to its actual accommodation/water/electricity deduction.
		return [("水电费及扣款", _amount_for_type(row, "宿舍扣款"))]
	if source_kind == "reward_punishment":
		return [(_reward_punishment_variable_type(row), _amount_for_type(row, "其他奖金"))]
	if source_kind == "housing_fund":
		return [
			("公积金个人", flt(_first(row, "个人承担"))),
			("公积金公司", flt(_first(row, "公司承担"))),
		]
	if source_kind == "social_insurance":
		return [
			("社保个人", _amount_for_type(row, "社保个人")),
			("社保公司", flt(_first(row, "公司合计承担25.38%", "公司合计承担", "公司承担"))),
		]
	if source_kind == "continuing_service":
		return [("继续服务奖", _amount_for_type(row, "继续服务奖"))]
	return []


def _can_manage_payroll_rules():
	# Administrator is the platform's highest-privilege account.  It must not be
	# blocked merely because its role list was not hydrated in the browser session.
	if frappe.session.user == "Administrator":
		return True
	roles = set(frappe.get_roles(frappe.session.user))
	return bool({"Administrator", "System Manager", "HR Manager"} & roles)


def _require_payroll_master_manager():
	if not _can_manage_payroll_rules():
		frappe.throw(_("仅系统管理员或人事管理员可以维护薪资架构和员工定薪。"))


def _default_rule(rule_code):
	return next((rule for rule in DEFAULT_PAYROLL_RULES if rule["rule_code"] == rule_code), {})


def _rule_parameters(value):
	if isinstance(value, dict):
		return dict(value)
	if not value:
		return {}
	try:
		decoded = json.loads(value)
	except (TypeError, ValueError) as exc:
		frappe.throw(_("薪资规则参数 JSON 格式错误：{0}").format(exc))
	if not isinstance(decoded, dict):
		frappe.throw(_("薪资规则参数必须是 JSON 对象。"))
	return decoded


def _rule_parameter_errors(rule_code, parameters):
	if rule_code not in EXECUTABLE_PAYROLL_RULES:
		return []
	errors = []
	for key in EXECUTABLE_PAYROLL_RULES[rule_code]["parameters"]:
		if key not in parameters:
			errors.append(f"缺少参数 {key}")
	for key in ("standard_hours_divisor", "weekday", "weekend", "holiday", "deep_night_shift", "large_night_shift", "small_night_shift", "absenteeism_multiplier"):
		if key in parameters and flt(parameters[key]) <= 0:
			errors.append(f"参数 {key} 必须大于 0")
	for key in ("red_apples_per_record", "amount_per_apple"):
		if key in parameters and flt(parameters[key]) < 0:
			errors.append(f"参数 {key} 不能小于 0")
	if "red_apples_per_record" in parameters and flt(parameters["red_apples_per_record"]) != cint(parameters["red_apples_per_record"]):
		errors.append("参数 red_apples_per_record 必须是整数")
	if "thresholds" in parameters:
		thresholds = parameters["thresholds"]
		if not isinstance(thresholds, list) or not thresholds:
			errors.append("参数 thresholds 必须是非空数组")
		else:
			for item in thresholds:
				if not isinstance(item, (list, tuple)) or len(item) != 2:
					errors.append("thresholds 每项必须为 [缺勤上限, 全勤奖金额]")
					break
	if "ranges" in parameters and not isinstance(parameters["ranges"], list):
		errors.append("参数 ranges 必须是数组")
	if rule_code == "PAYROLL_SETTLEMENT_NIGHT_SHIFT":
		pattern = re.compile(r"^(?:[01]\d|2[0-3]):[0-5]\d$")
		tiers = (("深夜班", "deep_night_shift", True), ("大夜班", "large_night_shift", False), ("小夜班", "small_night_shift", False))
		ranges = []
		for label, key, required in tiers:
			start, end = str(parameters.get(f"{key}_start") or ""), str(parameters.get(f"{key}_end") or "")
			# 大夜班、小夜班可先沿用终稿次数；只有填写其中一个时间时，
			# 才要求成对启用该档位的时间匹配。深夜班是当前已启用的默认规则。
			if not start and not end and not required:
				continue
			if not pattern.fullmatch(start) or not pattern.fullmatch(end):
				errors.append(f"{label}需要同时设置上班和下班时间（HH:MM）")
				continue
			start_minutes, end_minutes = _clock_time_minutes(start), _clock_time_minutes(end)
			if start_minutes == end_minutes:
				errors.append(f"{label}的上下班时间不能相同")
				continue
			ranges.append((label, start_minutes, end_minutes))
		for index, (label, start, end) in enumerate(ranges):
			for other_label, other_start, other_end in ranges[index + 1:]:
				if _night_shift_ranges_overlap(start, end, other_start, other_end):
					errors.append(f"{label}与{other_label}的匹配时段重叠，请调整后再保存")
	return errors


def _rule_is_effective(rule, payroll_month):
	if not payroll_month:
		return True
	month_end = getdate(_month_end(payroll_month))
	if rule.effective_from and getdate(rule.effective_from) > month_end:
		return False
	if rule.effective_to and getdate(rule.effective_to) < month_end:
		return False
	return True


def _rule_version_sort_key(rule):
	"""Newest applicable start date wins; legacy rows are the oldest baseline."""
	return (getdate(rule.effective_from) if rule.get("effective_from") else getdate("1900-01-01"), str(rule.get("modified") or ""))


def _effective_rule_row(rule_code, payroll_month, company):
	"""Select one company rule version for a month without relying on row order."""
	rows = frappe.get_all(
		PAYROLL_RULE_DOCTYPE,
		filters={"company": company, "rule_code": rule_code},
		fields=["name", "rule_code", "rule_name", "status", "editable", "effective_from", "effective_to", "formula_expression", "parameters_json", "modified"],
		limit_page_length=500,
	)
	if not rows:
		# Read-only compatibility for installations whose original rules predate
		# company isolation. New saves never write these global records.
		rows = frappe.get_all(
			PAYROLL_RULE_DOCTYPE,
			filters={"company": ["is", "not set"], "rule_code": rule_code},
			fields=["name", "rule_code", "rule_name", "status", "editable", "effective_from", "effective_to", "formula_expression", "parameters_json", "modified"],
			limit_page_length=500,
		)
	applicable = [row for row in rows if _rule_is_effective(row, payroll_month)]
	return max(applicable, key=_rule_version_sort_key) if applicable else None


def _effective_rule_config(rule_code, payroll_month="", company="", allow_incomplete_night_times=False):
	"""Return one validated executable rule and its immutable calculation snapshot."""
	company = _require_company(company)
	default = _default_rule(rule_code)
	if not default:
		frappe.throw(_("未注册的薪资执行规则：{0}").format(rule_code))
	default_parameters = _rule_parameters(default.get("parameters_json"))
	rule = _effective_rule_row(rule_code, payroll_month, company)
	if not rule:
		errors = _rule_parameter_errors(rule_code, default_parameters)
		if errors and not allow_incomplete_night_times:
			frappe.throw(_("执行规则 {0} 参数无效：{1}").format(rule_code, "；".join(errors)))
		return {
			"rule_code": rule_code,
			"rule_name": default.get("rule_name"),
			"formula_expression": default.get("formula_expression"),
			"parameters": default_parameters,
			"source": f"{company} / 内置默认规则",
		}
	rule = frappe.get_doc(PAYROLL_RULE_DOCTYPE, rule.name)
	if rule.status != "已启用":
		frappe.throw(_("执行规则 {0} 当前不是已启用状态，不能生成薪资结算。").format(rule_code))
	if not _rule_is_effective(rule, payroll_month):
		frappe.throw(_("执行规则 {0} 不在薪资月份 {1} 的生效区间内。").format(rule_code, payroll_month))
	saved_parameters = _rule_parameters(rule.parameters_json)
	# Existing installations stored the single matching range under the former
	# ``large_night_*`` keys.  Move it to deep night on upgrade; the new large
	# and small tiers deliberately stay empty until HR configures their ranges.
	if rule_code == "PAYROLL_SETTLEMENT_NIGHT_SHIFT":
		if "deep_night_shift_start" not in saved_parameters and "large_night_shift_start" in saved_parameters:
			saved_parameters["deep_night_shift_start"] = saved_parameters.pop("large_night_shift_start")
			saved_parameters["deep_night_shift_end"] = saved_parameters.pop("large_night_shift_end", "")
	parameters = {**default_parameters, **{key: value for key, value in saved_parameters.items() if value not in (None, "")}}
	errors = _rule_parameter_errors(rule_code, parameters)
	if allow_incomplete_night_times and rule_code == "PAYROLL_SETTLEMENT_NIGHT_SHIFT":
		errors = [error for error in errors if not error.startswith("请设置")]
	if errors:
		frappe.throw(_("执行规则 {0} 参数无效：{1}").format(rule_code, "；".join(errors)))
	return {
		"name": rule.name,
		"rule_code": rule_code,
		"rule_name": rule.rule_name,
		"formula_expression": rule.formula_expression or default.get("formula_expression"),
		"parameters": parameters,
		"company": company,
		"source": f"{company} / 规则中心" if rule.company else "历史全局规则（待迁移）",
	}


def _payroll_calculation_rules(company, payroll_month):
	return {
		rule_code: _effective_rule_config(rule_code, payroll_month, company)
		for rule_code in EXECUTABLE_PAYROLL_RULES
	}


def _night_shift_formula_expression(parameters):
	"""Build the one controlled night-shift formula from the saved rule rates."""
	return "[深夜班次数] * {0} + [大夜班次数] * {1} + [小夜班次数] * {2}".format(
		flt(parameters.get("deep_night_shift")),
		flt(parameters.get("large_night_shift")),
		flt(parameters.get("small_night_shift")),
	)


def _attendance_rule_formula_overrides(calculation_rules):
	"""Build the fixed calculator formulas from the business-rule parameters.

	The attendance-rule page deliberately exposes rates and divisors rather than
	free-form formulas.  Keeping the generated expressions here makes those
	settings executable while preserving the formula engine's restricted syntax.
	"""
	absence_rule = calculation_rules["PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION"]
	overtime_rule = calculation_rules["PAYROLL_SETTLEMENT_OVERTIME_PAY"]
	absence_divisor = _rule_number(absence_rule, "standard_hours_divisor", PAYROLL_STANDARD_HOURS_DIVISOR)
	absenteeism_multiplier = _rule_number(absence_rule, "absenteeism_multiplier", 3)
	overtime_divisor = _rule_number(overtime_rule, "standard_hours_divisor", PAYROLL_STANDARD_HOURS_DIVISOR)
	weekday_rate = _rule_number(overtime_rule, "weekday", 1.5)
	weekend_rate = _rule_number(overtime_rule, "weekend", 2)
	holiday_rate = _rule_number(overtime_rule, "holiday", 3)

	def number(value):
		return format(flt(value), "g")

	return {
		"full_salary_hourly_rate": {
			"expression": "ROUND([薪资小计] / {0}, 8)".format(number(absence_divisor)),
			"description": "全薪小时单价按当前缺勤扣款标准计薪工时计算。",
			"source": absence_rule["source"],
		},
		"base_salary_hourly_rate": {
			"expression": "ROUND([底薪] / {0}, 8)".format(number(overtime_divisor)),
			"description": "底薪小时单价按当前加班工资标准计薪工时计算。",
			"source": overtime_rule["source"],
		},
		"absenteeism_deduction": {
			"expression": "ROUND([全薪时薪] * [旷工工时] * {0}, 2)".format(number(absenteeism_multiplier)),
			"description": "旷工扣款按当前旷工工时倍率计算。",
			"source": absence_rule["source"],
		},
		"weekday_overtime_pay": {
			"expression": "ROUND([底薪时薪] * [平日加班时数] * {0}, 2)".format(number(weekday_rate)),
			"description": "平日加班费按当前考勤计薪规则倍率计算。",
			"source": overtime_rule["source"],
		},
		"weekend_overtime_pay": {
			"expression": "ROUND([底薪时薪] * [调整后周末加班] * {0}, 2)".format(number(weekend_rate)),
			"description": "周末加班费按当前考勤计薪规则倍率计算。",
			"source": overtime_rule["source"],
		},
		"holiday_overtime_pay": {
			"expression": "ROUND([底薪时薪] * [节假日加班时数] * {0}, 2)".format(number(holiday_rate)),
			"description": "节假日加班费按当前考勤计薪规则倍率计算。",
			"source": overtime_rule["source"],
		},
	}


def _apply_attendance_rule_parameters(payroll_formulas, calculation_rules):
	"""Inject auditable, validated attendance rates into the fixed formula set.

	The formula engine remains restricted to registered expressions; the rule
	centre changes only approved parameters.  This prevents a saved rule from
	becoming display-only while preserving an explicit formula trace in every
	generated settlement.
	"""
	night_rule = calculation_rules["PAYROLL_SETTLEMENT_NIGHT_SHIFT"]
	formula_overrides = _attendance_rule_formula_overrides(calculation_rules)
	formulas = []
	for formula in payroll_formulas:
		item = dict(formula)
		override = formula_overrides.get(item["output_field"])
		if override:
			item.update(override)
			compile_formula(item["expression"])
		elif item["output_field"] == "night_shift_allowance":
			item["expression"] = _night_shift_formula_expression(night_rule["parameters"])
			item["description"] = "夜班津贴按当前公司夜班规则计算。"
			item["source"] = night_rule["source"]
			compile_formula(item["expression"])
		formulas.append(item)
	return formulas


def _formula_rule_code(output_field):
	return f"FORMULA_{str(output_field or '').upper()}"


def _effective_payroll_formulas(company, payroll_month=""):
	"""Return the ordered built-in formula set with company overrides applied."""
	company = _require_company(company)
	rows = frappe.get_all(
		PAYROLL_RULE_DOCTYPE,
		filters={"company": company, "rule_code": ["like", "FORMULA_%"], "status": "已启用"},
		fields=["name", "rule_code", "output_field", "formula_expression", "formula_version", "effective_from", "effective_to", "modified"],
		order_by="modified desc",
		limit_page_length=500,
	)
	overrides = {}
	for row in rows:
		if not row.output_field or row.output_field in overrides or not _rule_is_effective(row, payroll_month):
			continue
		overrides[row.output_field] = row
	formulas = []
	for index, template in enumerate(FORMULA_TEMPLATES, start=1):
		formula = dict(template)
		formula["order"] = index
		formula["version"] = 1
		formula["source"] = f"{company} / 内置公式模板"
		override = overrides.get(template["output_field"])
		if override:
			formula.update(
				{
					"expression": override.formula_expression,
					"version": int(override.formula_version or 1),
					"source": f"{company} / {override.name}",
					"rule_name": override.name,
				}
			)
		compile_formula(formula["expression"])
		formulas.append(formula)
	return formulas


def _formula_catalog_row(formula, company, order):
	field = FIELD_BY_NAME[formula["output_field"]]
	_, dependencies = compile_formula(formula["expression"])
	return {
		**formula,
		"company": company,
		"order": order,
		"rule_code": _formula_rule_code(formula["output_field"]),
		"output_label": field["label"],
		"dependencies": [FIELD_BY_NAME[name]["label"] for name in dependencies],
		"status": "已启用",
	}


@frappe.whitelist()
def get_payroll_formula_catalog(company: str, payroll_month: str = ""):
	company = _require_company(company)
	formulas = _effective_payroll_formulas(company, payroll_month)
	return {
		"company": company,
		"payroll_month": payroll_month,
		"fields": FIELD_DEFINITIONS,
		"functions": FUNCTIONS,
		"formulas": [_formula_catalog_row(formula, company, index) for index, formula in enumerate(formulas, start=1)],
		"groups": ["薪资字段", "考勤字段", "月度变量", "计算结果"],
	}


def _effective_payroll_field_mappings():
	"""Merge the controlled defaults with explicit field-mapping overrides."""
	persisted_rows = frappe.get_all(
		PAYROLL_FIELD_MAPPING_DOCTYPE,
		fields=["*"],
		limit_page_length=1000,
	)
	persisted_by_code = {row.get("mapping_code"): row for row in persisted_rows if row.get("mapping_code")}
	mappings = []
	default_codes = set()
	for default in PAYROLL_SETTLEMENT_FIELD_MAPPINGS:
		mapping = dict(default)
		mapping["system_doctype"] = mapping.get("system_doctype") or PAYROLL_SETTLEMENT_DOCTYPE
		mapping["required_for_settlement"] = flt(mapping.get("required_for_settlement", 1))
		mapping["status"] = mapping.get("status") or "已启用"
		stored = persisted_by_code.get(mapping["mapping_code"])
		if stored:
			mapping.update(dict(stored))
			mapping["mapping_origin"] = "公司配置"
		else:
			mapping["mapping_origin"] = "系统默认"
		mappings.append(mapping)
		default_codes.add(mapping["mapping_code"])
	for row in persisted_rows:
		if row.get("mapping_code") not in default_codes:
			mapping = dict(row)
			mapping["mapping_origin"] = "公司自定义"
			mappings.append(mapping)
	return mappings


@frappe.whitelist()
def get_payroll_calculation_audit(company: str, payroll_month: str = "", attendance_lock_version: str = ""):
	"""Audit the bridge from formula configuration to settlement records.

	The response deliberately distinguishes *structural* validation (a formula
	compiles, its result has a settlement field, and every spreadsheet mapping has
	a valid target) from *execution* validation (an existing settlement row contains
	the formula trace).  A new month with no settlement output should be ready, not
	incorrectly marked failed merely because it has not been generated yet.
	"""
	company = _require_company(company)
	blockers = []
	warnings = []
	formula_rows = []
	try:
		formulas = _effective_payroll_formulas(company, payroll_month)
	except Exception as exc:
		formulas = []
		blockers.append("公司薪资公式无法加载：{0}".format(exc))

	formula_positions = {formula.get("output_field"): index for index, formula in enumerate(formulas)}
	for index, formula in enumerate(formulas):
		output_field = formula.get("output_field")
		row = {
			"output_field": output_field,
			"output_label": FIELD_BY_NAME.get(output_field, {}).get("label", output_field),
			"expression": formula.get("expression") or "",
			"dependencies": [],
			"participates_in_settlement": False,
			"valid": 0,
			"message": "",
		}
		try:
			_, dependencies = compile_formula(row["expression"])
			row["dependencies"] = dependencies
			future_dependencies = [name for name in dependencies if name in formula_positions and formula_positions[name] >= index]
			unbound_dependencies = [name for name in dependencies if name not in PAYROLL_FORMULA_CONTEXT_FIELDS and name not in formula_positions]
			if future_dependencies:
				raise FormulaError("引用的计算结果必须排在当前结果之前：{0}".format("、".join(FIELD_BY_NAME[name]["label"] for name in future_dependencies)))
			if unbound_dependencies:
				raise FormulaError("缺少结算上下文的字段：{0}".format("、".join(unbound_dependencies)))
			if not _doctype_has_field(PAYROLL_SETTLEMENT_DOCTYPE, output_field):
				raise FormulaError("薪资结算记录没有字段：{0}".format(output_field))
			if output_field not in PAYROLL_SETTLEMENT_FORMULA_OUTPUT_FIELDS:
				raise FormulaError("该计算结果未写入薪资结算记录：{0}".format(output_field))
			row["participates_in_settlement"] = True
			row["valid"] = 1
			row["message"] = "公式可执行，结果会写入薪资结算"
		except Exception as exc:
			row["message"] = str(exc)
			blockers.append("公式 {0}：{1}".format(row["output_label"], row["message"]))
		formula_rows.append(row)

	mapping_rows = []
	for mapping in _effective_payroll_field_mappings():
		target_doctype = mapping.get("system_doctype") or PAYROLL_SETTLEMENT_DOCTYPE
		target_field = mapping.get("system_field") or ""
		row = {
			"mapping_code": mapping.get("mapping_code"),
			"excel_column": mapping.get("excel_column"),
			"excel_label": mapping.get("excel_label"),
			"system_doctype": target_doctype,
			"system_field": target_field,
			"source_module": mapping.get("source_module"),
			"mapping_origin": mapping.get("mapping_origin"),
			"valid": 0,
			"message": "",
		}
		if mapping.get("status") != "已启用":
			row["message"] = "字段映射未启用"
		elif not _doctype_exists(target_doctype):
			row["message"] = "目标 DocType 不存在：{0}".format(target_doctype)
		elif not target_field or not _doctype_has_field(target_doctype, target_field):
			row["message"] = "目标字段不存在：{0}.{1}".format(target_doctype, target_field or "（未填写）")
		else:
			row["valid"] = 1
			row["message"] = "映射有效"
		if not row["valid"] and flt(mapping.get("required_for_settlement", 1)):
			blockers.append("字段映射 {0}：{1}".format(row["excel_label"] or row["mapping_code"], row["message"]))
		mapping_rows.append(row)

	settlement_filters = {"company": company}
	if payroll_month:
		settlement_filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		settlement_filters["attendance_lock_version"] = attendance_lock_version
	settlements = _safe_get_all(
		PAYROLL_SETTLEMENT_DOCTYPE,
		filters=settlement_filters,
		fields=["name", "source_trace_json", "attendance_lock_version"],
		limit_page_length=100000,
	)
	trace_failures = []
	if settlements:
		expected_outputs = set(formula_positions)
		for settlement in settlements:
			try:
				trace = json.loads(settlement.get("source_trace_json") or "{}")
				actual_outputs = {item.get("output_field") for item in trace.get("formula_trace", [])}
			except (TypeError, ValueError):
				actual_outputs = set()
			missing_outputs = expected_outputs - actual_outputs
			if missing_outputs:
				trace_failures.append("{0} 缺少公式追溯：{1}".format(settlement.get("name"), "、".join(sorted(missing_outputs))))
		if trace_failures:
			blockers.append("已有结算记录未完整保留公式执行追溯。")
			warnings.extend(trace_failures[:10])
		actual_execution = {
			"available": 1,
			"record_count": len(settlements),
			"trace_record_count": len(settlements) - len(trace_failures),
			"valid": 0 if trace_failures else 1,
			"message": "已核验 {0} 条结算记录：{1}".format(len(settlements), "全部包含本月公式追溯" if not trace_failures else "存在缺失的公式追溯"),
		}
	else:
		actual_execution = {
			"available": 0,
			"record_count": 0,
			"trace_record_count": 0,
			"valid": 1,
			"message": "尚无本月结算记录；当前已完成公式与字段映射的结构核查。",
		}
		warnings.append("尚未生成本月结算记录，无法进行实际公式追溯比对。")

	return {
		"company": company,
		"payroll_month": payroll_month,
		"attendance_lock_version": attendance_lock_version,
		"valid": not blockers,
		"summary": {
			"formula_total": len(formula_rows),
			"formula_valid": sum(1 for row in formula_rows if row["valid"]),
			"formula_participating": sum(1 for row in formula_rows if row["participates_in_settlement"]),
			"mapping_total": len(mapping_rows),
			"mapping_valid": sum(1 for row in mapping_rows if row["valid"]),
		},
		"formulas": formula_rows,
		"mappings": mapping_rows,
		"actual_execution": actual_execution,
		"blockers": blockers,
		"warnings": warnings,
	}


@frappe.whitelist()
def validate_payroll_formula(company: str, output_field: str, expression: str, sample_values: str = ""):
	company = _require_company(company)
	if output_field not in FIELD_BY_NAME or FIELD_BY_NAME[output_field]["group"] != "计算结果":
		frappe.throw(_("请选择系统允许的计算结果项目"))
	try:
		_, dependencies = compile_formula(expression)
		if output_field in dependencies:
			raise FormulaError("结果项目不能引用自身")
		if isinstance(sample_values, str):
			sample_values = json.loads(sample_values or "{}")
		value, _ = evaluate_formula(expression, sample_values or {})
	except (FormulaError, ValueError, TypeError) as exc:
		return {"valid": 0, "message": str(exc), "result": None, "dependencies": []}
	return {
		"valid": 1,
		"message": "公式校验通过，可参与薪资试算",
		"result": _money(value),
		"dependencies": [{"fieldname": name, "label": FIELD_BY_NAME[name]["label"]} for name in dependencies],
	}


@frappe.whitelist()
def upsert_payroll_formula(**kwargs: object):
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有维护薪资公式的权限"))
	data = dict(kwargs)
	company = _require_company(data.get("company"))
	output_field = data.get("output_field")
	expression = data.get("formula_expression") or data.get("expression")
	validation = validate_payroll_formula(company, output_field, expression)
	if not validation.get("valid"):
		frappe.throw(_("公式无效：{0}").format(validation.get("message")))
	field = FIELD_BY_NAME[output_field]
	rule_code = _formula_rule_code(output_field)
	name = frappe.db.get_value(PAYROLL_RULE_DOCTYPE, {"company": company, "rule_code": rule_code}, "name")
	current_version = int(frappe.db.get_value(PAYROLL_RULE_DOCTYPE, name, "formula_version") or 0) if name else 0
	values = {
		"company": company,
		"rule_code": rule_code,
		"rule_name": field["label"],
		"rule_category": data.get("rule_category") or "薪资结算",
		"rule_scope": data.get("rule_scope") or "薪资结算表",
		"status": data.get("status") or "已启用",
		"editable": 1,
		"effective_from": data.get("effective_from"),
		"effective_to": data.get("effective_to"),
		"calculation_mode": "公式",
		"output_field": output_field,
		"formula_expression": expression,
		"formula_version": current_version + 1,
		"parameters_json": "{}",
		"rule_text": data.get("rule_text") or data.get("description"),
		"source_file": data.get("source_file") or "公司薪资公式模板",
		"source_sheet": data.get("source_sheet"),
		"source_cell": data.get("source_cell"),
		"last_reviewed_by": frappe.session.user,
		"last_reviewed_on": now_datetime(),
	}
	if name:
		doc = frappe.get_doc(PAYROLL_RULE_DOCTYPE, name)
		if doc.company != company:
			frappe.throw(_("不能跨公司修改薪资公式"))
		doc.update(values)
		doc.save(ignore_permissions=True)
	else:
		doc = frappe.get_doc({"doctype": PAYROLL_RULE_DOCTYPE, **values})
		doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "formula_version": doc.formula_version, "validation": validation}


@frappe.whitelist()
def ensure_default_payroll_formulas(company: str, force: int = 0):
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有维护薪资公式的权限"))
	company = _require_company(company)
	created, updated = 0, 0
	for template in FORMULA_TEMPLATES:
		name = frappe.db.get_value(PAYROLL_RULE_DOCTYPE, {"company": company, "rule_code": _formula_rule_code(template["output_field"])}, "name")
		if name and not flt(force):
			continue
		upsert_payroll_formula(
			company=company,
			output_field=template["output_field"],
			formula_expression=template["expression"],
			rule_category="薪资结算",
			description=template["description"],
			source_file="5.2人资考勤.xlsx",
			source_sheet="薪资结算表",
		)
		updated += 1 if name else 0
		created += 0 if name else 1
	return {"company": company, "created": created, "updated": updated}


@frappe.whitelist()
def create_payroll_formula_template_file(company: str):
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from frappe.utils.file_manager import save_file

	company = _require_company(company)
	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "计薪公式"
	headers = ["结果项目", "计算公式", "规则说明", "适用范围", "生效开始", "生效结束", "状态"]
	sheet.append(headers)
	for formula in _effective_payroll_formulas(company):
		sheet.append([FIELD_BY_NAME[formula["output_field"]]["label"], formula["expression"], formula.get("description"), "薪资结算表", "", "", "已启用"])
	fill = PatternFill("solid", fgColor="D9EAF7")
	for cell in sheet[1]:
		cell.fill = fill
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center")
	for index, width in enumerate([24, 92, 56, 22, 16, 16, 12], start=1):
		sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
	fields = workbook.create_sheet("可用字段")
	fields.append(["分组", "显示名称", "系统字段", "唯一来源", "公式写法"])
	for item in FIELD_DEFINITIONS:
		fields.append([item["group"], item["label"], item["fieldname"], item["source"], f'[{item["label"]}]'])
	for cell in fields[1]:
		cell.fill = fill
		cell.font = Font(bold=True)
	functions = workbook.create_sheet("可用函数")
	functions.append(["函数", "用途", "写法"])
	for item in FUNCTIONS:
		functions.append([item["name"], item["label"], item["signature"]])
	for cell in functions[1]:
		cell.fill = fill
		cell.font = Font(bold=True)
	output = BytesIO()
	workbook.save(output)
	file_doc = save_file(f"{company}-计薪公式导入模板-{datetime.today().strftime('%Y%m%d%H%M%S')}.xlsx", output.getvalue(), None, None, is_private=1)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name}


def _formula_workbook_rows(file_url):
	workbook = _load_workbook(file_url)
	sheet = _matching_sheet(workbook, "计薪公式")
	if not sheet:
		frappe.throw(_("未找到“计薪公式”工作表"))
	rows = list(sheet.iter_rows(values_only=True))
	if not rows:
		return []
	headers = [_text(value) for value in rows[0]]
	result = []
	for values in rows[1:]:
		row = {headers[index]: _text(value) for index, value in enumerate(values) if index < len(headers) and headers[index]}
		if row.get("结果项目") or row.get("计算公式"):
			result.append(row)
	return result


# Legacy settlement sheets are written for people to read, so their headers are
# not always identical to the internal field labels.  Keep this translation at
# the import boundary: the calculation engine still receives only controlled
# business fields such as ``[标准工时]`` and ``[基本出勤工时]``.
PAYROLL_FORMULA_HEADER_ALIASES = {
	"base_salary": ("底薪",),
	"function_allowance": ("职能津贴", "职务津贴"),
	"certificate_skill_allowance": ("证书及多能工津贴", "证书、多能工津贴", "证书津贴", "多能工津贴"),
	"standard_hours": ("标准工时",),
	"basic_attendance_hours": ("基本出勤工时", "基本出勤", "调整前基本出勤工时"),
	"raw_weekend_overtime_hours": ("调整前周末加班", "调整前周末加班工时"),
	"weekday_overtime_hours": ("平日加班时数", "平日加班工时"),
	"holiday_overtime_hours": ("节假日加班时数", "节假日加班工时"),
	"deep_night_shift_count": ("深夜班次数", "深夜班"),
	"large_night_shift_count": ("大夜班次数", "大夜班(5)", "大夜班"),
	"small_night_shift_count": ("小夜班次数", "小夜班(2)", "小夜班"),
	"absenteeism_hours": ("旷工工时",),
	"proposal_improvement_bonus": ("提案改善奖", "提案改善奖金额"),
	"apple_reward_amount": ("红绿苹果", "红绿苹果金额"),
	"full_attendance_bonus": ("全勤奖",),
	"housing_subsidy": ("住房补贴",),
	"education_subsidy": ("学历补贴",),
	"other_bonus": ("其他奖金",),
	"production_bonus": ("生产奖",),
	"late_full_attendance_deduction": ("迟到及全勤奖扣款", "迟到金额及全勤奖扣款", "迟到金额+全勤奖扣款"),
	"other_deduction": ("其他扣款",),
	"social_security_personal": ("社保个人", "保险基金员工负担额"),
	"housing_fund_personal": ("公积金个人", "住房公积金"),
	"paid_proposal_birthday_welfare": ("已发福利", "提案改善奖&生日福利金（已发）"),
	"continuing_service_bonus": ("继续服务奖",),
	"income_tax": ("所得税", "所得税代扣款"),
	"year_end_bonus_tax": ("年终奖所得税",),
	"utilities_deduction": ("水电费及扣款", "水电费及扣款金额"),
	"manual_social_security_company": ("社保公司手工金额",),
	"manual_housing_fund_company": ("公积金公司手工金额",),
	"salary_subtotal": ("薪资小计", "全薪"),
	"missing_hours": ("缺勤工时",),
	"adjusted_absence_hours": ("调整后缺勤工时",),
	"weekend_overtime_hours": ("调整后周末加班", "调整后周末加班工时"),
	"full_salary_hourly_rate": ("全薪时薪", "全薪小时单价"),
	"base_salary_hourly_rate": ("底薪时薪", "底薪小时单价"),
	"absence_deduction_amount": ("缺勤扣款", "缺勤工时对应的扣除金额"),
	"weekday_overtime_pay": ("平日加班费", "加班费/平日"),
	"weekend_overtime_pay": ("周末加班费", "加班费/周末"),
	"holiday_overtime_pay": ("节假日加班费", "加班费/节假日"),
	"overtime_pay_total": ("加班费小计",),
	"night_shift_allowance": ("夜班津贴",),
	"subsidy_bonus_total": ("补贴小计", "全勤奖,住房学历补贴"),
	"bonus_total": ("奖金小计",),
	"absenteeism_deduction": ("旷工扣款",),
	"punishment_total": ("惩处小计",),
	"attendance_wage": ("出勤工资",),
	"gross_pay": ("应付工资",),
	"taxable_salary": ("计税工资",),
	"net_pay": ("实发工资",),
	"social_security_company": ("社保公司", "保险基金公司负担额"),
	"housing_fund_company": ("公积金公司", "住房公积金公司负担"),
	"company_cost_total": ("公司实际负担", "公司实际负担总计"),
	"export_tax_adjusted_net_pay": ("导出校验工资",),
}


def _normalise_formula_header(value):
	return re.sub(r"[\s\n\r/／,，、&＋+()（）]", "", _text(value)).lower()


PAYROLL_FORMULA_FIELD_BY_HEADER = {
	_normalise_formula_header(alias): fieldname
	for fieldname, aliases in PAYROLL_FORMULA_HEADER_ALIASES.items()
	for alias in aliases
}


def _load_formula_workbook(file_url):
	from openpyxl import load_workbook

	return load_workbook(BytesIO(_get_file_content(file_url)), data_only=False, read_only=False)


def _formula_field_from_excel_header(value):
	return PAYROLL_FORMULA_FIELD_BY_HEADER.get(_normalise_formula_header(value))


def _excel_column_headers_for_formula_row(sheet, formula_row):
	"""Find the nearest recognised header above each formula cell."""
	headers = {}
	for column in range(1, min(sheet.max_column or 0, 220) + 1):
		for row in range(formula_row - 1, max(formula_row - 8, 0), -1):
			fieldname = _formula_field_from_excel_header(sheet.cell(row=row, column=column).value)
			if fieldname:
				headers[column] = fieldname
				break
	return headers


def _excel_formula_to_business_expression(expression, headers):
	"""Convert ``=K6-I6`` to ``[缺勤工时] - [标准工时]`` using sheet headers."""
	from openpyxl.utils.cell import column_index_from_string, get_column_letter

	formula = _text(expression).strip()
	if formula.startswith("="):
		formula = formula[1:]
	if not formula:
		raise FormulaError("Excel 单元格没有公式")

	def field_token(column):
		fieldname = headers.get(column_index_from_string(column.replace("$", "").upper()))
		if not fieldname:
			raise FormulaError("引用列没有可识别的薪资字段：{0}".format(column.replace("$", "").upper()))
		return "[{0}]".format(FIELD_BY_NAME[fieldname]["label"])

	def replace_range(match):
		start, _, end = match.groups()
		start_index = column_index_from_string(start.replace("$", "").upper())
		end_index = column_index_from_string(end.replace("$", "").upper())
		if end_index < start_index:
			start_index, end_index = end_index, start_index
		tokens = [field_token(get_column_letter(index)) for index in range(start_index, end_index + 1)]
		return "(" + " + ".join(tokens) + ")"

	# Ranges must be expanded before standalone references.  The formula engine
	# deliberately treats a range as explicit additions, which keeps each input
	# visible in the card editor and avoids hidden Excel-only range semantics.
	formula = re.sub(r"\$?([A-Z]{1,3})\$?\d+\s*:\s*\$?([A-Z]{1,3})\$?\d+", replace_range, formula, flags=re.I)
	formula = re.sub(r"(?<![A-Z0-9_])(?:'[^']+'!)?\$?([A-Z]{1,3})\$?\d+", lambda match: field_token(match.group(1)), formula, flags=re.I)
	formula = formula.replace(";", ",")
	return formula


def _source_payroll_formula_rows(file_url):
	workbook = _load_formula_workbook(file_url)
	candidates = []
	for sheet in workbook.worksheets:
		for row in range(1, min(sheet.max_row or 0, 160) + 1):
			headers = _excel_column_headers_for_formula_row(sheet, row)
			if not headers:
				continue
			for column in range(1, min(sheet.max_column or 0, 220) + 1):
				cell = sheet.cell(row=row, column=column)
				if not (isinstance(cell.value, str) and cell.value.startswith("=")):
					continue
				output_field = headers.get(column)
				if not output_field or FIELD_BY_NAME[output_field]["group"] != "计算结果":
					continue
				try:
					business_expression = _excel_formula_to_business_expression(cell.value, headers)
					message = ""
				except FormulaError as exc:
					business_expression = ""
					message = str(exc)
				candidates.append({
					"结果项目": FIELD_BY_NAME[output_field]["label"],
					"output_field": output_field,
					"Excel公式": cell.value,
					"计算公式": business_expression,
					"source_sheet": sheet.title,
					"source_cell": cell.coordinate,
					"_column": column,
					"_row": row,
					"_message": message,
				})
	if not candidates:
		return []
	# A monthly settlement sheet has one calculation row with the most recognised
	# output formulas.  Ignore sample rows below it to avoid duplicate imports.
	best_sheet, best_row = max(
		((sheet, row) for sheet, row in {(item["source_sheet"], item["_row"]) for item in candidates}),
		key=lambda key: sum(1 for item in candidates if (item["source_sheet"], item["_row"]) == key),
	)
	selected = []
	seen_fields = set()
	for item in candidates:
		if (item["source_sheet"], item["_row"]) != (best_sheet, best_row) or item["output_field"] in seen_fields:
			continue
		seen_fields.add(item["output_field"])
		item.pop("_column", None)
		selected.append(item)
	return selected


@frappe.whitelist()
def preview_payroll_formula_source_workbook(file_url: str, company: str):
	company = _require_company(company)
	rows = _source_payroll_formula_rows(file_url)
	if not rows:
		# The existing, system-generated template remains importable.
		return {**preview_payroll_formula_workbook(file_url, company), "source_type": "公式模板"}
	for row in rows:
		validation = validate_payroll_formula(company, row["output_field"], row.get("计算公式")) if not row.get("_message") else {"valid": 0, "message": row["_message"]}
		row.update(validation)
		row.pop("_row", None)
		row.pop("_message", None)
	return {
		"company": company,
		"source_type": "Excel薪资结算表",
		"row_count": len(rows),
		"valid_count": sum(1 for row in rows if row.get("valid")),
		"rows": rows,
	}


@frappe.whitelist()
def import_payroll_formula_source_workbook(file_url: str, company: str):
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有导入薪资公式的权限"))
	preview = preview_payroll_formula_source_workbook(file_url, company)
	if preview.get("source_type") == "公式模板":
		return import_payroll_formula_workbook(file_url, company)
	invalid = [row for row in preview["rows"] if not row.get("valid")]
	if invalid:
		frappe.throw(_("Excel 公式存在 {0} 项错误，请先修正。首个错误：{1}").format(len(invalid), invalid[0].get("message")))
	for row in preview["rows"]:
		upsert_payroll_formula(
			company=company,
			output_field=row["output_field"],
			formula_expression=row["计算公式"],
			description="由 Excel {0}!{1} 转换：{2}".format(row["source_sheet"], row["source_cell"], row["Excel公式"]),
			rule_scope="薪资结算表",
			status="已启用",
			source_file=file_url,
			source_sheet=row["source_sheet"],
			source_cell=row["source_cell"],
		)
	return {"imported": len(preview["rows"]), "company": preview["company"], "source_type": preview["source_type"]}


@frappe.whitelist()
def preview_payroll_formula_workbook(file_url: str, company: str):
	company = _require_company(company)
	rows = []
	for row in _formula_workbook_rows(file_url):
		field = next((item for item in FIELD_DEFINITIONS if item["label"] == row.get("结果项目") and item["group"] == "计算结果"), None)
		validation = validate_payroll_formula(company, field["fieldname"] if field else "", row.get("计算公式"))
		rows.append({**row, "output_field": field["fieldname"] if field else "", **validation})
	return {"company": company, "row_count": len(rows), "valid_count": sum(1 for row in rows if row.get("valid")), "rows": rows}


@frappe.whitelist()
def import_payroll_formula_workbook(file_url: str, company: str):
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有导入薪资公式的权限"))
	preview = preview_payroll_formula_workbook(file_url, company)
	invalid = [row for row in preview["rows"] if not row.get("valid")]
	if invalid:
		frappe.throw(_("公式导入存在 {0} 行错误，请先修正预览结果。首个错误：{1}").format(len(invalid), invalid[0].get("message")))
	for row in preview["rows"]:
		upsert_payroll_formula(
			company=company,
			output_field=row["output_field"],
			formula_expression=row.get("计算公式"),
			description=row.get("规则说明"),
			rule_scope=row.get("适用范围"),
			effective_from=row.get("生效开始"),
			effective_to=row.get("生效结束"),
			status=row.get("状态") or "已启用",
			source_file=file_url,
			source_sheet="计薪公式",
		)
	return {"imported": len(preview["rows"]), "company": company}


def _rule_number(rule, key, fallback=0):
	return flt((rule or {}).get("parameters", {}).get(key, fallback)) or flt(fallback)


def _full_attendance_bonus(attendance, rule):
	"""Apply the confirmed threshold table unless a monthly variable overrides it."""
	parameters = rule.get("parameters", {})
	absence_basis = max(
		flt(getattr(attendance, "standard_hours", 0))
		- flt(getattr(attendance, "actual_attendance_hours", 0))
		- flt(getattr(attendance, "rest_leave_hours", 0))
		+ flt(getattr(attendance, "sick_leave_hours", 0)) * 0.5,
		0,
	)
	thresholds = sorted(parameters.get("thresholds") or [], key=lambda item: flt(item[0]))
	for maximum, amount in thresholds:
		if absence_basis <= flt(maximum):
			return flt(amount), absence_basis
	return 0, absence_basis


def _company_social_security_from_rule(personal_amount, rule):
	amount = flt(personal_amount)
	if amount <= 0:
		return 0
	ranges = list((rule or {}).get("parameters", {}).get("ranges") or [])
	# Existing data uses [minimum, maximum, company_amount]. Resolve exact and
	# narrow ranges before broad historical fallback ranges.
	def width(item):
		if not isinstance(item, (list, tuple)) or len(item) < 3:
			return float("inf")
		lower, upper = item[0], item[1]
		if upper in (None, ""):
			return float("inf")
		return max(flt(upper) - flt(lower), 0)
	for item in sorted(ranges, key=width):
		if not isinstance(item, (list, tuple)) or len(item) < 3:
			continue
		lower, upper, company_amount = item[0], item[1], item[2]
		if amount >= flt(lower) and (upper in (None, "") or amount <= flt(upper)):
			return flt(company_amount)
	return 0


def _rule_doc_values(rule, company):
	values = dict(rule)
	values["company"] = _require_company(company)
	values["parameters_json"] = json.dumps(values.get("parameters_json") or {}, ensure_ascii=False, default=str)
	values.setdefault("status", "已启用")
	values.setdefault("editable", 1)
	values.setdefault("last_reviewed_by", frappe.session.user)
	values.setdefault("last_reviewed_on", now_datetime())
	return values


@frappe.whitelist()
def can_edit_payroll_rules():
	return _can_manage_payroll_rules()


@frappe.whitelist()
def ensure_default_payroll_rules(company: str, force: int = 0):
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有维护薪资规则的权限"))
	company = _require_company(company)
	created = []
	updated = []
	for rule in DEFAULT_PAYROLL_RULES:
		existing = frappe.db.get_value(PAYROLL_RULE_DOCTYPE, {"company": company, "rule_code": rule["rule_code"]}, "name")
		values = _rule_doc_values(rule, company)
		if existing:
			if not flt(force):
				continue
			doc = frappe.get_doc(PAYROLL_RULE_DOCTYPE, existing)
			doc.update(values)
			doc.save(ignore_permissions=True)
			updated.append(doc.name)
			continue
		doc = frappe.get_doc({"doctype": PAYROLL_RULE_DOCTYPE, **values})
		doc.insert(ignore_permissions=True)
		created.append(doc.name)
	frappe.db.commit()
	return {"company": company, "created": len(created), "updated": len(updated)}


@frappe.whitelist()
def list_payroll_rules(company: str, rule_category: str = "", status: str = "", page_length: int = 200, payroll_month: str = ""):
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month) if payroll_month else ""
	filters = {"company": company}
	if rule_category:
		filters["rule_category"] = rule_category
	if status:
		filters["status"] = status
	rows = frappe.get_all(
		PAYROLL_RULE_DOCTYPE,
		filters=filters,
		fields=["*"],
		order_by="rule_category asc, rule_code asc",
		limit_page_length=int(page_length or 200),
	)
	if payroll_month:
		by_code = defaultdict(list)
		for row in rows:
			by_code[row.rule_code].append(row)
		rows = [
			max([row for row in versions if _rule_is_effective(row, payroll_month)], key=_rule_version_sort_key)
			for versions in by_code.values()
			if any(_rule_is_effective(row, payroll_month) for row in versions)
		]
	persisted_codes = {row.rule_code for row in rows}
	for default in DEFAULT_PAYROLL_RULES:
		if default["rule_code"] in persisted_codes:
			continue
		rows.append(frappe._dict({**_rule_doc_values(default, company), "rule_origin": "内置默认（未保存）"}))
	for row in rows:
		row["rule_origin"] = row.get("rule_origin") or "公司规则"
		try:
			parameters = _rule_parameters(row.parameters_json)
			errors = _rule_parameter_errors(row.rule_code, parameters)
		except Exception as exc:
			parameters, errors = {}, [str(exc)]
		if row.rule_code in EXECUTABLE_PAYROLL_RULES:
			row["execution_mode"] = "参数化公式"
			row["execution_status"] = "参数有效，可参与结算" if row.status == "已启用" and not errors else ("参数无效：" + "；".join(errors) if errors else "不参与结算")
		elif row.rule_code in FIXED_PAYROLL_CALCULATION_RULES:
			row["execution_mode"] = "固定计算器公式"
			row["execution_status"] = "由结算计算器执行，结构变更需受控发布"
		else:
			row["execution_mode"] = "来源/说明规则"
			row["execution_status"] = "作为导入、确认或审计依据，不直接计算"
		row["parameters"] = parameters
		row["payroll_month"] = payroll_month
		row["version_label"] = (
			"{0} 起生效".format(row.effective_from)
			if row.get("effective_from") else "历史默认版本"
		)
	return sorted(rows, key=lambda row: (row.get("rule_category") or "", row.get("rule_code") or ""))


@frappe.whitelist()
def validate_payroll_rule_execution(company: str, payroll_month: str = ""):
	"""Expose which formula drivers will be used by a monthly calculation."""
	company = _require_company(company)
	results = []
	for rule_code in EXECUTABLE_PAYROLL_RULES:
		try:
			config = _effective_rule_config(rule_code, payroll_month, company)
			results.append({**config, "valid": 1, "message": "参数有效，可参与结算"})
		except Exception as exc:
			results.append({"rule_code": rule_code, "valid": 0, "message": str(exc)})
	return {"company": company, "payroll_month": payroll_month, "rules": results, "valid": all(row["valid"] for row in results)}


@frappe.whitelist()
def upsert_payroll_rule(**kwargs):
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有维护薪资规则的权限"))
	data = dict(kwargs)
	company = _require_company(data.get("company"))
	if not data.get("rule_code"):
		frappe.throw(_("请填写规则编码"))
	if not data.get("rule_name"):
		frappe.throw(_("请填写规则名称"))
	parameters = _rule_parameters(data.get("parameters_json") or "{}")
	errors = _rule_parameter_errors(data["rule_code"], parameters)
	if errors:
		frappe.throw(_("规则参数无效：{0}").format("；".join(errors)))
	values = {
		"doctype": PAYROLL_RULE_DOCTYPE,
		"company": company,
		"rule_code": data.get("rule_code"),
		"rule_name": data.get("rule_name"),
		"rule_category": data.get("rule_category") or "其他",
		"rule_scope": data.get("rule_scope"),
		"status": data.get("status") or "已启用",
		"editable": flt(data.get("editable", 1)),
		"effective_from": data.get("effective_from"),
		"effective_to": data.get("effective_to"),
		"formula_expression": data.get("formula_expression"),
		"parameters_json": json.dumps(parameters, ensure_ascii=False, sort_keys=True),
		"rule_text": data.get("rule_text"),
		"source_file": data.get("source_file"),
		"source_sheet": data.get("source_sheet"),
		"source_cell": data.get("source_cell"),
		"missing_rule_note": data.get("missing_rule_note"),
		"last_reviewed_by": frappe.session.user,
		"last_reviewed_on": now_datetime(),
	}
	name = data.get("name") or frappe.db.get_value(PAYROLL_RULE_DOCTYPE, {"company": company, "rule_code": values["rule_code"]}, "name")
	if name:
		doc = frappe.get_doc(PAYROLL_RULE_DOCTYPE, name)
		if doc.company and doc.company != company:
			frappe.throw(_("不能跨公司修改薪资规则"))
		if not flt(doc.editable):
			frappe.throw(_("该薪资规则不允许修改"))
		doc.update(values)
		doc.save(ignore_permissions=True)
	else:
		doc = frappe.get_doc(values)
		doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return doc.name


@frappe.whitelist()
def save_payroll_rule_version(company: str, payroll_month: str, rule_code: str, parameters_json=None, status: str = "已启用"):
	"""Save a new company/month rule version instead of overwriting prior months.

	A rule changed for 2026-08 starts on 2026-08-01.  The previous version is
	closed on 2026-07-31, while an already locked attendance/payroll snapshot
	remains immutable and keeps its recorded calculation trace.
	"""
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有维护薪资规则的权限"))
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	default = _default_rule(rule_code)
	if not default:
		frappe.throw(_("未注册的薪资规则：{0}").format(rule_code))
	parameters = _rule_parameters(parameters_json)
	base_parameters = _rule_parameters(default.get("parameters_json"))
	parameters = {**base_parameters, **parameters}
	errors = _rule_parameter_errors(rule_code, parameters)
	if errors:
		frappe.throw(_("规则参数无效：{0}").format("；".join(errors)))
	month_start = f"{payroll_month}-01"
	rows = frappe.get_all(
		PAYROLL_RULE_DOCTYPE,
		filters={"company": company, "rule_code": rule_code},
		fields=["name", "effective_from", "effective_to", "editable", "modified"],
		limit_page_length=500,
	)
	exact = next((row for row in rows if str(row.effective_from or "") == month_start), None)
	values = _rule_doc_values(default, company)
	values.update({
		"status": status or "已启用",
		"effective_from": month_start,
		"effective_to": None,
		"parameters_json": json.dumps(parameters, ensure_ascii=False, sort_keys=True),
		"last_reviewed_by": frappe.session.user,
		"last_reviewed_on": now_datetime(),
	})
	if exact:
		doc = frappe.get_doc(PAYROLL_RULE_DOCTYPE, exact.name)
		if not flt(doc.editable):
			frappe.throw(_("该薪资规则不允许修改"))
		values["effective_to"] = doc.effective_to
		doc.update(values)
		doc.save(ignore_permissions=True)
	else:
		start_date = getdate(month_start)
		previous = [row for row in rows if (not row.effective_from or getdate(row.effective_from) < start_date) and (not row.effective_to or getdate(row.effective_to) >= start_date)]
		if previous:
			prior = max(previous, key=_rule_version_sort_key)
			prior_doc = frappe.get_doc(PAYROLL_RULE_DOCTYPE, prior.name)
			if not flt(prior_doc.editable):
				frappe.throw(_("当前生效的薪资规则不允许修改"))
			prior_doc.effective_to = (start_date - timedelta(days=1)).isoformat()
			prior_doc.save(ignore_permissions=True)
		future = [row for row in rows if row.effective_from and getdate(row.effective_from) > start_date]
		if future:
			values["effective_to"] = (min(getdate(row.effective_from) for row in future) - timedelta(days=1)).isoformat()
		doc = frappe.get_doc({"doctype": PAYROLL_RULE_DOCTYPE, **values})
		doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "rule_code": rule_code, "effective_from": doc.effective_from, "effective_to": doc.effective_to}


def _field_mapping_values(mapping):
	values = {
		"doctype": PAYROLL_FIELD_MAPPING_DOCTYPE,
		"display_order": mapping.get("display_order"),
		"mapping_code": mapping.get("mapping_code"),
		"excel_column": mapping.get("excel_column"),
		"excel_label": mapping.get("excel_label"),
		"system_doctype": mapping.get("system_doctype") or PAYROLL_SETTLEMENT_DOCTYPE,
		"system_field": mapping.get("system_field"),
		"source_module": mapping.get("source_module"),
		"source_detail": mapping.get("source_detail"),
		"formula_expression": mapping.get("formula_expression"),
		"rule_code": mapping.get("rule_code"),
		"required_for_settlement": flt(mapping.get("required_for_settlement", 1)),
		"status": mapping.get("status") or "已启用",
		"remarks": mapping.get("remarks"),
	}
	return values


@frappe.whitelist()
def ensure_default_payroll_field_mappings(force: int = 0):
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有维护薪资字段映射的权限"))
	created = []
	updated = []
	for mapping in PAYROLL_SETTLEMENT_FIELD_MAPPINGS:
		existing = frappe.db.get_value(PAYROLL_FIELD_MAPPING_DOCTYPE, {"mapping_code": mapping["mapping_code"]}, "name")
		values = _field_mapping_values(mapping)
		if existing:
			if not flt(force):
				continue
			doc = frappe.get_doc(PAYROLL_FIELD_MAPPING_DOCTYPE, existing)
			doc.update(values)
			doc.save(ignore_permissions=True)
			updated.append(doc.name)
			continue
		doc = frappe.get_doc(values)
		doc.insert(ignore_permissions=True)
		created.append(doc.name)
	frappe.db.commit()
	return {"created": len(created), "updated": len(updated)}


@frappe.whitelist()
def list_payroll_field_mappings(status: str = "", page_length: int = 200):
	filters = {}
	if status:
		filters["status"] = status
	return frappe.get_all(
		PAYROLL_FIELD_MAPPING_DOCTYPE,
		filters=filters,
		fields=["*"],
		order_by="display_order asc, excel_column asc",
		limit_page_length=int(page_length or 200),
	)


def _payroll_configuration_category(mapping):
	order = int(mapping.get("display_order") or 0)
	if order <= 4:
		return "员工资料"
	if order <= 8:
		return "固定薪资"
	if order <= 25:
		return "考勤结算"
	if order <= 30:
		return "奖金补贴"
	if order <= 34:
		return "扣款税费"
	if order <= 45:
		return "应付与实发"
	if order <= 48:
		return "公司成本"
	return "导出辅助"


def _payroll_configuration_data_type(fieldname, label):
	value = f"{fieldname or ''} {label or ''}".lower()
	if "hours" in value or "工时" in value or "时数" in value:
		return "工时"
	if "count" in value or "次数" in value:
		return "次数"
	if fieldname in ("department", "employee_code", "employee_name"):
		return "文本"
	return "金额"


@frappe.whitelist()
def list_payroll_configuration_items(company: str):
	"""Return the configurable salary-item catalogue used by the setup guide.

	The catalogue exposes every effective settlement field plus atomic options
	that are combined in the legacy Excel layout.  It does not create a second
	calculation engine: each option points to an existing controlled rule,
	monthly source type or settlement mapping.
	"""
	company = _require_company(company)
	persisted_mappings = {
		row.mapping_code: row
		for row in frappe.get_all(PAYROLL_FIELD_MAPPING_DOCTYPE, fields=["*"], limit_page_length=500)
	}
	persisted_rules = {
		row.rule_code: row
		for row in frappe.get_all(
			PAYROLL_RULE_DOCTYPE,
			filters={"company": company},
			fields=["name", "rule_code", "rule_name", "status", "editable"],
			limit_page_length=500,
		)
	}
	items = []
	for default_mapping in PAYROLL_SETTLEMENT_FIELD_MAPPINGS:
		if default_mapping.get("source_module") == "导出辅助":
			continue
		mapping = dict(default_mapping)
		if default_mapping["mapping_code"] in persisted_mappings:
			mapping.update(dict(persisted_mappings[default_mapping["mapping_code"]]))
		rule_code = mapping.get("rule_code") or ""
		rule = persisted_rules.get(rule_code)
		if rule_code in EXECUTABLE_PAYROLL_RULES:
			calculation_mode = "参数化规则"
		elif mapping.get("formula_expression"):
			calculation_mode = "固定计算器"
		else:
			calculation_mode = "来源字段"
		items.append(
			{
				"item_code": mapping.get("mapping_code"),
				"item_name": mapping.get("excel_label"),
				"category": _payroll_configuration_category(mapping),
				"data_type": _payroll_configuration_data_type(mapping.get("system_field"), mapping.get("excel_label")),
				"direction": "参考" if int(mapping.get("display_order") or 0) <= 4 else ("公司承担" if int(mapping.get("display_order") or 0) >= 46 else "参与结算"),
				"source_module": mapping.get("source_module") or "未设置",
				"result_field": mapping.get("system_field") or "",
				"aggregate_target": "",
				"mapping_code": mapping.get("mapping_code"),
				"rule_code": rule_code,
				"calculation_mode": calculation_mode,
				"configuration_status": "已映射" if mapping.get("system_field") else "待映射",
				"rule_status": rule.status if rule else ("使用内置规则" if rule_code else "无需规则"),
			}
		)
	for atomic in PAYROLL_ATOMIC_CONFIGURATION_ITEMS:
		item = dict(atomic)
		rule_code = item.get("rule_code") or ""
		rule = persisted_rules.get(rule_code)
		item.update(
			{
				"mapping_code": "",
				"calculation_mode": "参数化规则" if rule_code in EXECUTABLE_PAYROLL_RULES else "月度来源汇总",
				"configuration_status": "已连接汇总字段" if item.get("result_field") else "待映射",
				"rule_status": rule.status if rule else ("使用内置规则" if rule_code else "来源确认后参与"),
			}
		)
		items.append(item)
	category_order = ["员工资料", "固定薪资", "考勤结算", "奖金补贴", "扣款税费", "应付与实发", "公司成本"]
	items.sort(key=lambda row: (category_order.index(row["category"]) if row["category"] in category_order else 99, row["item_name"] or ""))
	return {
		"company": company,
		"categories": category_order,
		"items": items,
		"summary": {
			"item_count": len(items),
			"mapped_count": sum(1 for row in items if row["configuration_status"] != "待映射"),
			"rule_count": len({row.get("rule_code") for row in items if row.get("rule_code")}),
		},
	}


@frappe.whitelist()
def upsert_payroll_field_mapping(**kwargs):
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有维护薪资字段映射的权限"))
	data = dict(kwargs)
	if not data.get("mapping_code"):
		frappe.throw(_("请填写映射编码"))
	if not data.get("excel_column") or not data.get("excel_label"):
		frappe.throw(_("请填写 Excel 列和字段名"))
	values = _field_mapping_values(data)
	name = data.get("name") or frappe.db.get_value(PAYROLL_FIELD_MAPPING_DOCTYPE, {"mapping_code": values["mapping_code"]}, "name")
	if name:
		doc = frappe.get_doc(PAYROLL_FIELD_MAPPING_DOCTYPE, name)
		doc.update(values)
		doc.save(ignore_permissions=True)
	else:
		doc = frappe.get_doc(values)
		doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return doc.name


def _month_end(payroll_month):
	match = re.match(r"^(\d{4})-(\d{2})$", payroll_month or "")
	if not match:
		return None
	year, month = [int(value) for value in match.groups()]
	return f"{year:04d}-{month:02d}-{monthrange(year, month)[1]:02d}"


def _require_company(company):
	company = (company or "").strip()
	if not company:
		frappe.throw(_("薪资试算必须传入公司，不允许按月份全局读取或删除。"))
	if not frappe.db.exists("Company", company):
		frappe.throw(_("公司 {0} 不存在").format(company))
	return company


def _current_payroll_attendance_lock(company, payroll_month):
	result = list_available_payroll_attendance_locks(company, payroll_month)
	locks = result.get("locks") or []
	return next((row for row in locks if row.get("is_current")), None)


@frappe.whitelist()
def get_payroll_attendance_dependency(company: str, payroll_month: str):
	"""Return payroll's implicit attendance dependency for one company/month."""
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	current = _current_payroll_attendance_lock(company, payroll_month)
	if not current:
		return {
			"company": company,
			"payroll_month": payroll_month,
			"ready": False,
			"attendance_lock_version": "",
			"message": "请先在考勤假期完成并锁定本月考勤终稿",
		}
	version = str(current.get("attendance_lock_version") or "")
	if version.startswith(PROCESSING_ATTENDANCE_LOCK_PREFIX):
		# Materialize the locked attendance snapshot as an internal payroll cache.
		# This is automatic integration, not a payroll-side upload/review step.
		try:
			sync_locked_attendance_final_to_payroll(company, payroll_month, version)
		except Exception as exc:
			return {
				"company": company,
				"payroll_month": payroll_month,
				"ready": False,
				"attendance_lock_version": "",
				"message": _("考勤终稿自动继承失败：{0}").format(exc),
			}
	return {
		"company": company,
		"payroll_month": payroll_month,
		"ready": True,
		"attendance_lock_version": version,
		"summary_count": current.get("summary_count") or 0,
		"locked_on": current.get("locked_on"),
		"status": current.get("status"),
		"message": "已自动继承考勤假期锁定的本月考勤终稿",
	}


@frappe.whitelist()
def reload_payroll_participation_population(company: str, payroll_month: str):
	"""Reload the current locked attendance population and invalidate only unconfirmed downstream trials."""
	_require_payroll_master_manager()
	dependency = get_payroll_attendance_dependency(company, payroll_month)
	if not dependency.get("ready"):
		frappe.throw(_("无法重新加载人员范围：{0}").format(dependency.get("message") or "请先锁定考勤终稿"))
	invalidation = _invalidate_unconfirmed_payroll_trial(
		dependency["company"],
		dependency["payroll_month"],
		dependency["attendance_lock_version"],
		reason=_("重新加载当前锁定考勤人员范围"),
	)
	frappe.db.commit()
	return {**dependency, "invalidation": invalidation}


def _require_payroll_scope(company, payroll_month, attendance_lock_version=""):
	company = _require_company(company)
	payroll_month = (payroll_month or "").strip()
	if not re.match(r"^\d{4}-\d{2}$", payroll_month):
		frappe.throw(_("薪资月份必须为 YYYY-MM"))
	attendance_lock_version = (attendance_lock_version or "").strip()
	if not attendance_lock_version:
		current = _current_payroll_attendance_lock(company, payroll_month)
		attendance_lock_version = str((current or {}).get("attendance_lock_version") or "")
	if not attendance_lock_version:
		frappe.throw(_("请先在考勤假期完成并锁定本月考勤终稿"))
	return company, payroll_month, attendance_lock_version


@frappe.whitelist()
def list_available_payroll_attendance_locks(company: str, payroll_month: str):
	"""List only immutable attendance versions that payroll may consume.

	A reopened attendance month creates a new version.  Older locked summaries
	remain available for audit/reconciliation, while the current version is
	marked for the UI to select by default.  Draft and cross-company summaries
	are deliberately excluded.
	"""
	company = _require_company(company)
	payroll_month = (payroll_month or "").strip()
	if not re.match(r"^\d{4}-\d{2}$", payroll_month):
		frappe.throw(_("薪资月份必须为 YYYY-MM"))

	rows = frappe.get_all(
		MONTHLY_ATTENDANCE_DOCTYPE,
		filters={"company": company, "attendance_month": payroll_month, "lock_status": "已锁定"},
		fields=["attendance_lock_version", "locked_on"],
		limit_page_length=100000,
	)
	by_version = {}
	for row in rows:
		version = str(row.get("attendance_lock_version") or "").strip()
		if not version:
			continue
		item = by_version.setdefault(version, {"attendance_lock_version": version, "summary_count": 0, "locked_on": None})
		item["summary_count"] += 1
		if row.get("locked_on") and (not item["locked_on"] or row.get("locked_on") > item["locked_on"]):
			item["locked_on"] = row.get("locked_on")

	month_lock = frappe.db.get_value(
		"HRMS Attendance Month Lock",
		{"company": company, "attendance_month": payroll_month, "status": "已锁定"},
		["name", "active_version", "locked_on"],
		as_dict=True,
	)
	active_version = str((month_lock or {}).get("active_version") or "")
	# The attendance processing centre is the current source of truth for Yongxin's
	# signed final: attendance facts, apples, missed punches, full attendance and
	# special hours are frozen there together.  Expose that frozen
	# snapshot to payroll without asking HR to download and re-upload it as Excel.
	try:
		from hrms.api import attendance_processing_center

		state = attendance_processing_center.get_processing_batch(company, payroll_month)
		snapshot = str(state.get("locked_snapshot_version") or "")
		if snapshot:
			preview = attendance_processing_center.get_monthly_final_preview(company, payroll_month, "finance")
			if preview.get("available") and str(preview.get("locked_snapshot_version") or "") == snapshot:
				version = f"{PROCESSING_ATTENDANCE_LOCK_PREFIX}{snapshot}"
				by_version[version] = {
					"attendance_lock_version": version,
					"summary_count": len(preview.get("rows") or []),
					"locked_on": (state.get("final_outputs") or {}).get("generated_on"),
					"processing_final": 1,
				}
	except Exception:
		# Existing legacy attendance locks remain usable when the processing centre
		# is not configured or the caller is not an attendance-processing manager.
		pass

	def version_sort_key(item):
		value = item["attendance_lock_version"]
		return (0, int(value)) if value.isdigit() else (1, value)

	locks = []
	for item in sorted(by_version.values(), key=version_sort_key, reverse=True):
		item["is_current"] = bool(item.get("processing_final")) or bool(month_lock and item["attendance_lock_version"] == active_version)
		item["month_lock"] = (month_lock or {}).get("name") or ""
		item["status"] = "处理中心已锁定终稿" if item.get("processing_final") else ("当前已锁定版本" if item["is_current"] else "历史已锁定版本")
		locks.append(item)
	return {"company": company, "payroll_month": payroll_month, "locks": locks}


def _payroll_scope_filters(company, payroll_month, attendance_lock_version=""):
	filters = {"company": _require_company(company), "payroll_month": payroll_month}
	if attendance_lock_version:
		filters["attendance_lock_version"] = attendance_lock_version
	return filters


def _monthly_variable_scope(payroll_month):
	return f"{MONTHLY_VARIABLE_SCOPE_PREFIX}{payroll_month}"


def _default_variable_source_rows():
	rows = []
	for index, item in enumerate(DEFAULT_PAYROLL_VARIABLE_SOURCE_TYPES, 1):
		rows.append({**item, "enabled": 1, "sort_order": index * 10, "required_for_payroll": int(item.get("required_for_payroll") or 0), "target_area": item.get("target_area") or "月度增减项"})
	return rows


@frappe.whitelist()
def list_payroll_variable_source_types():
	"""List administrator-maintainable monthly payroll source definitions."""
	defaults = _default_variable_source_rows()
	# Attendance final is a system-provided, read-only source.  It is displayed
	# with the other payroll inputs, but is not a source type an administrator
	# can maintain.  In particular, do not persist its ``考勤继承`` display
	# marker in target_area: older sites only allow the two editable areas in
	# that Select field and reject the whole catalog on first load.
	system_sources = [item for item in defaults if item["source_code"] == "attendance_final"]
	editable_defaults = [item for item in defaults if item["source_code"] != "attendance_final"]
	# A running site can temporarily have the Python/JS code ahead of its DocType
	# schema (for example immediately after pulling an update, before migrate has
	# finished).  Do not let that make the payroll entry page unusable: the built-in
	# catalog is enough to keep imports available until the administrator migrates.
	if not _doctype_exists(VARIABLE_SOURCE_TYPE_DOCTYPE) or not _doctype_has_field(VARIABLE_SOURCE_TYPE_DOCTYPE, "target_area"):
		return defaults
	try:
		for item in editable_defaults:
			if frappe.db.exists(VARIABLE_SOURCE_TYPE_DOCTYPE, item["source_code"]):
				continue
			frappe.get_doc({"doctype": VARIABLE_SOURCE_TYPE_DOCTYPE, **item}).insert(ignore_permissions=True, ignore_if_duplicate=True)
		rows = frappe.get_all(
			VARIABLE_SOURCE_TYPE_DOCTYPE,
			# Salary changes belong to the preceding employee-salary step.  Full
			# attendance and housing allowance are inherited from the locked attendance
			# final, so none of them is a monthly-upload card in payroll.
			filters={"enabled": 1, "source_code": ["not in", ["attendance_final", "salary_change", "attendance_bonus", "housing_allowance"]]},
			fields=["name", "source_code", "source_name", "purpose", "required_for_payroll", "enabled", "sort_order", "required_fields", "template_notes", "target_area"],
			order_by="sort_order asc, source_name asc",
			limit_page_length=200,
		)
		return system_sources + rows
	except Exception:
		# The next successful migration restores the administrator-maintained
		# catalog.  Meanwhile the defaults are deliberately safe and writable.
		frappe.log_error(frappe.get_traceback(), "Payroll variable source catalog fallback")
		return defaults


def _attendance_scope_filters(company, attendance_month, attendance_lock_version):
	return {
		"company": _require_company(company),
		"attendance_month": attendance_month,
		"attendance_lock_version": attendance_lock_version,
		"lock_status": "已锁定",
	}


def _employee_identity_key(row):
	return getattr(row, "employee", None) or getattr(row, "employee_code", None) or getattr(row, "employee_name", None)


def _employee_population_labels(rows, allowed_keys):
	"""Return readable labels for records outside the locked attendance population.

	Payroll inputs are normally generated from attendance, but this guard also
	protects calculation and confirmation from manually-created or legacy input
	records.  A monthly payroll population is a subset of the locked attendance
	final; a variable or input record must never add a recipient on its own.
	"""
	labels = []
	for row in rows:
		key = _employee_identity_key(row)
		if not key or key in allowed_keys:
			continue
		code = str(getattr(row, "employee_code", "") or "").strip()
		name = str(getattr(row, "employee_name", "") or "").strip()
		labels.append(" ".join(value for value in (code, name) if value) or str(key))
	return sorted(set(labels))


def _assert_row_company(row, company, source_label):
	row_company = getattr(row, "company", None)
	if row_company and row_company != company:
		frappe.throw(_("{0} 存在跨公司数据：{1} 不属于 {2}").format(source_label, row_company, company))


def _source_trace_hash(trace):
	payload = json.dumps(trace, ensure_ascii=False, sort_keys=True, default=str)
	return payload, hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _payroll_run_snapshot(company, payroll_month, attendance_lock_version):
	"""Return the immutable upstream snapshot consumed by one payroll trial.

	The snapshot deliberately contains values as well as record hashes because a
	manual review may change a payroll variable without changing its original
	file hash.  A settlement can only be confirmed while this snapshot still
	matches the current attendance, salary, variable and rule state.
	"""
	company, payroll_month, attendance_lock_version = _require_payroll_scope(
		company, payroll_month, attendance_lock_version
	)
	allowed_versions = {attendance_lock_version, _monthly_variable_scope(payroll_month), ""}
	attendance = frappe.get_all(
		MONTHLY_ATTENDANCE_DOCTYPE,
		filters=_attendance_scope_filters(company, payroll_month, attendance_lock_version),
		# The attendance dependency helper may harmlessly re-sync the same locked
		# snapshot and therefore update ``modified``.  The lock checksum is the
		# stable business identity; using timestamps here would invalidate a trial
		# merely because the page was refreshed.
		fields=["name", "employee", "employee_code", "source_checksum"],
		order_by="name asc",
		limit_page_length=100000,
	)
	variable_rows = frappe.get_all(
		VARIABLE_RECORD_DOCTYPE,
		filters={"company": company, "payroll_month": payroll_month, "review_status": "已确认", "excluded": 0},
		fields=["name", "import_batch", "attendance_lock_version", "employee", "employee_code", "employee_name", "variable_type", "amount", "source_sheet", "source_hash", "modified"],
		order_by="name asc",
		limit_page_length=100000,
	)
	variables = [
		row for row in variable_rows
		if str(row.attendance_lock_version or "") in allowed_versions
		and str(row.source_sheet or "") != "考勤终稿锁定快照"
	]
	batch_rows = frappe.get_all(
		VARIABLE_BATCH_DOCTYPE,
		filters={"company": company, "payroll_month": payroll_month, "status": "已确认"},
		fields=["name", "attendance_lock_version", "source_type", "source_file", "variable_rows", "confirmed_on", "modified"],
		order_by="name asc",
		limit_page_length=10000,
	)
	batches = [
		row for row in batch_rows
		if str(row.attendance_lock_version or "") in allowed_versions
		and not str(row.source_file or "").startswith("attendance-processing-final:")
	]
	salary_filters = {"company": company, "status": ["!=", "已作废"]}
	month_end = _month_end(payroll_month)
	if month_end:
		salary_filters["effective_date"] = ["<=", month_end]
	salary_changes = frappe.get_all(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		filters=salary_filters,
		fields=["name", "employee", "employee_code", "effective_date", "base_salary", "function_allowance", "certificate_allowance", "multi_skill_allowance", "full_salary", "exclude_from_payroll", "modified"],
		order_by="name asc",
		limit_page_length=100000,
	)
	participation_decisions = []
	if _doctype_exists(MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE):
		participation_decisions = frappe.get_all(
			MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE,
			filters={"company": company, "payroll_month": payroll_month, "attendance_lock_version": attendance_lock_version},
			fields=["employee", "decision", "decision_reason", "settlement_basis", "review_status", "approved_by", "approved_on", "modified"],
			order_by="employee asc, modified asc",
			limit_page_length=100000,
		)
	payload = {
		"company": company,
		"payroll_month": payroll_month,
		"attendance_lock_version": attendance_lock_version,
		"attendance": [dict(row) for row in attendance],
		"confirmed_batches": [dict(row) for row in batches],
		"confirmed_variables": [dict(row) for row in variables],
		"salary_changes": [dict(row) for row in salary_changes],
		"participation_decisions": [dict(row) for row in participation_decisions],
		"calculation_rules": _payroll_calculation_rules(company, payroll_month),
		"payroll_formulas": _effective_payroll_formulas(company, payroll_month),
	}
	_trace, snapshot_hash = _source_trace_hash(payload)
	return snapshot_hash


def _trace_snapshot_hash(row):
	try:
		return str((json.loads(getattr(row, "source_trace_json", "") or "{}") or {}).get("payroll_run_snapshot_hash") or "")
	except (TypeError, ValueError):
		return ""


def _invalidate_unconfirmed_payroll_trial(company, payroll_month, attendance_lock_version="", reason=""):
	"""Remove derived, unconfirmed results after an upstream data change."""
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	if not attendance_lock_version:
		current = _current_payroll_attendance_lock(company, payroll_month) or {}
		attendance_lock_version = current.get("attendance_lock_version") or ""
	if not attendance_lock_version:
		return {"deleted_inputs": 0, "deleted_settlements": 0, "reason": reason}
	scope = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
	locked = frappe.get_all(
		PAYROLL_SETTLEMENT_DOCTYPE,
		filters={**scope, "calculation_status": ["in", ["已确认", "已生成工资单"]]},
		pluck="name",
	)
	if locked:
		frappe.throw(_("本月薪资已正式确认，不能直接修改上游数据；请创建薪资调整批次。"))
	input_names = frappe.get_all(PAYROLL_INPUT_DOCTYPE, filters=scope, pluck="name")
	settlement_names = frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters=scope, pluck="name")
	for name in settlement_names:
		frappe.delete_doc(PAYROLL_SETTLEMENT_DOCTYPE, name, ignore_permissions=True, force=True)
	for name in input_names:
		frappe.delete_doc(PAYROLL_INPUT_DOCTYPE, name, ignore_permissions=True, force=True)
	return {
		"deleted_inputs": len(input_names),
		"deleted_settlements": len(settlement_names),
		"reason": reason,
	}


def _salary_structure_sheet(workbook):
	sheet = _matching_sheet(workbook, "薪资架构")
	if sheet:
		return sheet
	# Yongxin's current architecture workbook uses the version itself as the
	# sheet name (for example ``2I``), while the level matrix starts with a
	# horizontal ``序号 / 底薪 / 职能津贴`` block.
	for sheet_name in workbook.sheetnames:
		candidate = workbook[sheet_name]
		rows = _read_rows(candidate, max_rows=12)
		header_rows = [{_normalise(value) for value in row} for row in rows]
		if (
			any("序号" in headers for headers in header_rows)
			and any(any("底薪" in header for header in headers) for headers in header_rows)
		):
			return candidate
	return None


def _row_value(row, index):
	return row[index] if index < len(row) else ""


def _parse_salary_grade_rows(sheet):
	rows = _read_rows(sheet)
	level_rows = _parse_sequential_salary_level_rows(rows)
	if level_rows:
		return level_rows
	grades = []
	current_section = ""
	current_job_family = ""
	header_start = None
	post_headers = []
	for raw in rows:
		normalised = [_normalise(value) for value in raw]
		joined = "".join(normalised)
		if any(marker in joined for marker in ("生产类直接人员", "生产类间接人员", "间接人员", "专业技术类", "管理类", "职员类")):
			current_section = next((_text(value) for value in raw if _text(value)), joined)
			current_job_family = ""
			continue
		if "岗性" in normalised and "岗级" in normalised:
			header_start = normalised.index("岗性")
			post_headers = [_text(value) for value in raw[header_start + 7 : header_start + 14] if _text(value)]
			continue
		if header_start is None:
			continue
		family = _text(_row_value(raw, header_start))
		if family:
			current_job_family = family
		grade_number = _text(_row_value(raw, header_start + 1))
		if not grade_number:
			continue
		if not current_job_family and not any(flt(_row_value(raw, index)) for index in range(header_start + 2, min(len(raw), header_start + 7))):
			continue
		base_salary = flt(_row_value(raw, header_start + 2))
		function_allowance = flt(_row_value(raw, header_start + 3))
		full_salary = flt(_row_value(raw, header_start + 4)) or base_salary + function_allowance
		job_grade = f"{current_job_family}-{grade_number}" if current_job_family else grade_number
		education_allowance = max([flt(value) for value in raw[header_start + 15 : header_start + 19]] or [0])
		grades.append(
			{
				"salary_level": cint(grade_number) or len(grades) + 1,
				"job_nature": current_section,
				"job_grade": job_grade,
				"post_category": "、".join(post_headers),
				"base_salary": base_salary,
				"function_allowance": function_allowance,
				"full_salary": full_salary,
				"grade_difference": flt(_row_value(raw, header_start + 5)),
				"grade_difference_ratio": flt(_row_value(raw, header_start + 6)),
				"education_allowance": education_allowance,
				"multi_skill_allowance": flt(_row_value(raw, header_start + 19)),
				"full_attendance_bonus_standard": flt(_row_value(raw, header_start + 20)),
				"rental_subsidy_standard": flt(_row_value(raw, header_start + 21)),
				"large_night_shift_allowance": flt(_row_value(raw, header_start + 22)) or 45,
				"small_night_shift_allowance": 24,
				"certificate_allowance": flt(_row_value(raw, header_start + 23)),
				"raw_row_json": json.dumps(raw, ensure_ascii=False, default=str),
			}
		)
	return grades


def _parse_sequential_salary_level_rows(rows):
	"""Parse the compact level matrix: 序号 / 底薪 / 职能津贴 / 全薪 / 级差."""
	if not rows:
		return []
	labels, label_column = {}, None
	for row in rows:
		for column, value in enumerate(row):
			if "序号" in _normalise(value):
				labels["levels"] = row
				label_column = column
				break
		if label_column is not None:
			break
	if label_column is None:
		return []
	for row in rows:
		label = _normalise(_row_value(row, label_column))
		if "底薪" in label:
			labels["base_salary"] = row
		elif "职能津贴" in label or "职务津贴" in label:
			labels["function_allowance"] = row
		elif "全薪" in label or "薪资小计" in label:
			labels["full_salary"] = row
		elif "级差" in label:
			labels["grade_difference"] = row
	if not {"levels", "base_salary", "function_allowance"}.issubset(labels):
		return []

	grades = []
	for column, raw_level in enumerate(labels["levels"][label_column + 1 :], start=label_column + 1):
		level = cint(raw_level)
		if level <= 0:
			continue
		base_salary = flt(_row_value(labels["base_salary"], column))
		function_allowance = flt(_row_value(labels["function_allowance"], column))
		full_salary = flt(_row_value(labels.get("full_salary", []), column)) or base_salary + function_allowance
		previous_full = grades[-1]["full_salary"] if grades else full_salary
		grades.append(
			{
				"salary_level": level,
				"job_grade": str(level),
				"base_salary": base_salary,
				"function_allowance": function_allowance,
				"full_salary": full_salary,
				"grade_difference": flt(_row_value(labels.get("grade_difference", []), column)) or full_salary - previous_full,
				"raw_row_json": json.dumps({"column": column, "level": level}, ensure_ascii=False),
			}
		)
	return grades


def _suggest_salary_structure_version(sheet):
	"""Use a version-named sheet such as ``2I`` as the import default."""
	title = _text(sheet.title).strip()
	if re.fullmatch(r"[A-Za-z0-9_-]+", title):
		return title
	for row in _read_rows(sheet, max_rows=4):
		for value in row:
			match = re.search(r"(?:薪资架构表[-－]?)?([0-9]+[A-Za-z]+)版?", _text(value))
			if match:
				return match.group(1)
	return ""


@frappe.whitelist()
def preview_salary_structure_workbook(file_url: str):
	workbook = _load_workbook(file_url)
	sheet = _salary_structure_sheet(workbook)
	if not sheet:
		return {"found": False, "sheet_name": "薪资架构", "row_count": 0, "grade_rows": 0, "sample_grades": []}
	grades = _parse_salary_grade_rows(sheet)
	return {
		"found": True,
		"sheet_name": sheet.title,
		"suggested_structure_version": _suggest_salary_structure_version(sheet),
		"row_count": len(_read_rows(sheet)),
		"grade_rows": len(grades),
		"sample_grades": grades[:10],
	}


@frappe.whitelist()
def import_salary_structure_workbook(file_url: str, structure_version: str, effective_from: str, effective_to: str = ""):
	_require_payroll_master_manager()
	if not structure_version:
		frappe.throw(_("请填写薪资架构版本"))
	if not effective_from:
		frappe.throw(_("请填写生效开始日期"))
	workbook = _load_workbook(file_url)
	sheet = _salary_structure_sheet(workbook)
	if not sheet:
		frappe.throw(_("未找到薪资架构工作表"))
	grades = _parse_salary_grade_rows(sheet)
	if not grades:
		frappe.throw(_("薪资架构工作表未识别到薪资档位"))

	version_name = frappe.db.get_value(SALARY_STRUCTURE_VERSION_DOCTYPE, {"structure_version": structure_version}, "name")
	if version_name:
		version = frappe.get_doc(SALARY_STRUCTURE_VERSION_DOCTYPE, version_name)
		version.effective_from = effective_from
		version.effective_to = effective_to
		version.source_file = file_url
		version.status = "已启用"
		version.save(ignore_permissions=True)
	else:
		version = frappe.get_doc(
			{
				"doctype": SALARY_STRUCTURE_VERSION_DOCTYPE,
				"structure_version": structure_version,
				"effective_from": effective_from,
				"effective_to": effective_to,
				"source_file": file_url,
				"status": "已启用",
			}
		)
		version.insert(ignore_permissions=True)

	for name in frappe.get_all(SALARY_GRADE_DOCTYPE, filters={"salary_structure_version": version.name}, pluck="name"):
		frappe.delete_doc(SALARY_GRADE_DOCTYPE, name, ignore_permissions=True, force=True)
	for grade in grades:
		doc = frappe.get_doc({"doctype": SALARY_GRADE_DOCTYPE, "salary_structure_version": version.name, **grade})
		doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"version": version.name, "grade_rows": len(grades)}


@frappe.whitelist()
def list_salary_structure_versions(page_length: int = 0):
	"""Return every salary-structure version for the version-management page.

	A newly created version intentionally starts empty, so filtering out headers
	without grade rows makes it impossible to continue editing that version.  The
	grade count is returned separately: an empty version can be managed, but it
	cannot be selected as an employee salary grade until levels are saved.
	"""
	versions = frappe.get_all(
		SALARY_STRUCTURE_VERSION_DOCTYPE,
		fields=["name", "structure_version", "status", "effective_from", "effective_to", "source_file", "remarks", "modified"],
		order_by="effective_from desc, modified desc",
		limit_page_length=100000,
	)
	grade_counts = defaultdict(int)
	for row in frappe.get_all(SALARY_GRADE_DOCTYPE, fields=["salary_structure_version"], limit_page_length=100000):
		if row.get("salary_structure_version"):
			grade_counts[row.salary_structure_version] += 1
	for row in versions:
		row["grade_count"] = grade_counts[row.name]
	limit = cint(page_length)
	return versions[:limit] if limit > 0 else versions


@frappe.whitelist()
def list_salary_grades(structure_version: str = "", page_length: int = 0):
	filters = {"salary_structure_version": structure_version} if structure_version else {}
	rows = frappe.get_all(
		SALARY_GRADE_DOCTYPE,
		filters=filters,
		fields=[
			"name",
			"salary_structure_version",
			"salary_level",
			"job_nature",
			"job_grade",
			"base_salary",
			"function_allowance",
			"certificate_allowance",
			"multi_skill_allowance",
			"full_salary",
			"full_attendance_bonus_standard",
			"rental_subsidy_standard",
		],
		order_by="salary_structure_version desc, modified asc",
		limit_page_length=100000,
	)

	def level_order(row):
		level = cint(row.get("salary_level"))
		if not level:
			match = re.search(r"(\d+)$", _text(row.get("job_grade")))
			level = cint(match.group(1)) if match else 999999
		return level

	# Do not render placeholder records created with an empty imported row.  A
	# zero-valued salary is still valid data when the level or job grade is set.
	rows = [
		row
		for row in rows
		if any(
			row.get(field) not in (None, "")
			for field in ("salary_level", "job_grade", "base_salary", "function_allowance", "full_salary")
		)
	]
	rows.sort(key=lambda row: (level_order(row), _text(row.get("job_grade")), _text(row.get("name"))))
	limit = cint(page_length)
	return rows[:limit] if limit > 0 else rows


@frappe.whitelist()
def list_assignable_salary_grades(payroll_month: str = ""):
	"""Return enabled salary levels for employee salary decisions.

	All versions applicable to the selected month are returned.  A saved employee
	may legitimately be linked to an earlier version (for example ``2I``) while a
	newer version (``2H``) is also enabled; hiding the earlier option makes a
	valid Link render as “手动定薪”.
	"""
	_require_payroll_master_manager()
	versions = _active_salary_structure_versions(payroll_month)
	if not versions:
		versions = frappe.get_all(
			SALARY_STRUCTURE_VERSION_DOCTYPE,
			filters={"status": "已启用"},
			fields=["name", "structure_version", "effective_from", "effective_to", "source_file"],
			order_by="effective_from desc, modified desc",
			limit_page_length=1000,
		)
	if not versions:
		return []
	version_by_name = {version.name: version for version in versions}
	version_order = {version.name: index for index, version in enumerate(versions)}
	rows = frappe.get_all(
		SALARY_GRADE_DOCTYPE,
		filters={"salary_structure_version": ["in", list(version_by_name)]},
		fields=["name", "salary_structure_version", "salary_level", "job_grade", "base_salary", "function_allowance", "full_salary", "modified"],
		order_by="modified desc",
		limit_page_length=500,
	)

	def level_order(row):
		level = cint(row.get("salary_level"))
		if not level:
			match = re.search(r"(\d+)$", _text(row.get("job_grade")))
			level = cint(match.group(1)) if match else 999999
		return level

	unique_rows = {}
	for row in rows:
		# The newest record wins only for duplicate levels within the same version.
		# Identical level numbers across 2H/2I are separate, valid choices.
		unique_rows.setdefault((row.salary_structure_version, level_order(row)), row)

	options = []
	for (version_name, level), row in sorted(
		unique_rows.items(), key=lambda item: (version_order.get(item[0][0], 999999), item[0][1])
	):
		version_code = version_by_name[version_name].structure_version
		options.append(
			{
				"name": row.name,
				"structure_version": version_code,
				"salary_level": level,
				"base_salary": flt(row.base_salary),
				"function_allowance": flt(row.function_allowance),
				"full_salary": flt(row.full_salary) or flt(row.base_salary) + flt(row.function_allowance),
				"label": _("{0} · {1}（底薪 {2}，职能津贴 {3}，全薪 {4}）").format(
					version_code,
					level,
					flt(row.base_salary),
					flt(row.function_allowance),
					flt(row.full_salary) or flt(row.base_salary) + flt(row.function_allowance),
				),
			}
		)
	return options


@frappe.whitelist()
def create_salary_level_structure_version(structure_version: str, effective_from: str, effective_to: str = ""):
	"""Create a blank, standalone salary-level version for the matrix editor."""
	_require_payroll_master_manager()
	structure_version = _text(structure_version)
	if not structure_version:
		frappe.throw(_("请填写薪级表版本名称"))
	if not effective_from:
		frappe.throw(_("请填写生效开始日期"))
	if frappe.db.exists(SALARY_STRUCTURE_VERSION_DOCTYPE, {"structure_version": structure_version}):
		frappe.throw(_("薪级表版本已存在，请使用不同名称"))
	version = frappe.get_doc(
		{
			"doctype": SALARY_STRUCTURE_VERSION_DOCTYPE,
			"structure_version": structure_version,
			"effective_from": effective_from,
			"effective_to": effective_to,
			"status": "已启用",
		}
	)
	version.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": version.name}


@frappe.whitelist()
def delete_salary_structure_version(name: str):
	"""Delete an unused salary-table version and all of its salary grades.

	It is intentionally blocked once any employee salary-change row references a
	grade from the version.  Deleting a referenced master record would make paid
	or imported history look like a manual salary and cannot be recovered safely.
	"""
	_require_payroll_master_manager()
	if not name or not frappe.db.exists(SALARY_STRUCTURE_VERSION_DOCTYPE, name):
		frappe.throw(_("未找到薪级表版本"))
	version = frappe.get_doc(SALARY_STRUCTURE_VERSION_DOCTYPE, name)
	grade_names = frappe.get_all(SALARY_GRADE_DOCTYPE, filters={"salary_structure_version": name}, pluck="name")
	if grade_names:
		reference_count = frappe.db.count(
			EMPLOYEE_SALARY_CHANGE_DOCTYPE,
			filters={"salary_grade": ["in", grade_names]},
		)
		if reference_count:
			frappe.throw(
				_("不能删除版本 {0}：已有 {1} 条员工定薪记录引用该版本。请保留该历史版本。").format(
					version.structure_version, reference_count
				)
			)
	for grade_name in grade_names:
		frappe.delete_doc(SALARY_GRADE_DOCTYPE, grade_name, ignore_permissions=True, force=True)
	frappe.delete_doc(SALARY_STRUCTURE_VERSION_DOCTYPE, name, ignore_permissions=True, force=True)
	frappe.db.commit()
	return {"name": name, "structure_version": version.structure_version, "deleted_grades": len(grade_names)}


@frappe.whitelist()
def save_salary_level_structure(structure_version: str, levels=None):
	"""Persist an ordered level matrix without assigning it to any employees."""
	_require_payroll_master_manager()
	if not frappe.db.exists(SALARY_STRUCTURE_VERSION_DOCTYPE, structure_version):
		frappe.throw(_("未找到薪级表版本"))
	if isinstance(levels, str):
		levels = json.loads(levels or "[]")
	if not isinstance(levels, list) or not levels:
		frappe.throw(_("请至少保留一个薪级"))

	clean_levels = []
	for index, raw in enumerate(levels, start=1):
		level = cint((raw or {}).get("level") or (raw or {}).get("salary_level")) or index
		base_salary = flt((raw or {}).get("base_salary"))
		function_allowance = flt((raw or {}).get("function_allowance"))
		full_salary = base_salary + function_allowance
		previous_full = clean_levels[-1]["full_salary"] if clean_levels else full_salary
		clean_levels.append(
			{
				"salary_level": level,
				"job_grade": str(level),
				"base_salary": base_salary,
				"function_allowance": function_allowance,
				"full_salary": full_salary,
				"grade_difference": full_salary - previous_full,
				"raw_row_json": json.dumps({"source": "salary-level-matrix", "salary_level": level}, ensure_ascii=False),
			}
		)

	for name in frappe.get_all(SALARY_GRADE_DOCTYPE, filters={"salary_structure_version": structure_version}, pluck="name"):
		frappe.delete_doc(SALARY_GRADE_DOCTYPE, name, ignore_permissions=True, force=True)
	for level in clean_levels:
		frappe.get_doc({"doctype": SALARY_GRADE_DOCTYPE, "salary_structure_version": structure_version, **level}).insert(ignore_permissions=True)
	frappe.db.commit()
	return {"saved": len(clean_levels)}


def _employee_context(employee):
	if not employee:
		return {}
	# Employee installations differ: some retain ``confirmation_date`` while
	# others expose only ``final_confirmation_date``.  Context lookup is shared
	# by every payroll import, so request only fields that exist in this site.
	fields = _safe_fields(
		"Employee",
		["custom_employee_code", "employee_number", "employee_name", "department", "designation", "date_of_joining", "relieving_date", "company", "employment_type", "status", "custom_is_confirmed", "final_confirmation_date", "confirmation_date"],
	)
	row = frappe.db.get_value("Employee", employee, fields, as_dict=True)
	if row:
		row["employee_code"] = row.get("custom_employee_code") or row.get("employee_number") or employee
	return row or {}


def _salary_contribution_defaults(employee_context, effective_date=""):
	"""Default contribution flags without taking away HR's per-person override."""
	month_end = ""
	try:
		parsed = getdate(effective_date) if effective_date else None
		month_end = _month_end(parsed.strftime("%Y-%m") if parsed else "")
	except Exception:
		month_end = ""
	stage = _employment_stage(employee_context or {}, month_end)
	is_active = (employee_context or {}).get("status") == "Active"
	# 永新规则：仅正式员工默认缴纳社保、公积金；试用员工默认均不缴纳。
	# 两项开关仍可由 HR 在员工行内按实际情况单独调整。
	contribution_enabled = int(bool(is_active and stage == "正式"))
	social_insurance_enabled = contribution_enabled
	housing_fund_enabled = contribution_enabled
	return {"social_insurance_enabled": social_insurance_enabled, "housing_fund_enabled": housing_fund_enabled, "employment_stage": stage}


def _social_insurance_payroll_policy(employee, payroll_month):
	"""Return whether this employee's social-insurance variables apply this month.

	The monthly social-insurance list remains the normal source of truth.  An
	employee master record only overrides that list when HR has explicitly marked
	the employee as not participating/stopped, or a confirmed start date has not
	yet been reached.  This intentionally does not infer non-participation from
	probation status.
	"""
	policy = {"apply": True, "reason": "按社保名单"}
	if not employee or not re.match(r"^\d{4}-\d{2}$", str(payroll_month or "")):
		return policy

	meta = frappe.get_meta("Employee")
	fieldnames = [
		fieldname
		for fieldname in (
			"custom_social_insurance_status",
			"custom_social_insurance_start_date",
			"custom_social_insurance_end_date",
		)
		if meta.get_field(fieldname)
	]
	if not fieldnames:
		return policy

	values = frappe.db.get_value("Employee", employee, fieldnames, as_dict=True) or {}
	status = str(values.get("custom_social_insurance_status") or "按社保名单").strip()
	if status in SOCIAL_INSURANCE_MANUAL_EXCLUSION_STATUSES:
		return {"apply": False, "reason": status}
	if status != "参保中":
		return policy

	year, month = (int(part) for part in payroll_month.split("-"))
	month_start = date(year, month, 1)
	month_end = date(year, month, monthrange(year, month)[1])
	start_date = getdate(values.get("custom_social_insurance_start_date")) if values.get("custom_social_insurance_start_date") else None
	end_date = getdate(values.get("custom_social_insurance_end_date")) if values.get("custom_social_insurance_end_date") else None
	if start_date and start_date > month_end:
		return {"apply": False, "reason": "社保起缴日期未到"}
	if end_date and end_date < month_start:
		return {"apply": False, "reason": "社保已停缴"}
	return {"apply": True, "reason": "参保中"}


def _apply_social_insurance_payroll_policy(values, employee, payroll_month):
	"""Suppress only social-insurance amounts when the employee is not covered."""
	policy = _social_insurance_payroll_policy(employee, payroll_month)
	if not policy["apply"]:
		for variable_type in SOCIAL_INSURANCE_VARIABLE_TYPES:
			values[variable_type] = 0
	return policy


def _grade_context(salary_grade):
	if not salary_grade:
		return {}
	return frappe.db.get_value(
		SALARY_GRADE_DOCTYPE,
		salary_grade,
		["base_salary", "function_allowance", "certificate_allowance", "multi_skill_allowance", "full_salary"],
		as_dict=True,
	) or {}


@frappe.whitelist()
def create_employee_salary_change(**kwargs):
	_require_payroll_master_manager()
	data = dict(kwargs)
	company = _require_company(data.get("company"))
	# 员工定薪不再有草稿或审批状态：每次保存都直接提交并作为算薪依据。
	# Keep the legacy field internally so existing records and historical queries
	# remain compatible with the rest of the payroll module.
	status = "已批准"
	employee = data.get("employee")
	if not employee:
		frappe.throw(_("请先选择员工。"))
	if not data.get("effective_date"):
		frappe.throw(_("请填写生效日期。"))
	employee_context = _employee_context(employee)
	if employee_context.get("company") and employee_context.get("company") != company:
		frappe.throw(_("员工 {0} 不属于公司 {1}").format(employee, company))
	grade_context = _grade_context(data.get("salary_grade"))
	contribution_defaults = _salary_contribution_defaults(employee_context, data.get("effective_date"))
	base_salary = flt(data.get("base_salary")) or flt(grade_context.get("base_salary"))
	function_allowance = flt(data.get("function_allowance")) or flt(grade_context.get("function_allowance"))
	certificate_allowance = flt(data.get("certificate_allowance")) or flt(grade_context.get("certificate_allowance"))
	multi_skill_allowance = flt(data.get("multi_skill_allowance")) or flt(grade_context.get("multi_skill_allowance"))
	# The supplied certificate/multi-skill allowance register confirms these
	# amounts ceased to be part of full salary on 2026-05-01.  Always calculate
	# the stored hourly-salary base from the two fixed-pay items instead of
	# trusting a legacy full-salary column that may still include them.
	full_salary = base_salary + function_allowance
	doc = frappe.get_doc(
		{
			"doctype": EMPLOYEE_SALARY_CHANGE_DOCTYPE,
			"company": company,
			"employee": employee,
			"employee_code": data.get("employee_code") or employee,
			"employee_name": data.get("employee_name") or employee_context.get("employee_name"),
			"department": data.get("department") or employee_context.get("department"),
			"designation": data.get("designation") or employee_context.get("designation"),
			"education_level_text": data.get("education_level_text"),
			"date_of_joining": data.get("date_of_joining") or employee_context.get("date_of_joining"),
			"effective_date": data.get("effective_date"),
			"salary_grade": data.get("salary_grade"),
			"base_salary": base_salary,
			"function_allowance": function_allowance,
			"certificate_allowance": certificate_allowance,
			"multi_skill_allowance": multi_skill_allowance,
			"full_salary": full_salary,
			"housing_fund_enabled": flt(data.get("housing_fund_enabled")) if "housing_fund_enabled" in data else contribution_defaults["housing_fund_enabled"],
			"social_insurance_enabled": flt(data.get("social_insurance_enabled")) if "social_insurance_enabled" in data else contribution_defaults["social_insurance_enabled"],
			"company_cost_total": flt(data.get("company_cost_total")),
			"prepared_by": data.get("prepared_by"),
			"reviewed_by": data.get("reviewed_by"),
			"approved_by": data.get("approved_by"),
			"status": status,
			"source_file": data.get("source_file"),
			"remarks": data.get("remarks"),
		}
	)
	doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return doc.name


@frappe.whitelist()
def update_employee_salary_change(name: str = "", company: str = "", employee: str = "", values: str | dict | None = None):
	"""Save one row from the employee-salary web table.

	Employee, department and post stay owned by the roster.  This endpoint only
	updates the salary-adjustment fields that users also maintain in Excel.
	"""
	_require_payroll_master_manager()
	company = _require_company(company)
	if isinstance(values, str):
		try:
			values = json.loads(values or "{}")
		except (TypeError, ValueError) as exc:
			frappe.throw(_("填写内容无法识别：{0}").format(exc))
	values = values or {}
	if not isinstance(values, dict):
		frappe.throw(_("填写内容格式不正确"))
	# Some form integrations submit nested JSON one level deeper.  Decode that
	# shape explicitly so changed values are never silently ignored.
	if isinstance(values.get("values"), str) and len(values) == 1:
		try:
			values = json.loads(values["values"] or "{}")
		except (TypeError, ValueError) as exc:
			frappe.throw(_("填写内容无法识别：{0}").format(exc))
	if not isinstance(values, dict):
		frappe.throw(_("填写内容格式不正确"))
	if not name:
		if not employee:
			frappe.throw(_("请先选择员工"))
		created_name = create_employee_salary_change(company=company, employee=employee, **values)
		created = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, created_name)
		return {"name": created.name, "full_salary": created.full_salary, "status": created.status, "created": 1}
	doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name)
	if doc.company != company:
		frappe.throw(_("不能修改其他公司的员工定薪记录"))
	allowed = {"effective_date", "salary_grade", "base_salary", "function_allowance", "certificate_allowance", "multi_skill_allowance", "social_insurance_enabled", "housing_fund_enabled", "remarks"}
	if not values.get("effective_date"):
		frappe.throw(_("请填写生效日期"))
	for fieldname in allowed:
		if fieldname not in values:
			continue
		if fieldname in {"base_salary", "function_allowance", "certificate_allowance", "multi_skill_allowance", "social_insurance_enabled", "housing_fund_enabled"}:
			setattr(doc, fieldname, flt(values[fieldname]))
		else:
			setattr(doc, fieldname, values[fieldname])
	if "salary_grade" in values and values.get("salary_grade"):
		grade_context = _grade_context(doc.salary_grade)
		if not grade_context:
			frappe.throw(_("所选薪级不存在，请重新选择"))
		# API callers may send only a salary grade.  The web table sends the three
		# displayed values as well, which keeps its explicit manual override valid.
		if "base_salary" not in values:
			doc.base_salary = flt(grade_context.get("base_salary"))
		if "function_allowance" not in values:
			doc.function_allowance = flt(grade_context.get("function_allowance"))
	doc.full_salary = flt(doc.base_salary) + flt(doc.function_allowance)
	# A legacy draft becomes effective as soon as it is edited in the new grid.
	doc.status = "已批准"
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "full_salary": doc.full_salary, "status": doc.status}


def _is_salary_excluded(row):
	return bool(flt(row.get("exclude_from_payroll")))


@frappe.whitelist()
def set_employee_payroll_participation(employee: str, company: str, payroll_month: str, participates: int = 1):
	"""Exclude or restore one employee for the selected payroll month.

	An exclusion is an approved, zero-value marker effective from this month. It
	does not delete the employee's normal salary history; restoring participation
	voids only that marker and resumes the previous approved salary decision.
	"""
	_require_payroll_master_manager()
	company = _require_company(company)
	month_end = _month_end(payroll_month)
	if not month_end:
		frappe.throw(_("请选择有效的薪资月份"))
	context = _employee_context(employee)
	if not context:
		frappe.throw(_("未找到员工"))
	if context.get("company") and context.get("company") != company:
		frappe.throw(_("员工不属于当前公司"))
	filters = {"company": company, "employee": employee, "exclude_from_payroll": 1, "status": "已批准"}
	markers = frappe.get_all(EMPLOYEE_SALARY_CHANGE_DOCTYPE, filters=filters, fields=["name", "effective_date"], order_by="effective_date desc, modified desc", limit_page_length=20)
	if flt(participates):
		for marker in markers:
			if str(marker.effective_date or "") <= month_end:
				doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, marker.name)
				doc.status = "已作废"
				doc.save(ignore_permissions=True)
				break
		frappe.db.commit()
		return {"employee": employee, "participates": 1}

	effective_date = f"{payroll_month}-01"
	existing = next((marker for marker in markers if str(marker.effective_date or "") == effective_date), None)
	if existing:
		doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, existing.name)
	else:
		doc = frappe.get_doc({"doctype": EMPLOYEE_SALARY_CHANGE_DOCTYPE})
	doc.update({
		"company": company,
		"employee": employee,
		"employee_code": _employee_code(context) or employee,
		"employee_name": context.get("employee_name") or employee,
		"department": context.get("department"),
		"designation": context.get("designation"),
		"date_of_joining": context.get("date_of_joining"),
		"effective_date": effective_date,
		"change_reason": "本月不参与薪资计算",
		"exclude_from_payroll": 1,
		"exclude_reason": "薪资中心标记，本月不参与计算",
		"status": "已批准",
		"remarks": "本月不参与计算；恢复后自动沿用此前已批准的定薪。",
	})
	if existing:
		doc.status = "已批准"
		doc.exclude_from_payroll = 1
		doc.exclude_reason = "薪资中心标记，本月不参与计算"
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"employee": employee, "participates": 0, "name": doc.name}


def _monthly_payroll_participation_decision_map(company, payroll_month, attendance_lock_version):
	"""Return the latest explicit decision for every person in one locked payroll population."""
	if not _doctype_exists(MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE):
		return {}
	rows = frappe.get_all(
		MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE,
		filters={
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
		},
		fields=["name", "employee", "employee_code", "employee_name", "decision", "decision_reason", "settlement_basis", "review_status", "approved_by", "approved_on", "approval_note", "modified"],
		order_by="modified desc, name desc",
		limit_page_length=100000,
	)
	by_key = {}
	for row in rows:
		for key in (row.get("employee"), row.get("employee_code"), row.get("employee_name")):
			key = _text(key).strip()
			if key and key not in by_key:
				by_key[key] = row
	return by_key


def _participation_decision_for_row(decisions, row):
	for key in (row.get("employee"), row.get("employee_code"), row.get("employee_name")):
		decision = decisions.get(_text(key).strip()) if key else None
		if decision:
			return decision
	return None


def _participation_decision_blocks_calculation(decision):
	if not decision:
		return False
	if decision.get("decision") == "异常待审核":
		return True
	return decision.get("decision") in {"离职结算", "不参与计算"} and decision.get("review_status") != PAYROLL_PARTICIPATION_APPROVED_STATUS


def _participation_decision_excludes(decision):
	return bool(
		decision
		and decision.get("decision") == "不参与计算"
		and decision.get("review_status") == PAYROLL_PARTICIPATION_APPROVED_STATUS
	)


def _employee_left_in_payroll_month(employee_row, payroll_month):
	"""Use the exit date when available so historical closed months stay calculable."""
	if not employee_row or _text(employee_row.get("status")) not in {"Left", "离职"}:
		return False
	relieving_date = _date_or_none(employee_row.get("relieving_date"))
	return not relieving_date or relieving_date <= _month_end(payroll_month)


def _attendance_employee_context_map(attendance_rows):
	employee_names = sorted({_text(row.get("employee")).strip() for row in attendance_rows if row.get("employee")})
	if not employee_names:
		return {}
	fields = _safe_fields("Employee", ["name", "status", "relieving_date", "employee_name", "custom_employee_code", "department"])
	return {
		row.name: row
		for row in _safe_get_all("Employee", fields=fields, filters={"name": ["in", employee_names]}, limit_page_length=100000)
	}


@frappe.whitelist()
def save_monthly_payroll_participation_decision(
	company: str,
	payroll_month: str,
	attendance_lock_version: str,
	employee: str,
	decision: str,
	decision_reason: str = "",
	settlement_basis: str = "",
	approval_note: str = "",
	approved: int = 0,
):
	"""Record an auditable monthly handling decision for one locked-attendance employee."""
	_require_payroll_master_manager()
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	if not _doctype_exists(MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE):
		frappe.throw(_("月度薪资参与决策数据表尚未安装，请先执行站点迁移。"))
	decision = _text(decision).strip()
	decision_reason = _text(decision_reason).strip()
	settlement_basis = _text(settlement_basis).strip()
	approval_note = _text(approval_note).strip()
	if decision not in PAYROLL_PARTICIPATION_DECISIONS:
		frappe.throw(_("请选择有效的本月处理方式。"))
	attendance = frappe.db.get_value(
		MONTHLY_ATTENDANCE_DOCTYPE,
		{**_attendance_scope_filters(company, payroll_month, attendance_lock_version), "employee": employee},
		["employee", "employee_code", "employee_name", "department"],
		as_dict=True,
	)
	if not attendance:
		frappe.throw(_("该员工不在当前锁定考勤终稿中，不能加入薪资计算范围。"))
	if decision == "异常待审核" and not decision_reason:
		frappe.throw(_("异常待审核必须填写异常说明。"))
	if decision == "不参与计算" and not decision_reason:
		frappe.throw(_("不参与计算必须填写原因。"))
	if decision == "离职结算":
		if not decision_reason:
			frappe.throw(_("离职结算必须填写处理说明。"))
		if not settlement_basis:
			frappe.throw(_("离职结算必须填写结算依据，例如离职审批单或已批准结算标准。"))
		if not cint(approved):
			frappe.throw(_("离职结算需勾选审核通过后才能参与计算。"))
	if decision == "不参与计算" and not cint(approved):
		frappe.throw(_("不参与计算需勾选审核通过后才能生效。"))

	context = _employee_context(employee)
	filters = {
		"company": company,
		"payroll_month": payroll_month,
		"attendance_lock_version": attendance_lock_version,
		"employee": employee,
	}
	existing_name = frappe.db.get_value(MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE, filters, "name")
	doc = frappe.get_doc(MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE, existing_name) if existing_name else frappe.get_doc({"doctype": MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE})
	review_status = "待审核" if decision == "异常待审核" else (PAYROLL_PARTICIPATION_APPROVED_STATUS if cint(approved) else "无需审核")
	doc.update({
		**filters,
		"employee_code": attendance.get("employee_code") or context.get("employee_code") or employee,
		"employee_name": attendance.get("employee_name") or context.get("employee_name") or employee,
		"department": attendance.get("department") or context.get("department"),
		"employee_status": context.get("status") or "",
		"relieving_date": context.get("relieving_date"),
		"decision": decision,
		"decision_reason": decision_reason,
		"settlement_basis": settlement_basis,
		"review_status": review_status,
		"approval_note": approval_note,
		"approved_by": frappe.session.user if review_status == PAYROLL_PARTICIPATION_APPROVED_STATUS else "",
		"approved_on": now_datetime() if review_status == PAYROLL_PARTICIPATION_APPROVED_STATUS else None,
	})
	if existing_name:
		doc.save(ignore_permissions=True)
	else:
		doc.insert(ignore_permissions=True)
	# A previous quick exclusion used an employee-salary marker.  A subsequent
	# decision of "正常计薪" must restore that employee instead of leaving a hidden
	# marker that contradicts the visible personnel-range decision.
	if decision == "正常计薪":
		marker_name = frappe.db.get_value(
			EMPLOYEE_SALARY_CHANGE_DOCTYPE,
			{
				"company": company,
				"employee": employee,
				"exclude_from_payroll": 1,
				"status": "已批准",
				"effective_date": ["<=", _month_end(payroll_month)],
			},
			"name",
			order_by="effective_date desc, modified desc",
		)
		if marker_name:
			marker = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, marker_name)
			marker.status = "已作废"
			marker.save(ignore_permissions=True)
	_invalidate_unconfirmed_payroll_trial(
		company,
		payroll_month,
		attendance_lock_version,
		reason=_("更新人员范围决策：{0} / {1}").format(attendance.get("employee_code") or attendance.get("employee_name"), decision),
	)
	frappe.db.commit()
	return {"name": doc.name, "decision": decision, "review_status": review_status}


@frappe.whitelist()
def list_employee_salary_changes(company: str, employee: str = "", payroll_month: str = "", page_length: int = 50):
	company = _require_company(company)
	filters = {"company": company}
	if employee:
		filters["employee"] = employee
	month_end = _month_end(payroll_month)
	if month_end:
		filters["effective_date"] = ["<=", month_end]
	return frappe.get_all(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		filters=filters,
		fields=["*"],
		order_by="effective_date desc, modified desc",
		limit_page_length=int(page_length or 50),
	)


@frappe.whitelist()
def list_employee_salary_change_grid(company: str, payroll_month: str = "", page_length: int = 1000):
	"""Return one editable salary row for every active employee, including blanks."""
	company = _require_company(company)
	month_end = _month_end(payroll_month)
	employee_fields = _safe_fields(
		"Employee",
		["name", "employee_name", "custom_employee_code", "department", "designation", "employment_type", "status", "custom_is_confirmed", "date_of_joining", "final_confirmation_date", "confirmation_date", "company"],
	)
	filters = {"company": company} if _doctype_has_field("Employee", "company") else {}
	if _doctype_has_field("Employee", "status"):
		filters["status"] = "Active"
	employees = _safe_get_all("Employee", filters=filters, fields=employee_fields, order_by="employee_name asc", limit_page_length=100000)
	change_filters = {"company": company}
	if month_end:
		change_filters["effective_date"] = ["<=", month_end]
	changes = frappe.get_all(EMPLOYEE_SALARY_CHANGE_DOCTYPE, filters=change_filters, fields=["*"], order_by="effective_date desc, modified desc", limit_page_length=100000)
	changes_by_employee = {}
	excluded_employees = set()
	for change in changes:
		if not change.employee or change.employee in changes_by_employee or change.employee in excluded_employees:
			continue
		if _is_salary_excluded(change):
			excluded_employees.add(change.employee)
			continue
		changes_by_employee[change.employee] = change
	department_values = {row.get("department") for row in employees if row.get("department")}
	department_names = {
		row.name: row.department_name
		for row in frappe.get_all(
			"Department",
			filters={"name": ["in", sorted(department_values)]},
			fields=["name", "department_name"],
			limit_page_length=100000,
		)
	}
	# A stopped/expired version must remain visible on employees who already use
	# it.  The editable dropdown only offers currently applicable grades, but the
	# saved Link gets a labelled historical option instead of being rendered as
	# “手动定薪”.
	grade_ids = {change.salary_grade for change in changes_by_employee.values() if change.get("salary_grade")}
	grade_labels = {}
	if grade_ids:
		grade_rows = frappe.get_all(
			SALARY_GRADE_DOCTYPE,
			filters={"name": ["in", sorted(grade_ids)]},
			fields=["name", "salary_structure_version", "salary_level", "job_grade", "base_salary", "function_allowance", "full_salary"],
			limit_page_length=100000,
		)
		version_ids = {row.salary_structure_version for row in grade_rows if row.get("salary_structure_version")}
		version_details = {
			row.name: row
			for row in frappe.get_all(
				SALARY_STRUCTURE_VERSION_DOCTYPE,
				filters={"name": ["in", sorted(version_ids)]},
				fields=["name", "structure_version", "status"],
				limit_page_length=100000,
			)
		}
		for grade in grade_rows:
			version = version_details.get(grade.salary_structure_version)
			version_code = version.structure_version if version else _("历史版本")
			level = cint(grade.salary_level) or _text(grade.job_grade) or "—"
			status_hint = "" if not version or version.status == "已启用" else _(" · {0}").format(version.status)
			grade_labels[grade.name] = _("{0} · {1}（底薪 {2}，职能津贴 {3}，全薪 {4}{5}）").format(
				version_code,
				level,
				flt(grade.base_salary),
				flt(grade.function_allowance),
				flt(grade.full_salary) or flt(grade.base_salary) + flt(grade.function_allowance),
				status_hint,
			)
	rows = []
	for employee_row in employees[: int(page_length or 1000)]:
		if employee_row.name in excluded_employees:
			continue
		change = changes_by_employee.get(employee_row.name)
		defaults = _salary_contribution_defaults(employee_row, f"{payroll_month}-01" if payroll_month else "")
		rows.append(
			{
				"name": change.name if change else "",
				"employee": employee_row.name,
				"employee_name": employee_row.get("employee_name") or employee_row.name,
				"employee_code": _employee_code(employee_row) or employee_row.name,
				# Show the business label immediately, including before the one-time
				# Department rename patch has been applied to legacy records.
				"department": department_names.get(employee_row.get("department"))
				or re.sub(r"\s+-\s+[^-]+$", "", employee_row.get("department") or "").strip(),
				"employment_type": "在职·{0}".format(defaults["employment_stage"]),
				"effective_date": change.effective_date if change else f"{payroll_month}-01" if payroll_month else "",
				"salary_grade": change.salary_grade if change else "",
				"salary_grade_label": grade_labels.get(change.salary_grade, _("已绑定历史薪级")) if change else "",
				"base_salary": change.base_salary if change else "",
				"function_allowance": change.function_allowance if change else 0,
				"certificate_allowance": change.certificate_allowance if change else 0,
				"multi_skill_allowance": change.multi_skill_allowance if change else 0,
				"full_salary": change.full_salary if change else 0,
				"social_insurance_enabled": int(bool(flt(change.social_insurance_enabled))) if change else defaults["social_insurance_enabled"],
				"housing_fund_enabled": int(bool(flt(change.housing_fund_enabled))) if change else defaults["housing_fund_enabled"],
				"contribution_default": "社保、公积金：正式默认缴纳，试用默认不缴纳；可按员工实际情况调整",
				"is_new": int(not change),
			}
		)
	# Missing base salaries are the only values requiring immediate action in
	# this grid, so keep them at the top instead of duplicating a separate alert
	# list above the editable table.
	rows.sort(key=lambda row: (0 if not flt(row.get("base_salary")) else 1, _text(row.get("employee_name")), _text(row.get("employee_code"))))
	return {"rows": rows, "total_count": len(rows), "company": company, "payroll_month": payroll_month}


@frappe.whitelist()
def get_active_salary_change_for_employee(employee: str | None = None, employee_code: str = "", payroll_month: str = "", company: str = ""):
	company = _require_company(company)
	filters = {"status": ["!=", "已作废"], "company": company}
	if employee:
		filters["employee"] = employee
	elif employee_code:
		filters["employee_code"] = employee_code
	month_end = _month_end(payroll_month)
	if month_end:
		filters["effective_date"] = ["<=", month_end]
	rows = frappe.get_all(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		filters=filters,
		fields=["*"],
		order_by="effective_date desc, modified desc",
		limit_page_length=1,
	)
	return rows[0] if rows else None


def _welfare_rule(source_type):
	for rule in WELFARE_SOURCE_RULES:
		if rule["source_type"] == source_type:
			return rule
	return {}


def _source_category(source_type):
	if source_type in ("学历补贴", "租房补贴", "高温补贴", "手机话费补贴", "油费补贴"):
		return "补贴"
	if source_type in ("薪资构成", "证书多能工津贴"):
		return "薪资主数据"
	if source_type in ("宿舍住宿费", "宿舍水电费"):
		return "宿舍"
	if source_type in ("社保个人", "社保公司", "公积金个人", "公积金公司"):
		return "社保公积金"
	if source_type in ("提案改善奖", "继续服务奖", "已发福利", "生产奖", "其他奖金", "奖惩提报", "苹果树"):
		return "奖金福利"
	if source_type in ("所得税", "年终奖所得税", "水电费及扣款", "其他扣款", "离职薪资结算"):
		return "个税扣款"
	return "其他"


@frappe.whitelist()
def list_payroll_welfare_source_rules():
	return WELFARE_SOURCE_RULES


@frappe.whitelist()
def upsert_payroll_welfare_source_record(**kwargs):
	data = dict(kwargs)
	company, payroll_month, attendance_lock_version = _require_payroll_scope(
		data.get("company"),
		data.get("payroll_month"),
		data.get("attendance_lock_version"),
	)
	source_type = data.get("source_type")
	if not source_type:
		frappe.throw(_("请选择来源类型"))
	employee_code = data.get("employee_code")
	employee_name = data.get("employee_name")
	employee = data.get("employee") or _payroll_employee_lookup(employee_code, employee_name)
	employee_context = _employee_context(employee)
	if employee_context.get("company") and employee_context.get("company") != company:
		frappe.throw(_("福利扣款来源员工 {0} 不属于公司 {1}").format(employee_code or employee_name or employee, company))
	rule = _welfare_rule(source_type)
	variable_type = data.get("variable_type") or WELFARE_SOURCE_VARIABLE_TYPE_MAP.get(source_type)
	direction = data.get("direction") or rule.get("direction")
	department = _department_lookup(
		data.get("department") or employee_context.get("department"), company
	) or data.get("department") or employee_context.get("department")
	trace_payload, trace_hash = _source_trace_hash(
		{
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
			"source_type": source_type,
			"source_reference": data.get("source_reference"),
			"source_file": data.get("source_file"),
			"employee": employee or employee_code or employee_name,
		}
	)
	values = {
		"doctype": WELFARE_SOURCE_DOCTYPE,
		"company": company,
		"payroll_month": payroll_month,
		"attendance_lock_version": attendance_lock_version,
		"source_category": data.get("source_category") or _source_category(source_type),
		"source_type": source_type,
		"variable_type": variable_type,
		"direction": direction,
		"employee": employee,
		"employee_code": employee_code,
		"employee_name": employee_name or employee_context.get("employee_name"),
		"department": department,
		"amount": flt(data.get("amount")),
		"eligibility_status": data.get("eligibility_status") or "符合",
		# There is no review stage.  Newly entered welfare rows wait only for the
		# single "confirm and sync" action, which validates employee matching.
		"confirmation_status": data.get("confirmation_status") or "待确认",
		"source_reference": data.get("source_reference"),
		"source_file": data.get("source_file"),
		"rule_snapshot": data.get("rule_snapshot") or rule.get("rule"),
		"remarks": data.get("remarks"),
		"raw_row_json": json.dumps(data, ensure_ascii=False, default=str),
		"source_trace_json": trace_payload,
		"source_hash": trace_hash,
	}
	if values["confirmation_status"] == "已确认":
		values["confirmed_by"] = frappe.session.user
		values["confirmed_on"] = now_datetime()
	record_name = data.get("name")
	if record_name and frappe.db.exists(WELFARE_SOURCE_DOCTYPE, record_name):
		doc = frappe.get_doc(WELFARE_SOURCE_DOCTYPE, record_name)
		doc.update(values)
		doc.save(ignore_permissions=True)
	else:
		doc = frappe.get_doc(values)
		doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return doc.name


@frappe.whitelist()
def list_payroll_welfare_source_records(company: str, payroll_month: str = "", source_type: str = "", attendance_lock_version: str = "", page_length: int = 100):
	company = _require_company(company)
	filters = {"company": company}
	if payroll_month:
		filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		filters["attendance_lock_version"] = attendance_lock_version
	if source_type:
		filters["source_type"] = source_type
	return frappe.get_all(
		WELFARE_SOURCE_DOCTYPE,
		filters=filters,
		fields=["*"],
		order_by="modified desc",
		limit_page_length=int(page_length or 100),
	)


@frappe.whitelist()
def sync_welfare_sources_to_payroll_variables(company: str, payroll_month: str, attendance_lock_version: str):
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	filters = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
	for name in frappe.get_all(
		VARIABLE_RECORD_DOCTYPE,
		filters={**filters, "source_sheet": WELFARE_SOURCE_SYNC_SHEET},
		pluck="name",
	):
		frappe.delete_doc(VARIABLE_RECORD_DOCTYPE, name, ignore_permissions=True, force=True)

	created = []
	source_rows = frappe.get_all(
		WELFARE_SOURCE_DOCTYPE,
		filters={**filters, "confirmation_status": "已确认", "eligibility_status": "符合"},
		fields=["*"],
		order_by="modified asc",
	)
	for row in source_rows:
		_assert_row_company(row, company, _("福利扣款来源"))
		variable_type = row.variable_type or WELFARE_SOURCE_VARIABLE_TYPE_MAP.get(row.source_type)
		if not variable_type:
			continue
		trace_payload, trace_hash = _source_trace_hash(
			{
				"company": company,
				"payroll_month": payroll_month,
				"attendance_lock_version": attendance_lock_version,
				"welfare_source_record": row.name,
				"source_type": row.source_type,
				"employee": row.employee or row.employee_code or row.employee_name,
			}
		)
		doc = frappe.get_doc(
			{
				"doctype": VARIABLE_RECORD_DOCTYPE,
				"company": company,
				"payroll_month": payroll_month,
				"attendance_lock_version": attendance_lock_version,
				"employee": row.employee,
				"employee_code": row.employee_code,
				"employee_name": row.employee_name,
				"department": _department_lookup(row.department, company) or row.department,
				"variable_type": variable_type,
				"amount": flt(row.amount),
				"review_status": "已确认",
				"validation_status": "通过",
				"validation_message": "来源记录已审核确认",
				"excluded": 0,
				"source_sheet": WELFARE_SOURCE_SYNC_SHEET,
				"remarks": f"{row.source_type}：{row.source_reference or row.name}",
				"raw_row_json": json.dumps(row, ensure_ascii=False, default=str),
				"source_trace_json": trace_payload,
				"source_hash": trace_hash,
			}
		)
		doc.insert(ignore_permissions=True)
		created.append(doc.name)
	frappe.db.commit()
	return {"created": len(created), "records": created}


@frappe.whitelist()
def confirm_all_payroll_welfare_sources(company: str, payroll_month: str, attendance_lock_version: str):
	"""Confirm legacy pending welfare rows and immediately sync them to payroll."""
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	pending = frappe.get_all(
		WELFARE_SOURCE_DOCTYPE,
		filters={
			**_payroll_scope_filters(company, payroll_month, attendance_lock_version),
			"confirmation_status": ["in", ["草稿", "待确认"]],
		},
		fields=["name", "employee", "employee_code", "employee_name", "department"],
		limit_page_length=100000,
	)
	if not pending:
		return {"confirmed": 0, "synced": 0, "message": _("当前月份没有待确认的福利/扣款来源。")}
	invalid = []
	resolved_people = {}
	for row in pending:
		employee = row.employee or _payroll_employee_lookup(row.employee_code, row.employee_name)
		context = _employee_context(employee)
		if not employee or (context.get("company") and context.get("company") != company):
			invalid.append(row.employee_code or row.employee_name or row.name)
			continue
		resolved_people[row.name] = (employee, context)
	if invalid:
		frappe.throw(_("以下福利/扣款来源未匹配当前公司员工，请先修正：{0}").format("、".join(invalid[:10])))
	invalidation = _invalidate_unconfirmed_payroll_trial(
		company, payroll_month, attendance_lock_version, reason=_("一键确认福利/扣款来源")
	)
	for row in pending:
		employee, context = resolved_people[row.name]
		frappe.db.set_value(
			WELFARE_SOURCE_DOCTYPE,
			row.name,
			{
				"employee": employee,
				"employee_code": context.get("employee_code") or row.employee_code,
				"employee_name": context.get("employee_name") or row.employee_name,
				"department": context.get("department") or row.department,
				"confirmation_status": "已确认",
				"confirmed_by": frappe.session.user,
				"confirmed_on": now_datetime(),
			},
			update_modified=False,
		)
	# The sync reads the uncommitted confirmation changes in this request and
	# commits both steps together, so no source is left confirmed-but-unsynced.
	sync_result = sync_welfare_sources_to_payroll_variables(company, payroll_month, attendance_lock_version)
	return {
		"confirmed": len(pending),
		"synced": sync_result.get("created", 0),
		"invalidated_trial": invalidation,
		"message": _("已确认 {0} 条福利/扣款来源并同步 {1} 条薪资变量。").format(len(pending), sync_result.get("created", 0)),
	}


@frappe.whitelist()
def list_payroll_import_templates():
	templates = []
	for template in PAYROLL_IMPORT_TEMPLATES:
		item = dict(template)
		item["columns"] = [
			{"excel_column": column[0], "system_field": column[1], "description": column[2]}
			for column in template.get("columns", [])
		]
		templates.append(item)
	return {"templates": templates, "settlement_fields": PAYROLL_SETTLEMENT_FIELD_MAPPINGS}


@frappe.whitelist()
def create_payroll_data_closure_template_file():
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from frappe.utils.file_manager import save_file

	workbook = Workbook()
	workbook.remove(workbook.active)
	header_fill = PatternFill("solid", fgColor="D9EAF7")
	header_font = Font(bold=True)

	for template in PAYROLL_IMPORT_TEMPLATES:
		sheet = workbook.create_sheet(template["sheet_name"])
		headers = [column[0] for column in template.get("columns", [])]
		sheet.append(headers)
		for cell in sheet[1]:
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = Alignment(horizontal="center", vertical="center")
		for index, column in enumerate(headers, start=1):
			sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(max(len(column) + 8, 14), 32)

	field_sheet = workbook.create_sheet("字段说明")
	field_sheet.append(["工作表", "目标表", "Excel字段", "系统字段", "说明"])
	for template in PAYROLL_IMPORT_TEMPLATES:
		for column in template.get("columns", []):
			field_sheet.append([template["sheet_name"], template["target_doctype"], column[0], column[1], column[2]])
	for cell in field_sheet[1]:
		cell.fill = header_fill
		cell.font = header_font
	for width_index, width in enumerate([22, 34, 22, 28, 58], start=1):
		field_sheet.column_dimensions[field_sheet.cell(1, width_index).column_letter].width = width

	mapping_sheet = workbook.create_sheet("薪资结算字段对应")
	mapping_sheet.append(["Excel列", "Excel字段名", "系统字段", "来源模块", "公式/来源", "对应规则"])
	for mapping in PAYROLL_SETTLEMENT_FIELD_MAPPINGS:
		mapping_sheet.append(
			[
				mapping.get("excel_column"),
				mapping.get("excel_label"),
				mapping.get("system_field"),
				mapping.get("source_module"),
				mapping.get("formula_expression") or mapping.get("source_detail"),
				mapping.get("rule_code"),
			]
		)
	for cell in mapping_sheet[1]:
		cell.fill = header_fill
		cell.font = header_font
	for width_index, width in enumerate([10, 26, 32, 18, 52, 34], start=1):
		mapping_sheet.column_dimensions[mapping_sheet.cell(1, width_index).column_letter].width = width

	output = BytesIO()
	workbook.save(output)
	file_doc = save_file(
		f"薪资数据闭环导入模板-{datetime.today().strftime('%Y%m%d%H%M%S')}.xlsx",
		output.getvalue(),
		None,
		None,
		is_private=1,
	)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name}


@frappe.whitelist()
def create_employee_salary_change_template_file():
	"""Create the focused template used for employee salary master-data changes.

	The combined data-closure workbook remains useful for a monthly payroll run.
	This smaller file is deliberately kept next to the salary architecture workflow so
	HR can maintain joiner, confirmation, promotion and adjustment records without
	accidentally importing monthly attendance or welfare data.
	"""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from frappe.utils.file_manager import save_file

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "员工薪资异动导入"
	template = next(item for item in PAYROLL_IMPORT_TEMPLATES if item["template_key"] == "employee_salary_change")
	headers = [column[0] for column in template["columns"]]
	sheet.append(headers)
	header_fill = PatternFill("solid", fgColor="D9EAF7")
	for cell in sheet[1]:
		cell.fill = header_fill
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
	for index, column in enumerate(headers, start=1):
		sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(max(len(column) + 8, 14), 28)

	notes = workbook.create_sheet("填写说明")
	notes.append(["字段", "填写规则"])
	for cell in notes[1]:
		cell.fill = header_fill
		cell.font = Font(bold=True)
	for column in template["columns"]:
		notes.append([column[0], column[2]])
	notes.append(["提交规则", "导入后立即提交并生效；生效日期不晚于算薪月份的记录会进入正式薪资试算。"])
	notes.append(["薪资小计", "可留空，系统会按底薪 + 职能津贴计算。"])
	notes.append(["公司", "模板不填写公司；导入时以页面顶部当前公司为准，跨公司数据会被阻断。"])
	notes.column_dimensions["A"].width = 22
	notes.column_dimensions["B"].width = 88

	output = BytesIO()
	workbook.save(output)
	file_doc = save_file(
		f"员工薪资异动导入模板-{datetime.today().strftime('%Y%m%d%H%M%S')}.xlsx",
		output.getvalue(),
		None,
		None,
		is_private=1,
	)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name}


@frappe.whitelist()
def create_housing_allowance_base_data_template_file(company: str, payroll_month: str = ""):
	"""Create the first-level housing-eligibility template used for system calculation."""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from frappe.utils.file_manager import save_file

	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month or datetime.today().strftime("%Y-%m"))
	rule = _effective_rule_config("WELFARE_RENTAL_SUBSIDY", payroll_month, company)
	parameters = rule["parameters"]
	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "住房补贴一阶数据"
	headers = [
		"工号", "姓名", "部门", "是否非苏州户籍", "是否在苏州租房", "是否在苏州购房",
		"是否住公司宿舍", "入职日期", "离职日期", "离职当月是否满勤", "备注",
	]
	sheet.append(headers)
	# A sample makes the yes/no convention unambiguous, but is marked so it cannot
	# accidentally be imported as a real employee.
	sheet.append(["示例工号", "示例（导入前删除本行）", "", "是", "是", "否", "否", "2026-07-08", "", "", "仅示例"])
	header_fill = PatternFill("solid", fgColor="D9EAF7")
	for cell in sheet[1]:
		cell.fill = header_fill
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center", vertical="center")
	for index, header in enumerate(headers, start=1):
		sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(max(len(header) + 8, 16), 28)

	notes = workbook.create_sheet("填写说明")
	notes.append(["字段", "填写规则"])
	for cell in notes[1]:
		cell.fill = header_fill
		cell.font = Font(bold=True)
	for field in ("是否非苏州户籍", "是否在苏州租房", "是否在苏州购房", "是否住公司宿舍"):
		notes.append([field, "必填，只能填写“是”或“否”。系统据此判断资格；空白会作为错误阻断确认。"])
	notes.append(["入职日期", "格式 yyyy-mm-dd；仅入职当月用于计算：10 日及以前 {0} 元、11–20 日 {1} 元、21 日及以后 {2} 元。".format(flt(parameters.get("before_or_on_day_10")), flt(parameters.get("day_11_to_20")), flt(parameters.get("after_or_on_day_21")))])
	notes.append(["离职日期 / 离职当月是否满勤", "离职当月必须填写满勤“是/否”；满勤 {0} 元，未满勤 0 元。".format(flt(parameters.get("resignation_full_attendance")))])
	notes.append(["不符合资格", "苏州户籍、未在苏州租房、已在苏州购房或住公司宿舍的人员会保留明细并自动标为“不参与计算”。"])
	notes.append(["二阶表识别", "若上传表已含“住房补贴”金额列，系统不重复计算，只校验员工与金额后直接进入待确认明细。"])
	notes.append(["规则快照", "{0}；{1}".format(rule.get("source"), rule.get("rule_name"))])
	notes.column_dimensions["A"].width = 28
	notes.column_dimensions["B"].width = 96

	output = BytesIO()
	workbook.save(output)
	file_doc = save_file(
		f"{payroll_month}-住房补贴一阶数据模板-{datetime.today().strftime('%Y%m%d%H%M%S')}.xlsx",
		output.getvalue(),
		None,
		None,
		is_private=1,
	)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name, "rule": rule}


def _employee_salary_change_import_rows(workbook, payroll_month):
	"""Return rows from either the simple system template or Yongxin's form."""
	template = next(item for item in PAYROLL_IMPORT_TEMPLATES if item["template_key"] == "employee_salary_change")
	standard_sheet = _matching_sheet(workbook, template["sheet_name"])
	if standard_sheet:
		return _template_rows_as_dicts(standard_sheet, template), standard_sheet.title
	for sheet_name in workbook.sheetnames:
		sheet = workbook[sheet_name]
		if _raw_payroll_source_kind(sheet) == "salary_change":
			return [_salary_change_row_from_source(row, payroll_month) for row in _salary_change_form_rows(sheet)], sheet.title
	return [], ""


def _validate_employee_salary_change_import_rows(rows, company, payroll_month):
	preview_rows, valid_rows = [], []
	for index, row in enumerate(rows, start=1):
		# Rows from the original form are normalized before this point; rows from the
		# system template already use the same business labels.
		mapped = _salary_change_row_from_source(row, payroll_month) if "调整后底薪" in row else row
		employee_code = _text(_first(mapped, "工号", "员工编号", "employee_code"))
		employee_name = _text(_first(mapped, "姓名", "employee_name"))
		structure_version = _text(_first(mapped, "薪资架构版本", "版本"))
		salary_level = cint(_first(mapped, "薪资序号"))
		grade, grade_match_mode = _salary_grade_from_import_row(mapped, payroll_month)
		has_source_salary = bool(flt(_first(mapped, "底薪")) or flt(_first(mapped, "职能津贴")))
		errors = []
		if not employee_code and not employee_name:
			errors.append("缺少工号或姓名")
		employee = _payroll_employee_lookup(employee_code, employee_name)
		code_employee = _employee_lookup(employee_code, "") if employee_code else None
		name_employee = _employee_lookup("", employee_name) if employee_name else None
		if employee_code and employee_name and code_employee and name_employee and code_employee != name_employee:
			errors.append("工号与姓名不一致：工号对应姓名为 {0}".format(_employee_context(code_employee).get("employee_name") or "-"))
		elif not employee:
			errors.append("未匹配到员工花名册")
		else:
			context = _employee_context(employee)
			if context.get("company") and context.get("company") != company:
				errors.append("员工不属于当前公司")
			elif not grade and not (structure_version and salary_level):
				history_grade = _salary_grade_from_matching_history(
					employee,
					_first(mapped, "底薪"),
					_first(mapped, "职能津贴"),
					_date_or_none(_first(mapped, "生效日期")) or (f"{payroll_month}-01" if payroll_month else ""),
				)
				if history_grade:
					grade, grade_match_mode = history_grade, "沿用既有档位"
		if not _date_or_none(_first(mapped, "生效日期")) and not payroll_month:
			errors.append("缺少生效日期")
		if structure_version and salary_level:
			if not grade:
				errors.append("未匹配到薪资架构：{0} / {1}".format(structure_version, salary_level))
		elif not has_source_salary:
			errors.append("缺少薪资架构匹配信息或表内定薪金额")
		preview_rows.append({
			"row_number": index,
			"employee_code": employee_code,
			"employee_name": employee_name,
			"department": _first(mapped, "部门"),
			"effective_date": _first(mapped, "生效日期") or (f"{payroll_month}-01" if payroll_month else ""),
			"base_salary": flt(_first(mapped, "底薪")),
			"function_allowance": flt(_first(mapped, "职能津贴")),
			"full_salary": flt(_first(mapped, "薪资小计")),
			"structure_version": structure_version,
			"salary_level": salary_level or "",
			"match_status": (
				"已匹配薪资架构"
				if grade_match_mode == "薪资架构"
				else (
					"按表内金额匹配薪资架构"
					if grade_match_mode == "表内金额唯一匹配"
					else ("沿用既有薪资档位" if grade_match_mode else ("使用表内定薪金额" if has_source_salary else "未匹配"))
				)
			),
			"errors": errors,
		})
		if not errors:
			valid_rows.append(mapped)
	return preview_rows, valid_rows


@frappe.whitelist()
def preview_employee_salary_change_workbook(file_url: str, company: str, payroll_month: str = ""):
	company = _require_company(company)
	workbook = _load_workbook(file_url)
	rows, sheet_name = _employee_salary_change_import_rows(workbook, payroll_month)
	if not sheet_name:
		return {"found": False, "sheet_name": "", "total_rows": 0, "valid_rows": 0, "failed_rows": 0, "rows": [], "message": "未找到员工薪资调整表"}
	preview_rows, valid_rows = _validate_employee_salary_change_import_rows(rows, company, payroll_month)
	return {
		"found": True,
		"sheet_name": sheet_name,
		"total_rows": len(preview_rows),
		"valid_rows": len(valid_rows),
		"failed_rows": len(preview_rows) - len(valid_rows),
		# The client paginates the complete result set and brings invalid rows to the
		# top. Returning only the first 50 rows made a later invalid row impossible
		# to find while still preventing the user from confirming the import.
		"rows": preview_rows,
		"message": "已识别人员薪资调整表；系统只读取调整后的底薪、职能津贴和全薪。校验异常行会显示原因，确认时仅导入通过的行。",
	}


def _employee_salary_change_snapshot(doc):
	"""Return only business fields, so a batch can restore an overwritten row."""
	meta = frappe.get_meta(EMPLOYEE_SALARY_CHANGE_DOCTYPE)
	return {
		field.fieldname: doc.get(field.fieldname)
		for field in meta.fields
		if field.fieldname and field.fieldtype not in {"Section Break", "Column Break", "Tab Break", "HTML"}
	}


def _employee_salary_change_signature(doc):
	"""Detect edits made after import before allowing a destructive rollback."""
	payload = _employee_salary_change_snapshot(doc)
	payload.pop("salary_import_batch", None)
	return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str).encode()).hexdigest()


def _employee_salary_change_name_from_row(row, payroll_month, company):
	employee_code = _first(row, "工号", "员工编号", "employee_code")
	employee_name = _first(row, "姓名", "employee_name")
	employee = _payroll_employee_lookup(employee_code, employee_name)
	effective_date = _date_or_none(_first(row, "生效日期")) or (f"{payroll_month}-01" if payroll_month else None)
	if not employee or not effective_date:
		return ""
	return frappe.db.get_value(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		{"company": company, "employee_code": _text(employee_code) or employee, "effective_date": effective_date},
		"name",
	)


def _salary_change_import_batch(company, file_url, payroll_month, total_rows, valid_rows, failed_rows=0, error_summary=""):
	return frappe.get_doc(
		{
			"doctype": FORM_IMPORT_BATCH_DOCTYPE,
			"company": company,
			"module_name": "薪酬",
			"template_key": "employee_salary_change",
			"template_name": "员工定薪 Excel 导入",
			"source_file": file_url,
			"status": "已导入待处理",
			"total_rows": total_rows,
			"valid_rows": valid_rows,
			"failed_rows": failed_rows,
			"error_summary": error_summary,
			"imported_by": frappe.session.user,
			"imported_on": now_datetime(),
		}
	)


@frappe.whitelist()
def import_employee_salary_change_workbook(file_url: str, company: str, payroll_month: str = ""):
	_require_payroll_master_manager()
	company = _require_company(company)
	workbook = _load_workbook(file_url)
	rows, sheet_name = _employee_salary_change_import_rows(workbook, payroll_month)
	if not sheet_name:
		frappe.throw(_("未找到员工薪资调整表；请上传《人员薪资调整模板（月）》或系统下载的模板。"))
	preview_rows, valid_rows = _validate_employee_salary_change_import_rows(rows, company, payroll_month)
	invalid_rows = [row for row in preview_rows if row["errors"]]
	if not valid_rows:
		frappe.throw(_("没有可导入的员工定薪记录。请根据预览中的校验原因修正 Excel 或员工花名册后重试。"))
	error_summary = "；".join(f"第{row['row_number']}行：{'、'.join(row['errors'])}" for row in invalid_rows[:10])
	batch = _salary_change_import_batch(
		company,
		file_url,
		payroll_month,
		len(rows),
		len(valid_rows),
		len(invalid_rows),
		error_summary,
	)
	batch.insert(ignore_permissions=True)
	rollback = {"created": {}, "updated": {}}
	changes = []
	for row in valid_rows:
		existing_name = _employee_salary_change_name_from_row(row, payroll_month, company)
		if existing_name:
			existing_doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, existing_name)
			rollback["updated"][existing_name] = _employee_salary_change_snapshot(existing_doc)
		name = _upsert_employee_salary_change_from_row(
			row, payroll_month, company, source_file=file_url, salary_import_batch=batch.name
		)
		if name:
			changes.append(name)
			imported_doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name)
			if existing_name:
				rollback["updated"][name] = {
					"before": rollback["updated"][name],
					"after_signature": _employee_salary_change_signature(imported_doc),
				}
			else:
				rollback["created"][name] = _employee_salary_change_signature(imported_doc)
	batch.status = "部分失败" if invalid_rows else "已处理"
	batch.mapping_json = json.dumps(rollback, ensure_ascii=False, default=str)
	batch.notes = _("可撤销：新建记录将删除，覆盖记录将恢复至导入前。已被后续修改或已进入正式薪资结算的数据不能撤销。")
	batch.save(ignore_permissions=True)
	frappe.db.commit()
	return {
		"sheet_name": sheet_name,
		"imported_rows": len(changes),
		"total_rows": len(rows),
		"skipped_rows": len(invalid_rows),
		"error_summary": error_summary,
		"batch": batch.name,
		"created_rows": len(rollback["created"]),
		"updated_rows": len(rollback["updated"]),
	}


def _salary_change_batch_months(batch_name):
	rows = frappe.get_all(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		filters={"salary_import_batch": batch_name},
		pluck="effective_date",
		limit_page_length=100000,
	)
	# Database Date fields are returned as ``datetime.date`` objects.  Normalize
	# them before deriving the payroll month; _workflow_month intentionally only
	# accepts the YYYY-MM string used by the payroll workflow.
	months = set()
	for effective_date in rows:
		normalized_date = _date_or_none(effective_date)
		if normalized_date:
			months.add(_workflow_month(normalized_date[:7]))
	return sorted(months)


def _assert_salary_change_batch_reversible(batch):
	if batch.status not in {"已处理", "部分失败"}:
		frappe.throw(_("只有未撤销的员工定薪导入批次可以撤销。"))
	for month in _salary_change_batch_months(batch.name):
		lock = _current_payroll_attendance_lock(batch.company, month) or {}
		version = lock.get("attendance_lock_version") or ""
		if not version:
			continue
		locked = frappe.db.exists(
			PAYROLL_SETTLEMENT_DOCTYPE,
			{**_payroll_scope_filters(batch.company, month, version), "calculation_status": ["in", ["已确认", "已生成工资单"]]},
		)
		if locked:
			frappe.throw(_("{0} 已正式确认或生成工资单，不能撤销上游员工定薪导入。请创建新的薪资调整记录。 ").format(month))


@frappe.whitelist()
def list_employee_salary_change_import_batches(company: str, page_length: int = 10):
	company = _require_company(company)
	batches = frappe.get_all(
		FORM_IMPORT_BATCH_DOCTYPE,
		filters={"company": company, "module_name": "薪酬", "template_key": "employee_salary_change"},
		fields=["name", "source_file", "status", "total_rows", "valid_rows", "imported_by", "imported_on", "notes", "modified"],
		order_by="modified desc",
		limit_page_length=cint(page_length) or 10,
	)
	for batch in batches:
		batch["affected_rows"] = frappe.db.count(EMPLOYEE_SALARY_CHANGE_DOCTYPE, {"salary_import_batch": batch.name})
		batch["can_rollback"] = int(batch.status in {"已处理", "部分失败"})
	return batches


@frappe.whitelist()
def rollback_employee_salary_change_import_batch(batch_name: str, company: str, reason: str = ""):
	_require_payroll_master_manager()
	company = _require_company(company)
	if not batch_name or not frappe.db.exists(FORM_IMPORT_BATCH_DOCTYPE, batch_name):
		frappe.throw(_("员工定薪导入批次不存在。"))
	batch = frappe.get_doc(FORM_IMPORT_BATCH_DOCTYPE, batch_name)
	if batch.company != company or batch.module_name != "薪酬" or batch.template_key != "employee_salary_change":
		frappe.throw(_("导入批次与当前公司不匹配。"))
	_assert_salary_change_batch_reversible(batch)
	reason = _text(reason).strip()
	if not reason:
		frappe.throw(_("请填写撤销原因。"))
	try:
		rollback = json.loads(batch.mapping_json or "{}")
	except (TypeError, ValueError):
		frappe.throw(_("该批次缺少可验证的撤销快照，无法安全撤销。"))
	created, updated = rollback.get("created") or {}, rollback.get("updated") or {}
	changed_after_import = []
	for name, signature in created.items():
		if not frappe.db.exists(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name):
			continue
		doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name)
		if doc.salary_import_batch != batch.name or _employee_salary_change_signature(doc) != signature:
			changed_after_import.append(name)
	for name, state in updated.items():
		if not frappe.db.exists(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name):
			changed_after_import.append(name)
			continue
		doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name)
		if doc.salary_import_batch != batch.name or _employee_salary_change_signature(doc) != state.get("after_signature"):
			changed_after_import.append(name)
	if changed_after_import:
		frappe.throw(_("以下记录在导入后已被修改，不能自动撤销：{0}").format("、".join(changed_after_import[:10])))
	invalidations = [_invalidate_unconfirmed_payroll_trial(company, month, reason=_("撤销员工定薪导入批次 {0}").format(batch.name)) for month in _salary_change_batch_months(batch.name)]
	for name in created:
		if frappe.db.exists(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name):
			frappe.delete_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name, ignore_permissions=True, force=True)
	for name, state in updated.items():
		if not frappe.db.exists(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name):
			continue
		doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name)
		doc.update(state.get("before") or {})
		doc.save(ignore_permissions=True)
	batch.status = "已作废"
	batch.error_summary = _("撤销原因：{0}").format(reason)
	batch.notes = "{0}\n{1}".format(batch.notes or "", _("已于 {0} 撤销，新建记录已删除，覆盖记录已恢复。 ").format(now_datetime())).strip()
	batch.save(ignore_permissions=True)
	frappe.db.commit()
	return {
		"batch": batch.name,
		"created_rows_removed": len(created),
		"updated_rows_restored": len(updated),
		"invalidated_trials": invalidations,
	}


@frappe.whitelist()
def preview_payroll_data_closure_workbook(file_url: str):
	workbook = _load_workbook(file_url)
	sheets = []
	for template in PAYROLL_IMPORT_TEMPLATES:
		sheet = _matching_sheet(workbook, template["sheet_name"])
		if not sheet:
			sheets.append({"sheet_name": template["sheet_name"], "target_doctype": template["target_doctype"], "found": False, "row_count": 0})
			continue
		rows = _template_rows_as_dicts(sheet, template)
		sheets.append({"sheet_name": template["sheet_name"], "target_doctype": template["target_doctype"], "found": True, "row_count": len(rows), "sample_rows": rows[:3]})
	settlement_preview = preview_payroll_settlement_workbook(file_url)
	sheets.append(
		{
			"sheet_name": "完整薪资结算表",
			"target_doctype": PAYROLL_SETTLEMENT_DOCTYPE,
			"found": settlement_preview.get("found"),
			"row_count": settlement_preview.get("row_count", 0),
			"sample_rows": settlement_preview.get("sample_rows", []),
		}
	)
	return {"sheets": sheets}


def _upsert_employee_salary_change_from_row(row, payroll_month="", company="", source_file="Excel导入", salary_import_batch=""):
	company = _require_company(company)
	employee_code = _first(row, "工号", "员工编号", "employee_code")
	employee_name = _first(row, "姓名", "employee_name")
	if not employee_code and not employee_name:
		return None
	employee = _payroll_employee_lookup(employee_code, employee_name)
	employee_context = _employee_context(employee)
	if employee_context.get("company") and employee_context.get("company") != company:
		frappe.throw(_("员工薪资异动导入存在跨公司员工：{0}").format(employee_code or employee_name or employee))
	effective_date = _date_or_none(_first(row, "生效日期")) or (f"{payroll_month}-01" if payroll_month else None)
	if not effective_date:
		frappe.throw(_("员工薪资异动导入缺少生效日期"))

	structure_version = _text(_first(row, "薪资架构版本", "版本"))
	salary_level = cint(_first(row, "薪资序号"))
	matched_grade, _ = _salary_grade_from_import_row(row, payroll_month)
	if not matched_grade and not (structure_version and salary_level):
		matched_grade = _salary_grade_from_matching_history(
			employee,
			_first(row, "底薪"),
			_first(row, "职能津贴"),
			effective_date,
		)
	base_salary = flt(matched_grade.base_salary) if matched_grade else flt(_first(row, "底薪"))
	function_allowance = flt(matched_grade.function_allowance) if matched_grade else flt(_first(row, "职能津贴"))
	certificate_allowance = flt(_first(row, "证书津贴", "证书及多能工津贴"))
	multi_skill_allowance = flt(_first(row, "多能工津贴"))
	# Keep the imported source value as a traceable source row, but calculate the
	# payroll hourly base with the current policy: certificate and multi-skill
	# allowances are paid as monthly bonuses, not fixed full salary.
	full_salary = base_salary + function_allowance
	salary_grade = matched_grade.name if matched_grade else _first(row, "薪资档位")
	# A value not resolved above can be a legacy label rather than a grade DocType
	# name, so never create a broken Link while importing otherwise valid pay.
	if salary_grade and not frappe.db.exists(SALARY_GRADE_DOCTYPE, salary_grade):
		salary_grade = ""
	designation = _first(row, "岗位") or employee_context.get("designation")
	if designation and not frappe.db.exists("Designation", designation):
		designation = employee_context.get("designation") or ""
	values = {
		"company": company,
		"employee": employee,
		"employee_code": employee_code or employee,
		"employee_name": employee_name or employee_context.get("employee_name"),
		"department": _department_lookup(_first(row, "部门")) or employee_context.get("department"),
		"designation": designation,
		"date_of_joining": employee_context.get("date_of_joining"),
		"effective_date": effective_date,
		"salary_grade": salary_grade,
		"base_salary": base_salary,
		"function_allowance": function_allowance,
		"certificate_allowance": certificate_allowance,
		"multi_skill_allowance": multi_skill_allowance,
		"full_salary": full_salary,
		"housing_fund_enabled": int(bool(flt(_first(row, "住房公积金", "公积金")) or _bool_value(_first(row, "住房公积金", "公积金")))),
		"social_insurance_enabled": int(bool(flt(_first(row, "社保费用", "社保")) or _bool_value(_first(row, "社保费用", "社保")))),
		"company_cost_total": flt(_first(row, "公司总承担")),
		# Ignore legacy spreadsheet status values: imports are directly submitted.
		"status": "已批准",
		"salary_import_batch": salary_import_batch or "",
		"source_file": source_file or "Excel导入",
		"remarks": _first(row, "备注"),
	}
	name = frappe.db.get_value(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		{"company": company, "employee_code": values["employee_code"], "effective_date": values["effective_date"]},
		"name",
	)
	if name:
		doc = frappe.get_doc(EMPLOYEE_SALARY_CHANGE_DOCTYPE, name)
		doc.update(values)
		doc.save(ignore_permissions=True)
	else:
		doc = frappe.get_doc({"doctype": EMPLOYEE_SALARY_CHANGE_DOCTYPE, **values})
		doc.insert(ignore_permissions=True)
	return doc.name


def _upsert_welfare_source_from_row(row, payroll_month="", company="", attendance_lock_version=""):
	month = _first(row, "薪资月份") or payroll_month
	company, month, attendance_lock_version = _require_payroll_scope(company, month, attendance_lock_version)
	return upsert_payroll_welfare_source_record(
		company=company,
		payroll_month=month,
		attendance_lock_version=attendance_lock_version,
		source_type=_first(row, "来源类型"),
		employee_code=_first(row, "工号", "员工编号"),
		employee_name=_first(row, "姓名"),
		department=_first(row, "部门"),
		amount=flt(_first(row, "金额")),
		direction=_first(row, "方向"),
		eligibility_status=_first(row, "资格状态") or "符合",
		confirmation_status=_first(row, "确认状态") or "已确认",
		source_reference=_first(row, "来源单据/说明", "来源说明"),
		remarks=_first(row, "备注"),
	)


def _upsert_attendance_summary_from_row(row, payroll_month="", company="", attendance_lock_version=""):
	attendance_month = _first(row, "考勤月份", "薪资月份") or payroll_month
	company, attendance_month, attendance_lock_version = _require_payroll_scope(company, attendance_month, attendance_lock_version)
	employee_code = _first(row, "工号", "员工编号")
	employee_name = _first(row, "姓名")
	if not employee_code and not employee_name:
		return None
	employee = _employee_lookup(employee_code, employee_name)
	employee_context = _employee_context(employee)
	if employee_context.get("company") and employee_context.get("company") != company:
		frappe.throw(_("月度考勤终稿导入存在跨公司员工：{0}").format(employee_code or employee_name or employee))
	values = {
		"company": company,
		"attendance_month": attendance_month,
		"attendance_lock_version": attendance_lock_version,
		"employee": employee,
		"employee_code": employee_code or employee,
		"employee_name": employee_name or employee_context.get("employee_name"),
		"department": _department_lookup(_first(row, "部门")) or employee_context.get("department"),
		"date_of_joining": _date_or_none(_first(row, "入职日期")) or employee_context.get("date_of_joining"),
		"standard_hours": flt(_first(row, "标准工时")),
		"actual_attendance_hours": flt(_first(row, "基本出勤工时", "实际出勤工时")),
		"adjusted_working_hours": flt(_first(row, "调整后工时")),
		"overtime_1_5_hours": flt(_first(row, "1.5倍加班", "平日加班")),
		"overtime_2_hours": flt(_first(row, "2倍加班", "周末加班")),
		"overtime_3_hours": flt(_first(row, "3倍加班", "节假日加班")),
		"leave_hours": flt(_first(row, "请假工时")),
		"absent_hours": flt(_first(row, "旷工工时")),
		"large_night_shift_count": flt(_first(row, "大夜班次数")),
		"small_night_shift_count": flt(_first(row, "小夜班次数")),
		"apple_reward_amount": flt(_first(row, "红绿苹果金额", "苹果树金额")),
		"full_attendance_deduction": flt(_first(row, "全勤扣款")),
		"status": _first(row, "状态") or "已确认",
		"lock_status": _first(row, "锁定状态") or "已锁定",
		"locked_by": frappe.session.user,
		"locked_on": now_datetime(),
	}
	name = frappe.db.get_value(
		MONTHLY_ATTENDANCE_DOCTYPE,
		{"company": company, "attendance_month": attendance_month, "attendance_lock_version": attendance_lock_version, "employee_code": values["employee_code"]},
		"name",
	)
	if name:
		doc = frappe.get_doc(MONTHLY_ATTENDANCE_DOCTYPE, name)
		doc.update(values)
		doc.save(ignore_permissions=True)
	else:
		doc = frappe.get_doc({"doctype": MONTHLY_ATTENDANCE_DOCTYPE, **values})
		doc.insert(ignore_permissions=True)
	return doc.name


def _settlement_cell(row, index):
	return row[index - 1] if index - 1 < len(row) else None


def _settlement_float(row, index):
	return flt(_settlement_cell(row, index))


def _payroll_settlement_sheet(workbook):
	return _matching_sheet(workbook, PAYROLL_SETTLEMENT_IMPORT_SHEET)


def _parse_settlement_sheet_rows(sheet):
	rows = []
	for raw in sheet.iter_rows(min_row=4, values_only=True):
		department = _text(_settlement_cell(raw, 2))
		employee_code = _text(_settlement_cell(raw, 3))
		employee_name = _text(_settlement_cell(raw, 4))
		if not employee_code and not employee_name:
			continue
		if department == "小计" or employee_code == "小计" or employee_name == "小计":
			continue
		rows.append(
			{
				"department": department,
				"employee_code": employee_code,
				"employee_name": employee_name,
				"base_salary": _settlement_float(raw, 5),
				"function_allowance": _settlement_float(raw, 6),
				"certificate_skill_allowance": _settlement_float(raw, 7),
				"salary_subtotal": _settlement_float(raw, 8),
				"standard_hours": _settlement_float(raw, 9),
				"basic_attendance_hours": _settlement_float(raw, 10),
				"missing_hours": _settlement_float(raw, 11),
				"raw_weekend_overtime_hours": _settlement_float(raw, 12),
				"adjusted_absence_hours": _settlement_float(raw, 13),
				"absence_deduction_amount": _settlement_float(raw, 14),
				"weekend_overtime_hours": _settlement_float(raw, 15),
				"weekday_overtime_hours": _settlement_float(raw, 16),
				"holiday_overtime_hours": _settlement_float(raw, 17),
				"weekday_overtime_pay": _settlement_float(raw, 18),
				"weekend_overtime_pay": _settlement_float(raw, 19),
				"holiday_overtime_pay": _settlement_float(raw, 20),
				"overtime_pay_total": _settlement_float(raw, 21),
				"large_night_shift_count": _settlement_float(raw, 22),
				"small_night_shift_count": _settlement_float(raw, 23),
				"night_shift_allowance": _settlement_float(raw, 24),
				"attendance_wage": _settlement_float(raw, 25),
				"proposal_improvement_bonus": _settlement_float(raw, 26),
				"apple_reward_amount": _settlement_float(raw, 27),
				"subsidy_bonus_total": _settlement_float(raw, 28),
				"production_bonus": _settlement_float(raw, 29),
				"bonus_total": _settlement_float(raw, 30),
				"absenteeism_hours": _settlement_float(raw, 31),
				"absenteeism_deduction": _settlement_float(raw, 32),
				"late_full_attendance_deduction": _settlement_float(raw, 33),
				"punishment_total": _settlement_float(raw, 34),
				"gross_pay": _settlement_float(raw, 35),
				"social_security_personal": _settlement_float(raw, 36),
				"housing_fund_personal": _settlement_float(raw, 37),
				"paid_proposal_birthday_welfare": _settlement_float(raw, 38),
				"taxable_salary": _settlement_float(raw, 39),
				"continuing_service_bonus": _settlement_float(raw, 40),
				"income_tax": _settlement_float(raw, 42),
				"year_end_bonus_tax": _settlement_float(raw, 43),
				"utilities_deduction": _settlement_float(raw, 44),
				"net_pay": _settlement_float(raw, 45),
				"social_security_company": _settlement_float(raw, 46),
				"housing_fund_company": _settlement_float(raw, 47),
				"company_cost_total": _settlement_float(raw, 48),
				"export_tax_adjusted_net_pay": _settlement_float(raw, 57),
			}
		)
	return rows


def _upsert_by_employee_month(doctype, month_field, payroll_month, employee_code, values, company="", attendance_lock_version=""):
	company = _require_company(company or values.get("company"))
	filters = {"company": company, month_field: payroll_month}
	if attendance_lock_version:
		filters["attendance_lock_version"] = attendance_lock_version
	if employee_code:
		filters["employee_code"] = employee_code
	if values.get("employee_name"):
		filters["employee_name"] = values.get("employee_name")
	name = frappe.db.get_value(doctype, filters, "name")
	if name:
		doc = frappe.get_doc(doctype, name)
		doc.update(values)
		doc.save(ignore_permissions=True)
	else:
		doc = frappe.get_doc({"doctype": doctype, **values})
		doc.insert(ignore_permissions=True)
	return doc.name


def _insert_settlement_welfare_source(company, payroll_month, attendance_lock_version, settlement_row, source_type, amount, reference):
	if not flt(amount):
		return None
	return upsert_payroll_welfare_source_record(
		company=company,
		payroll_month=payroll_month,
		attendance_lock_version=attendance_lock_version,
		source_type=source_type,
		employee_code=settlement_row.get("employee_code"),
		employee_name=settlement_row.get("employee_name"),
		department=settlement_row.get("department"),
		amount=amount,
		eligibility_status="符合",
		confirmation_status="已确认",
		source_reference=reference,
		source_file=PAYROLL_SETTLEMENT_IMPORT_SOURCE,
	)


def _upsert_sources_from_settlement_row(company, payroll_month, attendance_lock_version, row):
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	employee = _payroll_employee_lookup(row.get("employee_code"), row.get("employee_name"))
	employee_context = _employee_context(employee)
	if employee_context.get("company") and employee_context.get("company") != company:
		frappe.throw(_("完整薪资结算表导入存在跨公司员工：{0}").format(row.get("employee_code") or row.get("employee_name") or employee))
	full_salary = row.get("salary_subtotal") or row.get("base_salary") + row.get("function_allowance") + row.get("certificate_skill_allowance")
	salary_values = {
		"company": company,
		"employee": employee,
		"employee_code": row.get("employee_code"),
		"employee_name": row.get("employee_name") or employee_context.get("employee_name"),
		"department": _ensure_department(row.get("department")) or employee_context.get("department"),
		"designation": employee_context.get("designation"),
		"date_of_joining": employee_context.get("date_of_joining"),
		"effective_date": f"{payroll_month}-01",
		"change_reason": PAYROLL_SETTLEMENT_IMPORT_SOURCE,
		"base_salary": row.get("base_salary"),
		"function_allowance": row.get("function_allowance"),
		"certificate_allowance": row.get("certificate_skill_allowance"),
		"multi_skill_allowance": 0,
		"full_salary": full_salary,
		"status": "已批准",
		"source_file": PAYROLL_SETTLEMENT_IMPORT_SOURCE,
	}
	_upsert_by_employee_month(EMPLOYEE_SALARY_CHANGE_DOCTYPE, "effective_date", f"{payroll_month}-01", row.get("employee_code"), salary_values, company)

	attendance_values = {
		"company": company,
		"attendance_month": payroll_month,
		"attendance_lock_version": attendance_lock_version,
		"employee": employee,
		"employee_code": row.get("employee_code"),
		"employee_name": row.get("employee_name") or employee_context.get("employee_name"),
		"department": _ensure_department(row.get("department")),
		"date_of_joining": employee_context.get("date_of_joining"),
		"standard_hours": row.get("standard_hours"),
		"actual_attendance_hours": row.get("basic_attendance_hours"),
		"adjusted_working_hours": row.get("standard_hours") - row.get("adjusted_absence_hours"),
		"overtime_1_5_hours": row.get("weekday_overtime_hours"),
		"overtime_2_hours": row.get("raw_weekend_overtime_hours"),
		"overtime_3_hours": row.get("holiday_overtime_hours"),
		"absent_hours": row.get("absenteeism_hours"),
		"large_night_shift_count": row.get("large_night_shift_count"),
		"small_night_shift_count": row.get("small_night_shift_count"),
		"apple_reward_amount": row.get("apple_reward_amount"),
		"full_attendance_deduction": row.get("late_full_attendance_deduction"),
		"status": "已确认",
		"lock_status": "已锁定",
		"locked_by": frappe.session.user,
		"locked_on": now_datetime(),
	}
	_upsert_by_employee_month(MONTHLY_ATTENDANCE_DOCTYPE, "attendance_month", payroll_month, row.get("employee_code"), attendance_values, company, attendance_lock_version)

	input_values = {
		"company": company,
		"payroll_month": payroll_month,
		"attendance_lock_version": attendance_lock_version,
		"employee": employee,
		"employee_code": row.get("employee_code"),
		"employee_name": row.get("employee_name") or employee_context.get("employee_name"),
		"department": _ensure_department(row.get("department")),
		"date_of_joining": employee_context.get("date_of_joining"),
		"standard_hours": row.get("standard_hours"),
		"actual_attendance_hours": row.get("basic_attendance_hours"),
		"adjusted_working_hours": row.get("standard_hours") - row.get("adjusted_absence_hours"),
		"overtime_1_5_hours": row.get("weekday_overtime_hours"),
		"overtime_2_hours": row.get("raw_weekend_overtime_hours"),
		"overtime_3_hours": row.get("holiday_overtime_hours"),
		"absent_hours": row.get("absenteeism_hours"),
		"large_night_shift_count": row.get("large_night_shift_count"),
		"small_night_shift_count": row.get("small_night_shift_count"),
		"apple_reward_amount": row.get("apple_reward_amount"),
		"full_attendance_bonus": row.get("subsidy_bonus_total"),
		"social_security_personal": row.get("social_security_personal"),
		"housing_fund_personal": row.get("housing_fund_personal"),
		"other_bonus": row.get("proposal_improvement_bonus") + row.get("production_bonus"),
		"other_deduction": row.get("late_full_attendance_deduction"),
		"preliminary_earning_total": row.get("bonus_total"),
		"preliminary_deduction_total": row.get("social_security_personal") + row.get("housing_fund_personal") + row.get("utilities_deduction"),
		"settlement_status": "已生成工资表",
	}
	_upsert_by_employee_month(PAYROLL_INPUT_DOCTYPE, "payroll_month", payroll_month, row.get("employee_code"), input_values, company, attendance_lock_version)

	for source_type, fieldname, reference in (
		("提案改善奖", "proposal_improvement_bonus", "完整薪资结算表导入: 提案改善奖"),
		("其他奖金", "subsidy_bonus_total", "完整薪资结算表导入: 全勤奖,住房学历补贴"),
		("生产奖", "production_bonus", "完整薪资结算表导入: 生产奖"),
		("社保个人", "social_security_personal", "完整薪资结算表导入: 保险基金员工负担额"),
		("公积金个人", "housing_fund_personal", "完整薪资结算表导入: 住房公积金"),
		("已发福利", "paid_proposal_birthday_welfare", "完整薪资结算表导入: 提案改善奖&生日福利金已发"),
		("继续服务奖", "continuing_service_bonus", "完整薪资结算表导入: 继续服务奖"),
		("所得税", "income_tax", "完整薪资结算表导入: 所得税代扣款"),
		("年终奖所得税", "year_end_bonus_tax", "完整薪资结算表导入: 年终奖所得税"),
		("水电费及扣款", "utilities_deduction", "完整薪资结算表导入: 水电费及扣款"),
		("社保公司", "social_security_company", "完整薪资结算表导入: 保险基金公司负担额"),
		("公积金公司", "housing_fund_company", "完整薪资结算表导入: 住房公积金公司负担"),
	):
		_insert_settlement_welfare_source(company, payroll_month, attendance_lock_version, row, source_type, row.get(fieldname), reference)


@frappe.whitelist()
def preview_payroll_settlement_workbook(file_url: str):
	workbook = _load_workbook(file_url)
	sheet = _payroll_settlement_sheet(workbook)
	if not sheet:
		return {"found": False, "sheet_name": PAYROLL_SETTLEMENT_IMPORT_SHEET, "row_count": 0, "sample_rows": []}
	rows = _parse_settlement_sheet_rows(sheet)
	return {
		"found": True,
		"sheet_name": sheet.title,
		"row_count": len(rows),
		"sample_rows": rows[:5],
		"totals": {
			"gross_pay": round(sum(flt(row.get("gross_pay")) for row in rows), 2),
			"net_pay": round(sum(flt(row.get("net_pay")) for row in rows), 2),
			"company_cost_total": round(sum(flt(row.get("company_cost_total")) for row in rows), 2),
		},
	}


@frappe.whitelist()
def import_payroll_settlement_workbook(file_url: str, payroll_month: str = "", company: str = "", attendance_lock_version: str = ""):
	_require_payroll_scope(company, payroll_month or datetime.today().strftime("%Y-%m"), attendance_lock_version)
	frappe.throw(_("完整薪资结算表只能用于预览核对，不允许从 Excel 终稿直接覆盖薪资结算；请先导入/确认同公司同锁定版本变量，再由系统生成薪资输入表和薪资结算表。"))


@frappe.whitelist()
def import_payroll_data_closure_workbook(file_url: str, payroll_month: str = "", company: str = "", attendance_lock_version: str = ""):
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month or datetime.today().strftime("%Y-%m"), attendance_lock_version)
	workbook = _load_workbook(file_url)
	result = {"created_or_updated": defaultdict(int), "skipped_sheets": []}
	importers = {
		"employee_salary_change": _upsert_employee_salary_change_from_row,
		"welfare_source": _upsert_welfare_source_from_row,
		"monthly_attendance_summary": _upsert_attendance_summary_from_row,
	}
	for template in PAYROLL_IMPORT_TEMPLATES:
		sheet = _matching_sheet(workbook, template["sheet_name"])
		if not sheet:
			result["skipped_sheets"].append(template["sheet_name"])
			continue
		rows = _template_rows_as_dicts(sheet, template)
		importer = importers.get(template["template_key"])
		for row in rows:
			month = _first(row, "薪资月份", "考勤月份") or payroll_month
			if template["template_key"] == "employee_salary_change":
				name = importer(row, month, company)
			else:
				name = importer(row, month, company, attendance_lock_version)
			if name:
				result["created_or_updated"][template["sheet_name"]] += 1
	if _payroll_settlement_sheet(workbook):
		result["skipped_sheets"].append("完整薪资结算表（仅预览核对，不直接写入结算）")
	frappe.db.commit()
	result["created_or_updated"] = dict(result["created_or_updated"])
	return result


def _payroll_variable_validation(company, employee_code="", employee_name="", amount=0):
	errors = []
	warnings = []
	if not employee_code and not employee_name:
		errors.append("缺少工号和姓名")
	employee = _payroll_employee_lookup(employee_code, employee_name) if employee_code or employee_name else None
	matched_by_code = _employee_lookup(employee_code, "") if employee_code else None
	matched_by_name = _employee_lookup("", employee_name) if employee_name else None
	identity_conflict = bool(employee_code and employee_name and matched_by_code and matched_by_name and matched_by_code != matched_by_name)
	if identity_conflict:
		name_by_code = _employee_context(matched_by_code).get("employee_name")
		errors.append("工号与姓名不一致：工号对应姓名为 {0}".format(name_by_code or "-"))
	elif (employee_code or employee_name) and not employee:
		errors.append("无法匹配员工花名册")
	if employee and company:
		employee_context = _employee_context(employee)
		if employee_context.get("company") and employee_context.get("company") != company:
			errors.append("员工不属于当前公司")
	if not flt(amount):
		warnings.append("金额为 0，请在确认入账前核对")
	status = "错误" if errors else ("警告" if warnings else "通过")
	return {"employee": employee, "status": status, "errors": errors, "warnings": warnings, "message": "；".join(errors or warnings)}


def _payroll_variable_preview_row(company, sheet_name, source_kind, row, variable_type, amount, note=""):
	employee_code = _text(_first(row, "工号", "受奖惩人工号"))
	employee_name = _first(row, "姓名", "受奖/惩人", "受奖惩人姓名")
	validation = _payroll_variable_validation(company, employee_code, employee_name, amount)
	if note:
		validation["status"] = "错误"
		validation["errors"].append(note)
		validation["message"] = "；".join(validation["errors"])
	return {
		"sheet_name": sheet_name,
		"source_kind": source_kind,
		"employee_code": employee_code,
		"employee_name": employee_name,
		"variable_type": variable_type,
		"amount": round(flt(amount), 2),
		"validation_status": validation["status"],
		"validation_message": validation["message"],
	}


def _housing_base_preview_row(company, payroll_month, sheet_name, row):
	calculation = _housing_allowance_calculation(row, payroll_month, company)
	preview = _payroll_variable_preview_row(
		company,
		sheet_name,
		"housing_allowance_base",
		row,
		"住房补贴",
		calculation["amount"],
		calculation.get("error") or "",
	)
	preview.update(
		{
			"calculation_mode": calculation["mode"],
			"calculation_reason": calculation["reason"],
			"participation_status": "不参与计算" if calculation["excluded"] else "待确认",
		}
	)
	return preview


@frappe.whitelist()
def preview_payroll_variable_workbook(file_url: str, company: str = "", payroll_month: str = ""):
	company = _require_company(company) if company else ""
	payroll_month = _workflow_month(payroll_month or datetime.today().strftime("%Y-%m"))
	workbook = _load_workbook(file_url)
	sheets = []
	preview_rows = []
	handled_sheets = set()
	blocked_sources = []
	for sheet_name in workbook.sheetnames:
		sheet = workbook[sheet_name]
		source_kind = _raw_payroll_source_kind(sheet)
		if not source_kind:
			continue
		handled_sheets.add(sheet.title)
		rows = _raw_payroll_source_rows(sheet, source_kind)
		if source_kind in {"attendance_bonus", "housing_allowance", "housing_allowance_base"}:
			blocked_sources.append(source_kind)
			is_housing = source_kind in {"housing_allowance", "housing_allowance_base"}
			sheets.append({
				"sheet_name": sheet.title,
				"source_kind": source_kind,
				"target_doctype": MONTHLY_ATTENDANCE_DOCTYPE,
				"found": True,
				"row_count": len(rows),
				"mapped_rows": 0,
				"review_rows": len(rows),
				"note": "住房补贴应在考勤补充来源中导入、校验并随终稿锁定；薪酬端不重复导入。" if is_housing else "全勤奖由考勤终稿锁定结果和全勤规则自动计算，薪酬端不重复导入。",
			})
			continue
		if source_kind == "housing_allowance_base":
			mapped_rows = 0
			review_rows = 0
			for row in rows:
				preview = _housing_base_preview_row(company, payroll_month, sheet.title, row)
				preview_rows.append(preview)
				mapped_rows += 1 if not preview.get("participation_status") == "不参与计算" and preview["validation_status"] != "错误" else 0
				review_rows += 1 if preview["validation_status"] == "错误" else 0
			sheet_note = "已识别为一阶基础数据：系统将按当前租房补贴规则生成二阶应发金额；不符合资格者会保留明细但不参与计算。"
		elif source_kind == "housing_allowance":
			entries = [entry for row in rows for entry in _raw_variable_entries(source_kind, row)]
			mapped_rows = sum(1 for _variable_type, amount in entries if flt(amount))
			review_rows = 0
			for row in rows:
				for variable_type, amount in _raw_variable_entries(source_kind, row):
					preview = _payroll_variable_preview_row(company, sheet.title, source_kind, row, variable_type, amount)
					preview.update({"calculation_mode": "二阶金额直用", "calculation_reason": "已检测到“住房补贴”金额列，仅校验员工与金额，确认后直接参与计算", "participation_status": "待确认"})
					preview_rows.append(preview)
			sheet_note = "已识别为二阶数据：已含住房补贴金额，不重新计算；通过校验并确认后直接使用。"
		elif source_kind == "salary_change":
			mapped_rows = sum(1 for row in rows if _first(row, "工号", "姓名"))
			review_rows = 0
			for row in rows:
				preview_rows.append(_payroll_variable_preview_row(company, sheet.title, source_kind, row, "员工定薪", _first(row, "调整后全薪", "调整后底薪", "底薪")))
		elif source_kind == "proposal":
			# A proposal may name several awardees but has only one total amount.  It
			# must be allocated explicitly before it can affect payroll.
			mapped_rows = 0
			review_rows = sum(1 for row in rows if sum(1 for key in ("奖励人1", "奖励人2", "奖励人3") if _first(row, key)) > 1)
			for row in rows:
				recipients = [name for key in ("奖励人1", "奖励人2", "奖励人3") if (name := _first(row, key))]
				for recipient in recipients:
					entry = {**row, "姓名": recipient}
					allocation_note = "多奖励人需为每人填写分配金额" if len(recipients) > 1 else ""
					amount = 0 if allocation_note else flt(_first(row, "奖金(元)", "奖金（元）", "金额"))
					preview_rows.append(_payroll_variable_preview_row(company, sheet.title, source_kind, entry, "提案改善奖", amount, allocation_note))
		else:
			entries = [entry for row in rows for entry in _raw_variable_entries(source_kind, row)]
			mapped_rows = sum(1 for _variable_type, amount in entries if flt(amount))
			review_rows = 0
			for row in rows:
				for variable_type, amount in _raw_variable_entries(source_kind, row):
					preview_rows.append(_payroll_variable_preview_row(company, sheet.title, source_kind, row, variable_type, amount))
		sheets.append({
			"sheet_name": sheet.title,
			"source_kind": source_kind,
			"target_doctype": EMPLOYEE_SALARY_CHANGE_DOCTYPE if source_kind == "salary_change" else VARIABLE_RECORD_DOCTYPE,
			"found": True,
			"row_count": len(rows),
			"mapped_rows": mapped_rows,
			"review_rows": review_rows,
			"note": sheet_note if source_kind in {"housing_allowance_base", "housing_allowance"} else ("提案改善有多个奖励人，需先分配各人金额" if review_rows else ""),
		})
	for sheet_name in PAYROLL_VARIABLE_SHEETS:
		sheet = _matching_sheet(workbook, sheet_name)
		if sheet and sheet.title in handled_sheets:
			continue
		rows = _rows_as_dicts(sheet) if sheet else []
		variable_type = SHEET_VARIABLE_TYPES[sheet_name]
		mapped_rows = sum(1 for row in rows if _amount_for_type(row, variable_type))
		sheets.append({"sheet_name": sheet_name, "found": bool(sheet), "row_count": len(rows), "mapped_rows": mapped_rows})
		if sheet:
			for row in rows:
				preview_rows.append(_payroll_variable_preview_row(company, sheet.title, "standard", row, variable_type, _amount_for_type(row, variable_type)))
	settlement_sheet = _payroll_settlement_sheet(workbook)
	if settlement_sheet:
		settlement_rows = _parse_settlement_sheet_rows(settlement_sheet)
		sheets.append({"sheet_name": "完整薪资结算表", "found": True, "row_count": len(settlement_rows), "mapped_rows": len(settlement_rows)})
	return {
		"sheets": sheets,
		"preview_rows": preview_rows,
		"found_sheets": [sheet["sheet_name"] for sheet in sheets if sheet["found"]],
		"valid_rows": sum(1 for row in preview_rows if row["validation_status"] == "通过"),
		"warning_rows": sum(1 for row in preview_rows if row["validation_status"] == "警告"),
		"error_rows": sum(1 for row in preview_rows if row["validation_status"] == "错误"),
		"blocked": bool(blocked_sources),
		"blocked_message": (
			"住房补贴属于考勤补充来源，请在考勤假期模块导入、校验并锁定；薪酬模块只继承终稿。"
			if any(source in {"housing_allowance", "housing_allowance_base"} for source in blocked_sources)
			else "全勤奖由已锁定考勤终稿和全勤规则自动计算，请勿在薪酬模块重复导入。"
		) if blocked_sources else "",
	}


def _insert_variable(batch_name, company, payroll_month, attendance_lock_version, sheet_name, row, variable_type=None, amount=None, allow_zero=False, review_status="待确认", validation_note=""):
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	variable_type = variable_type or SHEET_VARIABLE_TYPES[sheet_name]
	amount = round(_amount_for_type(row, variable_type) if amount is None else flt(amount), 2)
	if not amount and not allow_zero:
		return None
	employee_code = _text(_first(row, "工号", "受奖惩人工号"))
	employee_name = _first(row, "姓名", "受奖/惩人", "受奖惩人姓名")
	if not employee_code and not employee_name:
		return None
	employee = _payroll_employee_lookup(employee_code, employee_name)
	employee_context = _employee_context(employee)
	resolved_employee_code = employee_context.get("employee_code") or employee_code
	validation = _payroll_variable_validation(company, resolved_employee_code, employee_name, amount)
	if validation_note:
		validation["status"] = "错误"
		validation["errors"].append(validation_note)
		validation["message"] = "；".join(validation["errors"])
	trace_payload, trace_hash = _source_trace_hash(
		{
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
			"import_batch": batch_name,
			"source_sheet": sheet_name,
			"employee": employee or employee_code or employee_name,
			"raw_row": row,
		}
	)
	doc = frappe.get_doc(
		{
			"doctype": VARIABLE_RECORD_DOCTYPE,
			"import_batch": batch_name,
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
			"employee": employee,
			"employee_code": resolved_employee_code,
			"employee_name": employee_name,
			"department": _department_lookup(_first(row, "部门", "单位", "受奖/惩人部门")),
			"variable_type": variable_type,
			"amount": amount,
			"review_status": review_status,
			"validation_status": validation["status"],
			"validation_message": validation["message"],
			"excluded": 0,
			"source_sheet": sheet_name,
			"remarks": _first(row, "备注"),
			"raw_row_json": json.dumps(row, ensure_ascii=False, default=str),
			"source_trace_json": trace_payload,
			"source_hash": trace_hash,
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


@frappe.whitelist()
def import_payroll_variable_workbook(file_url: str, payroll_month: str = "", company: str = "", attendance_lock_version: str = "", source_type: str = ""):
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month or datetime.today().strftime("%Y-%m"))
	# Monthly additions/deductions belong to company + payroll month, not to one
	# attendance snapshot.  A later attendance re-lock must not require re-import.
	attendance_lock_version = _monthly_variable_scope(payroll_month)
	workbook = _load_workbook(file_url)
	if any(_raw_payroll_source_kind(workbook[sheet_name]) == "salary_change" for sheet_name in workbook.sheetnames):
		frappe.throw(_("薪资异动属于“员工定薪”，请在员工定薪区域完成预览、审核与批准；月度增减项不会直接写入员工定薪。"))
	if any(_raw_payroll_source_kind(workbook[sheet_name]) == "attendance_bonus" for sheet_name in workbook.sheetnames):
		frappe.throw(_("全勤奖由已锁定考勤终稿和全勤规则自动计算，请勿在薪酬模块重复导入。"))
	if any(_raw_payroll_source_kind(workbook[sheet_name]) in {"housing_allowance", "housing_allowance_base"} for sheet_name in workbook.sheetnames):
		frappe.throw(_("住房补贴属于“考勤补充来源”，请在考勤假期模块导入、校验并锁定；薪酬模块只继承已锁定考勤终稿。"))
	invalidation = _invalidate_unconfirmed_payroll_trial(
		company, payroll_month, reason=_("新增月度增减项导入批次")
	)
	batch_source_type = source_type if _doctype_exists(VARIABLE_SOURCE_TYPE_DOCTYPE) and frappe.db.exists(VARIABLE_SOURCE_TYPE_DOCTYPE, source_type) else ""
	batch = frappe.get_doc(
		{
			"doctype": VARIABLE_BATCH_DOCTYPE,
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
			"source_type": batch_source_type,
			"source_file": file_url,
			"is_selected": 1,
			"status": "待确认",
			"imported_by": frappe.session.user,
			"imported_on": now_datetime(),
		}
	)
	batch.insert(ignore_permissions=True)
	if batch_source_type and _doctype_has_field(VARIABLE_BATCH_DOCTYPE, "is_selected"):
		for name in frappe.get_all(
			VARIABLE_BATCH_DOCTYPE,
			filters={"company": company, "payroll_month": payroll_month, "source_type": batch_source_type, "name": ["!=", batch.name]},
			pluck="name",
			limit_page_length=1000,
		):
			frappe.db.set_value(VARIABLE_BATCH_DOCTYPE, name, "is_selected", 0, update_modified=False)

	created = []
	salary_changes = []
	requires_allocation = []
	handled_sheets = set()
	handled_source_kinds = set()
	for actual_sheet_name in workbook.sheetnames:
		sheet = workbook[actual_sheet_name]
		source_kind = _raw_payroll_source_kind(sheet)
		if not source_kind:
			continue
		handled_sheets.add(sheet.title)
		handled_source_kinds.add(source_kind)
		rows = _raw_payroll_source_rows(sheet, source_kind)
		if source_kind == "salary_change":
			continue
		if source_kind == "housing_allowance_base":
			for row in rows:
				calculation = _housing_allowance_calculation(row, payroll_month, company)
				entry = dict(row)
				entry["金额"] = calculation["amount"]
				entry["备注"] = "{0}；{1}".format(calculation["mode"], calculation["reason"])
				name = _insert_variable(
					batch.name,
					company,
					payroll_month,
					attendance_lock_version,
					f"{sheet.title}（一阶数据系统计算）",
					entry,
					"住房补贴",
					calculation["amount"],
					True,
					validation_note=calculation.get("error") or "",
				)
				if name:
					created.append(name)
					if calculation["excluded"]:
						frappe.db.set_value(
							VARIABLE_RECORD_DOCTYPE,
							name,
							{"excluded": 1, "review_status": "已剔除", "validation_status": "通过", "validation_message": calculation["reason"]},
							update_modified=False,
						)
			continue
		if source_kind == "proposal":
			for row in rows:
				recipients = [name for key in ("奖励人1", "奖励人2", "奖励人3") if (name := _first(row, key))]
				if len(recipients) > 1:
					requires_allocation.append(_first(row, "编号") or "提案改善")
					for recipient in recipients:
						entry = dict(row)
						entry["姓名"] = recipient
						entry["备注"] = "提案奖金总额 {0}；请为每位奖励人填写分配金额".format(flt(_first(row, "奖金(元)", "奖金（元）", "金额")))
						name = _insert_variable(batch.name, company, payroll_month, attendance_lock_version, sheet.title, entry, "提案改善奖", 0, True, validation_note="多奖励人需为每人填写分配金额")
						if name:
							created.append(name)
					continue
				# A single named awardee can be imported safely.  An unassigned proposal
				# is intentionally ignored rather than paying the proposer by assumption.
				if not recipients:
					continue
				entry = dict(row)
				entry["姓名"] = recipients[0]
				name = _insert_variable(batch.name, company, payroll_month, attendance_lock_version, sheet.title, entry, "提案改善奖", flt(_first(row, "奖金(元)", "奖金（元）", "金额")))
				if name:
					created.append(name)
			continue
		for row in rows:
			for variable_type, amount in _raw_variable_entries(source_kind, row):
				name = _insert_variable(batch.name, company, payroll_month, attendance_lock_version, sheet.title, row, variable_type, amount)
				if name:
					created.append(name)
	for sheet_name in PAYROLL_VARIABLE_SHEETS:
		sheet = _matching_sheet(workbook, sheet_name)
		if not sheet or sheet.title in handled_sheets:
			continue
		handled_sheets.add(sheet.title)
		for row in _rows_as_dicts(sheet):
			name = _insert_variable(batch.name, company, payroll_month, attendance_lock_version, sheet_name, row)
			if name:
				created.append(name)
	if not handled_sheets:
		frappe.throw(_("未识别到可导入的薪酬来源工作表。请核对所选来源类型、表头及模板说明后重新上传。"))
	if not batch_source_type and len(handled_source_kinds) == 1 and _doctype_exists(VARIABLE_SOURCE_TYPE_DOCTYPE):
		inferred_source_type = next(iter(handled_source_kinds))
		if frappe.db.exists(VARIABLE_SOURCE_TYPE_DOCTYPE, inferred_source_type):
			batch_source_type = inferred_source_type
			batch.source_type = inferred_source_type
	if batch_source_type and handled_source_kinds:
		selected_code = frappe.db.get_value(VARIABLE_SOURCE_TYPE_DOCTYPE, batch_source_type, "source_code") or batch_source_type
		accepted_source_kinds = {selected_code}
		if selected_code == "housing_allowance":
			accepted_source_kinds.add("housing_allowance_base")
		if not accepted_source_kinds.intersection(handled_source_kinds):
			actual_labels = "、".join(RAW_PAYROLL_SOURCE_LABELS.get(code, code) for code in sorted(handled_source_kinds))
			selected_label = frappe.db.get_value(VARIABLE_SOURCE_TYPE_DOCTYPE, batch_source_type, "source_name") or batch_source_type
			frappe.throw(_("所选来源类型为“{0}”，但文件识别为“{1}”，请更正选择后重新上传。").format(selected_label, actual_labels))
	batch.variable_rows = len(created)
	recognized_labels = "、".join(sorted(handled_sheets))
	batch.notes = _("已识别来源：{0}。上传后已完成校验；修正异常后一次确认入账。").format(recognized_labels)
	batch.save(ignore_permissions=True)
	settlement_result = {}
	if _payroll_settlement_sheet(workbook):
		frappe.throw(_("完整薪资结算表只能作为来源映射核对，不允许在变量导入时直接覆盖薪资结算。请使用数据闭环导入并指定公司、月份和锁定版本。"))
	frappe.db.commit()
	return {
		"batch": batch.name,
		"status": batch.status,
		"variable_rows": len(created),
		"salary_changes": len(salary_changes),
		"requires_allocation": requires_allocation,
		"invalidated_trial": invalidation,
		**settlement_result,
	}


def _source_file_label(file_url):
	return (file_url or "").split("/")[-1] or file_url or ""


def _safe_payroll_export_filename_part(value):
	"""Keep display labels intact while making generated Excel filenames portable."""
	label = re.sub(r'[\\/:*?"<>|]+', "、", _text(value)).strip(" .、")
	return label[:80] or _("薪资来源")


def _payroll_signature_export_source(batch):
	"""Return the configured source code/name when a batch may be signed by staff."""
	if not batch.source_type or not _doctype_exists(VARIABLE_SOURCE_TYPE_DOCTYPE):
		frappe.throw(_("该来源没有可导出的员工签字表。"))
	source = frappe.db.get_value(
		VARIABLE_SOURCE_TYPE_DOCTYPE,
		batch.source_type,
		["source_code", "source_name"],
		as_dict=True,
	)
	source_code = _text((source or {}).get("source_code") or batch.source_type)
	if source_code not in SIGNABLE_PAYROLL_SOURCE_CODES:
		frappe.throw(_("{0} 不需要生成员工签字表。").format((source or {}).get("source_name") or batch.source_type))
	return source_code, _text((source or {}).get("source_name") or source_code)


def _signature_sheet_rows(rows, source_code):
	"""Keep one sign-off line per person for contribution sources, per item otherwise."""
	if source_code not in {"social_insurance", "housing_fund"}:
		return [
			[
				index,
				_text(row.employee_name),
				_text(row.employee_code),
				_text(row.department),
				_text(row.variable_type),
				flt(row.amount, 2),
				_text(row.remarks),
				"",
				"",
			]
			for index, row in enumerate(rows, start=1)
		]

	personal_type = "社保个人" if source_code == "social_insurance" else "公积金个人"
	company_type = "社保公司" if source_code == "social_insurance" else "公积金公司"
	grouped = {}
	for row in rows:
		key = _text(row.employee) or _text(row.employee_code) or _text(row.employee_name) or row.name
		entry = grouped.setdefault(key, {
			"employee_name": _text(row.employee_name), "employee_code": _text(row.employee_code),
			"department": _text(row.department), "personal": 0, "company": 0, "remarks": [],
		})
		if row.variable_type == personal_type:
			entry["personal"] += flt(row.amount, 2)
		elif row.variable_type == company_type:
			entry["company"] += flt(row.amount, 2)
		if _text(row.remarks):
			entry["remarks"].append(_text(row.remarks))
	return [
		[index, item["employee_name"], item["employee_code"], item["department"], item["personal"], item["company"], "；".join(dict.fromkeys(item["remarks"])), "", ""]
		for index, item in enumerate(grouped.values(), start=1)
	]


def _contribution_department_export_rows(rows, source_code):
	"""Match the contribution screen's department totals for department exports."""
	personal_type = "社保个人" if source_code == "social_insurance" else "公积金个人"
	company_type = "社保公司" if source_code == "social_insurance" else "公积金公司"
	people = {}
	for row in rows:
		key = _text(row.employee) or _text(row.employee_code) or _text(row.employee_name) or row.name
		person = people.setdefault(key, {
			"department": _text(row.department), "personal": 0, "company": 0,
		})
		if row.variable_type == personal_type:
			person["personal"] += flt(row.amount, 2)
		elif row.variable_type == company_type:
			person["company"] += flt(row.amount, 2)

	departments = {}
	for person in people.values():
		# The employee master may store the leaf department as "部门 - 公司";
		# the screen intentionally shows the department portion only.
		department = re.sub(r"\s+-\s+[^-]+$", "", person["department"] or _("未填写部门")).strip()
		summary = departments.setdefault(department, {"department": department, "headcount": 0, "personal": 0, "company": 0})
		summary["headcount"] += 1
		summary["personal"] += person["personal"]
		summary["company"] += person["company"]

	data_rows = [
		[index, item["department"], item["headcount"], flt(item["company"], 2), flt(item["personal"], 2), flt(item["company"] + item["personal"], 2)]
		for index, item in enumerate(departments.values(), start=1)
	]
	company_total = sum(item["company"] for item in departments.values())
	personal_total = sum(item["personal"] for item in departments.values())
	data_rows.append([_("合计"), "", len(people), flt(company_total, 2), flt(personal_total, 2), flt(company_total + personal_total, 2)])
	return data_rows


@frappe.whitelist()
def download_payroll_source_signature_sheet(batch_name: str, company: str, payroll_month: str = "", export_view: str = "personal"):
	"""Export the selected source's personal sign-off lines or contribution department summary."""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter
	from frappe.utils.file_manager import save_file

	if not batch_name or not frappe.db.exists(VARIABLE_BATCH_DOCTYPE, batch_name):
		frappe.throw(_("导入批次不存在"))
	batch = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, batch_name)
	company = _require_company(company or batch.company)
	if batch.company != company or (payroll_month and batch.payroll_month != payroll_month):
		frappe.throw(_("导入批次与当前公司或月份不一致。"))
	if batch.status == "已作废":
		frappe.throw(_("已作废批次不能导出员工签字表。"))
	source_code, source_name = _payroll_signature_export_source(batch)
	rows = frappe.get_all(
		VARIABLE_RECORD_DOCTYPE,
		filters={"import_batch": batch.name, "company": company, "excluded": 0},
		fields=["name", "employee", "employee_name", "employee_code", "department", "variable_type", "amount", "remarks"],
		order_by="department asc, employee_name asc, modified asc",
		limit_page_length=100000,
	)
	if not rows:
		frappe.throw(_("当前来源没有可供员工签字的有效记录。"))

	contribution = source_code in {"social_insurance", "housing_fund"}
	export_view = _text(export_view).lower()
	department_export = contribution and export_view == "department"
	department_total_label = "五险合计（元）" if source_code == "social_insurance" else "个人及公司合计（元）"
	headers = (
		["序号", "部门", "人数", "公司承担（元）", "个人承担（元）", department_total_label]
		if department_export
		else (["序号", "姓名", "工号", "部门", "个人承担（元）", "公司承担（元）", "备注", "员工签字", "签字日期"] if contribution
			else ["序号", "姓名", "工号", "部门", "项目", "金额（元）", "备注", "员工签字", "签字日期"])
	)
	data_rows = _contribution_department_export_rows(rows, source_code) if department_export else _signature_sheet_rows(rows, source_code)
	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "部门汇总" if department_export else "员工签字确认"
	title = _("{0} {1}部门汇总表").format(batch.payroll_month, source_name) if department_export else _("{0} {1}员工签字确认表").format(batch.payroll_month, source_name)
	sheet.append([title])
	sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
	sheet["A1"].font = Font(name="新宋体", size=16, bold=True)
	sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
	sheet.row_dimensions[1].height = 28
	sheet.append(headers)
	fill = PatternFill("solid", fgColor="D9EAF7")
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for cell in sheet[2]:
		cell.font = Font(name="新宋体", bold=True)
		cell.fill = fill
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		cell.border = border
	for row in data_rows:
		sheet.append(row)
		for cell in sheet[sheet.max_row]:
			cell.font = Font(name="新宋体")
			cell.alignment = Alignment(vertical="center", wrap_text=True)
			cell.border = border
		sheet.row_dimensions[sheet.max_row].height = 32
	widths = [8, 20, 12, 18, 18, 20] if department_export else [8, 14, 16, 18, 18, 16, 30, 22, 16]
	for index, width in enumerate(widths, start=1):
		sheet.column_dimensions[get_column_letter(index)].width = width
	sheet.freeze_panes = "A3"
	sheet.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{max(sheet.max_row, 2)}"

	output = BytesIO()
	workbook.save(output)
	# Source labels such as “证书/多能工津贴” are valid on screen but `/` is a
	# path separator.  Sanitise only the saved filename, never the displayed
	# source name or the signature sheet title.
	filename_suffix = _("部门汇总") if department_export else _("员工签字确认表")
	filename = f"{batch.payroll_month}_{_safe_payroll_export_filename_part(source_name)}_{filename_suffix}.xlsx"
	file_doc = save_file(filename, output.getvalue(), None, None, is_private=1)
	return {"file_url": file_doc.file_url, "file_name": file_doc.file_name, "rows": len(data_rows), "export_view": "department" if department_export else "personal"}


@frappe.whitelist()
def list_payroll_variable_import_batches(company: str, payroll_month: str = "", attendance_lock_version: str = "", page_length: int = 20):
	company = _require_company(company)
	source_labels = {
		str(row.get("source_code") or row.get("name") or ""): str(row.get("source_name") or row.get("source_code") or row.get("name") or "")
		for row in list_payroll_variable_source_types()
	}
	filters = {"company": company}
	if payroll_month:
		filters["payroll_month"] = payroll_month
	batches = frappe.get_all(
		VARIABLE_BATCH_DOCTYPE,
		filters=filters,
		fields=["name", "company", "payroll_month", "attendance_lock_version", "source_type", "source_file", "is_selected", "status", "variable_rows", "imported_by", "imported_on", "reviewed_by", "reviewed_on", "confirmed_by", "confirmed_on", "voided_by", "voided_on", "void_reason", "replacement_batch", "notes", "modified"],
		order_by="imported_on desc, modified desc",
		limit_page_length=int(page_length or 20),
	)
	# Keep obsolete pre-workflow uploads out of the current source checklist.
	# They have no mapped source type, use the retired “已导入” state and contain
	# no rows; showing them alongside the ten current source files makes the
	# monthly checklist look as if it has extra inputs.  The documents remain in
	# the database for audit and can still be reached from the DocType list.
	batches = [
		batch for batch in batches
		if not str(batch.source_file or "").startswith("attendance-processing-final:")
		and not (not batch.source_type and batch.status == "已导入" and not cint(batch.variable_rows))
	]
	for batch in batches:
		rows = frappe.get_all(
			VARIABLE_RECORD_DOCTYPE,
			filters={"company": company, "payroll_month": batch.payroll_month, "import_batch": batch.name},
			fields=["source_sheet", "review_status", "validation_status", "validation_message", "excluded", "employee", "employee_code", "employee_name"],
			limit_page_length=1000,
		)
		source_sheets = sorted({row.source_sheet for row in rows if row.source_sheet})
		active_rows = [row for row in rows if not row.excluded]
		unmatched_rows = [row for row in active_rows if not row.employee]
		unmatched_people = {
			row.employee_code or _normalise(row.employee_name) or _("未提供员工标识")
			for row in unmatched_rows
		}
		unmatched_reasons = defaultdict(int)
		for row in unmatched_rows:
			reason = row.validation_message or _("无法匹配员工花名册")
			unmatched_reasons[reason] += 1
		batch["source_file_label"] = _source_file_label(batch.source_file)
		batch["source_type_label"] = source_labels.get(str(batch.source_type or ""), batch.source_type or _("未分类"))
		batch["source_sheets"] = "、".join(source_sheets)
		batch["actual_variable_rows"] = len(rows)
		batch["included_rows"] = len(active_rows)
		batch["excluded_rows"] = len(rows) - len(active_rows)
		batch["matched_rows"] = sum(1 for row in active_rows if row.employee)
		batch["unmatched_rows"] = len(unmatched_rows)
		batch["unmatched_people"] = len(unmatched_people)
		batch["unmatched_reason_summary"] = [
			{"reason": reason, "count": count}
			for reason, count in sorted(unmatched_reasons.items(), key=lambda item: (-item[1], item[0]))
		]
		batch["error_rows"] = sum(1 for row in active_rows if row.validation_status == "错误")
		batch["warning_rows"] = sum(1 for row in active_rows if row.validation_status == "警告")
		batch["confirmed_rows"] = sum(1 for row in active_rows if row.review_status == "已确认")
		batch["is_pending_confirmation"] = int(batch.status in PENDING_VARIABLE_BATCH_STATUSES)
		batch["can_confirm_empty"] = int(batch["is_pending_confirmation"] and not rows and bool(batch.source_type) and "已识别来源：" in str(batch.notes or ""))
		batch["can_confirm"] = int(batch["is_pending_confirmation"] and not batch["error_rows"] and (bool(rows) or batch["can_confirm_empty"]))
		batch["can_delete"] = int(batch.status not in {"已确认", "已作废"})
		batch["can_void"] = int(batch.status == "已确认")
	return batches


@frappe.whitelist()
def select_payroll_variable_import_batch(batch_name: str, company: str = "", payroll_month: str = ""):
	"""Choose one version of a monthly source for the top editing workspace.

	The selection is deliberately scoped to one company, month and source type so
	a newer upload never leaves two versions of (for example) social insurance
	looking active at the same time.
	"""
	if not batch_name or not frappe.db.exists(VARIABLE_BATCH_DOCTYPE, batch_name):
		frappe.throw(_("导入批次不存在"))
	batch = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, batch_name)
	company = _require_company(company or batch.company)
	if batch.company != company or (payroll_month and batch.payroll_month != payroll_month):
		frappe.throw(_("导入批次与当前公司或月份不一致。"))
	if not batch.source_type:
		frappe.throw(_("历史未分类批次不能作为本月使用版本，请重新按来源类型导入。"))
	if batch.status == "已作废":
		frappe.throw(_("已作废批次不能作为本月使用版本。"))
	if not _doctype_has_field(VARIABLE_BATCH_DOCTYPE, "is_selected"):
		frappe.throw(_("系统尚未更新“本月选定版本”字段，请完成站点迁移后重试。"))
	for name in frappe.get_all(
		VARIABLE_BATCH_DOCTYPE,
		filters={"company": company, "payroll_month": batch.payroll_month, "source_type": batch.source_type},
		pluck="name",
		limit_page_length=1000,
	):
		frappe.db.set_value(VARIABLE_BATCH_DOCTYPE, name, "is_selected", int(name == batch.name), update_modified=False)
	frappe.db.commit()
	return {"batch": batch.name, "source_type": batch.source_type, "message": _("已选择 {0} 作为本月使用版本，可在上方编辑明细。 ").format(_source_file_label(batch.source_file))}


@frappe.whitelist()
def create_editable_payroll_variable_batch_version(batch_name: str, company: str = "", payroll_month: str = ""):
	"""Clone a confirmed source into a new auditable, editable monthly version."""
	if not batch_name or not frappe.db.exists(VARIABLE_BATCH_DOCTYPE, batch_name):
		frappe.throw(_("导入批次不存在"))
	source_batch = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, batch_name)
	company = _require_company(company or source_batch.company)
	if source_batch.company != company or (payroll_month and source_batch.payroll_month != payroll_month):
		frappe.throw(_("导入批次与当前公司或月份不一致。"))
	if source_batch.status != "已确认" or not source_batch.source_type:
		frappe.throw(_("仅已确认且已分类的来源版本可以复制后编辑。"))
	if not _doctype_has_field(VARIABLE_BATCH_DOCTYPE, "is_selected"):
		frappe.throw(_("系统尚未更新“本月选定版本”字段，请完成站点迁移后重试。"))

	invalidation = _invalidate_unconfirmed_payroll_trial(
		company, source_batch.payroll_month, reason=_("基于已确认来源创建可编辑版本")
	)
	editable_batch = frappe.get_doc({
		"doctype": VARIABLE_BATCH_DOCTYPE,
		"company": company,
		"payroll_month": source_batch.payroll_month,
		"attendance_lock_version": source_batch.attendance_lock_version,
		"source_type": source_batch.source_type,
		"source_file": source_batch.source_file,
		"is_selected": 1,
			"status": "待确认",
		"imported_by": frappe.session.user,
		"imported_on": now_datetime(),
		"notes": _("基于已确认批次 {0} 创建的可编辑版本；原批次保持不变。").format(source_batch.name),
	})
	editable_batch.insert(ignore_permissions=True)
	for name in frappe.get_all(
		VARIABLE_BATCH_DOCTYPE,
		filters={"company": company, "payroll_month": source_batch.payroll_month, "source_type": source_batch.source_type, "name": ["!=", editable_batch.name]},
		pluck="name",
		limit_page_length=1000,
	):
		frappe.db.set_value(VARIABLE_BATCH_DOCTYPE, name, "is_selected", 0, update_modified=False)

	rows = frappe.get_all(VARIABLE_RECORD_DOCTYPE, filters={"import_batch": source_batch.name}, fields=["*"], limit_page_length=100000)
	for row in rows:
		values = {
			"doctype": VARIABLE_RECORD_DOCTYPE,
			"import_batch": editable_batch.name,
			"company": company,
			"payroll_month": source_batch.payroll_month,
			"attendance_lock_version": source_batch.attendance_lock_version,
			"employee": row.employee,
			"employee_code": row.employee_code,
			"employee_name": row.employee_name,
			"department": row.department,
			"variable_type": row.variable_type,
			"amount": row.amount,
			"review_status": "已剔除" if row.excluded else "待确认",
			"validation_status": row.validation_status,
			"validation_message": row.validation_message,
			"excluded": row.excluded,
			"source_sheet": row.source_sheet,
			"remarks": row.remarks,
			"raw_row_json": row.raw_row_json,
			"source_trace_json": row.source_trace_json,
			"source_hash": row.source_hash,
		}
		frappe.get_doc(values).insert(ignore_permissions=True)
	editable_batch.variable_rows = len(rows)
	editable_batch.save(ignore_permissions=True)
	frappe.db.commit()
	return {"batch": editable_batch.name, "source_type": editable_batch.source_type, "rows": len(rows), "invalidated_trial": invalidation, "message": _("已创建可修改版本；完成修改后确认入账即可。")}


def _assert_no_singleton_variable_conflicts(batch, rows):
	"""Block accidental double-charging for one-row-per-employee sources.

	A replacement upload intentionally makes its predecessor a non-current
	version of the *same* source.  That predecessor remains immutable audit
	evidence but must not block confirmation of its selected replacement.
	"""
	candidates = [row for row in rows if not row.excluded and row.variable_type in SINGLETON_MONTHLY_VARIABLE_TYPES]
	if not candidates:
		return
	seen = {}
	conflicts = []
	for row in candidates:
		identity = row.employee or row.employee_code or ("name:" + str(row.employee_name or ""))
		key = (identity, row.variable_type)
		if key in seen:
			conflicts.append(_("批次内重复：{0} / {1}").format(row.employee_code or row.employee_name or identity, row.variable_type))
		else:
			seen[key] = row.name
	existing = frappe.get_all(
		VARIABLE_RECORD_DOCTYPE,
		filters={
			"company": batch.company,
			"payroll_month": batch.payroll_month,
			"review_status": "已确认",
			"excluded": 0,
			"variable_type": ["in", sorted(SINGLETON_MONTHLY_VARIABLE_TYPES)],
			"import_batch": ["!=", batch.name],
		},
		fields=["name", "import_batch", "attendance_lock_version", "employee", "employee_code", "employee_name", "variable_type"],
		limit_page_length=100000,
	)
	allowed_versions = {_monthly_variable_scope(batch.payroll_month), batch.attendance_lock_version, ""}
	batch_selection = {
		row.name: row
		for row in frappe.get_all(
			VARIABLE_BATCH_DOCTYPE,
			filters={"company": batch.company, "payroll_month": batch.payroll_month, "status": "已确认"},
			fields=["name", "source_type", "is_selected"],
			limit_page_length=100000,
		)
	}
	for row in existing:
		if str(row.attendance_lock_version or "") not in allowed_versions:
			continue
		prior_batch = batch_selection.get(row.import_batch)
		if (
			batch.source_type
			and prior_batch
			and prior_batch.source_type == batch.source_type
			and not cint(prior_batch.is_selected)
		):
			continue
		identity = row.employee or row.employee_code or ("name:" + str(row.employee_name or ""))
		if (identity, row.variable_type) in seen:
			conflicts.append(_("与已确认批次 {0} 重复：{1} / {2}").format(row.import_batch, row.employee_code or row.employee_name or identity, row.variable_type))
	if conflicts:
		frappe.throw(_("发现不应重复入账的月度项目：{0}。请作废旧批次或剔除重复记录后再确认。").format("；".join(conflicts[:10])))


def _confirm_payroll_variable_import_batch(batch, company: str, confirm_empty: int = 0):
	"""Confirm one already-validated batch without adding a review/lock stage."""
	if batch.status == "已确认":
		return {"batch": batch.name, "status": batch.status, "confirmed_rows": batch.variable_rows or 0, "already_confirmed": 1}
	if batch.status not in PENDING_VARIABLE_BATCH_STATUSES:
		frappe.throw(_("导入批次 {0} 当前状态为 {1}，不能确认入账。").format(batch.name, batch.status or "-"))
	rows = frappe.get_all(VARIABLE_RECORD_DOCTYPE, filters={"import_batch": batch.name}, fields=["name", "employee", "employee_code", "employee_name", "variable_type", "validation_status", "validation_message", "excluded"], limit_page_length=100000)
	errors = [row for row in rows if not row.excluded and row.validation_status == "错误"]
	active = [row for row in rows if not row.excluded]
	if errors:
		frappe.throw(_("批次 {0} 仍有 {1} 条异常记录，请修正或剔除后再确认：{2}").format(batch.name, len(errors), "；".join((row.validation_message or row.name) for row in errors[:5])))
	can_confirm_empty = bool(batch.source_type and "已识别来源：" in str(batch.notes or ""))
	if not active and not (cint(confirm_empty) and can_confirm_empty):
		frappe.throw(_("批次没有可入账的月度增减项记录；如来源文件明确为本月无数据，请使用“确认无数据”。"))
	_assert_no_singleton_variable_conflicts(batch, active)
	invalidation = _invalidate_unconfirmed_payroll_trial(
		company, batch.payroll_month, reason=_("确认月度增减项批次 {0}").format(batch.name)
	)
	for row in active:
		frappe.db.set_value(VARIABLE_RECORD_DOCTYPE, row.name, {"review_status": "已确认", "excluded": 0}, update_modified=False)
	for row in rows:
		if row.excluded:
			frappe.db.set_value(VARIABLE_RECORD_DOCTYPE, row.name, "review_status", "已剔除", update_modified=False)
	batch.status = "已确认"
	batch.confirmed_by = frappe.session.user
	batch.confirmed_on = now_datetime()
	if not active:
		batch.notes = _("已人工确认来源文件本月无数据；保留批次和原文件用于追溯。")
	batch.save(ignore_permissions=True)
	message = _("已确认该来源本月无数据，原文件已留痕。") if not active else _("{0} 条月度增减项已确认，将参与下次薪资计算。").format(len(active))
	return {"batch": batch.name, "status": batch.status, "confirmed_rows": len(active), "invalidated_trial": invalidation, "message": message}


@frappe.whitelist()
def confirm_payroll_variable_import_batch(batch_name: str, company: str = "", payroll_month: str = "", confirm_empty: int = 0):
	if not batch_name or not frappe.db.exists(VARIABLE_BATCH_DOCTYPE, batch_name):
		frappe.throw(_("导入批次不存在"))
	batch = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, batch_name)
	company = _require_company(company or batch.company)
	if batch.company != company or (payroll_month and batch.payroll_month != payroll_month):
		frappe.throw(_("导入批次与当前公司或月份不一致。"))
	result = _confirm_payroll_variable_import_batch(batch, company, confirm_empty)
	frappe.db.commit()
	return result


@frappe.whitelist()
def confirm_all_payroll_variable_import_batches(company: str, payroll_month: str, attendance_lock_version: str = ""):
	"""Confirm every selected, error-free source in a month as one atomic action.

	The selected version is important: older replacement uploads stay available for
	audit/deletion and must never be silently brought into payroll by a bulk click.
	"""
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	filters = {"company": company, "payroll_month": payroll_month, "status": ["in", sorted(PENDING_VARIABLE_BATCH_STATUSES)]}
	batches = frappe.get_all(VARIABLE_BATCH_DOCTYPE, filters=filters, fields=["name", "company", "payroll_month", "attendance_lock_version", "source_type", "source_file", "is_selected", "status"], order_by="imported_on asc, name asc", limit_page_length=1000)
	allowed_versions = {_monthly_variable_scope(payroll_month), attendance_lock_version, ""}
	batches = [
		frappe.get_doc(VARIABLE_BATCH_DOCTYPE, row.name)
		for row in batches
		if cint(row.is_selected)
		and str(row.attendance_lock_version or "") in allowed_versions
		and not str(row.source_file or "").startswith("attendance-processing-final:")
	]
	if not batches:
		frappe.throw(_("当前月份没有可一键确认的已录入批次。"))

	# No commit occurs inside the loop.  Any invalid row or duplicate therefore
	# rolls the whole request back instead of leaving a partially confirmed month.
	results = [_confirm_payroll_variable_import_batch(batch, company, confirm_empty=1) for batch in batches]
	frappe.db.commit()
	confirmed_rows = sum(cint(result.get("confirmed_rows")) for result in results)
	return {
		"batches": [result["batch"] for result in results],
		"confirmed_batches": len(results),
		"confirmed_rows": confirmed_rows,
		"message": _("已一键确认 {0} 个来源批次、{1} 条月度增减项。").format(len(results), confirmed_rows),
	}


@frappe.whitelist()
def void_payroll_variable_import_batch(batch_name: str, company: str = "", payroll_month: str = "", reason: str = "", replacement_batch: str = ""):
	"""Void a confirmed batch without deleting its audit evidence.

	A void changes participation status only.  The original file, parsed rows,
	confirmation details and the operator's reason remain available for audit.
	"""
	if not _can_manage_payroll_rules():
		frappe.throw(_("仅系统管理员或人事管理员可作废已确认的月度增减项批次。"))
	if not batch_name or not frappe.db.exists(VARIABLE_BATCH_DOCTYPE, batch_name):
		frappe.throw(_("导入批次不存在"))
	batch = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, batch_name)
	company = _require_company(company or batch.company)
	if batch.company != company or (payroll_month and batch.payroll_month != payroll_month):
		frappe.throw(_("导入批次与当前公司或月份不一致。"))
	if batch.status != "已确认":
		frappe.throw(_("仅已确认入账的批次可作废。"))
	if str(batch.source_file or "").startswith("attendance-processing-final:"):
		frappe.throw(_("考勤终稿由考勤假期模块管理，不得在薪酬模块作废。"))
	reason = str(reason or "").strip()
	if not reason:
		frappe.throw(_("请填写作废原因。"))
	if replacement_batch:
		if replacement_batch == batch.name or not frappe.db.exists(VARIABLE_BATCH_DOCTYPE, replacement_batch):
			frappe.throw(_("替代批次无效。"))
		replacement = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, replacement_batch)
		if replacement.company != company or replacement.payroll_month != batch.payroll_month:
			frappe.throw(_("替代批次必须属于同一公司和薪资月份。"))
	invalidation = _invalidate_unconfirmed_payroll_trial(
		company, batch.payroll_month, reason=_("作废月度增减项批次 {0}").format(batch.name)
	)
	rows = frappe.get_all(VARIABLE_RECORD_DOCTYPE, filters={"import_batch": batch.name}, pluck="name", limit_page_length=100000)
	for name in rows:
		frappe.db.set_value(VARIABLE_RECORD_DOCTYPE, name, {"review_status": "已作废", "excluded": 1}, update_modified=False)
	batch.status = "已作废"
	batch.voided_by = frappe.session.user
	batch.voided_on = now_datetime()
	batch.void_reason = reason
	batch.replacement_batch = replacement_batch or ""
	batch.notes = "{0}\n{1}".format(batch.notes or "", _("作废原因：{0}").format(reason)).strip()
	batch.save(ignore_permissions=True)
	frappe.db.commit()
	return {
		"batch": batch.name,
		"status": batch.status,
		"voided_rows": len(rows),
		"invalidated_trial": invalidation,
		"message": _("批次已留痕作废；原文件与确认记录仍可追溯，旧薪资输入表和未确认试算已失效。"),
	}


@frappe.whitelist()
def set_payroll_variable_record_excluded(name: str, excluded: int = 1):
	if not name or not frappe.db.exists(VARIABLE_RECORD_DOCTYPE, name):
		frappe.throw(_("薪资变量记录不存在"))
	doc = frappe.get_doc(VARIABLE_RECORD_DOCTYPE, name)
	batch = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, doc.import_batch) if doc.import_batch else None
	if batch and batch.status == "已确认":
		frappe.throw(_("已确认入账的批次不能直接修改或剔除。"))
	invalidation = _invalidate_unconfirmed_payroll_trial(
		doc.company, doc.payroll_month, reason=_("调整月度增减项剔除状态")
	)
	doc.excluded = int(excluded or 0)
	doc.review_status = "已剔除" if doc.excluded else "待确认"
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "excluded": doc.excluded, "review_status": doc.review_status, "invalidated_trial": invalidation}


@frappe.whitelist()
def delete_payroll_variable_import_batches(batch_names, company: str = "", payroll_month: str = "", attendance_lock_version: str = ""):
	"""Atomically delete selected, unconfirmed monthly import batches.

	Confirmed batches are immutable audit evidence.  Validate every selection
	before deleting anything so one invalid or cross-company batch cannot leave a
	partially-cleared month behind.
	"""
	if isinstance(batch_names, str):
		try:
			batch_names = json.loads(batch_names)
		except (TypeError, ValueError):
			batch_names = [name.strip() for name in batch_names.split(",") if name.strip()]
	batch_names = list(dict.fromkeys(batch_names or []))
	if not batch_names:
		frappe.throw(_("请至少选择一个待删除的导入批次。"))
	if len(batch_names) > 100:
		frappe.throw(_("单次最多删除 100 个导入批次。"))

	company = _require_company(company)
	batches = []
	for batch_name in batch_names:
		if not frappe.db.exists(VARIABLE_BATCH_DOCTYPE, batch_name):
			frappe.throw(_("导入批次不存在：{0}").format(batch_name))
		batch = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, batch_name)
		if batch.company != company:
			frappe.throw(_("导入批次 {0} 与当前公司不一致，已阻断删除。").format(batch.name))
		if payroll_month and batch.payroll_month != payroll_month:
			frappe.throw(_("导入批次 {0} 与当前薪资月份不一致，已阻断删除。").format(batch.name))
		if attendance_lock_version and not str(batch.attendance_lock_version or "").startswith(MONTHLY_VARIABLE_SCOPE_PREFIX) and batch.attendance_lock_version != attendance_lock_version:
			frappe.throw(_("导入批次 {0} 的锁定版本与当前锁定版本不一致，已阻断删除。").format(batch.name))
		if batch.status in {"已确认", "已作废"}:
			frappe.throw(_("已确认或已作废的批次 {0} 是审计证据，不能直接删除。").format(batch.name))
		batches.append(batch)

	variable_names = []
	invalidated_inputs = 0
	invalidated_settlements = 0
	invalidated_scopes = set()
	for batch in batches:
		variable_names.extend(
			frappe.get_all(
				VARIABLE_RECORD_DOCTYPE,
				filters={"company": company, "payroll_month": batch.payroll_month, "import_batch": batch.name},
				pluck="name",
			)
		)
		scope_key = (batch.company, batch.payroll_month)
		if scope_key not in invalidated_scopes:
			invalidation = _invalidate_unconfirmed_payroll_trial(
				batch.company, batch.payroll_month, reason=_("删除未确认月度增减项批次")
			)
			invalidated_inputs += cint(invalidation.get("deleted_inputs"))
			invalidated_settlements += cint(invalidation.get("deleted_settlements"))
			invalidated_scopes.add(scope_key)

	for name in variable_names:
		frappe.delete_doc(VARIABLE_RECORD_DOCTYPE, name, ignore_permissions=True, force=True)
	for batch in batches:
		frappe.delete_doc(VARIABLE_BATCH_DOCTYPE, batch.name, ignore_permissions=True, force=True)
	frappe.db.commit()

	return {
		"deleted_batches": [batch.name for batch in batches],
		"payroll_months": sorted({batch.payroll_month for batch in batches}),
		"deleted_variable_records": len(variable_names),
		"deleted_payroll_input_records": invalidated_inputs,
		"settlement_records_deleted": invalidated_settlements,
		"message": _("已删除 {0} 个未确认批次及 {1} 条明细；旧薪资输入表和未确认试算已失效，请重新生成。").format(len(batches), len(variable_names)),
	}


@frappe.whitelist()
def delete_payroll_variable_import_batch(batch_name: str, company: str = "", attendance_lock_version: str = ""):
	result = delete_payroll_variable_import_batches([batch_name], company, attendance_lock_version=attendance_lock_version)
	return {
		**result,
		"deleted_batch": result["deleted_batches"][0],
		"payroll_month": result["payroll_months"][0] if result["payroll_months"] else "",
	}


def _require_test_monthly_reset_access():
	"""Restrict destructive monthly resets to an explicitly acknowledged test run."""
	if "System Manager" not in frappe.get_roles():
		frappe.throw(_("只有系统管理员可以清空测试月度数据。"), frappe.PermissionError)


def _test_monthly_reset_confirmation(area, payroll_month, department):
	scope_label = department or "全公司"
	return "TEST 清空 {0} {1} {2}".format(TEST_MONTHLY_RESET_AREAS[area], payroll_month, scope_label)


def _test_monthly_reset_scope(company, payroll_month, department, area):
	_require_test_monthly_reset_access()
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	department = _text(department)
	if area not in TEST_MONTHLY_RESET_AREAS:
		frappe.throw(_("请选择要清空的测试数据范围。"))
	if department and not frappe.db.exists("Department", department):
		frappe.throw(_("请选择存在的部门。"))
	if area == "attendance" and not department:
		frappe.throw(_("请选择存在的部门；考勤清空不支持未限定部门。"))
	return company, payroll_month, department, area


def _test_monthly_reset_names(doctype, filters):
	return frappe.get_all(doctype, filters=filters, pluck="name", limit_page_length=100000)


def _test_monthly_reset_targets(company, payroll_month, department, area):
	"""Build a test reset dependency list without touching raw files."""
	month_start = "{0}-01".format(payroll_month)
	month_end = _month_end(payroll_month)
	targets = {}
	if area == "payroll":
		targets = {
			PAYROLL_SETTLEMENT_DOCTYPE: _test_monthly_reset_names(PAYROLL_SETTLEMENT_DOCTYPE, {"company": company, "payroll_month": payroll_month}),
			PAYROLL_INPUT_DOCTYPE: _test_monthly_reset_names(PAYROLL_INPUT_DOCTYPE, {"company": company, "payroll_month": payroll_month}),
			VARIABLE_RECORD_DOCTYPE: _test_monthly_reset_names(VARIABLE_RECORD_DOCTYPE, {"company": company, "payroll_month": payroll_month}),
			WELFARE_SOURCE_DOCTYPE: _test_monthly_reset_names(WELFARE_SOURCE_DOCTYPE, {"company": company, "payroll_month": payroll_month}),
			EMPLOYEE_SALARY_CHANGE_DOCTYPE: _test_monthly_reset_names(EMPLOYEE_SALARY_CHANGE_DOCTYPE, {"company": company, "effective_date": ["between", [month_start, month_end]]}),
		}
		if _doctype_exists(MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE):
			targets[MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE] = _test_monthly_reset_names(
				MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE,
				{"company": company, "payroll_month": payroll_month},
			)
	else:
		batch_names = _test_monthly_reset_names(
			"HRMS Attendance Import Batch", {"company": company, "attendance_month": payroll_month}
		)
		batch_filter = {"import_batch": ["in", batch_names or [""]], "department": department}
		targets = {
			PAYROLL_SETTLEMENT_DOCTYPE: _test_monthly_reset_names(PAYROLL_SETTLEMENT_DOCTYPE, {"company": company, "payroll_month": payroll_month, "department": department}),
			PAYROLL_INPUT_DOCTYPE: _test_monthly_reset_names(PAYROLL_INPUT_DOCTYPE, {"company": company, "payroll_month": payroll_month, "department": department}),
			VARIABLE_RECORD_DOCTYPE: _test_monthly_reset_names(VARIABLE_RECORD_DOCTYPE, {"company": company, "payroll_month": payroll_month, "department": department, "source_sheet": "考勤终稿锁定快照"}),
			"HRMS Monthly Attendance Summary": _test_monthly_reset_names("HRMS Monthly Attendance Summary", {"company": company, "attendance_month": payroll_month, "department": department}),
			"HRMS Attendance Department Confirmation": _test_monthly_reset_names("HRMS Attendance Department Confirmation", {"company": company, "attendance_month": payroll_month, "department": department}),
			"HRMS Attendance Processing Record": _test_monthly_reset_names("HRMS Attendance Processing Record", {"company": company, "attendance_month": payroll_month, "department": department}),
			"HRMS Attendance Exception": _test_monthly_reset_names("HRMS Attendance Exception", {**batch_filter, "attendance_date": ["between", [month_start, month_end]]}),
			"HRMS Attendance Day Check": _test_monthly_reset_names("HRMS Attendance Day Check", {**batch_filter, "attendance_date": ["between", [month_start, month_end]]}),
			"HRMS Attendance Leave Evidence": _test_monthly_reset_names("HRMS Attendance Leave Evidence", batch_filter),
			"HRMS Apple Reward Record": _test_monthly_reset_names("HRMS Apple Reward Record", batch_filter),
		}
		if _doctype_exists(MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE):
			targets[MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE] = _test_monthly_reset_names(
				MONTHLY_PAYROLL_PARTICIPATION_DOCTYPE,
				{"company": company, "payroll_month": payroll_month, "department": department},
			)
	return {doctype: names for doctype, names in targets.items() if names}


def _test_monthly_reset_batch_candidates(variable_names):
	if not variable_names:
		return []
	return list(
		{
			row.import_batch
			for row in frappe.get_all(
				VARIABLE_RECORD_DOCTYPE,
				filters={"name": ["in", variable_names]},
				fields=["import_batch"],
				limit_page_length=100000,
			)
			if row.import_batch
		}
	)


def _test_monthly_reset_salary_import_batch_candidates(salary_change_names):
	"""Keep no empty salary-import batch visible after a full monthly payroll reset."""
	if not salary_change_names:
		return []
	return list(
		{
			row.salary_import_batch
			for row in frappe.get_all(
				EMPLOYEE_SALARY_CHANGE_DOCTYPE,
				filters={"name": ["in", salary_change_names]},
				fields=["salary_import_batch"],
				limit_page_length=100000,
			)
			if row.salary_import_batch
		}
	)


def _reopen_test_attendance_month(company, attendance_month, department):
	lock_name = frappe.db.get_value("HRMS Attendance Month Lock", {"company": company, "attendance_month": attendance_month}, "name")
	if not lock_name:
		return False
	lock = frappe.get_doc("HRMS Attendance Month Lock", lock_name)
	if lock.status != "已锁定":
		return False
	lock.status = "已重开"
	lock.reopened_by = frappe.session.user
	lock.reopened_on = now_datetime()
	lock.remarks = "{0}\n{1}".format(
		lock.remarks or "",
		"测试清空：{0} {1} 部门 {2}".format(TEST_MONTHLY_RESET_AREAS["attendance"], attendance_month, department),
	).strip()
	lock.save(ignore_permissions=True)
	if frappe.db.exists("DocType", "HRMS Attendance Lock Audit"):
		frappe.get_doc({
			"doctype": "HRMS Attendance Lock Audit",
			"month_lock": lock.name,
			"company": company,
			"attendance_month": attendance_month,
			"action": "解锁",
			"lock_version": lock.active_version,
			"reason": "测试清空部门数据：{0}".format(department),
			"operator": frappe.session.user,
			"occurred_on": now_datetime(),
			"source_checksum": lock.source_checksum,
		}).insert(ignore_permissions=True)
	return True


@frappe.whitelist()
def preview_test_monthly_data_reset(company: str, payroll_month: str, department: str = "", area: str = "payroll"):
	"""Show the exact test-only monthly reset impact before any record is removed."""
	company, payroll_month, department, area = _test_monthly_reset_scope(company, payroll_month, department, area)
	targets = _test_monthly_reset_targets(company, payroll_month, department, area)
	scope_label = department or "全公司"
	warnings = [
		"仅清理所选范围和月份的派生业务数据，不删除员工花名册、部门、薪资架构或原始上传附件。",
		"确认语必须完全输入：{0}".format(_test_monthly_reset_confirmation(area, payroll_month, department)),
	]
	if area == "attendance":
		warnings.append("考勤清空会同步删除本部门本月的薪资输入、试算结果和考勤继承变量，并将该公司该月考勤锁定重开；其他部门需重新锁定后再试算。")
	else:
		warnings.append("薪酬清空会删除全公司本月生效的定薪、福利来源、月度增减项、薪资输入和结算结果；早于本月生效的历史定薪、导入附件与批次追溯保留。")
	return {
		"company": company,
		"payroll_month": payroll_month,
		"department": department,
		"scope_label": scope_label,
		"area": area,
		"area_label": TEST_MONTHLY_RESET_AREAS[area],
		"confirmation": _test_monthly_reset_confirmation(area, payroll_month, department),
		"records": [{"doctype": doctype, "count": len(names), "sample_names": names[:5]} for doctype, names in targets.items()],
		"total_count": sum(len(names) for names in targets.values()),
		"warnings": warnings,
	}


@frappe.whitelist()
def reset_test_monthly_data(company: str, payroll_month: str, department: str = "", area: str = "payroll", confirmation: str = "", test_mode: int = 0):
	"""Delete one test month scope after an explicit, non-reusable acknowledgement."""
	company, payroll_month, department, area = _test_monthly_reset_scope(company, payroll_month, department, area)
	expected_confirmation = _test_monthly_reset_confirmation(area, payroll_month, department)
	if not cint(test_mode) or _text(confirmation) != expected_confirmation:
		frappe.throw(_("测试清空确认不匹配；请重新预览影响范围并输入完整确认语。"))
	targets = _test_monthly_reset_targets(company, payroll_month, department, area)
	variable_names = list(targets.get(VARIABLE_RECORD_DOCTYPE) or [])
	salary_change_names = list(targets.get(EMPLOYEE_SALARY_CHANGE_DOCTYPE) or [])
	batch_candidates = _test_monthly_reset_batch_candidates(variable_names)
	salary_import_batch_candidates = _test_monthly_reset_salary_import_batch_candidates(salary_change_names) if area == "payroll" else []
	deleted = {}
	try:
		for doctype, names in targets.items():
			deleted[doctype] = 0
			for name in names:
				if frappe.db.exists(doctype, name):
					frappe.delete_doc(doctype, name, ignore_permissions=True, force=True)
					deleted[doctype] += 1
		deleted_batches = 0
		for batch_name in batch_candidates:
			if frappe.db.exists(VARIABLE_BATCH_DOCTYPE, batch_name) and not frappe.db.exists(VARIABLE_RECORD_DOCTYPE, {"import_batch": batch_name}):
				frappe.delete_doc(VARIABLE_BATCH_DOCTYPE, batch_name, ignore_permissions=True, force=True)
				deleted_batches += 1
		for batch_name in salary_import_batch_candidates:
			if frappe.db.exists(FORM_IMPORT_BATCH_DOCTYPE, batch_name) and not frappe.db.exists(EMPLOYEE_SALARY_CHANGE_DOCTYPE, {"salary_import_batch": batch_name}):
				frappe.delete_doc(FORM_IMPORT_BATCH_DOCTYPE, batch_name, ignore_permissions=True, force=True)
				deleted_batches += 1
		attendance_reopened = _reopen_test_attendance_month(company, payroll_month, department) if area == "attendance" else False
		frappe.db.commit()
	except Exception:
		frappe.db.rollback()
		raise
	return {
		"company": company,
		"payroll_month": payroll_month,
		"department": department,
		"area": area,
		"deleted": deleted,
		"deleted_batches": deleted_batches,
		"attendance_reopened": attendance_reopened,
		"count": sum(deleted.values()) + deleted_batches,
		"message": _("已清空测试范围：{0} / {1} / {2}；原始附件、花名册和薪资架构均未删除。").format(payroll_month, department or "全公司", TEST_MONTHLY_RESET_AREAS[area]),
	}


@frappe.whitelist()
def update_payroll_variable_record(
	name: str,
	employee: str = "",
	employee_code: str = "",
	employee_name: str = "",
	department: str = "",
	variable_type: str = "",
	amount: float = 0,
	source_sheet: str = "",
	remarks: str = "",
):
	if not name or not frappe.db.exists(VARIABLE_RECORD_DOCTYPE, name):
		frappe.throw(_("薪资变量记录不存在"))

	# When HR selects an employee in the correction dialog, the source form's
	# stale code/name must not win on save.  Otherwise the Link looks resolved
	# but the next validation still reports the original unmatched identifier.
	selected_employee = employee if employee and frappe.db.exists("Employee", employee) else None
	resolved_employee = selected_employee or _payroll_employee_lookup(employee_code, employee_name)
	if not resolved_employee:
		frappe.throw(_("请先选择或填写可匹配到员工档案的员工。"))

	employee_context = _employee_context(resolved_employee)
	doc = frappe.get_doc(VARIABLE_RECORD_DOCTYPE, name)
	company = _require_company(doc.company)
	batch = frappe.get_doc(VARIABLE_BATCH_DOCTYPE, doc.import_batch) if doc.import_batch else None
	if batch and batch.status == "已确认":
		frappe.throw(_("已确认入账的批次不能直接修改。"))
	if batch and batch.source_type and variable_type and variable_type != doc.variable_type:
		source_name = frappe.db.get_value(VARIABLE_SOURCE_TYPE_DOCTYPE, batch.source_type, "source_name") or batch.source_type
		frappe.throw(_("{0} 来源的变量类型已锁定为“{1}”，只能调整金额。").format(source_name, doc.variable_type or "-"))
	if employee_context.get("company") and employee_context.get("company") != company:
		frappe.throw(_("薪资变量员工 {0} 不属于公司 {1}").format(employee_code or employee_name or resolved_employee, company))
	invalidation = _invalidate_unconfirmed_payroll_trial(
		company, doc.payroll_month, reason=_("人工修正月度增减项明细")
	)
	doc.employee = resolved_employee
	doc.employee_code = employee_context.get("employee_code") if selected_employee else (employee_code or employee_context.get("employee_code") or resolved_employee)
	doc.employee_name = employee_context.get("employee_name") if selected_employee else (employee_name or employee_context.get("employee_name"))
	doc.department = _department_lookup(department) or employee_context.get("department")
	if variable_type:
		doc.variable_type = variable_type
	doc.amount = flt(amount)
	doc.source_sheet = source_sheet or doc.source_sheet
	doc.remarks = remarks
	validation = _payroll_variable_validation(company, doc.employee_code, doc.employee_name, doc.amount)
	doc.validation_status = validation["status"]
	doc.validation_message = validation["message"]
	doc.review_status = "待确认"
	doc.excluded = 0
	doc.save(ignore_permissions=True)

	frappe.db.commit()
	result = frappe.get_value(VARIABLE_RECORD_DOCTYPE, doc.name, ["name", "employee", "employee_code", "employee_name", "department", "variable_type", "amount", "source_sheet", "remarks", "review_status", "validation_status", "validation_message", "excluded"], as_dict=True)
	result["deleted_payroll_input_records"] = cint(invalidation.get("deleted_inputs"))
	result["deleted_settlement_records"] = cint(invalidation.get("deleted_settlements"))
	result["message"] = _("薪资变量明细已保存；旧薪资输入表和未确认试算已失效，请重新生成。")
	return result


def _variable_totals(company, payroll_month, attendance_lock_version=""):
	company = _require_company(company)
	filters = {"company": company, "payroll_month": payroll_month}
	if _doctype_has_field(VARIABLE_RECORD_DOCTYPE, "review_status"):
		filters["review_status"] = "已确认"
	if _doctype_has_field(VARIABLE_RECORD_DOCTYPE, "excluded"):
		filters["excluded"] = 0
	totals = defaultdict(lambda: defaultdict(float))
	identity = {}
	sources = defaultdict(list)
	allowed_versions = {attendance_lock_version, _monthly_variable_scope(payroll_month), ""}
	selected_batch_by_source = {}
	batch_source_by_name = {}
	if _doctype_has_field(VARIABLE_BATCH_DOCTYPE, "is_selected"):
		for batch in frappe.get_all(
			VARIABLE_BATCH_DOCTYPE,
			filters={"company": company, "payroll_month": payroll_month, "status": "已确认"},
			fields=["name", "source_type", "is_selected"],
			limit_page_length=100000,
		):
			if batch.source_type:
				batch_source_by_name[batch.name] = batch.source_type
				if batch.is_selected:
					selected_batch_by_source[batch.source_type] = batch.name
	for row in frappe.get_all(VARIABLE_RECORD_DOCTYPE, filters=filters, fields=["*"]):
		if str(row.get("attendance_lock_version") or "") not in allowed_versions:
			continue
		# Housing allowance is an attendance-support fact.  Historic payroll-side
		# uploads stay auditable, but are deliberately excluded so they cannot be
		# paid alongside the amount inherited from the locked attendance final.
		if row.variable_type == "住房补贴" and row.source_sheet != "考勤终稿锁定快照":
			continue
		source_type = batch_source_by_name.get(row.get("import_batch"))
		selected_batch = selected_batch_by_source.get(source_type)
		if source_type and selected_batch and selected_batch != row.get("import_batch"):
			continue
		_assert_row_company(row, company, _("薪资变量"))
		key = _employee_identity_key(row)
		if not key:
			continue
		identity.setdefault(key, row)
		totals[key][row.variable_type] += flt(row.amount)
		sources[key].append(
			{
				"name": row.name,
				"variable_type": row.variable_type,
				"amount": flt(row.amount),
				"source_sheet": row.source_sheet,
				"source_hash": row.source_hash,
			}
		)
	return totals, identity, sources


def _workflow_month(payroll_month):
	payroll_month = (payroll_month or "").strip()
	if not re.match(r"^\d{4}-\d{2}$", payroll_month):
		frappe.throw(_("薪资月份必须为 YYYY-MM"))
	return payroll_month


def _workflow_snapshot(step_key, metrics, blockers=None, warnings=None, evidence=None):
	blockers = [str(item) for item in (blockers or []) if item]
	warnings = [str(item) for item in (warnings or []) if item]
	payload = {
		"step_key": step_key,
		"metrics": metrics or [],
		"blockers": blockers,
		"warnings": warnings,
		"evidence": evidence or [],
	}
	raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
	return {
		"key": step_key,
		"label": PAYROLL_WORKFLOW_STEP_LABELS[step_key],
		"ready": not blockers,
		"metrics": metrics or [],
		"blockers": blockers,
		"warnings": warnings,
		"validation_hash": hashlib.sha256(raw.encode("utf-8")).hexdigest(),
		"snapshot_json": raw,
	}


def _workflow_rows(doctype, filters, fields):
	if not _doctype_exists(doctype):
		return []
	available = _safe_fields(doctype, list(dict.fromkeys(["name", "modified", *fields])))
	return [dict(row) for row in frappe.get_all(doctype, filters=filters, fields=available, order_by="name asc", limit_page_length=100000)]


def _validate_master_step(company, payroll_month):
	filters = {"company": company} if _doctype_has_field("Employee", "company") else {}
	if _doctype_has_field("Employee", "status"):
		filters["status"] = "Active"
	fields = ["employee_name", "custom_employee_code", "department", "designation", "status", "date_of_joining", "final_confirmation_date", "confirmation_date"]
	rows = _workflow_rows("Employee", filters, fields)
	missing = []
	for row in rows:
		missing_fields = []
		if not _employee_code(row):
			missing_fields.append("工号")
		for fieldname, label in (("employee_name", "姓名"), ("department", "部门"), ("designation", "岗位"), ("status", "状态"), ("date_of_joining", "入职日期")):
			if _doctype_has_field("Employee", fieldname) and not row.get(fieldname):
				missing_fields.append(label)
		if missing_fields:
			missing.append(f"{row.get('employee_name') or row.get('name')}：{'/'.join(missing_fields)}")
	blockers = []
	if not rows:
		blockers.append("当前公司没有在职员工。")
	if missing:
		blockers.append("有 {0} 位员工的花名册关键字段不完整：{1}".format(len(missing), "；".join(missing[:8])))
	return _workflow_snapshot(
		"master",
		[{"label": "在职员工", "value": len(rows)}, {"label": "完整资料", "value": len(rows) - len(missing)}, {"label": "待补资料", "value": len(missing)}],
		blockers,
		[],
		rows,
	)


def _validate_salary_step(company, payroll_month, attendance_lock_version=""):
	workbench = get_salary_architecture_workbench(company, payroll_month)
	coverage = dict(workbench.get("coverage") or {})
	missing = list(workbench.get("missing_profiles") or [])
	trial = list(workbench.get("trial_profiles") or [])
	population_label = "在职员工"
	if attendance_lock_version:
		attendance_rows = _workflow_rows(
			MONTHLY_ATTENDANCE_DOCTYPE,
			_attendance_scope_filters(company, payroll_month, attendance_lock_version),
			["employee", "employee_code", "employee_name", "department"],
		)
		profiles = _active_salary_changes_for_month(company, payroll_month)
		decisions = _monthly_payroll_participation_decision_map(company, payroll_month, attendance_lock_version)
		employee_contexts = _attendance_employee_context_map(attendance_rows)
		missing, trial, pending_decisions, excluded = [], [], [], []
		population_keys = set()
		for row in attendance_rows:
			keys = [str(row.get(field) or "").strip() for field in ("employee", "employee_code", "employee_name")]
			population_keys.update(key for key in keys if key)
			decision = _participation_decision_for_row(decisions, row)
			label = row.get("employee_code") or row.get("employee_name") or row.get("employee")
			if _participation_decision_blocks_calculation(decision):
				pending_decisions.append(label)
				continue
			if _participation_decision_excludes(decision):
				excluded.append(label)
				continue
			if not decision and _employee_left_in_payroll_month(employee_contexts.get(row.get("employee")), payroll_month):
				pending_decisions.append(label)
				continue
			profile = next((profiles.get(key) for key in keys if key and profiles.get(key)), None)
			if profile and _is_salary_excluded(profile):
				excluded.append(label)
			elif not profile:
				missing.append(row)
			elif _is_trial_salary_change(profile):
				trial.append(profile)
		coverage.update({
			"active_employee_count": len(attendance_rows) - len(excluded),
			"approved_profile_count": len(attendance_rows) - len(excluded) - len(missing),
			"coverage_percent": round(100 * (len(attendance_rows) - len(excluded) - len(missing)) / (len(attendance_rows) - len(excluded)), 1) if len(attendance_rows) - len(excluded) else 100,
		})
		population_label = "本月考勤终稿人员"
	blockers = []
	if not coverage.get("active_employee_count"):
		blockers.append("没有可用的在职员工。")
	if missing:
		missing_labels = [
			str(row.get("employee_code") or row.get("employee") or row.get("employee_name") or row)
			for row in missing[:10]
		]
		blockers.append("有 {0} 位员工缺少当月有效定薪：{1}{2}".format(
			len(missing), "、".join(missing_labels), "等" if len(missing) > 10 else ""
		))
	if attendance_lock_version and pending_decisions:
		blockers.append("有 {0} 位离职或异常员工尚未完成审核决定：{1}{2}".format(
			len(pending_decisions), "、".join(str(item) for item in pending_decisions[:10]), "等" if len(pending_decisions) > 10 else ""
		))
	if trial and company != LOCAL_PAYROLL_TEST_COMPANY:
		blockers.append("有 {0} 位员工仍使用试运营/测试定薪。".format(len(trial)))
	# Only records that are actually effective for this month belong
	# to the lock snapshot. A future salary decision must not invalidate a closed
	# historical month.
	approved_changes = _latest_salary_change_map(payroll_month, company)
	evidence_by_name = {}
	for row in approved_changes.values():
		if row.get("name"):
			evidence_by_name[row.get("name")] = dict(row)
	evidence = list(evidence_by_name.values())
	return _workflow_snapshot(
		"salary",
		[
			{"label": population_label, "value": coverage.get("active_employee_count") or 0},
			{"label": "已提交定薪", "value": coverage.get("approved_profile_count") or 0},
			{"label": "定薪覆盖率", "value": f"{coverage.get('coverage_percent') or 0}%"},
			{"label": "测试值", "value": len(trial)},
		],
		blockers,
		[],
		evidence,
	)


def _validate_rules_step(company, payroll_month, attendance_lock_version=""):
	blockers = []
	warnings = []
	# Historical or superseded settlement versions remain auditable, but they
	# must not block a new calculation scope.  Execution-trace validation is
	# therefore limited to the selected immutable attendance version.
	audit = get_payroll_calculation_audit(company, payroll_month, attendance_lock_version)
	formulas = audit.get("formulas") or []
	mapping_count = (audit.get("summary") or {}).get("mapping_valid", 0)
	blockers.extend(audit.get("blockers") or [])
	warnings.extend(audit.get("warnings") or [])
	component_count = _safe_count("Salary Component", {"disabled": 0} if _doctype_has_field("Salary Component", "disabled") else {})
	if not formulas:
		blockers.append("没有可执行的薪资计算公式。")
	if not mapping_count:
		blockers.append("薪资结算字段尚未建立映射。")
	if not component_count:
		warnings.append("未检测到启用的标准工资项，请确认是否仅使用本项目公式引擎。")
	evidence = []
	for doctype, filters, fields in (
		(PAYROLL_RULE_DOCTYPE, {"company": company, "status": "已启用", "rule_code": ["like", "FORMULA_%"]}, ["rule_code", "formula_expression", "parameters_json", "effective_from", "effective_to"]),
		(PAYROLL_FIELD_MAPPING_DOCTYPE, {"status": "已启用"}, ["mapping_code", "system_field", "formula_expression", "rule_code"]),
	):
		evidence.extend(_workflow_rows(doctype, filters, fields))
	return _workflow_snapshot(
		"rules",
		[{"label": "核算公式", "value": len(formulas)}, {"label": "字段映射", "value": mapping_count}, {"label": "标准工资项", "value": component_count}],
		blockers,
		warnings,
		evidence,
	)


def _attendance_rule_cards(company, payroll_month):
	labels = {
		"ATTENDANCE_FULL_ATTENDANCE_BONUS": ("迟到与全勤奖", "工作日迟到次数扣减全勤奖；实际出勤工时已由考勤事实参与全勤梯度。", "影响全勤奖及迟到扣款"),
		"PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION": ("缺勤与旷工扣款", "早退的实际时长和工作日旷工都会成为旷工工时；按工资小计和当前倍率扣款。", "影响缺勤扣款、旷工扣款和出勤工资"),
		"ATTENDANCE_MISSED_PUNCH": ("忘打卡红苹果", "仅由“忘打卡”来源产生红苹果，不与原始考勤的缺卡标记重复扣罚。", "影响红苹果金额；仅重新加工后的来源使用新值"),
		"PAYROLL_SETTLEMENT_OVERTIME_PAY": ("加班工资倍率", "按平日、周末、法定节假日工时分别计算。", "影响加班费小计"),
		"PAYROLL_SETTLEMENT_NIGHT_SHIFT": ("夜班津贴", "深夜班按默认时段匹配；大夜班、小夜班填写完整时段后也按时段匹配，未填写时沿用终稿次数。", "影响夜班津贴和应发工资"),
	}
	cards = []
	for rule_code in PAYROLL_ATTENDANCE_RULE_CODES:
		config = _effective_rule_config(rule_code, payroll_month, company, allow_incomplete_night_times=True)
		title, description, effect = labels[rule_code]
		cards.append({
			"rule_code": rule_code,
			"title": title,
			"description": description,
			"effect": effect,
			"parameters": config.get("parameters") or {},
			"source": config.get("source"),
			"rule_name": config.get("rule_name"),
		})
	return cards


def _validate_attendance_rule_step(company, payroll_month):
	blockers = []
	try:
		cards = _attendance_rule_cards(company, payroll_month)
	except Exception as exc:
		cards = []
		blockers.append("考勤计薪规则无法执行：{0}".format(exc))
	if len(cards) != len(PAYROLL_ATTENDANCE_RULE_CODES):
		blockers.append("考勤计薪规则不完整。")
	return _workflow_snapshot(
		"attendance",
		[{"label": "缺勤规则", "value": 1 if cards else 0}, {"label": "加班规则", "value": 1 if cards else 0}, {"label": "夜班规则", "value": 1 if cards else 0}, {"label": "全勤奖规则", "value": 1 if cards else 0}],
		blockers,
		[],
		cards,
	)


def _validate_sources_step(company, payroll_month, attendance_lock_version):
	blockers = []
	warnings = []
	if not attendance_lock_version:
		return _workflow_snapshot("sources", [{"label": "考勤终稿", "value": 0}], ["请先选择当前月份的已锁定考勤终稿版本。"])
	attendance = _workflow_rows(MONTHLY_ATTENDANCE_DOCTYPE, _attendance_scope_filters(company, payroll_month, attendance_lock_version), ["employee", "employee_code", "employee_name", "department", "source_hash", "lock_status", "locked_on"])
	if not attendance:
		blockers.append("当前考勤版本不存在已锁定月度终稿。")
	attendance_keys = [_employee_identity_key(frappe._dict(row)) for row in attendance]
	if len([key for key in attendance_keys if key]) != len(set(key for key in attendance_keys if key)):
		blockers.append("考勤终稿存在重复员工。")
	if any(not key for key in attendance_keys):
		blockers.append("考勤终稿存在无法识别员工的记录。")
	pending_welfare = _safe_count(WELFARE_SOURCE_DOCTYPE, {"company": company, "payroll_month": payroll_month, "attendance_lock_version": attendance_lock_version, "confirmation_status": ["in", ["草稿", "待确认"]]})
	if pending_welfare:
		blockers.append("仍有 {0} 条福利/扣款来源待确认。".format(pending_welfare))
	variable_filters = {"company": company, "payroll_month": payroll_month}
	if _doctype_has_field(VARIABLE_RECORD_DOCTYPE, "review_status"):
		variable_filters["review_status"] = "已确认"
	if _doctype_has_field(VARIABLE_RECORD_DOCTYPE, "excluded"):
		variable_filters["excluded"] = 0
	variables = _workflow_rows(VARIABLE_RECORD_DOCTYPE, variable_filters, ["employee", "employee_code", "employee_name", "variable_type", "amount", "source_sheet", "source_hash", "review_status", "attendance_lock_version"])
	allowed_versions = {attendance_lock_version, _monthly_variable_scope(payroll_month), ""}
	variables = [row for row in variables if str(row.get("attendance_lock_version") or "") in allowed_versions]
	user_variables = [row for row in variables if str(row.get("source_sheet") or "") != "考勤终稿锁定快照"]
	pending_batch_filters = {"company": company, "payroll_month": payroll_month, "status": ["in", ["待解析", *sorted(PENDING_VARIABLE_BATCH_STATUSES)]]}
	if _doctype_has_field(VARIABLE_BATCH_DOCTYPE, "source_type"):
		pending_batch_filters["source_type"] = ["!=", "salary_change"]
	pending_batch_rows = _workflow_rows(VARIABLE_BATCH_DOCTYPE, pending_batch_filters, ["attendance_lock_version", "source_file", "is_selected"])
	pending_batches = len([
		row for row in pending_batch_rows
		if str(row.get("attendance_lock_version") or "") in allowed_versions
		and not str(row.get("source_file") or "").startswith("attendance-processing-final:")
		and cint(row.get("is_selected"))
	])
	if pending_batches:
		blockers.append("仍有 {0} 个月度增减项批次待确认。".format(pending_batches))
	confirmed_batch_rows = _workflow_rows(
		VARIABLE_BATCH_DOCTYPE,
		{"company": company, "payroll_month": payroll_month, "status": "已确认"},
		["name", "attendance_lock_version", "source_type", "source_file", "variable_rows", "confirmed_by", "confirmed_on", "notes", "is_selected"],
	)
	confirmed_batch_rows = [
		row for row in confirmed_batch_rows
		if str(row.get("attendance_lock_version") or "") in allowed_versions
		and not str(row.get("source_file") or "").startswith("attendance-processing-final:")
		and (not row.get("source_type") or cint(row.get("is_selected")))
	]
	confirmed_source_types = {str(row.get("source_type") or "") for row in confirmed_batch_rows}
	required_sources = [
		row for row in list_payroll_variable_source_types()
		if cint(row.get("required_for_payroll")) and row.get("target_area") == "月度增减项"
	]
	missing_required_sources = [
		row for row in required_sources
		if str(row.get("source_code") or row.get("name") or "") not in confirmed_source_types
	]
	if missing_required_sources:
		blockers.append("以下每月必须明确状态的来源尚未确认入账：{0}。有数据时请确认批次；本月确实无数据时请上传对应空表并“确认无数据”。".format(
			"、".join(str(row.get("source_name") or row.get("source_code") or row.get("name")) for row in missing_required_sources)
		))
	attendance_key_set = set(key for key in attendance_keys if key)
	variable_by_key = {}
	for row in user_variables:
		key = _employee_identity_key(frappe._dict(row))
		if key and key not in variable_by_key:
			variable_by_key[key] = row
	extra_keys = sorted(set(variable_by_key) - attendance_key_set)
	if extra_keys:
		extra_labels = []
		for key in extra_keys[:10]:
			row = variable_by_key[key]
			code = str(row.get("employee_code") or row.get("employee") or "").strip()
			name = str(row.get("employee_name") or "").strip()
			extra_labels.append(" ".join(value for value in (code, name) if value) or key)
		warnings.append("月度变量中有 {0} 位员工不在当前考勤终稿：{1}{2}；原始批次保留，但本月不计入也不会创建额外薪资人员。".format(
			len(extra_keys), "、".join(extra_labels), "等" if len(extra_keys) > 10 else ""
		))
	confirmed_welfare = _safe_count(WELFARE_SOURCE_DOCTYPE, {"company": company, "payroll_month": payroll_month, "attendance_lock_version": attendance_lock_version, "confirmation_status": "已确认", "eligibility_status": "符合"})
	if not user_variables and not confirmed_welfare:
		warnings.append("当前没有月度奖金、补贴或扣款；如果本月确实为零，可继续锁定。")
	evidence = attendance + user_variables + confirmed_batch_rows + _workflow_rows(WELFARE_SOURCE_DOCTYPE, {"company": company, "payroll_month": payroll_month, "attendance_lock_version": attendance_lock_version}, ["employee", "employee_code", "source_type", "amount", "eligibility_status", "confirmation_status"])
	return _workflow_snapshot(
		"sources",
		[{"label": "考勤终稿", "value": len(attendance)}, {"label": "已确认月度增减项", "value": len(user_variables)}, {"label": "已确认来源", "value": len(confirmed_source_types)}, {"label": "福利/扣款", "value": confirmed_welfare}, {"label": "待确认", "value": pending_welfare + pending_batches}],
		blockers,
		warnings,
		evidence,
	)


def _validate_calculation_step(company, payroll_month, attendance_lock_version):
	if not attendance_lock_version:
		return _workflow_snapshot("calculation", [{"label": "试算结果", "value": 0}], ["请先选择考勤锁定版本。"])
	scope = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
	attendance_evidence = _workflow_rows(
		MONTHLY_ATTENDANCE_DOCTYPE,
		_attendance_scope_filters(company, payroll_month, attendance_lock_version),
		["employee", "employee_code", "employee_name"],
	)
	input_evidence = _workflow_rows(PAYROLL_INPUT_DOCTYPE, scope, ["employee", "employee_code", "employee_name", "source_hash", "source_trace_json", "settlement_status"])
	settlement_evidence = _workflow_rows(PAYROLL_SETTLEMENT_DOCTYPE, scope, ["employee", "employee_code", "employee_name", "source_hash", "source_trace_json", "gross_pay", "net_pay", "company_cost_total", "calculation_status"])
	attendance_count = len(attendance_evidence)
	input_count = len(input_evidence)
	settlement_count = len(settlement_evidence)
	attendance_keys = {_employee_identity_key(frappe._dict(row)) for row in attendance_evidence}
	attendance_keys.discard(None)
	blockers = []
	if not input_count:
		blockers.append("请先生成薪资输入表。")
	if input_count != attendance_count:
		blockers.append("薪资输入表人数与考勤终稿人数不一致。")
	if not settlement_count:
		blockers.append("请先完成本月工资试算。")
	if settlement_count != input_count:
		blockers.append("薪资结算表人数与输入表人数不一致。")
	for label, rows in (("薪资输入表", input_evidence), ("薪资结算表", settlement_evidence)):
		outside_labels = _employee_population_labels([frappe._dict(row) for row in rows], attendance_keys)
		if outside_labels:
			blockers.append("{0}包含不在锁定考勤终稿中的员工：{1}。".format(label, "、".join(outside_labels[:10])))
	current_snapshot = _payroll_run_snapshot(company, payroll_month, attendance_lock_version)
	if input_evidence and any(_trace_snapshot_hash(frappe._dict(row)) != current_snapshot for row in input_evidence):
		blockers.append("上游考勤、定薪、月度增减项或规则已变化，请重新生成薪资输入表。")
	if settlement_evidence and any(_trace_snapshot_hash(frappe._dict(row)) != current_snapshot for row in settlement_evidence):
		blockers.append("当前试算结果已失效，请重新试算后再确认。")
	evidence = input_evidence + settlement_evidence
	return _workflow_snapshot("calculation", [{"label": "考勤人数", "value": attendance_count}, {"label": "输入表", "value": input_count}, {"label": "试算结果", "value": settlement_count}], blockers, [], evidence)


def _validate_delivery_step(company, payroll_month, attendance_lock_version):
	if not attendance_lock_version:
		return _workflow_snapshot("delivery", [{"label": "已确认结算", "value": 0}], ["请先选择考勤锁定版本。"])
	scope = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
	settlements = _workflow_rows(PAYROLL_SETTLEMENT_DOCTYPE, scope, ["employee", "employee_code", "net_pay", "calculation_status", "payment_status", "confirmation_status"])
	confirmed = [row for row in settlements if row.get("calculation_status") in ("已确认", "已生成工资单")]
	blockers = []
	if not settlements:
		blockers.append("当前没有薪资结算结果。")
	elif len(confirmed) != len(settlements):
		blockers.append("仍有 {0} 条结算结果未完成复核确认。".format(len(settlements) - len(confirmed)))
	return _workflow_snapshot("delivery", [{"label": "结算人数", "value": len(settlements)}, {"label": "已确认结算", "value": len(confirmed)}], blockers, [], settlements)


def _validate_payroll_workflow_step(company, payroll_month, step_key, attendance_lock_version=""):
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	if step_key not in PAYROLL_WORKFLOW_STEP_LABELS:
		frappe.throw(_("未知的薪酬流程步骤：{0}").format(step_key))
	validators = {
		"master": lambda: _validate_master_step(company, payroll_month),
		"salary": lambda: _validate_salary_step(company, payroll_month, attendance_lock_version),
		"rules": lambda: _validate_rules_step(company, payroll_month, attendance_lock_version),
		"attendance": lambda: _validate_attendance_rule_step(company, payroll_month),
		"sources": lambda: _validate_sources_step(company, payroll_month, attendance_lock_version),
		"calculation": lambda: _validate_calculation_step(company, payroll_month, attendance_lock_version),
		"delivery": lambda: _validate_delivery_step(company, payroll_month, attendance_lock_version),
	}
	return validators[step_key]()


def _step_lock_doc(company, payroll_month, step_key):
	if not _doctype_exists(PAYROLL_STEP_LOCK_DOCTYPE):
		return None
	name = frappe.db.get_value(PAYROLL_STEP_LOCK_DOCTYPE, {"company": company, "payroll_month": payroll_month, "step_key": step_key}, "name")
	return frappe.get_doc(PAYROLL_STEP_LOCK_DOCTYPE, name) if name else None


def _save_step_lock(company, payroll_month, attendance_lock_version, validation):
	doc = _step_lock_doc(company, payroll_month, validation["key"])
	if not doc:
		doc = frappe.get_doc({"doctype": PAYROLL_STEP_LOCK_DOCTYPE, "company": company, "payroll_month": payroll_month, "step_key": validation["key"]})
	doc.step_label = validation["label"]
	doc.lock_status = "已锁定"
	doc.attendance_lock_version = attendance_lock_version if validation["key"] in ("sources", "calculation", "delivery") else ""
	doc.validation_hash = validation["validation_hash"]
	doc.validation_summary = "系统校验通过；人工锁定当前快照。"
	doc.blocker_count = len(validation["blockers"])
	doc.warning_count = len(validation["warnings"])
	doc.locked_by = frappe.session.user
	doc.locked_on = now_datetime()
	doc.invalidated_by = None
	doc.invalidated_on = None
	doc.invalidation_reason = None
	doc.snapshot_json = validation["snapshot_json"]
	if doc.is_new():
		doc.insert(ignore_permissions=True)
	else:
		doc.save(ignore_permissions=True)
	return doc


def _workflow_status(company, payroll_month, attendance_lock_version=""):
	steps = []
	prior_locked = True
	for step_key, label in PAYROLL_WORKFLOW_STEPS:
		validation = _validate_payroll_workflow_step(company, payroll_month, step_key, attendance_lock_version)
		doc = _step_lock_doc(company, payroll_month, step_key)
		stored_locked = bool(doc and doc.lock_status == "已锁定")
		version_matches = not doc or step_key not in ("sources", "calculation", "delivery") or str(doc.attendance_lock_version or "") == str(attendance_lock_version or "")
		stale = bool(stored_locked and (doc.validation_hash != validation["validation_hash"] or not version_matches))
		locked = bool(stored_locked and not stale and prior_locked)
		step = {
			**{key: value for key, value in validation.items() if key != "snapshot_json"},
			"label": label,
			"locked": locked,
			"stale": stale,
			"prerequisites_locked": prior_locked,
			"lock_status": ("已锁定" if locked else ("已失效" if stale or (doc and doc.lock_status == "已失效") else "待锁定")),
			"locked_by": doc.locked_by if doc else "",
			"locked_on": doc.locked_on if doc else "",
			"attendance_lock_version": doc.attendance_lock_version if doc else "",
		}
		steps.append(step)
		prior_locked = locked
	return steps


@frappe.whitelist()
def get_payroll_workflow_status(company: str, payroll_month: str, attendance_lock_version: str = ""):
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	return {"company": company, "payroll_month": payroll_month, "attendance_lock_version": attendance_lock_version or "", "steps": _workflow_status(company, payroll_month, attendance_lock_version)}


@frappe.whitelist()
def save_attendance_pay_rule(company: str, payroll_month: str, rule_code: str, settings=None):
	"""Save an attendance-pay rule from the plain-language inline editor.

	The API deliberately receives only business fields.  Rule codes, formulas,
	JSON and source-cell references remain an internal implementation detail.
	"""
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有维护考勤计薪规则的权限"))
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	if rule_code not in PAYROLL_ATTENDANCE_RULE_CODES:
		frappe.throw(_("该项目不属于可编辑的考勤计薪规则"))
	if isinstance(settings, str):
		try:
			settings = json.loads(settings or "{}")
		except (TypeError, ValueError) as exc:
			frappe.throw(_("填写内容无法识别：{0}").format(exc))
	settings = settings or {}
	if not isinstance(settings, dict):
		frappe.throw(_("填写内容格式不正确"))

	default = _default_rule(rule_code)
	parameters = dict((_effective_rule_config(rule_code, payroll_month, company, allow_incomplete_night_times=True).get("parameters") or {}))
	allowed = {
		"ATTENDANCE_FULL_ATTENDANCE_BONUS": {"thresholds", "late_deduction"},
		"PAYROLL_SETTLEMENT_ABSENCE_DEDUCTION": {"standard_hours_divisor", "absenteeism_multiplier"},
		"ATTENDANCE_MISSED_PUNCH": {"red_apples_per_record", "amount_per_apple"},
		"PAYROLL_SETTLEMENT_OVERTIME_PAY": {"standard_hours_divisor", "weekday", "weekend", "holiday"},
		"PAYROLL_SETTLEMENT_NIGHT_SHIFT": {"deep_night_shift", "large_night_shift", "small_night_shift", "deep_night_shift_start", "deep_night_shift_end", "large_night_shift_start", "large_night_shift_end", "small_night_shift_start", "small_night_shift_end"},
	}[rule_code]
	for key in allowed:
		if key in settings:
			parameters[key] = settings[key]
	errors = _rule_parameter_errors(rule_code, parameters)
	if errors:
		frappe.throw(_("请检查填写内容：{0}").format("；".join(errors)))
	save_payroll_rule_version(
		company=company,
		payroll_month=payroll_month,
		rule_code=rule_code,
		parameters_json=json.dumps(parameters, ensure_ascii=False),
	)
	return get_payroll_attendance_rule_overview(company, payroll_month)


@frappe.whitelist()
def get_payroll_attendance_rule_overview(company: str, payroll_month: str):
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	validation = _validate_attendance_rule_step(company, payroll_month)
	return {"company": company, "payroll_month": payroll_month, "valid": validation["ready"], "blockers": validation["blockers"], "rules": _attendance_rule_cards(company, payroll_month) if validation["ready"] else []}


def get_attendance_processing_rule_settings(company: str, attendance_month: str):
	"""Return approved company/month values used while processing sources.

	Attendance import owns source facts.  This module owns the rule values, so a
	missing marker in the raw DingTalk attendance sheet can never create a second
	red-apple penalty beside the dedicated missed-punch source.
	"""
	company = _require_company(company)
	attendance_month = _workflow_month(attendance_month)
	rule = _effective_rule_config("ATTENDANCE_MISSED_PUNCH", attendance_month, company)
	return dict(rule.get("parameters") or {})


@frappe.whitelist()
def lock_payroll_workflow_step(company: str, payroll_month: str, step_key: str, attendance_lock_version: str = ""):
	if not _can_manage_payroll_rules():
		frappe.throw(_("仅系统管理员或人事管理员可以锁定薪酬流程。"))
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	steps = _workflow_status(company, payroll_month, attendance_lock_version)
	target = next((step for step in steps if step["key"] == step_key), None)
	if not target:
		frappe.throw(_("未知的薪酬流程步骤。"))
	if not target["prerequisites_locked"]:
		frappe.throw(_("请先锁定上一步，系统不允许跳过前置确认。"))
	if not target["ready"]:
		frappe.throw(_("当前步骤校验未通过：{0}").format("；".join(target["blockers"])))
	if step_key == "calculation":
		scope = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
		for name in frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters=scope, pluck="name"):
			doc = frappe.get_doc(PAYROLL_SETTLEMENT_DOCTYPE, name)
			if doc.calculation_status not in ("已确认", "已生成工资单"):
				doc.calculation_status = "已确认"
				doc.save(ignore_permissions=True)
		validation = _validate_payroll_workflow_step(company, payroll_month, step_key, attendance_lock_version)
	else:
		validation = _validate_payroll_workflow_step(company, payroll_month, step_key, attendance_lock_version)
	_save_step_lock(company, payroll_month, attendance_lock_version, validation)
	frappe.db.commit()
	return get_payroll_workflow_status(company, payroll_month, attendance_lock_version)


@frappe.whitelist()
def unlock_payroll_workflow_step(company: str, payroll_month: str, step_key: str, reason: str, attendance_lock_version: str = ""):
	if not _can_manage_payroll_rules():
		frappe.throw(_("仅系统管理员或人事管理员可以解锁薪酬流程。"))
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	reason = (reason or "").strip()
	if not reason:
		frappe.throw(_("解锁必须填写原因。"))
	keys = [key for key, _label in PAYROLL_WORKFLOW_STEPS]
	if step_key not in keys:
		frappe.throw(_("未知的薪酬流程步骤。"))
	start = keys.index(step_key)
	if start <= keys.index("calculation") and attendance_lock_version:
		scope = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
		generated_slips = _safe_count(PAYROLL_SETTLEMENT_DOCTYPE, {**scope, "calculation_status": "已生成工资单"})
		if generated_slips:
			frappe.throw(_("当前结算已生成工资单，不能直接解锁上游流程。请先走工资单撤销流程。"))
		for name in frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters={**scope, "calculation_status": "已确认"}, pluck="name"):
			doc = frappe.get_doc(PAYROLL_SETTLEMENT_DOCTYPE, name)
			doc.calculation_status = "已生成"
			doc.save(ignore_permissions=True)
	for key in keys[start:]:
		doc = _step_lock_doc(company, payroll_month, key)
		if not doc or doc.lock_status != "已锁定":
			continue
		doc.lock_status = "已失效"
		doc.invalidated_by = frappe.session.user
		doc.invalidated_on = now_datetime()
		doc.invalidation_reason = reason if key == step_key else "上游步骤解锁：{0}".format(reason)
		doc.save(ignore_permissions=True)
	frappe.db.commit()
	return get_payroll_workflow_status(company, payroll_month, attendance_lock_version)


def _assert_workflow_locked_for_generation(company, payroll_month, attendance_lock_version):
	"""Validate the current payroll data when calculation is requested.

	Payroll areas are not a manual seven-step workflow.  Users may maintain them
	in any order; generation is the point where the system checks the conditions
	that materially affect the result.
	"""
	checks = (
		("员工定薪", _validate_salary_step(company, payroll_month, attendance_lock_version)),
		("薪资核算规则", _validate_rules_step(company, payroll_month, attendance_lock_version)),
		("考勤计薪规则", _validate_attendance_rule_step(company, payroll_month)),
		("考勤与月度增减项", _validate_sources_step(company, payroll_month, attendance_lock_version)),
	)
	blockers = []
	for label, validation in checks:
		blockers.extend(f"{label}：{message}" for message in validation.get("blockers") or [])
	if blockers:
		frappe.throw(_("薪资试算前请处理：{0}").format("；".join(blockers)))


def _clock_time_minutes(value):
	"""Extract an HH:MM clock value from a DingTalk detail cell."""
	match = re.search(r"(?<!\d)((?:[01]\d|2[0-3]):[0-5]\d)(?!\d)", str(value or ""))
	if not match:
		return None
	clock = match.group(1)
	hour, minute = clock.split(":")
	return int(hour) * 60 + int(minute)


def _night_shift_range_segments(start, end):
	"""Split a same-day or cross-midnight range into comparable day segments."""
	return [(start, end)] if start < end else [(start, 24 * 60), (0, end)]


def _night_shift_ranges_overlap(start, end, other_start, other_end):
	return any(
		max(segment_start, other_segment_start) < min(segment_end, other_segment_end)
		for segment_start, segment_end in _night_shift_range_segments(start, end)
		for other_segment_start, other_segment_end in _night_shift_range_segments(other_start, other_end)
	)


def _attendance_detail_matches_night_shift(detail, start, end):
	"""Return whether one complete clock-in/out record fits one configured tier."""
	clock_in = _clock_time_minutes((detail or {}).get("clock_in"))
	clock_out = _clock_time_minutes((detail or {}).get("clock_out"))
	start_minutes, end_minutes = _clock_time_minutes(start), _clock_time_minutes(end)
	if None in {clock_in, clock_out, start_minutes, end_minutes} or start_minutes == end_minutes:
		return False
	if start_minutes < end_minutes:
		return start_minutes <= clock_in and clock_out <= end_minutes and clock_in <= clock_out
	# Cross-midnight shift: clock-in belongs to the evening and clock-out to the
	# following morning. Requiring both values makes incomplete punches ineligible.
	return clock_in >= start_minutes and clock_out <= end_minutes


def _locked_night_shift_matches(company, payroll_month, parameters):
	"""Classify complete locked attendance rows into the three configured tiers."""
	from hrms.api import attendance_processing_center

	batch = attendance_processing_center._latest_batch(company, payroll_month, "attendance_draft")
	if not batch:
		return {}
	records = frappe.get_all(
		attendance_processing_center.PROCESSING_RECORD_DOCTYPE,
		filters={"import_batch": batch.name},
		fields=["employee_code", "employee_name", "confirmed_value_json", "processed_value_json"],
		limit_page_length=5000,
	)
	tiers = (
		("deep_night_shift_count", "deep_night_shift_start", "deep_night_shift_end"),
		("large_night_shift_count", "large_night_shift_start", "large_night_shift_end"),
		("small_night_shift_count", "small_night_shift_start", "small_night_shift_end"),
	)
	matches = {}
	for record in records:
		try:
			values = json.loads(record.confirmed_value_json or record.processed_value_json or "{}")
		except (TypeError, ValueError):
			continue
		if not isinstance(values, dict):
			continue
		counts = {fieldname: 0 for fieldname, _start, _end in tiers}
		for detail in values.get("attendance_details") or []:
			for fieldname, start_key, end_key in tiers:
				if _attendance_detail_matches_night_shift(detail, parameters.get(start_key), parameters.get(end_key)):
					counts[fieldname] += 1
					break
		if any(counts.values()):
			for key in (str(record.employee_code or "").strip(), f"name:{str(record.employee_name or '').strip()}"):
				if key:
					matches[key] = counts
	return matches


@frappe.whitelist()
def sync_locked_attendance_final_to_payroll(company: str, payroll_month: str, attendance_lock_version: str):
	"""Copy a *locked* processing-centre final into payroll's immutable input scope.

	This is an integration adapter, not a second attendance calculation.  The
	attendance centre remains responsible for all attendance facts and its
	locked snapshot is verified again before any payroll record is written.
	"""
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	if not attendance_lock_version.startswith(PROCESSING_ATTENDANCE_LOCK_PREFIX):
		return {"synced": False, "reason": "selected_lock_is_not_processing_final"}

	from hrms.api import attendance_processing_center

	expected_snapshot = attendance_lock_version.removeprefix(PROCESSING_ATTENDANCE_LOCK_PREFIX)
	preview = attendance_processing_center.get_monthly_final_preview(company, payroll_month, "finance")
	if not preview.get("available"):
		frappe.throw(_("考勤处理中心终稿不可用：{0}").format(preview.get("reason") or "请先完成并锁定终稿"))
	if str(preview.get("locked_snapshot_version") or "") != expected_snapshot:
		frappe.throw(_("考勤终稿版本已变化，请重新选择当前锁定版本后再同步薪资。"))

	rows = preview.get("rows") or []
	if not rows:
		frappe.throw(_("锁定考勤终稿没有可参与薪资计算的员工。"))
	night_parameters = _effective_rule_config("PAYROLL_SETTLEMENT_NIGHT_SHIFT", payroll_month, company)["parameters"]
	full_attendance_rule = _effective_rule_config("ATTENDANCE_FULL_ATTENDANCE_BONUS", payroll_month, company)
	night_shift_matches = _locked_night_shift_matches(company, payroll_month, night_parameters)
	scope = _attendance_scope_filters(company, payroll_month, attendance_lock_version)
	existing = {
		_employee_identity_key(row): row.name
		for row in frappe.get_all(MONTHLY_ATTENDANCE_DOCTYPE, filters=scope, fields=["name", "employee", "employee_code", "employee_name"])
		if _employee_identity_key(row)
	}
	created_or_updated = 0
	for row in rows:
		# The preview intentionally returns the immutable source facts.  Its
		# derived fields are generated only in the downloadable workbook, so derive
		# the payroll-facing one-times settlement value from the same calculation
		# chain instead of treating the missing preview key as zero.
		attendance_calculation = attendance_processing_center._final_calculation(row)
		employee_code = _text(row.get("employee_code"))
		employee_name = _text(row.get("employee_name"))
		if not employee_code and not employee_name:
			continue
		employee = _employee_lookup(employee_code, employee_name)
		employee_context = _employee_context(employee)
		if employee_context.get("company") and employee_context.get("company") != company:
			frappe.throw(_("锁定考勤终稿存在跨公司员工：{0}").format(employee_code or employee_name))
		key = employee or employee_code or f"name:{employee_name}"
		night_counts = night_shift_matches.get(employee_code) or night_shift_matches.get(f"name:{employee_name}")
		# 每个档位独立决定是否启用时段匹配：深夜班默认启用；大夜班、小
		# 夜班尚未设置完整时段时，继续沿用终稿次数，避免空的可选配置阻断薪酬。
		large_time_matching_enabled = bool(night_parameters.get("large_night_shift_start") and night_parameters.get("large_night_shift_end"))
		small_time_matching_enabled = bool(night_parameters.get("small_night_shift_start") and night_parameters.get("small_night_shift_end"))
		deep_night_shift_count = flt((night_counts or {}).get("deep_night_shift_count"))
		large_night_shift_count = (
			flt((night_counts or {}).get("large_night_shift_count"))
			if large_time_matching_enabled else max(flt(row.get("large_night_shifts")) - deep_night_shift_count, 0)
		)
		small_night_shift_count = (
			flt((night_counts or {}).get("small_night_shift_count"))
			if small_time_matching_enabled else flt(row.get("small_night_shifts"))
		)
		values = {
			"company": company,
			"attendance_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
			"lock_status": "已锁定",
			"locked_by": frappe.session.user,
			"locked_on": now_datetime(),
			"source_batch_ids": f"attendance-processing-final:{expected_snapshot}",
			"source_checksum": expected_snapshot,
			"employee": employee,
			"employee_code": employee_code or employee,
			"employee_name": employee_name or employee_context.get("employee_name"),
			"department": _department_lookup(_text(row.get("department")), company) or employee_context.get("department"),
			"date_of_joining": employee_context.get("date_of_joining"),
			"standard_hours": flt(row.get("standard_hours")),
			"actual_attendance_hours": flt(row.get("actual_attendance_hours")),
			"adjusted_working_hours": flt(attendance_calculation["adjusted_one"]),
			"overtime_1_5_hours": flt(row.get("workday_overtime_hours")) + flt(row.get("special_workday_hours")),
			"overtime_2_hours": flt(row.get("restday_overtime_hours")) + flt(row.get("special_restday_hours")),
			"overtime_3_hours": flt(row.get("holiday_overtime_hours")) + flt(row.get("special_holiday_hours")),
			"absent_hours": flt(row.get("absence_hours")),
			"deep_night_shift_count": deep_night_shift_count,
			"large_night_shift_count": large_night_shift_count,
			"small_night_shift_count": small_night_shift_count,
			"green_apples": flt(row.get("green_apples")),
			"red_apples": flt(row.get("red_apples")),
			"apple_reward_amount": flt(row.get("green_apple_amount")) - flt(row.get("red_apple_amount")),
			"red_apple_penalty": flt(row.get("red_apple_amount")),
			"full_attendance_deduction": (
				flt(row.get("late_count"))
				* _rule_number(full_attendance_rule, "late_deduction")
			),
			"status": "已确认",
		}
		name = existing.get(key)
		if name:
			doc = frappe.get_doc(MONTHLY_ATTENDANCE_DOCTYPE, name)
			# This adapter is called while the payroll page reads its attendance
			# dependency.  Re-saving an identical locked summary changes Frappe's
			# ``modified`` timestamp and makes users with that summary open receive a
			# false "modified after you opened it" conflict.  Preserve the original
			# lock audit data and only write when the actual payroll-facing snapshot
			# differs.
			values["locked_by"] = doc.locked_by or values["locked_by"]
			values["locked_on"] = doc.locked_on or values["locked_on"]
			numeric_fields = {
				"standard_hours", "actual_attendance_hours", "adjusted_working_hours",
				"overtime_1_5_hours", "overtime_2_hours", "overtime_3_hours",
				"absent_hours", "deep_night_shift_count", "large_night_shift_count", "small_night_shift_count",
				"green_apples", "red_apples", "apple_reward_amount",
				"red_apple_penalty", "full_attendance_deduction",
			}
			changed = any(
				flt(doc.get(fieldname)) != flt(value)
				if fieldname in numeric_fields
				else str(doc.get(fieldname) or "") != str(value or "")
				for fieldname, value in values.items()
			)
			if changed:
				doc.update(values)
				doc.save(ignore_permissions=True)
		else:
			frappe.get_doc({"doctype": MONTHLY_ATTENDANCE_DOCTYPE, **values}).insert(ignore_permissions=True)
		created_or_updated += 1

	batch_source = f"attendance-processing-final:{expected_snapshot}"
	batch_name = frappe.db.get_value(
		VARIABLE_BATCH_DOCTYPE,
		{"company": company, "payroll_month": payroll_month, "attendance_lock_version": attendance_lock_version, "source_file": batch_source},
		"name",
	)
	# Page initialization can request the same dependency more than once.  A
	# locked snapshot is immutable, so when the mirrored batch already contains
	# the same attendance-derived variable results there is nothing to rewrite.
	# Besides being faster, this prevents a harmless refresh from changing the
	# batch timestamp.
	expected_attendance_variables = {
		(_text(row.get("employee_code")) or _text(row.get("employee_name")), variable_type): amount
		for row in rows
		if _text(row.get("employee_code")) or _text(row.get("employee_name"))
		for variable_type, amount in (
			("全勤奖", flt(row.get("full_attendance_award"))),
			("住房补贴", flt(row.get("housing_allowance"))),
		)
	}
	if batch_name:
		existing_attendance_rows = frappe.get_all(
			VARIABLE_RECORD_DOCTYPE,
			filters={"import_batch": batch_name, "source_sheet": "考勤终稿锁定快照", "variable_type": ["in", ["全勤奖", "住房补贴"]]},
			fields=["employee_code", "employee_name", "variable_type", "amount", "review_status", "excluded"],
			limit_page_length=100000,
		)
		existing_attendance_variables = {
			(_text(row.employee_code) or _text(row.employee_name), row.variable_type): row
			for row in existing_attendance_rows
			if _text(row.employee_code) or _text(row.employee_name)
		}
		if (
			len(existing_attendance_variables) == len(expected_attendance_variables)
			and all(
				key in existing_attendance_variables
				and flt(existing_attendance_variables[key].amount) == amount
				and existing_attendance_variables[key].review_status == "已确认"
				and not existing_attendance_variables[key].excluded
				for key, amount in expected_attendance_variables.items()
			)
		):
			return {"synced": False, "reason": "already_current", "attendance_rows": created_or_updated, "variable_rows": len(existing_attendance_rows), "attendance_lock_version": attendance_lock_version}
	if batch_name:
		frappe.db.delete(VARIABLE_RECORD_DOCTYPE, {"import_batch": batch_name})
		# Do not save a stale document object here.  This adapter can be called by
		# parallel page-load requests, and a direct update keeps the immutable
		# attendance cache from raising a misleading "has been modified" error.
		frappe.db.set_value(
			VARIABLE_BATCH_DOCTYPE,
			batch_name,
			{
				"status": "已确认",
				"imported_on": now_datetime(),
				"confirmed_by": frappe.session.user,
				"confirmed_on": now_datetime(),
			},
			update_modified=False,
		)
	else:
		batch = frappe.get_doc({
			"doctype": VARIABLE_BATCH_DOCTYPE,
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
			"source_file": batch_source,
			"status": "已确认",
			"imported_by": frappe.session.user,
			"imported_on": now_datetime(),
			"confirmed_by": frappe.session.user,
			"confirmed_on": now_datetime(),
			"notes": "由考勤处理中心已锁定终稿自动同步；不允许编辑原始考勤事实。",
		})
		batch.insert(ignore_permissions=True)
		batch_name = batch.name
	variable_rows = 0
	for row in rows:
		base_row = {"工号": _text(row.get("employee_code")), "姓名": _text(row.get("employee_name")), "部门": _text(row.get("department"))}
		# Full attendance and housing allowance are independently checked in the
		# attendance centre.  Both are inherited from its locked snapshot, never
		# uploaded again or calculated from one another in payroll.
		for variable_type, amount, allow_zero in (
			("全勤奖", flt(row.get("full_attendance_award")), True),
			("住房补贴", flt(row.get("housing_allowance")), True),
		):
			name = _insert_variable(batch_name, company, payroll_month, attendance_lock_version, "考勤终稿锁定快照", base_row, variable_type, amount, allow_zero, review_status="已确认")
			if name:
				variable_rows += 1
	frappe.db.set_value(
		VARIABLE_BATCH_DOCTYPE,
		batch_name,
		{
			"variable_rows": variable_rows,
			"imported_on": now_datetime(),
			"confirmed_by": frappe.session.user,
			"confirmed_on": now_datetime(),
		},
		update_modified=False,
	)
	frappe.db.commit()
	return {"synced": True, "attendance_rows": created_or_updated, "variable_rows": variable_rows, "attendance_lock_version": attendance_lock_version}


@frappe.whitelist()
def generate_payroll_input_records(company: str, payroll_month: str, attendance_lock_version: str):
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	sync_locked_attendance_final_to_payroll(company, payroll_month, attendance_lock_version)
	_assert_workflow_locked_for_generation(company, payroll_month, attendance_lock_version)
	calculation_rules = _payroll_calculation_rules(company, payroll_month)
	payroll_run_snapshot_hash = _payroll_run_snapshot(company, payroll_month, attendance_lock_version)
	input_filters = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
	variables, variable_identity, variable_sources = _variable_totals(company, payroll_month, attendance_lock_version)
	summary_rows = frappe.get_all(MONTHLY_ATTENDANCE_DOCTYPE, filters=_attendance_scope_filters(company, payroll_month, attendance_lock_version), fields=["*"])
	if not summary_rows:
		frappe.throw(_("未找到公司 {0}、月份 {1}、锁定版本 {2} 的已锁定月度考勤终稿。").format(company, payroll_month, attendance_lock_version))

	attendance_by_key = {}
	for row in summary_rows:
		_assert_row_company(row, company, _("月度考勤终稿"))
		key = _employee_identity_key(row)
		if not key:
			frappe.throw(_("月度考勤终稿存在无法识别员工的记录。"))
		if key in attendance_by_key:
			frappe.throw(_("月度考勤终稿存在重复员工：{0}").format(key))
		attendance_by_key[key] = row

	# The locked attendance final defines this month's payroll population.
	# Confirmed additions for people outside that population remain traceable in
	# their source batch, but must never create an extra payroll recipient.
	extra_variable_keys = sorted(set(variable_identity) - set(attendance_by_key))

	active_salary_changes = _active_salary_changes_for_month(company, payroll_month)
	participation_decisions = _monthly_payroll_participation_decision_map(company, payroll_month, attendance_lock_version)
	employee_contexts = _attendance_employee_context_map(summary_rows)
	missing_salary_profiles = []
	trial_salary_profiles = []
	pending_participation_decisions = []
	excluded_salary_keys = set()
	for key, attendance in attendance_by_key.items():
		profile = (
			active_salary_changes.get(getattr(attendance, "employee", None))
			or active_salary_changes.get(getattr(attendance, "employee_code", None))
			or active_salary_changes.get(getattr(attendance, "employee_name", None))
		)
		label = getattr(attendance, "employee_code", None) or getattr(attendance, "employee_name", None) or key
		decision = _participation_decision_for_row(participation_decisions, attendance)
		if _participation_decision_blocks_calculation(decision):
			pending_participation_decisions.append(label)
			continue
		if _participation_decision_excludes(decision):
			excluded_salary_keys.add(key)
			continue
		if not decision and _employee_left_in_payroll_month(employee_contexts.get(getattr(attendance, "employee", None)), payroll_month):
			pending_participation_decisions.append(label)
			continue
		if profile and _is_salary_excluded(profile):
			excluded_salary_keys.add(key)
			continue
		if not profile:
			missing_salary_profiles.append(label)
		elif _is_trial_salary_change(profile):
			trial_salary_profiles.append(label)
	if missing_salary_profiles:
		frappe.throw(_("无法生成薪资输入表：以下员工缺少本月有效定薪：{0}").format(", ".join(missing_salary_profiles[:10])))
	if pending_participation_decisions:
		frappe.throw(_("无法生成薪资输入表：以下离职或异常人员尚未完成审核决定：{0}").format(", ".join(pending_participation_decisions[:10])))
	if trial_salary_profiles and company != LOCAL_PAYROLL_TEST_COMPANY:
		frappe.throw(_("无法生成薪资输入表：以下员工仍使用本地试运营/测试薪资数据，请先导入并批准正式薪资异动：{0}").format(", ".join(trial_salary_profiles[:10])))

	for name in frappe.get_all(PAYROLL_INPUT_DOCTYPE, filters=input_filters, pluck="name"):
		frappe.delete_doc(PAYROLL_INPUT_DOCTYPE, name, ignore_permissions=True, force=True)

	created = []
	for key in sorted(k for k in attendance_by_key if k):
		if key in excluded_salary_keys:
			continue
		attendance = attendance_by_key.get(key)
		source = attendance
		values = defaultdict(float, variables.get(key, {}))
		social_insurance_policy = _apply_social_insurance_payroll_policy(
			values, getattr(source, "employee", None), payroll_month
		)
		calculated_full_attendance_bonus, full_attendance_absence_basis = _full_attendance_bonus(
			attendance, calculation_rules["ATTENDANCE_FULL_ATTENDANCE_BONUS"]
		)
		manual_full_attendance_bonus = "全勤奖" in values
		full_attendance_bonus = values["全勤奖"] if manual_full_attendance_bonus else calculated_full_attendance_bonus
		apple_reward_amount = flt(getattr(attendance, "apple_reward_amount", 0)) + values["苹果树"]
		earnings = (
			apple_reward_amount
			+ full_attendance_bonus
			+ values["住房补贴"]
			+ values["学历补贴"]
			+ values["其他奖金"]
		)
		deductions = values["宿舍扣款"] + values["社保个人"] + values["公积金个人"] + values["其他扣款"]
		trace, source_hash = _source_trace_hash({
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
			"payroll_run_snapshot_hash": payroll_run_snapshot_hash,
			"attendance_summary": getattr(attendance, "name", ""),
			"variable_records": variable_sources.get(key, []),
			"social_insurance_policy": social_insurance_policy,
			"calculation_rules": calculation_rules,
			"full_attendance_bonus_source": "确认变量" if manual_full_attendance_bonus else "考勤规则自动计算",
			"full_attendance_absence_basis": full_attendance_absence_basis,
		})
		doc = frappe.get_doc(
			{
				"doctype": PAYROLL_INPUT_DOCTYPE,
				"company": company,
				"payroll_month": payroll_month,
				"attendance_lock_version": attendance_lock_version,
				"employee": getattr(source, "employee", None),
				"employee_code": getattr(source, "employee_code", ""),
				"employee_name": getattr(source, "employee_name", ""),
				"department": getattr(source, "department", None),
				"date_of_joining": getattr(attendance, "date_of_joining", None),
				"standard_hours": flt(getattr(attendance, "standard_hours", 0)),
				"actual_attendance_hours": flt(getattr(attendance, "actual_attendance_hours", 0)),
				"adjusted_working_hours": flt(getattr(attendance, "adjusted_working_hours", 0)),
				"overtime_1_5_hours": flt(getattr(attendance, "overtime_1_5_hours", 0)),
				"overtime_2_hours": flt(getattr(attendance, "overtime_2_hours", 0)),
			"overtime_3_hours": flt(getattr(attendance, "overtime_3_hours", 0)),
			"leave_hours": flt(getattr(attendance, "leave_hours", 0)),
			"absent_hours": flt(getattr(attendance, "absent_hours", 0)),
			"deep_night_shift_count": flt(getattr(attendance, "deep_night_shift_count", 0)),
			"large_night_shift_count": flt(getattr(attendance, "large_night_shift_count", 0)),
				"small_night_shift_count": flt(getattr(attendance, "small_night_shift_count", 0)),
				"apple_reward_amount": apple_reward_amount,
				"attendance_full_deduction": flt(getattr(attendance, "full_attendance_deduction", 0)),
				"full_attendance_bonus": full_attendance_bonus,
				"housing_subsidy": values["住房补贴"],
				"education_subsidy": values["学历补贴"],
				"dormitory_deduction": values["宿舍扣款"],
				"social_security_personal": values["社保个人"],
				"housing_fund_personal": values["公积金个人"],
				"other_bonus": values["其他奖金"],
				"other_deduction": values["其他扣款"],
				"preliminary_earning_total": earnings,
				"preliminary_deduction_total": deductions,
				"settlement_status": "待结算",
				"source_trace_json": trace,
				"source_hash": source_hash,
			}
		)
		doc.insert(ignore_permissions=True)
		created.append(doc.name)

	frappe.db.commit()
	return {"created": len(created), "records": created}


@frappe.whitelist()
def list_payroll_variable_records(company: str, payroll_month: str = "", import_batch: str = "", attendance_lock_version: str = "", page_length: int = 50, start: int = 0):
	company = _require_company(company)
	filters = {"company": company}
	if payroll_month:
		filters["payroll_month"] = payroll_month
	if import_batch:
		filters["import_batch"] = import_batch
	if payroll_month:
		filters["attendance_lock_version"] = ["in", list({attendance_lock_version, _monthly_variable_scope(payroll_month), ""})]
	filters["source_sheet"] = ["!=", "考勤终稿锁定快照"]
	page_length = min(max(cint(page_length) or 50, 1), 5000)
	return frappe.get_all(
		VARIABLE_RECORD_DOCTYPE,
		filters=filters,
		fields=["*"],
		order_by="modified desc",
		limit_start=max(cint(start), 0),
		limit_page_length=page_length,
	)


@frappe.whitelist()
def list_payroll_input_records(company: str, payroll_month: str = "", attendance_lock_version: str = "", page_length: int = 50):
	company = _require_company(company)
	filters = {"company": company}
	if payroll_month:
		filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		filters["attendance_lock_version"] = attendance_lock_version
	return frappe.get_all(PAYROLL_INPUT_DOCTYPE, filters=filters, fields=["*"], order_by="modified desc", limit_page_length=int(page_length or 50))


def _rate(amount, divisor=PAYROLL_STANDARD_HOURS_DIVISOR):
	return flt(amount) / flt(divisor) if flt(amount) and flt(divisor) else 0


def _money(value):
	return flt(value, 2)


def _company_social_security(personal_amount, rule=None):
	if rule:
		return _company_social_security_from_rule(personal_amount, rule)
	amount = flt(personal_amount)
	if amount <= 0 or amount < 524.96:
		return 0
	if amount == 524.96:
		return 1256.82
	if 520 < amount < 531:
		return 1269
	if 531 < amount < 636:
		return 1522.8
	if amount > 636:
		return 1649.7
	return 0


@frappe.whitelist()
def generate_payroll_settlement_records(company: str, payroll_month: str, attendance_lock_version: str):
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	_assert_workflow_locked_for_generation(company, payroll_month, attendance_lock_version)
	payroll_run_snapshot_hash = _payroll_run_snapshot(company, payroll_month, attendance_lock_version)
	calculation_rules = _payroll_calculation_rules(company, payroll_month)
	payroll_formulas = _apply_attendance_rule_parameters(
		_effective_payroll_formulas(company, payroll_month), calculation_rules
	)
	settlement_filters = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
	locked = frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters={**settlement_filters, "calculation_status": ["in", ["已确认", "已生成工资单"]]}, pluck="name")
	if locked:
		frappe.throw(_("公司 {0} 月份 {1} 锁定版本 {2} 已存在锁定薪资结算，不允许覆盖。").format(company, payroll_month, attendance_lock_version))

	variables, variable_identity, variable_sources = _variable_totals(company, payroll_month, attendance_lock_version)
	input_rows = frappe.get_all(PAYROLL_INPUT_DOCTYPE, filters=_payroll_scope_filters(company, payroll_month, attendance_lock_version), fields=["*"])
	if not input_rows:
		frappe.throw(_("未找到公司 {0}、月份 {1}、锁定版本 {2} 的薪资输入表。").format(company, payroll_month, attendance_lock_version))
	stale_inputs = [row for row in input_rows if _trace_snapshot_hash(row) != payroll_run_snapshot_hash]
	if stale_inputs:
		frappe.throw(_("薪资输入表对应的上游数据已经变化，请重新生成薪资输入表后再试算。"))
	input_by_key = {}
	for row in input_rows:
		_assert_row_company(row, company, _("薪资输入表"))
		key = _employee_identity_key(row)
		if not key:
			frappe.throw(_("薪资输入表存在无法识别员工的记录。"))
		if key in input_by_key:
			frappe.throw(_("薪资输入表存在重复员工：{0}").format(key))
		input_by_key[key] = row
	attendance_rows = frappe.get_all(
		MONTHLY_ATTENDANCE_DOCTYPE,
		filters=_attendance_scope_filters(company, payroll_month, attendance_lock_version),
		fields=["employee", "employee_code", "employee_name"],
		limit_page_length=100000,
	)
	attendance_keys = {_employee_identity_key(row) for row in attendance_rows}
	attendance_keys.discard(None)
	outside_labels = _employee_population_labels(input_rows, attendance_keys)
	if outside_labels:
		frappe.throw(
			_("无法试算：薪资输入表包含不在锁定考勤终稿中的员工：{0}。请重新生成薪资输入表。").format(
				"、".join(outside_labels[:10])
			)
		)

	# Variables outside the locked attendance/input population are audit-only for
	# this month and cannot create settlement recipients.
	extra_variable_keys = sorted(set(variable_identity) - set(input_by_key))
	active_salary_changes = _active_salary_changes_for_month(company, payroll_month)
	missing_salary_profiles = []
	trial_salary_profiles = []
	for key, input_row in input_by_key.items():
		profile = (
			active_salary_changes.get(getattr(input_row, "employee", None))
			or active_salary_changes.get(getattr(input_row, "employee_code", None))
			or active_salary_changes.get(getattr(input_row, "employee_name", None))
		)
		if not profile:
			missing_salary_profiles.append(getattr(input_row, "employee_code", None) or getattr(input_row, "employee_name", None) or key)
		elif _is_trial_salary_change(profile):
			trial_salary_profiles.append(getattr(input_row, "employee_code", None) or getattr(input_row, "employee_name", None) or key)
	if missing_salary_profiles:
		frappe.throw(_("无法试算：以下员工缺少本月有效定薪：{0}").format(", ".join(missing_salary_profiles[:10])))
	if trial_salary_profiles and company != LOCAL_PAYROLL_TEST_COMPANY:
		frappe.throw(
			_("无法试算：以下员工仍使用本地试运营/测试薪资数据，请先导入并批准正式薪资异动：{0}").format(
				", ".join(trial_salary_profiles[:10])
			)
		)

	for name in frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters=settlement_filters, pluck="name"):
		frappe.delete_doc(PAYROLL_SETTLEMENT_DOCTYPE, name, ignore_permissions=True, force=True)

	created = []
	for key in sorted(k for k in input_by_key if k):
		input_row = input_by_key.get(key)
		source = input_row
		values = defaultdict(float, variables.get(key, {}))
		social_insurance_policy = _apply_social_insurance_payroll_policy(
			values, getattr(source, "employee", None), payroll_month
		)
		salary_change = get_active_salary_change_for_employee(
			employee=getattr(source, "employee", None),
			employee_code=getattr(source, "employee_code", ""),
			payroll_month=payroll_month,
			company=company,
		) or {}

		base_salary = values["底薪"] or flt(salary_change.get("base_salary"))
		function_allowance = values["职能津贴"] or values["职务津贴"] or flt(salary_change.get("function_allowance"))
		certificate_skill_allowance = values["证书及多能工津贴"] or values["证书津贴"] + values["多能工津贴"] or flt(salary_change.get("certificate_allowance")) + flt(salary_change.get("multi_skill_allowance"))
		formula_context = {
			"base_salary": base_salary,
			"function_allowance": function_allowance,
			"certificate_skill_allowance": certificate_skill_allowance,
			"standard_hours": flt(getattr(input_row, "standard_hours", 0)),
			"basic_attendance_hours": flt(getattr(input_row, "actual_attendance_hours", 0)),
			"raw_weekend_overtime_hours": flt(getattr(input_row, "overtime_2_hours", 0)),
			"weekday_overtime_hours": flt(getattr(input_row, "overtime_1_5_hours", 0)),
			"holiday_overtime_hours": flt(getattr(input_row, "overtime_3_hours", 0)),
			"deep_night_shift_count": flt(getattr(input_row, "deep_night_shift_count", 0)),
			"large_night_shift_count": flt(getattr(input_row, "large_night_shift_count", 0)),
			"small_night_shift_count": flt(getattr(input_row, "small_night_shift_count", 0)),
			"absenteeism_hours": flt(getattr(input_row, "absent_hours", 0)),
			"proposal_improvement_bonus": values["提案改善奖"],
			"apple_reward_amount": flt(getattr(input_row, "apple_reward_amount", 0)),
			"full_attendance_bonus": flt(getattr(input_row, "full_attendance_bonus", 0)),
			"housing_subsidy": flt(getattr(input_row, "housing_subsidy", 0)),
			"education_subsidy": flt(getattr(input_row, "education_subsidy", 0)),
			"other_bonus": flt(getattr(input_row, "other_bonus", 0)),
			"production_bonus": values["生产奖"],
			"late_full_attendance_deduction": values["迟到金额+全勤奖扣款"] or flt(getattr(input_row, "attendance_full_deduction", 0)),
			"other_deduction": flt(getattr(input_row, "other_deduction", 0)),
			"social_security_personal": (
				flt(getattr(input_row, "social_security_personal", 0)) if social_insurance_policy["apply"] else 0
			),
			"housing_fund_personal": flt(getattr(input_row, "housing_fund_personal", 0)),
			"paid_proposal_birthday_welfare": values["已发福利"],
			"continuing_service_bonus": values["继续服务奖"],
			"income_tax": values["所得税"],
			"year_end_bonus_tax": values["年终奖所得税"],
			"utilities_deduction": _money(values["水电费及扣款"] + flt(getattr(input_row, "dormitory_deduction", 0))),
			"manual_social_security_company": values["社保公司"],
			"manual_housing_fund_company": values["公积金公司"],
		}
		try:
			calculated, formula_trace = evaluate_formula_set(payroll_formulas, formula_context)
		except FormulaError as exc:
			frappe.throw(_("员工 {0} 薪资公式执行失败：{1}").format(getattr(source, "employee_name", key), exc))
		# Explicit local names keep the settlement document construction readable.
		salary_subtotal = calculated["salary_subtotal"]
		standard_hours = formula_context["standard_hours"]
		basic_attendance_hours = formula_context["basic_attendance_hours"]
		raw_weekend_overtime_hours = formula_context["raw_weekend_overtime_hours"]
		weekday_overtime_hours = formula_context["weekday_overtime_hours"]
		holiday_overtime_hours = formula_context["holiday_overtime_hours"]
		deep_night_shift_count = formula_context["deep_night_shift_count"]
		large_night_shift_count = formula_context["large_night_shift_count"]
		small_night_shift_count = formula_context["small_night_shift_count"]
		absenteeism_hours = formula_context["absenteeism_hours"]
		proposal_improvement_bonus = formula_context["proposal_improvement_bonus"]
		apple_reward_amount = formula_context["apple_reward_amount"]
		production_bonus = formula_context["production_bonus"]
		late_full_attendance_deduction = formula_context["late_full_attendance_deduction"]
		social_security_personal = formula_context["social_security_personal"]
		housing_fund_personal = formula_context["housing_fund_personal"]
		paid_proposal_birthday_welfare = formula_context["paid_proposal_birthday_welfare"]
		continuing_service_bonus = formula_context["continuing_service_bonus"]
		income_tax = formula_context["income_tax"]
		year_end_bonus_tax = formula_context["year_end_bonus_tax"]
		utilities_deduction = formula_context["utilities_deduction"]
		missing_hours = calculated["missing_hours"]
		adjusted_absence_hours = calculated["adjusted_absence_hours"]
		weekend_overtime_hours = calculated["weekend_overtime_hours"]
		full_salary_hourly_rate = calculated["full_salary_hourly_rate"]
		base_salary_hourly_rate = calculated["base_salary_hourly_rate"]
		absence_deduction_amount = calculated["absence_deduction_amount"]
		weekday_overtime_pay = calculated["weekday_overtime_pay"]
		weekend_overtime_pay = calculated["weekend_overtime_pay"]
		holiday_overtime_pay = calculated["holiday_overtime_pay"]
		overtime_pay_total = calculated["overtime_pay_total"]
		night_shift_allowance = calculated["night_shift_allowance"]
		subsidy_bonus_total = calculated["subsidy_bonus_total"]
		bonus_total = calculated["bonus_total"]
		absenteeism_deduction = calculated["absenteeism_deduction"]
		punishment_total = calculated["punishment_total"]
		attendance_wage = calculated["attendance_wage"]
		gross_pay = calculated["gross_pay"]
		taxable_salary = calculated["taxable_salary"]
		net_pay = calculated["net_pay"]
		social_security_company = calculated["social_security_company"]
		housing_fund_company = calculated["housing_fund_company"]
		company_cost_total = calculated["company_cost_total"]
		export_tax_adjusted_net_pay = calculated["export_tax_adjusted_net_pay"]
		trace, source_hash = _source_trace_hash({
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": attendance_lock_version,
			"payroll_run_snapshot_hash": payroll_run_snapshot_hash,
			"payroll_input_record": getattr(input_row, "name", ""),
			"variable_records": variable_sources.get(key, []),
			"salary_change": salary_change.get("name") if salary_change else "",
			"calculation_rules": calculation_rules,
			"formula_trace": formula_trace,
			"salary_subtotal_source": "公司公式：底薪、职能津贴、证书及多能工津贴",
		})

		doc = frappe.get_doc(
			{
				"doctype": PAYROLL_SETTLEMENT_DOCTYPE,
				"company": company,
				"payroll_month": payroll_month,
				"attendance_lock_version": attendance_lock_version,
				"employee": getattr(source, "employee", None),
				"employee_code": getattr(source, "employee_code", ""),
				"employee_name": getattr(source, "employee_name", ""),
				"department": getattr(source, "department", None),
				"base_salary": base_salary,
				"function_allowance": function_allowance,
				"certificate_skill_allowance": certificate_skill_allowance,
				"salary_subtotal": salary_subtotal,
				"standard_hours": standard_hours,
				"basic_attendance_hours": basic_attendance_hours,
				"missing_hours": missing_hours,
				"raw_weekend_overtime_hours": raw_weekend_overtime_hours,
				"adjusted_absence_hours": adjusted_absence_hours,
				"absence_deduction_amount": absence_deduction_amount,
				"weekday_overtime_hours": weekday_overtime_hours,
				"weekend_overtime_hours": weekend_overtime_hours,
				"holiday_overtime_hours": holiday_overtime_hours,
				"full_salary_hourly_rate": full_salary_hourly_rate,
				"base_salary_hourly_rate": base_salary_hourly_rate,
				"weekday_overtime_pay": weekday_overtime_pay,
				"weekend_overtime_pay": weekend_overtime_pay,
				"holiday_overtime_pay": holiday_overtime_pay,
				"overtime_pay_total": overtime_pay_total,
				"deep_night_shift_count": deep_night_shift_count,
				"large_night_shift_count": large_night_shift_count,
				"small_night_shift_count": small_night_shift_count,
				"night_shift_allowance": night_shift_allowance,
				"attendance_wage": attendance_wage,
				"proposal_improvement_bonus": proposal_improvement_bonus,
				"apple_reward_amount": apple_reward_amount,
				"subsidy_bonus_total": subsidy_bonus_total,
				"production_bonus": production_bonus,
				"bonus_total": bonus_total,
				"absenteeism_hours": absenteeism_hours,
				"absenteeism_deduction": absenteeism_deduction,
				"late_full_attendance_deduction": late_full_attendance_deduction,
				"punishment_total": punishment_total,
				"gross_pay": gross_pay,
				"social_security_personal": social_security_personal,
				"housing_fund_personal": housing_fund_personal,
				"paid_proposal_birthday_welfare": paid_proposal_birthday_welfare,
				"taxable_salary": taxable_salary,
				"continuing_service_bonus": continuing_service_bonus,
				"income_tax": income_tax,
				"year_end_bonus_tax": year_end_bonus_tax,
				"utilities_deduction": utilities_deduction,
				"net_pay": net_pay,
				"social_security_company": social_security_company,
				"housing_fund_company": housing_fund_company,
				"company_cost_total": company_cost_total,
				"export_tax_adjusted_net_pay": export_tax_adjusted_net_pay,
				"calculation_status": "已生成",
				"source_trace_json": trace,
				"source_hash": source_hash,
			}
		)
		doc.insert(ignore_permissions=True)
		created.append(doc.name)

	frappe.db.commit()
	return {"created": len(created), "records": created}


@frappe.whitelist()
def list_payroll_settlement_records(company: str, payroll_month: str = "", attendance_lock_version: str = "", page_length: int = 50):
	company = _require_company(company)
	filters = {"company": company}
	if payroll_month:
		filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		filters["attendance_lock_version"] = attendance_lock_version
	return frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters=filters, fields=["*"], order_by="modified desc", limit_page_length=int(page_length or 50))


def _previous_payroll_month(payroll_month):
	"""Return the immediately preceding YYYY-MM value, or an empty string."""
	match = re.match(r"^(\d{4})-(\d{2})$", str(payroll_month or ""))
	if not match:
		return ""
	year, month = (int(value) for value in match.groups())
	if month == 1:
		return f"{year - 1:04d}-12"
	return f"{year:04d}-{month - 1:02d}"


@frappe.whitelist()
def get_payroll_home_dashboard(company: str, payroll_month: str = "", attendance_lock_version: str = ""):
	"""Build the data-backed overview shown above the monthly payroll table.

	The dashboard deliberately reads only settlement records.  It therefore never
	mixes draft inputs, live employee data, or an arbitrary company-wide total into
	a financial comparison.  The prior month uses its current locked attendance
	version where available, so re-opened attendance periods do not double-count
	historical settlement runs.
	"""
	company = _require_company(company)
	payroll_month = str(payroll_month or "").strip()
	if payroll_month and not attendance_lock_version:
		current_lock = _current_payroll_attendance_lock(company, payroll_month)
		attendance_lock_version = str((current_lock or {}).get("attendance_lock_version") or "")
	settlement_fields = [
		"department", "employee", "employee_code", "employee_name", "base_salary",
		"function_allowance", "certificate_skill_allowance", "overtime_pay_total",
		"bonus_total", "gross_pay", "net_pay", "company_cost_total",
		"social_security_company", "housing_fund_company", "calculation_status",
	]

	def settlement_rows(month, lock_version=""):
		if not month:
			return []
		filters = {"company": company, "payroll_month": month}
		if lock_version:
			filters["attendance_lock_version"] = lock_version
		return frappe.get_all(
			PAYROLL_SETTLEMENT_DOCTYPE,
			filters=filters,
			fields=settlement_fields,
			limit_page_length=100000,
		)

	def summarize(rows):
		totals = defaultdict(float)
		amount_fields = (
			"base_salary", "function_allowance", "certificate_skill_allowance",
			"overtime_pay_total", "bonus_total", "gross_pay", "net_pay",
			"company_cost_total", "social_security_company", "housing_fund_company",
		)
		for row in rows:
			for field in amount_fields:
				totals[field] += flt(row.get(field))
		headcount = len(rows)
		confirmed_count = sum(1 for row in rows if row.get("calculation_status") in {"已确认", "已生成工资单"})
		return {
			"headcount": headcount,
			"confirmed_count": confirmed_count,
			"confirmation_rate": round(confirmed_count / headcount * 100, 1) if headcount else 0,
			"average_net_pay": round(totals["net_pay"] / headcount, 2) if headcount else 0,
			**{field: round(totals[field], 2) for field in amount_fields},
		}

	current_rows = settlement_rows(payroll_month, attendance_lock_version)
	previous_month = _previous_payroll_month(payroll_month)
	previous_lock = _current_payroll_attendance_lock(company, previous_month) if previous_month else None
	previous_rows = settlement_rows(previous_month, str((previous_lock or {}).get("attendance_lock_version") or ""))
	current = summarize(current_rows)
	previous = summarize(previous_rows)

	def comparison(field):
		current_value = flt(current.get(field))
		previous_value = flt(previous.get(field))
		available = bool(previous_rows)
		return {
			"comparison_available": available,
			"change_amount": round(current_value - previous_value, 2) if available else 0,
			"change_percent": round((current_value - previous_value) / previous_value * 100, 1) if available and previous_value else 0,
		}

	composition_source = [
		("固定薪资", current["base_salary"] + current["function_allowance"] + current["certificate_skill_allowance"], "#168a5b"),
		("加班工资", current["overtime_pay_total"], "#3b82f6"),
		("奖金补贴", current["bonus_total"], "#d99013"),
		("公司社保公积金", current["social_security_company"] + current["housing_fund_company"], "#8b5cf6"),
	]
	known_composition = sum(amount for _label, amount, _color in composition_source)
	composition_source.append(("出勤及其他", max(0, current["company_cost_total"] - known_composition), "#94a3b8"))
	cost_total = current["company_cost_total"]
	composition = [
		{
			"label": label,
			"amount": round(amount, 2),
			"percent": round(amount / cost_total * 100, 1) if cost_total else 0,
			"color": color,
		}
		for label, amount, color in composition_source
		if amount > 0
	]

	department_totals = defaultdict(lambda: {"company_cost_total": 0.0, "headcount": 0})
	for row in current_rows:
		department = row.get("department") or "未维护部门"
		department_totals[department]["company_cost_total"] += flt(row.get("company_cost_total"))
		department_totals[department]["headcount"] += 1
	department_rows = sorted(department_totals.items(), key=lambda item: item[1]["company_cost_total"], reverse=True)
	if len(department_rows) > 5:
		top_rows, other_rows = department_rows[:5], department_rows[5:]
		department_rows = top_rows + [("其他部门", {
			"company_cost_total": sum(row["company_cost_total"] for _name, row in other_rows),
			"headcount": sum(row["headcount"] for _name, row in other_rows),
		})]
	colors = ["#168a5b", "#3b82f6", "#d99013", "#8b5cf6", "#e05d44", "#64748b"]
	departments, cumulative_percent = [], 0.0
	for index, (department, values) in enumerate(department_rows):
		amount = values["company_cost_total"]
		percent = amount / cost_total * 100 if cost_total else 0
		departments.append({
			"department": department,
			"amount": round(amount, 2),
			"headcount": values["headcount"],
			"percent": round(percent, 1),
			"start_percent": round(cumulative_percent, 2),
			"end_percent": round(cumulative_percent + percent, 2),
			"color": colors[index % len(colors)],
		})
		cumulative_percent += percent

	if not current_rows:
		insight = "暂未生成本月薪资结算；完成试算后即可查看本月成本、部门占比和环比变化。"
	elif previous_rows:
		net_change = comparison("net_pay")["change_percent"]
		direction = "上升" if net_change > 0 else "下降" if net_change < 0 else "持平"
		insight = "本月已结算 {0} 人，实发工资较 {1} {2} {3:.1f}%。".format(current["headcount"], previous_month, direction, abs(net_change))
	else:
		insight = "本月已结算 {0} 人；尚无 {1} 的可比结算数据。".format(current["headcount"], previous_month or "上月")

	return {
		"payroll_month": payroll_month,
		"previous_month": previous_month,
		"summary": current,
		"previous_summary": previous,
		"metrics": {
			"net_pay": comparison("net_pay"),
			"average_net_pay": comparison("average_net_pay"),
			"company_cost_total": comparison("company_cost_total"),
		},
		"composition": composition,
		"departments": departments,
		"insight": insight,
	}


def _latest_salary_change_map(payroll_month="", company=""):
	company = _require_company(company)
	# Employee salary records are submitted as soon as they are saved. The legacy
	# status field only keeps historic voided records out of payroll calculations.
	filters = {"company": company, "status": ["!=", "已作废"]}
	month_end = _month_end(payroll_month)
	if month_end:
		filters["effective_date"] = ["<=", month_end]
	rows = frappe.get_all(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		filters=filters,
		fields=["*"],
		order_by="effective_date desc, modified desc",
		limit_page_length=100000,
	)
	by_key = {}
	for row in rows:
		for key in (row.employee, row.employee_code, row.employee_name):
			if key and key not in by_key:
				by_key[key] = row
	return by_key


def _is_trial_salary_change(row):
	"""Return true for the deliberately seeded/local trial salary values.

	This is a guardrail, not a financial rule.  Trial values are allowed in local
	testing, but a payroll operator must replace them with an approved real salary
	change before a formal run.
	"""
	text = " ".join(
		[
			_text(row.get("change_reason")),
			_text(row.get("remarks")),
			_text(row.get("source_file")),
		]
	).upper()
	return any(marker in text for marker in ("TEST", "试运营", "本地试运行", "本地试运营", "DEMO", "SEED"))


def _active_salary_structure_versions(payroll_month=""):
	"""Return enabled versions that cover the selected payroll month."""
	month_end = _month_end(payroll_month)
	versions = frappe.get_all(
		SALARY_STRUCTURE_VERSION_DOCTYPE,
		filters={"status": "已启用"},
		fields=["name", "structure_version", "effective_from", "effective_to", "source_file"],
		order_by="effective_from desc, modified desc",
		limit_page_length=1000,
	)
	if not month_end:
		return versions
	return [
		version
		for version in versions
		if (not version.effective_from or str(version.effective_from) <= month_end)
		and (not version.effective_to or str(version.effective_to) >= month_end)
	]


@frappe.whitelist()
def get_salary_architecture_workbench(company: str, payroll_month: str = ""):
	"""Return the operational readiness view for the salary architecture tab.

	The page uses this to show the exact gap between imported salary structures,
	employee salary decisions and the payroll calculation, instead of implying that
	a visible "salary" value is automatically eligible for payment.
	"""
	company = _require_company(company)
	month_end = _month_end(payroll_month)
	employee_filters = {"company": company} if _doctype_has_field("Employee", "company") else {}
	if _doctype_has_field("Employee", "status"):
		employee_filters["status"] = "Active"
	employee_fields = _safe_fields("Employee", ["name", "employee_name", "custom_employee_code", "department", "designation"])
	employees = _safe_get_all("Employee", filters=employee_filters, fields=employee_fields, order_by="employee_name asc", limit_page_length=100000)
	approved_changes = _latest_salary_change_map(payroll_month, company)
	missing_profiles = []
	excluded_profiles = []
	trial_profiles = []
	for employee in employees:
		code = _employee_code(employee)
		change = approved_changes.get(employee.name) or approved_changes.get(code) or approved_changes.get(employee.get("employee_name"))
		profile = {
			"employee": employee.name,
			"employee_name": employee.get("employee_name") or employee.name,
			"employee_code": code,
			"department": employee.get("department"),
			"designation": employee.get("designation"),
		}
		if change and _is_salary_excluded(change):
			profile.update({"salary_change": change.get("name"), "effective_date": change.get("effective_date"), "exclude_reason": change.get("exclude_reason")})
			excluded_profiles.append(profile)
		elif not change:
			missing_profiles.append(profile)
		elif _is_trial_salary_change(change):
			profile.update({"salary_change": change.get("name"), "effective_date": change.get("effective_date")})
			trial_profiles.append(profile)

	versions = _active_salary_structure_versions(payroll_month)
	version_names = [row.name for row in versions]
	grade_count = _safe_count(SALARY_GRADE_DOCTYPE, {"salary_structure_version": ["in", version_names]}) if version_names else 0
	enabled_rule_count = _safe_count(PAYROLL_RULE_DOCTYPE, {"company": company, "status": "已启用"})
	formula_count = _safe_count(PAYROLL_RULE_DOCTYPE, {"company": company, "rule_code": ["like", "FORMULA_%"], "status": "已启用"})
	mapping_count = _safe_count(PAYROLL_FIELD_MAPPING_DOCTYPE, {"status": "已启用"})
	standard_template_filters = {}
	if _doctype_has_field("Salary Structure", "company"):
		standard_template_filters["company"] = company
	if _doctype_has_field("Salary Structure", "is_active"):
		standard_template_filters["is_active"] = 1
	standard_template_count = _safe_count("Salary Structure", standard_template_filters)
	standard_assignment_filters = {}
	if _doctype_has_field("Salary Structure Assignment", "company"):
		standard_assignment_filters["company"] = company
	if _doctype_has_field("Salary Structure Assignment", "docstatus"):
		standard_assignment_filters["docstatus"] = 1
	month_end = _month_end(payroll_month)
	if month_end and _doctype_has_field("Salary Structure Assignment", "from_date"):
		standard_assignment_filters["from_date"] = ["<=", month_end]
	standard_assignment_count = _safe_count("Salary Structure Assignment", standard_assignment_filters)
	payroll_participant_count = len(employees) - len(excluded_profiles)
	approved_count = payroll_participant_count - len(missing_profiles)
	coverage_percent = round((approved_count / payroll_participant_count * 100), 1) if payroll_participant_count else 100

	stages = [
		{
			"key": "employee",
			"title": "员工基础资料",
			"status": "已就绪" if employees else "待维护",
			"tone": "ready" if employees else "blocked",
			"count": len(employees),
			"unit": "人",
			"detail": "员工花名册必须包含公司、工号、姓名、部门、岗位、在职状态和入职/转正信息。",
		},
		{
			"key": "structure",
			"title": "薪资架构版本",
			"status": "已就绪" if versions and grade_count else "待配置",
			"tone": "ready" if versions and grade_count else "warning",
			"count": grade_count,
			"unit": "个薪资档位",
			"detail": "已启用且适用于当前月份的薪资架构版本。" if versions else "请先导入并启用薪资架构表。",
		},
		{
			"key": "rules",
			"title": "薪资规则与字段",
			"status": "已连接" if enabled_rule_count and formula_count and mapping_count else "待配置",
			"tone": "ready" if enabled_rule_count and formula_count and mapping_count else "warning",
			"count": formula_count,
			"unit": "个公式",
			"detail": "结算字段映射、加班/扣款规则和公式必须启用；否则只能展示数据，不能形成可信结算。",
		},
		{
			"key": "profile",
			"title": "员工定薪覆盖",
			"status": "已覆盖" if not missing_profiles else "待补齐",
			"tone": "ready" if not missing_profiles else "warning",
			"count": approved_count,
			"unit": f" / {payroll_participant_count} 人",
			"detail": "仅统计参与本月计算且生效日期不晚于当前算薪月份的已提交定薪。",
		},
		{
			"key": "trial",
			"title": "试运营值检查",
			"status": "可用于正式算薪" if not trial_profiles else "需要替换",
			"tone": "ready" if not trial_profiles else "blocked",
			"count": len(trial_profiles),
			"unit": "人",
			"detail": "本地试运营或测试工资值不能作为正式发薪依据。",
		},
	]
	return {
		"company": company,
		"payroll_month": payroll_month,
		"coverage": {"active_employee_count": len(employees), "payroll_participant_count": payroll_participant_count, "excluded_profile_count": len(excluded_profiles), "approved_profile_count": approved_count, "missing_profile_count": len(missing_profiles), "coverage_percent": coverage_percent},
		"rules": {"enabled_rule_count": enabled_rule_count, "formula_count": formula_count, "mapping_count": mapping_count},
		"standard_payroll": {"template_count": standard_template_count, "assignment_count": standard_assignment_count},
		"stages": stages,
		"active_versions": versions,
		"missing_profiles": missing_profiles[:100],
		"excluded_profiles": excluded_profiles[:100],
		"trial_profiles": trial_profiles[:100],
	}


def _settlement_map(payroll_month="", company="", attendance_lock_version=""):
	company = _require_company(company)
	filters = {"company": company}
	if payroll_month:
		filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		filters["attendance_lock_version"] = attendance_lock_version
	rows = frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters=filters, fields=["*"], limit_page_length=100000)
	by_key = {}
	for row in rows:
		for key in (row.employee, row.employee_code, row.employee_name):
			if key and key not in by_key:
				by_key[key] = row
	return by_key


def _status_label(row):
	status = row.get("status") or ""
	if status == "Active":
		return "在职"
	if status == "Left":
		return "已离职"
	if status == "Inactive":
		return "待离职"
	return status or "未维护"


def _employment_stage(row, month_end=""):
	employment_type = row.get("employment_type") or ""
	# Yongxin's roster contract: imported “在职” is Full-time underneath;
	# “是否转正=否” is the authoritative trial-stage flag for both the roster
	# and payroll.  Keep the legacy Probation value compatible with this rule.
	if row.get("custom_is_confirmed") == "否" or "试用" in employment_type or "Probation" in employment_type:
		return "试用"
	confirmation_date = row.get("final_confirmation_date") or row.get("confirmation_date")
	if confirmation_date and month_end and str(confirmation_date) > month_end:
		return "试用"
	if confirmation_date and (not month_end or str(confirmation_date) <= month_end):
		return "正式"
	if row.get("status") == "Active":
		return "正式"
	return _status_label(row)


@frappe.whitelist()
def list_employee_salary_profiles(company: str, payroll_month: str = "", attendance_lock_version: str = "", page_length: int = 100):
	company = _require_company(company)
	employee_fields = _safe_fields(
		"Employee",
		[
			"name",
			"employee_name",
			"custom_employee_code",
			"department",
			"designation",
			"employment_type",
			"status",
			"date_of_joining",
			"final_confirmation_date",
			"confirmation_date",
			"relieving_date",
			"company",
		],
	)
	employee_filters = {"company": company} if _doctype_has_field("Employee", "company") else {}
	employees = _safe_get_all("Employee", filters=employee_filters, fields=employee_fields, order_by="employee_name asc", limit_page_length=100000)
	salary_changes = _latest_salary_change_map(payroll_month, company)
	settlements = _settlement_map(payroll_month, company, attendance_lock_version)
	month_end = _month_end(payroll_month)
	rows = []
	counts = {"active": 0, "regular": 0, "probation": 0, "pending_exit": 0}

	for employee in employees:
		code = _employee_code(employee)
		change = salary_changes.get(employee.name) or salary_changes.get(code) or salary_changes.get(employee.get("employee_name")) or frappe._dict()
		settlement = settlements.get(employee.name) or settlements.get(code) or settlements.get(employee.get("employee_name")) or frappe._dict()
		status_label = _status_label(employee)
		stage = _employment_stage(employee, month_end)
		if employee.get("status") == "Active" or status_label == "在职":
			counts["active"] += 1
		if stage == "正式":
			counts["regular"] += 1
		if stage == "试用":
			counts["probation"] += 1
		if employee.get("relieving_date") or status_label == "待离职":
			counts["pending_exit"] += 1

	for employee in employees[: int(page_length or 100)]:
		code = _employee_code(employee)
		change = salary_changes.get(employee.name) or salary_changes.get(code) or salary_changes.get(employee.get("employee_name")) or frappe._dict()
		settlement = settlements.get(employee.name) or settlements.get(code) or settlements.get(employee.get("employee_name")) or frappe._dict()
		status_label = _status_label(employee)
		stage = _employment_stage(employee, month_end)

		salary_risk = ""
		if status_label == "在职" and not change:
			salary_risk = "缺少已提交定薪"
		elif change and _is_trial_salary_change(change):
			salary_risk = "当前为试运营测试值，不可用于正式发薪"

		rows.append(
			{
				"employee": employee.name,
				"employee_name": employee.get("employee_name") or employee.name,
				"employee_code": code,
				"department": employee.get("department") or change.get("department") or settlement.get("department"),
				"designation": employee.get("designation") or change.get("designation"),
				"employment_type": employee.get("employment_type") or stage,
				"employee_status": status_label,
				"fixed_salary": flt(change.get("base_salary")) or flt(settlement.get("base_salary")),
				"total_salary": flt(change.get("full_salary")) or flt(settlement.get("salary_subtotal")),
				"date_of_joining": employee.get("date_of_joining") or change.get("date_of_joining"),
				"confirmation_date": employee.get("final_confirmation_date") or employee.get("confirmation_date"),
				"latest_adjustment_date": change.get("effective_date"),
				"adjustment_reason": change.get("change_reason"),
				"salary_grade": change.get("salary_grade"),
				"salary_change_status": change.get("status") or ("已批准" if change else "未维护"),
				"salary_risk": salary_risk,
				"salary_source": change.get("name") or settlement.get("name") or "",
				"settlement_status": settlement.get("calculation_status") or "未结算",
			}
		)

	return {"counts": counts, "rows": rows}


@frappe.whitelist()
def get_payroll_participation_preview(company: str, payroll_month: str, attendance_lock_version: str = ""):
	"""Show the locked-attendance population that payroll is allowed to calculate.

	The employee master is deliberately not used as the calculation population:
	only records in the selected immutable attendance final can enter payroll.
	"""
	company = _require_company(company)
	payroll_month = _workflow_month(payroll_month)
	attendance_lock_version = (attendance_lock_version or "").strip()
	if not attendance_lock_version:
		current = _current_payroll_attendance_lock(company, payroll_month)
		attendance_lock_version = str((current or {}).get("attendance_lock_version") or "")
	if not attendance_lock_version:
		return {
			"available": False,
			"company": company,
			"payroll_month": payroll_month,
			"attendance_lock_version": "",
			"columns": [],
			"rows": [],
			"reason": _("请先在考勤假期完成并锁定本月考勤终稿。"),
		}
	if attendance_lock_version.startswith(PROCESSING_ATTENDANCE_LOCK_PREFIX):
		sync_locked_attendance_final_to_payroll(company, payroll_month, attendance_lock_version)
	fields = _safe_fields(
		MONTHLY_ATTENDANCE_DOCTYPE,
		[
			"employee",
			"employee_code",
			"employee_name",
			"department",
			"standard_hours",
			"actual_attendance_hours",
			"adjusted_working_hours",
			"absent_hours",
			"attendance_lock_version",
			"locked_on",
		],
	)
	rows = frappe.get_all(
		MONTHLY_ATTENDANCE_DOCTYPE,
		filters=_attendance_scope_filters(company, payroll_month, attendance_lock_version),
		fields=fields,
		order_by="department asc, employee_code asc, employee_name asc",
		limit_page_length=100000,
	)
	employee_contexts = _attendance_employee_context_map(rows)
	decisions = _monthly_payroll_participation_decision_map(company, payroll_month, attendance_lock_version)
	salary_profiles = _active_salary_changes_for_month(company, payroll_month)
	counts = defaultdict(int)
	decision_rows = []
	for source_row in rows:
		row = dict(source_row)
		employee_context = employee_contexts.get(row.get("employee")) or {}
		profile = next(
			(
				salary_profiles.get(key)
				for key in (row.get("employee"), row.get("employee_code"), row.get("employee_name"))
				if key and salary_profiles.get(key)
			),
			None,
		)
		decision_record = _participation_decision_for_row(decisions, row)
		legacy_exclusion = bool(profile and _is_salary_excluded(profile))
		left_in_month = _employee_left_in_payroll_month(employee_context, payroll_month)
		if decision_record:
			decision = decision_record.get("decision")
			review_status = decision_record.get("review_status")
			decision_reason = decision_record.get("decision_reason")
			settlement_basis = decision_record.get("settlement_basis")
			approval_note = decision_record.get("approval_note")
		elif legacy_exclusion:
			decision = "不参与计算"
			review_status = "历史标记"
			decision_reason = profile.get("exclude_reason") or "员工定薪页标记"
			settlement_basis = ""
			approval_note = ""
		elif left_in_month:
			decision = "待处理"
			review_status = "需决策"
			decision_reason = "花名册显示离职，请选择离职结算或不参与计算。"
			settlement_basis = ""
			approval_note = ""
		else:
			decision = "正常计薪"
			review_status = "无需审核"
			decision_reason = ""
			settlement_basis = ""
			approval_note = ""

		if decision == "不参与计算" and (review_status in {PAYROLL_PARTICIPATION_APPROVED_STATUS, "历史标记"}):
			calculation_status = "不参与计算"
			counts["excluded"] += 1
		elif decision in {"异常待审核", "待处理"} or _participation_decision_blocks_calculation(decision_record):
			calculation_status = "阻塞：等待人员范围决策"
			counts["pending"] += 1
		elif decision == "离职结算":
			calculation_status = "离职结算参与计算" if profile and not legacy_exclusion else "阻塞：离职结算缺少有效定薪"
			counts["termination"] += 1
		elif not profile:
			calculation_status = "阻塞：缺少有效定薪"
			counts["missing_salary"] += 1
		else:
			calculation_status = "参与计算"
			counts["normal"] += 1
		row.update({
			"employee_status": employee_context.get("status") or "未知",
			"relieving_date": employee_context.get("relieving_date") or "",
			"decision": decision,
			"review_status": review_status,
			"decision_reason": decision_reason,
			"settlement_basis": settlement_basis,
			"approval_note": approval_note,
			"salary_status": "已有有效定薪" if profile and not legacy_exclusion else "缺少有效定薪",
			"calculation_status": calculation_status,
			"can_decide": 1,
		})
		decision_rows.append(row)
	rows = decision_rows
	columns = [
		("employee_code", "工号"),
		("employee_name", "姓名"),
		("department", "部门"),
		("employee_status", "花名册状态"),
		("relieving_date", "离职日期"),
		("decision", "本月处理"),
		("review_status", "审核状态"),
		("salary_status", "定薪"),
		("calculation_status", "计算状态"),
		("standard_hours", "标准工时"),
		("actual_attendance_hours", "实际出勤工时"),
		("adjusted_working_hours", "调整后工时"),
		("absent_hours", "缺勤工时"),
	]
	return {
		"available": bool(rows),
		"company": company,
		"payroll_month": payroll_month,
		"attendance_lock_version": attendance_lock_version,
		"locked_on": next((row.get("locked_on") for row in rows if row.get("locked_on")), None),
		"columns": [{"field": field, "label": label} for field, label in columns],
		"rows": rows,
		"counts": dict(counts),
		"reason": "" if rows else _("当前锁定考勤终稿没有可参与薪资计算的员工。"),
	}


@frappe.whitelist()
def list_monthly_payroll_overview(company: str, payroll_month: str = "", attendance_lock_version: str = ""):
	company = _require_company(company)
	# A month can retain several immutable attendance/payroll trial versions for
	# audit.  The overview is an operational dashboard, so it must aggregate only
	# the currently selected locked attendance version instead of silently adding
	# every historical version for that month.
	if payroll_month:
		company, payroll_month, attendance_lock_version = _require_payroll_scope(
			company, payroll_month, attendance_lock_version
		)
	employee_filters = {"company": company} if _doctype_has_field("Employee", "company") else {}
	if _doctype_has_field("Employee", "status"):
		employee_filters["status"] = "Active"
	employee_count = _safe_count("Employee", employee_filters)
	attendance_filters = {"company": company}
	input_filters = {"company": company}
	settlement_filters = {"company": company}
	variable_filters = {"company": company}
	batch_filters = {"company": company, "status": "已确认"}
	if payroll_month:
		attendance_filters["attendance_month"] = payroll_month
		input_filters["payroll_month"] = payroll_month
		settlement_filters["payroll_month"] = payroll_month
		variable_filters["payroll_month"] = payroll_month
		batch_filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		attendance_filters["attendance_lock_version"] = attendance_lock_version
		attendance_filters["lock_status"] = "已锁定"
		input_filters["attendance_lock_version"] = attendance_lock_version
		settlement_filters["attendance_lock_version"] = attendance_lock_version
	attendance_count = _safe_count(MONTHLY_ATTENDANCE_DOCTYPE, attendance_filters)
	input_count = _safe_count(PAYROLL_INPUT_DOCTYPE, input_filters)
	settlement_count = _safe_count(PAYROLL_SETTLEMENT_DOCTYPE, settlement_filters)
	if _doctype_has_field(VARIABLE_RECORD_DOCTYPE, "review_status"):
		variable_filters["review_status"] = "已确认"
	if _doctype_has_field(VARIABLE_RECORD_DOCTYPE, "excluded"):
		variable_filters["excluded"] = 0
	variable_rows = frappe.get_all(VARIABLE_RECORD_DOCTYPE, filters=variable_filters, fields=["attendance_lock_version", "source_sheet"], limit_page_length=100000)
	allowed_variable_versions = {attendance_lock_version, _monthly_variable_scope(payroll_month), ""} if payroll_month else None
	variable_count = len([
		row for row in variable_rows
		if (not allowed_variable_versions or str(row.attendance_lock_version or "") in allowed_variable_versions)
		and str(row.source_sheet or "") != "考勤终稿锁定快照"
	])
	confirmed_batch_rows = frappe.get_all(
		VARIABLE_BATCH_DOCTYPE,
		filters=batch_filters,
		fields=["attendance_lock_version", "source_file"],
		limit_page_length=10000,
	)
	confirmed_batch_count = len([
		row for row in confirmed_batch_rows
		if (not allowed_variable_versions or str(row.attendance_lock_version or "") in allowed_variable_versions)
		and not str(row.source_file or "").startswith("attendance-processing-final:")
	])
	settlements = frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters=settlement_filters, fields=["gross_pay", "net_pay", "company_cost_total"]) if payroll_month else []
	gross_total = sum(flt(row.gross_pay) for row in settlements)
	net_total = sum(flt(row.net_pay) for row in settlements)
	company_cost_total = sum(flt(row.company_cost_total) for row in settlements)
	coverage_base = attendance_count or employee_count
	coverage = round(settlement_count / coverage_base * 100, 2) if coverage_base else 0
	return {
		"cards": [
			{"label": "在职员工", "value": employee_count},
			{"label": "考勤终稿", "value": attendance_count},
			{"label": "薪资输入表", "value": input_count},
			{"label": "薪资结算表", "value": settlement_count},
			{"label": "变量记录", "value": variable_count},
			{"label": "已确认导入批次", "value": confirmed_batch_count},
			{"label": "考勤终稿结算覆盖率", "value": f"{coverage}%"},
			{"label": "公司实际负担总计", "value": round(company_cost_total, 2)},
		],
		"totals": {"gross_pay": round(gross_total, 2), "net_pay": round(net_total, 2), "company_cost_total": round(company_cost_total, 2)},
	}


def _active_salary_changes_for_month(company, payroll_month):
	filters = {"company": _require_company(company), "status": ["!=", "已作废"]}
	month_end = _month_end(payroll_month)
	if month_end:
		filters["effective_date"] = ["<=", month_end]
	rows = frappe.get_all(
		EMPLOYEE_SALARY_CHANGE_DOCTYPE,
		filters=filters,
		fields=[
			"name",
			"employee",
			"employee_code",
			"employee_name",
			"effective_date",
			"full_salary",
			"change_reason",
			"remarks",
			"source_file",
			"exclude_from_payroll",
			"exclude_reason",
		],
		order_by="effective_date desc, modified desc",
		limit_page_length=100000,
	)
	by_key = {}
	for row in rows:
		for key in (row.employee, row.employee_code, row.employee_name):
			if key and key not in by_key:
				by_key[key] = row
	return by_key


@frappe.whitelist()
def get_payroll_month_runbook(company: str, payroll_month: str, attendance_lock_version: str):
	"""Return the controlled monthly payroll checklist for one company/month/attendance lock."""
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	attendance_rows = frappe.get_all(
		MONTHLY_ATTENDANCE_DOCTYPE,
		filters=_attendance_scope_filters(company, payroll_month, attendance_lock_version),
		fields=["name", "employee", "employee_code", "employee_name"],
		limit_page_length=100000,
	)
	active_changes = _active_salary_changes_for_month(company, payroll_month)
	missing_salary_profiles = []
	trial_salary_profiles = []
	for row in attendance_rows:
		keys = (row.employee, row.employee_code, row.employee_name)
		profile = next((active_changes[key] for key in keys if key and key in active_changes), None)
		if profile and _is_salary_excluded(profile):
			continue
		if not profile:
			missing_salary_profiles.append(row.employee_code or row.employee_name or row.name)
		elif _is_trial_salary_change(profile):
			trial_salary_profiles.append(row.employee_code or row.employee_name or row.name)

	scope_filters = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
	input_count = _safe_count(PAYROLL_INPUT_DOCTYPE, scope_filters)
	settlement_count = _safe_count(PAYROLL_SETTLEMENT_DOCTYPE, scope_filters)
	confirmed_settlement_count = _safe_count(PAYROLL_SETTLEMENT_DOCTYPE, {**scope_filters, "calculation_status": "已确认"})
	confirmed_variable_filters = {"company": company, "payroll_month": payroll_month}
	if _doctype_has_field(VARIABLE_RECORD_DOCTYPE, "review_status"):
		confirmed_variable_filters["review_status"] = "已确认"
	if _doctype_has_field(VARIABLE_RECORD_DOCTYPE, "excluded"):
		confirmed_variable_filters["excluded"] = 0
	confirmed_variable_rows = frappe.get_all(VARIABLE_RECORD_DOCTYPE, filters=confirmed_variable_filters, fields=["attendance_lock_version", "source_sheet"], limit_page_length=100000)
	allowed_variable_versions = {attendance_lock_version, _monthly_variable_scope(payroll_month), ""}
	variable_count = len([
		row for row in confirmed_variable_rows
		if str(row.attendance_lock_version or "") in allowed_variable_versions
		and str(row.source_sheet or "") != "考勤终稿锁定快照"
	])
	confirmed_welfare_count = _safe_count(WELFARE_SOURCE_DOCTYPE, {**scope_filters, "confirmation_status": "已确认", "eligibility_status": "符合"})
	pending_welfare_count = _safe_count(WELFARE_SOURCE_DOCTYPE, {**scope_filters, "confirmation_status": ["in", ["草稿", "待确认"]]})
	variable_batch_rows = frappe.get_all(
		VARIABLE_BATCH_DOCTYPE,
		filters={"company": company, "payroll_month": payroll_month},
		fields=["status", "attendance_lock_version", "source_file", "source_type", "is_selected"],
		limit_page_length=10000,
	)
	variable_batch_rows = [
		row for row in variable_batch_rows
		if str(row.attendance_lock_version or "") in allowed_variable_versions
		and not str(row.source_file or "").startswith("attendance-processing-final:")
		and (not row.source_type or cint(row.is_selected))
	]
	confirmed_variable_batch_count = sum(1 for row in variable_batch_rows if row.status == "已确认")
	pending_variable_batch_count = sum(1 for row in variable_batch_rows if row.status in {"待解析", "待确认", "待审核", "已导入"})
	enabled_rule_count = _safe_count(PAYROLL_RULE_DOCTYPE, {"company": company, "status": "已启用"})
	formula_count = _safe_count(PAYROLL_RULE_DOCTYPE, {"company": company, "rule_code": ["like", "FORMULA_%"], "status": "已启用"})
	mapping_count = _safe_count(PAYROLL_FIELD_MAPPING_DOCTYPE, {"status": "已启用"})
	standard_template_filters = {}
	if _doctype_has_field("Salary Structure", "company"):
		standard_template_filters["company"] = company
	if _doctype_has_field("Salary Structure", "is_active"):
		standard_template_filters["is_active"] = 1
	standard_template_count = _safe_count("Salary Structure", standard_template_filters)
	standard_assignment_filters = {}
	if _doctype_has_field("Salary Structure Assignment", "company"):
		standard_assignment_filters["company"] = company
	if _doctype_has_field("Salary Structure Assignment", "docstatus"):
		standard_assignment_filters["docstatus"] = 1
	month_end = _month_end(payroll_month)
	if month_end and _doctype_has_field("Salary Structure Assignment", "from_date"):
		standard_assignment_filters["from_date"] = ["<=", month_end]
	standard_assignment_count = _safe_count("Salary Structure Assignment", standard_assignment_filters)

	def stage(title, summary, count, status, tone, route, action_label, detail=""):
		return {
			"title": title,
			"summary": summary,
			"count": count,
			"unit": " 条",
			"status": status,
			"tone": tone,
			"route": route,
			"action_label": action_label,
			"detail": detail,
		}

	attendance_count = len(attendance_rows)
	master_ready = attendance_count > 0 and not missing_salary_profiles and not trial_salary_profiles
	input_ready = attendance_count > 0 and master_ready and input_count == attendance_count
	settlement_ready = input_count > 0 and settlement_count == input_count
	stages = [
		stage(
			"薪资档案与异动",
			"每位进入考勤终稿的员工必须有当月有效的已批准薪资异动。",
			attendance_count - len(missing_salary_profiles) - len(trial_salary_profiles),
			"已就绪" if master_ready else "待补齐",
			"ready" if master_ready else "blocked",
			"salary-assignments",
			"处理员工分配",
			(
				"缺少 {0} 位员工的有效薪资异动；另有 {1} 位员工仍使用试运营/测试薪资。".format(
					len(missing_salary_profiles), len(trial_salary_profiles)
				)
				if missing_salary_profiles or trial_salary_profiles
				else "底薪、职能、证书/多能工与全薪从异动记录读取。"
			),
		),
		stage(
			"月度变量与福利扣款",
			"奖金、补贴、宿舍水电、社保公积金、个税和离职结算按已确认来源进入变量。",
			variable_count + confirmed_welfare_count,
			"待复核" if pending_welfare_count + pending_variable_batch_count else "已就绪",
			"warning" if pending_welfare_count + pending_variable_batch_count else "ready",
			"variables",
			"维护月度变量",
			"已确认导入批次 {0} 个、月度明细 {1} 条；待处理批次 {2} 个。".format(
				confirmed_variable_batch_count, variable_count + confirmed_welfare_count, pending_variable_batch_count + pending_welfare_count
			),
		),
		stage(
			"生成薪资输入表",
			"将锁定考勤与月度变量按员工汇总，先核对工时、加班、苹果树和扣款。",
			input_count,
			"已生成" if input_ready else "待生成",
			"ready" if input_ready else "pending",
			"inputs",
			"查看输入表",
			"输入表需与锁定考勤人数一致。",
		),
		stage(
			"试算与差异复核",
			"按公司结算公式生成所有细项、应付工资、实发工资及公司实际负担。",
			settlement_count,
			"已试算" if settlement_ready else "待试算",
			"ready" if settlement_ready else "pending",
			"settlements",
			"查看结算表",
			"结算表必须与输入表人数一致；可展开显示全部字段及来源。",
		),
		stage(
			"确认与发放",
			"复核无差异后确认本次试算；确认后不得被重新生成覆盖。",
			confirmed_settlement_count,
			"已确认" if settlement_count and confirmed_settlement_count == settlement_count else "待确认",
			"ready" if settlement_count and confirmed_settlement_count == settlement_count else "pending",
			"payroll-disbursement",
			"复核与发放",
			"工资条和发放仅消费已确认的结算结果。",
		),
	]
	warnings = []
	if missing_salary_profiles:
		warnings.append("待补薪资异动：{0}".format("、".join(missing_salary_profiles[:10])))
	if trial_salary_profiles:
		warnings.append("待替换试运营/测试薪资：{0}".format("、".join(trial_salary_profiles[:10])))
	if pending_welfare_count:
		warnings.append("存在 {0} 条待确认福利/扣款来源".format(pending_welfare_count))
	if attendance_count and input_count and input_count != attendance_count:
		warnings.append("薪资输入表人数与锁定考勤不一致")
	if input_count and settlement_count and settlement_count != input_count:
		warnings.append("薪资结算表人数与薪资输入表不一致")
	master_validation = _validate_master_step(company, payroll_month)
	salary_validation = _validate_salary_step(company, payroll_month, attendance_lock_version)
	rules_validation = _validate_rules_step(company, payroll_month, attendance_lock_version)
	attendance_rule_validation = _validate_attendance_rule_step(company, payroll_month)
	sources_validation = _validate_sources_step(company, payroll_month, attendance_lock_version)
	calculation_validation = _validate_calculation_step(company, payroll_month, attendance_lock_version)
	delivery_validation = _validate_delivery_step(company, payroll_month, attendance_lock_version)

	def readiness_area(key, title, validation, *, warning_only=False, status=""):
		blockers = validation.get("blockers") or []
		area_warnings = validation.get("warnings") or []
		if blockers:
			tone = "warning" if warning_only else "blocked"
			area_status = status or "需处理"
		elif area_warnings:
			tone, area_status = "warning", status or "请留意"
		else:
			tone, area_status = "ready", status or "数据正常"
		detail = "；".join(blockers or area_warnings) or "当前数据可用，无需人工逐步锁定。"
		return {"key": key, "title": title, "summary": detail, "status": area_status, "tone": tone, "detail": detail}

	combined_salary_validation = {
		"blockers": (salary_validation.get("blockers") or []) + (rules_validation.get("blockers") or []) + (attendance_rule_validation.get("blockers") or []),
		"warnings": (salary_validation.get("warnings") or []) + (rules_validation.get("warnings") or []) + (attendance_rule_validation.get("warnings") or []),
	}
	readiness_areas = [
		# The monthly payroll population is defined by the locked attendance
		# final, not by the live employee master.  Showing the live active count
		# here made HR compare two different scopes (for example 201 active vs
		# 192 employees in the July signed final).
		readiness_area("master", "人员范围", master_validation, warning_only=True, status=f"{attendance_count} 人"),
		readiness_area("salary", "员工定薪", combined_salary_validation),
		readiness_area("sources", "月度增减项", sources_validation),
		readiness_area("calculation", "薪资试算", calculation_validation),
		readiness_area("delivery", "确认与发放", delivery_validation),
	]
	return {
		"scope": {"company": company, "payroll_month": payroll_month, "attendance_lock_version": attendance_lock_version},
		"cards": [
			{"label": "锁定考勤", "value": attendance_count},
			{"label": "有效薪资档案", "value": attendance_count - len(missing_salary_profiles) - len(trial_salary_profiles)},
			{"label": "确认变量来源", "value": variable_count + confirmed_welfare_count},
			{"label": "薪资输入表", "value": input_count},
			{"label": "薪资结算表", "value": settlement_count},
			{"label": "已确认结算", "value": confirmed_settlement_count},
		],
		"readiness_areas": readiness_areas,
		"process_steps": readiness_areas,
		"stages": stages,
		"warning": "；".join(warnings),
	}


@frappe.whitelist()
def confirm_payroll_settlement_records(company: str, payroll_month: str, attendance_lock_version: str):
	"""Lock a reviewed payroll trial so later recalculation cannot overwrite it."""
	if not _can_manage_payroll_rules():
		frappe.throw(_("您没有确认薪资结算的权限"))
	company, payroll_month, attendance_lock_version = _require_payroll_scope(company, payroll_month, attendance_lock_version)
	_assert_workflow_locked_for_generation(company, payroll_month, attendance_lock_version)
	calculation_validation = _validate_calculation_step(company, payroll_month, attendance_lock_version)
	if calculation_validation.get("blockers"):
		frappe.throw(_("薪资确认前请处理：{0}").format("；".join(calculation_validation["blockers"])))
	scope_filters = _payroll_scope_filters(company, payroll_month, attendance_lock_version)
	attendance_count = _safe_count(MONTHLY_ATTENDANCE_DOCTYPE, _attendance_scope_filters(company, payroll_month, attendance_lock_version))
	input_count = _safe_count(PAYROLL_INPUT_DOCTYPE, scope_filters)
	settlement_count = _safe_count(PAYROLL_SETTLEMENT_DOCTYPE, scope_filters)
	pending_welfare_count = _safe_count(WELFARE_SOURCE_DOCTYPE, {**scope_filters, "confirmation_status": ["in", ["草稿", "待确认"]]})
	if not attendance_count:
		frappe.throw(_("未找到已锁定月度考勤终稿，不能确认薪资结算。"))
	if input_count != attendance_count or settlement_count != input_count:
		frappe.throw(_("薪资确认前，锁定考勤、薪资输入表和薪资结算表人数必须一致。"))
	if pending_welfare_count:
		frappe.throw(_("仍有 {0} 条福利/扣款来源待确认，不能确认薪资结算。").format(pending_welfare_count))
	for name in frappe.get_all(PAYROLL_SETTLEMENT_DOCTYPE, filters=scope_filters, pluck="name"):
		doc = frappe.get_doc(PAYROLL_SETTLEMENT_DOCTYPE, name)
		if doc.calculation_status not in ("已确认", "已生成工资单"):
			doc.calculation_status = "已确认"
			doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"confirmed": settlement_count, "company": company, "payroll_month": payroll_month, "attendance_lock_version": attendance_lock_version}


@frappe.whitelist()
def list_payroll_disbursement_records(company: str, payroll_month: str = "", attendance_lock_version: str = "", page_length: int = 100):
	company = _require_company(company)
	filters = {"company": company}
	if payroll_month:
		filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		filters["attendance_lock_version"] = attendance_lock_version
	rows = frappe.get_all(
		PAYROLL_SETTLEMENT_DOCTYPE,
		filters=filters,
		fields=["employee", "employee_code", "employee_name", "department", "gross_pay", "net_pay", "company_cost_total", "calculation_status", "modified"],
		order_by="department asc, employee_name asc",
		limit_page_length=int(page_length or 100),
	)
	for row in rows:
		row["payment_status"] = "待发放" if row.get("calculation_status") in ("已确认", "已生成工资单") else "待结算"
		row["confirmation_status"] = "已确认" if row.get("calculation_status") in ("已确认", "已生成工资单") else "待确认"
	return rows


@frappe.whitelist()
def list_payroll_report_summary(company: str, payroll_month: str = "", attendance_lock_version: str = ""):
	company = _require_company(company)
	filters = {"company": company}
	if payroll_month:
		filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		filters["attendance_lock_version"] = attendance_lock_version
	rows = frappe.get_all(
		PAYROLL_SETTLEMENT_DOCTYPE,
		filters=filters,
		fields=[
			"department",
			"gross_pay",
			"net_pay",
			"company_cost_total",
			"overtime_pay_total",
			"bonus_total",
			"punishment_total",
			"social_security_personal",
			"housing_fund_personal",
			"social_security_company",
			"housing_fund_company",
		],
		limit_page_length=100000,
	)
	summary = defaultdict(lambda: defaultdict(float))
	for row in rows:
		key = row.department or "未维护部门"
		summary[key]["headcount"] += 1
		for field in (
			"gross_pay",
			"net_pay",
			"company_cost_total",
			"overtime_pay_total",
			"bonus_total",
			"punishment_total",
			"social_security_personal",
			"housing_fund_personal",
			"social_security_company",
			"housing_fund_company",
		):
			summary[key][field] += flt(row.get(field))
	return [
		{"department": department, **{field: round(value, 2) for field, value in values.items()}}
		for department, values in sorted(summary.items(), key=lambda item: item[0])
	]


@frappe.whitelist()
def list_payroll_analysis(company: str, payroll_month: str = "", attendance_lock_version: str = ""):
	report_rows = list_payroll_report_summary(company, payroll_month, attendance_lock_version)
	totals = defaultdict(float)
	for row in report_rows:
		for key, value in row.items():
			if key != "department":
				totals[key] += flt(value)
	cost_buckets = [
		{"label": "应付工资", "value": round(totals["gross_pay"], 2)},
		{"label": "实发工资", "value": round(totals["net_pay"], 2)},
		{"label": "加班工资", "value": round(totals["overtime_pay_total"], 2)},
		{"label": "奖金福利", "value": round(totals["bonus_total"], 2)},
		{"label": "惩处扣款", "value": round(totals["punishment_total"], 2)},
		{"label": "公司社保公积金", "value": round(totals["social_security_company"] + totals["housing_fund_company"], 2)},
		{"label": "公司实际负担总计", "value": round(totals["company_cost_total"], 2)},
	]
	return {"cost_buckets": cost_buckets, "department_rows": report_rows}


@frappe.whitelist()
def list_payroll_dependency_status(company: str, payroll_month: str = "", attendance_lock_version: str = ""):
	company = _require_company(company)
	attendance_filters = {"company": company}
	input_filters = {"company": company}
	settlement_filters = {"company": company}
	welfare_filters = {"company": company}
	if payroll_month:
		attendance_filters["attendance_month"] = payroll_month
		input_filters["payroll_month"] = payroll_month
		settlement_filters["payroll_month"] = payroll_month
		welfare_filters["payroll_month"] = payroll_month
	if attendance_lock_version:
		attendance_filters["attendance_lock_version"] = attendance_lock_version
		input_filters["attendance_lock_version"] = attendance_lock_version
		settlement_filters["attendance_lock_version"] = attendance_lock_version
		welfare_filters["attendance_lock_version"] = attendance_lock_version
	return [
		{"source": "员工花名册", "doctype": "Employee", "count": _safe_count("Employee", {"company": company} if _doctype_has_field("Employee", "company") else {}), "status": "已联动"},
		{"source": "薪资主数据/薪资异动", "doctype": EMPLOYEE_SALARY_CHANGE_DOCTYPE, "count": _safe_count(EMPLOYEE_SALARY_CHANGE_DOCTYPE, {"company": company}), "status": "已联动"},
		{"source": "月度考勤终稿", "doctype": MONTHLY_ATTENDANCE_DOCTYPE, "count": _safe_count(MONTHLY_ATTENDANCE_DOCTYPE, attendance_filters), "status": "仅已锁定可入薪资"},
		{"source": "福利扣款来源", "doctype": WELFARE_SOURCE_DOCTYPE, "count": _safe_count(WELFARE_SOURCE_DOCTYPE, welfare_filters), "status": "已联动"},
		{"source": "薪资输入表", "doctype": PAYROLL_INPUT_DOCTYPE, "count": _safe_count(PAYROLL_INPUT_DOCTYPE, input_filters), "status": "公司隔离"},
		{"source": "薪资结算表", "doctype": PAYROLL_SETTLEMENT_DOCTYPE, "count": _safe_count(PAYROLL_SETTLEMENT_DOCTYPE, settlement_filters), "status": "公司隔离"},
		{"source": "薪资规则/字段映射", "doctype": PAYROLL_RULE_DOCTYPE, "count": _safe_count(PAYROLL_RULE_DOCTYPE), "status": "已联动"},
	]
