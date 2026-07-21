"""Repeatable end-to-end acceptance for the HR form intake pipeline.

This module deliberately exercises the *same* APIs exposed by the Desk:
template generation, completed-workbook preview, staging import, approval,
formal-document generation, and activation.  It only permits the isolated
``TEST-HRMS`` company and never uses real-company employees or source files.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import frappe
from frappe import _
from openpyxl import load_workbook

from hrms.api import employee_field_template as roster
from hrms.api import form_data_intake as intake
from hrms.api import form_import_demo_seed as demo_seed


TEST_COMPANY = "TEST-HRMS"
TEST_TAG = "TEST-FORM-E2E-20260717-V8"
PROTECTED_COMPANIES = ("永新", "1")
SEPARATION_EMPLOYEE_CODE = f"{TEST_TAG}-SEPARATION-001"
PERFORMANCE_EMPLOYEE_CODE = f"{TEST_TAG}-PERFORMANCE-001"


def _protected_snapshot():
	return {
		company: {
			"employees": frappe.db.count("Employee", {"company": company}),
			"latest_employee_modified": str(
				frappe.db.get_value("Employee", {"company": company}, "modified", order_by="modified desc") or ""
			),
		}
		for company in PROTECTED_COMPANIES
	}


def _assert_test_company():
	if not frappe.db.exists("Company", TEST_COMPANY):
		frappe.throw(_("端到端验收只能在 {0} 执行。").format(TEST_COMPANY))
	demo_seed._assert_test_foundation()


def _create_uploaded_file(file_name: str, content: bytes):
	return frappe.get_doc(
		{
			"doctype": "File",
			"file_name": file_name,
			"content": content,
			"is_private": 1,
		}
	).insert(ignore_permissions=True)


def _template_file_content(file_url: str):
	name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not name:
		frappe.throw(_("未找到刚生成的测试模板文件。"))
	content = frappe.get_doc("File", name).get_content()
	return content.encode() if isinstance(content, str) else content


def _filled_staging_workbook(profile):
	"""Download the real template, fill its blank data row, then upload it."""
	template = intake.create_form_import_template_file(profile["key"])
	workbook = load_workbook(BytesIO(_template_file_content(template["file_url"])))
	if workbook.sheetnames != ["填写说明", "数据"]:
		frappe.throw(_("{0} 模板工作表不符合约定：{1}").format(profile["label"], "、".join(workbook.sheetnames)))
	data_sheet = workbook["数据"]
	headers = [data_sheet.cell(row=1, column=index).value for index in range(1, len(profile["columns"]) + 1)]
	expected = [column["label"] for column in profile["columns"]]
	if headers != expected:
		frappe.throw(_("{0} 模板字段与导入契约不一致。").format(profile["label"]))
	values = demo_seed._values_for(profile)
	# Keep every V3 payroll/attendance source inside its own unlocked test month.
	for key in ("attendance_month", "payroll_month", "effective_month"):
		values[key] = "2099-07"
	for key in ("attendance_date", "occurred_on"):
		values[key] = "2099-07-06"
	values["summary_date"] = "2099-07-07"
	values["start_time"] = "2099-07-06 08:00:00"
	values["end_time"] = "2099-07-06 17:00:00"
	values["completed_at"] = "2099-07-06 18:00:00"
	# Formal HR transactions cannot become effective before their effective date.
	# Use today's date only in the isolated E2E data set.
	today = str(date.today())
	if profile["key"] == "employee_transfer":
		values["transfer_date"] = today
	if profile["key"] == "qualification_review":
		values["review_date"] = today
	if profile["key"] == "training_registration":
		values["training_content"] = f"{TEST_TAG} 培训课程"
	if profile["key"] == "performance_summary":
		values["employee_code"] = PERFORMANCE_EMPLOYEE_CODE
		values["employee_name"] = PERFORMANCE_EMPLOYEE_CODE
	if profile["key"] == "resignation_application":
		values["employee_code"] = SEPARATION_EMPLOYEE_CODE
		values["employee_name"] = SEPARATION_EMPLOYEE_CODE
	for index, column in enumerate(profile["columns"], start=1):
		data_sheet.cell(row=2, column=index, value=values.get(column["key"], f"TEST-E2E-{profile['key']}-{column['key']}"))
	output = BytesIO()
	workbook.save(output)
	return _create_uploaded_file(f"{TEST_TAG}-{profile['key']}.xlsx", output.getvalue())


def _run_roster_template_acceptance():
	"""The roster has a dedicated import route, so verify that route separately."""
	content = roster.build_employee_import_template()
	workbook = load_workbook(BytesIO(content))
	if workbook.sheetnames[:3] != ["员工花名册", "说明", "枚举字段"]:
		frappe.throw(_("员工花名册下载模板工作表不完整。"))
	fields = roster._get_employee_import_fields(roster._get_template_doc())
	sheet = workbook["员工花名册"]
	headers = [sheet.cell(row=1, column=index).value for index in range(1, len(fields) + 1)]
	if not all(str(header or "").replace(" *", "") == field["field_label"] for header, field in zip(headers, fields)):
		frappe.throw(_("员工花名册模板字段与当前字段模板不一致。"))

	code = "TEST-E2E-ROSTER-001"
	existing = frappe.db.get_value("Employee", {"custom_employee_code": code}, "name")
	if existing:
		employee = frappe.get_doc("Employee", existing)
		return {
			"template_downloaded": True,
			"template_fields": len(fields),
			"employee": employee.name,
			"employee_status": employee.status,
			"result": {"status": "existing"},
		}
	values = {
		"first_name": code,
		"employee_name": code,
		"custom_employee_code": code,
		"employee_number": code,
		"company": TEST_COMPANY,
		"department": demo_seed.TEST_DEPARTMENT,
		"designation": demo_seed.TEST_DESIGNATION,
		"employment_type": "Full-time",
		"date_of_joining": "2026-01-01",
		"date_of_birth": "1990-01-01",
		"gender": "Male",
		"cell_number": "13900000117",
		"status": "Active",
		"naming_series": "HR-EMP-",
	}
	for index, field in enumerate(fields, start=1):
		sheet.cell(row=2, column=index, value=values.get(field["fieldname"], ""))
	output = BytesIO()
	workbook.save(output)
	file_doc = _create_uploaded_file(f"{TEST_TAG}-employee-roster.xlsx", output.getvalue())
	preview = roster.preview_employee_roster_import(file_doc.file_url, mode="insert", match_by="employee_code")
	if preview.get("failed") or not preview.get("can_import"):
		frappe.throw(_("员工花名册预校验失败：{0}").format(frappe.as_json(preview)))
	imported = roster.import_employee_roster(file_doc.file_url, mode="insert", match_by="employee_code")
	if imported.get("failed") or not imported.get("inserted"):
		frappe.throw(_("员工花名册导入失败：{0}").format(frappe.as_json(imported)))
	employee = frappe.get_doc("Employee", frappe.db.get_value("Employee", {"custom_employee_code": code}, "name"))
	return {
		"template_downloaded": True,
		"template_fields": len(fields),
		"employee": employee.name,
		"employee_status": employee.status,
		"result": imported,
	}


def _approval_to_final(row_name: str):
	"""Advance every configured approval step using the test administrator."""
	steps = 0
	while True:
		row = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, row_name)
		if row.review_status == "已批准":
			return steps
		if row.review_status in ("已驳回",) or row.status in ("处理失败", "已忽略"):
			frappe.throw(_("测试行 {0} 无法继续审核，当前状态：{1}/{2}").format(row_name, row.status, row.review_status))
		intake.review_form_import_row(row_name, "批准", f"{TEST_TAG}：第 {steps + 1} 个审批节点通过")
		steps += 1
		if steps > 5:
			frappe.throw(_("测试审批节点数量异常。"))


def _active_appraisal_cycle():
	return frappe.db.get_value("Appraisal Cycle", {"status": "In Progress"}, "name", order_by="start_date desc") or ""


def _verify_activated_target(row):
	target = frappe.get_doc(row.target_doctype, row.target_name)
	if frappe.get_meta(target.doctype).is_submittable:
		if target.docstatus != 1:
			frappe.throw(_("{0} 未提交。").format(target.doctype))
		if target.doctype == "Employee Onboarding":
			if not target.activities or any(not activity.task for activity in target.activities):
				frappe.throw(_("入职单未生成完整办理任务。"))
			if target.boarding_status not in ("Pending", "In Process", "Completed"):
				frappe.throw(_("入职单状态异常：{0}").format(target.boarding_status))
	elif target.doctype == intake.BUSINESS_PROCESS_RECORD_DOCTYPE and target.status not in ("待跟进", "已生效"):
		frappe.throw(_("正式业务记录未生效。"))
	elif target.doctype == "HRMS Employee Salary Change" and target.status != "已批准":
		frappe.throw(_("薪资异动没有进入已批准状态。"))
	elif target.doctype == "HRMS Payroll Welfare Source Record" and target.confirmation_status != "已确认":
		frappe.throw(_("薪资来源没有确认。"))
	elif target.doctype == "HRMS Attendance Exception" and target.confirmation_status != "已确认":
		frappe.throw(_("考勤异常没有确认。"))
	elif target.doctype == "HRMS Monthly Attendance Summary" and target.status != "已确认":
		frappe.throw(_("月度考勤终稿没有确认。"))
	return target


def _activation_order(profiles):
	"""Keep employee-state transactions last, after all inputs have been consumed."""
	last = {"employee_transfer", "qualification_review", "resignation_application"}
	return [item for item in profiles if item["key"] not in last] + [item for item in profiles if item["key"] in last]


def _ensure_test_attendance_lock(attendance_month):
	from hrms.api import attendance_import

	lock = frappe.db.get_value(
		"HRMS Attendance Month Lock",
		{"company": TEST_COMPANY, "attendance_month": attendance_month},
		["name", "status", "active_version"],
		as_dict=True,
	)
	if lock and lock.status == "已锁定":
		return str(lock.active_version)
	return attendance_import.lock_attendance_month(TEST_COMPANY, attendance_month, f"{TEST_TAG}：考勤终稿审核通过后锁定")[
		"attendance_lock_version"
	]


def _ensure_test_holiday_list(employee):
	"""Employee Separation validates the actual effective day against Holiday List."""
	name = "TEST-HRMS E2E 2026 Holiday List"
	if not frappe.db.exists("Holiday List", name):
		frappe.get_doc(
			{
				"doctype": "Holiday List",
				"holiday_list_name": name,
				"from_date": "2026-01-01",
				"to_date": "2026-12-31",
			}
		).insert(ignore_permissions=True)
	holiday_list = frappe.get_doc("Holiday List", name)
	if not any(str(row.holiday_date) == str(date.today()) for row in holiday_list.holidays):
		holiday_list.append("holidays", {"holiday_date": str(date.today()), "description": f"{TEST_TAG} separation activation fixture"})
		holiday_list.save(ignore_permissions=True)
	if frappe.db.get_value("Employee", employee, "holiday_list") != name:
		frappe.db.set_value("Employee", employee, "holiday_list", name)
	assignment = frappe.db.get_value(
		"Holiday List Assignment", {"assigned_to": employee, "holiday_list": name, "docstatus": 1}, "name"
	)
	if not assignment:
		assignment_doc = frappe.get_doc(
			{
				"doctype": "Holiday List Assignment",
				"naming_series": "HR-HLA-.YYYY.-",
				"applicable_for": "Employee",
				"assigned_to": employee,
				"holiday_list": name,
				"from_date": "2026-01-01",
			}
		).insert(ignore_permissions=True)
		assignment_doc.submit()


def _ensure_separation_employee():
	return _ensure_test_employee(SEPARATION_EMPLOYEE_CODE, with_holiday_list=True)


def _ensure_test_employee(code, with_holiday_list=False):
	name = frappe.db.get_value("Employee", {"custom_employee_code": code}, "name")
	if not name:
		name = frappe.get_doc(
			{
				"doctype": "Employee",
				"naming_series": "HR-EMP-",
				"first_name": code,
				"employee_name": code,
				"custom_employee_code": code,
				"employee_number": code,
				"company": TEST_COMPANY,
				"department": demo_seed.TEST_DEPARTMENT,
				"designation": demo_seed.TEST_DESIGNATION,
				"employment_type": "Full-time",
				"date_of_joining": str(date.today()),
				"date_of_birth": "1990-01-01",
				"gender": "Male",
				"status": "Active",
			}
		).insert(ignore_permissions=True).name
	if with_holiday_list:
		_ensure_test_holiday_list(name)
	return name


@frappe.whitelist()
def run_form_import_e2e_acceptance(tag: str = TEST_TAG):
	"""Execute the full 28-template acceptance suite in TEST-HRMS.

	The suite is resumable. A repeated call reuses the tagged batches and only
	continues incomplete review/generation/activation steps.
	"""
	if tag != TEST_TAG:
		frappe.throw(_("本验收仅允许固定测试标识 {0}。").format(TEST_TAG))
	_assert_test_company()
	protected_before = _protected_snapshot()
	intake.ensure_default_form_approval_matrices(TEST_COMPANY)
	result = {
		"company": TEST_COMPANY,
		"tag": tag,
		"roster": _run_roster_template_acceptance(),
		"templates": {},
		"errors": [],
		"protected_before": protected_before,
	}
	result["separation_employee"] = _ensure_separation_employee()
	result["performance_employee"] = _ensure_test_employee(PERFORMANCE_EMPLOYEE_CODE)

	# 1. Exact download -> fill -> preview -> import for every staging template.
	for profile in intake.FORM_IMPORT_PROFILES:
		if profile.get("entry_mode") == "employee_roster":
			continue
		batch_name = frappe.db.get_value(
			intake.FORM_IMPORT_BATCH_DOCTYPE,
			{"company": TEST_COMPANY, "template_key": profile["key"], "notes": tag},
			"name",
		)
		try:
			if not batch_name:
				file_doc = _filled_staging_workbook(profile)
				preview = intake.preview_form_import(file_doc.file_url, profile["key"], TEST_COMPANY)
				if preview.get("missing_required") or preview.get("failed_rows") or preview.get("valid_rows") != 1:
					frappe.throw(_("模板填写后的预校验失败：{0}").format(frappe.as_json(preview)))
				imported = intake.import_form_workbook(file_doc.file_url, profile["key"], TEST_COMPANY, notes=tag)
				if imported.get("failed_rows") or imported.get("valid_rows") != 1:
					frappe.throw(_("模板导入失败：{0}").format(frappe.as_json(imported)))
				batch_name = imported["batch_name"]
			batch = frappe.get_doc(intake.FORM_IMPORT_BATCH_DOCTYPE, batch_name)
			row_name = frappe.db.get_value(intake.FORM_IMPORT_ROW_DOCTYPE, {"import_batch": batch_name, "row_number": 2}, "name")
			if not row_name:
				frappe.throw(_("导入批次未创建可审核行。"))
			result["templates"][profile["key"]] = {"label": profile["label"], "batch": batch_name, "row": row_name, "download_fill_preview_import": "passed"}
		except Exception as error:
			result["errors"].append({"template_key": profile["key"], "stage": "download_fill_preview_import", "error": str(error)})

	# 2. Approve -> formal document -> activate, then verify the persisted target.
	cycle = _active_appraisal_cycle()
	payroll_lock_version = ""
	for profile in _activation_order([item for item in intake.FORM_IMPORT_PROFILES if item.get("entry_mode") != "employee_roster"]):
		item = result["templates"].get(profile["key"])
		if not item:
			continue
		try:
			if profile["key"] in intake.PAYROLL_TEMPLATE_KEYS and not payroll_lock_version:
				if result["templates"].get("attendance_final", {}).get("activation") != "passed":
					frappe.throw(_("薪资资料验收前必须先完成月度考勤终稿确认。"))
				payroll_lock_version = _ensure_test_attendance_lock("2099-07")
			if profile["key"] == "resignation_application":
				row_for_holiday = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, item["row"])
				_ensure_test_holiday_list(row_for_holiday.employee)
			item["approval_steps"] = _approval_to_final(item["row"])
			row = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, item["row"])
			if not row.target_name:
				generated = intake.generate_form_import_target(
					item["row"],
					attendance_lock_version=payroll_lock_version if profile["key"] in intake.PAYROLL_TEMPLATE_KEYS else "",
					appraisal_cycle=cycle if profile["key"] == "performance_summary" else "",
				)
				item["target_doctype"] = generated["target_doctype"]
				item["target_name"] = generated["target_name"]
			else:
				item["target_doctype"] = row.target_doctype
				item["target_name"] = row.target_name
			row = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, item["row"])
			if row.status != "已提交生效":
				intake.activate_form_import_target(item["row"])
			row = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, item["row"])
			if row.status != "已提交生效":
				frappe.throw(_("导入行没有进入已提交生效状态。"))
			target = _verify_activated_target(row)
			item.update({"target_doctype": target.doctype, "target_name": target.name, "activation": "passed"})
		except Exception as error:
			result["errors"].append({"template_key": profile["key"], "stage": "approval_generate_activate", "error": str(error)})

	result["protected_after"] = _protected_snapshot()
	if protected_before != result["protected_after"]:
		frappe.throw(_("端到端验收触碰了受保护公司：{0}").format(frappe.as_json(result)))
	result["summary"] = {
		"profile_count": len(intake.FORM_IMPORT_PROFILES),
		"staging_templates": len(intake.FORM_IMPORT_PROFILES) - 1,
		"download_fill_preview_import_passed": sum(1 for item in result["templates"].values() if item.get("download_fill_preview_import") == "passed"),
		"activation_passed": sum(1 for item in result["templates"].values() if item.get("activation") == "passed"),
		"failed": len(result["errors"]),
	}
	frappe.db.commit()
	return result
