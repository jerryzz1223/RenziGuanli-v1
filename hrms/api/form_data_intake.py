"""Template-driven intake for the HR forms supplied by Yongxin.

The import layer deliberately stores signed / workflow forms as normalized staging
records.  A spreadsheet upload must not silently change an employee to left,
promoted or transferred before the responsible HR workflow has confirmed it.
"""

import hashlib
import json
import re
from datetime import date, datetime
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt, getdate, now_datetime


FORM_IMPORT_BATCH_DOCTYPE = "HRMS Form Import Batch"
FORM_IMPORT_ROW_DOCTYPE = "HRMS Form Import Row"
FORM_APPROVAL_MATRIX_DOCTYPE = "HRMS Form Approval Matrix"
BUSINESS_PROCESS_RECORD_DOCTYPE = "HRMS Business Process Record"
EMPLOYEE_ROSTER_TEMPLATE_KEY = "employee_roster"


DEFAULT_FORM_APPROVAL_ROUTES = {
	"payroll": [
		{"step_no": 1, "step_label": "人事复核", "approver_role": "HR Manager"},
		{"step_no": 2, "step_label": "财务复核", "approver_role": "Accounts Manager"},
	],
	"organization": [
		{"step_no": 1, "step_label": "人事复核", "approver_role": "HR Manager"},
		{"step_no": 2, "step_label": "组织管理员确认", "approver_role": "System Manager"},
	],
	"standard": [{"step_no": 1, "step_label": "人事复核", "approver_role": "HR Manager"}],
}

PAYROLL_TEMPLATE_KEYS = {
	"salary_structure_change", "reward_punishment", "skill_certificate_allowance", "full_attendance_bonus",
	"housing_allowance", "education_allowance", "dormitory_fee", "social_insurance", "service_award",
	"exit_payroll_settlement",
}
ORGANIZATION_TEMPLATE_KEYS = {"org_structure"}
BUSINESS_PROCESS_TEMPLATE_CONFIG = {
	"org_structure": {"record_type": "组织变更", "date_keys": ("effective_date",), "title": lambda data: _("组织变更：{0}").format(data.get("department") or "")},
	"contract_intent": {"record_type": "合同续签意愿", "date_keys": ("contract_end_date", "survey_date"), "title": lambda data: _("合同意愿：{0}").format(data.get("employee_name") or "")},
	"certificate_management": {"record_type": "员工证书档案", "date_keys": ("validity_period", "next_review_due", "first_issue_date"), "title": lambda data: _("证书档案：{0}").format(data.get("certificate_no") or data.get("employee_name") or "")},
	"attendance_department_summary": {"record_type": "部门考勤核对", "date_keys": ("summary_date",), "title": lambda data: _("部门考勤核对：{0}").format(data.get("department") or "")},
	"proposal_improvement": {"record_type": "提案改善", "date_keys": ("proposal_date",), "title": lambda data: _("提案改善：{0}").format(data.get("subject") or data.get("proposal_no") or "")},
	"system_feedback": {"record_type": "系统反馈", "date_keys": ("followup_date", "completed_date"), "title": lambda data: _("系统反馈：{0}").format(data.get("feedback_no") or data.get("description") or "")},
}


def _column(key, label, required=False, aliases=None):
	return {"key": key, "label": label, "required": required, "aliases": aliases or []}


# One profile per usable data form.  Several source sheets are intentionally
# consolidated where the business meaning is identical (e.g. finance/sign-off
# attendance final drafts).  The source_sheets values make every workbook sheet
# traceable to its receiving business area.
FORM_IMPORT_PROFILES = [
	{
		"key": EMPLOYEE_ROSTER_TEMPLATE_KEY,
		"module": "人事",
		"label": "员工花名册",
		"description": "员工主档、联系方式、教育和合同保险基础资料。此表通过现有智能花名册导入直接写入 Employee。",
		"source_sheets": ["花名册"],
		"processing_target": "员工主档（直接导入）",
		"entry_mode": "employee_roster",
		"columns": [
			_column("employee_code", "工号", True, ["员工编码", "员工编号"]),
			_column("employee_name", "姓名", True, ["员工姓名"]),
			_column("department", "部门", True, ["单位"]),
			_column("date_of_joining", "入职日期", True, ["到职日期"]),
			_column("designation", "岗位", True, ["职位", "职务"]),
			_column("employment_type", "工作性质", False, ["员工类型", "雇佣类型"]),
			_column("cell_number", "联系电话", False, ["手机号", "电话"]),
			_column("id_card", "身份证号码", False, ["身份证", "证件号码"]),
			_column("education_level", "学历", False),
			_column("graduation_school", "毕业院校", False),
			_column("major", "科系", False, ["专业"]),
		],
	},
	{
		"key": "org_structure",
		"module": "组织",
		"label": "组织架构与编制",
		"description": "公司、部门、上级部门、负责人、岗位及编制；先入组织数据池，再由组织管理员确认发布。",
		"source_sheets": ["26Q3组织架构图"],
		"processing_target": "组织变更正式记录 / 部门发布（确认后）",
		"target_doctype": "HRMS Business Process Record",
		"columns": [
			_column("company", "公司", True), _column("department", "部门", True), _column("parent_department", "上级部门"),
			_column("department_head_code", "负责人工号"), _column("department_head_name", "负责人姓名"),
			_column("designation", "岗位"), _column("headcount", "编制人数"), _column("effective_date", "生效日期"), _column("remarks", "备注"),
		],
	},
	{
		"key": "employee_transfer",
		"module": "人事",
		"label": "员工职务调动申请",
		"description": "转岗、调部门和岗位异动申请；入库后由人事异动流程确认，确认前不改员工当前任职。",
		"source_sheets": ["员工职务调动申请表"],
		"processing_target": "人事异动（确认后）",
		"target_doctype": "Employee Transfer",
		"columns": [
			_column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("transfer_date", "调动日期", True),
			_column("from_department", "调出部门", True), _column("from_designation", "原职务"),
			_column("to_department", "调入部门", True), _column("to_designation", "调动后职务", True),
			_column("reason", "调动理由", True, ["理由"]), _column("remarks", "备注"),
		],
	},
	{
		"key": "qualification_review",
		"module": "人事",
		"label": "人员职能资格认定",
		"description": "试用、转正、晋升及资格认定资料；通过后再生成转正/晋升业务单。",
		"source_sheets": ["人员职能资格认定表"],
		"processing_target": "转正 / 晋升（确认后）",
		"target_doctype": "Employee Promotion",
		"columns": [
			_column("employee_code", "工号", True), _column("employee_name", "被考核人", True, ["姓名"]), _column("department", "部门", True),
			_column("designation", "岗位", True), _column("grade", "职等"), _column("review_date", "认定日期", True),
			_column("evaluation", "最终评价", True), _column("interview_record", "面谈记录"), _column("remarks", "备注"),
		],
	},
	{
		"key": "contract_intent",
		"module": "人事",
		"label": "劳动合同到期意愿调查",
		"description": "合同续签、终止或离职意愿；仅生成待办资料，不自动变更合同。",
		"source_sheets": ["人事组员工劳动合同到期意愿调查表"],
		"processing_target": "合同续签意愿正式记录（确认后）",
		"target_doctype": "HRMS Business Process Record",
		"columns": [
			_column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门", True),
			_column("contract_end_date", "合同到期日", True), _column("contract_type", "合同类型"), _column("employee_intent", "员工意愿", True),
			_column("survey_date", "调查日期", True), _column("remarks", "备注"),
		],
	},
	{
		"key": "resignation_application",
		"module": "人事",
		"label": "员工辞职申请",
		"description": "离职申请和交接信息；必须经离职流程确认才会变更员工状态。",
		"source_sheets": ["人事组员工辞职申请单"],
		"processing_target": "离职管理 / 离职面谈（确认后）",
		"target_doctype": "Employee Separation",
		"columns": [
			_column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门", True),
			_column("designation", "职务", False, ["岗位"]), _column("application_date", "申请日期", True),
			_column("last_working_date", "最后工作日", True, ["离职日期"]), _column("reason", "离职原因", True),
			_column("handover_status", "交接状态"), _column("remarks", "备注"),
		],
	},
	{
		"key": "recruitment_interview",
		"module": "招聘",
		"label": "候选人面试清单",
		"description": "简历筛选、面试、录用、报到与离职追踪，入库后供招聘漏斗处理。",
		"source_sheets": ["2026年度人员面试清单"],
		"processing_target": "候选人 / 面试（确认后）",
		"target_doctype": "Job Applicant",
		"columns": [
			_column("candidate_name", "姓名", True), _column("phone", "联系电话", True), _column("source_channel", "信息渠道"),
			_column("gender", "性别"), _column("education", "学历"), _column("interview_date", "面试时间", True),
			_column("applied_designation", "面试岗位", True), _column("interviewer", "面试官"), _column("decision", "录用与否"),
			_column("assigned_department", "分配部门"), _column("expected_join_date", "预计上班日期"), _column("remarks", "备注"),
		],
	},
	{
		"key": "attendance_daily",
		"module": "考勤",
		"label": "每日考勤与人工调整",
		"description": "钉钉日考勤导出或人工调整。原始与调整数据都会保留，后续由考勤导入中心生成异常和月度终稿。",
		"source_sheets": ["每日统计（钉钉导出）", "每日统计（修改后）"],
		"processing_target": "考勤日核验（后续处理）",
		"target_doctype": "HRMS Attendance Day Check",
		"columns": [
			_column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("attendance_date", "日期", True),
			_column("department", "实际部门", False, ["部门", "单位"]), _column("shift_name", "班次"), _column("actual_in_time", "上班时间"),
			_column("actual_out_time", "下班时间"), _column("standard_hours", "标准工时"), _column("actual_hours", "实际出勤（小时）", False, ["实际出勤"]),
			_column("workday_overtime_hours", "工作日加班（小时）"), _column("restday_overtime_hours", "休息日加班（小时）"),
			_column("holiday_overtime_hours", "节假日加班（小时）"), _column("approval_reference", "关联审批单"), _column("remarks", "备注"),
		],
	},
	{
		"key": "attendance_department_summary",
		"module": "考勤",
		"label": "部门出勤明细",
		"description": "按部门日报的现有人数、出勤、请假和备注，用于与个人考勤交叉核验。",
		"source_sheets": ["出勤明细"],
		"processing_target": "部门考勤核对正式记录",
		"target_doctype": "HRMS Business Process Record",
		"columns": [
			_column("summary_date", "统计日期", True), _column("department", "部门", True), _column("regular_headcount", "正式员工人数"),
			_column("probation_headcount", "试用员工人数"), _column("total_headcount", "合计人数", True), _column("attendance_count", "出勤人数", True),
			_column("leave_count", "请假人数"), _column("leave_employee_names", "请假人员"), _column("remarks", "备注"),
		],
	},
	{
		"key": "leave_export",
		"module": "考勤",
		"label": "请假审批导出",
		"description": "钉钉请假明细及审批状态；只有审批通过且已结束的记录才能成为有效请假证据。",
		"source_sheets": ["请假单（钉钉导出）"],
		"processing_target": "请假审批凭证（后续处理）",
		"target_doctype": "HRMS Attendance Leave Evidence",
		"columns": [
			_column("external_id", "数据id", True), _column("employee_name", "创建人", True, ["姓名"]), _column("department", "创建人部门"),
			_column("leave_type", "请假类型（实际）", True, ["请假类型"]), _column("start_time", "开始时间", True), _column("end_time", "结束时间", True),
			_column("duration", "时长", True), _column("reason", "请假事由"), _column("approval_no", "审批编号"),
			_column("approval_result", "审批结果"), _column("approval_status", "审批状态", True), _column("completed_at", "完成时间"),
		],
	},
	{
		"key": "attendance_exception",
		"module": "考勤",
		"label": "出勤异常确认",
		"description": "迟到、早退、旷工、未打卡和补卡确认，入库后供异常处理人员闭环。",
		"source_sheets": ["出勤异常"],
		"processing_target": "考勤异常（确认后）",
		"target_doctype": "HRMS Attendance Exception",
		"columns": [
			_column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("attendance_date", "出勤日期", True),
			_column("department", "单位", True, ["部门"]), _column("shift_name", "应上班时间"), _column("actual_in_time", "实际上班时间"),
			_column("actual_out_time", "实际下班时间"), _column("exception_type", "异常类型", True), _column("handling", "处理方式"), _column("remarks", "备注"),
		],
	},
	{
		"key": "apple_reward",
		"module": "考勤",
		"label": "苹果树奖惩记录",
		"description": "钉钉苹果树原始记录及人工调整，作为月度奖惩和薪酬变量来源。",
		"source_sheets": ["苹果树（钉钉导出）", "苹果树（修改后）"],
		"processing_target": "苹果树奖惩明细（后续处理）",
		"target_doctype": "HRMS Apple Reward Record",
		"columns": [
			_column("external_id", "数据id"), _column("employee_code", "工号", False), _column("employee_name", "姓名", True),
			_column("department", "部门", False, ["单位"]), _column("occurred_on", "发生日期", True, ["日期"]),
			_column("apple_type", "苹果类型", True), _column("quantity", "数量", True), _column("reason", "事由"),
			_column("approval_status", "审批状态"), _column("approval_result", "审批结果"), _column("remarks", "备注"),
		],
	},
	{
		"key": "attendance_final",
		"module": "考勤",
		"label": "月度考勤终稿",
		"description": "签字版、财务版和初稿的月度工时结果，用于对账，不覆盖原始日考勤。",
		"source_sheets": ["考勤初稿", "考勤终稿（签字版）", "考勤终稿（财务版）"],
		"processing_target": "月度考勤终稿（确认后）",
		"target_doctype": "HRMS Monthly Attendance Summary",
		"columns": [
			_column("attendance_month", "考勤月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True),
			_column("department", "部门", True), _column("standard_hours", "标准工时（小时）"), _column("actual_hours", "实际出勤（小时）"),
			_column("workday_overtime_hours", "工作日加班（小时）"), _column("restday_overtime_hours", "休息日加班（小时）"),
			_column("holiday_overtime_hours", "节假日加班（小时）"), _column("absence_hours", "调整后缺勤工时"), _column("green_apple_amount", "绿苹果"),
			_column("red_apple_amount", "红苹果"), _column("housing_allowance", "住房补贴"), _column("full_attendance_bonus", "全勤"), _column("remarks", "备注"),
		],
	},
	{
		"key": "salary_structure_change",
		"module": "薪酬",
		"label": "薪资构成调整",
		"description": "员工薪资版本、底薪、津贴及社保承担调整，进入薪资主数据审核队列。",
		"source_sheets": ["薪资构成"],
		"processing_target": "员工薪资主数据 / 薪资结构分配（确认后）",
		"target_doctype": "HRMS Employee Salary Change",
		"columns": [
			_column("effective_month", "调整月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门", True),
			_column("change_reason", "异动原因", True), _column("previous_base_salary", "调整前底薪"), _column("base_salary", "调整后底薪", True),
			_column("functional_allowance", "职能津贴"), _column("position_allowance", "职务津贴"), _column("certificate_allowance", "证书津贴"),
			_column("multi_skill_allowance", "多能工津贴"), _column("gross_salary", "全薪"), _column("remarks", "备注"),
		],
	},
	{
		"key": "reward_punishment",
		"module": "薪酬",
		"label": "奖惩提报",
		"description": "奖惩申请、人事确认和财务金额提报；审核后生成薪酬福利/扣款来源，并同步为薪资变量。",
		"source_sheets": ["奖惩提报单（提交人事）", "奖惩提报单（提交财务）"],
		"processing_target": "薪酬福利/扣款来源 / 薪资变量（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门", True), _column("occurred_on", "发生日", True),
			_column("subject", "主旨", True), _column("reward_punishment_type", "奖惩类型", True), _column("rule", "奖惩条例"), _column("standard", "标准"), _column("amount", "金额（元）"), _column("remarks", "备注"),
		],
	},
	{
		"key": "skill_certificate_allowance",
		"module": "薪酬",
		"label": "证书与多能工津贴",
		"description": "证书、多能工及职务津贴月度名单，作为福利扣款来源资料。",
		"source_sheets": ["证书、多能工津贴名单"],
		"processing_target": "福利扣款来源 / 薪酬变量（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("payroll_month", "月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门", True),
			_column("multi_skill_allowance", "多能工津贴"), _column("certificate_allowance", "证书津贴"), _column("remarks", "备注"),
		],
	},
	{
		"key": "full_attendance_bonus",
		"module": "薪酬",
		"label": "全勤奖",
		"description": "全勤奖月度名单；金额可由规则计算或通过导入确认。",
		"source_sheets": ["全勤奖"],
		"processing_target": "薪酬变量（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("payroll_month", "月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "单位", True, ["部门"]),
			_column("amount", "全勤奖", True), _column("remarks", "备注"),
		],
	},
	{
		"key": "housing_allowance",
		"module": "薪酬",
		"label": "住房补贴",
		"description": "住房补贴月度名单及住房情况，作为薪酬变量来源。",
		"source_sheets": ["住房补贴终稿"],
		"processing_target": "薪酬变量（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("payroll_month", "月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "单位", True, ["部门"]),
			_column("amount", "住房补贴", True), _column("housing_status", "住房情况"), _column("remarks", "备注"),
		],
	},
	{
		"key": "education_allowance",
		"module": "薪酬",
		"label": "学历补贴",
		"description": "学历及补贴资格名单，可回填员工档案并作为薪酬变量来源。",
		"source_sheets": ["学历补贴"],
		"processing_target": "员工档案 / 薪酬变量（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("payroll_month", "月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门", True),
			_column("designation", "岗位"), _column("education_category", "学历类别"), _column("education_level", "学历"), _column("major", "专业"),
			_column("amount", "补贴金额（元）", True), _column("remarks", "备注"),
		],
	},
	{
		"key": "dormitory_fee",
		"module": "薪酬",
		"label": "宿舍费与水电扣款",
		"description": "住宿、水电、搭伙及补贴抵扣明细，作为薪酬扣款来源。",
		"source_sheets": ["宿舍费"],
		"processing_target": "福利扣款来源（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("payroll_month", "月份", True), _column("employee_code", "工号", False), _column("employee_name", "姓名", True), _column("department", "部门", True),
			_column("building", "幢数"), _column("floor", "楼层"), _column("dormitory_type", "地点"), _column("accommodation_days", "住宿天数"),
			_column("rent_amount", "应收金额"), _column("utilities_amount", "水电费"), _column("deduction_amount", "当月扣款", True), _column("remarks", "备注"),
		],
	},
	{
		"key": "social_insurance",
		"module": "薪酬",
		"label": "社保公积金名单",
		"description": "社保、医疗及个人/公司承担额，进入月度薪资扣款和成本资料池。",
		"source_sheets": ["2606社保名单"],
		"processing_target": "社保公积金记录 / 薪酬变量（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("payroll_month", "月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门", True),
			_column("social_security_base", "社保基数", True), _column("medical_base", "医疗基数"), _column("company_amount", "公司合计承担"),
			_column("employee_amount", "个人合计承担", True), _column("remarks", "备注"),
		],
	},
	{
		"key": "service_award",
		"module": "薪酬",
		"label": "继续服务奖",
		"description": "按月发放的继续服务奖，作为薪酬变量来源。",
		"source_sheets": ["6月继续服务奖"],
		"processing_target": "薪酬变量（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("payroll_month", "月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门", True), _column("amount", "金额", True), _column("remarks", "备注"),
		],
	},
	{
		"key": "proposal_improvement",
		"module": "绩效",
		"label": "提案改善",
		"description": "改善提案、问题、效益和跟进状态，可作为绩效和奖金依据。",
		"source_sheets": ["提案改善表"],
		"processing_target": "提案改善正式记录 / 奖金依据",
		"target_doctype": "HRMS Business Process Record",
		"columns": [
			_column("proposal_no", "提案编号", True), _column("proposal_date", "提案日期", True), _column("employee_name", "提案人", True), _column("department", "部门", True),
			_column("subject", "提案主题", True), _column("background", "提案背景"), _column("improvement", "改善内容"), _column("expected_benefit", "预期效益"), _column("status", "解决状态"), _column("remarks", "备注"),
		],
	},
	{
		"key": "exit_payroll_settlement",
		"module": "薪酬",
		"label": "离职人员薪资结算",
		"description": "离职工资、加班、扣款、社保和实发金额；审核后作为离职薪资来源进入结算，正式实发工资仍由薪资引擎生成。",
		"source_sheets": ["离职人员薪资结算"],
		"processing_target": "离职薪资来源 / 薪资变量（确认后）",
		"target_doctype": "HRMS Payroll Welfare Source Record",
		"columns": [
			_column("payroll_month", "月份", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True), _column("department", "部门"),
			_column("join_date", "到职日期"), _column("exit_date", "离职日期", True), _column("base_salary", "底薪"), _column("allowance", "津贴"),
			_column("overtime_pay", "加班费"), _column("gross_pay", "应付工资"), _column("deduction_amount", "扣款合计"), _column("net_pay", "实发工资", True), _column("remarks", "备注"),
		],
	},
	{
		"key": "training_registration",
		"module": "培训",
		"label": "教育训练登记",
		"description": "培训课程、对象、课时、成绩和评价，支持后续培训记录和技能分析。",
		"source_sheets": ["教育训练登记表"],
		"processing_target": "培训活动 / 培训结果（确认后）",
		"target_doctype": "Training Event",
		"columns": [
			_column("training_month", "月份", True), _column("employee_name", "姓名", True), _column("department", "部门", True), _column("course_type", "课程类型"),
			_column("training_content", "培训内容", True), _column("owner_department", "课程归属部门"), _column("training_mode", "内/外训"), _column("hours", "课时"),
			_column("trainer", "授课人"), _column("location", "地点"), _column("score", "成绩"), _column("training_date", "实际上课时间", True), _column("remarks", "备注（评价标准）", False, ["备注"]),
		],
	},
	{
		"key": "certificate_management",
		"module": "培训",
		"label": "证书管理",
		"description": "安全管理、特种作业证书及复审计划，可用于证书到期提醒和津贴资格。",
		"source_sheets": ["证书管理清单"],
		"processing_target": "员工证书正式档案（确认后）",
		"target_doctype": "HRMS Business Process Record",
		"columns": [
			_column("employee_name", "姓名", True), _column("unit_type", "单位类型"), _column("person_type", "人员类型"), _column("certificate_no", "证书编号", True),
			_column("first_issue_date", "初次取证日期"), _column("validity_period", "证书有效期", True), _column("review_frequency", "复审频次"),
			_column("next_review_due", "下次复审到期月份"), _column("review_status", "复审完成情况"), _column("new_certificate_price", "新证价格"), _column("review_price", "复审价格"), _column("remarks", "备注"),
		],
	},
	{
		"key": "performance_summary",
		"module": "绩效",
		"label": "半年度/年度工作总结",
		"description": "员工工作总结、绩效事迹、问题和下一期计划，进入绩效评估资料池。",
		"source_sheets": ["年度工作总结及展望"],
		"processing_target": "绩效考核 / 绩效反馈（确认后）",
		"target_doctype": "Appraisal",
		"columns": [
			_column("period_type", "周期类型", True), _column("period", "考核周期", True), _column("employee_code", "工号", True), _column("employee_name", "姓名", True),
			_column("department", "部门", True), _column("join_date", "入职时间"), _column("summary_date", "填写日期", True), _column("achievements", "绩效事迹", True),
			_column("improvements", "待改善事项"), _column("next_plan", "下一期计划"), _column("remarks", "备注"),
		],
	},
	{
		"key": "system_feedback",
		"module": "系统反馈",
		"label": "系统问题与改进清单",
		"description": "系统使用问题、截图、处理状态和跟进日期，供产品迭代闭环。",
		"source_sheets": ["第一次260707"],
		"processing_target": "系统反馈正式工单",
		"target_doctype": "HRMS Business Process Record",
		"columns": [
			_column("feedback_no", "序号", True), _column("category", "大类"), _column("page_or_feature", "图片/问题"), _column("description", "问题描述", True),
			_column("status", "解决状态"), _column("followup_date", "跟进日期"), _column("completed_date", "日期"), _column("remarks", "备注"),
		],
	},
]


def _normalise_text(value):
	if value is None:
		return ""
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d %H:%M:%S")
	if isinstance(value, date):
		return value.isoformat()
	return str(value).strip()


def _normalise_header(value):
	return re.sub(r"\s+", "", _normalise_text(value).replace("（", "(").replace("）", ")").replace("\n", ""))


def _profile_map():
	return {profile["key"]: profile for profile in FORM_IMPORT_PROFILES}


def _profile_or_throw(template_key):
	profile = _profile_map().get(template_key)
	if not profile:
		frappe.throw(_("未找到该人资表单模板"))
	return profile


def _get_file_content(file_url):
	file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not file_name:
		frappe.throw(_("未找到上传文件"))
	content = frappe.get_doc("File", file_name).get_content()
	return content.encode() if isinstance(content, str) else content


def _load_workbook(file_url):
	from openpyxl import load_workbook

	return load_workbook(BytesIO(_get_file_content(file_url)), data_only=True, read_only=True)


def _profile_alias_map(profile):
	aliases = {}
	for column in profile["columns"]:
		for label in [column["label"], *column.get("aliases", [])]:
			aliases[_normalise_header(label)] = column["key"]
	return aliases


def _find_sheet_and_header(workbook, profile):
	aliases = _profile_alias_map(profile)
	required = {column["key"] for column in profile["columns"] if column.get("required")}
	candidates = []
	ordered_names = [name for name in profile.get("source_sheets", []) if name in workbook.sheetnames]
	ordered_names.extend(name for name in workbook.sheetnames if name not in ordered_names)
	for sheet_name in ordered_names:
		sheet = workbook[sheet_name]
		for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True), start=1):
			matched = {aliases.get(_normalise_header(value)) for value in row if aliases.get(_normalise_header(value))}
			if not matched:
				continue
			score = len(matched & required) * 10 + len(matched)
			candidates.append((score, sheet, row_index, row, matched))
	if not candidates:
		frappe.throw(_("未识别到模板表头，请使用系统下载的模板填写"))
	return max(candidates, key=lambda item: item[0])


def _read_plan(file_url, profile):
	workbook = _load_workbook(file_url)
	_score, sheet, header_row, headers, _matched = _find_sheet_and_header(workbook, profile)
	aliases = _profile_alias_map(profile)
	mapping = {}
	for index, value in enumerate(headers, start=1):
		key = aliases.get(_normalise_header(value))
		if key and key not in mapping:
			mapping[key] = index
	missing_required = [column["label"] for column in profile["columns"] if column.get("required") and column["key"] not in mapping]
	rows = []
	for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
		raw = {
			f"{index}:{_normalise_text(header) or '未命名列'}": _normalise_text(values[index - 1] if len(values) >= index else "")
			for index, header in enumerate(headers, start=1)
		}
		normalized = {key: _normalise_text(values[column_index - 1] if len(values) >= column_index else "") for key, column_index in mapping.items()}
		if not any(normalized.values()):
			continue
		rows.append({"row_number": row_number, "raw": raw, "normalized": normalized})
	return {
		"sheet_name": sheet.title,
		"header_row": header_row,
		"headers": [_normalise_text(value) for value in headers],
		"mapping": mapping,
		"missing_required": missing_required,
		"rows": rows,
	}


def _employee_by_code(company, employee_code):
	if not employee_code:
		return ""
	filters = {"company": company, "status": "Active"}
	for fieldname in ("custom_employee_code", "employee_number"):
		if frappe.db.has_column("Employee", fieldname):
			value = frappe.db.get_value("Employee", {**filters, fieldname: employee_code}, "name")
			if value:
				return value
	return ""


def _employee_by_name(company, employee_name):
	if not employee_name:
		return ""
	return frappe.db.get_value("Employee", {"company": company, "status": "Active", "employee_name": employee_name}, "name") or ""


def _department_exists(company, department):
	if not department:
		return ""
	return frappe.db.get_value("Department", {"company": company, "name": department}, "name") or ""


def _business_date(data):
	for key in ("attendance_date", "transfer_date", "review_date", "application_date", "occurred_on", "proposal_date", "training_date", "summary_date", "exit_date"):
		value = data.get(key)
		if not value:
			continue
		try:
			return getdate(value)
		except Exception:
			continue
	return None


def _record_key(profile, data, row_number):
	bits = [profile["key"], data.get("external_id") or data.get("employee_code") or data.get("candidate_name") or data.get("proposal_no") or data.get("feedback_no") or "ROW", data.get("attendance_date") or data.get("transfer_date") or data.get("payroll_month") or data.get("training_date") or row_number]
	return "-".join(_normalise_text(value).replace(" ", "")[:40] for value in bits)


def _validate_rows(profile, company, plan):
	result = []
	for item in plan["rows"]:
		data = item["normalized"]
		errors = [_("缺少必填值：{0}").format(column["label"]) for column in profile["columns"] if column.get("required") and not data.get(column["key"])]
		employee = _employee_by_code(company, data.get("employee_code")) or _employee_by_name(company, data.get("employee_name"))
		if data.get("employee_code") and not employee:
			errors.append(_("未匹配到当前公司在职员工工号：{0}").format(data["employee_code"]))
		department = _department_exists(company, data.get("department"))
		if data.get("department") and not department:
			errors.append(_("未匹配到当前公司部门：{0}").format(data["department"]))
		result.append({
			**item,
			"employee": employee,
			"department": department,
			"record_key": _record_key(profile, data, item["row_number"]),
			"errors": errors,
		})
	return result


@frappe.whitelist()
def list_form_import_templates(module_name: str = ""):
	profiles = FORM_IMPORT_PROFILES
	if module_name:
		profiles = [profile for profile in profiles if profile["module"] == module_name]
	return [{
		"key": profile["key"], "module": profile["module"], "label": profile["label"], "description": profile["description"],
		"source_sheets": profile["source_sheets"], "processing_target": profile["processing_target"], "entry_mode": profile.get("entry_mode", "staging"),
		"columns": profile["columns"],
	} for profile in profiles]


@frappe.whitelist()
def create_form_import_template_file(template_key: str):
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill

	profile = _profile_or_throw(template_key)
	workbook = Workbook()
	instructions = workbook.active
	instructions.title = "填写说明"
	instructions.append(["模板名称", profile["label"]])
	instructions.append(["所属模块", profile["module"]])
	instructions.append(["后续处理", profile["processing_target"]])
	instructions.append(["填写规则", "仅填写“数据”页；工号、部门必须与当前公司员工/组织资料完全一致；日期建议使用 YYYY-MM-DD。"])
	instructions.append(["安全规则", "上传先做字段和数据校验。签核型表单仅进入待处理数据池，不会自动变更员工状态或生成薪资。"])
	instructions.column_dimensions["A"].width = 16
	instructions.column_dimensions["B"].width = 100
	data_sheet = workbook.create_sheet("数据")
	headers = [column["label"] for column in profile["columns"]]
	data_sheet.append(headers)
	data_sheet.append([""] * len(headers))
	for cell in data_sheet[1]:
		cell.font = Font(bold=True, color="FFFFFF")
		cell.fill = PatternFill("solid", fgColor="1677FF")
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		data_sheet.column_dimensions[cell.column_letter].width = max(14, min(28, len(str(cell.value)) * 2 + 4))
	data_sheet.freeze_panes = "A2"
	data_sheet.auto_filter.ref = f"A1:{data_sheet.cell(row=1, column=len(headers)).column_letter}2"
	output = BytesIO()
	workbook.save(output)
	filename = f"{profile['label']}导入模板.xlsx"
	file_doc = frappe.get_doc({"doctype": "File", "file_name": filename, "content": output.getvalue(), "is_private": 0}).insert(ignore_permissions=True)
	return {"file_url": file_doc.file_url, "file_name": filename}


@frappe.whitelist()
def preview_form_import(file_url: str, template_key: str, company: str):
	profile = _profile_or_throw(template_key)
	if profile.get("entry_mode") == "employee_roster":
		return {"entry_mode": "employee_roster", "redirect_route": "employee-roster-import", "message": _("员工花名册需要使用智能花名册导入，以便安全写入 Employee 主档。")}
	if not company or not frappe.db.exists("Company", company):
		frappe.throw(_("请选择有效公司"))
	plan = _read_plan(file_url, profile)
	rows = _validate_rows(profile, company, plan) if not plan["missing_required"] else []
	failed = sum(1 for row in rows if row["errors"])
	return {
		"entry_mode": "staging", "template": {"key": profile["key"], "label": profile["label"], "module": profile["module"], "processing_target": profile["processing_target"]},
		"sheet_name": plan["sheet_name"], "header_row": plan["header_row"], "headers": plan["headers"], "mapping": plan["mapping"],
		"missing_required": plan["missing_required"], "total_rows": len(plan["rows"]), "valid_rows": len(rows) - failed, "failed_rows": failed,
		"preview_rows": [{"row_number": row["row_number"], "record_key": row["record_key"], "employee": row["employee"], "errors": row["errors"], "normalized": row["normalized"]} for row in rows[:100]],
	}


@frappe.whitelist()
def import_form_workbook(file_url: str, template_key: str, company: str, notes: str = ""):
	profile = _profile_or_throw(template_key)
	if profile.get("entry_mode") == "employee_roster":
		frappe.throw(_("员工花名册请使用智能花名册导入。"))
	if not company or not frappe.db.exists("Company", company):
		frappe.throw(_("请选择有效公司"))
	plan = _read_plan(file_url, profile)
	if plan["missing_required"]:
		frappe.throw(_("缺少必填列：{0}").format("、".join(plan["missing_required"])))
	rows = _validate_rows(profile, company, plan)
	failed = sum(1 for row in rows if row["errors"])
	batch = frappe.get_doc({
		"doctype": FORM_IMPORT_BATCH_DOCTYPE, "company": company, "module_name": profile["module"], "template_key": profile["key"], "template_name": profile["label"],
		"source_file": file_url, "status": "部分失败" if failed else "已导入待处理", "total_rows": len(rows), "valid_rows": len(rows) - failed, "failed_rows": failed,
		"mapping_json": json.dumps({"sheet_name": plan["sheet_name"], "header_row": plan["header_row"], "mapping": plan["mapping"]}, ensure_ascii=False),
		"error_summary": _("{0} 行校验失败").format(failed) if failed else "", "imported_by": frappe.session.user, "imported_on": now_datetime(), "notes": notes,
	}).insert(ignore_permissions=True)
	for row in rows:
		data = row["normalized"]
		frappe.get_doc({
			"doctype": FORM_IMPORT_ROW_DOCTYPE, "import_batch": batch.name, "company": company, "module_name": profile["module"], "template_key": profile["key"],
			"row_number": row["row_number"], "record_key": row["record_key"], "employee": row["employee"], "employee_code": data.get("employee_code"),
			"employee_name": data.get("employee_name") or data.get("candidate_name"), "department": row["department"], "business_date": _business_date(data),
		"status": "处理失败" if row["errors"] else "待人事审核", "target_doctype": profile.get("target_doctype", ""),
			"raw_data_json": json.dumps(row["raw"], ensure_ascii=False), "normalized_data_json": json.dumps(data, ensure_ascii=False), "error_message": "；".join(row["errors"]),
		}).insert(ignore_permissions=True)
	return {"batch_name": batch.name, "total_rows": len(rows), "valid_rows": len(rows) - failed, "failed_rows": failed, "status": batch.status}


@frappe.whitelist()
def list_form_import_batches(company: str, module_name: str = "", page_length: int = 50):
	filters = {"company": company}
	if module_name:
		filters["module_name"] = module_name
	return frappe.get_all(FORM_IMPORT_BATCH_DOCTYPE, filters=filters, fields=["name", "module_name", "template_key", "template_name", "status", "total_rows", "valid_rows", "failed_rows", "source_file", "imported_by", "imported_on"], order_by="modified desc", limit_page_length=min(int(page_length or 50), 200))


# ---------------------------------------------------------------------------
# Import review and controlled activation
# ---------------------------------------------------------------------------
# A source workbook is evidence, not an instruction to alter employee, attendance
# or salary data.  The following actions intentionally keep the three decisions
# separate: human review -> draft creation -> formal activation.


def _require_form_import_reviewer():
	frappe.only_for(("System Manager", "HR Manager"))
	frappe.has_permission(FORM_IMPORT_ROW_DOCTYPE, "write", throw=True)


def _default_approval_route(row):
	if row.template_key in PAYROLL_TEMPLATE_KEYS:
		return "薪资两级审批", DEFAULT_FORM_APPROVAL_ROUTES["payroll"]
	if row.template_key in ORGANIZATION_TEMPLATE_KEYS:
		return "组织两级审批", DEFAULT_FORM_APPROVAL_ROUTES["organization"]
	return "人事审核", DEFAULT_FORM_APPROVAL_ROUTES["standard"]


def _approval_route_for_row(row):
	"""Use an enabled company matrix first; otherwise use the safe default route."""
	matrix_name = frappe.db.get_value(
		FORM_APPROVAL_MATRIX_DOCTYPE,
		{"company": row.company, "template_key": row.template_key, "enabled": 1},
		"name",
	)
	matrix_name = matrix_name or frappe.db.get_value(
		FORM_APPROVAL_MATRIX_DOCTYPE,
		{"company": ["is", "not set"], "template_key": row.template_key, "enabled": 1},
		"name",
	)
	if matrix_name:
		matrix = frappe.get_doc(FORM_APPROVAL_MATRIX_DOCTYPE, matrix_name)
		steps = sorted(
			[{"step_no": int(step.step_no or 0), "step_label": step.step_label, "approver_role": step.approver_role} for step in matrix.steps if step.approver_role],
			key=lambda item: item["step_no"],
		)
		if steps:
			return matrix.route_name, steps
	return _default_approval_route(row)


def _current_approval_step(row):
	route_name, steps = _approval_route_for_row(row)
	step_no = int(row.approval_step or 1)
	step = next((item for item in steps if item["step_no"] == step_no), None)
	if not step:
		step = steps[0]
		step_no = step["step_no"]
	return route_name, steps, step_no, step


def _can_approve_role(role):
	roles = set(frappe.get_roles(frappe.session.user))
	return role in roles or "System Manager" in roles


def _approval_history(row):
	try:
		value = json.loads(row.approval_history_json or "[]")
		return value if isinstance(value, list) else []
	except (TypeError, json.JSONDecodeError):
		return []


def _append_approval_history(row, step, decision, note):
	history = _approval_history(row)
	history.append(
		{
			"step_no": step.get("step_no"),
			"step_label": step.get("step_label"),
			"required_role": step.get("approver_role"),
			"decision": decision,
			"approver": frappe.session.user,
			"occurred_on": str(now_datetime()),
			"note": note or "",
		}
	)
	row.approval_history_json = json.dumps(history, ensure_ascii=False)


@frappe.whitelist()
def ensure_default_form_approval_matrices(company: str):
	"""Create editable company approval settings from the safe built-in defaults."""
	_require_form_import_reviewer()
	if not company or not frappe.db.exists("Company", company):
		frappe.throw(_("请选择有效公司"))
	created, existing = [], []
	for profile in FORM_IMPORT_PROFILES:
		if profile.get("entry_mode") == "employee_roster":
			continue
		if frappe.db.exists(FORM_APPROVAL_MATRIX_DOCTYPE, {"company": company, "template_key": profile["key"]}):
			existing.append(profile["key"])
			continue
		stub = frappe._dict({"company": company, "template_key": profile["key"]})
		route_name, steps = _default_approval_route(stub)
		doc = frappe.get_doc(
			{
				"doctype": FORM_APPROVAL_MATRIX_DOCTYPE,
				"company": company,
				"template_key": profile["key"],
				"route_name": route_name,
				"enabled": 1,
				"steps": steps,
				"remarks": _("系统初始化默认路线；可按公司实际审批职责调整。"),
			}
		).insert(ignore_permissions=True)
		created.append(doc.name)
	frappe.db.commit()
	return {"company": company, "created": len(created), "existing": len(existing), "matrices": created}


def _get_form_import_row(row_name):
	if not row_name or not frappe.db.exists(FORM_IMPORT_ROW_DOCTYPE, row_name):
		frappe.throw(_("导入明细不存在"))
	return frappe.get_doc(FORM_IMPORT_ROW_DOCTYPE, row_name)


def _row_data(row):
	try:
		data = json.loads(row.normalized_data_json or "{}")
	except (TypeError, json.JSONDecodeError):
		frappe.throw(_("导入明细的规范数据无法读取，请重新导入该行。"))
	if not isinstance(data, dict):
		frappe.throw(_("导入明细的规范数据格式不正确。"))
	return data


def _row_source_file(row):
	return frappe.db.get_value(FORM_IMPORT_BATCH_DOCTYPE, row.import_batch, "source_file") or ""


def _month_from(value):
	text = _normalise_text(value)
	match = re.match(r"^(\d{4}-\d{2})", text)
	if match:
		return match.group(1)
	return ""


def _date_from(value, label):
	if not value:
		frappe.throw(_("缺少{0}，不能生成正式草稿。 ").format(label))
	try:
		return getdate(value)
	except Exception:
		frappe.throw(_("{0}格式不正确：{1}").format(label, value))


def _employee_context_for_row(row):
	if not row.employee:
		frappe.throw(_("该行尚未匹配在职员工，不能生成正式单据。"))
	context = frappe.db.get_value(
		"Employee",
		row.employee,
		["employee_name", "company", "department", "designation", "date_of_joining"],
		as_dict=True,
	) or frappe._dict()
	if context.company != row.company:
		frappe.throw(_("员工 {0} 不属于当前公司 {1}。 ").format(row.employee, row.company))
	return context


def _require_link(doctype, value, label):
	if not value or not frappe.db.exists(doctype, value):
		frappe.throw(_("{0}不存在或尚未维护：{1}").format(label, value or _("未填写")))
	return value


def _active_attendance_lock_version(company, payroll_month):
	if not payroll_month:
		frappe.throw(_("薪资来源必须填写月份。"))
	lock = frappe.db.get_value(
		"HRMS Attendance Month Lock",
		{"company": company, "attendance_month": payroll_month, "status": "已锁定"},
		["name", "active_version"],
		as_dict=True,
	)
	if not lock:
		frappe.throw(_("公司 {0} 的 {1} 考勤尚未锁定，不能生成会影响薪资的正式记录。").format(company, payroll_month))
	return str(lock.active_version)


def _attendance_batch_for_row(row, attendance_month):
	"""Attach workflow-generated attendance evidence to an auditable batch."""
	checksum = hashlib.sha256(f"form-import:{row.import_batch}:{attendance_month}".encode()).hexdigest()
	name = frappe.db.get_value(
		"HRMS Attendance Import Batch",
		{"company": row.company, "attendance_month": attendance_month, "source_checksum": checksum},
		"name",
	)
	if name:
		return frappe.get_doc("HRMS Attendance Import Batch", name)
	batch = frappe.get_doc(
		{
			"doctype": "HRMS Attendance Import Batch",
			"company": row.company,
			"attendance_month": attendance_month,
			"source_file": _row_source_file(row),
			"source_type": "表单导入审核生效",
			"source_checksum": checksum,
			"status": "已导入",
			"imported_by": frappe.session.user,
			"imported_on": now_datetime(),
			"notes": _("来自表单导入批次 {0}，经人事审核后生成。").format(row.import_batch),
		}
	).insert(ignore_permissions=True)
	return batch


def _insert_target(row, data, payroll_month="", attendance_lock_version="", appraisal_cycle=""):
	"""Create one editable formal draft for a reviewed row.

	The returned document is never submitted here.  Calling ``activate`` is the
	only operation allowed to submit an HR transaction or expose a payroll source.
	"""
	target = row.target_doctype
	# Older staged rows were created before reward/exit forms were routed to the
	# payroll source ledger.  Preserve their source data but upgrade their target
	# at generation time instead of requiring a re-import.
	if row.template_key in ("reward_punishment", "exit_payroll_settlement"):
		target = "HRMS Payroll Welfare Source Record"
	# Upgrade historical staging rows to their dedicated, formal business record.
	if row.template_key in BUSINESS_PROCESS_TEMPLATE_CONFIG:
		target = BUSINESS_PROCESS_RECORD_DOCTYPE
	context = _employee_context_for_row(row) if row.employee else frappe._dict()

	if target == BUSINESS_PROCESS_RECORD_DOCTYPE:
		config = BUSINESS_PROCESS_TEMPLATE_CONFIG.get(row.template_key)
		if not config:
			frappe.throw(_("该表单尚未配置业务正式记录类型。"))
		effective_date = None
		for key in config.get("date_keys", ()):
			if data.get(key):
				try:
					effective_date = getdate(data.get(key))
				except Exception:
					pass
				break
		summary_fields = {
			"org_structure": ["parent_department", "designation", "headcount", "department_head_name", "remarks"],
			"contract_intent": ["contract_type", "contract_end_date", "employee_intent", "survey_date", "remarks"],
			"certificate_management": ["certificate_no", "unit_type", "person_type", "first_issue_date", "validity_period", "review_frequency", "next_review_due", "review_status", "remarks"],
			"attendance_department_summary": ["regular_headcount", "probation_headcount", "total_headcount", "attendance_count", "leave_count", "leave_employee_names", "remarks"],
			"proposal_improvement": ["proposal_no", "background", "improvement", "expected_benefit", "status", "remarks"],
			"system_feedback": ["category", "page_or_feature", "description", "status", "followup_date", "remarks"],
		}.get(row.template_key, [])
		summary = "；".join(f"{key}：{data.get(key)}" for key in summary_fields if data.get(key) not in (None, ""))
		return frappe.get_doc(
			{
				"doctype": target,
				"company": row.company,
				"record_type": config["record_type"],
				"source_import_row": row.name,
				"source_import_batch": row.import_batch,
				"module_name": row.module_name,
				"employee": row.employee,
				"department": row.department or context.get("department"),
				"title": config["title"](data),
				"effective_date": effective_date,
				"status": "草稿",
				"owner_user": frappe.session.user,
				"follow_up_due_date": effective_date if row.template_key in ("contract_intent", "certificate_management", "system_feedback") else None,
				"summary": summary,
				"normalized_data_json": json.dumps(data, ensure_ascii=False),
				"approval_history_json": row.approval_history_json,
				"remarks": data.get("remarks"),
			}
		).insert(ignore_permissions=True)

	if target == "Employee Transfer":
		to_department = _require_link("Department", data.get("to_department"), _("调入部门"))
		to_designation = _require_link("Designation", data.get("to_designation"), _("调动后职务"))
		current_department = data.get("from_department") or context.department
		transfer_details = [{"property": _("Department"), "fieldname": "department", "current": current_department, "new": to_department}]
		if to_designation:
			transfer_details.append({"property": _("Designation"), "fieldname": "designation", "current": data.get("from_designation") or context.designation, "new": to_designation})
		return frappe.get_doc(
			{
				"doctype": target,
				"employee": row.employee,
				"employee_name": context.employee_name,
				"company": row.company,
				"department": context.department,
				"transfer_date": _date_from(data.get("transfer_date"), _("调动日期")),
				"transfer_details": transfer_details,
			}
		).insert(ignore_permissions=True)

	if target == "Employee Promotion":
		promotion_details = []
		if data.get("designation") and frappe.db.exists("Designation", data.get("designation")):
			promotion_details.append({"property": _("Designation"), "fieldname": "designation", "current": context.designation, "new": data.get("designation")})
		return frappe.get_doc(
			{
				"doctype": target,
				"employee": row.employee,
				"employee_name": context.employee_name,
				"company": row.company,
				"department": context.department,
				"promotion_date": _date_from(data.get("review_date"), _("认定日期")),
				"promotion_details": promotion_details,
			}
		).insert(ignore_permissions=True)

	if target == "Employee Separation":
		return frappe.get_doc(
			{
				"doctype": target,
				"employee": row.employee,
				"company": row.company,
				"boarding_begins_on": _date_from(data.get("last_working_date"), _("最后工作日")),
				"resignation_letter_date": _date_from(data.get("application_date"), _("申请日期")),
				"exit_interview": _("离职原因：{0}<br>交接状态：{1}<br>备注：{2}").format(data.get("reason") or "", data.get("handover_status") or "", data.get("remarks") or ""),
			}
		).insert(ignore_permissions=True)

	if target == "Job Applicant":
		candidate_name = data.get("candidate_name") or row.employee_name
		if not candidate_name:
			frappe.throw(_("候选人姓名为空，不能生成候选人档案。"))
		email = data.get("email") or f"import-{hashlib.sha1(row.name.encode()).hexdigest()[:12]}@pending.invalid"
		decision = data.get("decision") or ""
		status = "Accepted" if "录用" in decision else "Open"
		return frappe.get_doc(
			{
				"doctype": target,
				"applicant_name": candidate_name,
				"email_id": email,
				"phone_number": data.get("phone"),
				"designation": data.get("applied_designation") if frappe.db.exists("Designation", data.get("applied_designation")) else "",
				"status": status,
				"notes": _("来源：表单导入 {0}；面试时间：{1}；面试官：{2}；决定：{3}；请补充真实邮箱。").format(row.name, data.get("interview_date") or "", data.get("interviewer") or "", decision),
			}
		).insert(ignore_permissions=True)

	if target == "Training Event":
		training_date = _date_from(data.get("training_date"), _("实际上课时间"))
		start_time = f"{training_date} 09:00:00"
		hours = max(flt(data.get("hours")), 1)
		end_hour = min(23, 9 + int(hours))
		end_time = f"{training_date} {end_hour:02d}:00:00"
		return frappe.get_doc(
			{
				"doctype": target,
				"event_name": data.get("training_content"),
				"event_status": "Scheduled",
				"type": "Theory",
				"company": row.company,
				"trainer_name": data.get("trainer"),
				"course": data.get("course_type"),
				"location": data.get("location"),
				"start_time": start_time,
				"end_time": end_time,
				"introduction": data.get("remarks") or data.get("training_content"),
				"employees": [{"employee": row.employee, "department": context.department, "status": "Invited"}] if row.employee else [],
			}
		).insert(ignore_permissions=True)

	if target == "Employee Skill Map":
		if not row.employee:
			frappe.throw(_("证书记录必须先匹配员工。"))
		doc = frappe.get_doc({"doctype": target, "employee": row.employee}).insert(ignore_permissions=True)
		return doc

	if target == "Appraisal":
		cycle = appraisal_cycle or frappe.db.get_value("Appraisal Cycle", {"status": "In Progress"}, "name", order_by="start_date desc")
		if not cycle:
			frappe.throw(_("请先维护进行中的绩效周期，或在生成草稿时选择绩效周期。"))
		return frappe.get_doc(
			{
				"doctype": target,
				"naming_series": "HR-APR-.YYYY.-",
				"employee": row.employee,
				"employee_name": context.employee_name,
				"company": row.company,
				"department": context.department,
				"designation": context.designation,
				"appraisal_cycle": cycle,
				"reflections": data.get("achievements"),
				"remarks": _("待改善：{0}\n下一期计划：{1}\n来源：{2}").format(data.get("improvements") or "", data.get("next_plan") or "", row.name),
			}
		).insert(ignore_permissions=True)

	if target == "HRMS Employee Salary Change":
		from hrms.api import payroll_input

		month = payroll_month or _month_from(data.get("effective_month"))
		effective_date = f"{month}-01" if month else data.get("effective_date")
		name = payroll_input.create_employee_salary_change(
			company=row.company,
			employee=row.employee,
			employee_code=data.get("employee_code") or row.employee_code,
			employee_name=data.get("employee_name") or row.employee_name,
			department=context.department,
			designation=context.designation,
			effective_date=effective_date,
			change_reason=data.get("change_reason"),
			base_salary=flt(data.get("base_salary")),
			function_allowance=flt(data.get("functional_allowance")) + flt(data.get("position_allowance")),
			certificate_allowance=flt(data.get("certificate_allowance")),
			multi_skill_allowance=flt(data.get("multi_skill_allowance")),
			full_salary=flt(data.get("gross_salary")),
			prepared_by=frappe.session.user,
			status="草稿",
			source_file=_row_source_file(row),
			remarks=data.get("remarks"),
		)
		return frappe.get_doc(target, name)

	if target == "HRMS Payroll Welfare Source Record":
		from hrms.api import payroll_input

		month = payroll_month or _month_from(data.get("payroll_month") or data.get("occurred_on"))
		lock_version = attendance_lock_version or _active_attendance_lock_version(row.company, month)
		profile_to_source = {
			"skill_certificate_allowance": ("证书多能工津贴", "证书及多能工津贴", flt(data.get("multi_skill_allowance")) + flt(data.get("certificate_allowance"))),
			"full_attendance_bonus": ("其他奖金", "全勤奖", flt(data.get("amount"))),
			"housing_allowance": ("租房补贴", "住房补贴", flt(data.get("amount"))),
			"education_allowance": ("学历补贴", "学历补贴", flt(data.get("amount"))),
			"dormitory_fee": ("宿舍住宿费", "宿舍扣款", flt(data.get("deduction_amount"))),
			"social_insurance": ("社保个人", "社保个人", flt(data.get("employee_amount"))),
			"service_award": ("继续服务奖", "继续服务奖", flt(data.get("amount"))),
			"reward_punishment": (("其他扣款" if "惩" in (data.get("reward_punishment_type") or "") else "其他奖金"), ("其他扣款" if "惩" in (data.get("reward_punishment_type") or "") else "其他奖金"), flt(data.get("amount"))),
			"exit_payroll_settlement": ("其他扣款", "其他扣款", flt(data.get("deduction_amount"))),
		}
		if row.template_key not in profile_to_source:
			frappe.throw(_("该薪资表单尚未配置正式薪资来源映射。"))
		source_type, variable_type, amount = profile_to_source[row.template_key]
		name = payroll_input.upsert_payroll_welfare_source_record(
			company=row.company,
			payroll_month=month,
			attendance_lock_version=lock_version,
			source_type=source_type,
			variable_type=variable_type,
			direction="应扣" if variable_type in ("宿舍扣款", "社保个人", "其他扣款") else "应发",
			employee=row.employee,
			employee_code=data.get("employee_code") or row.employee_code,
			employee_name=data.get("employee_name") or row.employee_name,
			department=context.department,
			amount=amount,
			eligibility_status="符合",
			confirmation_status="草稿",
			source_reference=_('表单导入审核行 {0}').format(row.name),
			source_file=_row_source_file(row),
			remarks=data.get("remarks"),
		)
		return frappe.get_doc(target, name)

	if target == "HRMS Monthly Attendance Summary":
		from hrms.api import attendance_import

		month = _month_from(data.get("attendance_month"))
		if not month:
			frappe.throw(_("月度考勤终稿必须填写考勤月份。"))
		lock = attendance_import._prepare_month_lock_for_generation(row.company, month)
		lock_version = str(lock.active_version)
		values = {
			"company": row.company,
			"attendance_month": month,
			"attendance_lock_version": lock_version,
			"lock_status": "草稿",
			"source_batch_ids": _attendance_batch_for_row(row, month).name,
			"source_checksum": hashlib.sha256(f"form-import-final:{row.import_batch}:{month}".encode()).hexdigest(),
			"employee": row.employee,
			"employee_code": data.get("employee_code") or row.employee_code,
			"employee_name": data.get("employee_name") or row.employee_name,
			"department": row.department or context.get("department"),
			"date_of_joining": context.get("date_of_joining"),
			"standard_hours": flt(data.get("standard_hours")),
			"actual_attendance_hours": flt(data.get("actual_hours")),
			"adjusted_working_hours": flt(data.get("actual_hours")),
			"overtime_1_5_hours": flt(data.get("workday_overtime_hours")),
			"overtime_2_hours": flt(data.get("restday_overtime_hours")),
			"overtime_3_hours": flt(data.get("holiday_overtime_hours")),
			"absent_hours": flt(data.get("absence_hours")),
			"green_apples": flt(data.get("green_apple_amount")),
			"red_apples": flt(data.get("red_apple_amount")),
			"apple_reward_amount": flt(data.get("green_apple_amount")) - flt(data.get("red_apple_amount")),
			"full_attendance_deduction": flt(data.get("full_attendance_bonus")),
			"status": "草稿",
		}
		existing = frappe.db.get_value(
			target,
			{
				"company": row.company,
				"attendance_month": month,
				"attendance_lock_version": lock_version,
				"employee_code": values["employee_code"],
			},
			"name",
		)
		if existing:
			doc = frappe.get_doc(target, existing)
			doc.update(values)
			doc.save(ignore_permissions=True)
			return doc
		return frappe.get_doc({"doctype": target, **values}).insert(ignore_permissions=True)

	if target in ("HRMS Attendance Day Check", "HRMS Attendance Leave Evidence", "HRMS Apple Reward Record", "HRMS Attendance Exception"):
		from hrms.api import attendance_import

		attendance_date = _date_from(data.get("attendance_date") or data.get("occurred_on") or data.get("start_time"), _("考勤日期"))
		batch = _attendance_batch_for_row(row, attendance_date.strftime("%Y-%m"))
		if target == "HRMS Attendance Day Check":
			name = attendance_import._insert_day_check(
				batch.name,
				{
					"工号": data.get("employee_code") or row.employee_code,
					"姓名": data.get("employee_name") or row.employee_name,
					"日期": str(attendance_date),
					"实际部门": context.department,
					"班次": data.get("shift_name"),
					"上班时间": data.get("actual_in_time"),
					"下班时间": data.get("actual_out_time"),
					"标准工时": data.get("standard_hours"),
					"实际出勤（小时）": data.get("actual_hours"),
					"工作日加班（小时）": data.get("workday_overtime_hours"),
					"休息日加班（小时）": data.get("restday_overtime_hours"),
					"节假日加班（小时）": data.get("holiday_overtime_hours"),
					"关联审批单": data.get("approval_reference"),
				},
				row.company,
				"人工调整",
				_("表单导入审核行 {0}").format(row.name),
				1,
			)
			if not name:
				frappe.throw(_("考勤明细未能匹配员工或公司，不能生成草稿。"))
			batch.daily_sheet_rows = int(batch.daily_sheet_rows or 0) + 1
			batch.save(ignore_permissions=True)
			return frappe.get_doc(target, name)
		if target == "HRMS Attendance Leave Evidence":
			name = attendance_import._insert_leave_evidence(
				batch.name,
				{"创建人": data.get("employee_name") or row.employee_name, "请假类型（实际）": data.get("leave_type"), "开始时间": data.get("start_time"), "结束时间": data.get("end_time"), "时长": data.get("duration"), "请假事由": data.get("reason"), "审批编号": data.get("approval_no"), "审批结果": data.get("approval_result"), "审批状态": data.get("approval_status")},
			)
			batch.leave_sheet_rows = int(batch.leave_sheet_rows or 0) + 1
			batch.save(ignore_permissions=True)
			return frappe.get_doc(target, name)
		if target == "HRMS Apple Reward Record":
			apple_type = data.get("apple_type") or ""
			quantity = flt(data.get("quantity"))
			name = attendance_import._insert_apple_record(
				batch.name,
				{"奖/惩日期": str(attendance_date), "受奖/惩人": data.get("employee_name") or row.employee_name, "部门": context.department, "绿苹果": quantity if "绿" in apple_type else 0, "红苹果": quantity if "红" in apple_type else 0, "奖/惩项目": data.get("reason"), "审批编号": data.get("approval_no"), "审批结果": data.get("approval_result"), "审批状态": data.get("approval_status")},
			)
			batch.apple_sheet_rows = int(batch.apple_sheet_rows or 0) + 1
			batch.save(ignore_permissions=True)
			return frappe.get_doc(target, name)
		day_check = frappe.db.get_value("HRMS Attendance Day Check", {"company": row.company, "employee": row.employee, "attendance_date": attendance_date}, "name")
		exception_type = data.get("exception_type")
		if exception_type not in ("忘打卡", "迟到", "早退", "旷工", "未申请加班"):
			exception_type = "忘打卡"
		return frappe.get_doc(
			{
				"doctype": target, "import_batch": batch.name, "day_check": day_check, "attendance_date": attendance_date,
				"employee": row.employee, "employee_code": data.get("employee_code") or row.employee_code,
				"employee_name": data.get("employee_name") or row.employee_name, "department": context.department,
				"exception_type": exception_type, "expected_shift": data.get("shift_name"), "actual_in_time": data.get("actual_in_time"),
				"actual_out_time": data.get("actual_out_time"), "handling_method": data.get("handling"), "confirmation_status": "待确认", "remarks": data.get("remarks"),
			}
		).insert(ignore_permissions=True)

	frappe.throw(_("模板 {0} 尚未配置可自动生成的正式单据。审核意见已保留，但请在对应业务模块补充处理。 ").format(row.template_key))


def _refresh_import_batch_status(batch_name):
	rows = frappe.get_all(FORM_IMPORT_ROW_DOCTYPE, filters={"import_batch": batch_name}, fields=["status"])
	if not rows:
		return
	statuses = {row.status for row in rows}
	status = "审核中"
	if statuses <= {"已提交生效", "已忽略", "处理失败", "已驳回"}:
		status = "已处理"
	elif "处理失败" in statuses:
		status = "部分失败"
	frappe.db.set_value(FORM_IMPORT_BATCH_DOCTYPE, batch_name, "status", status)


@frappe.whitelist()
def list_form_import_review_rows(company: str, module_name: str = "", review_status: str = "", page_length: int = 100):
	_require_form_import_reviewer()
	if not company or not frappe.db.exists("Company", company):
		frappe.throw(_("请选择有效公司"))
	filters = {"company": company}
	if module_name:
		filters["module_name"] = module_name
	if review_status:
		filters["review_status"] = review_status
	return frappe.get_all(
		FORM_IMPORT_ROW_DOCTYPE,
		filters=filters,
		fields=["name", "import_batch", "module_name", "template_key", "record_key", "employee", "employee_name", "department", "business_date", "status", "review_status", "approval_route", "approval_step", "approval_step_label", "target_doctype", "target_name", "reviewed_by", "reviewed_on", "review_note", "processing_error"],
		order_by="modified desc",
		limit_page_length=min(int(page_length or 100), 500),
	)


@frappe.whitelist()
def review_form_import_row(row_name: str, decision: str, review_note: str = ""):
	"""Approve or reject an imported row before any formal document is created."""
	_require_form_import_reviewer()
	row = _get_form_import_row(row_name)
	decision = (decision or "").strip()
	if decision not in ("批准", "驳回"):
		frappe.throw(_("审核决定只能是“批准”或“驳回”。"))
	if row.status in ("处理失败", "已忽略", "已提交生效"):
		frappe.throw(_("当前状态不允许审核。"))
	if decision == "驳回" and not (review_note or "").strip():
		frappe.throw(_("驳回时必须填写审核意见。"))
	route_name, steps, step_no, step = _current_approval_step(row)
	if not _can_approve_role(step["approver_role"]):
		frappe.throw(_("当前审批节点“{0}”要求角色：{1}。 ").format(step["step_label"], step["approver_role"]))
	row.approval_route = route_name
	row.approval_step = step_no
	row.approval_step_label = step["step_label"]
	_append_approval_history(row, step, decision, review_note)
	if decision == "批准":
		next_step = next((item for item in steps if item["step_no"] > step_no), None)
		if next_step:
			row.review_status = "审批中"
			row.status = "待人事审核"
			row.approval_step = next_step["step_no"]
			row.approval_step_label = next_step["step_label"]
			next_action = _("等待 {0}（{1}）审批").format(next_step["step_label"], next_step["approver_role"])
		else:
			row.review_status = "已批准"
			row.status = "已审核"
			next_action = "生成正式草稿"
	else:
		row.review_status = "已驳回"
		row.status = "已驳回"
		next_action = "结束"
	row.reviewed_by = frappe.session.user
	row.reviewed_on = now_datetime()
	row.review_note = review_note
	row.processing_error = ""
	row.save(ignore_permissions=True)
	_refresh_import_batch_status(row.import_batch)
	frappe.db.commit()
	return {"name": row.name, "status": row.status, "review_status": row.review_status, "approval_route": row.approval_route, "approval_step": row.approval_step, "approval_step_label": row.approval_step_label, "next_action": next_action}


@frappe.whitelist()
def generate_form_import_target(row_name: str, payroll_month: str = "", attendance_lock_version: str = "", appraisal_cycle: str = ""):
	"""Generate an editable target draft from an approved row, without activation."""
	_require_form_import_reviewer()
	row = _get_form_import_row(row_name)
	if row.review_status != "已批准":
		frappe.throw(_("请先完成人事审核并批准该行。"))
	if row.target_name:
		return {"name": row.name, "target_doctype": row.target_doctype, "target_name": row.target_name, "status": row.status, "existing": 1}
	try:
		target = _insert_target(row, _row_data(row), payroll_month, attendance_lock_version, appraisal_cycle)
	except Exception as error:
		row.processing_error = str(error)
		row.save(ignore_permissions=True)
		frappe.db.commit()
		raise
	row.target_doctype = target.doctype
	row.target_name = target.name
	row.status = "已生成草稿"
	row.generated_by = frappe.session.user
	row.generated_on = now_datetime()
	row.processing_error = ""
	row.save(ignore_permissions=True)
	_refresh_import_batch_status(row.import_batch)
	frappe.db.commit()
	return {"name": row.name, "target_doctype": target.doctype, "target_name": target.name, "target_docstatus": target.docstatus, "status": row.status}


def _activate_non_submittable_target(row, target):
	"""Apply the minimum explicit business confirmation for non-submittable doctypes."""
	if target.doctype == BUSINESS_PROCESS_RECORD_DOCTYPE:
		follow_up_types = {"组织变更", "合同续签意愿", "提案改善", "系统反馈"}
		target.status = "待跟进" if target.record_type in follow_up_types else "已生效"
		target.approval_history_json = row.approval_history_json
		target.save(ignore_permissions=True)
		return _("正式业务记录已建立，当前状态为“{0}”。").format(target.status)
	if target.doctype == "HRMS Employee Salary Change":
		target.status = "已批准"
		target.reviewed_by = target.reviewed_by or frappe.session.user
		target.approved_by = target.approved_by or frappe.session.user
		target.save(ignore_permissions=True)
		return _("薪资异动已批准；后续薪资结算会读取该员工在生效月份前的最新版本。")
	if target.doctype == "HRMS Payroll Welfare Source Record":
		from hrms.api import payroll_input

		target.confirmation_status = "已确认"
		target.confirmed_by = frappe.session.user
		target.confirmed_on = now_datetime()
		target.save(ignore_permissions=True)
		sync = payroll_input.sync_welfare_sources_to_payroll_variables(
			target.company, target.payroll_month, target.attendance_lock_version
		)
		return _("福利/扣款来源已确认，已同步 {0} 条薪资变量；请重新生成薪资输入表和薪资结算表。 ").format(sync.get("created", 0))
	if target.doctype == "HRMS Attendance Exception":
		target.confirmation_status = "已确认"
		target.confirmed_by = frappe.session.user
		target.confirmed_on = now_datetime()
		target.save(ignore_permissions=True)
		return _("考勤异常已确认；生成月度考勤终稿时会读取日考勤与异常处理结果。")
	if target.doctype == "HRMS Monthly Attendance Summary":
		target.status = "已确认"
		target.save(ignore_permissions=True)
		return _("月度考勤终稿已确认；仍需在考勤中心锁定月份后才可进入薪资。")
	return _("正式记录已建立。该单据不需要提交，后续请在所属模块完成业务处理。")


@frappe.whitelist()
def activate_form_import_target(row_name: str):
	"""Submit a formal draft or explicitly confirm a non-submittable target."""
	_require_form_import_reviewer()
	row = _get_form_import_row(row_name)
	if row.review_status != "已批准" or not row.target_doctype or not row.target_name:
		frappe.throw(_("请先审核通过并生成正式草稿。"))
	if row.status == "已提交生效":
		return {"name": row.name, "target_doctype": row.target_doctype, "target_name": row.target_name, "status": row.status, "existing": 1}
	if not frappe.db.exists(row.target_doctype, row.target_name):
		frappe.throw(_("正式草稿不存在，不能生效。"))
	target = frappe.get_doc(row.target_doctype, row.target_name)
	if frappe.get_meta(target.doctype).is_submittable:
		frappe.has_permission(target.doctype, "submit", throw=True)
		if target.docstatus == 0:
			target.submit()
		message = _("正式单据已提交生效。")
	else:
		message = _activate_non_submittable_target(row, target)
	row.status = "已提交生效"
	row.activated_by = frappe.session.user
	row.activated_on = now_datetime()
	row.processing_error = ""
	row.save(ignore_permissions=True)
	_refresh_import_batch_status(row.import_batch)
	frappe.db.commit()
	return {"name": row.name, "target_doctype": row.target_doctype, "target_name": row.target_name, "status": row.status, "message": message}
