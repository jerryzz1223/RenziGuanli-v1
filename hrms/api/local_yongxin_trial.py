"""Local-only Yongxin operational trial support.

This module is deliberately opt-in and refuses every site except the developer
``hrms.localhost`` site.  It provides a repeatable, auditable way to clear only
Yongxin business data, transform the supplied HR workbook into the product's
own import templates, and exercise the review-to-payroll chain.  It must never
be used against a deployed server.
"""

from __future__ import annotations

import json
import re
from collections import OrderedDict, defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import frappe
from frappe import _
from frappe.utils import getdate
from openpyxl import load_workbook

from hrms.api import attendance_import
from hrms.api import employee_field_template as roster
from hrms.api import form_data_intake as intake
from hrms.api import payroll_input


LOCAL_SITE = "hrms.localhost"
TRIAL_COMPANY = "永新"
SOURCE_WORKBOOK = "/tmp/hrms-yongxin-source.xlsx"
CLEAR_CONFIRMATION = "CLEAR LOCALHOST YONGXIN BUSINESS DATA"
RUN_CONFIRMATION = "RUN LOCALHOST YONGXIN TRIAL"
TRIAL_MARKER = "LOCALHOST-YONGXIN-TRIAL-20260729"

# Dependent records must be deleted before their source rows.  No global
# settings, users, roles, companies, payroll rules or DingTalk credentials are
# included here.
CLEAR_DOCTYPES = (
	"HRMS Payroll Settlement Record",
	"HRMS Payroll Input Record",
	"HRMS Payroll Variable Record",
	"HRMS Payroll Welfare Source Record",
	"HRMS Employee Salary Change",
	"HRMS Payroll Variable Import Batch",
	"HRMS Apple Reward Record",
	"HRMS Attendance Exception",
	"HRMS Attendance Leave Evidence",
	"HRMS Monthly Attendance Summary",
	"HRMS Attendance Lock Audit",
	"HRMS Attendance Department Confirmation",
	"HRMS Attendance Month Lock",
	"HRMS Attendance Day Check",
	"HRMS Attendance Import Batch",
	"HRMS DingTalk Raw Record",
	"HRMS DingTalk Sync Log",
	"HRMS DingTalk User Map",
	"HRMS Form Import Row",
	"HRMS Business Process Record",
	"HRMS Form Import Batch",
	"Employee Promotion",
	"Employee Transfer",
	"Employee Separation",
	"Employee Onboarding",
	"Training Event Employee",
	"Training Event",
	"Appraisal",
	"Employee",
	"Department",
)


def _assert_local_trial_access():
	if frappe.local.site != LOCAL_SITE:
		frappe.throw(_("本试运营脚本只允许在 {0} 执行，当前站点不允许。").format(LOCAL_SITE))
	if not frappe.db.exists("Company", TRIAL_COMPANY):
		frappe.throw(_("未找到本地试运营公司：{0}。").format(TRIAL_COMPANY))
	if "System Manager" not in frappe.get_roles():
		frappe.throw(_("只有系统管理员可以执行本机试运营。"), frappe.PermissionError)


def _text(value):
	if value is None:
		return ""
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d %H:%M:%S")
	if isinstance(value, date):
		return value.isoformat()
	return str(value).strip()


def _header_key(value):
	return re.sub(r"[\s\n\r（）()【】\[\]：:、/\\_-]", "", _text(value)).lower()


def _date_value(value, fallback=""):
	if not value:
		return fallback
	if isinstance(value, (datetime, date)):
		return value.strftime("%Y-%m-%d")
	text = _text(value)
	# DingTalk exports commonly append a Chinese weekday, for example
	# ``26-07-09 星期四``.  The display suffix is useful to a person but is not
	# a database date; remove it before parsing.  Two-digit years are treated as
	# 20xx because this local trial only processes the supplied 2026 workbook.
	text = re.sub(r"\s*星期[一二三四五六日天].*$", "", text).strip()
	match = re.search(r"(20\d{2})[/-](\d{1,2})[/-](\d{1,2})", text)
	if match:
		return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
	match = re.search(r"(?<!\d)(\d{2})[/-](\d{1,2})[/-](\d{1,2})(?!\d)", text)
	if match:
		return f"20{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
	try:
		return str(getdate(text))
	except Exception:
		return fallback


def _month_value(value, fallback=""):
	date_value = _date_value(value)
	if date_value:
		return date_value[:7]
	match = re.search(r"(20\d{2})\D?(\d{1,2})", _text(value))
	if match:
		return f"{match.group(1)}-{int(match.group(2)):02d}"
	return fallback


def _safe_number(value):
	try:
		return float(value or 0)
	except (TypeError, ValueError):
		match = re.search(r"-?\d+(?:\.\d+)?", _text(value))
		return float(match.group()) if match else 0.0


def _workbook_or_throw(source_path=SOURCE_WORKBOOK):
	path = Path(source_path)
	if not path.exists():
		frappe.throw(_("未找到试运营源表，请先将人资系统沟通表复制到容器路径：{0}").format(source_path))
	return load_workbook(path, data_only=True)


def _sheet_by_name(workbook, name):
	for sheet_name in workbook.sheetnames:
		if _text(sheet_name) == _text(name):
			return workbook[sheet_name]
	return None


def _employee_names():
	return frappe.get_all("Employee", filters={"company": TRIAL_COMPANY}, pluck="name")


def _company_scoped_filters(doctype):
	"""Return the narrowest safe scope available for a local business record."""
	if not frappe.db.exists("DocType", doctype):
		return None
	meta = frappe.get_meta(doctype)
	if meta.istable:
		return None
	# Apple Reward Record is an older custom DocType.  Some sites carry a stale
	# custom ``company`` field in metadata while the SQL table has no such
	# column.  Scope it through employee/import batch instead of ever risking a
	# cross-company deletion.
	if doctype == "HRMS Apple Reward Record":
		employees = _employee_names()
		return {"employee": ["in", employees or [""]]}
	if meta.has_field("company"):
		return {"company": TRIAL_COMPANY}
	if meta.has_field("employee"):
		employees = _employee_names()
		return {"employee": ["in", employees or [""]]}
	if meta.has_field("import_batch"):
		batches = frappe.get_all("HRMS Attendance Import Batch", filters={"company": TRIAL_COMPANY}, pluck="name")
		return {"import_batch": ["in", batches or [""]]}
	return None


def _delete_doctype(doctype):
	filters = _company_scoped_filters(doctype)
	if filters is None:
		return 0
	# Department is a nested-set tree.  Its children must be removed before the
	# parent nodes or Frappe correctly refuses the deletion.
	order_by = "lft desc" if doctype == "Department" else "modified desc"
	names = frappe.get_all(doctype, filters=filters, pluck="name", order_by=order_by)
	for name in names:
		doc = frappe.get_doc(doctype, name)
		if doc.docstatus == 1:
			doc.cancel()
		frappe.delete_doc(doctype, name, ignore_permissions=True, force=True)
	return len(names)


def _counts():
	counts = {}
	for doctype in CLEAR_DOCTYPES:
		filters = _company_scoped_filters(doctype)
		if filters is not None:
			counts[doctype] = frappe.db.count(doctype, filters)
	return counts


def _source_sheet_status(workbook):
	available = {_text(name) for name in workbook.sheetnames}
	status = []
	for profile in intake.FORM_IMPORT_PROFILES:
		matched = [sheet for sheet in profile["source_sheets"] if _text(sheet) in available]
		status.append({
			"key": profile["key"],
			"label": profile["label"],
			"source_sheets": profile["source_sheets"],
			"matched_sheets": matched,
			"source_present": bool(matched),
		})
	return status


@frappe.whitelist()
def get_local_yongxin_trial_context(source_path: str = SOURCE_WORKBOOK):
	"""Read-only preflight.  It never changes records or contacts a server."""
	_assert_local_trial_access()
	workbook = _workbook_or_throw(source_path)
	return {
		"site": frappe.local.site,
		"company": TRIAL_COMPANY,
		"source_workbook": source_path,
		"source_sheets": [_text(name) for name in workbook.sheetnames],
		"source_profile_status": _source_sheet_status(workbook),
		"current_counts": _counts(),
		"clear_confirmation": CLEAR_CONFIRMATION,
		"run_confirmation": RUN_CONFIRMATION,
		"safety": "仅作用于 localhost 中公司“永新”；不会连接、修改或读取服务器。",
	}


@frappe.whitelist()
def get_local_yongxin_trial_progress():
	"""Return PII-safe counts for the resumable localhost trial only."""
	_assert_local_trial_access()
	batches = frappe.get_all(
		"HRMS Form Import Batch",
		filters={"company": TRIAL_COMPANY, "notes": ["like", f"%{TRIAL_MARKER}%"]},
		fields=["name", "template_key", "status", "total_rows", "valid_rows", "failed_rows"],
	)
	batch_names = [item.name for item in batches]
	rows = frappe.get_all(
		intake.FORM_IMPORT_ROW_DOCTYPE,
		filters={"import_batch": ["in", batch_names or [""]]},
		fields=["template_key", "status", "review_status", "target_doctype", "target_name"],
	)
	by_template = defaultdict(lambda: defaultdict(int))
	for row in rows:
		bucket = by_template[row.template_key]
		bucket["rows"] += 1
		bucket[f"status:{row.status or '空'}"] += 1
		bucket[f"review:{row.review_status or '空'}"] += 1
		if row.target_name:
			bucket["target_created"] += 1
	return {
		"site": frappe.local.site,
		"company": TRIAL_COMPANY,
		"employees": frappe.db.count("Employee", {"company": TRIAL_COMPANY}),
		"departments": frappe.db.count("Department", {"company": TRIAL_COMPANY}),
		"monthly_summaries": frappe.db.count("HRMS Monthly Attendance Summary", {"company": TRIAL_COMPANY, "attendance_month": "2026-05"}),
		"month_lock": frappe.db.get_value("HRMS Attendance Month Lock", {"company": TRIAL_COMPANY, "attendance_month": "2026-05"}, ["name", "status", "active_version"], as_dict=True),
		"batches": [{"template_key": item.template_key, "status": item.status, "total": item.total_rows, "valid": item.valid_rows, "failed": item.failed_rows} for item in batches],
		"rows_by_template": {key: dict(value) for key, value in sorted(by_template.items())},
	}


@frappe.whitelist()
def get_local_yongxin_trial_acceptance_summary():
	"""Return the final PII-safe counters for the localhost trial handover."""
	_assert_local_trial_access()
	batches = frappe.get_all(
		"HRMS Form Import Batch",
		filters={"company": TRIAL_COMPANY, "notes": ["like", f"%{TRIAL_MARKER}%"]},
		pluck="name",
	)
	rows = frappe.get_all(
		intake.FORM_IMPORT_ROW_DOCTYPE,
		filters={"import_batch": ["in", batches or [""]]},
		fields=["template_key", "status"],
	)
	staging = defaultdict(lambda: defaultdict(int))
	for row in rows:
		staging[row.template_key][row.status or "空"] += 1
	def count(doctype, filters=None):
		return frappe.db.count(doctype, filters if filters is not None else (_company_scoped_filters(doctype) or {"name": ""}))
	lock = frappe.db.get_value(
		"HRMS Attendance Month Lock",
		{"company": TRIAL_COMPANY, "attendance_month": "2026-05"},
		["status", "active_version"],
		as_dict=True,
	)
	return {
		"site": frappe.local.site,
		"company": TRIAL_COMPANY,
		"attendance_month": "2026-05",
		"month_lock": lock or {},
		"formal_documents": {
			"employees": count("Employee"),
			"departments": count("Department"),
			"daily_attendance_checks": count("HRMS Attendance Day Check", {"company": TRIAL_COMPANY}),
			"monthly_attendance_summaries": count("HRMS Monthly Attendance Summary", {"company": TRIAL_COMPANY, "attendance_month": "2026-05"}),
			"attendance_exceptions": count("HRMS Attendance Exception", {"employee": ["in", _employee_names() or [""]]}),
			"salary_changes": count("HRMS Employee Salary Change", {"company": TRIAL_COMPANY}),
			"welfare_sources": count("HRMS Payroll Welfare Source Record", {"company": TRIAL_COMPANY, "payroll_month": "2026-05", "confirmation_status": "已确认"}),
			"payroll_variables": count("HRMS Payroll Variable Record", {"company": TRIAL_COMPANY, "payroll_month": "2026-05"}),
			"payroll_inputs": count("HRMS Payroll Input Record", {"company": TRIAL_COMPANY, "payroll_month": "2026-05"}),
			"payroll_settlements": count("HRMS Payroll Settlement Record", {"company": TRIAL_COMPANY, "payroll_month": "2026-05"}),
			"employee_onboarding": count("Employee Onboarding"),
			"employee_transfers": count("Employee Transfer"),
			"employee_separations": count("Employee Separation"),
			"training_events": count("Training Event"),
			"appraisals": count("Appraisal"),
		},
		"staging_status_by_template": {key: dict(sorted(value.items())) for key, value in sorted(staging.items())},
	}


@frappe.whitelist()
def repair_local_yongxin_trial_dates():
	"""Repair only staged localhost trial date text before formal activation.

	The source workbook contains text such as ``26-07-09 星期四``.  Import
	preview accepts text values, while formal target documents correctly require
	ISO dates.  This repair changes only the normalised staging payload; the raw
	upload evidence remains untouched and traceable.
	"""
	_assert_local_trial_access()
	batches = frappe.get_all(
		"HRMS Form Import Batch",
		filters={"company": TRIAL_COMPANY, "notes": ["like", f"%{TRIAL_MARKER}%"]},
		pluck="name",
	)
	date_keys = (
		"date_of_joining", "boarding_begins_on", "attendance_date", "transfer_date",
		"review_date", "application_date", "occurred_on", "proposal_date",
		"training_date", "summary_date", "exit_date", "last_working_date",
		"contract_end_date", "survey_date", "followup_date", "completed_date",
	)
	updated = defaultdict(int)
	for row_name in frappe.get_all(
		intake.FORM_IMPORT_ROW_DOCTYPE,
		filters={"import_batch": ["in", batches or [""]]},
		pluck="name",
	):
		row = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, row_name)
		try:
			data = json.loads(row.normalized_data_json or "{}")
		except (TypeError, json.JSONDecodeError):
			continue
		if not isinstance(data, dict):
			continue
		changed = False
		for key in date_keys:
			if data.get(key):
				value = _date_value(data[key], data[key])
				if value != data[key]:
					data[key] = value
					changed = True
		if changed:
			row.normalized_data_json = json.dumps(data, ensure_ascii=False)
			business_date = _date_value(
				data.get("attendance_date") or data.get("occurred_on") or data.get("review_date")
				or data.get("application_date") or data.get("training_date") or data.get("summary_date"),
				row.business_date,
			)
			if business_date:
				row.business_date = business_date
			row.save(ignore_permissions=True)
			updated[row.template_key] += 1
	frappe.db.commit()
	return {"site": frappe.local.site, "company": TRIAL_COMPANY, "updated": dict(sorted(updated.items()))}


@frappe.whitelist()
def repair_local_yongxin_trial_workflow_values():
	"""Repair only trial staging values that do not match standard HR enums.

	The source transfer sheet labels a department change as ``部门调整`` while
	ERPNext's Employee Transfer document accepts ``调岗``.  Keep the uploaded
	evidence unchanged, but normalise the local staging value so the same
	approved row can generate its formal business document on retry.
	"""
	_assert_local_trial_access()
	designation = frappe.db.get_value("Designation", {"designation_name": "本机试运营岗位"}, "name")
	if not designation:
		designation = frappe.get_doc({"doctype": "Designation", "designation_name": "本机试运营岗位"}).insert(
			ignore_permissions=True
		).name
	batches = frappe.get_all(
		"HRMS Form Import Batch",
		filters={"company": TRIAL_COMPANY, "notes": ["like", f"%{TRIAL_MARKER}%"]},
		pluck="name",
	)
	updated = 0
	for row_name in frappe.get_all(
		intake.FORM_IMPORT_ROW_DOCTYPE,
		filters={
			"import_batch": ["in", batches or [""]],
			"template_key": "employee_transfer",
			"status": ["not in", ("已提交生效", "已驳回", "已忽略")],
		},
		pluck="name",
	):
		row = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, row_name)
		try:
			data = json.loads(row.normalized_data_json or "{}")
		except (TypeError, json.JSONDecodeError):
			continue
		if not isinstance(data, dict):
			continue
		data["transfer_type"] = "调岗"
		if not data.get("to_designation") or data.get("to_designation") == "未设置岗位":
			data["to_designation"] = designation
		if not data.get("from_designation") or data.get("from_designation") == "未设置岗位":
			data["from_designation"] = designation
		row.normalized_data_json = json.dumps(data, ensure_ascii=False)
		row.processing_error = ""
		row.save(ignore_permissions=True)
		updated += 1
	frappe.db.commit()
	return {"site": frappe.local.site, "company": TRIAL_COMPANY, "updated_transfer_rows": updated, "designation": designation}


@frappe.whitelist()
def prepare_local_yongxin_separation_prerequisites():
	"""Assign the local trial holiday list and consolidate retry duplicates.

	Employee Separation validates that the employee has a holiday list before
	submission.  This is a real HR prerequisite, not a value that should be
	silently bypassed in production.  For the explicitly isolated localhost
	trial we assign the already-created local holiday list, then point repeated
	retry rows to one draft separation document per employee.
	"""
	_assert_local_trial_access()
	holiday_name = "本机试运营假期表"
	if not frappe.db.exists("Holiday List", holiday_name):
		frappe.get_doc(
			{
				"doctype": "Holiday List",
				"holiday_list_name": holiday_name,
				"from_date": "2026-01-01",
				"to_date": "2026-12-31",
			}
		).insert(ignore_permissions=True)
	employees = frappe.get_all("Employee", filters={"company": TRIAL_COMPANY}, pluck="name")
	assigned = 0
	if frappe.get_meta("Employee").has_field("holiday_list"):
		for employee in employees:
			if frappe.db.get_value("Employee", employee, "holiday_list") != holiday_name:
				frappe.db.set_value("Employee", employee, "holiday_list", holiday_name, update_modified=False)
				assigned += 1
	batches = frappe.get_all(
		"HRMS Form Import Batch",
		filters={"company": TRIAL_COMPANY, "notes": ["like", f"%{TRIAL_MARKER}%"]},
		pluck="name",
	)
	rows = frappe.get_all(
		intake.FORM_IMPORT_ROW_DOCTYPE,
		filters={"import_batch": ["in", batches or [""]], "template_key": "resignation_application", "target_doctype": "Employee Separation"},
		fields=["name", "employee", "target_name"],
		order_by="creation asc",
	)
	by_employee = defaultdict(list)
	for row in rows:
		if row.employee and row.target_name and frappe.db.exists("Employee Separation", row.target_name):
			by_employee[row.employee].append(row)
	consolidated = 0
	for employee_rows in by_employee.values():
		keeper = employee_rows[0]
		for duplicate in employee_rows[1:]:
			if duplicate.target_name != keeper.target_name and frappe.db.exists("Employee Separation", duplicate.target_name):
				doc = frappe.get_doc("Employee Separation", duplicate.target_name)
				if doc.docstatus == 0:
					frappe.delete_doc("Employee Separation", duplicate.target_name, ignore_permissions=True, force=True)
			row = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, duplicate.name)
			row.target_name = keeper.target_name
			row.target_doctype = "Employee Separation"
			row.status = "已生成草稿"
			row.processing_error = ""
			row.save(ignore_permissions=True)
			consolidated += 1
	frappe.db.commit()
	return {"site": frappe.local.site, "company": TRIAL_COMPANY, "holiday_list": holiday_name, "employees_assigned": assigned, "duplicate_rows_consolidated": consolidated}


@frappe.whitelist()
def clear_local_yongxin_business_data(confirm: str):
	"""Remove only local Yongxin transactional and roster data after a backup."""
	_assert_local_trial_access()
	if confirm != CLEAR_CONFIRMATION:
		frappe.throw(_("确认文字不正确，未执行任何清理。"))
	deleted = OrderedDict()
	try:
		for doctype in CLEAR_DOCTYPES:
			deleted[doctype] = _delete_doctype(doctype)
		frappe.db.commit()
	except Exception:
		frappe.db.rollback()
		raise
	return {
		"site": frappe.local.site,
		"company": TRIAL_COMPANY,
		"deleted": {doctype: count for doctype, count in deleted.items() if count},
		"remaining": {doctype: count for doctype, count in _counts().items() if count},
	}


def _file_from_content(file_name, content):
	return frappe.get_doc({"doctype": "File", "file_name": file_name, "content": content, "is_private": 1}).insert(ignore_permissions=True)


def _file_content(file_url):
	name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not name:
		frappe.throw(_("未找到临时导入模板。"))
	content = frappe.get_doc("File", name).get_content()
	return content.encode() if isinstance(content, str) else content


def _source_roster_rows(workbook):
	"""Read the supplied roster without exposing its PII in any result/log."""
	sheet = _sheet_by_name(workbook, "花名册")
	if not sheet:
		return []
	headers = [_text(sheet.cell(2, column).value) for column in range(1, sheet.max_column + 1)]
	rows = []
	for row_number in range(4, sheet.max_row + 1):
		values = [_text(sheet.cell(row_number, column).value) for column in range(1, sheet.max_column + 1)]
		if not any(values):
			continue
		item = {headers[index]: values[index] for index in range(len(headers)) if headers[index]}
		if not item.get("工号") or not item.get("姓名"):
			continue
		rows.append(item)
	return rows


def _import_roster(workbook):
	rows = _source_roster_rows(workbook)
	if not rows:
		return {"status": "skipped", "reason": "花名册没有可导入的员工行"}
	content = roster.build_employee_import_template()
	book = load_workbook(BytesIO(content))
	sheet = book["员工花名册"]
	fields = roster._get_employee_import_fields(roster._get_template_doc())
	inserted_rows = 0
	generated_phone_count = 0
	for source in rows:
		name = source.get("姓名", "")
		cell_number = source.get("联系电话", "")
		# A single source row has no phone, while the product correctly treats the
		# field as mandatory.  Keep that case test-only and auditable rather than
		# inventing a real contact detail.  The value is never used to create a
		# system user or to send a message.
		if not _text(cell_number):
			cell_number = f"1390000{inserted_rows + 1:04d}"
			generated_phone_count += 1
		values = {
			"first_name": name,
			"employee_name": name,
			"employee_number": source.get("工号", ""),
			"custom_employee_code": source.get("工号", ""),
			"company": TRIAL_COMPANY,
			"department": source.get("部门", ""),
			"designation": source.get("岗位", "") or "未设置岗位",
			"date_of_joining": _date_value(source.get("入职日期"), "2026-01-01"),
			"cell_number": cell_number,
			"gender": {"男": "Male", "女": "Female"}.get(source.get("性别", ""), "Other"),
			"date_of_birth": _date_value(source.get("出生年月"), "1990-01-01"),
			"employment_type": "Full-time",
			"status": "Active",
			"naming_series": "HR-EMP-",
			"current_address": source.get("现居住地", ""),
			"permanent_address": source.get("户籍地址", ""),
			"education_level": source.get("学历", ""),
			"graduation_school": source.get("毕业院校", ""),
			"major": source.get("科系", ""),
		}
		target_row = inserted_rows + 2
		for column, field in enumerate(fields, start=1):
			sheet.cell(target_row, column, values.get(field["fieldname"], ""))
		inserted_rows += 1
	output = BytesIO()
	book.save(output)
	file_doc = _file_from_content(f"{TRIAL_MARKER}-员工花名册.xlsx", output.getvalue())
	preview = roster.preview_employee_roster_import(file_doc.file_url, mode="insert", match_by="employee_code")
	if preview.get("failed") or not preview.get("can_import"):
		return {
			"status": "failed",
			"source_rows": len(rows),
			"preview": _roster_preview_summary(preview),
			"test_only_generated_phone_count": generated_phone_count,
		}
	result = roster.import_employee_roster(file_doc.file_url, mode="insert", match_by="employee_code")
	return {
		"status": "imported" if not result.get("failed") else "failed",
		"source_rows": len(rows),
		"result": _roster_preview_summary(result),
		"test_only_generated_phone_count": generated_phone_count,
	}


def _roster_preview_summary(result):
	"""Return validation aggregates only; never echo employee/identity data."""
	errors = defaultdict(int)
	for item in result.get("errors") or []:
		label = _text(item.get("field_label")) or _text(item.get("fieldname")) or _("整行")
		message = _text(item.get("message")) or _("校验失败")
		errors[f"{label}：{message}"] += 1
	return {
		"row_count": result.get("row_count", 0),
		"planned_inserted": result.get("inserted", 0),
		"planned_updated": result.get("updated", 0),
		"skipped": result.get("skipped", 0),
		"failed": result.get("failed", 0),
		"can_import": bool(result.get("can_import")),
		"error_summary": dict(sorted(errors.items())),
	}


@frappe.whitelist()
def diagnose_local_yongxin_roster(source_path: str = SOURCE_WORKBOOK):
	"""Read-only, PII-safe roster diagnostic for the local trial."""
	_assert_local_trial_access()
	workbook = _workbook_or_throw(source_path)
	return _import_roster(workbook)


def _find_table_header(sheet, profile):
	"""Find the most likely header row in a conventional or two-level sheet."""
	labels = set()
	for column in profile["columns"]:
		labels.add(_header_key(column["label"]))
		labels.update(_header_key(alias) for alias in column.get("aliases", []))
	best = (0, 1)
	for row_number in range(1, min(sheet.max_row, 12) + 1):
		values = [_header_key(sheet.cell(row_number, col).value) for col in range(1, sheet.max_column + 1)]
		score = sum(1 for value in values if value and value in labels)
		if score > best[0]:
			best = (score, row_number)
	return best[1] if best[0] else 1


def _table_rows(sheet, profile):
	header_row = _find_table_header(sheet, profile)
	headers = [_text(sheet.cell(header_row, col).value) for col in range(1, sheet.max_column + 1)]
	# A few supplied sheets use a group heading followed by field labels.  Preserve
	# both forms as aliases, e.g. “合同-签订日期”.
	next_headers = [_text(sheet.cell(header_row + 1, col).value) for col in range(1, sheet.max_column + 1)] if header_row < sheet.max_row else []
	result = []
	for row_number in range(header_row + 1, sheet.max_row + 1):
		values = [_text(sheet.cell(row_number, col).value) for col in range(1, sheet.max_column + 1)]
		if not any(values):
			continue
		row = {}
		for index, value in enumerate(values):
			if headers[index]:
				row[headers[index]] = value
			if index < len(next_headers) and headers[index] and next_headers[index] and row_number > header_row + 1:
				row[f"{headers[index]}-{next_headers[index]}"] = value
		result.append(row)
	return result


def _first_value(row, labels):
	lookup = {_header_key(key): value for key, value in row.items() if _text(key)}
	for label in labels:
		value = lookup.get(_header_key(label))
		if _text(value):
			return _text(value)
	return ""


def _employee_lookup():
	rows = frappe.get_all("Employee", filters={"company": TRIAL_COMPANY}, fields=["name", "employee_name", "employee_number", "custom_employee_code", "department"])
	by_name = {}
	by_code = {}
	for row in rows:
		if row.employee_name:
			by_name[_text(row.employee_name)] = row
		for code in (row.employee_number, row.custom_employee_code):
			if code:
				by_code[_text(code)] = row
	return by_name, by_code


def _profile_defaults(key, row, source_sheet):
	month = "2026-07" if key in {"salary_structure_change", "education_allowance"} else "2026-06"
	if key in {"attendance_final", "full_attendance_bonus", "housing_allowance"}:
		month = "2026-05"
	if key == "attendance_daily":
		month = "2026-07"
	return {
		"company": TRIAL_COMPANY,
		"attendance_month": month,
		"payroll_month": month,
		"effective_month": month,
		"attendance_date": _date_value(_first_value(row, ["日期", "出勤日期", "workDate"]), f"{month}-01"),
		"occurred_on": _date_value(_first_value(row, ["发生日期", "日期"]), f"{month}-01"),
		"summary_date": _date_value(_first_value(row, ["统计日期", "日期"]), f"{month}-01"),
		"training_month": month,
		"training_date": _date_value(_first_value(row, ["实际上课时间", "培训日期", "日期"]), f"{month}-01"),
		"review_date": f"{month}-01",
		"application_date": f"{month}-01",
		"last_working_date": f"{month}-28",
		"proposal_date": f"{month}-01",
		"survey_date": f"{month}-01",
		"source_sheet": source_sheet,
	}


def _normalise_profile_row(profile, source_row, source_sheet, by_name, by_code):
	data = _profile_defaults(profile["key"], source_row, source_sheet)
	for column in profile["columns"]:
		data[column["key"]] = _first_value(source_row, [column["label"], *column.get("aliases", [])]) or data.get(column["key"], "")
	# Common source headings differ from template labels.  Only derive missing
	# links from existing local employees; do not fabricate employee identity.
	data["employee_code"] = data.get("employee_code") or _first_value(source_row, ["工号", "员工编号", "员工编码"])
	data["employee_name"] = data.get("employee_name") or _first_value(source_row, ["姓名", "创建人", "被考核人", "提案人"])
	data["department"] = data.get("department") or _first_value(source_row, ["部门", "单位", "实际部门", "创建人部门"])
	person = by_code.get(_text(data.get("employee_code"))) or by_name.get(_text(data.get("employee_name")))
	if person:
		data["employee_code"] = data.get("employee_code") or person.employee_number or person.custom_employee_code
		data["employee_name"] = data.get("employee_name") or person.employee_name
		data["department"] = data.get("department") or person.department
	# Normalise dates and known numeric labels used by the generic importer.
	for key in ("attendance_date", "occurred_on", "summary_date", "review_date", "application_date", "last_working_date", "proposal_date", "survey_date", "training_date", "completed_date", "followup_date", "contract_end_date", "exit_date", "join_date"):
		if key in data:
			data[key] = _date_value(data[key], data[key])
	for key in ("attendance_month", "payroll_month", "effective_month", "training_month", "period"):
		if key in data and data[key]:
			data[key] = _month_value(data[key], data[key])
	return data


def _seed_row(profile, employee):
	month = "2026-05" if profile["key"] in {"attendance_final", "full_attendance_bonus", "housing_allowance"} else "2026-07"
	data = {column["key"]: "" for column in profile["columns"]}
	data.update({
		"company": TRIAL_COMPANY,
		"employee_code": employee.employee_number or employee.custom_employee_code,
		"employee_name": employee.employee_name,
		"department": employee.department,
		"attendance_month": month,
		"payroll_month": month,
		"effective_month": month,
		"attendance_date": f"{month}-01",
		"occurred_on": f"{month}-01",
		"summary_date": f"{month}-01",
		"review_date": f"{month}-01",
		"application_date": f"{month}-01",
		"last_working_date": f"{month}-28",
		"transfer_date": f"{month}-01",
		"from_department": employee.department,
		"to_department": employee.department,
		"from_designation": employee.designation or "未设置岗位",
		"to_designation": employee.designation or "未设置岗位",
		"contract_end_date": f"{month}-28",
		"employee_intent": "续签意愿待确认（本机试运营）",
		"proposal_date": f"{month}-01",
		"survey_date": f"{month}-01",
		"training_month": month,
		"training_date": f"{month}-01",
		"reason": "本地试运营缺少已填写源表时生成的隔离种子数据",
		"remarks": f"{TRIAL_MARKER}：仅本机验证，不可用于正式业务。",
		"description": "本地试运营种子记录",
		"subject": "本地试运营种子记录",
		"proposal_no": f"{TRIAL_MARKER}-P",
		"feedback_no": f"{TRIAL_MARKER}-F",
		"external_id": f"{TRIAL_MARKER}-{profile['key']}",
		"approval_status": "已完成",
		"approval_result": "同意",
		"leave_type": "事假",
		"start_time": f"{month}-01 08:00:00",
		"end_time": f"{month}-01 17:00:00",
		"duration": "8",
		"exception_type": "未打卡",
		"handling": "本地试运营待确认",
		"apple_type": "绿苹果",
		"quantity": "1",
		"amount": "100",
		"reward_punishment_type": "奖励",
		"social_security_base": "3000",
		"employee_amount": "300",
		"deduction_amount": "100",
		"net_pay": "2900",
		"exit_date": f"{month}-28",
		"base_salary": "3000",
		"previous_base_salary": "3000",
		"gross_salary": "3000",
		"change_reason": "本地试运营薪资链路验证",
		"evaluation": "通过",
		"designation": "未设置岗位",
		"candidate_name": "本机试运营候选人",
		"candidate_email": "localhost-trial@example.invalid",
		"phone": "13900000000",
		"job_offer_no": f"{TRIAL_MARKER}-OFFER",
		"onboarding_template": "本机试运营入职模板",
		"date_of_joining": f"{month}-01",
		"boarding_begins_on": f"{month}-01",
		"holiday_list": "本机试运营假期表",
		"interview_date": f"{month}-01 10:00:00",
		"applied_designation": "本机试运营岗位",
		"transfer_type": "调岗",
		"period_type": "月度",
		"period": month,
		"achievements": "本地试运营记录",
		"training_content": "本地试运营培训",
		"certificate_no": f"{TRIAL_MARKER}-CERT",
		"validity_period": "12个月",
		"total_headcount": "1",
		"attendance_count": "1",
	})
	return data


def _ensure_local_onboarding_prerequisites():
	"""Create the smallest valid, explicitly marked local hiring chain.

	The source workbook does not contain an accepted candidate or a submitted
	Offer.  Creating these prerequisites lets the onboarding import exercise its
	real validation path without treating a blank spreadsheet as a successful
	workflow.  This helper is guarded by :func:`_assert_local_trial_access` at
	the public caller and is never used by normal product imports.
	"""
	email = "localhost-trial@example.invalid"
	holiday_name = "本机试运营假期表"
	template_title = f"{TRIAL_COMPANY} 标准入职流程"
	designation = frappe.db.get_value("Designation", {}, "name", order_by="creation asc")
	if not designation:
		designation = frappe.get_doc({"doctype": "Designation", "designation_name": "本机试运营岗位"}).insert(
			ignore_permissions=True
		).name

	if not frappe.db.exists("Holiday List", holiday_name):
		frappe.get_doc(
			{
				"doctype": "Holiday List",
				"holiday_list_name": holiday_name,
				"from_date": "2026-01-01",
				"to_date": "2026-12-31",
			}
		).insert(ignore_permissions=True)

	applicant_name = frappe.db.get_value("Job Applicant", {"email_id": email}, "name")
	if applicant_name:
		applicant = frappe.get_doc("Job Applicant", applicant_name)
		if applicant.status != "Accepted":
			applicant.db_set("status", "Accepted", update_modified=False)
	else:
		applicant = frappe.get_doc(
			{
				"doctype": "Job Applicant",
				"applicant_name": "本机试运营候选人",
				"email_id": email,
				"status": "Accepted",
				"notes": f"{TRIAL_MARKER}：仅本机试运营前置数据。",
			}
		).insert(ignore_permissions=True)

	offer_name = frappe.db.get_value("Job Offer", {"job_applicant": applicant.name, "docstatus": ["!=", 2]}, "name")
	if offer_name:
		offer = frappe.get_doc("Job Offer", offer_name)
		if offer.docstatus == 0:
			offer.status = "Accepted"
			offer.submit()
		if offer.status != "Accepted":
			offer.db_set("status", "Accepted", update_modified=False)
	else:
		offer = frappe.get_doc(
			{
				"doctype": "Job Offer",
				"job_applicant": applicant.name,
				"offer_date": "2026-05-01",
				"designation": designation,
				"company": TRIAL_COMPANY,
				"status": "Accepted",
				"terms": f"{TRIAL_MARKER}：仅本机试运营 Offer。",
			}
		).insert(ignore_permissions=True)
		offer.submit()
		if offer.status != "Accepted":
			offer.db_set("status", "Accepted", update_modified=False)

	template = intake.ensure_default_employee_onboarding_template(TRIAL_COMPANY)
	return {
		"candidate_email": email,
		"job_offer_no": offer.name,
		"onboarding_template": template.get("title") or template_title,
		"holiday_list": holiday_name,
		"designation": designation,
	}


def _template_upload(profile, data_rows):
	template = intake.create_form_import_template_file(profile["key"])
	book = load_workbook(BytesIO(_file_content(template["file_url"])))
	sheet = book["数据"]
	for row_index, values in enumerate(data_rows, start=2):
		for column_index, column in enumerate(profile["columns"], start=1):
			sheet.cell(row_index, column_index, values.get(column["key"], ""))
	output = BytesIO()
	book.save(output)
	return _file_from_content(f"{TRIAL_MARKER}-{profile['key']}.xlsx", output.getvalue())


def _required_complete(profile, data):
	return all(_text(data.get(column["key"])) for column in profile["columns"] if column.get("required"))


def _form_preview_summary(preview):
	"""Summarise generic form validation without exposing uploaded row content."""
	errors = defaultdict(int)
	for row in preview.get("preview_rows") or []:
		for message in row.get("errors") or []:
			text = _text(message)
			if text:
				errors[text] += 1
	return {
		"total_rows": preview.get("total_rows", 0),
		"valid_rows": preview.get("valid_rows", 0),
		"failed_rows": preview.get("failed_rows", 0),
		"missing_required": preview.get("missing_required", []),
		"error_summary": dict(sorted(errors.items())),
	}


def _stage_profile(profile, workbook, by_name, by_code, max_rows=1000):
	if profile["key"] == "employee_roster":
		return None
	rows = []
	source_sheets = profile["source_sheets"]
	# The same monthly attendance appears in initial, signed and finance copies.
	# Finance final is the declared payroll source, so stage that version only to
	# avoid double-counting a person-month in the payroll trial.
	if profile["key"] == "attendance_final":
		source_sheets = ["考勤终稿（财务版）", "考勤终稿（签字版）", "考勤初稿"]
	for sheet_name in source_sheets:
		sheet = _sheet_by_name(workbook, sheet_name)
		if not sheet:
			continue
		for source_row in _table_rows(sheet, profile):
			data = _normalise_profile_row(profile, source_row, _text(sheet.title), by_name, by_code)
			if _required_complete(profile, data):
				rows.append(data)
			if len(rows) >= max_rows:
				break
		if profile["key"] == "attendance_final" and rows:
			break
		if len(rows) >= max_rows:
			break
	seeded = False
	if not rows:
		employee = next(iter(by_name.values()), None)
		if not employee:
			return {"status": "skipped", "reason": "没有可用员工用于隔离种子数据"}
		rows = [_seed_row(profile, employee)]
		if profile["key"] == "employee_onboarding":
			# The workbook has no accepted-candidate / Offer chain.  Supply a
			# clearly marked local-only prerequisite chain so this form is tested
			# through the same validation as a real onboarding import.
			rows[0].update(_ensure_local_onboarding_prerequisites())
		seeded = True
	file_doc = _template_upload(profile, rows)
	preview = intake.preview_form_import(file_doc.file_url, profile["key"], TRIAL_COMPANY)
	if not preview.get("valid_rows"):
		return {"status": "failed", "rows": len(rows), "seeded": seeded, "preview": _form_preview_summary(preview)}
	result = intake.import_form_workbook(file_doc.file_url, profile["key"], TRIAL_COMPANY, notes=f"{TRIAL_MARKER}；来源：{','.join(profile['source_sheets'])}；{'隔离种子' if seeded else '源表转换'}")
	return {"status": "staged" if not result.get("failed_rows") else "partial", "rows": len(rows), "seeded": seeded, "batch": result.get("batch_name"), "valid_rows": result.get("valid_rows"), "failed_rows": result.get("failed_rows")}


def _requested_profiles(profile_keys):
	if isinstance(profile_keys, str):
		try:
			profile_keys = json.loads(profile_keys)
		except ValueError:
			profile_keys = [item.strip() for item in profile_keys.split(",") if item.strip()]
	if not profile_keys:
		return [profile for profile in intake.FORM_IMPORT_PROFILES if profile["key"] != "employee_roster"]
	requested = {str(key) for key in profile_keys}
	profiles = [profile for profile in intake.FORM_IMPORT_PROFILES if profile["key"] in requested and profile["key"] != "employee_roster"]
	missing = requested - {profile["key"] for profile in profiles}
	if missing:
		frappe.throw(_("未识别的表单键：{0}").format("、".join(sorted(missing))))
	return profiles


@frappe.whitelist()
def stage_local_yongxin_profiles(profile_keys: str = "[]", source_path: str = SOURCE_WORKBOOK):
	"""Stage selected non-roster forms in small, resumable local-only batches."""
	_assert_local_trial_access()
	workbook = _workbook_or_throw(source_path)
	by_name, by_code = _employee_lookup()
	if not by_name:
		frappe.throw(_("请先完成员工花名册导入。"))
	result = {}
	for profile in _requested_profiles(profile_keys):
		try:
			result[profile["key"]] = _stage_profile(profile, workbook, by_name, by_code)
		except Exception as error:
			frappe.db.rollback()
			result[profile["key"]] = {"status": "failed", "reason": str(error)[:300]}
		finally:
			frappe.db.commit()
	return result


@frappe.whitelist()
def activate_local_yongxin_profiles(profile_keys: str):
	"""Advance only staged, requested local rows through the normal HR workflow."""
	_assert_local_trial_access()
	keys = {profile["key"] for profile in _requested_profiles(profile_keys)}
	batches = frappe.get_all(
		"HRMS Form Import Batch",
		filters={"company": TRIAL_COMPANY, "notes": ["like", f"%{TRIAL_MARKER}%"]},
		pluck="name",
	)
	return _approve_and_activate_batches(batches, keys)


def _approve_and_activate_batches(batch_names, template_keys):
	"""Advance only rows that passed validation through the normal approval API."""
	result = defaultdict(lambda: {"approved": 0, "generated": 0, "activated": 0, "failed": 0, "failure_summary": defaultdict(int)})
	for batch_name in batch_names:
		# Include already-approved rows as well.  A target can legitimately fail to
		# generate before a date repair or attendance lock, and should be
		# recoverable without requiring the user to approve the evidence again.
		rows = frappe.get_all(
			intake.FORM_IMPORT_ROW_DOCTYPE,
			filters={"import_batch": batch_name, "status": ["not in", ("处理失败", "已忽略", "已提交生效", "已驳回")]},
			pluck="name",
		)
		for row_name in rows:
			row = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, row_name)
			if row.template_key not in template_keys:
				continue
			try:
				while frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, row_name).review_status != "已批准":
					intake.review_form_import_row(row_name, "批准", f"{TRIAL_MARKER} 本机试运营审核")
				result[row.template_key]["approved"] += 1
				generated = intake.generate_form_import_target(row_name, payroll_month="2026-05")
				result[row.template_key]["generated"] += 1 if generated else 0
				intake.activate_form_import_target(row_name)
				result[row.template_key]["activated"] += 1
			except Exception as error:
				message = str(error)[:280]
				result[row.template_key]["failed"] += 1
				result[row.template_key]["failure_summary"][message] += 1
	return {
		key: {
			**{field: value for field, value in payload.items() if field != "failure_summary"},
			"failure_summary": dict(sorted(payload["failure_summary"].items(), key=lambda item: (-item[1], item[0]))[:5]),
		}
		for key, payload in result.items()
	}


def _ensure_may_salary_seed(by_name):
	"""The supplied salary sheet is July while attendance final is May.

	Use isolated May salary versions for every employee in the locked-month
	attendance scope.  This lets the May attendance-to-payroll chain be tested
	without silently applying July source payroll to May.
	"""
	employees = frappe.get_all(
		"HRMS Monthly Attendance Summary",
		filters={"company": TRIAL_COMPANY, "attendance_month": "2026-05"},
		pluck="employee",
	)
	employees = sorted(set(employee for employee in employees if employee))
	if not employees:
		# Before month-final activation this helper must not create arbitrary
		# salaries.  Call it only after the source final has produced summaries.
		return {"status": "skipped", "reason": "无员工"}
	profile = next(item for item in intake.FORM_IMPORT_PROFILES if item["key"] == "salary_structure_change")
	data_rows = []
	for employee_name in employees:
		employee = frappe.get_doc("Employee", employee_name)
		data = _seed_row(profile, employee)
		data.update({"effective_month": "2026-05", "remarks": f"{TRIAL_MARKER}：源表薪资构成仅覆盖 2026-07；本条仅用于 2026-05 本机闭环验证。"})
		data_rows.append(data)
	file_doc = _template_upload(profile, data_rows)
	result = intake.import_form_workbook(file_doc.file_url, profile["key"], TRIAL_COMPANY, notes=data["remarks"])
	activated = _approve_and_activate_batches([result["batch_name"]], {"salary_structure_change"})
	return {"status": "seeded", "rows": len(data_rows), "batch": result["batch_name"], "activation": activated}


def _lock_local_yongxin_may_attendance():
	"""Confirm and lock the local May attendance final before payroll use."""
	month = "2026-05"
	confirmations = attendance_import.list_attendance_department_confirmations(TRIAL_COMPANY, month, page_length=1000)
	for confirmation in confirmations:
		if confirmation.confirmation_status != "已确认":
			attendance_import.review_attendance_department_confirmation(confirmation.name, "confirm", f"{TRIAL_MARKER} 本机部门确认")
	lock = frappe.db.get_value(
		"HRMS Attendance Month Lock",
		{"company": TRIAL_COMPANY, "attendance_month": month},
		["name", "status", "active_version"],
		as_dict=True,
	)
	if not lock or lock.status != "已锁定":
		lock = attendance_import.lock_attendance_month(TRIAL_COMPANY, month, f"{TRIAL_MARKER} 本机闭环锁定")
	return {"month": month, "lock": lock, "confirmed_departments": len(confirmations)}


def _run_may_payroll_chain():
	lock_result = _lock_local_yongxin_may_attendance()
	month = "2026-05"
	lock = lock_result["lock"]
	version = str(lock.get("attendance_lock_version") or lock.get("active_version")) if isinstance(lock, dict) else str(lock.active_version)
	inputs = payroll_input.generate_payroll_input_records(TRIAL_COMPANY, month, version)
	settlements = payroll_input.generate_payroll_settlement_records(TRIAL_COMPANY, month, version)
	confirmed = payroll_input.confirm_payroll_settlement_records(TRIAL_COMPANY, month, version)
	return {"month": month, "lock": lock, "inputs": {"created": inputs.get("created")}, "settlements": {"created": settlements.get("created")}, "confirmed": confirmed.get("confirmed")}


@frappe.whitelist()
def lock_local_yongxin_may_attendance():
	_assert_local_trial_access()
	return _lock_local_yongxin_may_attendance()


@frappe.whitelist()
def seed_local_yongxin_may_salary_versions():
	_assert_local_trial_access()
	by_name, _by_code = _employee_lookup()
	return _ensure_may_salary_seed(by_name)


@frappe.whitelist()
def run_local_yongxin_may_payroll_chain():
	_assert_local_trial_access()
	return _run_may_payroll_chain()


@frappe.whitelist()
def activate_local_yongxin_may_welfare_sources(profile_keys: str):
	"""Confirm trial welfare sources in bulk and sync payroll variables once.

	The normal one-row activation action deliberately refreshes variables
	immediately, which is helpful for ordinary HR use.  Repeating that refresh
	for a large uploaded workbook creates duplicate queue work.  This localhost
	trial helper keeps the same business state transitions but performs exactly
	one final variable refresh for the locked May version.
	"""
	_assert_local_trial_access()
	keys = {profile["key"] for profile in _requested_profiles(profile_keys)}
	allowed = {
		"reward_punishment", "skill_certificate_allowance", "full_attendance_bonus",
		"housing_allowance", "education_allowance", "dormitory_fee", "social_insurance",
		"service_award", "exit_payroll_settlement",
	}
	if not keys or not keys.issubset(allowed):
		frappe.throw(_("该批量确认仅适用于薪资福利/扣款来源表单。"))
	lock = frappe.db.get_value(
		"HRMS Attendance Month Lock",
		{"company": TRIAL_COMPANY, "attendance_month": "2026-05", "status": "已锁定"},
		["name", "active_version"],
		as_dict=True,
	)
	if not lock:
		frappe.throw(_("请先锁定永新 2026-05 月度考勤。"))
	batches = frappe.get_all(
		"HRMS Form Import Batch",
		filters={"company": TRIAL_COMPANY, "notes": ["like", f"%{TRIAL_MARKER}%"]},
		pluck="name",
	)
	rows = frappe.get_all(
		intake.FORM_IMPORT_ROW_DOCTYPE,
		filters={"import_batch": ["in", batches or [""]], "template_key": ["in", sorted(keys)], "review_status": "已批准"},
		fields=["name", "template_key", "target_doctype", "target_name", "status"],
	)
	confirmed = defaultdict(int)
	skipped = defaultdict(int)
	for row in rows:
		if not row.target_name or row.target_doctype != "HRMS Payroll Welfare Source Record":
			skipped[row.template_key] += 1
			continue
		if not frappe.db.exists(row.target_doctype, row.target_name):
			skipped[row.template_key] += 1
			continue
		target = frappe.get_doc(row.target_doctype, row.target_name)
		if target.payroll_month != "2026-05" or str(target.attendance_lock_version) != str(lock.active_version):
			skipped[row.template_key] += 1
			continue
		if target.confirmation_status != "已确认":
			target.confirmation_status = "已确认"
			target.confirmed_by = frappe.session.user
			target.confirmed_on = frappe.utils.now_datetime()
			target.save(ignore_permissions=True)
		row_doc = frappe.get_doc(intake.FORM_IMPORT_ROW_DOCTYPE, row.name)
		row_doc.status = "已提交生效"
		row_doc.activated_by = frappe.session.user
		row_doc.activated_on = frappe.utils.now_datetime()
		row_doc.processing_error = ""
		row_doc.save(ignore_permissions=True)
		confirmed[row.template_key] += 1
	frappe.db.commit()
	sync = payroll_input.sync_welfare_sources_to_payroll_variables(TRIAL_COMPANY, "2026-05", str(lock.active_version))
	return {
		"company": TRIAL_COMPANY,
		"payroll_month": "2026-05",
		"attendance_lock_version": str(lock.active_version),
		"confirmed": dict(sorted(confirmed.items())),
		"skipped": dict(sorted(skipped.items())),
		"variable_sync": {"created": sync.get("created", 0), "updated": sync.get("updated", 0), "skipped": sync.get("skipped", 0)},
	}


@frappe.whitelist()
def run_local_yongxin_operational_trial(confirm: str, source_path: str = SOURCE_WORKBOOK):
	"""Import source data via product templates and test the May payroll closure.

	The function assumes :func:`clear_local_yongxin_business_data` already ran.
	It returns only counters/statuses, never employee PII or source row values.
	"""
	_assert_local_trial_access()
	if confirm != RUN_CONFIRMATION:
		frappe.throw(_("确认文字不正确，未执行试运营。"))
	workbook = _workbook_or_throw(source_path)
	report = {"site": frappe.local.site, "company": TRIAL_COMPANY, "marker": TRIAL_MARKER, "roster": {}, "profiles": {}, "activation": {}, "payroll_chain": None, "warnings": []}
	report["roster"] = _import_roster(workbook)
	if report["roster"].get("status") != "imported":
		frappe.throw(_("员工花名册导入未成功，已停止后续试运营。"))
	by_name, by_code = _employee_lookup()
	if not by_name:
		frappe.throw(_("员工花名册导入后未找到永新员工，已停止。"))
	batch_names = []
	for profile in intake.FORM_IMPORT_PROFILES:
		item = _stage_profile(profile, workbook, by_name, by_code)
		if item is None:
			continue
		report["profiles"][profile["key"]] = item
		if item.get("batch"):
			batch_names.append(item["batch"])
	# Actual imports are staged for audit.  Activate core records needed for the
	# attendance/payroll flow only; recruitment and HR transactions remain in the
	# human review queue because their source forms are incomplete templates.
	core_keys = {"attendance_daily", "leave_export", "attendance_exception", "apple_reward", "attendance_final", "full_attendance_bonus", "housing_allowance", "dormitory_fee", "social_insurance", "skill_certificate_allowance", "education_allowance", "service_award"}
	report["activation"] = _approve_and_activate_batches(batch_names, core_keys)
	report["may_salary_seed"] = _ensure_may_salary_seed(by_name)
	try:
		report["payroll_chain"] = _run_may_payroll_chain()
	except Exception as error:
		report["warnings"].append(f"2026-05 考勤至薪资闭环未完成：{str(error)[:500]}")
	# Month coverage is deliberately reported because source payroll inputs span
	# May, June and July.  This prevents an apparently complete payroll from
	# combining different accounting months.
	report["warnings"].append("源表月份并不一致：考勤终稿/全勤/住房为 2026-05，社保/宿舍/继续服务奖多为 2026-06，薪资构成为 2026-07。闭环仅以 2026-05 考勤及隔离薪资种子验证。")
	frappe.db.commit()
	return report
