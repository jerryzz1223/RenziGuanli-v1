import hashlib
import json
import os
import re
import zipfile
from datetime import date, datetime
from io import BytesIO
from xml.etree import ElementTree

import frappe
from frappe import _
from frappe.custom.doctype.custom_field.custom_field import create_custom_field
from frappe.custom.doctype.property_setter.property_setter import make_property_setter
from frappe.utils import cint, flt


TEMPLATE_DOCTYPE = "HRMS Employee Field Template"
TEMPLATE_CHILD_TABLE = "HRMS Employee Field Template Item"
HRMS_EMPLOYEE_REPORT_DOCTYPE = "HRMS Employee Report"
EMPLOYEE_DOCTYPE = "Employee"
EMPLOYEE_IMPORT_TEMPLATE_FILENAME = "员工导入模板.xlsx"
EMPLOYEE_EXPORT_FILENAME = "员工花名册导出.xlsx"
EMPLOYEE_FAILED_ROWS_FILENAME = "员工花名册失败行.xlsx"
EMPLOYEE_FALLBACK_DATE_OF_BIRTH = "1905-01-01"
EMPLOYEE_MATERIAL_FIELD_PREFIX = "hrms_material_"
EMPLOYEE_MATERIAL_FILE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".pdf"}
PAYROLL_WELFARE_SOURCE_DOCTYPE = "HRMS Payroll Welfare Source Record"
PAYROLL_SOCIAL_INSURANCE_SOURCE_TYPES = (
	"社保个人",
	"社保公司",
	"公积金个人",
	"公积金公司",
)
EMPLOYEE_MATERIAL_GROUPS = [
	{
		"label": "员工基本资料",
		"description": "身份证、学历证明、个人证件照等入职基础材料。",
		"types": [
			("identity_card_photo", "身份证照片"),
			("education_certificate", "学历证明"),
			("personal_id_photo", "个人证件照"),
			("identity_card_copy", "身份证复印件"),
		],
	},
	{
		"label": "员工档案资料",
		"description": "劳动合同、入职简历、体检单、入职记录等档案材料。",
		"types": [
			("labor_contract", "劳动合同"),
			("onboarding_resume", "入职简历"),
			("onboarding_record", "入职记录"),
			("onboarding_health_check", "入职体检单"),
		],
	},
	{
		"label": "员工离职资料",
		"description": "离职审批、离职证明、离职申请、工作交接表等材料。",
		"types": [
			("separation_approval", "离职审批"),
			("separation_certificate", "离职证明"),
			("separation_application", "离职申请"),
			("handover_form", "工作交接表"),
		],
	},
]
EMPLOYEE_IMPORT_NON_DEFERRABLE_FIELDS = {
	"custom_employee_code",
	"first_name",
	"employee_name",
	"department",
	"date_of_joining",
	"designation",
}
HR_SETTINGS_MANAGER_ROLES = ("HR Manager", "System Manager")
HR_SETTINGS_PAGE_ROLES = list(HR_SETTINGS_MANAGER_ROLES)
HRMS_DEVELOPER_PAGE_ROLES = ["System Manager"]
HRMS_ACCESS_PAGE_ROLES = ["System Manager"]

PERSONNEL_PAGE_DEFINITIONS = [
	{"name": "employee-detail", "title": "员工档案详情", "icon": "user"},
	{"name": "employee-roster-import", "title": "导入花名册", "icon": "upload"},
	{"name": "employee-roster-export", "title": "导出花名册", "icon": "download"},
	{"name": "personnel-reports", "title": "人事报表", "icon": "bar-chart"},
	{
		"name": "staff-attribute-settings",
		"title": "员工属性设置",
		"icon": "settings",
		"roles": HR_SETTINGS_PAGE_ROLES,
	},
	{
		"name": "hr-settings-center",
		"title": "设置中心",
		"icon": "settings",
		"roles": HR_SETTINGS_PAGE_ROLES,
	},
	{
		"name": "hrms-developer-center",
		"title": "开发中心",
		"icon": "code",
		"roles": HRMS_DEVELOPER_PAGE_ROLES,
	},
	{
		"name": "hrms-access-center",
		"title": "账户与权限中心",
		"icon": "key",
		"roles": HRMS_ACCESS_PAGE_ROLES,
	},
	{"name": "employee-property-history", "title": "异动记录", "icon": "timeline"},
	{"name": "cross-department-support", "title": "跨部门支援", "icon": "users"},
	{"name": "recruitment-center", "title": "招聘中心", "icon": "briefcase", "roles": ["HR User", "HR Manager", "System Manager", "Interviewer"]},
	{"name": "attendance-import-center", "title": "考勤导入中心", "icon": "upload"},
	{"name": "payroll-input-center", "title": "薪资输入中心", "icon": "database"},
	{"name": "form-data-intake", "title": "人资表单导入中心", "icon": "upload"},
]
LEGACY_PERSONNEL_PAGE_SLUGS = {
	"employee-property-history": "employee-property-hi",
}


def _require_hr_settings_manager():
	"""Protect schema/configuration changes from ordinary HR users.

	The settings centre is deliberately a business-admin surface.  Creating or
	changing fields ultimately writes Custom Field and Property Setter records,
	so hiding the navigation link alone is not a sufficient control.
	"""
	frappe.only_for(HR_SETTINGS_MANAGER_ROLES)


def _can_manage_personnel_pages():
	return frappe.session.user == "Administrator" or "System Manager" in frappe.get_roles(frappe.session.user)


def _require_system_manager():
	"""Keep account and permission summaries out of ordinary HR accounts."""
	frappe.only_for("System Manager")

PROPERTY_HISTORY_FIELD_CONTRACT = ["property", "current", "new", "fieldname"]
PROPERTY_HISTORY_SOURCES = [
	{
		"doctype": "Employee Transfer",
		"label": "人事异动",
		"date_field": "transfer_date",
		"detail_field": "transfer_details",
	},
	{
		"doctype": "Employee Promotion",
		"label": "转正/晋升",
		"date_field": "promotion_date",
		"detail_field": "promotion_details",
	},
]

EMPLOYEE_TEMPLATE_CATEGORIES = ["在职信息", "个人信息", "联系信息", "教育信息", "合同保险", "工资社保", "个税申报", "附件"]
EMPLOYEE_DETAIL_BLOCK_TABS = {
	"任职记录": "在职信息",
	"奖惩记录": "在职信息",
	"考察期信息": "在职信息",
	"退休信息": "在职信息",
	"档案信息": "在职信息",
	"教育经历": "个人信息",
	"工作经历": "个人信息",
	"语言能力": "个人信息",
	"工作技能": "个人信息",
	"合同记录": "合同信息",
	"社保公积金记录": "工资社保",
	"材料附件": "材料附件",
	"背景调查": "背景调查",
}
DEFAULT_DETAIL_BLOCK_DESCRIPTIONS = {
	"任职记录": "记录员工部门、岗位、职级、工作地点、任职起止日期等历史变化。",
	"教育经历": "记录学历、毕业院校、专业、学习形式、毕业时间等教育背景。",
	"工作经历": "记录入职前工作单位、岗位、起止时间和工作内容。",
	"语言能力": "记录语种、熟练程度、证书和备注。",
	"工作技能": "记录技能名称、熟练程度、证书和备注。",
	"奖惩记录": "记录奖励、处分、奖惩日期、奖惩原因、处理结果和附件。",
	"考察期信息": "记录试用期、考察开始/结束日期、考察结果和转正办理记录。",
	"退休信息": "记录退休日期、退休年龄、办理状态和退休备注。",
	"档案信息": "记录档案编号、档案位置、档案备注和归档状态。",
	"合同记录": "记录合同编号、签订日期、签订次数、合同期限和附件。",
	"社保公积金记录": "记录社保、医保、公积金缴纳情况。",
	"材料附件": "记录身份证、学历证明、合同扫描件、离职证明等员工材料。",
	"背景调查": "记录背调状态、背调结果、背调机构和附件。",
}
EXCEL_ERROR_VALUES = {"#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A", "#GETTING_DATA"}
MIN_REASONABLE_EMPLOYEE_DATE_YEAR = 1905
PHONE_FIELDNAMES = {"cell_number", "emergency_phone_number"}
GENDER_VALUE_ALIASES = {
	"男": "Male",
	"男性": "Male",
	"male": "Male",
	"m": "Male",
	"Male": "Male",
	"女": "Female",
	"女性": "Female",
	"female": "Female",
	"f": "Female",
	"Female": "Female",
	"其他": "Other",
	"其它": "Other",
	"other": "Other",
	"Other": "Other",
}
EMPLOYEE_DUPLICATE_MATCH_FIELDS = {
	"employee_code": ("custom_employee_code",),
	"id_card": ("passport_number", "custom_id_number"),
	"phone": ("cell_number",),
	"auto": ("custom_employee_code", "passport_number", "custom_id_number", "cell_number"),
}

DEFAULT_EMPLOYEE_REPORTS = [
	{
		"id": "active_roster",
		"report_name": "在职员工花名册",
		"description": "在职员工的人事信息表",
		"group_name": "人事档案",
		"fields": [
			"employee_name",
			"custom_employee_code",
			"department",
			"designation",
			"employment_type",
			"date_of_joining",
			"cell_number",
		],
		"filters": {"status": "Active"},
	},
	{
		"id": "employee_contact",
		"report_name": "员工通讯录",
		"description": "在职员工的通讯方式，包括联系电话、邮箱、居住地址等",
		"group_name": "人事档案",
		"fields": ["employee_name", "cell_number", "company_email", "personal_email", "current_address"],
		"filters": {"status": "Active"},
	},
	{
		"id": "contract_status",
		"report_name": "员工劳动合同签订情况",
		"description": "在职员工劳动合同签订与到期情况",
		"group_name": "人事档案",
		"fields": [
			"employee_name",
			"custom_employee_code",
			"department",
			"custom_contract_sign_date",
			"custom_contract_no",
			"contract_end_date",
		],
		"filters": {"status": "Active"},
	},
	{
		"id": "work_years",
		"report_name": "员工工作年限信息表",
		"description": "在职员工工作年限与入职时间统计",
		"group_name": "人事档案",
		"fields": ["employee_name", "custom_employee_code", "department", "date_of_joining", "custom_entry_age"],
		"filters": {"status": "Active"},
	},
	{
		"id": "birthday",
		"report_name": "员工生日信息表",
		"description": "在职员工生日信息统计",
		"group_name": "人事档案",
		"fields": ["employee_name", "department", "date_of_birth", "custom_age", "cell_number"],
		"filters": {"status": "Active"},
	},
	{
		"id": "education",
		"report_name": "员工学历信息表",
		"description": "在职员工学历、毕业院校、专业等信息",
		"group_name": "人事档案",
		"fields": [
			"employee_name",
			"department",
			"custom_education_category",
			"custom_education_level",
			"custom_graduation_school",
			"custom_major",
		],
		"filters": {"status": "Active"},
	},
	{
		"id": "left_employee",
		"report_name": "离职员工信息表",
		"description": "离职员工资料归档",
		"group_name": "人事档案",
		"fields": ["employee_name", "custom_employee_code", "department", "designation", "relieving_date", "reason_for_leaving"],
		"filters": {"status": "Left"},
	},
	{
		"id": "department_headcount",
		"report_name": "各部门员工花名册",
		"description": "按部门导出的在职员工基础信息",
		"group_name": "人事统计",
		"fields": ["employee_name", "custom_employee_code", "department", "designation", "employment_type", "date_of_joining"],
		"filters": {"status": "Active"},
	},
	{
		"id": "employee_age",
		"report_name": "各部门员工年龄段统计表",
		"description": "员工年龄段统计基础数据",
		"group_name": "人事统计",
		"fields": ["employee_name", "department", "date_of_birth", "custom_age", "employment_type"],
		"filters": {"status": "Active"},
	},
	{
		"id": "employee_certificate",
		"report_name": "员工个人证件照存档情况表",
		"description": "员工证件与证照材料存档情况",
		"group_name": "人事统计",
		"fields": ["employee_name", "department", "custom_id_number", "passport_number"],
		"filters": {"status": "Active"},
	},
	{
		"id": "administrative_roster",
		"report_name": "行政花名册",
		"description": "行政常用员工基础信息表",
		"group_name": "行政报表",
		"fields": ["employee_name", "department", "designation", "cell_number", "current_address"],
		"filters": {"status": "Active"},
	},
]

MULTI_RECORD_EXPORT_TABLE_MAP = {
	"教育经历": "education",
	"工作经历": "external_work_history",
	"任职记录": "internal_work_history",
}

MULTI_RECORD_EXPORT_CATEGORIES = [
	{"label": "紧急联系人", "description": "包括紧急联系人姓名、紧急联系人关系、紧急联系人电话等全部字段"},
	{"label": "家庭情况", "description": "包括成员姓名、关系、工作单位等全部字段"},
	{"label": "教育经历", "description": "包括学历、毕业院校、毕业专业等全部字段"},
	{"label": "工作经历", "description": "包括工作单位、职务、工作起止日期等全部字段"},
	{"label": "合同信息", "description": "包括合同编号、合同类型、合同期限等全部字段"},
	{"label": "培训经历", "description": "包括培训名称、单位、培训起止日期等全部字段"},
	{"label": "任职记录", "description": "包括任职起止日期、部门、职务等全部字段"},
	{"label": "兼职信息", "description": "包括员工的兼职部门、兼职岗位"},
	{"label": "奖惩记录", "description": "包括奖惩类型、奖惩日期、奖惩内容等全部字段"},
	{"label": "职称信息", "description": "包括职称名称、职称级别、职称获取时间等全部字段"},
	{"label": "证书/证件", "description": "包括证书/证件类型、证件名称、级别等全部字段"},
	{"label": "与本司员工关系记录", "description": "包括姓名、关系、任职单位等全部字段"},
]

COMPANY_ROSTER_CUSTOM_FIELDS = [
	{
		"category": "在职信息",
		"field_label": "序号",
		"fieldname": "custom_roster_sequence",
		"fieldtype": "Int",
		"description": "来自公司花名册的原始序号",
		"insert_after": "naming_series",
	},
	{
		"category": "在职信息",
		"field_label": "工号",
		"fieldname": "custom_employee_code",
		"fieldtype": "Data",
		"description": "公司内部员工工号",
		"reqd": 1,
		"insert_after": "naming_series",
	},
	{
		"category": "个人信息",
		"field_label": "证件类型",
		"fieldname": "custom_id_type",
		"fieldtype": "Select",
		"options": "身份证\n护照\n港澳通行证\n台胞证\n其他",
		"description": "员工证件类型",
		"insert_after": "passport_number",
	},
	{
		"category": "个人信息",
		"field_label": "籍贯",
		"fieldname": "custom_native_place",
		"fieldtype": "Data",
		"description": "员工籍贯",
		"insert_after": "passport_number",
	},
	{
		"category": "在职信息",
		"field_label": "直间接",
		"fieldname": "custom_direct_indirect",
		"fieldtype": "Select",
		"options": "直接人员\n间接人员",
		"description": "直接人员或间接人员",
		"insert_after": "employment_type",
	},
	{
		"category": "个人信息",
		"field_label": "民族",
		"fieldname": "custom_ethnicity",
		"fieldtype": "Data",
		"description": "员工民族",
		"insert_after": "passport_number",
	},
	{
		"category": "个人信息",
		"field_label": "婚姻状况",
		"fieldname": "custom_marital_status_text",
		"fieldtype": "Select",
		"options": "未\n已\n离异\n丧偶",
		"description": "公司花名册中的婚姻状态",
		"insert_after": "gender",
	},
	{
		"category": "个人信息",
		"field_label": "年龄",
		"fieldname": "custom_age",
		"fieldtype": "Int",
		"description": "按出生日期计算或导入的年龄",
		"insert_after": "date_of_birth",
	},
	{
		"category": "教育信息",
		"field_label": "学历类别",
		"fieldname": "custom_education_category",
		"fieldtype": "Data",
		"description": "学历取得类别",
		"insert_after": "date_of_birth",
	},
	{
		"category": "教育信息",
		"field_label": "学习形式",
		"fieldname": "custom_study_mode",
		"fieldtype": "Data",
		"description": "全日制、函授、自考等学习形式",
		"insert_after": "custom_education_category",
	},
	{
		"category": "教育信息",
		"field_label": "学历",
		"fieldname": "custom_education_level",
		"fieldtype": "Data",
		"description": "最高学历",
		"insert_after": "custom_study_mode",
	},
	{
		"category": "教育信息",
		"field_label": "毕业院校",
		"fieldname": "custom_graduation_school",
		"fieldtype": "Data",
		"description": "毕业院校",
		"insert_after": "custom_education_level",
	},
	{
		"category": "教育信息",
		"field_label": "科系",
		"fieldname": "custom_major",
		"fieldtype": "Data",
		"description": "专业或科系",
		"insert_after": "custom_graduation_school",
	},
	{
		"category": "联系信息",
		"field_label": "交通工具",
		"fieldname": "custom_transport",
		"fieldtype": "Data",
		"description": "员工日常通勤交通工具",
		"insert_after": "current_address",
	},
	{
		"category": "在职信息",
		"field_label": "试用期",
		"fieldname": "custom_probation_months",
		"fieldtype": "Int",
		"description": "试用期月数",
		"insert_after": "date_of_joining",
	},
	{
		"category": "在职信息",
		"field_label": "是否转正",
		"fieldname": "custom_is_confirmed",
		"fieldtype": "Select",
		"options": "是\n否",
		"description": "是否已转正",
		"insert_after": "final_confirmation_date",
	},
	{
		"category": "合同保险",
		"field_label": "合同-签订日期",
		"fieldname": "custom_contract_sign_date",
		"fieldtype": "Date",
		"description": "劳动合同签订日期",
		"insert_after": "contract_end_date",
	},
	{
		"category": "合同保险",
		"field_label": "合同-合同编号",
		"fieldname": "custom_contract_no",
		"fieldtype": "Data",
		"description": "劳动合同编号",
		"insert_after": "custom_contract_sign_date",
	},
	{
		"category": "合同保险",
		"field_label": "合同-签订次数",
		"fieldname": "custom_contract_sign_count",
		"fieldtype": "Int",
		"description": "劳动合同签订次数",
		"insert_after": "custom_contract_no",
	},
	{
		"category": "合同保险",
		"field_label": "保险-社保",
		"fieldname": "custom_social_insurance",
		"fieldtype": "Data",
		"description": "社保缴纳情况",
		"insert_after": "custom_contract_sign_count",
	},
	{
		"category": "工资社保",
		"field_label": "社保参保状态",
		"fieldname": "custom_social_insurance_status",
		"fieldtype": "Select",
		"options": "按社保名单\n参保中\n不参保（已确认）\n停缴",
		"default": "按社保名单",
		"description": "薪资计算默认以社保名单为准；仅“不参保（已确认）”或“停缴”会将当月社保个人和公司承担置零。",
		"insert_after": "custom_social_insurance",
	},
	{
		"category": "工资社保",
		"field_label": "社保起缴日期",
		"fieldname": "custom_social_insurance_start_date",
		"fieldtype": "Date",
		"description": "参保中员工从该日期所在月份起参与社保薪资计算；留空时按社保名单处理。",
		"insert_after": "custom_social_insurance_status",
	},
	{
		"category": "工资社保",
		"field_label": "社保停缴日期",
		"fieldname": "custom_social_insurance_end_date",
		"fieldtype": "Date",
		"description": "停缴日期所在月份之后不再参与社保薪资计算。",
		"insert_after": "custom_social_insurance_start_date",
	},
	{
		"category": "合同保险",
		"field_label": "保险-医保",
		"fieldname": "custom_medical_insurance",
		"fieldtype": "Data",
		"description": "医保缴纳情况",
		"insert_after": "custom_social_insurance",
	},
	{
		"category": "合同保险",
		"field_label": "保险-公积金",
		"fieldname": "custom_housing_fund",
		"fieldtype": "Data",
		"description": "公积金缴纳情况",
		"insert_after": "custom_medical_insurance",
	},
]

COMPANY_ROSTER_FIELD_ORDER = [
	"custom_roster_sequence",
	"custom_employee_code",
	"first_name",
	"department",
	"date_of_joining",
	"cell_number",
	"custom_id_type",
	"passport_number",
	"permanent_address",
	"custom_native_place",
	"designation",
	"employment_type",
	"custom_direct_indirect",
	"custom_ethnicity",
	"custom_marital_status_text",
	"date_of_birth",
	"custom_age",
	"gender",
	"custom_education_category",
	"custom_study_mode",
	"custom_education_level",
	"custom_graduation_school",
	"custom_major",
	"current_address",
	"custom_transport",
	"person_to_be_contacted",
	"emergency_phone_number",
	"custom_probation_months",
	"final_confirmation_date",
	"custom_is_confirmed",
	"custom_contract_sign_date",
	"custom_contract_no",
	"custom_contract_sign_count",
	"contract_end_date",
	"custom_social_insurance",
	"custom_social_insurance_status",
	"custom_social_insurance_start_date",
	"custom_social_insurance_end_date",
	"custom_medical_insurance",
	"custom_housing_fund",
	"company",
	"status",
]

HEADER_FIELD_ALIASES = {
	"员工编号": "custom_employee_code",
	"出生日期": "date_of_birth",
	"出生年月": "date_of_birth",
	"现居住地": "current_address",
	"现住址": "current_address",
	"联系电话": "cell_number",
	"身份证号码": "passport_number",
	"证件号码": "passport_number",
	"证件类型": "custom_id_type",
	"直/间接": "custom_direct_indirect",
	"紧急联系人": "person_to_be_contacted",
	"紧急联系电话": "emergency_phone_number",
	"院校": "custom_graduation_school",
	"现职务": "designation",
	"职位": "designation",
	"职务": "designation",
	"岗位": "designation",
	"上级主管": "reports_to",
	"直接上级": "reports_to",
	"汇报对象": "reports_to",
	"职级": "grade",
	"员工等级": "grade",
	"分支机构": "branch",
	"分公司": "branch",
	"工作性质": "employment_type",
	"雇佣类型": "employment_type",
	"用工类型": "employment_type",
	"转正日期": "final_confirmation_date",
	"合同-结束月份": "contract_end_date",
}

# These labels appear in daily rosters, import files and business discussions.
# They are reserved so a later custom field or alias cannot silently hijack a
# stable mapping that is already consumed by imports and business rules.
RESERVED_EMPLOYEE_BUSINESS_FIELD_KEYS = {
	"工作性质": "employment_type",
	"雇佣类型": "employment_type",
	"用工类型": "employment_type",
}

EMPLOYEE_ROSTER_REQUIRED_COLUMNS = {
	"employee_name": "姓名",
	"custom_employee_code": "工号",
	"department": "部门",
	"designation": "岗位",
	"employment_type": "工作性质",
	"date_of_joining": "入职日期",
	"custom_id_type": "证件类型",
	"passport_number": "证件号码",
	"cell_number": "手机号码",
}

EMPLOYEE_MINIMUM_IMPORT_REQUIRED_COLUMNS = {
	"first_name": "姓名",
	"custom_employee_code": "工号",
	"department": "部门",
	"designation": "岗位",
	"date_of_joining": "入职日期",
	"cell_number": "手机号码",
}

EMPLOYEE_IMPORT_REQUIRED_ALTERNATIVES = {
	"first_name": ("first_name", "employee_name"),
}

FIELD_GOVERNANCE_DEFAULTS = {
	"employment_type": {"aliases": "工作性质\n雇佣类型\n用工类型"},
	"custom_social_insurance_status": {"aliases": "社保参保状态\n参保状态"},
	"custom_social_insurance_start_date": {"aliases": "社保起缴日期\n社保开始缴纳日期"},
	"custom_social_insurance_end_date": {"aliases": "社保停缴日期\n社保停止缴纳日期"},
	"custom_education_category": {"aliases": "学历类别\n学历取得类别", "detail_block": "教育经历", "record_type": "单行资料块"},
	"custom_study_mode": {"aliases": "学习形式\n学习方式", "detail_block": "教育经历", "record_type": "单行资料块"},
	"custom_education_level": {"aliases": "学历\n最高学历\n文化程度", "detail_block": "教育经历", "record_type": "单行资料块"},
	"custom_graduation_school": {"aliases": "毕业院校\n学校\n院校名称", "detail_block": "教育经历", "record_type": "单行资料块"},
	"custom_major": {"aliases": "科系\n专业\n专业或科系", "detail_block": "教育经历", "record_type": "单行资料块"},
	"custom_contract_no": {"aliases": "合同编号\n劳动合同编号", "detail_block": "合同记录", "record_type": "单行资料块"},
	"custom_contract_sign_date": {"aliases": "合同签订日期\n劳动合同签订日期", "detail_block": "合同记录", "record_type": "单行资料块"},
	"custom_contract_sign_count": {"aliases": "合同签订次数\n劳动合同签订次数", "detail_block": "合同记录", "record_type": "单行资料块"},
	"contract_end_date": {"aliases": "合同结束日期\n合同-结束月份", "detail_block": "合同记录", "record_type": "单行资料块"},
	"custom_social_insurance": {"aliases": "社保\n保险-社保", "detail_block": "社保公积金记录", "record_type": "单行资料块"},
	"custom_medical_insurance": {"aliases": "医保\n保险-医保", "detail_block": "社保公积金记录", "record_type": "单行资料块"},
	"custom_housing_fund": {"aliases": "公积金\n保险-公积金", "detail_block": "社保公积金记录", "record_type": "单行资料块"},
}

OPTION_LABEL_MAP = {
	"Active": "激活",
	"Inactive": "非激活",
	"Suspended": "停职",
	"Left": "已离职",
	"Male": "男",
	"Female": "女",
	"Other": "其他",
	"Full-time": "全职",
	"Part-time": "兼职",
	"Intern": "实习生",
	"Contract": "外包",
	"Retainer": "退休返聘",
	"Bank": "银行",
	"Cash": "现金",
	"Cheque": "支票",
}



IMPORT_EXAMPLE_VALUES = {
	"first_name": "示例员工",
	"gender": "男",
	"date_of_birth": "2000-01-01",
	"date_of_joining": "2026-07-01",
	"status": "激活",
	"company": "1 (Demo)",
	"department": "人力资源部",
	"designation": "人事专员",
	"employment_type": "全职",
	"cell_number": "13800000000",
	"personal_email": "employee@example.com",
	"salary_mode": "银行",
}

NON_CONFIGURABLE_FIELDTYPES = {
	"Section Break",
	"Column Break",
	"Tab Break",
	"HTML",
	"Button",
	"Fold",
	"Table",
	"Table MultiSelect",
}

NON_CONFIGURABLE_FIELDNAMES = {
	"employee",
	"image",
	"lft",
	"rgt",
	"old_parent",
	"connections_tab",
}

# Frappe implementation fields are not personnel data.  The company work
# number is the only employee identifier exposed to HR users.
EMPLOYEE_INTERNAL_FIELDNAMES = {"naming_series", "employee_number"}

DEFAULT_FIELD_CATEGORY_BY_SECTION = {
	"basic_details_tab": "个人信息",
	"basic_information": "个人信息",
	"joining_details": "在职信息",
	"employment_details": "在职信息",
	"address_and_contact_tab": "联系信息",
	"contact_details": "联系信息",
	"emergency_contact": "联系信息",
	"salary_tab": "工资社保",
	"salary_details": "工资社保",
	"payroll_cost_centers": "工资社保",
	"personal_details": "个人信息",
	"passport_details": "个人信息",
	"profile": "个人信息",
	"education": "教育信息",
	"educational_qualification": "教育信息",
	"exit_tab": "在职信息",
	"exit": "在职信息",
}

FIELD_OVERRIDES = {
	"first_name": {"category": "个人信息", "field_label": "姓名", "description": "员工真实姓名"},
	"middle_name": {"category": "个人信息", "field_label": "中间名"},
	"last_name": {"category": "个人信息", "field_label": "姓"},
	"employee_name": {"category": "个人信息", "field_label": "姓名", "description": "系统生成的员工完整姓名"},
	"gender": {"category": "个人信息", "field_label": "性别"},
	"date_of_birth": {"category": "个人信息", "field_label": "出生年月"},
	"salutation": {"category": "个人信息", "field_label": "称谓"},
	"date_of_joining": {"category": "在职信息", "field_label": "入职日期"},
	"status": {"category": "在职信息", "field_label": "系统状态", "description": "Frappe 内部单据状态，不作为人事业务口径"},
	"company": {"category": "在职信息", "field_label": "公司", "description": "员工所属公司"},
	"branch": {"category": "在职信息", "field_label": "分支机构（分公司）"},
	"department": {"category": "在职信息", "field_label": "部门", "description": "员工所属部门"},
	"designation": {"category": "在职信息", "field_label": "岗位", "description": "员工当前岗位"},
	"reports_to": {"category": "在职信息", "field_label": "上级主管", "description": "员工汇报对象"},
	"grade": {"category": "在职信息", "field_label": "员工等级"},
	"employment_type": {"category": "在职信息", "field_label": "工作性质", "description": "员工工作性质，统一使用实习、试用、全职、外包、返聘五类。"},
	"user_id": {"category": "联系信息", "field_label": "用户账号"},
	"create_user_automatically": {"category": "联系信息", "field_label": "自动创建用户", "description": "自动为员工创建系统用户"},
	"create_user_permission": {"category": "联系信息", "field_label": "自动创建用户权限"},
	"cell_number": {"category": "联系信息", "field_label": "手机号码", "description": "主要联系电话"},
	"company_email": {"category": "联系信息", "field_label": "公司邮箱"},
	"prefered_email": {"category": "联系信息", "field_label": "首选联系邮箱"},
	"prefered_contact_email": {"category": "联系信息", "field_label": "首选联系邮箱"},
	"personal_email": {"category": "联系信息", "field_label": "个人电子邮件"},
	"unsubscribed": {"category": "联系信息", "field_label": "已退订"},
	"current_address": {"category": "联系信息", "field_label": "当前地址"},
	"current_accommodation_type": {"category": "联系信息", "field_label": "当前地址性质"},
	"permanent_address": {"category": "联系信息", "field_label": "户籍地址"},
	"permanent_accommodation_type": {"category": "联系信息", "field_label": "永久地址类型"},
	"person_to_be_contacted": {"category": "联系信息", "field_label": "紧急联系"},
	"emergency_phone_number": {"category": "联系信息", "field_label": "紧急联系人电话"},
	"relation": {"category": "联系信息", "field_label": "关系"},
	"attendance_device_id": {"category": "在职信息", "field_label": "考勤设备编号"},
	"holiday_list": {"category": "在职信息", "field_label": "假期表"},
	"default_shift": {"category": "在职信息", "field_label": "默认班次"},
	"expense_approver": {"category": "在职信息", "field_label": "费用审批人"},
	"leave_approver": {"category": "在职信息", "field_label": "请假审批人"},
	"shift_request_approver": {"category": "在职信息", "field_label": "班次申请审批人"},
	"final_confirmation_date": {"category": "在职信息", "field_label": "转正日期"},
	"contract_end_date": {"category": "合同保险", "field_label": "合同-结束月份"},
	"notice_number_of_days": {"category": "在职信息", "field_label": "通知期天数"},
	"date_of_retirement": {"category": "在职信息", "field_label": "退休日期"},
	"relieving_date": {"category": "在职信息", "field_label": "离职日期"},
	"resignation_letter_date": {"category": "在职信息", "field_label": "离职申请日期"},
	"reason_for_leaving": {"category": "在职信息", "field_label": "离职原因"},
	"leave_encashed": {"category": "工资社保", "field_label": "已折现假期"},
	"encashment_date": {"category": "工资社保", "field_label": "折现日期"},
	"held_on": {"category": "在职信息", "field_label": "离职面谈日期"},
	"new_workplace": {"category": "在职信息", "field_label": "新工作单位"},
	"salary_mode": {"category": "工资社保", "field_label": "工资发放方式"},
	"payroll_cost_center": {"category": "工资社保", "field_label": "薪资成本中心"},
	"employee_advance_account": {"category": "工资社保", "field_label": "员工预支账户"},
	"bank_name": {"category": "工资社保", "field_label": "银行名称"},
	"bank_ac_no": {"category": "工资社保", "field_label": "银行账号"},
	"iban": {"category": "工资社保", "field_label": "IBAN"},
	"ctc": {"category": "工资社保", "field_label": "年度薪资"},
	"salary_currency": {"category": "工资社保", "field_label": "薪资币种"},
	"pan_number": {"category": "个税申报", "field_label": "税号"},
	"provident_fund_account": {"category": "工资社保", "field_label": "公积金账号"},
	"passport_number": {"category": "个人信息", "field_label": "证件号码", "description": "身份证、护照等证件号码"},
	"valid_upto": {"category": "个人信息", "field_label": "证件有效期"},
	"marital_status": {"category": "个人信息", "field_label": "婚姻状况"},
	"family_background": {"category": "个人信息", "field_label": "家庭背景"},
	"blood_group": {"category": "个人信息", "field_label": "血型"},
	"health_details": {"category": "个人信息", "field_label": "健康信息"},
	"health_insurance_provider": {"category": "合同保险", "field_label": "健康保险机构"},
	"health_insurance_no": {"category": "合同保险", "field_label": "健康保险编号"},
	"passport_date_of_issue": {"category": "个人信息", "field_label": "证件签发日期"},
	"place_of_issue": {"category": "个人信息", "field_label": "证件签发地点"},
	"bio": {"category": "个人信息", "field_label": "履历/求职信"},
}

FIELD_TYPE_MAP = {
	"文本格式": "Data",
	"日期格式": "Date",
	"自定义选项": "Select",
	"长文本格式": "Small Text",
	"Data": "Data",
	"Date": "Date",
	"Select": "Select",
	"Small Text": "Small Text",
	"Check": "Check",
	"Link": "Link",
}

CATEGORY_INSERT_AFTER = {
	"在职信息": "date_of_joining",
	"个人信息": "date_of_birth",
	"联系信息": "emergency_phone_number",
	"教育信息": "date_of_birth",
	"合同保险": "contract_end_date",
	"工资社保": "salary_mode",
	"个税申报": "salary_mode",
	"附件": "bio",
}


def _parse_json(value, fallback):
	if value is None:
		return fallback
	if isinstance(value, str):
		return json.loads(value) if value else fallback
	return value


def _validate_category(category):
	if category not in EMPLOYEE_TEMPLATE_CATEGORIES:
		frappe.throw(_("无效的员工属性分类: {0}").format(category))


def _normalise_fieldtype(fieldtype):
	fieldtype = FIELD_TYPE_MAP.get(fieldtype)
	if not fieldtype:
		frappe.throw(_("不支持的字段类型"))
	return fieldtype


def _get_field_category(field, current_category):
	return FIELD_OVERRIDES.get(field.fieldname, {}).get("category") or current_category or "在职信息"


def _get_field_label(field):
	return FIELD_OVERRIDES.get(field.fieldname, {}).get("field_label") or _(field.label or field.fieldname)


def _get_field_description(field):
	return FIELD_OVERRIDES.get(field.fieldname, {}).get("description") or field.description or _(
		"来自 Employee 标准字段"
	)


def _is_configurable_employee_field(field):
	return bool(
		field.fieldname
		and field.label
		and not field.hidden
		and field.fieldname not in NON_CONFIGURABLE_FIELDNAMES
		and field.fieldtype not in NON_CONFIGURABLE_FIELDTYPES
	)


def get_configurable_employee_fields():
	meta = frappe.get_meta(EMPLOYEE_DOCTYPE)
	current_category = "个人信息"
	fields = []

	for field in meta.fields:
		if field.fieldtype in {"Tab Break", "Section Break"}:
			current_category = DEFAULT_FIELD_CATEGORY_BY_SECTION.get(field.fieldname, current_category)
			continue

		if not _is_configurable_employee_field(field):
			continue

		fields.append(
			{
				"category": _get_field_category(field, current_category),
				"field_label": _get_field_label(field),
				"fieldname": field.fieldname,
				"fieldtype": field.fieldtype,
				"description": _get_field_description(field),
				"options": field.options,
				"insert_after": getattr(field, "insert_after", None),
			}
		)

	return fields


def _make_custom_fieldname(field_label):
	slug = re.sub(r"[^a-z0-9_]+", "_", frappe.scrub(field_label or "").lower()).strip("_")
	if not slug:
		slug = hashlib.sha1((field_label or "").encode("utf-8")).hexdigest()[:10]
	return f"custom_hrms_{slug}"[:140]


def _get_template_doc():
	doc = frappe.get_single(TEMPLATE_DOCTYPE)
	if doc.enabled is None:
		doc.enabled = 1
	_sync_employee_fields(doc)
	_sync_company_roster_fields(doc)
	_apply_field_governance_defaults(doc)
	_apply_company_roster_defaults(doc)
	_apply_employee_required_defaults(doc)
	_apply_employee_internal_field_policy(doc)
	ensure_required_roster_columns(doc)
	_retire_personnel_status_field(doc)

	return doc


def _apply_field_governance_defaults(doc):
	changed = False
	for row in doc.template_items:
		if _apply_field_governance_defaults_to_row(row):
			changed = True
	if changed:
		doc.save(ignore_permissions=True)


def _is_employee_internal_field(fieldname):
	return fieldname in EMPLOYEE_INTERNAL_FIELDNAMES


def _apply_employee_internal_field_policy(doc):
	"""Prevent internal document naming fields from leaking into HR-facing configuration."""
	changed = False
	for row in doc.template_items:
		if not _is_employee_internal_field(row.fieldname):
			continue

		for fieldname in (
			"enabled",
			"required",
			"search_enabled",
			"import_enabled",
			"export_enabled",
			"form_visible",
			"detail_visible",
			"roster_visible",
		):
			if not _template_item_supports_field(fieldname):
				continue
			if _template_row_int(row, fieldname) != 0:
				row.set(fieldname, 0)
				changed = True

		for fieldname in ("aliases", "detail_block", "record_type"):
			if _template_item_supports_field(fieldname) and row.get(fieldname):
				row.set(fieldname, "")
				changed = True
		if _template_item_supports_field("detail_block_order") and _template_row_int(row, "detail_block_order"):
			row.detail_block_order = 0
			changed = True

	if changed:
		doc.save(ignore_permissions=True)


def _template_item_exists(doc, fieldname):
	return any(row.fieldname == fieldname for row in doc.template_items)


def _template_row_value(row, fieldname, default=None):
	value = row.get(fieldname)
	return default if value is None else value


def _template_row_int(row, fieldname, default=0):
	return frappe.utils.cint(_template_row_value(row, fieldname, default))


def _template_item_supports_field(fieldname):
	return bool(frappe.get_meta(TEMPLATE_CHILD_TABLE).get_field(fieldname))


def _template_row_bool(row, fieldname, default=1):
	if not _template_item_supports_field(fieldname):
		return 1 if default else 0
	value = _template_row_value(row, fieldname, default)
	return 1 if frappe.utils.cint(value) else 0


def _field_flag_enabled(row, fieldname, default=1):
	return bool(_template_row_bool(row, fieldname, default))


def _field_aliases_for_row(row):
	aliases = []
	if _template_item_supports_field("aliases"):
		aliases = [alias.strip() for alias in (row.get("aliases") or "").splitlines() if alias.strip()]
	if row.get("field_label"):
		aliases.append(row.get("field_label"))
	return list(dict.fromkeys(aliases))


def _get_detail_block_definitions():
	return [
		{
			"label": label,
			"tab": EMPLOYEE_DETAIL_BLOCK_TABS.get(label, "概览"),
			"description": DEFAULT_DETAIL_BLOCK_DESCRIPTIONS.get(label, ""),
		}
		for label in EMPLOYEE_DETAIL_BLOCK_TABS
	]


def _apply_field_governance_defaults_to_row(row):
	defaults = FIELD_GOVERNANCE_DEFAULTS.get(row.fieldname) or {}
	changed = False
	for fieldname, value in defaults.items():
		if not _template_item_supports_field(fieldname):
			continue
		if row.get(fieldname):
			continue
		row.set(fieldname, value)
		changed = True
	for fieldname, default in (
		("import_enabled", 1),
		("export_enabled", 1),
		("form_visible", 1),
		("detail_visible", 1),
		("roster_visible", 0),
	):
		if _template_item_supports_field(fieldname) and row.get(fieldname) is None:
			row.set(fieldname, default)
			changed = True
	return changed


def _sync_company_roster_fields(doc):
	changed = False
	rows_by_fieldname = {row.fieldname: row for row in doc.template_items}
	meta_fields = _get_employee_meta_field_map()
	required_fieldnames = {fieldname for fieldname, field in meta_fields.items() if getattr(field, "reqd", 0)}
	supports_required = _template_item_supports_field("required")

	for item in COMPANY_ROSTER_CUSTOM_FIELDS:
		custom_field_name = f"{EMPLOYEE_DOCTYPE}-{item['fieldname']}"
		custom_field = {
			"fieldname": item["fieldname"],
			"label": item["field_label"],
			"fieldtype": item["fieldtype"],
			"insert_after": item.get("insert_after") or CATEGORY_INSERT_AFTER.get(item["category"], "date_of_joining"),
			"description": item.get("description"),
		}
		for property_name in ("read_only", "no_copy", "reqd"):
			if item.get(property_name) is not None:
				custom_field[property_name] = item[property_name]
		if item.get("options"):
			custom_field["options"] = item["options"]

		if not frappe.db.exists("Custom Field", custom_field_name):
			create_custom_field(EMPLOYEE_DOCTYPE, custom_field)
			changed = True
		else:
			for fieldname, value in custom_field.items():
				if fieldname == "fieldname":
					continue
				if frappe.db.get_value("Custom Field", custom_field_name, fieldname) != value:
					frappe.db.set_value("Custom Field", custom_field_name, fieldname, value, update_modified=False)
					changed = True

		row = rows_by_fieldname.get(item["fieldname"])
		row_values = {
			"category": item["category"],
			"field_label": item["field_label"],
			"fieldname": item["fieldname"],
			"fieldtype": item["fieldtype"],
			"description": item.get("description"),
			"source": "自定义",
			"enabled": 1,
			"search_enabled": 1 if item["fieldname"] in {"custom_employee_code"} else 0,
			"options": item.get("options"),
			"insert_after": item.get("insert_after") or CATEGORY_INSERT_AFTER.get(item["category"], "date_of_joining"),
		}
		if supports_required:
			row_values["required"] = 1 if item["fieldname"] in required_fieldnames else 0
		if row:
			managed_fields = ["fieldtype", "options", "insert_after", "source"]
			for fieldname in managed_fields:
				value = row_values.get(fieldname)
				if row.get(fieldname) != value:
					row.set(fieldname, value)
					changed = True
		else:
			doc.append("template_items", row_values)
			changed = True

	if changed:
		doc.save(ignore_permissions=True)
		frappe.clear_cache(doctype=EMPLOYEE_DOCTYPE)


def _apply_company_roster_defaults(doc):
	if frappe.db.get_default("hrms_company_roster_defaults_applied_v1"):
		return

	meta_fields = _get_employee_meta_field_map()
	required_fieldnames = {fieldname for fieldname, field in meta_fields.items() if getattr(field, "reqd", 0)}
	enabled_fieldnames = set(COMPANY_ROSTER_FIELD_ORDER) | required_fieldnames
	supports_required = _template_item_supports_field("required")
	search_fieldnames = {
		"custom_employee_code",
		"first_name",
		"cell_number",
		"passport_number",
		"department",
		"designation",
	}
	changed = False

	for row in doc.template_items:
		override = FIELD_OVERRIDES.get(row.fieldname)
		if override:
			for fieldname in ("category", "field_label", "description"):
				if override.get(fieldname) and row.get(fieldname) != override[fieldname]:
					row.set(fieldname, override[fieldname])
					changed = True

		enabled = 1 if row.fieldname in enabled_fieldnames else 0
		required = 1 if row.fieldname in required_fieldnames else 0
		search_enabled = 1 if row.fieldname in search_fieldnames else 0
		if _template_row_int(row, "enabled") != enabled:
			row.enabled = enabled
			changed = True
		if supports_required and _template_row_int(row, "required") != required:
			row.required = required
			changed = True
		if _template_row_int(row, "search_enabled") != search_enabled:
			row.search_enabled = search_enabled
			changed = True

	order_map = {fieldname: index for index, fieldname in enumerate(COMPANY_ROSTER_FIELD_ORDER)}
	category_map = {category: index for index, category in enumerate(EMPLOYEE_TEMPLATE_CATEGORIES)}

	def sort_key(row):
		if row.fieldname in order_map:
			return 0, order_map[row.fieldname]
		return 1, category_map.get(row.category, len(category_map)), row.field_label or row.fieldname

	sorted_rows = sorted(doc.template_items, key=sort_key)
	if [row.name for row in sorted_rows] != [row.name for row in doc.template_items]:
		doc.set("template_items", sorted_rows)
		changed = True

	if changed:
		doc.save(ignore_permissions=True)
	frappe.db.set_default("hrms_company_roster_defaults_applied_v1", "1")


def _apply_employee_required_defaults(doc):
	if frappe.db.get_default("hrms_employee_required_defaults_applied_v1"):
		return
	if not _template_item_supports_field("required"):
		return

	meta_fields = _get_employee_meta_field_map()
	changed = False
	for row in doc.template_items:
		if getattr(meta_fields.get(row.fieldname), "reqd", 0) and not row.get("required"):
			row.required = 1
			changed = True

	if changed:
		doc.save(ignore_permissions=True)
	frappe.db.set_default("hrms_employee_required_defaults_applied_v1", "1")


def _sync_employee_fields(doc):
	changed = False
	rows_by_fieldname = {row.fieldname: row for row in doc.template_items}
	field_items = get_configurable_employee_fields()
	meta_fields = _get_employee_meta_field_map()
	current_system_fieldnames = {item["fieldname"] for item in field_items}
	supports_required = _template_item_supports_field("required")

	for row in list(doc.template_items):
		if row.source == "系统" and row.fieldname not in current_system_fieldnames:
			doc.remove(row)
			changed = True

	rows_by_fieldname = {row.fieldname: row for row in doc.template_items}

	for item in field_items:
		row = rows_by_fieldname.get(item["fieldname"])
		if row:
			if row.source == "系统":
				for fieldname in ("fieldtype", "options", "insert_after"):
					if row.get(fieldname) != item.get(fieldname):
						row.set(fieldname, item.get(fieldname))
						changed = True
				if not row.get("description"):
					row.description = item.get("description")
					changed = True
			continue
		row_values = {
			**item,
			"source": "系统",
			"enabled": 1,
			"search_enabled": 0,
		}
		if supports_required:
			row_values["required"] = 1 if getattr(meta_fields.get(item["fieldname"]), "reqd", 0) else 0
		doc.append("template_items", row_values)
		changed = True

	if changed:
		doc.save(ignore_permissions=True)


def _serialize_item(row):
	return {
		"category": row.category,
		"field_label": row.field_label,
		"fieldname": row.fieldname,
		"fieldtype": row.fieldtype,
		"description": row.description,
		"source": row.source,
		"enabled": _template_row_int(row, "enabled"),
		"required": _template_row_int(row, "required"),
		"search_enabled": _template_row_int(row, "search_enabled"),
		"import_enabled": _template_row_bool(row, "import_enabled", 1),
		"export_enabled": _template_row_bool(row, "export_enabled", 1),
		"form_visible": _template_row_bool(row, "form_visible", 1),
		"detail_visible": _template_row_bool(row, "detail_visible", 1),
		"roster_visible": _template_row_bool(row, "roster_visible", 0),
		"aliases": "\n".join(_field_aliases_for_row(row)),
		"detail_block": _template_row_value(row, "detail_block", ""),
		"detail_block_order": _template_row_int(row, "detail_block_order", 0),
		"record_type": _template_row_value(row, "record_type", "单字段"),
		"options": row.options,
		"insert_after": row.insert_after,
		"idx": row.idx,
	}


def _get_employee_meta_field_map():
	return {field.fieldname: field for field in frappe.get_meta(EMPLOYEE_DOCTYPE).fields if field.fieldname}


def _legacy_personnel_page_sort_key(page_name):
	match = re.search(r"-(\d+)$", page_name or "")
	return (int(match.group(1)) if match else 0, page_name or "")


def _legacy_personnel_page_names(page_name):
	legacy_prefix = LEGACY_PERSONNEL_PAGE_SLUGS.get(page_name)
	if not legacy_prefix:
		return []

	pattern = re.compile(rf"^{re.escape(legacy_prefix)}(?:-\d+)?$")
	return sorted(
		[
			row.name
			for row in frappe.get_all(
				"Page",
				filters={"name": ["like", f"{legacy_prefix}%"]},
				fields=["name"],
			)
			if pattern.match(row.name or "")
		],
		key=_legacy_personnel_page_sort_key,
	)


def _existing_personnel_page_name(page_name):
	return frappe.db.exists("Page", page_name) or frappe.db.get_value("Page", {"page_name": page_name}, "name")


def _cleanup_legacy_personnel_pages(page_name):
	legacy_names = _legacy_personnel_page_names(page_name)
	if not legacy_names:
		return _existing_personnel_page_name(page_name), []

	cleaned = []
	for legacy_name in legacy_names:
		if frappe.db.exists("Page", legacy_name):
			frappe.delete_doc("Page", legacy_name, force=True, ignore_permissions=True)
			cleaned.append(legacy_name)

	return _existing_personnel_page_name(page_name), cleaned


@frappe.whitelist()
def ensure_personnel_pages():
	# Page registration writes Desk metadata.  It runs automatically during
	# migration; interactive calls from ordinary users are intentionally a no-op
	# so navigation cannot be used as a privilege-escalation path.
	if not _can_manage_personnel_pages():
		return {"created": [], "updated": [], "cleaned": [], "skipped": True}

	created = []
	updated = []
	cleaned = []
	for page in PERSONNEL_PAGE_DEFINITIONS:
		page_name = page["name"]
		values = {
			"doctype": "Page",
			"name": page_name,
			"page_name": page_name,
			"title": page["title"],
			"module": "HR",
			"icon": page.get("icon") or "file",
			"standard": "Yes",
			"system_page": 0,
		}
		existing_name, cleaned_names = _cleanup_legacy_personnel_pages(page_name)
		cleaned.extend(cleaned_names)

		if not existing_name:
			page_doc = frappe.get_doc(values)
			page_doc.set("roles", [{"role": role} for role in page.get("roles", [])])
			page_doc.insert(ignore_permissions=True)
			created.append(page_name)
			continue

		for fieldname in ("page_name", "title", "module", "icon", "standard", "system_page"):
			if frappe.db.get_value("Page", existing_name, fieldname) != values[fieldname]:
				frappe.db.set_value("Page", existing_name, fieldname, values[fieldname], update_modified=False)
				if page_name not in updated:
					updated.append(page_name)

		desired_roles = list(page.get("roles", []))
		page_doc = frappe.get_doc("Page", existing_name)
		current_roles = [row.role for row in page_doc.roles]
		if current_roles != desired_roles:
			page_doc.set("roles", [{"role": role} for role in desired_roles])
			page_doc.save(ignore_permissions=True)
			if page_name not in updated:
				updated.append(page_name)

	if created or updated or cleaned:
		frappe.clear_cache()

	return {"created": created, "updated": updated, "cleaned": cleaned}


def ensure_personnel_sidebar_links():
	"""Keep required custom personnel pages visible in the live sidebar.

	Workspace Sidebar records are database documents, so an exported JSON file
	does not always replace an already-customised sidebar during migration.
	"""
	if not frappe.db.exists("Workspace Sidebar", "Personnel"):
		return {"updated": False, "skipped": True}

	sidebar = frappe.get_doc("Workspace Sidebar", "Personnel")
	link = {
		"child": 1,
		"collapsible": 0,
		"indent": 0,
		"keep_closed": 0,
		"label": "跨部门支援",
		"link_to": "cross-department-support",
		"link_type": "Page",
		"open_in_new_tab": 0,
		"show_arrow": 0,
		"type": "Link",
	}
	items = [
		row
		for row in sidebar.items
		if row.get("link_to") not in {"cross-department-support", "Cross Department Support Capability"}
		and row.get("label") != "跨部门支援"
	]
	insert_after = next(
		(
			index
			for index, row in enumerate(items)
			if row.get("link_to") in {"Employee Skill Map", "Employee Training"}
		),
		len(items) - 1,
	)
	items.insert(insert_after + 1, link)
	sidebar.set("items", items)
	sidebar.flags.ignore_links = True
	sidebar.save(ignore_permissions=True)
	frappe.clear_cache()
	return {"updated": True}


def ensure_employee_work_nature_setup():
	"""Keep the public employee schema limited to the single work-nature field."""
	doc = _get_template_doc()
	_retire_personnel_status_field(doc)
	return {"updated": True}


def _property_history_doc_filters(employee=None, department=None, company=None):
	filters = {"docstatus": ["!=", 2]}
	if employee:
		filters["employee"] = employee
	if department:
		filters["department"] = department
	if company:
		filters["company"] = company
	return filters


def _load_property_history_children(source_doctype, parent_names):
	if not parent_names:
		return {}

	children = frappe.get_all(
		"Employee Property History",
		filters={"parenttype": source_doctype, "parent": ["in", parent_names]},
		fields=["parent", "property", "current", "new", "fieldname", "idx"],
		order_by="idx asc",
	)
	grouped = {}
	for row in children:
		grouped.setdefault(row.parent, []).append(
			{
				"property": row.property,
				"current": row.current,
				"new": row.new,
				"fieldname": row.fieldname,
			}
		)
	return grouped


def _property_history_matches_search(row, search):
	if not search:
		return True
	search = str(search).strip().lower()
	if not search:
		return True
	values = [
		row.get("employee"),
		row.get("employee_name"),
		row.get("department"),
		row.get("company"),
		row.get("source_label"),
		row.get("source_name"),
	]
	for change in row.get("changes") or []:
		values.extend([change.get("property"), change.get("current"), change.get("new"), change.get("fieldname")])
	return search in " ".join(str(value or "").lower() for value in values)


@frappe.whitelist()
def get_employee_property_history(employee: str | None = None, department: str | None = None, company: str | None = None, search: str | None = None, limit_start: int = 0, limit_page_length: int = 50):
	limit_start = max(frappe.utils.cint(limit_start), 0)
	limit_page_length = frappe.utils.cint(limit_page_length) or 50
	limit_page_length = max(min(limit_page_length, 200), 1)
	records = []

	for source in PROPERTY_HISTORY_SOURCES:
		source_doctype = source["doctype"]
		if not frappe.has_permission(source_doctype, "read"):
			continue

		date_field = source["date_field"]
		parent_rows = frappe.get_list(
			source_doctype,
			filters=_property_history_doc_filters(employee=employee, department=department, company=company),
			fields=["name", "employee", "employee_name", "department", "company", date_field, "docstatus", "modified"],
			order_by=f"{date_field} desc, modified desc",
			limit_page_length=1000,
		)
		children_by_parent = _load_property_history_children(source_doctype, [row.name for row in parent_rows])

		for parent in parent_rows:
			record = {
				"source_doctype": source_doctype,
				"source_label": source["label"],
				"source_name": parent.name,
				"employee": parent.employee,
				"employee_name": parent.employee_name,
				"department": parent.department,
				"company": parent.company,
				"effective_date": parent.get(date_field),
				"docstatus": parent.docstatus,
				"modified": parent.modified,
				"changes": children_by_parent.get(parent.name, []),
			}
			if _property_history_matches_search(record, search):
				records.append(record)

	records.sort(
		key=lambda row: (
			str(row.get("effective_date") or date.min),
			str(row.get("modified") or datetime.min),
		),
		reverse=True,
	)
	total = len(records)
	rows = records[limit_start : limit_start + limit_page_length]

	return {
		"rows": rows,
		"total": total,
		"limit_start": limit_start,
		"limit_page_length": limit_page_length,
		"sources": PROPERTY_HISTORY_SOURCES,
		"field_contract": PROPERTY_HISTORY_FIELD_CONTRACT,
	}


def _get_employee_import_fields(doc):
	meta_fields = _get_employee_meta_field_map()
	rows = []

	for row in doc.template_items:
		if _is_employee_internal_field(row.fieldname):
			continue
		if row.fieldname not in meta_fields:
			continue
		is_minimum_import_field = _is_employee_import_required_field(row.fieldname, row)
		if not row.get("enabled") and not is_minimum_import_field:
			continue
		if not _field_flag_enabled(row, "import_enabled", 1) and not is_minimum_import_field:
			continue

		meta_field = meta_fields[row.fieldname]
		rows.append(
			{
				"field_label": row.field_label or meta_field.label or row.fieldname,
				"fieldname": row.fieldname,
				"fieldtype": row.fieldtype or meta_field.fieldtype,
				"description": row.description or meta_field.description or "",
				"category": row.category,
				"required": 1 if is_minimum_import_field else 0,
				"options": row.options or meta_field.options or "",
				"enabled": _template_row_int(row, "enabled"),
				"aliases": "\n".join(_field_aliases_for_row(row)),
				"import_enabled": _template_row_bool(row, "import_enabled", 1),
				"export_enabled": _template_row_bool(row, "export_enabled", 1),
				"form_visible": _template_row_bool(row, "form_visible", 1),
				"detail_visible": _template_row_bool(row, "detail_visible", 1),
				"roster_visible": _template_row_bool(row, "roster_visible", 0),
				"detail_block": _template_row_value(row, "detail_block", ""),
				"detail_block_order": _template_row_int(row, "detail_block_order", 0),
				"record_type": _template_row_value(row, "record_type", "单字段"),
			}
		)

	def sort_key(field):
		company_order = {fieldname: index for index, fieldname in enumerate(COMPANY_ROSTER_FIELD_ORDER)}
		if field["fieldname"] in company_order:
			return 0, company_order[field["fieldname"]]

		category = field["category"]
		category_index = (
			EMPLOYEE_TEMPLATE_CATEGORIES.index(category)
			if category in EMPLOYEE_TEMPLATE_CATEGORIES
			else len(EMPLOYEE_TEMPLATE_CATEGORIES)
		)
		return 1, category_index, not field["required"], field["field_label"]

	rows.sort(key=sort_key)
	return rows


def _get_employee_export_fields(doc):
	meta_fields = _get_employee_meta_field_map()
	rows = []
	for row in doc.template_items:
		if _is_employee_internal_field(row.fieldname):
			continue
		if row.fieldname not in meta_fields:
			continue
		if not row.get("enabled"):
			continue
		if not _field_flag_enabled(row, "export_enabled", 1):
			continue
		field = _serialize_item(row)
		field["required"] = 1 if row.get("required") else 0
		rows.append(field)

	category_order = {category: index for index, category in enumerate(EMPLOYEE_TEMPLATE_CATEGORIES)}
	rows.sort(key=lambda field: (category_order.get(field["category"], len(category_order)), field["idx"]))
	return rows


def _display_option(value):
	return OPTION_LABEL_MAP.get(value, value)


def _get_field_options(field):
	options = []
	if field.get("fieldtype") == "Select":
		options = [option.strip() for option in (field.get("options") or "").splitlines() if option.strip()]
	if field.get("fieldname") == "gender" and not options:
		options = ["Male", "Female", "Other"]
	if field.get("fieldname") == "employment_type" and not options:
		options = ["Full-time", "Part-time", "Intern", "Contract", "Retainer"]
	if field.get("fieldname") == "status" and not options:
		options = ["Active", "Inactive", "Suspended", "Left"]
	return [_display_option(option) for option in options]


def _example_value(field):
	if field["fieldname"] in IMPORT_EXAMPLE_VALUES:
		return IMPORT_EXAMPLE_VALUES[field["fieldname"]]
	if field["fieldtype"] in {"Date", "Datetime"}:
		return "2026-07-01"
	if field["fieldtype"] in {"Int", "Float", "Currency"}:
		return 0
	options = _get_field_options(field)
	if options:
		return options[0]
	return ""


def _normalise_header(value):
	value = re.sub(r"\s+", "", str(value or ""))
	return value.replace("*", "").replace("＊", "").strip().lower()


def _field_lookup(fields):
	lookup = {}
	fields_by_name = {field["fieldname"]: field for field in fields}
	for field in fields:
		aliases = []
		if field.get("aliases"):
			aliases.extend(alias.strip() for alias in str(field.get("aliases")).splitlines() if alias.strip())
		for key in (field["field_label"], field["fieldname"], *aliases):
			lookup[_normalise_header(key)] = field
	for header, fieldname in HEADER_FIELD_ALIASES.items():
		field = fields_by_name.get(fieldname)
		if field:
			lookup[_normalise_header(header)] = field
	return lookup


def _is_employee_import_required_field(fieldname, field=None):
	return fieldname in EMPLOYEE_MINIMUM_IMPORT_REQUIRED_COLUMNS


def _employee_import_required_field_satisfied(fieldname, matched_fieldnames):
	accepted_fieldnames = EMPLOYEE_IMPORT_REQUIRED_ALTERNATIVES.get(fieldname, (fieldname,))
	return any(accepted_fieldname in matched_fieldnames for accepted_fieldname in accepted_fieldnames)


def _column_name_to_index(cell_reference):
	match = re.match(r"([A-Z]+)", cell_reference or "")
	if not match:
		return None

	index = 0
	for char in match.group(1):
		index = index * 26 + ord(char) - ord("A") + 1
	return index


def _read_xlsx_rows_from_xml(content, sheet_index=0):
	namespaces = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
	with zipfile.ZipFile(BytesIO(content)) as archive:
		shared_strings = []
		if "xl/sharedStrings.xml" in archive.namelist():
			shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
			for item in shared_root.findall(".//m:si", namespaces):
				text_parts = [node.text or "" for node in item.findall(".//m:t", namespaces)]
				shared_strings.append("".join(text_parts))

		sheet_name = f"xl/worksheets/sheet{sheet_index + 1}.xml"
		sheet_root = ElementTree.fromstring(archive.read(sheet_name))
		rows = []
		for row in sheet_root.findall(".//m:sheetData/m:row", namespaces):
			values = []
			for cell in row.findall("m:c", namespaces):
				column_index = _column_name_to_index(cell.attrib.get("r"))
				if column_index:
					while len(values) < column_index - 1:
						values.append("")

				value_node = cell.find("m:v", namespaces)
				inline_node = cell.find("m:is/m:t", namespaces)
				value = ""
				if inline_node is not None:
					value = inline_node.text or ""
				elif value_node is not None:
					value = value_node.text or ""
					if cell.attrib.get("t") == "s" and value.isdigit() and int(value) < len(shared_strings):
						value = shared_strings[int(value)]
				values.append(value)
			rows.append(values)
		return rows


def _read_xlsx_first_sheet_rows(content):
	try:
		from openpyxl import load_workbook

		workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
		sheet = workbook.worksheets[0]
		return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
	except Exception:
		return _read_xlsx_rows_from_xml(content)


def _make_employee_workbook(fields):
	from openpyxl import Workbook
	from openpyxl.comments import Comment
	from openpyxl.styles import Font, PatternFill
	from openpyxl.utils import get_column_letter
	from hrms.utils.export_watermark import save_workbook_with_logo_watermark

	workbook = Workbook()
	data_sheet = workbook.active
	data_sheet.title = "员工花名册"
	header_fill = PatternFill("solid", fgColor="F6F7F9")
	required_font = Font(color="D93025", bold=True)
	normal_font = Font(bold=True)

	for column, field in enumerate(fields, start=1):
		cell = data_sheet.cell(row=1, column=column, value=f"{field['field_label']}{' *' if field['required'] else ''}")
		cell.fill = header_fill
		cell.font = required_font if field["required"] else normal_font
		cell.comment = Comment(f"Employee 字段：{field['fieldname']}\n{field.get('description') or ''}", "HRMS")
		data_sheet.cell(row=2, column=column, value=_example_value(field))
		data_sheet.column_dimensions[get_column_letter(column)].width = max(14, min(28, len(field["field_label"]) + 8))

	instruction_sheet = workbook.create_sheet("说明")
	instructions = [
		["填写说明"],
		["1. 导入的 Excel 中表头名称建议和“员工花名册”字段名称一致；系统也支持“合同/保险”这类两行复合表头。"],
		["2. 带 * 的字段为必填字段，已在表头中标红并加粗。"],
		["3. 选项字段仅可填写模板支持的选项值，详细规则见“枚举字段”工作表。"],
		["4. 请不要在单元格中使用公式，导入时会读取单元格最终文本。"],
		["5. 公司、状态、单据编号模板如未在表格中提供，系统会使用默认值。"],
	]
	for row_index, row in enumerate(instructions, start=1):
		instruction_sheet.append(row)
		if row_index == 1:
			instruction_sheet.cell(row=row_index, column=1).font = Font(bold=True)
	instruction_sheet.column_dimensions["A"].width = 90

	enum_sheet = workbook.create_sheet("枚举字段")
	enum_sheet.append(["字段名称", "可选值"])
	enum_sheet["A1"].font = enum_sheet["B1"].font = Font(bold=True)
	for field in fields:
		options = _get_field_options(field)
		if options:
			enum_sheet.append([field["field_label"], "、".join(options)])
	enum_sheet.column_dimensions["A"].width = 24
	enum_sheet.column_dimensions["B"].width = 80

	output = BytesIO()
	save_workbook_with_logo_watermark(workbook, output)
	return output.getvalue()


def build_employee_import_template():
	doc = _get_template_doc()
	fields = _get_employee_import_fields(doc)
	return _make_employee_workbook(fields)


def _get_file_content(file_url):
	name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not name:
		frappe.throw(_("未找到上传文件"))
	file_doc = frappe.get_doc("File", name)
	content = file_doc.get_content()
	if isinstance(content, str):
		content = content.encode()
	return content


def _match_uploaded_headers(headers, fields):
	lookup = _field_lookup(fields)
	matches = []
	matched_fieldnames = set()
	for fallback_index, header in enumerate(headers):
		header_label = _header_label(header)
		field = lookup.get(_normalise_header(header_label))
		if field:
			matched_fieldnames.add(field["fieldname"])
		matches.append(
			{
				"header": header_label,
				"column_index": _header_column_index(header, fallback_index),
				"field_label": field["field_label"] if field else "",
				"fieldname": field["fieldname"] if field else "",
				"matched": bool(field),
			}
		)

	missing_required = [
		{"field_label": field["field_label"], "fieldname": field["fieldname"]}
		for field in fields
		if _is_employee_import_required_field(field["fieldname"], field)
		and not _employee_import_required_field_satisfied(field["fieldname"], matched_fieldnames)
		and field["fieldname"] not in IMPORT_EXAMPLE_VALUES
	]
	return matches, missing_required


def _is_blank_value(value):
	return value is None or not str(value).strip()


def _count_non_empty_values(row):
	return sum(1 for value in row if not _is_blank_value(value))


def _clean_header_label(value):
	value = str(value or "").replace("\r", "").replace("\n", "")
	return re.sub(r"\s+", " ", value).strip()


def _header_label(header):
	if isinstance(header, dict):
		return header.get("header") or header.get("label") or ""
	return header


def _header_column_index(header, fallback_index):
	if isinstance(header, dict) and header.get("column_index") is not None:
		return int(header.get("column_index"))
	return fallback_index


def _build_headers_from_rows(base_row, next_row):
	group_header_labels = {"合同", "保险"}
	base_row = list(base_row)
	next_row = list(next_row)
	has_group_header = any(_clean_header_label(value) in group_header_labels for value in base_row)
	max_columns = max(len(base_row), len(next_row))

	headers = []
	last_group = ""
	for column in range(max_columns):
		parent = _clean_header_label(base_row[column]) if column < len(base_row) else ""
		child = _clean_header_label(next_row[column]) if column < len(next_row) else ""

		if parent:
			last_group = parent

		if has_group_header and child and last_group in group_header_labels:
			header = f"{last_group}-{child}"
		else:
			header = parent

		if header:
			headers.append({"header": header, "column_index": column})

	return headers, has_group_header


def _score_header_candidate(headers, fields):
	if not fields:
		return len(headers)

	lookup = _field_lookup(fields)
	matched_fieldnames = set()
	for header in headers:
		field = lookup.get(_normalise_header(_header_label(header)))
		if field:
			matched_fieldnames.add(field["fieldname"])

	required_matches = sum(1 for field in fields if field["required"] and field["fieldname"] in matched_fieldnames)
	company_roster_matches = sum(1 for fieldname in COMPANY_ROSTER_FIELD_ORDER if fieldname in matched_fieldnames)
	return len(matched_fieldnames) * 10 + required_matches * 3 + company_roster_matches


def _detect_uploaded_headers(rows, fields=None):
	if not rows:
		return [], 0

	candidate_count = min(10, len(rows))
	if fields:
		best_headers = []
		best_data_start_index = 0
		best_score = -1
		for index in range(candidate_count):
			base_row = list(rows[index])
			next_row = list(rows[index + 1]) if index + 1 < len(rows) else []
			headers, has_group_header = _build_headers_from_rows(base_row, next_row)
			score = _score_header_candidate(headers, fields)
			if score > best_score:
				best_headers = headers
				best_data_start_index = index + (2 if has_group_header else 1)
				best_score = score

		if best_score > 0:
			return best_headers, best_data_start_index

	header_index = max(range(candidate_count), key=lambda index: _count_non_empty_values(rows[index]))
	base_row = list(rows[header_index])
	next_row = list(rows[header_index + 1]) if header_index + 1 < len(rows) else []
	headers, has_group_header = _build_headers_from_rows(base_row, next_row)

	data_start_index = header_index + (2 if has_group_header else 1)
	return headers, data_start_index


def _count_data_rows(rows, start_index=1):
	count = 0
	for row in rows[start_index:]:
		if any(not _is_blank_value(value) for value in row):
			count += 1
	return count


def _get_uploaded_roster_context(file_url):
	doc = _get_template_doc()
	fields = _get_employee_import_fields(doc)
	content = _get_file_content(file_url)
	rows = _read_xlsx_first_sheet_rows(content)
	headers, data_start_index = _detect_uploaded_headers(rows, fields)
	matches, missing_required = _match_uploaded_headers(headers, fields)
	return {
		"fields": fields,
		"rows": rows,
		"headers": headers,
		"matches": matches,
		"missing_required": missing_required,
		"data_start_index": data_start_index,
	}


def _apply_manual_header_mappings(context, manual_mappings=None):
	manual_mappings = _parse_json(manual_mappings, {}) or {}
	if not manual_mappings:
		return context

	fields_by_name = {field["fieldname"]: field for field in context["fields"]}
	matched_fieldnames = set()
	for match in context["matches"]:
		column_key = str(match.get("column_index"))
		fieldname = manual_mappings.get(column_key) or manual_mappings.get(match.get("header"))
		field = fields_by_name.get(fieldname)
		if field:
			match["field_label"] = field["field_label"]
			match["fieldname"] = field["fieldname"]
			match["matched"] = True
		if match.get("fieldname"):
			matched_fieldnames.add(match["fieldname"])

	context["missing_required"] = [
		{"field_label": field["field_label"], "fieldname": field["fieldname"]}
		for field in context["fields"]
		if _is_employee_import_required_field(field["fieldname"], field)
		and not _employee_import_required_field_satisfied(field["fieldname"], matched_fieldnames)
		and field["fieldname"] not in IMPORT_EXAMPLE_VALUES
	]
	return context


def _get_rows_by_category(doc):
	fields = [_serialize_item(row) for row in doc.template_items if not _is_employee_internal_field(row.fieldname)]
	categories = []
	for category in EMPLOYEE_TEMPLATE_CATEGORIES:
		categories.append(
			{
				"label": category,
				"fields": [field for field in fields if field["category"] == category],
			}
		)
	return categories


@frappe.whitelist()
def get_employee_field_template():
	doc = _get_template_doc()
	fields = [_serialize_item(row) for row in doc.template_items if not _is_employee_internal_field(row.fieldname)]
	return {
		"enabled": int(doc.enabled or 0),
		"categories": _get_rows_by_category(doc),
		"fields": fields,
	}


def _get_employee_field_center_payload():
	doc = _get_template_doc()
	fields = [_serialize_item(row) for row in doc.template_items if not _is_employee_internal_field(row.fieldname)]
	fields_by_block = {}
	for field in fields:
		block = field.get("detail_block")
		if not block:
			continue
		fields_by_block.setdefault(block, []).append(field)

	detail_blocks = []
	for block in _get_detail_block_definitions():
		block_fields = sorted(
			fields_by_block.get(block["label"], []),
			key=lambda field: (field.get("detail_block_order") or 0, field.get("idx") or 0),
		)
		detail_blocks.append({**block, "fields": block_fields, "enabled": 1})

	return {
		"enabled": int(doc.enabled or 0),
		"categories": _get_rows_by_category(doc),
		"fields": fields,
		"detail_blocks": detail_blocks,
		"import_mappings": [
			{
				"fieldname": field["fieldname"],
				"field_label": field["field_label"],
				"aliases": field.get("aliases") or "",
				"import_enabled": field.get("import_enabled", 1),
			}
			for field in fields
			if field.get("enabled")
		],
		"export_templates": _get_default_export_template_summaries(),
		"base_data_modules": _get_base_data_module_summaries(),
		"record_types": _get_record_type_summaries(fields),
	}


def _get_default_export_template_summaries():
	return [
		{"label": report["report_name"], "group": report["group_name"], "fields": report.get("fields") or []}
		for report in DEFAULT_EMPLOYEE_REPORTS
	]


def _get_base_data_module_summaries():
	"""Return the dictionaries a HR administrator can safely maintain.

	The dictionary is deliberately described together with its consuming field.
	A plain list of DocTypes made it far too easy to add a value without knowing
	which screen or rule would start using it.
	"""
	definitions = [
		{
			"label": "公司",
			"doctype": "Company",
			"linked_fields": ["Employee.company"],
			"description": "定义员工、部门、考勤和薪资的数据归属公司。",
			"scope": "全系统数据隔离边界；不是普通下拉选项。",
			"risk": "改名或停用前须评估关联数据与权限范围。",
		},
		{
			"label": "分支机构",
			"doctype": "Branch",
			"linked_fields": ["Employee.branch"],
			"description": "定义分公司、厂区或业务分支。",
			"scope": "员工档案、组织归属和部分报表筛选。",
			"risk": "新增后可在员工档案选择；历史记录不自动迁移。",
		},
		{
			"label": "部门",
			"doctype": "Department",
			"linked_fields": ["Employee.department"],
			"description": "定义组织层级及员工归属部门。",
			"scope": "员工档案、组织架构、招聘、审批与数据权限。",
			"risk": "调整层级或负责人会影响组织、审批和权限。",
		},
		{
			"label": "岗位",
			"doctype": "Designation",
			"linked_fields": ["Employee.designation"],
			"description": "定义员工当前职位/岗位名称。",
			"scope": "员工档案、招聘职位、任职异动和人事报表。",
			"risk": "新增后可选用；不要用改名代替岗位异动。",
		},
		{
			"label": "职级",
			"doctype": "Employee Grade",
			"linked_fields": ["Employee.grade"],
			"description": "定义员工等级或职级口径。",
			"scope": "员工档案、薪资定级和人事统计。",
			"risk": "如已被薪资规则引用，变更前请先做规则预览。",
		},
		{
			"label": "工作性质",
			"doctype": "Employment Type",
			"linked_fields": ["Employee.employment_type", "Job Opening.employment_type"],
			"description": "定义员工工作性质：在职·正式、在职·试用期、退休返聘、待离职、离职。",
			"scope": "员工档案、招聘职位，以及可能按用工性质配置的薪资或合同规则。",
			"risk": "仅使用实习、试用、全职、外包、返聘五类；不要新增、改名或删除已有取值。",
		},
	]

	for item in definitions:
		item["record_count"] = _model_catalog_record_count(item["doctype"])
		item["values"] = _get_base_data_value_samples(item["doctype"])
	return definitions


def _get_base_data_value_samples(doctype, limit=8):
	"""Read a small, safe preview for a dictionary card, not its full dataset."""
	if not frappe.db.exists("DocType", doctype):
		return []
	try:
		return [
			row.name
			for row in frappe.get_all(
				doctype,
				fields=["name"],
				order_by="modified desc",
				limit_page_length=limit,
			)
		]
	except Exception:
		# An optional upstream DocType must not prevent the settings page loading.
		return []


def _get_record_type_summaries(fields):
	record_types = []
	for block in _get_detail_block_definitions():
		block_fields = [
			field for field in fields if field.get("detail_block") == block["label"] and field.get("record_type") == "多行记录"
		]
		record_types.append(
			{
				"label": block["label"],
				"tab": block["tab"],
				"record_type": "多行记录" if block_fields else "单行资料块",
				"fields": [field["fieldname"] for field in block_fields],
			}
		)
	return record_types


@frappe.whitelist()
def get_employee_field_center():
	_require_hr_settings_manager()
	return _get_employee_field_center_payload()


@frappe.whitelist()
def get_hr_settings_center():
	_require_hr_settings_manager()
	field_center = _get_employee_field_center_payload()
	return {
		"modules": [
			{"label": "字段管理中心", "key": "field-center", "count": len(field_center["fields"])},
			{"label": "员工属性设置", "key": "staff-attributes", "count": len(field_center["fields"])},
			{"label": "字段别名配置", "key": "field-aliases", "count": len(field_center["import_mappings"])},
			{"label": "导入映射设置", "key": "import-mapping", "count": len(field_center["import_mappings"])},
			{"label": "详情资料块设置", "key": "detail-blocks", "count": len(field_center["detail_blocks"])},
			{"label": "导出模板设置", "key": "export-templates", "count": len(field_center["export_templates"])},
			{"label": "基础资料设置", "key": "base-data", "count": len(field_center["base_data_modules"])},
			{"label": "多行记录类型", "key": "record-types", "count": len(field_center["record_types"])},
		],
		"field_center": field_center,
	}


@frappe.whitelist()
def get_hrms_access_center():
	"""Return a deliberately small account/role summary for the admin landing page.

	Passwords are never queried or returned.  The only credential-related text is
	the documented initial local development account, so administrators understand
	why a user's existing password cannot appear in this screen.
	"""
	_require_system_manager()
	users = frappe.get_all(
		"User",
		fields=["name", "full_name", "user_type", "enabled", "last_login"],
		order_by="name asc",
		ignore_permissions=True,
	)
	user_names = [user.name for user in users]
	roles_by_user = {name: [] for name in user_names}
	if user_names:
		role_rows = frappe.get_all(
			"Has Role",
			filters={"parenttype": "User", "parent": ["in", user_names]},
			fields=["parent", "role"],
			order_by="parent asc, idx asc",
			ignore_permissions=True,
		)
		for row in role_rows:
			roles_by_user.setdefault(row.parent, []).append(row.role)

	role_users = {}
	for user_name, assigned_roles in roles_by_user.items():
		for role in assigned_roles:
			role_users.setdefault(role, set()).add(user_name)

	role_doctypes = {}
	for permission_doctype in ("DocPerm", "Custom DocPerm"):
		for row in frappe.get_all(
			permission_doctype,
			fields=["role", "parent"],
			ignore_permissions=True,
		):
			if row.role and row.parent:
				role_doctypes.setdefault(row.role, set()).add(row.parent)

	primary_roles = ("System Manager", "HR Manager", "HR User", "Employee", "Guest")
	role_labels = {
		"Administrator": "超级管理员",
		"System Manager": "系统管理员",
		"HR Manager": "人资管理员",
		"HR User": "人事专员",
		"Employee": "员工",
		"Employee Self Service": "员工自助",
		"Expense Approver": "费用审批人",
		"Leave Approver": "请假审批人",
		"Interviewer": "面试官",
		"Guest": "访客",
		"Sales Master Manager": "销售主数据管理员",
		"Maintenance Manager": "维护管理员",
	}
	accounts = []
	for user in users:
		roles = roles_by_user.get(user.name, [])
		accounts.append(
			{
				"user": user.name,
				"full_name": user.full_name or user.name,
				"user_type": user.user_type or "System User",
				"enabled": int(user.enabled or 0),
				"roles": [role for role in roles if role in primary_roles] or roles[:1],
				"assigned_roles": roles,
				"assigned_role_labels": [role_labels.get(role, _(role)) for role in roles],
				"role_count": len(roles),
				"last_login": user.last_login,
			}
		)

	role_descriptions = {
		"System Manager": "系统配置与开发管理角色，可维护账户、角色、权限、页面和数据模型。",
		"HR Manager": "人资管理角色，用于员工档案、人事配置、考勤和薪资等管理操作。",
		"HR User": "人事经办角色，用于日常录入、查询和处理人资业务，不应拥有系统开发权限。",
		"Employee": "员工自助角色，用于查看和办理与本人相关的员工业务。",
		"Employee Self Service": "员工自助扩展角色，用于请假、考勤、费用等个人业务入口。",
		"Guest": "匿名访客角色，不用于后台管理账户。",
	}
	roles = []
	for role in frappe.get_all(
		"Role",
		fields=["name", "is_custom", "disabled", "desk_access"],
		order_by="disabled asc, is_custom desc, name asc",
		ignore_permissions=True,
	):
		configured_doctypes = sorted(role_doctypes.get(role.name, set()))
		assigned_users = role_users.get(role.name, set())
		user_count = len(assigned_users - {"Administrator"})
		roles.append(
			{
				**role,
				"label": role_labels.get(role.name, _(role.name)),
				"user_count": user_count,
				"administrator_assigned": int("Administrator" in assigned_users),
				"permission_doctype_count": len(configured_doctypes),
				"permission_doctypes": [
					{"name": doctype, "label": _(doctype)} for doctype in configured_doctypes[:6]
				],
				"is_project_used": int(
					bool(user_count)
					or role.name in role_descriptions
					or role.name in {"Leave Approver", "Expense Approver", "Interviewer"}
				),
				"description": role_descriptions.get(
					role.name,
					"框架或业务模块提供的角色；只有分配给账户并配置业务单据权限后才会实际生效。",
				),
			}
		)

	return {
		"accounts": accounts,
		"roles": roles,
		"initial_local_account": {
			"user": "Administrator",
			"password": "admin",
			"notice": "仅为 README 中的本地开发初始凭据；无法查看或恢复任何账户的当前密码，部署前必须修改。",
		},
	}


def _configuration_record_count(doctype):
	meta = frappe.get_meta(doctype)
	if meta.issingle:
		return 1
	return frappe.db.count(doctype)


FIELD_RUNTIME_REFERENCES = {
	"employment_type": {
		"meaning": "员工工作性质，统一显示为在职·正式、在职·试用期、退休返聘、待离职、离职五类。",
		"scope": "员工档案、招聘职位、薪资条件和合同规则可按此字段引用。",
		"managed_by": "引用“工作性质（Employment Type）”基础字典；新增值后即可选择。",
	},
	"department": {
		"meaning": "员工当前组织归属。",
		"scope": "员工档案、组织架构、审批、招聘与数据权限。",
		"managed_by": "引用“部门”基础字典；组织调整请走人事异动。",
	},
	"designation": {
		"meaning": "员工当前岗位/职位。",
		"scope": "员工档案、招聘、任职异动和人事报表。",
		"managed_by": "引用“岗位（Designation）”基础字典。",
	},
	"grade": {
		"meaning": "员工等级或职级。",
		"scope": "员工档案、薪资定级和人事统计。",
		"managed_by": "引用“职级（Employee Grade）”基础字典。",
	},
	"company": {
		"meaning": "员工数据所属公司。",
		"scope": "权限、部门、考勤、薪资和报表的数据隔离边界。",
		"managed_by": "引用“公司”基础字典；不可当作一般分类随意修改。",
	},
}

# Only these small controlled dictionaries are previewed in the developer
# centre. Other Link fields can point at Employee, User, Account, etc.;
# rendering their values here would be both noisy and an unnecessary data
# exposure on a configuration page.
DEVELOPER_CENTER_PREVIEW_DICTIONARIES = {
	"Company",
	"Branch",
	"Department",
	"Designation",
	"Employee Grade",
	"Employment Type",
}


def _field_runtime_reference(field):
	"""Describe a field in business language and expose its actual UI reach."""
	field = dict(field)
	fieldname = field.get("fieldname")
	known_reference = FIELD_RUNTIME_REFERENCES.get(fieldname, {})
	used_in = []
	if field.get("form_visible"):
		used_in.append("员工档案")
	if field.get("detail_visible"):
		used_in.append("档案详情")
	if field.get("roster_visible"):
		used_in.append("员工花名册")
	if field.get("import_enabled"):
		used_in.append("花名册导入")
	if field.get("export_enabled"):
		used_in.append("花名册导出")

	fieldtype = field.get("fieldtype") or "Data"
	options = [option.strip() for option in str(field.get("options") or "").splitlines() if option.strip()]
	if fieldtype == "Link":
		value_source = "引用基础字典：{0}".format(field.get("options") or "未配置引用对象")
		allowed_values = (
			_get_base_data_value_samples(field.get("options"), limit=6)
			if field.get("options") in DEVELOPER_CENTER_PREVIEW_DICTIONARIES
			else []
		)
	elif fieldtype == "Select":
		value_source = "固定选项"
		allowed_values = [_display_option(option) for option in options]
	else:
		value_source = "自由录入"
		allowed_values = []

	return {
		"field_label": field.get("field_label"),
		"fieldname": fieldname,
		"fieldtype": fieldtype,
		"category": field.get("category"),
		"source": field.get("source"),
		"description": field.get("description") or "未补充业务说明",
		"meaning": known_reference.get("meaning") or field.get("description") or "员工档案业务字段。",
		"scope": known_reference.get("scope") or "、".join(used_in) or "当前未在员工前台展示",
		"managed_by": known_reference.get("managed_by") or value_source,
		"value_source": value_source,
		"allowed_values": allowed_values,
		"used_in": used_in,
		"enabled": field.get("enabled"),
	}


def _get_hrms_developer_field_catalog():
	field_center = _get_employee_field_center_payload()
	fields = [_field_runtime_reference(field) for field in field_center["fields"] if field.get("enabled")]
	fields.sort(key=lambda field: (field["category"] or "", field["field_label"] or ""))
	return fields


@frappe.whitelist()
def get_hrms_developer_configuration_map():
	"""Describe the supported no-code controls and their real runtime consumers."""
	_require_system_manager()
	definitions = [
		{
			"key": "employee-fields",
			"category": "员工档案",
			"label": "员工字段、导入导出与详情资料块",
			"doctype": TEMPLATE_DOCTYPE,
			"purpose": "控制员工字段显示、别名、导入映射、导出模板和详情资料块。",
			"where_used": "员工档案详情、员工花名册导入、导出和人事报表。",
			"storage": "HRMS Employee Field Template 与受控 Custom Field",
			"manage_route": "hr-settings-center",
			"verify_route": "employee-detail",
			"test_hint": "修改后打开员工档案、导入预览和导出模板，确认同一字段配置同时生效。",
		},
		{
			"key": "workflow",
			"category": "审批流程",
			"label": "工作流与审批状态",
			"doctype": "Workflow",
			"purpose": "配置单据状态、允许执行转换的角色以及审批动作。",
			"where_used": "启用了工作流的请假、人事、薪资等业务单据。",
			"storage": "Workflow / Workflow State / Workflow Transition",
			"manage_route": "List/Workflow",
			"verify_route": "List/Workflow",
			"test_hint": "使用申请人与审批人两个测试账户提交同一业务单据，验证状态和按钮随角色变化。",
		},
		{
			"key": "form-approval-matrix",
			"category": "导入审批",
			"label": "人资表单审批矩阵",
			"doctype": "HRMS Form Approval Matrix",
			"purpose": "按导入表单类型和业务条件确定审批步骤与审批角色。",
			"where_used": "人资表单导入中心；保存和提交导入批次时由 form_data_intake.py 读取。",
			"storage": "HRMS Form Approval Matrix",
			"manage_route": "List/HRMS Form Approval Matrix",
			"verify_route": "form-data-intake",
			"test_hint": "新建导入批次并进入提交预览，检查生成的审批步骤是否与矩阵一致。",
		},
		{
			"key": "attendance-rules",
			"category": "考勤规则",
			"label": "考勤自定义规则",
			"doctype": "HRMS Attendance Custom Rule",
			"purpose": "配置考勤导入时的匹配、校验、异常识别和处理规则。",
			"where_used": "考勤导入中心；attendance_import.py 在预检和入库时读取。",
			"storage": "HRMS Attendance Custom Rule",
			"manage_route": "List/HRMS Attendance Custom Rule",
			"verify_route": "attendance-import-center",
			"test_hint": "上传同一份考勤样例做预检，比较修改前后的命中规则和异常结果。",
		},
		{
			"key": "payroll-rules",
			"category": "薪资规则",
			"label": "薪资计算规则",
			"doctype": "HRMS Payroll Rule",
			"purpose": "配置全勤、加班、夜班、福利资格和结算参数。",
			"where_used": "薪资输入中心；payroll_input.py 计算与校验薪资输入记录时读取。",
			"storage": "HRMS Payroll Rule（按公司隔离）",
			"manage_route": "List/HRMS Payroll Rule",
			"verify_route": "payroll-input-center",
			"test_hint": "在薪资输入中心运行规则预览，确认计算明细引用了当前公司的规则版本。",
		},
		{
			"key": "payroll-mapping",
			"category": "薪资导入",
			"label": "薪资字段映射",
			"doctype": "HRMS Payroll Field Mapping",
			"purpose": "把外部薪资列名映射为系统字段，控制导入识别方式。",
			"where_used": "薪资输入中心的文件预览与导入；payroll_input.py 读取。",
			"storage": "HRMS Payroll Field Mapping",
			"manage_route": "List/HRMS Payroll Field Mapping",
			"verify_route": "payroll-input-center",
			"test_hint": "用固定样例文件执行预览，检查原始列是否映射到预期系统字段。",
		},
		{
			"key": "dingtalk",
			"category": "系统集成",
			"label": "钉钉连接与同步设置",
			"doctype": "HRMS DingTalk Settings",
			"purpose": "配置钉钉应用身份、同步开关和网关参数。",
			"where_used": "钉钉员工同步、考勤同步和网关接口；dingtalk_integration.py 与 dingtalk_employee_gateway.py 读取。",
			"storage": "HRMS DingTalk Settings（单例配置）",
			"manage_route": "Form/HRMS DingTalk Settings/HRMS DingTalk Settings",
			"verify_route": "attendance-import-center/dingtalk",
			"test_hint": "先执行连接测试和预览同步；确认成功后再开启正式同步，避免直接写入错误数据。",
		},
	]
	for item in definitions:
		item["record_count"] = _configuration_record_count(item["doctype"])
		item["status"] = "已接入业务" if item["record_count"] else "已接入，尚未配置记录"
	return {
		"items": definitions,
		"field_catalog": _get_hrms_developer_field_catalog(),
		"base_dictionaries": _get_base_data_module_summaries(),
		"boundary": {
			"no_code": "字段显示与映射、基础字典取值、已有规则参数、工作流、角色权限和用户数据范围可在系统内修改。",
			"requires_code": "新增计算算法、外部协议、新数据关系或绕过现有规则引擎的行为仍需代码、迁移和回归测试。",
		},
	}


@frappe.whitelist()
def create_employment_type_from_developer_center(employee_type_name: str):
	"""Reject ad-hoc public work-nature values; the five labels are fixed."""
	_require_system_manager()
	frappe.throw(_("工作性质固定为在职·正式、在职·试用期、退休返聘、待离职、离职五类，不能新增其他取值。"))


def _model_catalog_record_count(doctype):
	"""Return a safe catalogue count without making the guide depend on optional modules."""
	if not frappe.db.exists("DocType", doctype):
		return 0
	try:
		return _configuration_record_count(doctype)
	except Exception:
		return 0


@frappe.whitelist()
def get_hrms_model_governance_catalog():
	"""Expose the small, project-relevant model catalogue instead of the full framework registry."""
	_require_system_manager()
	definitions = [
		{
			"doctype": "Employee",
			"label": "员工档案",
			"category": "核心业务模型",
			"usage": "正在使用",
			"purpose": "保存员工身份、任职、组织归属和在职状态，是大多数人资业务的主档案。",
			"where_used": "员工档案、花名册导入导出、考勤、薪资、审批和人事报表。",
			"manage_route": "employee-detail",
			"manage_label": "进入员工档案",
			"safe_change": "字段显示、中文名称、导入映射请在“员工字段与导入导出”中维护。",
			"risk": "中",
			"origin": "HRMS 标准模型",
		},
		{
			"doctype": "Company",
			"label": "公司",
			"category": "核心业务模型",
			"usage": "正在使用",
			"purpose": "定义数据所属公司和业务隔离边界。",
			"where_used": "员工、部门、考勤、薪资、权限数据范围和所有公司级规则。",
			"manage_route": "List/Company",
			"manage_label": "管理公司",
			"safe_change": "可以维护公司资料；删除或改名会影响大量关联记录。",
			"risk": "高",
			"origin": "Frappe/ERP 标准模型",
		},
		{
			"doctype": "Department",
			"label": "部门",
			"category": "核心业务模型",
			"usage": "正在使用",
			"purpose": "保存组织层级、部门归属和部门负责人。",
			"where_used": "组织架构、员工档案、招聘计划、审批人和数据权限。",
			"manage_route": "List/Department",
			"manage_label": "管理部门",
			"safe_change": "日常组织调整可以维护；字段结构通过受控配置处理。",
			"risk": "中",
			"origin": "HRMS 标准模型",
		},
		{
			"doctype": "Designation",
			"label": "职位",
			"category": "核心业务模型",
			"usage": "正在使用",
			"purpose": "定义岗位/职位名称及其组织来源信息。",
			"where_used": "员工档案、招聘、编制计划、岗位分析和审批条件。",
			"manage_route": "List/Designation",
			"manage_label": "管理职位",
			"safe_change": "可维护职位资料；不要直接修改标准字段类型。",
			"risk": "中",
			"origin": "HRMS 标准模型",
		},
		{
			"doctype": "Job Applicant",
			"label": "候选人",
			"category": "核心业务模型",
			"usage": "按功能使用",
			"purpose": "保存应聘者、应聘职位和招聘阶段信息。",
			"where_used": "招聘、面试、录用和候选人报表。",
			"manage_route": "List/Job Applicant",
			"manage_label": "查看候选人",
			"safe_change": "业务数据在招聘模块维护；结构变更需要先评估导入和审批。",
			"risk": "中",
			"origin": "HRMS 标准模型",
		},
		{
			"doctype": "Attendance",
			"label": "考勤记录",
			"category": "核心业务模型",
			"usage": "正在使用",
			"purpose": "保存员工每日考勤结果。",
			"where_used": "考勤导入、异常检查、月度统计和薪资计算。",
			"manage_route": "attendance-import-center",
			"manage_label": "进入考勤中心",
			"safe_change": "匹配和异常逻辑在“考勤自定义规则”中维护，不直接改模型。",
			"risk": "高",
			"origin": "HRMS 标准模型",
		},
		{
			"doctype": "Leave Application",
			"label": "请假申请",
			"category": "核心业务模型",
			"usage": "按功能使用",
			"purpose": "保存员工请假期间、类型、审批和状态。",
			"where_used": "请假、考勤、审批和假期余额计算。",
			"manage_route": "List/Leave Application",
			"manage_label": "查看请假申请",
			"safe_change": "审批路径使用工作流配置；日期和余额算法仍由业务代码控制。",
			"risk": "高",
			"origin": "HRMS 标准模型",
		},
		{
			"doctype": "Salary Slip",
			"label": "工资单",
			"category": "核心业务模型",
			"usage": "按功能使用",
			"purpose": "保存员工每个薪资周期的计算和结算结果。",
			"where_used": "薪资计算、工资单、财务凭证和员工自助查询。",
			"manage_route": "payroll-input-center",
			"manage_label": "进入薪资中心",
			"safe_change": "计算参数和映射使用薪资规则页面维护，不直接改工资单结构。",
			"risk": "高",
			"origin": "HRMS 标准模型",
		},
		{
			"doctype": TEMPLATE_DOCTYPE,
			"label": "员工字段模板",
			"category": "无代码业务配置",
			"usage": "已接入业务",
			"purpose": "控制员工字段的显示、中文名称、必填、导入导出和详情资料块。",
			"where_used": "员工档案详情、花名册导入导出和人事报表。",
			"manage_route": "hr-settings-center",
			"manage_label": "管理员工字段",
			"safe_change": "这里是员工字段的首选入口，保存后由现有业务页面读取。",
			"risk": "低",
			"origin": "本项目扩展模型",
		},
		{
			"doctype": "HRMS Form Approval Matrix",
			"label": "人资表单审批矩阵",
			"category": "无代码业务配置",
			"usage": "已接入业务",
			"purpose": "按导入表单和条件选择审批步骤、审批角色。",
			"where_used": "人资表单导入中心的保存、提交和审批步骤生成。",
			"manage_route": "List/HRMS Form Approval Matrix",
			"manage_label": "管理审批矩阵",
			"safe_change": "可以页面化维护；修改后用固定导入样例验证审批步骤。",
			"risk": "中",
			"origin": "本项目扩展模型",
		},
		{
			"doctype": "HRMS Attendance Custom Rule",
			"label": "考勤自定义规则",
			"category": "无代码业务配置",
			"usage": "已接入业务",
			"purpose": "配置考勤导入的匹配、校验、异常识别和处理参数。",
			"where_used": "考勤导入中心预检和正式入库。",
			"manage_route": "List/HRMS Attendance Custom Rule",
			"manage_label": "管理考勤规则",
			"safe_change": "可以页面化维护；用同一考勤文件比较修改前后预检结果。",
			"risk": "中",
			"origin": "本项目扩展模型",
		},
		{
			"doctype": "HRMS Payroll Rule",
			"label": "薪资计算规则",
			"category": "无代码业务配置",
			"usage": "已接入业务",
			"purpose": "配置全勤、加班、夜班、福利资格和结算参数。",
			"where_used": "薪资输入中心的计算、校验和规则预览。",
			"manage_route": "List/HRMS Payroll Rule",
			"manage_label": "管理薪资规则",
			"safe_change": "可以页面化维护；发布前必须使用固定月份样例对比计算明细。",
			"risk": "高",
			"origin": "本项目扩展模型",
		},
		{
			"doctype": "HRMS Payroll Field Mapping",
			"label": "薪资字段映射",
			"category": "无代码业务配置",
			"usage": "已接入业务",
			"purpose": "把外部薪资表列名映射为系统字段。",
			"where_used": "薪资文件预览、导入和结果字段汇总。",
			"manage_route": "List/HRMS Payroll Field Mapping",
			"manage_label": "管理字段映射",
			"safe_change": "可以页面化维护；用固定文件验证每一列的目标字段。",
			"risk": "中",
			"origin": "本项目扩展模型",
		},
		{
			"doctype": "HRMS DingTalk Settings",
			"label": "钉钉连接设置",
			"category": "无代码业务配置",
			"usage": "按功能使用",
			"purpose": "保存钉钉应用身份、同步开关和网关参数。",
			"where_used": "钉钉员工、审批和考勤同步。",
			"manage_route": "Form/HRMS DingTalk Settings/HRMS DingTalk Settings",
			"manage_label": "管理钉钉连接",
			"safe_change": "先连接测试和预览，再开启正式同步；不要在说明页展示密钥。",
			"risk": "高",
			"origin": "本项目扩展模型",
		},
		{
			"doctype": "Workflow",
			"label": "工作流",
			"category": "无代码业务配置",
			"usage": "按功能使用",
			"purpose": "配置业务状态、转换动作和允许执行转换的角色。",
			"where_used": "已启用工作流的请假、人事、薪资和导入业务。",
			"manage_route": "List/Workflow",
			"manage_label": "管理工作流",
			"safe_change": "使用申请人与审批人两个测试账户验证完整状态流转。",
			"risk": "高",
			"origin": "Frappe 标准配置模型",
		},
		{
			"doctype": "HRMS Data Cleanup Log",
			"label": "数据清理日志",
			"category": "系统内部记录",
			"usage": "后台自动使用",
			"purpose": "记录后台关联数据清理任务的执行结果。",
			"where_used": "数据清理任务与运维审计，不是日常业务配置。",
			"manage_route": "hrms-data-operations",
			"manage_label": "查看运行状态",
			"safe_change": "只查看，不要手工新增、修改字段或改权限。",
			"risk": "禁止直接修改",
			"origin": "本项目内部模型",
		},
		{
			"doctype": "HRMS Business Process Record",
			"label": "业务流程运行记录",
			"category": "系统内部记录",
			"usage": "后台自动使用",
			"purpose": "保存自动化业务流程的运行状态和关联信息。",
			"where_used": "后台任务、流程恢复和运维排查。",
			"manage_route": "hrms-data-operations",
			"manage_label": "查看运行状态",
			"safe_change": "只查看；人工编辑会破坏流程状态一致性。",
			"risk": "禁止直接修改",
			"origin": "本项目内部模型",
		},
		{
			"doctype": "HRMS Form Import Batch",
			"label": "表单导入批次",
			"category": "系统内部记录",
			"usage": "后台自动使用",
			"purpose": "保存一次表单导入的批次、审批和处理状态。",
			"where_used": "人资表单导入中心。",
			"manage_route": "form-data-intake",
			"manage_label": "进入导入中心",
			"safe_change": "从导入中心操作，不在单据类型编辑器中修改。",
			"risk": "禁止直接修改",
			"origin": "本项目内部模型",
		},
		{
			"doctype": "HRMS Form Import Row",
			"label": "表单导入明细行",
			"category": "系统内部记录",
			"usage": "后台自动使用",
			"purpose": "保存导入批次中的逐行解析和校验结果。",
			"where_used": "人资表单导入预览、错误定位和正式入库。",
			"manage_route": "form-data-intake",
			"manage_label": "进入导入中心",
			"safe_change": "只由导入流程生成；不要人工编辑。",
			"risk": "禁止直接修改",
			"origin": "本项目内部模型",
		},
		{
			"doctype": "HRMS Attendance Month Lock",
			"label": "考勤月度锁定",
			"category": "系统内部记录",
			"usage": "后台自动使用",
			"purpose": "记录考勤月份是否已锁定，防止结算后数据继续变化。",
			"where_used": "考勤月结、撤回和薪资结算保护。",
			"manage_route": "attendance-import-center",
			"manage_label": "进入考勤中心",
			"safe_change": "通过考勤中心执行锁定或撤回，不直接编辑底层记录。",
			"risk": "禁止直接修改",
			"origin": "本项目内部模型",
		},
	]

	items = []
	for index, item in enumerate(definitions):
		item = dict(item)
		item["id"] = index + 1
		item["record_count"] = _model_catalog_record_count(item["doctype"])
		item["exists"] = int(bool(frappe.db.exists("DocType", item["doctype"])))
		item["structure_route"] = f"Form/DocType/{item['doctype']}"
		items.append(item)

	return {
		"items": items,
		"summary": {
			"project_model_count": len(items),
			"business_model_count": sum(1 for item in items if item["category"] == "核心业务模型"),
			"config_model_count": sum(1 for item in items if item["category"] == "无代码业务配置"),
			"internal_model_count": sum(1 for item in items if item["category"] == "系统内部记录"),
			"framework_model_count": frappe.db.count("DocType"),
		},
		"guidance": {
			"need_to_know": "不需要了解全部底层模型。日常管理只需要理解核心业务模型和无代码配置；系统内部记录只用于排查。",
			"doctype_meaning": "单据类型是数据结构定义：它同时决定保存哪些字段、表单如何呈现、数据存到哪里、权限和工作流如何挂接。",
			"raw_registry": "完整列表包含 Frappe、ERPNext、HRMS 和本项目扩展模型，很多库存、会计或框架模型并未进入本项目的人资导航。",
		},
	}


@frappe.whitelist()
def test_hrms_effective_permission(
	user: str,
	doctype: str,
	permission_type: str = "read",
	document_name: str | None = None,
):
	"""Run a read-only check through Frappe's real effective-permission engine."""
	_require_system_manager()
	allowed_permission_types = {
		"select",
		"read",
		"write",
		"create",
		"delete",
		"submit",
		"cancel",
		"amend",
		"report",
		"import",
		"export",
		"print",
		"email",
		"share",
	}
	if not frappe.db.exists("User", user):
		frappe.throw(_("账户 {0} 不存在").format(user))
	if not frappe.db.exists("DocType", doctype):
		frappe.throw(_("单据类型 {0} 不存在").format(doctype))
	if permission_type not in allowed_permission_types:
		frappe.throw(_("不支持的权限类型：{0}").format(permission_type))

	document_name = (document_name or "").strip()
	doc = None
	if document_name:
		if not frappe.db.exists(doctype, document_name):
			frappe.throw(_("{0} 记录 {1} 不存在").format(doctype, document_name))
		doc = frappe.get_doc(doctype, document_name)

	roles = frappe.get_roles(user)
	allowed = bool(frappe.has_permission(doctype, permission_type, doc=doc, user=user))
	user_permissions = frappe.get_all(
		"User Permission",
		filters={"user": user},
		fields=["allow", "for_value", "applicable_for", "is_default", "hide_descendants"],
		order_by="allow asc, for_value asc",
		ignore_permissions=True,
	)
	permission_field = {
		"select": "select",
		"read": "read",
		"write": "write",
		"create": "create",
		"delete": "delete",
		"submit": "submit",
		"cancel": "cancel",
		"amend": "amend",
		"report": "report",
		"import": "import",
		"export": "export",
		"print": "print",
		"email": "email",
		"share": "share",
	}[permission_type]
	granting_roles = set()
	for permission_doctype in ("DocPerm", "Custom DocPerm"):
		for row in frappe.get_all(
			permission_doctype,
			filters={"parent": doctype, "role": ["in", roles], permission_field: 1},
			fields=["role"],
			ignore_permissions=True,
		):
			granting_roles.add(row.role)

	return {
		"allowed": int(allowed),
		"user": user,
		"doctype": doctype,
		"document_name": document_name,
		"permission_type": permission_type,
		"roles": roles,
		"granting_roles": sorted(granting_roles),
		"user_permissions": user_permissions,
		"scope_mode": "具体记录" if document_name else "单据类型入口",
		"explanation": (
			"系统实际权限引擎允许该操作。" if allowed else "系统实际权限引擎拒绝该操作。"
		),
	}


@frappe.whitelist()
def save_employee_field_center(items: str):
	_require_hr_settings_manager()
	return save_employee_field_template(items)


@frappe.whitelist()
def get_employee_import_export_schema():
	doc = _get_template_doc()
	import_fields = _get_employee_import_fields(doc)
	fields = _get_employee_export_fields(doc)
	categories = []
	for category in EMPLOYEE_TEMPLATE_CATEGORIES:
		category_fields = [field for field in fields if field["category"] == category]
		categories.append({"label": category, "fields": category_fields})

	return {
		"fields": fields,
		"import_fields": import_fields,
		"export_fields": fields,
		"categories": categories,
		"multi_record_categories": MULTI_RECORD_EXPORT_CATEGORIES,
	}


EMPLOYEE_ROSTER_STATUS_CARDS = [
	{"label": "在职 · 正式", "filters": {"employment_type": "Full-time", "custom_is_confirmed": "是", "status": "Active"}},
	{"label": "在职 · 试用期", "filters": {"employment_type": "Full-time", "custom_is_confirmed": "否", "status": "Active"}},
	{"label": "退休返聘", "filters": {"employment_type": "Retainer", "status": "Active"}},
	{"label": "待离职", "filters": {"status": "Inactive"}},
	{"label": "离职", "filters": {"status": "Left"}},
]

EMPLOYEE_ROSTER_SORT_OPTIONS = {
	"date_of_joining": "date_of_joining",
	"modified": "modified",
	"employee_name": "employee_name",
	"custom_employee_code": "custom_employee_code",
	"name": "name",
}

EMPLOYEE_ROSTER_QUICK_EDIT_FIELDS = {
	"employee_name",
	"first_name",
	"custom_employee_code",
	"company",
	"branch",
	"department",
	"designation",
	"reports_to",
	"grade",
	"employment_type",
	"status",
	"date_of_joining",
	"cell_number",
}


def ensure_required_roster_columns(doc):
	changed = False
	rows_by_fieldname = {row.fieldname: row for row in doc.template_items}

	for fieldname, field_label in EMPLOYEE_ROSTER_REQUIRED_COLUMNS.items():
		row = rows_by_fieldname.get(fieldname)
		if not row:
			continue
		if row.get("field_label") != field_label:
			row.field_label = field_label
			changed = True
		for flag in ("enabled", "roster_visible", "export_enabled", "import_enabled", "detail_visible", "form_visible"):
			if _template_item_supports_field(flag) and frappe.utils.cint(row.get(flag)) != 1:
				row.set(flag, 1)
				changed = True

	if changed:
		doc.save(ignore_permissions=True)
	return doc


def _retire_personnel_status_field(doc):
	"""Remove the superseded public field from the template and Employee schema."""
	changed = False
	for row in list(doc.template_items):
		if row.fieldname == "custom_personnel_status":
			doc.remove(row)
			changed = True

	rows_by_fieldname = {row.fieldname: row for row in doc.template_items}
	employment_type = rows_by_fieldname.get("employment_type")
	if employment_type:
		if employment_type.get("field_label") != "工作性质":
			employment_type.set("field_label", "工作性质")
			changed = True
		if _template_item_supports_field("aliases"):
			aliases = [alias.strip() for alias in str(employment_type.get("aliases") or "").splitlines() if alias.strip()]
			for alias in ("工作性质", "雇佣类型", "用工类型"):
				if alias not in aliases:
					aliases.append(alias)
			new_aliases = "\n".join(aliases)
			if employment_type.get("aliases") != new_aliases:
				employment_type.set("aliases", new_aliases)
				changed = True

	for internal_fieldname in ("status",):
		row = rows_by_fieldname.get(internal_fieldname)
		if not row:
			continue
		for fieldname in ("roster_visible", "detail_visible"):
			if _template_item_supports_field(fieldname) and _template_row_int(row, fieldname) != 0:
				row.set(fieldname, 0)
				changed = True

	if changed:
		doc.save(ignore_permissions=True)

	custom_field_name = f"{EMPLOYEE_DOCTYPE}-custom_personnel_status"
	if frappe.db.exists("Custom Field", custom_field_name):
		frappe.delete_doc("Custom Field", custom_field_name, ignore_permissions=True, force=True)
		frappe.clear_cache(doctype=EMPLOYEE_DOCTYPE)
	return doc


def _get_employee_roster_columns():
	# search_fields: ["employee_name", "cell_number", "custom_employee_code"]
	# sort_options
	doc = _get_template_doc()
	ensure_required_roster_columns(doc)
	fields = [
		_serialize_item(row)
		for row in doc.template_items
		if not _is_employee_internal_field(row.fieldname)
		and row.get("enabled")
		and (_field_flag_enabled(row, "roster_visible", 0) or row.fieldname in EMPLOYEE_ROSTER_REQUIRED_COLUMNS)
	]
	fields_by_name = {field["fieldname"]: field for field in fields}
	columns = []

	for fieldname, field_label in EMPLOYEE_ROSTER_REQUIRED_COLUMNS.items():
		field = fields_by_name.get(fieldname)
		if field:
			field = {**field, "field_label": field_label}
			columns.append(field)

	column_fieldnames = {column["fieldname"] for column in columns}
	for field in fields:
		if field["fieldname"] not in column_fieldnames:
			columns.append(field)
			column_fieldnames.add(field["fieldname"])

	return columns


def _build_employee_roster_filters(filters=None):
	filters = _parse_json(filters, {}) or {}
	meta_fields = _get_employee_meta_field_map()
	allowed_filters = {
		"status",
		"employment_type",
		"custom_is_confirmed",
		"department",
		"designation",
		"company",
		"branch",
	}
	employee_filters = {}

	for fieldname, value in filters.items():
		if fieldname in allowed_filters and fieldname in meta_fields and value not in (None, ""):
			employee_filters[fieldname] = value

	# Every roster request is scoped to the active company.  The browser sends
	# it explicitly; a server-side default keeps direct API calls from silently
	# returning employees from every company.
	if "company" in meta_fields and not employee_filters.get("company"):
		default_company = frappe.defaults.get_user_default("Company")
		if default_company:
			employee_filters["company"] = default_company

	return employee_filters


def _build_employee_roster_or_filters(search):
	search = (search or "").strip()
	if not search:
		return []

	meta_fields = _get_employee_meta_field_map()
	or_filters = []
	for fieldname in ("employee_name", "cell_number", "custom_employee_code"):
		if fieldname == "name" or fieldname in meta_fields:
			or_filters.append([EMPLOYEE_DOCTYPE, fieldname, "like", f"%{search}%"])
	return or_filters


def _get_roster_fetch_fields(columns):
	fetch_fields = {"name", "modified", "image"}
	for field in columns:
		fetch_fields.add(field["fieldname"])
	for fieldname in [
		"employee_name",
		"custom_employee_code",
		"department",
		"designation",
		"cell_number",
		"custom_id_type",
		"passport_number",
		"status",
		"employment_type",
		"date_of_joining",
	]:
		fetch_fields.add(fieldname)

	meta_fields = _get_employee_meta_field_map()
	return [fieldname for fieldname in fetch_fields if fieldname == "name" or fieldname in meta_fields]


def _strip_department_company_suffix(value):
	text = str(value or "").strip()
	return re.sub(r"\s+-\s+[^-]+$", "", text).strip()


def _get_department_display_names(department_values):
	department_values = sorted({value for value in department_values if value})
	if not department_values:
		return {}
	return {
		department.name: department.department_name
		for department in frappe.get_all(
			"Department",
			filters={"name": ["in", department_values]},
			fields=["name", "department_name"],
			limit_page_length=0,
		)
		if department.department_name
	}


def _department_display_name(value, department_names=None):
	if not value:
		return ""
	if department_names is None:
		department_names = _get_department_display_names([value])
	if value in department_names and department_names[value]:
		return department_names[value]
	return _strip_department_company_suffix(value)


def _hydrate_employee_roster_display_values(rows):
	department_values = sorted({row.get("department") for row in rows if row.get("department")})
	department_names = _get_department_display_names(department_values)

	for row in rows:
		row["department_display"] = _department_display_name(row.get("department"), department_names)
		row["employee_code_display"] = row.get("custom_employee_code") or ""
	return rows


def _employee_business_code_value(row):
	return str(row.get("custom_employee_code") or "").strip()


def _employee_business_code_sort_key(row):
	value = _employee_business_code_value(row)
	return tuple(
		(0, int(part)) if part.isdigit() else (1, part.casefold())
		for part in re.split(r"(\d+)", value)
	)


def _sort_employee_roster_by_business_code(rows, sort_order):
	populated = [row for row in rows if _employee_business_code_value(row)]
	blank = [row for row in rows if not _employee_business_code_value(row)]
	populated.sort(
		key=_employee_business_code_sort_key,
		reverse=sort_order == "desc",
	)
	# Missing work numbers remain visible as data-quality exceptions, but never
	# appear before valid work numbers in either direction.
	return populated + blank


def _count_employee_rows(filters, or_filters=None):
	"""Count permission-filtered employees without materializing every name."""
	rows = frappe.get_list(
		EMPLOYEE_DOCTYPE,
		filters=filters,
		or_filters=or_filters,
		fields=[{"COUNT": "*", "as": "count"}],
		limit_page_length=1,
	)
	return frappe.utils.cint(rows[0].get("count")) if rows else 0


@frappe.whitelist()
def get_employee_by_business_code(employee_code: str, company: str = ""):
	"""Resolve the public company work number to the internal Employee link value."""
	if not frappe.has_permission(EMPLOYEE_DOCTYPE, "read"):
		frappe.throw(_("无权查询员工信息"), frappe.PermissionError)

	employee_code = str(employee_code or "").strip()
	if not employee_code:
		return None

	filters = {"status": "Active"}
	if company:
		filters["company"] = company

	rows = frappe.get_list(
		EMPLOYEE_DOCTYPE,
		filters={**filters, "custom_employee_code": employee_code},
		fields=["name", "employee_name", "custom_employee_code", "company"],
		limit_page_length=2,
	)
	if not rows:
		return None
	if len(rows) > 1:
		frappe.throw(_("工号 {0} 匹配到多名在职员工，请先在员工花名册中处理重复工号。").format(employee_code))

	employee = rows[0]
	return {
		"name": employee.name,
		"employee_name": employee.employee_name,
		"employee_code": employee.custom_employee_code,
		"company": employee.company,
	}


@frappe.whitelist()
def get_employee_roster(
	filters: str = "{}",
	search: str = "",
	sort_by: str = "modified",
	sort_order: str = "desc",
	page: int = 1,
	page_length: int = 20,
):
	# page_length
	# dynamic_columns
	columns = _get_employee_roster_columns()
	employee_filters = _build_employee_roster_filters(filters)
	or_filters = _build_employee_roster_or_filters(search)
	sort_field = EMPLOYEE_ROSTER_SORT_OPTIONS.get(sort_by) or "modified"
	sort_order = "asc" if str(sort_order).lower() == "asc" else "desc"
	page = max(frappe.utils.cint(page), 1)
	page_length = min(max(frappe.utils.cint(page_length) or 20, 10), 500)
	start = (page - 1) * page_length
	fields = _get_roster_fetch_fields(columns)

	if sort_field == "custom_employee_code":
		all_rows = frappe.get_list(
			EMPLOYEE_DOCTYPE,
			filters=employee_filters,
			or_filters=or_filters,
			fields=fields,
			limit_page_length=0,
		)
		total = len(all_rows)
		rows = _sort_employee_roster_by_business_code(all_rows, sort_order)[start : start + page_length]
	else:
		rows = frappe.get_list(
			EMPLOYEE_DOCTYPE,
			filters=employee_filters,
			or_filters=or_filters,
			fields=fields,
			order_by=f"{sort_field} {sort_order}",
			limit_start=start,
			limit_page_length=page_length,
		)
		total = _count_employee_rows(employee_filters, or_filters)
	rows = _hydrate_employee_roster_display_values(rows)

	return {
		"rows": rows,
		"columns": columns,
		"total": total,
		"page": page,
		"page_length": page_length,
		"sort_by": sort_field,
		"sort_order": sort_order,
		"status_cards": EMPLOYEE_ROSTER_STATUS_CARDS,
		"sort_options": [
			{"label": "入职日期", "value": "date_of_joining"},
			{"label": "更新时间", "value": "modified"},
			{"label": "姓名", "value": "employee_name"},
			{"label": "工号", "value": "custom_employee_code"},
		],
	}


@frappe.whitelist()
def get_employee_roster_summary(filters: str = "{}"):
	employee_filters = _build_employee_roster_filters(filters)
	summary = []
	for card in EMPLOYEE_ROSTER_STATUS_CARDS:
		card_filters = dict(employee_filters)
		card_filters.update(_build_employee_roster_filters(card["filters"]))
		summary.append(
			{
				"label": card["label"],
				"filters": card["filters"],
				"count": _count_employee_rows(card_filters),
			}
		)
	return summary


@frappe.whitelist()
def quick_update_employee_roster(employee: str, values: str = "{}"):
	values = _parse_json(values, {}) or {}
	if not employee:
		frappe.throw(_("请选择员工"))

	doc = frappe.get_doc(EMPLOYEE_DOCTYPE, employee)
	doc.check_permission("write")
	meta_fields = _get_employee_meta_field_map()
	template_fields = {field["fieldname"] for field in _get_employee_import_fields(_get_template_doc())}
	allowed_fields = EMPLOYEE_ROSTER_QUICK_EDIT_FIELDS | template_fields
	for fieldname, value in values.items():
		if fieldname in allowed_fields and fieldname in meta_fields:
			doc.set(fieldname, value)
	doc.save()
	return {"name": doc.name}


def _display_employee_field_value(fieldname, value):
	if fieldname == "department":
		return _department_display_name(value)
	return value


def _get_employee_detail_sections(doc, department_display=""):
	template = _get_template_doc()
	fields = [
		_serialize_item(row)
		for row in template.template_items
		if not _is_employee_internal_field(row.fieldname)
		and row.fieldname not in {"status"}
		and row.get("enabled")
		and _field_flag_enabled(row, "detail_visible", 1)
	]
	sections = []
	doc_values = doc.as_dict()
	for category in EMPLOYEE_TEMPLATE_CATEGORIES:
		category_fields = []
		for field in fields:
			if field["category"] != category or field["fieldname"] not in doc_values:
				continue
			value = doc.get(field["fieldname"])
			if field["fieldname"] == "department" and department_display:
				value = department_display
			category_fields.append({**field, "value": value})
		if category_fields:
			sections.append({"label": category, "fields": category_fields})
	return sections


def _get_employee_child_items(doc, child_fieldname, field_map, limit=5):
	items = []
	for row in list(doc.get(child_fieldname) or [])[:limit]:
		items.append(
			{
				"fields": [
					{"label": label, "value": _display_employee_field_value(fieldname, row.get(fieldname))}
					for label, fieldname in field_map
					if row.get(fieldname) not in (None, "")
				]
			}
		)
	return items


def _get_employee_flat_related_item(doc, field_map):
	fields = [
		{"label": label, "value": _display_employee_field_value(fieldname, doc.get(fieldname))}
		for label, fieldname in field_map
		if doc.get(fieldname) not in (None, "")
	]
	if not fields:
		return []
	return [{"fields": fields}]


def _get_employee_payroll_social_insurance_items(doc):
	"""Return the social-insurance amounts actually entered for this employee's payroll.

	Employee profile fields describe participation eligibility, but contribution
	amounts belong to the monthly payroll source records.  Keeping this lookup
	separate prevents an employee profile from showing a stale or inferred amount.
	"""
	if not frappe.db.exists("DocType", PAYROLL_WELFARE_SOURCE_DOCTYPE):
		return []

	rows = frappe.get_all(
		PAYROLL_WELFARE_SOURCE_DOCTYPE,
		filters={
			"employee": doc.name,
			"source_type": ["in", PAYROLL_SOCIAL_INSURANCE_SOURCE_TYPES],
			"confirmation_status": ["!=", "已驳回"],
		},
		fields=["payroll_month", "attendance_lock_version", "source_type", "amount", "confirmation_status"],
		order_by="payroll_month desc, modified desc",
		limit_page_length=1000,
	)
	grouped = {}
	for row in rows:
		month = row.payroll_month or "未设置月份"
		key = (month, row.attendance_lock_version or "")
		group = grouped.setdefault(
			key,
			{
				"amounts": {source_type: 0 for source_type in PAYROLL_SOCIAL_INSURANCE_SOURCE_TYPES},
				"entered_types": set(),
				"statuses": set(),
			},
		)
		group["amounts"][row.source_type] += flt(row.amount)
		group["entered_types"].add(row.source_type)
		if row.confirmation_status:
			group["statuses"].add(row.confirmation_status)

	items = []
	for (month, _lock_version), group in grouped.items():
		fields = [{"label": "薪资月份", "value": month}]
		for source_type in PAYROLL_SOCIAL_INSURANCE_SOURCE_TYPES:
			if source_type in group["entered_types"]:
				fields.append({"label": source_type, "value": "{0:.2f} 元".format(group["amounts"][source_type])})
		if group["statuses"]:
			fields.append({"label": "录入状态", "value": "、".join(sorted(group["statuses"]))})
		items.append({"fields": fields})
	return items


def _get_employee_doctype_items(doctype, filters, field_map, order_by="modified desc", limit=5):
	if not frappe.db.exists("DocType", doctype):
		return []
	meta = frappe.get_meta(doctype)
	fields = ["name"] + [fieldname for _, fieldname in field_map if meta.get_field(fieldname)]
	rows = frappe.get_all(doctype, filters=filters, fields=fields, order_by=order_by, limit_page_length=limit)
	return [
		{
			"name": row.name,
			"doctype": doctype,
			"fields": [
				{"label": label, "value": _display_employee_field_value(fieldname, row.get(fieldname))}
				for label, fieldname in field_map
				if row.get(fieldname) not in (None, "")
			],
		}
		for row in rows
	]


def _get_configured_detail_block_records(doc):
	template = _get_template_doc()
	records_by_tab = {}
	fields_by_block = {}
	for row in template.template_items:
		if not row.get("enabled") or not _field_flag_enabled(row, "detail_visible", 1):
			continue
		block = _template_row_value(row, "detail_block", "")
		if not block:
			continue
		fields_by_block.setdefault(block, []).append(row)

	for block_name, rows in fields_by_block.items():
		rows.sort(key=lambda row: (_template_row_int(row, "detail_block_order", 0), row.idx or 0))
		field_map = [(row.field_label or row.fieldname, row.fieldname) for row in rows]
		items = _get_employee_flat_related_item(doc, field_map)
		if not items:
			continue
		tab = EMPLOYEE_DETAIL_BLOCK_TABS.get(block_name, "概览")
		records_by_tab.setdefault(tab, []).append(
			_make_related_record(
				block_name,
				DEFAULT_DETAIL_BLOCK_DESCRIPTIONS.get(block_name, ""),
				[label for label, _fieldname in field_map],
				items,
				None,
				"在字段中心配置",
				"hr-settings-center",
			)
		)
	return records_by_tab


def _make_related_record(
	label,
	description,
	fields,
	items=None,
	action_doctype=None,
	action_label=None,
	action_route=None,
	compact=False,
):
	items = items or []
	return {
		"label": label,
		"description": description,
		"fields": fields,
		"items": items,
		"count": len(items),
		"action_doctype": action_doctype,
		"action_label": action_label or "新增记录",
		"action_route": action_route,
		"compact": compact,
	}


def _get_employee_related_records(doc):
	reward_punishment_items = _get_employee_doctype_items(
		"HRMS Employee Reward Punishment",
		{"employee": doc.name},
		[
			("奖惩类型", "reward_punishment_type"),
			("奖惩类别", "category"),
			("奖惩日期", "occurred_on"),
			("主旨", "subject"),
			("奖惩原因", "reason"),
			("处理结果", "decision_result"),
			("奖惩金额", "amount"),
			("经办人", "handled_by"),
			("附件", "attachment"),
			("状态", "status"),
		],
		order_by="occurred_on desc",
	)
	transfer_items = _get_employee_doctype_items(
		"Employee Transfer",
		{"employee": doc.name},
		[("异动日期", "transfer_date"), ("状态", "docstatus"), ("新员工编号", "new_employee_id")],
	)
	promotion_items = _get_employee_doctype_items(
		"Employee Promotion",
		{"employee": doc.name},
		[("转正/晋升日期", "promotion_date"), ("当前薪资", "current_ctc"), ("调整后薪资", "revised_ctc")],
	)
	separation_items = _get_employee_doctype_items(
		"Employee Separation",
		{"employee": doc.name},
		[("离职员工", "employee"), ("离职模板", "employee_separation_template"), ("办理状态", "boarding_status")],
	)
	education_items = _get_employee_child_items(
		doc,
		"education",
		[("学历", "qualification"), ("毕业院校", "school_univ"), ("专业", "major_opt_subj"), ("毕业年份", "year_of_passing")],
	) + _get_employee_flat_related_item(
		doc,
		[
			("学历类别", "custom_education_category"),
			("学习形式", "custom_study_mode"),
			("学历", "custom_education_level"),
			("毕业院校", "custom_graduation_school"),
			("科系", "custom_major"),
		],
	)
	contract_items = _get_employee_flat_related_item(
		doc,
		[
			("合同编号", "custom_contract_no"),
			("签订日期", "custom_contract_sign_date"),
			("签订次数", "custom_contract_sign_count"),
			("结束日期", "contract_end_date"),
		],
	)
	insurance_items = _get_employee_payroll_social_insurance_items(doc)

	records = {
		"在职信息": [
			_make_related_record(
				"任职记录",
				"记录员工部门、岗位、职级、工作地点、任职起止日期等历史变化。提交人事异动、转正或岗位调整后，会形成员工成长记录。",
				["起始日期", "结束日期", "部门", "岗位", "职级", "公司", "工作地点"],
				_get_employee_child_items(
					doc,
					"internal_work_history",
					[("起始日期", "from_date"), ("结束日期", "to_date"), ("部门", "department"), ("岗位", "designation"), ("公司", "company")],
				)
				+ transfer_items
				+ promotion_items,
				"Employee Transfer",
				"办理人事异动",
			),
			_make_related_record(
				"奖惩记录",
				"记录奖励、处分、奖惩日期、奖惩原因、处理结果、附件等信息，用于员工成长记录和人事档案留痕。",
				["奖惩类别", "奖惩日期", "奖惩原因", "处理结果", "经办人", "附件"],
				reward_punishment_items,
				"HRMS Employee Reward Punishment",
				"新增奖惩记录",
			),
			_make_related_record(
				"考察期信息",
				"记录试用期、考察开始/结束日期、考察结果、转正日期和转正办理记录。",
				["试用期", "考察开始日期", "考察结束日期", "考察结果", "转正日期", "转正意见"],
				promotion_items,
				"Employee Promotion",
				"办理转正",
			),
			_make_related_record(
				"退休信息",
				"记录退休日期、退休年龄、退休办理状态和退休备注。",
				["退休日期", "退休年龄", "办理状态", "退休备注"],
				[
					{
						"fields": [
							{"label": "退休日期", "value": doc.get("date_of_retirement")},
							{"label": "离职日期", "value": doc.get("relieving_date")},
							{"label": "离职原因", "value": doc.get("reason_for_leaving")},
						]
					}
				]
				if doc.get("date_of_retirement") or doc.get("relieving_date")
				else [],
			),
			_make_related_record(
				"档案信息",
				"记录员工档案编号、档案备注、材料完整度、档案所在位置等信息。",
				["档案编号", "档案备注", "材料完整度", "档案位置", "归档日期"],
				[],
				None,
				"添加档案字段",
				"staff-attribute-settings",
			),
		],
		"个人信息": [
			_make_related_record(
				"教育经历",
				"记录学历、毕业院校、专业、学习形式、毕业时间等教育背景。",
				["学历类别", "学习形式", "学历", "毕业院校", "科系", "毕业时间"],
				education_items,
			),
			_make_related_record(
				"工作经历",
				"记录入职前工作单位、岗位、起止时间、证明人等信息。",
				["工作单位", "岗位", "开始日期", "结束日期", "证明人"],
				_get_employee_child_items(doc, "external_work_history", [("工作单位", "company_name"), ("岗位", "designation"), ("开始日期", "from_date"), ("结束日期", "to_date")]),
			),
			_make_related_record("语言能力", "记录语种、熟练程度、证书和备注。", ["语种", "熟练程度", "证书", "备注"], [], None, "添加语言字段", "staff-attribute-settings"),
			_make_related_record("工作技能", "记录技能名称、熟练度、认证情况和备注。", ["技能名称", "熟练度", "认证情况", "备注"], [], None, "添加技能字段", "staff-attribute-settings"),
			_make_related_record("职称", "记录职称名称、等级、取得时间、有效期和附件。", ["职称名称", "等级", "取得时间", "有效期", "附件"], [], None, "添加职称字段", "staff-attribute-settings"),
			_make_related_record("证书/证件", "记录证书名称、证书编号、发证机构、有效期和附件。", ["证书名称", "证书编号", "发证机构", "有效期", "附件"], [], None, "添加证书字段", "staff-attribute-settings"),
			_make_related_record("培训经历", "记录培训名称、培训机构、培训日期、结果和证书。", ["培训名称", "培训机构", "培训日期", "结果", "证书"], [], None, "添加培训字段", "staff-attribute-settings"),
		],
		"联系信息": [
			_make_related_record("紧急联系人", "记录紧急联系人姓名、关系、电话、地址和备注。", ["姓名", "关系", "电话", "地址", "备注"], []),
			_make_related_record("家庭成员", "记录家庭成员姓名、关系、工作单位、电话和备注。", ["姓名", "关系", "工作单位", "电话", "备注"], []),
		],
		"合同信息": [
			_make_related_record(
				"合同记录",
				"记录合同编号、合同类型、签订日期、起止日期、签订次数和附件。",
				["合同编号", "合同类型", "签订日期", "开始日期", "结束日期", "签订次数", "附件"],
				contract_items,
			),
		],
		"工资社保": [
			_make_related_record(
				"社保公积金记录",
				"",
				[],
				insurance_items,
				compact=True,
			),
		],
		"材料附件": [
			_make_related_record("员工基本资料", "身份证、学历证明、个人证件照等入职基础材料。", ["身份证照片", "学历证明", "个人证件照", "身份证复印件"], []),
			_make_related_record("员工档案资料", "劳动合同、入职简历、体检单、入职记录等档案材料。", ["劳动合同", "入职简历", "入职记录", "入职体检单"], []),
			_make_related_record("员工离职资料", "离职审批、离职证明、离职申请、工作交接表等材料。", ["离职审批", "离职证明", "离职申请", "工作交接表"], []),
		],
		"背景调查": [
			_make_related_record("人事异动记录", "展示员工转岗、调薪、转正、离职等关键人事动作。", ["类型", "日期", "办理人", "结果"], transfer_items + promotion_items + separation_items, "Employee Transfer", "办理人事异动"),
			_make_related_record("操作记录", "记录档案查看、字段变更、材料上传等操作留痕。", ["操作时间", "操作人", "操作类型", "说明"], []),
		],
	}

	for tab, configured_records in _get_configured_detail_block_records(doc).items():
		existing_labels = {record["label"] for record in records.get(tab, [])}
		for record in configured_records:
			if record["label"] not in existing_labels:
				records.setdefault(tab, []).append(record)

	return records


def _get_employee_growth_records(doc):
	"""Return submitted work-nature changes and confirmations for the timeline."""
	transfers = frappe.get_all(
		"Employee Transfer",
		filters={"employee": doc.name, "docstatus": 1},
		fields=["name", "transfer_date"],
		order_by="transfer_date asc, creation asc",
		limit_page_length=0,
	)
	transfer_dates = {row.name: row.transfer_date for row in transfers}
	changes = (
		frappe.get_all(
			"Employee Property History",
			filters={
				"parenttype": "Employee Transfer",
				"parent": ["in", list(transfer_dates)],
				"fieldname": "employment_type",
			},
			fields=["parent", "current", "new"],
			order_by="idx asc",
			limit_page_length=0,
		)
		if transfer_dates
		else []
	)
	records = [
		{
			"date": transfer_dates[row.parent],
			"title": "工作性质调整",
			"from_value": row.current,
			"to_value": row.new,
		}
		for row in changes
	]

	promotions = frappe.get_all(
		"Employee Promotion",
		filters={"employee": doc.name, "docstatus": 1, "custom_confirmation_result": "转正通过"},
		fields=["promotion_date", "custom_confirmation_interview_notes"],
		order_by="promotion_date asc, creation asc",
		limit_page_length=0,
	)
	for promotion in promotions:
		is_automatic = "系统根据转正日期自动办理转正" in str(promotion.custom_confirmation_interview_notes or "")
		records.append(
			{
				"date": promotion.promotion_date,
				"title": "自动转正" if is_automatic else "转正通过",
				"from_value": "在职 · 试用期",
				"to_value": "在职 · 正式",
			}
		)
	return sorted(records, key=lambda record: str(record.get("date") or ""))


@frappe.whitelist()
def get_employee_detail(employee: str):
	doc = frappe.get_doc(EMPLOYEE_DOCTYPE, employee)
	doc.check_permission("read")
	department_display = _department_display_name(doc.get("department"))
	return {
		"header": {
			"name": doc.name,
			"employee_name": doc.get("employee_name"),
			"custom_employee_code": doc.get("custom_employee_code"),
			"company": doc.get("company"),
			"department": doc.get("department"),
			"department_display": department_display,
			"designation": doc.get("designation"),
			"employment_type": doc.get("employment_type"),
			"status": doc.get("status"),
			"custom_is_confirmed": doc.get("custom_is_confirmed"),
			"final_confirmation_date": doc.get("final_confirmation_date"),
			"date_of_joining": doc.get("date_of_joining"),
			"gender": doc.get("gender"),
			"age": doc.get("age"),
			"cell_number": doc.get("cell_number"),
			"image": doc.get("image"),
		},
		"growth_records": _get_employee_growth_records(doc),
		"sections": _get_employee_detail_sections(doc, department_display),
		"materials": _get_employee_materials(doc),
		"related_records": _get_employee_related_records(doc),
		"permissions": {
			"can_edit_employee_detail": _can_edit_employee_detail(),
		},
	}


def _can_edit_employee_detail():
	user = frappe.session.user
	return user == "Administrator" or "System Manager" in frappe.get_roles(user)


def _get_employee_material_type_map():
	return {
		key: {"key": key, "label": label, "fieldname": f"{EMPLOYEE_MATERIAL_FIELD_PREFIX}{key}"}
		for group in EMPLOYEE_MATERIAL_GROUPS
		for key, label in group["types"]
	}


def _get_employee_materials(doc):
	"""Return employee-linked files grouped by their HR material type."""
	type_map = _get_employee_material_type_map()
	files_by_fieldname = {material["fieldname"]: [] for material in type_map.values()}
	files = frappe.get_all(
		"File",
		filters={"attached_to_doctype": EMPLOYEE_DOCTYPE, "attached_to_name": doc.name},
		fields=["name", "file_name", "file_url", "attached_to_field", "is_private", "modified"],
		order_by="modified desc",
	)
	for file in files:
		if file.attached_to_field not in files_by_fieldname:
			continue
		files_by_fieldname[file.attached_to_field].append(
			{
				"name": file.name,
				"file_name": file.file_name,
				"file_url": file.file_url,
				"is_private": file.is_private,
				"modified": file.modified,
			}
		)

	return [
		{
			"label": group["label"],
			"description": group["description"],
			"types": [
				{
					**type_map[key],
					"files": files_by_fieldname[type_map[key]["fieldname"]],
				}
				for key, _label in group["types"]
			],
		}
		for group in EMPLOYEE_MATERIAL_GROUPS
	]


@frappe.whitelist()
def upload_employee_material(employee: str, material_type: str, file_url: str):
	"""Classify an uploaded file as a durable employee archive material."""
	if not _can_edit_employee_detail():
		frappe.throw(_("只有管理员可以上传员工档案材料"), frappe.PermissionError)
	if not employee or not material_type or not file_url:
		frappe.throw(_("请选择材料类型和要上传的文件"))

	material = _get_employee_material_type_map().get(material_type)
	if not material:
		frappe.throw(_("员工材料类型不正确"))
	doc = frappe.get_doc(EMPLOYEE_DOCTYPE, employee)
	doc.check_permission("write")
	file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not file_name:
		frappe.throw(_("未找到已上传的材料文件"))

	file_doc = frappe.get_doc("File", file_name)
	file_doc.check_permission("read")
	is_attached_to_employee = (
		file_doc.attached_to_doctype == EMPLOYEE_DOCTYPE and file_doc.attached_to_name == doc.name
	)
	if not is_attached_to_employee and file_doc.owner != frappe.session.user:
		frappe.throw(_("只能归档当前登录用户上传的材料"), frappe.PermissionError)
	extension = os.path.splitext((file_doc.file_name or file_url).split("?", 1)[0])[1].lower()
	if extension not in EMPLOYEE_MATERIAL_FILE_EXTENSIONS:
		frappe.throw(_("仅支持 JPG、PNG、WebP 或 PDF 格式的员工材料"))

	file_doc.db_set("attached_to_doctype", EMPLOYEE_DOCTYPE)
	file_doc.db_set("attached_to_name", doc.name)
	file_doc.db_set("attached_to_field", material["fieldname"])
	return {"materials": _get_employee_materials(doc)}


@frappe.whitelist()
def update_employee_photo(employee: str, file_url: str):
	"""Save a verified image uploaded through the employee detail page."""
	if not _can_edit_employee_detail():
		frappe.throw(_("只有管理员可以上传员工照片"), frappe.PermissionError)
	if not employee or not file_url:
		frappe.throw(_("请先选择要上传的照片"))

	doc = frappe.get_doc(EMPLOYEE_DOCTYPE, employee)
	doc.check_permission("write")
	file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not file_name:
		frappe.throw(_("未找到已上传的照片文件"))

	file_doc = frappe.get_doc("File", file_name)
	file_doc.check_permission("read")
	is_attached_to_employee = (
		file_doc.attached_to_doctype == EMPLOYEE_DOCTYPE
		and file_doc.attached_to_name == doc.name
		and file_doc.attached_to_field == "image"
	)
	if not is_attached_to_employee and file_doc.owner != frappe.session.user:
		frappe.throw(_("只能使用当前登录用户上传的照片"), frappe.PermissionError)
	extension = os.path.splitext((file_doc.file_name or file_url).split("?", 1)[0])[1].lower()
	if extension not in {".jpg", ".jpeg", ".png", ".webp"}:
		frappe.throw(_("仅支持 JPG、PNG 或 WebP 格式的照片"))
	if file_doc.file_size and int(file_doc.file_size) > 5 * 1024 * 1024:
		frappe.throw(_("照片大小不能超过 5MB"))

	try:
		from PIL import Image, UnidentifiedImageError

		with Image.open(BytesIO(file_doc.get_content())) as image:
			image.verify()
	except (OSError, UnidentifiedImageError, ValueError):
		frappe.throw(_("上传的文件不是有效的图片"))

	# FileUploader normally creates this association. Set it explicitly as a
	# fallback so uploaded photos remain on the employee's attachment timeline.
	if (
		file_doc.attached_to_doctype != EMPLOYEE_DOCTYPE
		or file_doc.attached_to_name != doc.name
		or file_doc.attached_to_field != "image"
	):
		file_doc.db_set("attached_to_doctype", EMPLOYEE_DOCTYPE)
		file_doc.db_set("attached_to_name", doc.name)
		file_doc.db_set("attached_to_field", "image")

	doc.image = file_doc.file_url
	doc.save()
	return {"name": doc.name, "image": doc.image}


@frappe.whitelist()
def get_employee_detail_navigation(employee: str, filters: str = "{}"):
	names = frappe.get_list(
		EMPLOYEE_DOCTYPE,
		filters=_build_employee_roster_filters(filters),
		pluck="name",
		order_by="modified desc",
		limit_page_length=0,
	)
	if employee not in names:
		return {"previous": None, "next": None}
	index = names.index(employee)
	return {
		"previous": names[index - 1] if index > 0 else None,
		"next": names[index + 1] if index + 1 < len(names) else None,
	}


@frappe.whitelist()
def parse_employee_roster_file(file_url: str):
	context = _get_uploaded_roster_context(file_url)

	return {
		"headers": context["matches"],
		"fields": context["fields"],
		"missing_required": context["missing_required"],
		"row_count": _count_data_rows(context["rows"], context["data_start_index"]),
		"can_import": not context["missing_required"],
	}


def _clean_import_value(value):
	if value is None:
		return None
	if isinstance(value, str):
		value = value.strip()
		upper_value = value.upper()
		if upper_value in EXCEL_ERROR_VALUES or any(upper_value.startswith(error_value) for error_value in EXCEL_ERROR_VALUES):
			return None
		if value.startswith("="):
			return None
		return value or None
	if isinstance(value, float) and value.is_integer():
		return int(value)
	return value


def _is_excel_error_or_formula(value):
	if not isinstance(value, str):
		return False
	value = value.strip()
	upper_value = value.upper()
	return value.startswith("=") or upper_value in EXCEL_ERROR_VALUES or any(
		upper_value.startswith(error_value) for error_value in EXCEL_ERROR_VALUES
	)


def _display_employee_import_value(value):
	if value is None:
		return ""
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()
	return str(value).strip()


def _field_error(row_index, field, message, suggestion=None, excel_cell=None, current_value=None):
	error = {
		"row": row_index,
		"fieldname": field.get("fieldname"),
		"field_label": field.get("field_label") or field.get("fieldname"),
		"fieldtype": field.get("fieldtype") or "Data",
		"message": message,
	}
	if suggestion:
		error["suggestion"] = suggestion
	if excel_cell:
		error["excel_cell"] = excel_cell
	if current_value is not None:
		error["current_value"] = _display_employee_import_value(current_value)
	return error


def _excel_cell_reference(row_index, column_index):
	"""Return a user-facing Excel cell reference from a zero-based column index."""
	if row_index is None or column_index is None:
		return ""
	column_number = int(column_index) + 1
	letters = ""
	while column_number:
		column_number, remainder = divmod(column_number - 1, 26)
		letters = chr(65 + remainder) + letters
	return f"{letters}{row_index}"


def _employee_import_fix_suggestion(row_index, field, issue):
	fieldname = field.get("fieldname")
	field_label = field.get("field_label") or fieldname or _("该字段")
	location = _("请修改 Excel 第 {0} 行“{1}”列：").format(row_index or "", field_label)
	if issue == "required":
		if fieldname == "cell_number":
			return location + _("这是必填项，请填写有效手机号（11 位大陆手机号）或 7 至 15 位座机号码后重新上传。")
		return location + _("这是必填项，请补充有效内容后重新上传。")
	if issue == "date":
		return location + _("请填写实际日期，例如 2027-12-31；若该项暂无资料且非必填，可直接清空该单元格。")
	if issue == "phone":
		if fieldname == "emergency_phone_number":
			return location + _("此列只能填写电话。若当前填写的是联系人姓名，请移到“紧急联系”列；没有电话号码时可清空本单元格。")
		return location + _("请填写有效手机号（11 位大陆手机号）或 7 至 15 位座机号码。")
	if issue == "excel_error":
		return location + _("请删除公式或 Excel 错误值，直接填写最终的文本或日期。")
	return location + _("请修正后重新上传。")


def _is_employee_import_deferred_placeholder(value):
	return isinstance(value, str) and value.strip() in {"-", "—"}


def _can_defer_employee_import_field(fieldname):
	return fieldname not in EMPLOYEE_IMPORT_NON_DEFERRABLE_FIELDS


def _date_to_iso_if_reasonable(value):
	if isinstance(value, datetime):
		value = value.date()
	if not isinstance(value, date):
		return None
	if value.year < MIN_REASONABLE_EMPLOYEE_DATE_YEAR:
		return None
	return value.isoformat()


def _normalise_date_value(value):
	value = _clean_import_value(value)
	if not value:
		return None
	if isinstance(value, datetime):
		return _date_to_iso_if_reasonable(value)
	if isinstance(value, date):
		return _date_to_iso_if_reasonable(value)
	if isinstance(value, (int, float)):
		try:
			from openpyxl.utils.datetime import from_excel

			return _date_to_iso_if_reasonable(from_excel(value).date())
		except Exception:
			return None
	text = str(value).replace("/", "-").strip()
	try:
		return _date_to_iso_if_reasonable(frappe.utils.getdate(text))
	except Exception:
		return None


def _normalise_number_value(value, fieldtype):
	value = _clean_import_value(value)
	if value is None:
		return None
	try:
		if fieldtype == "Int":
			return frappe.utils.cint(value)
		return float(value)
	except Exception:
		return value


def _reverse_option_label(value, fieldname):
	value = _clean_import_value(value)
	if value is None:
		return None
	text = str(value).strip()
	reverse = {label: option for option, label in OPTION_LABEL_MAP.items()}
	if text in reverse:
		return reverse[text]

	field_options = {
		"gender": {"男": "Male", "女": "Female", "其他": "Other"},
		"status": {"激活": "Active", "在职": "Active", "正式": "Active", "非激活": "Inactive", "停职": "Suspended", "离职": "Left", "已离职": "Left"},
		"employment_type": {
			"全职": "Full-time",
			"正式": "Full-time",
			"在职": "Full-time",
			"兼职": "Part-time",
			"实习": "Intern",
			"实习生": "Intern",
			"外包": "Contract",
			"试用": "Probation",
			"试用期": "Probation",
			"返聘": "Retainer",
			"退休返聘": "Retainer",
		},
		"salary_mode": {"银行": "Bank", "现金": "Cash", "支票": "Cheque"},
		"custom_marital_status_text": {
			"未": "未",
			"未婚": "未",
			"已": "已",
			"已婚": "已",
			"离": "离异",
			"离异": "离异",
			"丧": "丧偶",
			"丧偶": "丧偶",
			"分居": None,
		},
	}
	return field_options.get(fieldname, {}).get(text, text)


def _normalise_phone_value(value):
	value = _clean_import_value(value)
	if value is None:
		return None
	text = str(value).strip()
	if not text:
		return None

	candidates = re.findall(r"(?:\+?86[-\s]?)?1[3-9]\d{9}|0\d{2,3}[-\s]?\d{7,8}|\d{7,8}", text)
	if candidates:
		phone = re.sub(r"[\s-]", "", candidates[0])
		if phone.startswith("86") and len(phone) == 13:
			phone = phone[2:]
		if re.fullmatch(r"1[3-9]\d{9}", phone):
			return f"+86{phone}"
		return phone

	digits = re.sub(r"\D", "", text)
	if 7 <= len(digits) <= 15:
		return digits
	return None


def _normalise_gender_value(value):
	value = _clean_import_value(value)
	if value is None:
		return None
	text = str(value).strip()
	return GENDER_VALUE_ALIASES.get(text, GENDER_VALUE_ALIASES.get(text.lower(), text))


def _derive_identity_card_values(identity_card_number, today=None):
	"""Derive only non-sensitive structured values from a mainland China ID number."""
	identity_card_number = re.sub(r"\s+", "", str(identity_card_number or "")).upper()
	if not re.fullmatch(r"\d{17}[0-9X]", identity_card_number):
		return {}

	try:
		birth_date = datetime.strptime(identity_card_number[6:14], "%Y%m%d").date()
	except ValueError:
		return {}

	today = today or frappe.utils.getdate()
	if birth_date > today:
		return {}

	age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
	return {
		"date_of_birth": birth_date.isoformat(),
		"gender": "Male" if int(identity_card_number[16]) % 2 else "Female",
		"custom_age": age,
	}


def _apply_identity_card_derivatives(values, warnings, row_index=None):
	derived_values = _derive_identity_card_values(values.get("passport_number"))
	if not derived_values:
		return

	if not values.get("custom_id_type"):
		values["custom_id_type"] = "身份证"

	derived_labels = {"date_of_birth": "出生年月", "gender": "性别", "custom_age": "年龄"}
	for fieldname, derived_value in derived_values.items():
		current_value = values.get(fieldname)
		if current_value in (None, ""):
			values[fieldname] = derived_value
			warnings.append(
				_("第 {0} 行：已根据身份证号码补充{1}，未覆盖人工填写值。").format(
					row_index or "", derived_labels[fieldname]
				)
			)
		elif str(current_value) != str(derived_value):
			warnings.append(
				_("第 {0} 行：身份证推导的{1}与人工填写值不一致，已保留人工填写值。").format(
					row_index or "", derived_labels[fieldname]
				)
			)


def _normalise_import_value(fieldname, value, field):
	value = _clean_import_value(value)
	if value is None:
		return None

	# The source workbook uses “在职” for this column.  It is not persisted as
	# a second value: whether the employee is shown as 正式 or 试用期 is decided
	# by the accompanying “是否转正” field.
	if fieldname == "employment_type":
		return _reverse_option_label(value, fieldname)

	if fieldname == "gender":
		return _normalise_gender_value(value)

	if fieldname in PHONE_FIELDNAMES:
		return _normalise_phone_value(value)

	fieldtype = field.get("fieldtype")
	if fieldtype in {"Date", "Datetime"}:
		return _normalise_date_value(value)
	if fieldtype in {"Int", "Float", "Currency"}:
		return _normalise_number_value(value, fieldtype)
	if fieldtype == "Check":
		return 1 if str(value).strip().lower() in {"1", "true", "yes", "y", "是", "已", "有"} else 0
	if fieldtype == "Select":
		return _reverse_option_label(value, fieldname)
	if isinstance(value, float) and value.is_integer():
		return str(int(value))
	return str(value).strip()


def _get_default_company():
	default_company = frappe.defaults.get_user_default("Company") or frappe.defaults.get_global_default("company")
	if default_company and frappe.db.exists("Company", default_company):
		return default_company
	return frappe.db.get_value("Company", {}, "name")


def _resolve_company(value, default_company, warnings):
	value = _clean_import_value(value)
	if value:
		text = str(value).strip()
		company = frappe.db.exists("Company", text) or frappe.db.get_value("Company", {"company_name": text}, "name")
		if company:
			return company
		warnings.append(_("公司“{0}”不存在，已使用默认公司“{1}”。").format(text, default_company or ""))
	return default_company


def _resolve_roster_department(value, company):
	"""Resolve a roster value to an existing, explicitly assignable leaf node.

	Departments are the organization source of truth.  A roster import may never
	create one from free text because that would turn an accidental group label
	into a payroll and reporting dimension.
	"""
	value = _clean_import_value(value)
	if not value:
		return None, ""
	department_name = _strip_department_company_suffix(value)
	existing = (
		frappe.db.get_value("Department", {"department_name": department_name, "company": company}, "name")
		or frappe.db.get_value("Department", {"department_name": department_name}, "name")
		or frappe.db.exists("Department", str(value).strip())
	)
	if not existing:
		return None, _("部门“{0}”不存在；请先在部门管理中建立并同步组织层级。").format(department_name)

	department = frappe.get_cached_doc("Department", existing)
	if company and department.company and department.company != company:
		return None, _("部门“{0}”不属于当前公司。").format(department_name)
	if cint(department.is_group):
		return None, _("部门“{0}”是文件夹节点，花名册只能选择最末级组织。").format(department_name)
	if frappe.get_meta("Department").has_field("hrms_roster_assignable") and not cint(
		department.get("hrms_roster_assignable")
	):
		return None, _("部门“{0}”尚未启用花名册归属。").format(department_name)
	return department.name, ""


def _find_or_create_department(value, company, base_records):
	"""Compatibility wrapper retained for existing import integrations.

	The historical helper name remains callable, but its unsafe create behaviour
	is deliberately removed.  ``base_records['部门']`` must stay zero.
	"""
	department, error = _resolve_roster_department(value, company)
	if error:
		frappe.throw(error)
	return department


def _find_or_create_designation(value, base_records):
	value = _clean_import_value(value)
	if not value:
		return None
	designation_name = str(value).strip()
	existing = frappe.db.exists("Designation", designation_name) or frappe.db.get_value(
		"Designation", {"designation_name": designation_name}, "name"
	)
	if existing:
		return existing
	doc = frappe.get_doc({"doctype": "Designation", "designation_name": designation_name})
	doc.insert(ignore_permissions=True)
	base_records["岗位"] = base_records.get("岗位", 0) + 1
	return doc.name


def _find_or_create_employment_type(value, base_records):
	value = _clean_import_value(value)
	if not value or not frappe.db.exists("DocType", "Employment Type"):
		return value
	employment_type = str(value).strip()
	existing = frappe.db.exists("Employment Type", employment_type)
	if existing:
		return existing
	doc = frappe.get_doc({"doctype": "Employment Type", "employee_type_name": employment_type})
	doc.insert(ignore_permissions=True)
	base_records["工作性质"] = base_records.get("工作性质", 0) + 1
	return doc.name


def _find_or_create_gender(value, base_records):
	gender = _normalise_gender_value(value)
	if not gender or not frappe.db.exists("DocType", "Gender"):
		return gender

	existing = frappe.db.exists("Gender", gender) or frappe.db.get_value("Gender", {"gender": gender}, "name")
	if existing:
		return existing

	doc = frappe.get_doc({"doctype": "Gender", "gender": gender})
	doc.insert(ignore_permissions=True)
	base_records["性别"] = base_records.get("性别", 0) + 1
	return doc.name


def _ensure_employee_base_records(values, base_records, warnings):
	default_company = _get_default_company()
	values["company"] = _resolve_company(values.get("company"), default_company, warnings)
	if values.get("gender"):
		values["gender"] = _find_or_create_gender(values["gender"], base_records)
	if values.get("department") and values.get("company"):
		values["department"] = _find_or_create_department(values["department"], values["company"], base_records)
	if values.get("designation"):
		values["designation"] = _find_or_create_designation(values["designation"], base_records)
	if values.get("employment_type"):
		values["employment_type"] = _find_or_create_employment_type(values["employment_type"], base_records)


def _drop_invalid_employee_date_ranges(values, warnings, row_index):
	if values.get("date_of_birth"):
		try:
			date_of_birth = frappe.utils.getdate(values["date_of_birth"])
			today = frappe.utils.getdate()
			date_of_joining = frappe.utils.getdate(values["date_of_joining"]) if values.get("date_of_joining") else None
			if date_of_birth > today or (date_of_joining and date_of_birth >= date_of_joining):
				skipped_value = values.pop("date_of_birth")
				warnings.append(
					_("第 {0} 行：出生年月“{1}”不合理，已跳过该字段。").format(row_index, skipped_value)
				)
		except Exception:
			skipped_value = values.pop("date_of_birth", None)
			if skipped_value:
				warnings.append(
					_("第 {0} 行：出生年月“{1}”无法识别，已跳过该字段。").format(row_index, skipped_value)
				)

	if values.get("contract_end_date") and values.get("date_of_joining"):
		try:
			if frappe.utils.getdate(values["contract_end_date"]) <= frappe.utils.getdate(values["date_of_joining"]):
				skipped_value = values.pop("contract_end_date")
				warnings.append(
					_("第 {0} 行：合同结束日期“{1}”早于或等于入职日期，已跳过该字段。").format(
						row_index, skipped_value
					)
				)
		except Exception:
			skipped_value = values.pop("contract_end_date", None)
			if skipped_value:
				warnings.append(
					_("第 {0} 行：合同结束日期“{1}”无法识别，已跳过该字段。").format(row_index, skipped_value)
				)

	if values.get("final_confirmation_date") and values.get("date_of_joining"):
		try:
			if frappe.utils.getdate(values["final_confirmation_date"]) < frappe.utils.getdate(values["date_of_joining"]):
				skipped_value = values.pop("final_confirmation_date")
				warnings.append(
					_("第 {0} 行：转正日期“{1}”早于入职日期，已跳过该字段。").format(row_index, skipped_value)
				)
		except Exception:
			skipped_value = values.pop("final_confirmation_date", None)
			if skipped_value:
				warnings.append(
					_("第 {0} 行：转正日期“{1}”无法识别，已跳过该字段。").format(row_index, skipped_value)
				)


def _find_existing_employee(values, meta_fields, company=None):
	for fieldname in ("custom_employee_code", "passport_number"):
		if fieldname in meta_fields and values.get(fieldname):
			filters = {fieldname: values[fieldname]}
			if company:
				filters["company"] = company
			existing = frappe.db.get_value(EMPLOYEE_DOCTYPE, filters, "name")
			if existing:
				return existing
	return None


def _find_existing_employee_by_strategy(values, meta_fields, match_by="employee_code", company=None):
	match_by = match_by if match_by in EMPLOYEE_DUPLICATE_MATCH_FIELDS else "employee_code"
	for fieldname in EMPLOYEE_DUPLICATE_MATCH_FIELDS[match_by]:
		if fieldname in meta_fields and values.get(fieldname):
			filters = {fieldname: values[fieldname]}
			if company:
				filters["company"] = company
			existing = frappe.db.get_value(EMPLOYEE_DOCTYPE, filters, "name")
			if existing:
				return existing
	return None


def _row_to_employee_values(row, matches, fields_by_name, warnings, row_index=None, row_overrides=None):
	values = {}
	errors = []
	row_overrides = row_overrides if isinstance(row_overrides, dict) else {}
	for match in matches:
		fieldname = match.get("fieldname")
		column_index = match.get("column_index")
		if not fieldname or column_index is None or column_index >= len(row):
			continue
		field = fields_by_name.get(fieldname)
		if not field:
			continue
		raw_value = row_overrides[fieldname] if fieldname in row_overrides else row[column_index]
		if _is_employee_import_deferred_placeholder(raw_value) and _can_defer_employee_import_field(fieldname):
			values.setdefault("_employee_import_deferred_fields", set()).add(fieldname)
			continue
		value = _normalise_import_value(fieldname, raw_value, field)
		excel_cell = _excel_cell_reference(row_index, column_index)
		if _is_blank_value(raw_value) and _is_employee_import_required_field(fieldname, field):
			errors.append(
				_field_error(
					row_index,
					field,
					_("必填字段为空"),
					_employee_import_fix_suggestion(row_index, field, "required"),
					excel_cell,
					raw_value,
				)
			)
		elif not _is_blank_value(raw_value) and value is None:
			if _is_excel_error_or_formula(raw_value):
				errors.append(
					_field_error(
						row_index,
						field,
						_("单元格包含 Excel 错误值或公式结果不可用"),
						_employee_import_fix_suggestion(row_index, field, "excel_error"),
						excel_cell,
						raw_value,
					)
				)
			elif field.get("fieldtype") in {"Date", "Datetime"}:
				errors.append(
					_field_error(
						row_index,
						field,
						_("日期无法识别"),
						_employee_import_fix_suggestion(row_index, field, "date"),
						excel_cell,
						raw_value,
					)
				)
			elif fieldname in PHONE_FIELDNAMES:
				errors.append(
					_field_error(
						row_index,
						field,
						_("手机号无法识别"),
						_employee_import_fix_suggestion(row_index, field, "phone"),
						excel_cell,
						raw_value,
					)
				)
		if value is not None:
			values[fieldname] = value

	if values.get("first_name") and not values.get("employee_name"):
		values["employee_name"] = values["first_name"]
	if values.get("employee_name") and not values.get("first_name"):
		values["first_name"] = values["employee_name"]
	_apply_identity_card_derivatives(values, warnings, row_index)
	if not values.get("status"):
		values["status"] = "Active"
	if not values.get("date_of_birth") and fields_by_name.get("date_of_birth"):
		values["date_of_birth"] = EMPLOYEE_FALLBACK_DATE_OF_BIRTH
		warnings.append(
			_("第 {0} 行：出生年月为空或无法识别，已临时使用 {1}，请后续补正。").format(
				row_index or "", EMPLOYEE_FALLBACK_DATE_OF_BIRTH
			)
		)

	# 批量导入只写员工资料，不默认创建系统用户，避免因为邮箱缺失阻断花名册导入。
	values["create_user_automatically"] = 0
	values["create_user_permission"] = 0
	return values, errors


def _validate_employee_import_row(values, fields_by_name, meta_fields, row_index, parse_errors=None):
	errors = []
	deferred_fields = values.get("_employee_import_deferred_fields") or set()
	parse_error_fields = {error.get("fieldname") for error in parse_errors or []}
	for fieldname, field in fields_by_name.items():
		if (
			_is_employee_import_required_field(fieldname, field)
			and fieldname in meta_fields
			and fieldname not in deferred_fields
			and fieldname not in parse_error_fields
			and _is_blank_value(values.get(fieldname))
		):
			errors.append(_field_error(row_index, field, _("必填字段为空")))

	if not values.get("first_name") and "first_name" not in fields_by_name:
		name_field = fields_by_name.get("first_name") or fields_by_name.get("employee_name") or {
			"fieldname": "first_name",
			"field_label": _("姓名"),
		}
		errors.append(_field_error(row_index, name_field, _("缺少员工姓名")))

	if values.get("department") and "department" not in deferred_fields:
		department, error = _resolve_roster_department(values.get("department"), values.get("company"))
		if error:
			department_field = fields_by_name.get("department") or {
				"fieldname": "department",
				"field_label": _("部门"),
			}
			errors.append(_field_error(row_index, department_field, error))
		else:
			values["department"] = department

	return errors


def _preview_employee_action(values, meta_fields, mode, match_by):
	existing = _find_existing_employee_by_strategy(values, meta_fields, match_by, values.get("company"))
	if existing:
		if mode == "insert":
			return "skip", existing
		return "update", existing
	if mode == "update":
		return "skip", None
	return "insert", None


def _get_employee_roster_replace_candidates(planned_rows, company=""):
	"""Return only same-company staff omitted from a verified full-roster import.

	A first import contains only ``insert`` actions, so its new employee names do
	not exist when the preview is created.  Compare both the resolved employee
	IDs and business codes, otherwise those new rows are incorrectly archived at
	the end of the same import.  The company filter is mandatory: one company's
	full-roster replacement must never change another company's employees.
	"""
	company = company or _get_default_company()
	if not company:
		return []
	imported_employee_names = {row["existing"] for row in planned_rows if row.get("existing")}
	imported_employee_codes = {
		str(row.get("values", {}).get("custom_employee_code") or "").strip()
		for row in planned_rows
		if row.get("values", {}).get("custom_employee_code")
	}
	current_employees = frappe.get_all(
		EMPLOYEE_DOCTYPE,
		filters={"company": company, "status": ["!=", "Left"]},
		fields=["name", "custom_employee_code"],
		limit_page_length=0,
	)
	return [
		row.name
		for row in current_employees
		if row.name not in imported_employee_names
		and str(row.custom_employee_code or "").strip() not in imported_employee_codes
	]


def _get_employee_roster_preview_code(values, existing=None):
	"""Return the roster-facing employee code, never the internal Employee name."""
	code = values.get("custom_employee_code")
	if code:
		return str(code)
	if not existing:
		return ""

	employee = frappe.db.get_value(EMPLOYEE_DOCTYPE, existing, "custom_employee_code")
	return str(employee or "")


def _make_failed_row(row_index, field_error, source_row):
	return {
		"row": row_index,
		"fieldname": field_error.get("fieldname"),
		"field_label": field_error.get("field_label"),
		"message": field_error.get("message"),
		"suggestion": field_error.get("suggestion"),
		"excel_cell": field_error.get("excel_cell"),
		"values": ["" if value is None else value for value in source_row],
	}


def _dedupe_import_errors(errors):
	seen = set()
	deduped = []
	for error in errors:
		key = (error.get("row"), error.get("fieldname"), error.get("message"))
		if key in seen:
			continue
		seen.add(key)
		deduped.append(error)
	return deduped


def _build_employee_roster_import_plan(
	file_url, mode="insert", match_by="employee_code", manual_mappings=None, row_overrides=None
):
	mode = mode or "insert"
	if mode not in {"insert", "update", "replace"}:
		frappe.throw(_("导入模式不正确"))
	if match_by not in EMPLOYEE_DUPLICATE_MATCH_FIELDS:
		frappe.throw(_("重复员工匹配策略不正确"))

	context = _apply_manual_header_mappings(_get_uploaded_roster_context(file_url), manual_mappings)
	if context["missing_required"]:
		frappe.throw(
			_("必填字段尚未匹配：{0}").format(
				"、".join(field["field_label"] for field in context["missing_required"])
			)
		)

	fields_by_name = {field["fieldname"]: field for field in context["fields"]}
	meta_fields = _get_employee_meta_field_map()
	row_overrides = _parse_json(row_overrides, {}) or {}
	if not isinstance(row_overrides, dict):
		frappe.throw(_("人工校正数据格式不正确"))
	result = {
		"inserted": 0,
		"updated": 0,
		"skipped": 0,
		"archived": 0,
		"failed": 0,
		"row_count": 0,
		"base_records": {"性别": 0, "部门": 0, "岗位": 0, "工作性质": 0},
		"errors": [],
		"warnings": [],
		"failed_rows": [],
		"preview_rows": [],
		"deferred": 0,
		"manual_corrections": sum(
			len(overrides) for overrides in row_overrides.values() if isinstance(overrides, dict)
		),
	}
	planned_rows = []

	for row_index, row in enumerate(context["rows"][context["data_start_index"] :], start=context["data_start_index"] + 1):
		if not any(not _is_blank_value(value) for value in row):
			continue
		result["row_count"] += 1
		row_override = row_overrides.get(str(row_index), row_overrides.get(row_index, {}))
		values, parse_errors = _row_to_employee_values(
			row,
			context["matches"],
			fields_by_name,
			result["warnings"],
			row_index,
			row_override,
		)
		# Resolve the company before duplicate matching and replacement preview.
		# A blank company column means the selected/default company, never a
		# cross-company search of an administrator's entire employee table.
		values["company"] = _resolve_company(values.get("company"), _get_default_company(), result["warnings"])
		row_errors = _dedupe_import_errors(
			parse_errors + _validate_employee_import_row(values, fields_by_name, meta_fields, row_index, parse_errors)
		)
		action, existing = _preview_employee_action(values, meta_fields, mode, match_by)

		if mode == "update" and action == "skip" and not row_errors:
			row_errors.append(
				{
					"row": row_index,
					"fieldname": match_by,
					"field_label": _("重复匹配"),
					"message": _("未找到可更新的员工"),
				}
			)

		if row_errors:
			result["failed"] += 1
			for error in row_errors:
				result["errors"].append(error)
				result["failed_rows"].append(_make_failed_row(row_index, error, row))
			result["preview_rows"].append(
				{
					"row": row_index,
					"action": "failed",
					"employee_code": _get_employee_roster_preview_code(values, existing),
					"errors": row_errors,
				}
			)
			continue

		if action == "update":
			result["updated"] += 1
		elif action == "skip":
			result["skipped"] += 1
		else:
			result["inserted"] += 1
		result["deferred"] += len(values.get("_employee_import_deferred_fields") or [])

		result["preview_rows"].append(
			{
				"row": row_index,
				"action": action,
				"employee_code": _get_employee_roster_preview_code(values, existing),
				"errors": [],
			}
		)
		planned_rows.append({"row_index": row_index, "row": row, "values": values, "action": action, "existing": existing})

	if mode == "replace":
		target_companies = {row["values"].get("company") for row in planned_rows if row["values"].get("company")}
		if len(target_companies) > 1:
			frappe.throw(_("覆盖当前花名册一次只能处理一个公司，请按公司分别导入。"))
		result["archived"] = len(_get_employee_roster_replace_candidates(planned_rows, next(iter(target_companies), "")))

	return result, planned_rows, meta_fields


@frappe.whitelist()
def preview_employee_roster_import(
	file_url: str,
	mode: str = "insert",
	match_by: str = "employee_code",
	manual_mappings: str = "{}",
	row_overrides: str = "{}",
):
	result, _planned_rows, _meta_fields = _build_employee_roster_import_plan(
		file_url, mode, match_by, manual_mappings, row_overrides
	)
	result["can_import"] = not result["failed"]
	result["failed_rows_key"] = _store_employee_roster_failed_rows(result["failed_rows"])
	return result


@frappe.whitelist()
def import_employee_roster(
	file_url: str,
	mode: str = "insert",
	match_by: str = "employee_code",
	manual_mappings: str = "{}",
	row_overrides: str = "{}",
):
	preview_result, planned_rows, meta_fields = _build_employee_roster_import_plan(
		file_url, mode, match_by, manual_mappings, row_overrides
	)
	if mode == "replace" and preview_result["failed"]:
		frappe.throw(_("覆盖当前花名册前，请先修正所有错误行"))

	result = {
		**preview_result,
		"inserted": 0,
		"updated": 0,
		"skipped": preview_result["skipped"],
		"archived": 0,
		"base_records": {"性别": 0, "部门": 0, "岗位": 0, "工作性质": 0},
	}

	for planned_row in planned_rows:
		row_index = planned_row["row_index"]
		values = planned_row["values"]
		try:
			_drop_invalid_employee_date_ranges(values, result["warnings"], row_index)
			_ensure_employee_base_records(values, result["base_records"], result["warnings"])
			if planned_row["action"] == "update":
				doc = frappe.get_doc(EMPLOYEE_DOCTYPE, planned_row["existing"])
				for fieldname, value in values.items():
					if fieldname in meta_fields:
						doc.set(fieldname, value)
				doc.save(ignore_permissions=True)
				result["updated"] += 1
			elif planned_row["action"] == "insert":
				doc = frappe.new_doc(EMPLOYEE_DOCTYPE)
				for fieldname, value in values.items():
					if fieldname in meta_fields:
						doc.set(fieldname, value)
				doc.insert(ignore_permissions=True)
				result["inserted"] += 1
		except Exception as exc:
			frappe.log_error(frappe.get_traceback(), _("员工花名册导入失败"))
			result["failed"] += 1
			error = {
				"row": row_index,
				"fieldname": "",
				"field_label": _("整行"),
				"message": str(exc),
			}
			result["errors"].append(error)
			result["failed_rows"].append(_make_failed_row(row_index, error, planned_row["row"]))

	if mode == "replace" and not result["failed"]:
		target_companies = {row["values"].get("company") for row in planned_rows if row["values"].get("company")}
		for employee_name in _get_employee_roster_replace_candidates(planned_rows, next(iter(target_companies), "")):
			try:
				doc = frappe.get_doc(EMPLOYEE_DOCTYPE, employee_name)
				doc.status = "Left"
				doc.relieving_date = frappe.utils.today()
				doc.save(ignore_permissions=True)
				result["archived"] += 1
			except Exception as exc:
				frappe.log_error(frappe.get_traceback(), _("员工花名册覆盖失败"))
				result["failed"] += 1
				result["errors"].append(
					{
						"row": "",
						"fieldname": "",
						"field_label": _("覆盖当前花名册"),
						"message": _("员工 {0} 标记为已离职失败：{1}").format(employee_name, exc),
					}
				)

	frappe.db.commit()
	result["can_import"] = not result["failed"]
	result["failed_rows_key"] = _store_employee_roster_failed_rows(result["failed_rows"])
	return result


def _make_employee_roster_failure_workbook(failed_rows):
	from openpyxl import Workbook
	from hrms.utils.export_watermark import save_workbook_with_logo_watermark

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "失败行"
	rows = [["行号", "Excel 位置", "字段名", "字段", "问题", "修改建议", "原始行数据"]]
	for failed_row in failed_rows:
		rows.append(
			[
				failed_row.get("row"),
				failed_row.get("excel_cell"),
				failed_row.get("fieldname"),
				failed_row.get("field_label"),
				failed_row.get("message"),
				failed_row.get("suggestion"),
				" | ".join(str(value) for value in failed_row.get("values") or []),
			]
		)
	_write_sheet_rows(sheet, rows)

	output = BytesIO()
	save_workbook_with_logo_watermark(workbook, output)
	return output.getvalue()


def _store_employee_roster_failed_rows(failed_rows):
	if not failed_rows:
		return ""
	key = f"employee_roster_failed_rows:{frappe.generate_hash(length=16)}"
	frappe.cache().set_value(key, failed_rows, expires_in_sec=60 * 60)
	return key


@frappe.whitelist()
def download_employee_roster_failed_rows(failed_rows: str = "[]", failed_rows_key: str = ""):
	from frappe.desk.utils import provide_binary_file

	if failed_rows_key:
		failed_rows = frappe.cache().get_value(failed_rows_key) or []
	else:
		failed_rows = _parse_json(failed_rows, [])
	if not failed_rows:
		frappe.throw(_("没有可下载的失败行"))

	provide_binary_file(
		EMPLOYEE_FAILED_ROWS_FILENAME.removesuffix(".xlsx"),
		"xlsx",
		_make_employee_roster_failure_workbook(failed_rows),
	)


def _get_allowed_export_fields(fields):
	meta_fields = _get_employee_meta_field_map()
	allowed = {}
	for field in fields:
		if field["fieldname"] in meta_fields and field.get("export_enabled", 1):
			allowed[field["fieldname"]] = field
	return allowed


def _parse_selected_fields(fields, allowed_fields):
	fields = _parse_json(fields, [])
	if not fields:
		fields = [
			fieldname
			for fieldname, field in allowed_fields.items()
			if field["required"] or field.get("enabled")
		]
	selected = []
	for fieldname in fields:
		if fieldname in allowed_fields and fieldname not in selected:
			selected.append(fieldname)
	return selected


def _parse_selected_tables(tables):
	tables = _parse_json(tables, [])
	allowed = {row["label"] for row in MULTI_RECORD_EXPORT_CATEGORIES}
	selected = []
	for table in tables:
		if table in allowed and table not in selected:
			selected.append(table)
	return selected


def _safe_sheet_title(title, used_titles):
	title = re.sub(r"[\[\]\*\?/\\:]", "-", title or "工作表")[:31]
	if title not in used_titles:
		used_titles.add(title)
		return title

	base = title[:28]
	index = 2
	while f"{base}-{index}" in used_titles:
		index += 1
	title = f"{base}-{index}"[:31]
	used_titles.add(title)
	return title


def _write_sheet_rows(sheet, rows):
	from openpyxl.styles import Font, PatternFill
	from openpyxl.utils import get_column_letter

	header_fill = PatternFill("solid", fgColor="F6F7F9")
	header_font = Font(bold=True)
	for row_index, row in enumerate(rows, start=1):
		for column_index, value in enumerate(row, start=1):
			cell = sheet.cell(row=row_index, column=column_index, value=value)
			if row_index == 1:
				cell.fill = header_fill
				cell.font = header_font

	for column_index, column_cells in enumerate(sheet.columns, start=1):
		max_width = max(len(str(cell.value or "")) for cell in column_cells)
		sheet.column_dimensions[get_column_letter(column_index)].width = max(12, min(32, max_width + 4))


def _get_child_export_fields(child_doctype):
	return [
		field
		for field in frappe.get_meta(child_doctype).fields
		if field.fieldname
		and field.label
		and field.fieldtype not in NON_CONFIGURABLE_FIELDTYPES
		and field.fieldname not in {"name", "parent", "parenttype", "parentfield", "idx"}
	]


def _format_employee_export_value(fieldname, value, department_names):
	if fieldname == "department":
		return _department_display_name(value, department_names)
	return value


def _make_employee_export_workbook(selected_fields, allowed_fields, selected_tables, filters=None):
	from openpyxl import Workbook
	from hrms.utils.export_watermark import save_workbook_with_logo_watermark

	filters = filters or {}
	workbook = Workbook()
	used_titles = set()
	main_sheet = workbook.active
	main_sheet.title = _safe_sheet_title("员工花名册", used_titles)

	main_headers = [allowed_fields[fieldname]["field_label"] for fieldname in selected_fields]
	main_rows = [main_headers]
	employees = frappe.get_all(EMPLOYEE_DOCTYPE, filters=filters, fields=selected_fields, order_by="modified desc")
	department_names = _get_department_display_names([employee.get("department") for employee in employees])
	for employee in employees:
		main_rows.append(
			[_format_employee_export_value(fieldname, employee.get(fieldname), department_names) for fieldname in selected_fields]
		)
	_write_sheet_rows(main_sheet, main_rows)

	employee_rows = frappe.get_all(
		EMPLOYEE_DOCTYPE,
		filters=filters,
		fields=["name", "employee_name", "custom_employee_code"],
		order_by="modified desc",
	)
	employee_label = {row.name: row.employee_name for row in employee_rows}
	employee_code = {row.name: row.custom_employee_code or "" for row in employee_rows}
	employee_names = list(employee_label)

	for table_label in selected_tables:
		sheet = workbook.create_sheet(_safe_sheet_title(table_label, used_titles))
		child_fieldname = MULTI_RECORD_EXPORT_TABLE_MAP.get(table_label)
		if not child_fieldname:
			_write_sheet_rows(
				sheet,
				[
					["说明"],
					[f"{table_label} 暂未映射到 Employee 标准子表，后续可在员工属性设置中接入对应数据表。"],
				],
			)
			continue

		table_field = frappe.get_meta(EMPLOYEE_DOCTYPE).get_field(child_fieldname)
		child_doctype = table_field.options if table_field else None
		if not child_doctype:
			_write_sheet_rows(sheet, [["说明"], [f"{table_label} 未找到对应子表定义。"]])
			continue

		child_fields = _get_child_export_fields(child_doctype)
		headers = ["工号", "员工姓名"] + [field.label for field in child_fields]
		rows = [headers]
		if employee_names:
			for row in frappe.get_all(
				child_doctype,
				filters={
					"parenttype": EMPLOYEE_DOCTYPE,
					"parentfield": child_fieldname,
					"parent": ["in", employee_names],
				},
				fields=["parent"] + [field.fieldname for field in child_fields],
				order_by="parent asc, idx asc",
			):
				rows.append(
					[employee_code.get(row.parent, row.parent), employee_label.get(row.parent)]
					+ [row.get(field.fieldname) for field in child_fields]
				)
		_write_sheet_rows(sheet, rows)

	output = BytesIO()
	save_workbook_with_logo_watermark(workbook, output)
	return output.getvalue()


def _employee_export_records_cache_key(current_filters=None):
	filters = _parse_json(current_filters, {}) or {}
	company = filters.get("company") or frappe.defaults.get_user_default("Company") or ""
	scope = f"{frappe.session.user}|{company}"
	return f"employee_roster_export_records:{hashlib.sha256(scope.encode()).hexdigest()[:20]}"


def log_employee_export_record(filename, selected_fields, selected_tables, export_scope, current_filters):
	cache_key = _employee_export_records_cache_key(current_filters)
	records = frappe.cache().get_value(cache_key) or []
	records.insert(
		0,
		{
			"filename": filename,
			"fields": selected_fields,
			"tables": selected_tables,
			"export_scope": export_scope,
			"current_filters": current_filters,
			"created_at": frappe.utils.now_datetime().strftime("%Y-%m-%d %H:%M:%S"),
			"user": frappe.session.user,
		},
	)
	frappe.cache().set_value(cache_key, records[:20], expires_in_sec=60 * 60 * 24 * 30)


@frappe.whitelist()
def get_employee_export_records(current_filters: str = "{}"):
	return frappe.cache().get_value(_employee_export_records_cache_key(current_filters)) or []


def _report_file_stem(report_name):
	report_name = re.sub(r'[\\/:*?"<>|]+', "-", report_name or "员工报表").strip()
	return (report_name or "员工报表")[:80]


def _get_default_export_fields(allowed_fields):
	return [
		fieldname
		for fieldname, field in allowed_fields.items()
		if field.get("required") or field.get("enabled")
	][:10]


def _normalise_report_definition(definition, allowed_fields):
	selected_fields = _parse_selected_fields(json.dumps(definition.get("fields") or []), allowed_fields)
	if not selected_fields:
		selected_fields = _get_default_export_fields(allowed_fields)
	selected_tables = _parse_selected_tables(json.dumps(definition.get("tables") or []))
	return {
		"id": definition.get("id") or definition.get("name"),
		"name": definition.get("name"),
		"report_name": definition.get("report_name"),
		"description": definition.get("description") or "",
		"group_name": definition.get("group_name") or "人事档案",
		"fields": selected_fields,
		"tables": selected_tables,
		"filters": definition.get("filters") or {},
		"is_standard": 1 if definition.get("id") and not definition.get("name") else 0,
	}


def _get_saved_employee_reports(allowed_fields):
	if not frappe.db.exists("DocType", HRMS_EMPLOYEE_REPORT_DOCTYPE):
		return []

	reports = []
	for row in frappe.get_all(
		HRMS_EMPLOYEE_REPORT_DOCTYPE,
		filters={"disabled": 0},
		fields=["name", "report_name", "description", "group_name", "selected_fields", "selected_tables"],
		order_by="modified desc",
	):
		reports.append(
			_normalise_report_definition(
				{
					"name": row.name,
					"id": row.name,
					"report_name": row.report_name,
					"description": row.description,
					"group_name": row.group_name,
					"fields": _parse_json(row.selected_fields, []),
					"tables": _parse_json(row.selected_tables, []),
				},
				allowed_fields,
			)
		)
	return reports


def _get_employee_report_definitions(allowed_fields):
	reports = [_normalise_report_definition(report, allowed_fields) for report in DEFAULT_EMPLOYEE_REPORTS]
	reports.extend(_get_saved_employee_reports(allowed_fields))
	return reports


@frappe.whitelist()
def get_employee_report_center():
	doc = _get_template_doc()
	template_fields = _get_employee_export_fields(doc)
	allowed_fields = _get_allowed_export_fields(template_fields)
	reports = _get_employee_report_definitions(allowed_fields)
	group_order = ["人事档案", "人事统计", "行政报表"]
	groups = []

	for group_name in group_order + sorted({report["group_name"] for report in reports} - set(group_order)):
		group_reports = [report for report in reports if report["group_name"] == group_name]
		if group_reports:
			groups.append({"name": group_name, "reports": group_reports})

	return {"groups": groups}


@frappe.whitelist()
def save_employee_roster_report(
	report_name: str,
	description: str = "",
	group_name: str = "人事档案",
	fields: str = "[]",
	tables: str = "[]",
):
	if not frappe.db.exists("DocType", HRMS_EMPLOYEE_REPORT_DOCTYPE):
		frappe.throw(_("请先迁移数据库后再保存人事报表"))

	report_name = (report_name or "").strip()
	if not report_name:
		frappe.throw(_("请填写报表名称"))

	doc = _get_template_doc()
	template_fields = _get_employee_export_fields(doc)
	allowed_fields = _get_allowed_export_fields(template_fields)
	selected_fields = _parse_selected_fields(fields, allowed_fields)
	selected_tables = _parse_selected_tables(tables)
	if not selected_fields:
		frappe.throw(_("请至少选择一个字段"))

	report = frappe.get_doc(
		{
			"doctype": HRMS_EMPLOYEE_REPORT_DOCTYPE,
			"report_name": report_name,
			"description": description or "",
			"group_name": group_name or "人事档案",
			"selected_fields": json.dumps(selected_fields, ensure_ascii=False),
			"selected_tables": json.dumps(selected_tables, ensure_ascii=False),
		}
	)
	report.insert(ignore_permissions=True)
	frappe.db.commit()
	return {"name": report.name, "report_name": report.report_name}


@frappe.whitelist()
def download_employee_report(report_id: str):
	from frappe.desk.utils import provide_binary_file

	doc = _get_template_doc()
	template_fields = _get_employee_export_fields(doc)
	allowed_fields = _get_allowed_export_fields(template_fields)
	reports = {report["id"]: report for report in _get_employee_report_definitions(allowed_fields)}
	if report_id not in reports:
		frappe.throw(_("报表不存在或已被禁用"))

	report = reports[report_id]
	if not report["fields"]:
		frappe.throw(_("报表没有可导出的字段"))

	provide_binary_file(
		_report_file_stem(report["report_name"]),
		"xlsx",
		_make_employee_export_workbook(
			report["fields"],
			allowed_fields,
			report["tables"],
			report.get("filters") or {},
		),
	)


@frappe.whitelist()
# def download_employee_roster_export(fields: str
def download_employee_roster_export(
	fields: str = "[]",
	tables: str = "[]",
	export_scope: str = "all",
	current_filters: str = "{}",
):
	from frappe.desk.utils import provide_binary_file

	doc = _get_template_doc()
	template_fields = _get_employee_export_fields(doc)
	allowed_fields = _get_allowed_export_fields(template_fields)
	selected_fields = _parse_selected_fields(fields, allowed_fields)
	selected_tables = _parse_selected_tables(tables)
	if not selected_fields:
		frappe.throw(_("请至少选择一个导出字段"))

	export_scope = "current_filters" if export_scope == "current_filters" else "all"
	active_filters = _build_employee_roster_filters(current_filters)
	filters = active_filters if export_scope == "current_filters" else {
		"company": active_filters["company"]
	} if active_filters.get("company") else {}
	log_employee_export_record(EMPLOYEE_EXPORT_FILENAME, selected_fields, selected_tables, export_scope, filters)
	provide_binary_file(
		EMPLOYEE_EXPORT_FILENAME.removesuffix(".xlsx"),
		"xlsx",
		_make_employee_export_workbook(selected_fields, allowed_fields, selected_tables, filters),
	)


def _sync_employee_field_required_property(row):
	required = 1 if row.get("required") else 0
	custom_field_name = f"{EMPLOYEE_DOCTYPE}-{row.fieldname}"
	if row.source == "自定义" and frappe.db.exists("Custom Field", custom_field_name):
		frappe.db.set_value("Custom Field", custom_field_name, "reqd", required, update_modified=False)
		return

	if row.source == "系统" and frappe.get_meta(EMPLOYEE_DOCTYPE).get_field(row.fieldname):
		make_property_setter(
			EMPLOYEE_DOCTYPE,
			row.fieldname,
			"reqd",
			required,
			"Check",
			validate_fields_for_doctype=False,
		)


def _sync_employee_field_label_property(row):
	field_label = (row.field_label or "").strip()
	if not field_label:
		return

	custom_field_name = f"{EMPLOYEE_DOCTYPE}-{row.fieldname}"
	if row.source == "自定义" and frappe.db.exists("Custom Field", custom_field_name):
		frappe.db.set_value("Custom Field", custom_field_name, "label", field_label, update_modified=False)
		return

	if row.source == "系统" and frappe.get_meta(EMPLOYEE_DOCTYPE).get_field(row.fieldname):
		make_property_setter(
			EMPLOYEE_DOCTYPE,
			row.fieldname,
			"label",
			field_label,
			"Data",
			validate_fields_for_doctype=False,
		)


def _validate_reserved_employee_business_mapping(fieldname, field_label=None, aliases=None):
	"""Prevent a business-facing label from being mapped to two employee fields."""
	values = []
	if field_label is not None:
		values.append(field_label)
	if aliases is not None:
		values.extend(str(aliases or "").splitlines())

	for value in values:
		key = str(value or "").strip()
		if not key:
			continue
		owner = RESERVED_EMPLOYEE_BUSINESS_FIELD_KEYS.get(key)
		if owner and owner != fieldname:
			frappe.throw(
				_("业务名称“{0}”已唯一映射到字段 {1}，不能同时用于字段 {2}。").format(
					key, owner, fieldname
				)
			)


@frappe.whitelist()
def save_employee_field_template(items: str):
	_require_hr_settings_manager()
	items = _parse_json(items, [])
	doc = _get_template_doc()
	rows_by_fieldname = {row.fieldname: row for row in doc.template_items}
	supports_required = _template_item_supports_field("required")

	for item in items:
		fieldname = item.get("fieldname")
		if fieldname not in rows_by_fieldname:
			frappe.throw(_("字段不存在: {0}").format(fieldname))
		# Internal Frappe document naming is deliberately not configurable from
		# the HR field center. Ignore stale browser payloads that still contain it.
		if _is_employee_internal_field(fieldname):
			continue

		row = rows_by_fieldname[fieldname]
		if item.get("category"):
			_validate_category(item["category"])
			row.category = item["category"]

		if "field_label" in item:
			field_label = (item.get("field_label") or "").strip()
			if not field_label:
				frappe.throw(_("字段名称不能为空"))
			_validate_reserved_employee_business_mapping(fieldname, field_label=field_label)
			row.field_label = field_label
			_sync_employee_field_label_property(row)

		if "description" in item:
			row.description = item.get("description")
		if "enabled" in item:
			row.enabled = 1 if item.get("enabled") else 0
		if "required" in item and supports_required:
			row.required = 1 if item.get("required") else 0
			_sync_employee_field_required_property(row)
		if "search_enabled" in item:
			row.search_enabled = 1 if item.get("search_enabled") else 0
		for flag in ("import_enabled", "export_enabled", "form_visible", "detail_visible", "roster_visible"):
			if flag in item and _template_item_supports_field(flag):
				row.set(flag, 1 if item.get(flag) else 0)
		for config_fieldname in ("aliases", "detail_block", "record_type"):
			if config_fieldname in item and _template_item_supports_field(config_fieldname):
				if config_fieldname == "aliases":
					_validate_reserved_employee_business_mapping(fieldname, aliases=item.get(config_fieldname))
				row.set(config_fieldname, item.get(config_fieldname))
		if "detail_block_order" in item and _template_item_supports_field("detail_block_order"):
			row.detail_block_order = frappe.utils.cint(item.get("detail_block_order"))

	doc.save(ignore_permissions=True)
	_apply_employee_internal_field_policy(doc)
	frappe.clear_cache(doctype=EMPLOYEE_DOCTYPE)
	return get_employee_field_template()


@frappe.whitelist()
def create_employee_custom_field(
	category: str,
	field_label: str,
	fieldtype: str,
	description: str | None = None,
	options: str | None = None,
	required: bool | int | str = False,
	search_enabled: bool | int | str = False,
):
	_require_hr_settings_manager()
	_validate_category(category)
	fieldtype = _normalise_fieldtype(fieldtype)

	field_label = (field_label or "").strip()
	if not field_label or len(field_label) > 30:
		frappe.throw(_("字段名称不能为空且不能超过 30 个字符"))
	_validate_reserved_employee_business_mapping("", field_label=field_label)

	options = (options or "").strip()
	if fieldtype == "Select" and not options:
		frappe.throw(_("自定义选项字段必须填写选项"))

	doc = _get_template_doc()
	fieldname = _make_custom_fieldname(field_label)

	if _template_item_exists(doc, fieldname):
		frappe.throw(_("员工属性字段已存在: {0}").format(field_label))

	custom_field_name = f"{EMPLOYEE_DOCTYPE}-{fieldname}"
	if not frappe.db.exists("Custom Field", custom_field_name):
		custom_field = {
			"fieldname": fieldname,
			"label": field_label,
			"fieldtype": fieldtype,
			"insert_after": CATEGORY_INSERT_AFTER.get(category, "date_of_joining"),
			"description": description,
			"reqd": 1 if required else 0,
		}
		if fieldtype == "Select":
			custom_field["options"] = options
		create_custom_field(EMPLOYEE_DOCTYPE, custom_field)

	row_values = {
		"category": category,
		"field_label": field_label,
		"fieldname": fieldname,
		"fieldtype": fieldtype,
		"description": description,
		"source": "自定义",
		"enabled": 1,
		"search_enabled": 1 if search_enabled else 0,
		"options": options,
		"insert_after": CATEGORY_INSERT_AFTER.get(category, "date_of_joining"),
	}
	if _template_item_supports_field("required"):
		row_values["required"] = 1 if required else 0
	doc.append("template_items", row_values)
	doc.save(ignore_permissions=True)
	frappe.clear_cache(doctype=EMPLOYEE_DOCTYPE)
	return get_employee_field_template()


@frappe.whitelist()
def set_employee_template_field_enabled(fieldname: str, enabled: int | str):
	_require_hr_settings_manager()
	doc = _get_template_doc()
	if _is_employee_internal_field(fieldname):
		_apply_employee_internal_field_policy(doc)
		return get_employee_field_template()
	for row in doc.template_items:
		if row.fieldname == fieldname:
			row.enabled = 1 if frappe.utils.cint(enabled) else 0
			doc.save(ignore_permissions=True)
			return get_employee_field_template()

	frappe.throw(_("字段不存在: {0}").format(fieldname))


@frappe.whitelist()
def download_employee_import_template():
	from frappe.desk.utils import provide_binary_file

	provide_binary_file(
		EMPLOYEE_IMPORT_TEMPLATE_FILENAME.removesuffix(".xlsx"),
		"xlsx",
		build_employee_import_template(),
	)
