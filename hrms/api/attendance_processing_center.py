"""Persistent orchestration for manually uploaded attendance sources.

This module intentionally does not change ``attendance_import.py``.  It owns a
small, auditable workflow backed by existing import batches plus one unified
processing-record DocType.  Attendance-only monthly support files are import-only:
they are validated on upload but never enter the attendance exception/review
workflow or receive a second manual transformation.
"""

from __future__ import annotations

import hashlib
import json
import re
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import frappe
from frappe import _
from frappe.utils import cint, flt, now_datetime

from hrms.api.attendance_processors.apple_tree import AppleTreeRules, is_auto_excluded_apple_tree_row, preflight_apple_tree_rows, process_apple_tree_rows
from hrms.api.attendance_processors.attendance_draft import (
	dingtalk_daily_header_location,
	exception_lines_from_attendance_details,
	find_dingtalk_daily_sheet,
	precheck_attendance_draft_structure,
	process_attendance_draft_rows,
	rows_from_dingtalk_daily_sheet,
)
from hrms.api.attendance_processors.missed_punch import MissedPunchRules, precheck_missed_punch_structure, process_missed_punch_rows


IMPORT_BATCH_DOCTYPE = "HRMS Attendance Import Batch"
PROCESSING_RECORD_DOCTYPE = "HRMS Attendance Processing Record"
DEPARTMENT_MAPPING_DOCTYPE = "HRMS Attendance Department Mapping"
SOURCE_TYPES = ("attendance_draft", "apple_tree", "missing_card")
# These three sources are independent monthly facts used to produce the locked
# attendance final.  In particular, housing allowance is no longer a payroll
# monthly-variable upload: it is imported, checked and frozen with attendance.
MONTHLY_SUPPORT_SOURCE_TYPES = ("housing_allowance", "full_attendance", "special_hours")
# Finals locked before housing allowance returned to attendance did not include
# that source.  Keep them readable for audit and payroll, without weakening the
# gate for every newly generated final.
LEGACY_MONTHLY_SUPPORT_SOURCE_TYPES = ("full_attendance", "special_hours")
SOURCE_LABELS = {
	"attendance_draft": "考勤初稿",
	"apple_tree": "苹果树",
	"missing_card": "忘打卡",
	"housing_allowance": "住房补贴",
	"full_attendance": "全勤奖",
	"special_hours": "特殊工时",
}
MONTHLY_SUPPORT_SOURCE_CONFIG = {
	"housing_allowance": {
		"label": "住房补贴",
		"required_headers": ("工号", "姓名", "住房补贴"),
		"amount_headers": ("住房补贴",),
		"description": "一次性导入当月住房补贴明细；仅校验工号、姓名和住房补贴金额。",
		"mode": "monthly_amount",
		"value_field": "housing_allowance",
		"value_header": "住房补贴",
	},
	"full_attendance": {
		"label": "全勤奖",
		"required_headers": ("工号", "姓名", "全勤奖"),
		"amount_headers": ("全勤奖",),
		"description": "一次性导入当月全勤奖明细；仅校验工号、姓名和全勤奖金额。",
		"mode": "monthly_amount",
		"value_field": "full_attendance_award",
		"value_header": "全勤奖",
	},
	"special_hours": {
		"label": "特殊工时",
		"required_headers": ("工号", "姓名"),
		"description": "一次性导入当月特殊工时登记表；仅校验员工、日期与工时。",
		"mode": "special_hours_grid",
		"value_field": "special_hours",
	},
}

_DINGTALK_DEPARTMENT_IDENTIFIER_RE = re.compile(r"\s*[-－—–]\s*\d+\s*$")


def _display_department(value: Any) -> str:
	"""Hide DingTalk's non-business trailing department identifier in outputs."""
	return _DINGTALK_DEPARTMENT_IDENTIFIER_RE.sub("", str(value or "").strip()).strip()


_DAILY_ATTENDANCE_DATE_RE = re.compile(r"(?:(\d{4})|(\d{2}))[-/](\d{1,2})[-/](\d{1,2})")


def _daily_attendance_date(value: Any) -> str:
	"""Normalize DingTalk's ``26-06-01 星期一`` date value for filtering."""
	match = _DAILY_ATTENDANCE_DATE_RE.search(str(value or ""))
	if not match:
		return ""
	year = match.group(1) or f"20{match.group(2)}"
	try:
		return date(int(year), int(match.group(3)), int(match.group(4))).isoformat()
	except ValueError:
		return ""


# These identifiers deliberately remain stable in the database and in the
# processors.  They are implementation details though, not wording that an
# attendance administrator should have to interpret on the review screen.
PROCESSING_FIELD_LABELS = {
	"employee_code": "工号",
	"employee_name": "姓名",
	"department": "部门",
	"created_at": "创建时间",
	"punch_time": "补卡时间",
	"punch_type": "补卡类型",
	"reason": "补卡理由",
	"approval_result": "审批结果",
	"approval_status": "审批状态",
	"included": "是否计入",
	"red_apples": "红苹果",
	"amount": "红苹果金额",
	"housing_allowance": "住房补贴",
	"full_attendance_award": "全勤奖",
	"special_hours": "特殊工时（小时）",
	"special_hours_days": "特殊工时明细",
	"day": "日期",
	"hours": "工时",
	"standard_hours": "标准工时",
	"actual_attendance_hours": "实际出勤工时",
	"workday_overtime_hours": "工作日加班工时",
	"restday_overtime_hours": "休息日加班工时",
	"holiday_overtime_hours": "节假日加班工时",
	"large_night_shifts": "大夜班次数",
	"small_night_shifts": "小夜班次数",
	"personal_leave_hours": "事假工时",
	"sick_leave_hours": "病假工时",
	"annual_leave_hours": "特休工时",
	"work_injury_hours": "工伤工时",
	"rest_arrangement_hours": "排休工时",
	"absence_hours": "旷工工时",
	"absence_marker_count": "旷工标记次数",
	"clock_in_missing_count": "上班漏打卡次数",
	"clock_out_missing_count": "下班漏打卡次数",
	"late_count": "迟到次数",
	"early_count": "早退次数",
	"exception_events": "异常事件明细",
	"data_quality_events": "数据质量说明",
	"attendance_details": "涉及日期及打卡详情",
	"attendance_date": "考勤日期",
	"shift": "班次",
	"clock_in": "上班打卡",
	"clock_out": "下班打卡",
	"clock_in_missing": "上班缺卡次数",
	"clock_out_missing": "下班缺卡次数",
	"source_row_count": "来源明细行数",
	"eligible_for_downstream": "计入下游",
	"include_in_downstream": "计入下游",
	"工号": "工号",
	"姓名": "姓名",
	"部门": "部门",
	"苹果类型": "苹果类型",
	"有效苹果数": "有效苹果数",
}

# The completed attendance-draft dataset mirrors the human-readable monthly
# attendance summary: one employee per row and one attendance metric per
# column.  It intentionally excludes special hours, allowance and award data,
# because those belong to the later confirmed sources rather than this draft.
ATTENDANCE_DRAFT_RESULT_COLUMNS = (
	("department", "部门"),
	("employee_name", "姓名"),
	("employee_code", "工号"),
	("standard_hours", "标准工时（小时）"),
	("actual_attendance_hours", "实际出勤（小时）"),
	("workday_overtime_hours", "工作日加班（小时）"),
	("restday_overtime_hours", "休息日加班（小时）"),
	("holiday_overtime_hours", "节假日加班（小时）"),
	("large_night_shifts", "大夜班"),
	("small_night_shifts", "小夜班"),
	("personal_leave_hours", "事假（小时）"),
	("sick_leave_hours", "病假（小时）"),
	("annual_leave_hours", "特休（小时）"),
	("work_injury_hours", "工伤（小时）"),
	("rest_arrangement_hours", "排休（小时）"),
	("absence_hours", "旷工（小时）"),
	("clock_in_missing_count", "上班漏打卡次数"),
	("clock_out_missing_count", "下班漏打卡次数"),
)

# These are the editable source columns from DingTalk's daily statistics.  A
# correction changes one original daily row and then re-aggregates that
# employee's own monthly result; it never writes back to the uploaded workbook.
ATTENDANCE_DAILY_EDIT_FIELDS = (
	("班次", "班次", ("班次",)),
	("上班时间", "上班打卡", ("上班时间", "上班打卡", "上班打卡时间")),
	("下班时间", "下班打卡", ("下班时间", "下班打卡", "下班打卡时间")),
	("上班缺卡", "上班缺卡", ("上班缺卡", "上班未打卡次数")),
	("下班缺卡", "下班缺卡", ("下班缺卡", "下班未打卡次数")),
	("迟到次数", "迟到次数", ("迟到次数",)),
	("早退次数", "早退次数", ("早退次数",)),
	("旷工", "旷工标记", ("旷工", "旷工_2")),
	("旷工(小时)", "旷工工时", ("旷工(小时)", "请假/旷工(小时)")),
	("标准工时", "标准工时", ("标准工时", "标准工时（小时）")),
	("实际出勤（小时）", "实际出勤（小时）", ("实际出勤（小时）", "实际出勤")),
	("工作日加班（小时）", "工作日加班（小时）", ("工作日加班（小时）",)),
	("休息日加班（小时）", "休息日加班（小时）", ("休息日加班（小时）",)),
	("节假日加班（小时）", "节假日加班（小时）", ("节假日加班（小时）",)),
	("大夜班", "大夜班", ("大夜班",)),
	("小夜班", "小夜班", ("小夜班",)),
)

# The review page remains a complete, columnar view of one approval per row.
# The separate printable download contract is defined just below.
MISSED_PUNCH_RESULT_COLUMNS = (
	("employee_code", "工号"),
	("employee_name", "姓名"),
	("department", "部门"),
	("created_at", "创建时间"),
	("punch_time", "补卡时间"),
	("punch_type", "补卡类型"),
	("reason", "补卡理由"),
	("approval_result", "审批结果"),
	("approval_status", "审批状态"),
	("included", "是否计入"),
	("red_apples", "红苹果"),
	("amount", "红苹果金额"),
)

# The downloaded missed-punch workbook is a paper sign-off list, not an audit
# dump.  Approval and source identifiers remain persisted and available in the
# review UI, while the workbook contains only the fields HR needs to print.
MISSED_PUNCH_SIGNOFF_COLUMNS = (
	"序号",
	"部门",
	"创建时间",
	"补卡时间",
	"补卡类型",
	"补卡理由",
	"创建人",
	"签名",
	"备注",
)

# The first fields intentionally mirror the established Apple-tree sign-off
# table.  System-only processing columns follow them, so the page and the
# downloadable result use one familiar, usable header instead of two layouts.
APPLE_TREE_RESULT_COLUMNS = (
	("创建时间", "创建时间"),
	("奖惩日期", "奖/惩日期"),
	("部门", "受奖/惩人部门"),
	("姓名", "受奖/惩人"),
	("绿苹果", "绿苹果"),
	("红苹果", "红苹果"),
	("项目", "奖/惩项目"),
	("备注", "备注"),
	("创建人", "创建人"),
	("工号", "工号"),
	("苹果类型", "苹果类型"),
	("有效苹果数", "有效苹果数"),
	("审批编号", "审批编号"),
	("审批结果", "审批结果"),
	("审批状态", "审批状态"),
)

APPLE_TREE_RESULT_TRAIL_HEADERS = ("异常说明", "处理状态", "是否计入下游", "来源追溯")

# Apple-tree downloads are paper sign-off lists.  The complete approval,
# exception and provenance fields remain available in the review UI and audit
# records, but do not belong on this printable form.
APPLE_TREE_SIGNOFF_COLUMNS = (
	"序号",
	"创建时间",
	"奖/惩日期",
	"受奖/惩人部门",
	"受奖/惩人",
	"绿苹果",
	"红苹果",
	"奖/惩项目",
	"备注",
	"创建人",
	"签名",
	"备注",
)

EXCEPTION_LABELS = {
	"CLOCK_IN_MISSING": "上班漏打卡",
	"CLOCK_OUT_MISSING": "下班漏打卡",
	"SHIFT_MISSING": "班次缺失",
	"EMPLOYEE_DEPARTMENT_MISMATCH": "部门信息不一致",
	"EMPLOYEE_DEPARTMENT_CONFLICT": "来源部门信息冲突",
	"DEPARTMENT_CONFLICT": "部门信息不一致",
	"EMPLOYEE_NOT_FOUND": "花名册未找到员工",
	"EMPLOYEE_MATCH_PENDING": "员工匹配待确认",
	"EMPLOYEE_CODE_MISSING": "工号缺失",
	"EMPLOYEE_CODE_NAME_CONFLICT": "工号与姓名冲突",
	"EMPLOYEE_NAME_MISMATCH": "姓名信息不一致",
	"EMPLOYEE_NAME_CONFLICT": "姓名信息不一致",
	"EMPLOYEE_NAME_AMBIGUOUS": "姓名对应多名员工",
	"EMPLOYEE_AMBIGUOUS": "姓名对应多名员工",
	"ATTENDANCE_DATE_MISSING": "考勤日期缺失",
	"ATTENDANCE_DATE_INVALID": "考勤日期无效",
	"ATTENDANCE_DATE_DUPLICATE": "考勤日期重复",
	"ATTENDANCE_MONTH_MISMATCH": "考勤月份不一致",
	"SHIFT_MISSING": "班次缺失（数据质量）",
	"CLOCK_IN_MISSING": "上班缺卡",
	"CLOCK_OUT_MISSING": "下班缺卡",
	"LATE_MARKED": "迟到（钉钉标记）",
	"EARLY_MARKED": "早退（钉钉标记）",
	"ABSENCE_MARKED": "旷工标记待核验",
	"INVALID_NUMERIC_VALUE": "工时或次数格式无效",
	"SOURCE_FILE_MISSING": "来源文件定位缺失",
	"SOURCE_SHEET_MISSING": "来源工作表定位缺失",
	"SOURCE_ROW_MISSING": "来源行号缺失",
	"STRUCTURE_MISSING_REQUIRED_FIELD": "源表缺少必要字段",
	"INVALID_PUNCH_TIME": "补卡时间无效",
	"OUTSIDE_ATTENDANCE_MONTH": "补卡不属于本月",
	"APPROVAL_NO_MISSING": "补卡审批编号缺失",
	"APPROVAL_NOT_APPROVED": "补卡审批未通过",
	"APPROVAL_NOT_ENDED": "补卡审批未结束",
	"OFFLINE_ENTRY_REQUIRES_CONFIRMATION": "线下补录待确认",
	"MONTHLY_AMOUNT_MISSING": "月度金额缺失",
	"MONTHLY_AMOUNT_INVALID": "月度金额格式无效",
	"SPECIAL_HOURS_INVALID": "特殊工时格式无效",
	"DUPLICATE_EMPLOYEE_RECORD": "员工在来源表中重复",
	"OFFLINE_REASON_REQUIRED": "线下补录缺少原因",
	"OFFLINE_CONFIRMER_REQUIRED": "线下补录缺少确认人",
	"DUPLICATE_APPROVAL_NO": "审批编号重复",
	"MULTIPLE_APPROVALS_SAME_PUNCH_TIME": "同一补卡时间有多笔审批",
	"FORMER_EMPLOYEE_REQUIRES_CONFIRMATION": "离职人员记录待确认",
	"AMOUNT_MISSING": "苹果数量缺失",
	"AMOUNT_INVALID": "苹果数量无效",
	"AMOUNT_TEXT_CONFLICT": "苹果数量与项目说明不一致",
	"APPLE_TYPE_UNRECOGNIZED": "无法识别红苹果或绿苹果",
	"APPROVAL_NOT_FINISHED": "审批未结束",
	"APPROVAL_NOT_PASSED": "审批未通过",
	"MISSING_APPROVAL_NO": "审批编号缺失",
	"MISSING_APPROVAL_RESULT": "审批结果缺失",
	"MISSING_APPROVAL_STATUS": "审批状态缺失",
	"MONTH_MISMATCH": "记录不属于本月",
	"OFFLINE_ENTRY_REQUIRES_CONFIRMATION": "线下补录待确认",
}


def _review_guidance(exception_codes: list[str], source_type: str) -> list[str]:
	"""Return concrete, policy-safe choices for the shared human review queue."""
	codes = set(exception_codes or [])
	guidance = []
	if {"CLOCK_IN_MISSING", "CLOCK_OUT_MISSING"} & codes:
		guidance.append("先核对钉钉补卡审批；补卡审批已通过时，在“忘打卡”来源上传对应审批后再确认。")
		guidance.append("确认确实未打卡且没有有效补卡时，可选择“确认未打卡（不计入下游）”。")
	if "SHIFT_MISSING" in codes:
		guidance.append("核对是否为入职前、离职后或无排班日；确认无排班可通过，需补排班则保留待审核。")
	if {"LATE_MARKED", "EARLY_MARKED"} & codes:
		guidance.append("迟到、早退来自钉钉明确标记；先核对请假、主管说明或补卡，不得仅按次数自动扣款。")
	if "ABSENCE_MARKED" in codes:
		guidance.append("旷工字段在该来源中没有小时单位；先核对排班、有效请假及主管确认，再决定是否形成薪资缺勤工时。")
	if {"EMPLOYEE_DEPARTMENT_MISMATCH", "EMPLOYEE_DEPARTMENT_CONFLICT", "DEPARTMENT_CONFLICT"} & codes:
		guidance.append("以花名册当前部门为准；花名册无误可通过，需变更部门时先更新花名册或部门映射。")
	if {"EMPLOYEE_NOT_FOUND", "EMPLOYEE_MATCH_PENDING", "EMPLOYEE_CODE_MISSING", "EMPLOYEE_CODE_NAME_CONFLICT", "EMPLOYEE_NAME_MISMATCH", "EMPLOYEE_NAME_CONFLICT", "EMPLOYEE_NAME_AMBIGUOUS", "EMPLOYEE_AMBIGUOUS"} & codes:
		guidance.append("核对工号、姓名和花名册；无法唯一匹配前不要通过，也不要删除原始记录。")
	if {"OFFLINE_ENTRY_REQUIRES_CONFIRMATION", "OFFLINE_REASON_REQUIRED", "OFFLINE_CONFIRMER_REQUIRED"} & codes:
		guidance.append("线下补录须补齐原因和确认人；不能作为钉钉自动记录直接计入。")
	if {"APPROVAL_NOT_APPROVED", "APPROVAL_NOT_ENDED", "APPROVAL_NOT_FINISHED", "APPROVAL_NOT_PASSED"} & codes:
		guidance.append("等待审批“已通过且已结束”后重新上传该来源；当前记录不能自动计入。")
	if not guidance:
		guidance.append("核对来源追溯信息后，选择通过、驳回或保留待审核；原始导入行会继续保留。")
	return guidance


def _review_options(exception_codes: list[str], source_type: str) -> list[dict[str, str]]:
	"""Offer review decisions without changing a number merely to close a case."""
	codes = set(exception_codes or [])
	options = []
	if {"CLOCK_IN_MISSING", "CLOCK_OUT_MISSING"} & codes:
		options.extend([
			{"label": "忘记打卡，等待补卡审批", "review_status": "待审核", "reason": "等待员工补卡审批；当前不确认。"},
			{"label": "补卡审批已通过，确认当前考勤数据", "review_status": "已通过", "reason": "已核对补卡审批通过且已结束，确认当前考勤数据。"},
			{"label": "确认未打卡（不计入下游）", "review_status": "已驳回", "reason": "已核对无有效补卡审批，确认未打卡，本员工汇总不计入下游。"},
		])
	if "SHIFT_MISSING" in codes:
		options.extend([
			{"label": "确认无排班/入离职期间，保留当前数据", "review_status": "已通过", "reason": "已核对为无排班日或入离职期间，确认当前考勤数据。"},
			{"label": "等待补充或更正排班", "review_status": "待审核", "reason": "需补充或更正排班信息后再确认。"},
		])
	if {"LATE_MARKED", "EARLY_MARKED", "ABSENCE_MARKED"} & codes:
		options.extend([
			{"label": "已核对请假/主管说明，确认当前数据", "review_status": "已通过", "reason": "已核对请假或主管说明，确认当前考勤数据。"},
			{"label": "等待补充考勤证明", "review_status": "待审核", "reason": "需补充请假、补卡或主管说明后再确认。"},
			{"label": "确认异常不计入下游", "review_status": "已驳回", "reason": "已核对考勤异常，当前记录不计入下游。"},
		])
	if {"EMPLOYEE_DEPARTMENT_MISMATCH", "EMPLOYEE_DEPARTMENT_CONFLICT", "DEPARTMENT_CONFLICT"} & codes:
		options.extend([
			{"label": "花名册部门无误，按花名册确认", "review_status": "已通过", "reason": "已核对花名册，按花名册当前部门确认。"},
			{"label": "等待更新花名册或部门映射", "review_status": "待审核", "reason": "需先更新花名册或部门映射后再确认。"},
		])
	if not options:
		options.extend([
			{"label": "核对无误，确认当前数据", "review_status": "已通过", "reason": "已核对来源和相关资料，确认当前数据。"},
			{"label": "资料不足，保留待审核", "review_status": "待审核", "reason": "资料不足，保留待审核。"},
			{"label": "确认不应计入（不计入下游）", "review_status": "已驳回", "reason": "已核对该记录不应计入下游。"},
		])
	return options
def _require_processing_manager():
	"""Guard every whitelisted read/write path before any ignore_permissions call."""
	# ``frappe.has_role`` is only exposed on the browser-side Frappe object in
	# this deployment.  Server methods must resolve roles from the current
	# session explicitly.
	roles = set(frappe.get_roles(frappe.session.user))
	if not ({"System Manager", "HR Manager"} & roles):
		frappe.throw(_("只有 System Manager 或 HR Manager 可以访问考勤处理中心。"), frappe.PermissionError)


def _require_company(company: str) -> str:
	company = (company or "").strip()
	if not company:
		frappe.throw(_("请先选择公司。"))
	return company


def _require_month(attendance_month: str) -> str:
	attendance_month = (attendance_month or "").strip()
	if not attendance_month or len(attendance_month) != 7 or attendance_month[4:5] != "-":
		frappe.throw(_("考勤月份必须使用 YYYY-MM。"))
	return attendance_month


def _require_source_type(source_type: str) -> str:
	source_type = (source_type or "").strip()
	if source_type not in SOURCE_TYPES:
		frappe.throw(_("不支持的考勤来源类型。"))
	return source_type


def _require_monthly_support_source_type(source_type: str) -> str:
	source_type = (source_type or "").strip()
	if source_type not in MONTHLY_SUPPORT_SOURCE_TYPES:
		frappe.throw(_("不支持的月度补充来源类型。"))
	return source_type


def _require_processing_source_type(source_type: str) -> str:
	source_type = (source_type or "").strip()
	if source_type not in SOURCE_TYPES + MONTHLY_SUPPORT_SOURCE_TYPES:
		frappe.throw(_("不支持的考勤处理来源类型。"))
	return source_type


def _json(value: Any) -> str:
	return json.dumps(value, ensure_ascii=False, default=str, sort_keys=True)


def _loads(value: Any, default: Any):
	if value in (None, ""):
		return default
	try:
		return json.loads(value)
	except (TypeError, ValueError):
		return default


def _source_department_values(original_value: Any) -> list[str]:
	"""Read the unmodified department values from retained source rows."""
	if not isinstance(original_value, dict):
		return []
	# Attendance drafts retain a ``rows`` list; DingTalk forgot-punch records
	# retain one raw row directly, where the column is ``创建人部门``.
	rows = original_value.get("rows")
	if not isinstance(rows, list):
		rows = [original_value]
	values = []
	for row in rows:
		if not isinstance(row, dict):
			continue
		value = row.get("创建人部门") or row.get("实际部门") or row.get("部门") or row.get("department")
		value = str(value or "").strip()
		if value and value not in values:
			values.append(value)
	return values


def _department_comparison(record: dict[str, Any]) -> dict[str, str] | None:
	"""Expose both sides of an identity-department mismatch to the reviewer."""
	if "EMPLOYEE_DEPARTMENT_MISMATCH" not in (record.get("exception_codes") or []) and "DEPARTMENT_CONFLICT" not in (record.get("exception_codes") or []):
		return None
	source_departments = _source_department_values(record.get("original_value"))
	# The persisted proposed value is the roster department resolved at processing
	# time. Prefer it so reviewing old batches stays reproducible even after a
	# later roster update.
	roster_department = (record.get("proposed_value") or {}).get("department") or record.get("department") or ""
	if not source_departments and not roster_department:
		return None
	return {
		"source_department": "、".join(source_departments) or "（原表未提供）",
		"roster_department": str(roster_department).strip() or "（花名册未提供）",
	}


def _department_mapping_for(company: str, source_type: str = "missing_card") -> dict[str, str]:
	"""Return only reviewed, enabled aliases for the current company/source."""
	try:
		rows = frappe.get_all(
			DEPARTMENT_MAPPING_DOCTYPE,
			filters={"company": company, "source_type": source_type, "enabled": 1},
			fields=["source_department", "target_department"],
			limit_page_length=0,
		)
	except Exception:
		# A site that has not migrated the optional mapping DocType must still be
		# able to retain and review source rows without a hidden fallback mapping.
		return {}
	return {
		str(row.source_department or "").strip(): str(row.target_department or "").strip()
		for row in rows
		if str(row.source_department or "").strip() and str(row.target_department or "").strip()
	}


def _file_doc(file_url: str):
	name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not name:
		frappe.throw(_("找不到上传的来源文件。"))
	return frappe.get_doc("File", name)


def _file_checksum(file_url: str) -> str:
	return hashlib.sha256(_file_doc(file_url).get_content()).hexdigest()


def _load_workbook(file_url: str):
	from openpyxl import load_workbook

	return load_workbook(BytesIO(_file_doc(file_url).get_content()), data_only=True, read_only=True)


def _latest_batch(company: str, attendance_month: str, source_type: str):
	name = frappe.db.get_value(
		IMPORT_BATCH_DOCTYPE,
		{
			"company": company,
			"attendance_month": attendance_month,
			"source_type": source_type,
			"status": ["!=", "已撤销"],
		},
		"name",
		order_by="modified desc, creation desc",
	)
	return frappe.get_doc(IMPORT_BATCH_DOCTYPE, name) if name else None


def _batch_notes(batch) -> dict[str, Any]:
	notes = _loads(batch.notes, {})
	return notes if isinstance(notes, dict) else {}


def _save_batch_notes(batch, updates: dict[str, Any]):
	notes = _batch_notes(batch)
	notes.setdefault("attendance_processing_center", {}).update(updates)
	batch.notes = _json(notes)
	batch.save(ignore_permissions=True)


def _processing_meta(batch) -> dict[str, Any]:
	return _batch_notes(batch).get("attendance_processing_center", {})


def _normalized_header(value: Any) -> str:
	"""Compare wide Excel headers without being sensitive to line breaks/spaces."""
	return "".join(str(value or "").split())


def _month_marker(attendance_month: str) -> str:
	return f"{int(attendance_month[5:7])}月"


def _sheet_matches_month(sheet, attendance_month: str) -> bool:
	"""Ignore comparison/last-month tabs when a workbook contains several tabs."""
	marker = _month_marker(attendance_month)
	values = [sheet.title]
	for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
		values.extend(str(value or "") for value in row)
	return marker in "".join(values)


def _monthly_support_sheets(workbook, attendance_month: str):
	return [sheet for sheet in workbook.worksheets if _sheet_matches_month(sheet, attendance_month)]


def _support_header_matches(sheet, required_headers: tuple[str, ...]):
	required = tuple(_normalized_header(value) for value in required_headers)
	for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
		headers = [_normalized_header(value) for value in row]
		positions = {field: [index for index, value in enumerate(headers) if value == field] for field in required}
		if all(positions.values()):
			return row_number, positions
	return None, {}


def _monthly_amount_header_groups(sheet, config: dict[str, Any]):
	"""Find monthly amount tables without confusing notes with employee data.

	The supplied HR workbooks print the main list in two side-by-side tables and
	then add a separate "new hire / leaver" reference section.  The latter has a
	different column layout and contains employees already listed in the main
	table, so carrying the first header's column indexes to the bottom of the
	sheet reads names as employee codes and blank cells as amounts.
	"""
	amount_headers = {
		_normalized_header(value)
		for value in config.get("amount_headers") or (config["value_header"],)
	}
	matches = []
	for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
		headers = [_normalized_header(value) for value in row]
		code_indexes = [index for index, value in enumerate(headers) if value == "工号"]
		name_indexes = [index for index, value in enumerate(headers) if value == "姓名"]
		amount_indexes = [index for index, value in enumerate(headers) if value in amount_headers]
		if not code_indexes or not name_indexes or not amount_indexes:
			continue
		matches.append({
			"row_number": row_number,
			"code_indexes": code_indexes,
			"name_indexes": name_indexes,
			"amount_indexes": amount_indexes,
			"is_numbered_list": "序号" in headers,
		})
	# A numbered list is the main source of truth.  If a customer provides a
	# simple one-table template without serial numbers, preserve its first table
	# rather than rejecting a valid import.
	main_matches = [match for match in matches if match["is_numbered_list"]]
	return matches, (main_matches or matches[:1])


def _monthly_amount_rows(sheet, batch, config: dict[str, Any]):
	all_headers, main_headers = _monthly_amount_header_groups(sheet, config)
	if not main_headers:
		return []
	rows = []
	for header in main_headers:
		next_header_row = next(
			(item["row_number"] for item in all_headers if item["row_number"] > header["row_number"]),
			sheet.max_row + 1,
		)
		for source_row, values in enumerate(
			sheet.iter_rows(min_row=header["row_number"] + 1, max_row=next_header_row - 1, values_only=True),
			start=header["row_number"] + 1,
		):
			for table_index, code_index in enumerate(header["code_indexes"]):
				name_index = header["name_indexes"][min(table_index, len(header["name_indexes"]) - 1)]
				amount_index = header["amount_indexes"][min(table_index, len(header["amount_indexes"]) - 1)]
				department_index = name_index + 1 if name_index + 1 < len(values) else None
				code = values[code_index] if code_index < len(values) else None
				name = values[name_index] if name_index < len(values) else None
				amount = values[amount_index] if amount_index < len(values) else None
				department = values[department_index] if department_index is not None else None
				if not any(value not in (None, "") for value in (code, name, amount, department)):
					continue
				# Monthly sheets often place policy notes between the master list and
				# the appended reference tables.  Keep incomplete employee rows only
				# when they still contain a business value to review; otherwise notes,
				# totals and signature labels must not become fake employees.
				if code in (None, "") and name in (None, ""):
					continue
				if (
					(code in (None, "") or name in (None, ""))
					and amount in (None, "")
					and department in (None, "")
				):
					continue
				rows.append({
					"employee_code": str(code).strip() if code not in (None, "") else "",
					"employee_name": str(name).strip() if name not in (None, "") else "",
					"department": str(department).strip() if department not in (None, "") else "",
					config["value_field"]: amount,
					"source_file": batch.source_file,
					"source_sheet": sheet.title,
					"source_row": source_row,
					"source_id": f"{sheet.title}:{source_row}:{table_index + 1}",
				})
	return rows


def _special_hours_rows(sheet, batch):
	header_row, positions = _support_header_matches(sheet, ("工号", "姓名"))
	if not header_row:
		return []
	date_row = header_row + 1
	day_values = next(sheet.iter_rows(min_row=date_row, max_row=date_row, values_only=True), ())
	day_indexes = [(index, cint(value)) for index, value in enumerate(day_values) if 1 <= cint(value) <= 31]
	code_index, name_index = positions["工号"][0], positions["姓名"][0]
	department_index = name_index + 1
	rows = []
	for source_row, values in enumerate(sheet.iter_rows(min_row=date_row + 1, values_only=True), start=date_row + 1):
		code = values[code_index] if code_index < len(values) else None
		name = values[name_index] if name_index < len(values) else None
		department = values[department_index] if department_index < len(values) else None
		daily_values = [{"day": day, "hours": values[index] if index < len(values) else None} for index, day in day_indexes if index < len(values) and values[index] not in (None, "")]
		if not any(value not in (None, "") for value in (code, name, department)) and not daily_values:
			continue
		# The final "单日工时汇总" line contains daily totals but no employee.
		# It is a report footer, not a person with missing identifiers.
		if code in (None, "") and name in (None, ""):
			continue
		rows.append({
			"employee_code": str(code).strip() if code not in (None, "") else "",
			"employee_name": str(name).strip() if name not in (None, "") else "",
			"department": str(department).strip() if department not in (None, "") else "",
			"special_hours_days": daily_values,
			"source_file": batch.source_file,
			"source_sheet": sheet.title,
			"source_row": source_row,
			"source_id": f"{sheet.title}:{source_row}",
		})
	return rows


def _read_monthly_support_rows(batch):
	config = MONTHLY_SUPPORT_SOURCE_CONFIG[batch.source_type]
	workbook = _load_workbook(batch.source_file)
	sheets = _monthly_support_sheets(workbook, batch.attendance_month)
	rows = []
	for sheet in sheets:
		if config["mode"] == "monthly_amount":
			rows.extend(_monthly_amount_rows(sheet, batch, config))
		else:
			rows.extend(_special_hours_rows(sheet, batch))
	return rows, sheets


def _monthly_support_precheck(batch) -> dict[str, Any]:
	config = MONTHLY_SUPPORT_SOURCE_CONFIG[batch.source_type]
	workbook = _load_workbook(batch.source_file)
	matches = []
	monthly_sheets = _monthly_support_sheets(workbook, batch.attendance_month)
	for sheet in monthly_sheets:
		header_row, positions = _support_header_matches(sheet, config["required_headers"])
		if not header_row:
			continue
		if config["mode"] == "monthly_amount":
			record_count = len(_monthly_amount_rows(sheet, batch, config))
		else:
			record_count = len(_special_hours_rows(sheet, batch))
		matches.append({"sheet": sheet.title, "header_row": header_row, "record_count": record_count})
	if not matches:
		return {
			"is_valid": False,
			"status": "结构异常",
			"required_headers": list(config["required_headers"]),
			"message": _("未找到与处理月份匹配且同时包含 {0} 的工作表。").format("、".join(config["required_headers"])),
			"sheets": [sheet.title for sheet in monthly_sheets] or workbook.sheetnames,
		}
	record_count = sum(item["record_count"] for item in matches)
	if not record_count:
		return {
			"is_valid": False,
			"status": "结构异常",
			"required_headers": list(config["required_headers"]),
			"matching_sheets": matches,
			"record_count": 0,
			"message": _("已识别字段，但未找到任何工号记录。"),
		}
	return {
		"is_valid": True,
		"status": "结构通过",
		"required_headers": list(config["required_headers"]),
		"matching_sheets": matches,
		"record_count": record_count,
	}


def _employee_directory(company: str = ""):
	"""Return business employee codes, never Frappe's internal Employee name.

	The roster's visible 工号 is ``custom_employee_code`` in this deployment.
	Internal document names must not be used to match DingTalk rows.
	"""
	try:
		employees = frappe.get_all(
			"Employee",
			filters={"company": company} if company else None,
			fields=["custom_employee_code", "employee_name", "department", "status as employment_status", "date_of_joining", "relieving_date"],
			# A small default page would make most name+department matches look
			# missing even though those employees exist in the roster.
			limit_page_length=5000,
		)
		return [
			{
				"employee_code": (employee.custom_employee_code or "").strip(),
				"employee_name": employee.employee_name,
				"department": employee.department,
				"employment_status": employee.employment_status,
				"date_of_joining": employee.date_of_joining,
				"relieving_date": employee.relieving_date,
			}
			for employee in employees
			if (employee.custom_employee_code or "").strip()
		]
	except Exception:
		return []


def _as_nonnegative_number(value: Any):
	if isinstance(value, bool) or value in (None, ""):
		return None
	try:
		number = float(value)
	except (TypeError, ValueError):
		return None
	return number if number >= 0 else None


def _process_monthly_support_rows(batch):
	"""Validate and persist one-time monthly attendance-support datasets.

	Validation errors are deliberately retained with the import result so that
	HR can correct and re-upload the original workbook.  They are not attendance
	exceptions: no manual approval, rejection, or data adjustment is available
	for these sources.
	"""
	config = MONTHLY_SUPPORT_SOURCE_CONFIG[batch.source_type]
	raw_rows, sheets = _read_monthly_support_rows(batch)
	employee_rows = _employee_directory(batch.company)
	employees = {str(employee["employee_code"]).strip(): employee for employee in employee_rows}
	employees_by_name = defaultdict(list)
	for employee in employee_rows:
		name_key = re.sub(r"\s+", "", str(employee.get("employee_name") or ""))
		if name_key:
			employees_by_name[name_key].append(employee)
	attendance_batch = _latest_batch(batch.company, batch.attendance_month, "attendance_draft")
	attendance_people = frappe.get_all(
		PROCESSING_RECORD_DOCTYPE,
		filters={"import_batch": attendance_batch.name} if attendance_batch else {"name": "__none__"},
		fields=["employee_code", "employee_name"],
		limit_page_length=5000,
	)
	attendance_codes = {str(row.employee_code or "").strip() for row in attendance_people if str(row.employee_code or "").strip()}
	attendance_names = {re.sub(r"\s+", "", str(row.employee_name or "")) for row in attendance_people if str(row.employee_name or "").strip()}
	seen_codes = set()
	processed_rows = []
	last_day = monthrange(int(batch.attendance_month[:4]), int(batch.attendance_month[5:]))[1]
	for raw in raw_rows:
		source_code = str(raw.get("employee_code") or "").strip()
		code = source_code
		name = raw.get("employee_name") or ""
		department = raw.get("department") or ""
		exception_codes = []
		identity_resolution = ""
		exclusion_reason = ""
		coded_employee = employees.get(code)
		name_matches = employees_by_name.get(re.sub(r"\s+", "", str(name)), []) if name else []
		if len(name_matches) == 1 and (
			not coded_employee
			or (name and str(coded_employee.get("employee_name") or "").strip() != str(name).strip())
		):
			matched_by_name = name_matches[0]
			code = str(matched_by_name.get("employee_code") or "").strip()
			identity_resolution = _("已按唯一姓名 {0} 将来源工号 {1} 修正为花名册工号 {2}").format(name, source_code or "空", code)
		if not code:
			exception_codes.append("EMPLOYEE_CODE_MISSING")
		elif code in seen_codes:
			exception_codes.append("DUPLICATE_EMPLOYEE_RECORD")
		else:
			seen_codes.add(code)
		employee = employees.get(code)
		if employees and code and not employee:
			exception_codes.append("EMPLOYEE_NOT_FOUND")
		elif employee:
			if name and str(employee.get("employee_name") or "").strip() and name != str(employee["employee_name"]).strip():
				exception_codes.append("EMPLOYEE_CODE_NAME_CONFLICT")
			# Housing allowance, full-attendance and special-hours files are
			# verified by employee identity plus their respective amount/hours.
			# Source departments are helpful audit context, but historical labels
			# must not block a payroll input when the roster has since moved the
			# employee to another department.
			department = str(employee.get("department") or department).strip()
		in_attendance_population = bool(code and code in attendance_codes) or bool(name and re.sub(r"\s+", "", str(name)) in attendance_names)
		if attendance_people and not in_attendance_population:
			# A support-sheet template may retain departed or zero-value historical
			# rows. They remain traceable, but must neither block this month nor create
			# a payroll employee absent from the attendance population.
			exclusion_reason = _("不在本月考勤初稿人员范围，已保留来源但不进入月度终稿")
			exception_codes = [code for code in exception_codes if code not in {"EMPLOYEE_CODE_MISSING", "EMPLOYEE_NOT_FOUND", "EMPLOYEE_CODE_NAME_CONFLICT"}]

		if config["mode"] == "monthly_amount":
			amount = _as_nonnegative_number(raw.get(config["value_field"]))
			if raw.get(config["value_field"]) in (None, ""):
				exception_codes.append("MONTHLY_AMOUNT_MISSING")
			elif amount is None:
				exception_codes.append("MONTHLY_AMOUNT_INVALID")
			proposed = {"employee_code": code, "employee_name": name, "department": department, config["value_field"]: amount if amount is not None else raw.get(config["value_field"]), "eligible_for_downstream": not exception_codes and not exclusion_reason}
		else:
			daily_entries = []
			for entry in raw.get("special_hours_days") or []:
				hours = _as_nonnegative_number(entry.get("hours"))
				if hours is None or entry.get("day", 0) > last_day:
					exception_codes.append("SPECIAL_HOURS_INVALID")
					continue
				daily_entries.append({"day": entry["day"], "hours": hours})
			proposed = {"employee_code": code, "employee_name": name, "department": department, "special_hours": sum(entry["hours"] for entry in daily_entries), "special_hours_days": daily_entries, "eligible_for_downstream": not exception_codes and not exclusion_reason}

		exception_codes = list(dict.fromkeys(exception_codes))
		proposed["eligible_for_downstream"] = not exception_codes and not exclusion_reason
		if identity_resolution:
			proposed["identity_resolution"] = identity_resolution
		if exclusion_reason:
			proposed["exclusion_reason"] = exclusion_reason
		processed_rows.append({
			"employee_code": code,
			"employee_name": name,
			"department": department,
			"source_file": raw["source_file"],
			"source_sheet": raw["source_sheet"],
			"source_row": raw["source_row"],
			"source_id": raw["source_id"],
			"original_data": raw,
			"processed_value": proposed,
			"proposed_value": proposed,
			"exception_codes": exception_codes,
			"exception_message": "；".join(EXCEPTION_LABELS.get(code, code) for code in exception_codes),
			# Monthly support data is a one-time import.  A validation error blocks
			# this import version from the final, but must not create a task in the
			# shared attendance exception queue.
			"review_status": "无需审核",
			"eligible_for_downstream": not exception_codes and not exclusion_reason,
		})
	# This is an import-validation count, not a count of review tasks.
	exception_rows = sum(1 for row in processed_rows if row["exception_codes"])
	excluded_rows = sum(1 for row in processed_rows if (row.get("proposed_value") or {}).get("exclusion_reason"))
	return {
		"status": "导入异常" if exception_rows else "已确认",
		"processed_rows": processed_rows,
		"metrics": {"source_rows": len(raw_rows), "processed_rows": len(processed_rows), "exception_rows": exception_rows, "excluded_rows": excluded_rows},
		"source_sheets": [sheet.title for sheet in sheets],
	}


def _simple_sheet_rows(sheet, file_url: str, header_hints: tuple[str, ...] = ()) -> list[dict[str, Any]]:
	"""Read a flat form whose real header may follow one or more title rows."""
	values = list(sheet.iter_rows(values_only=True))
	if not values:
		return []
	header_index = 0
	if header_hints:
		header_index = max(
			range(min(12, len(values))),
			key=lambda index: sum(
				1 for hint in header_hints if hint in {str(value).strip() for value in values[index] if value not in (None, "")}
			),
		)
	headers = [str(value).strip() if value is not None else "" for value in values[header_index]]
	result = []
	for source_row, row_values in enumerate(values[header_index + 1 :], start=header_index + 2):
		if not any(value not in (None, "") for value in row_values):
			continue
		row = {headers[index]: row_values[index] if index < len(row_values) else None for index in range(len(headers)) if headers[index]}
		row.update({"source_file": file_url, "source_sheet": sheet.title, "source_row": source_row})
		result.append(row)
	return result


def _attendance_draft_sheet(workbook):
	return find_dingtalk_daily_sheet(workbook)


def _source_sheet(workbook, source_type: str):
	if source_type == "attendance_draft":
		sheet = _attendance_draft_sheet(workbook)
		if not sheet:
			frappe.throw(_("考勤初稿未找到可识别的钉钉每日数据表；请确认表内包含“姓名、工号、日期”等表头。"))
		return sheet
	if "钉钉导出数据" in workbook.sheetnames:
		return workbook["钉钉导出数据"]
	return workbook[workbook.sheetnames[0]]


def _read_source_rows(batch):
	workbook = _load_workbook(batch.source_file)
	sheet = _source_sheet(workbook, batch.source_type)
	if batch.source_type == "attendance_draft":
		rows = rows_from_dingtalk_daily_sheet(sheet, source_file=batch.source_file)
		return rows, sheet.title
	header_hints = (
		("创建时间", "奖/惩日期", "受奖/惩人", "绿苹果", "红苹果", "奖/惩项目")
		if batch.source_type == "apple_tree"
		else ("创建时间", "补卡时间", "补卡类型", "补卡理由", "创建人")
	)
	rows = _simple_sheet_rows(sheet, batch.source_file, header_hints)
	if batch.source_type == "apple_tree":
		preflight = preflight_apple_tree_rows(rows)
		if preflight.get("来源口径") == "人资月度汇总表":
			for row in rows:
				row["source_kind"] = "monthly_summary"
	elif batch.source_type == "missing_card":
		preflight = precheck_missed_punch_structure(list(rows[0]) if rows else [])
		if preflight.get("source_kind") == "monthly_summary":
			for row in rows:
				row["source_kind"] = "monthly_summary"
	return rows, sheet.title


def _precheck(batch):
	rows, sheet_name = _read_source_rows(batch)
	if batch.source_type == "attendance_draft":
		workbook = _load_workbook(batch.source_file)
		sheet = _source_sheet(workbook, batch.source_type)
		location = dingtalk_daily_header_location(sheet)
		headers = location["headers"] if location else []
		result = precheck_attendance_draft_structure(headers)
	elif batch.source_type == "apple_tree":
		result = preflight_apple_tree_rows(rows)
	else:
		headers = list(rows[0]) if rows else []
		result = precheck_missed_punch_structure(headers)
	return {"source_sheet": sheet_name, "row_count": len(rows), "result": result}


def _attendance_draft_exception_policy() -> dict[str, bool]:
	"""Map the visual rule-centre switches to reviewed import detectors only."""
	policy = {
		"missing_punch": True,
		"late": True,
		"early": True,
		"absence_marker": True,
	}
	rule_keys = {
		"ATT-DRAFT-MISSING-PUNCH": "missing_punch",
		"ATT-DRAFT-LATE": "late",
		"ATT-DRAFT-EARLY": "early",
		"ATT-DRAFT-ABSENCE-MARKER": "absence_marker",
	}
	try:
		rules = frappe.get_all(
			"HRMS Attendance Custom Rule",
			filters={"rule_code": ["in", list(rule_keys)]},
			fields=["rule_code", "enabled", "application_mode"],
			limit_page_length=len(rule_keys),
		)
	except Exception:
		return policy
	for rule in rules:
		key = rule_keys.get(rule.rule_code)
		if key:
			policy[key] = bool(rule.enabled and rule.application_mode == "异常提示")
	return policy


def _missed_punch_rules(company: str, attendance_month: str) -> MissedPunchRules:
	"""Use the rule centre for future missed-punch batches without rewriting history."""
	try:
		from hrms.api.payroll_input import get_attendance_processing_rule_settings

		settings = get_attendance_processing_rule_settings(company, attendance_month)
		red_apples = cint(settings.get("red_apples_per_record", 2))
		amount_per_apple = flt(settings.get("amount_per_apple", 5))
		return MissedPunchRules(
			red_apples_per_record=red_apples,
			amount_per_record=red_apples * amount_per_apple,
		)
	except Exception:
		# A phased rollout must not block an attendance upload on a legacy site.
		return MissedPunchRules()


def _process_batch(batch) -> dict[str, Any]:
	rows, sheet_name = _read_source_rows(batch)
	employees = _employee_directory(batch.company)
	if batch.source_type == "attendance_draft":
		exception_policy = _attendance_draft_exception_policy()
		return process_attendance_draft_rows(
			rows,
			attendance_month=batch.attendance_month,
			source_file=batch.source_file,
			source_sheet=sheet_name,
			employee_directory=employees or None,
			exception_policy=exception_policy,
		)
	if batch.source_type == "apple_tree":
		excluded_source_rows = sum(1 for row in rows if is_auto_excluded_apple_tree_row(row))
		processed_rows = process_apple_tree_rows(
			rows,
			rules=AppleTreeRules(target_month=batch.attendance_month),
			employees=employees or None,
			source_file=batch.source_file,
			source_sheet=sheet_name,
			start_row=2,
		)
		exception_rows = sum(1 for row in processed_rows if row.get("review_status") == "待审核")
		return {
			"status": "已确认" if rows and not processed_rows and excluded_source_rows == len(rows) else "待处理异常" if exception_rows else "待确认",
			"structure_precheck": preflight_apple_tree_rows(rows),
			"processed_rows": processed_rows,
			"metrics": {"source_rows": len(rows), "processed_rows": len(processed_rows), "excluded_source_rows": excluded_source_rows, "exception_rows": exception_rows},
		}
	return process_missed_punch_rows(
		rows,
		attendance_month=batch.attendance_month,
		source_file=batch.source_file,
		source_sheet=sheet_name,
		employee_directory=employees or None,
		department_mapping=_department_mapping_for(batch.company, "missing_card"),
		rules=_missed_punch_rules(batch.company, batch.attendance_month),
	)


def _row_value(row: dict[str, Any], *keys: str):
	for key in keys:
		if row.get(key) not in (None, ""):
			return row.get(key)
	return ""


def _confirmed_downstream_eligible(confirmed: Any) -> bool:
	"""Respect an approved reviewer's explicit inclusion decision.

	The three processors use slightly different field names while converging on
	the same review contract. A passed review whose confirmed value explicitly
	excludes a record must not enter final aggregation merely because it passed.
	"""
	if not isinstance(confirmed, dict):
		return True
	for fieldname in ("eligible_for_downstream", "include_in_downstream", "included"):
		if fieldname not in confirmed:
			continue
		value = confirmed[fieldname]
		if isinstance(value, bool):
			return value
		if isinstance(value, (int, float)):
			return value != 0
		return str(value or "").strip().lower() in {"1", "true", "yes", "y", "是", "计入"}
	return True


def _record_payload(batch, row: dict[str, Any]) -> dict[str, Any]:
	proposed = row.get("proposed_value") or {}
	processed = row.get("processed_value") or proposed
	# Retain the untouched DingTalk source row whenever the processor supplies
	# it.  The smaller original_value is an editable-field snapshot, not a
	# substitute for source-row traceability.
	original = row.get("original_data") or row.get("original_value") or {}
	# apple_tree uses include_in_downstream while the other processors use the
	# shared eligible_for_downstream field. Preserve either source contract.
	downstream_eligible = row.get("eligible_for_downstream")
	if downstream_eligible is None:
		downstream_eligible = row.get("include_in_downstream", False)
	# Normal records never enter a manual-review queue. A processor that omits
	# the display status must default to automatic inclusion; only a record with
	# a real exception remains pending for manual correction.
	review_status = row.get("review_status") or ("待审核" if row.get("exception_codes") else "无需审核")
	# An unresolved or rejected mismatch is visible and retained, but must never
	# contaminate the otherwise usable source result or stop other employees from
	# moving to the next calculation step.
	downstream_eligible = bool(downstream_eligible) and review_status in {"无需审核", "已通过"}
	return {
		"doctype": PROCESSING_RECORD_DOCTYPE,
		"import_batch": batch.name,
		"company": batch.company,
		"attendance_month": batch.attendance_month,
		"source_type": batch.source_type,
		"employee_code": _row_value(row, "employee_code", "工号"),
		"employee_name": _row_value(row, "employee_name", "姓名"),
		"department": _row_value(row, "department", "部门"),
		"processed_value_json": _json(processed),
		"original_value_json": _json(original),
		"exception_codes": _json(row.get("exception_codes") or []),
		"exception_message": row.get("exception_message") or "",
		"review_status": review_status,
		"proposed_value_json": _json(proposed),
		"confirmed_value_json": _json(row.get("confirmed_value")) if row.get("confirmed_value") is not None else "",
		"reviewer": row.get("reviewer") or "",
		"reviewed_on": row.get("reviewed_on") or None,
		"review_note": row.get("review_note") or "",
		"review_history_json": _json(row.get("review_history") or []),
		"eligible_for_downstream": 1 if downstream_eligible else 0,
		"source_file": row.get("source_file") or batch.source_file,
		"source_sheet": row.get("source_sheet") or "",
		"source_row": cint(row.get("source_row")),
		"source_id": row.get("source_id") or "",
		"approval_no": row.get("approval_no") or "",
	}


def _persist_processed_rows(batch, result):
	if frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": batch.name}):
		return
	for row in result["processed_rows"]:
		frappe.get_doc(_record_payload(batch, row)).insert(ignore_permissions=True)


def _result_rows(batch, page_length: int = 5000):
	records = frappe.get_all(
		PROCESSING_RECORD_DOCTYPE,
		filters={"import_batch": batch.name},
		fields=["name", "attendance_month", "employee_code", "employee_name", "department", "source_type", "processed_value_json", "original_value_json", "exception_codes", "exception_message", "review_status", "proposed_value_json", "confirmed_value_json", "reviewer", "reviewed_on", "review_note", "eligible_for_downstream", "source_file", "source_sheet", "source_row", "source_id", "approval_no"],
		order_by="employee_code asc, source_row asc",
		limit_page_length=page_length,
	)
	rows = [_serialize_record(row) for row in records]
	return _hydrate_apple_tree_result_rows(batch, rows) if batch.source_type == "apple_tree" else rows


def _restore_daily_exception_lines_from_source(record: dict[str, Any]) -> list[dict[str, Any]]:
	"""Rebuild date-level alerts from a historic record's original daily rows.

	Older batches can contain an employee-level ``ABSENCE_MARKED`` code while
	the then-current ``attendance_details`` projection did not retain the
	corresponding per-day marker.  The raw DingTalk rows are retained precisely
	for this audit case, so replay them read-only with the current field mapping.
	"""
	original = record.get("original_value") if isinstance(record.get("original_value"), dict) else {}
	source_rows = original.get("rows") if isinstance(original, dict) else []
	attendance_month = str(record.get("attendance_month") or "").strip()
	if not isinstance(source_rows, list) or not source_rows or not re.fullmatch(r"\d{4}-\d{2}", attendance_month):
		return []
	try:
		replayed = process_attendance_draft_rows(
			source_rows,
			attendance_month=attendance_month,
			source_file=record.get("source_file") or "",
			source_sheet=record.get("source_sheet") or "每日统计",
		)
	except Exception:
		return []
	matching_row = next(
		(
			row
			for row in replayed.get("processed_rows") or []
			if str(row.get("employee_code") or "") == str(record.get("employee_code") or "")
		),
		None,
	)
	if not matching_row:
		return []
	details = (matching_row.get("proposed_value") or {}).get("attendance_details") or []
	return exception_lines_from_attendance_details(details, record.get("exception_codes") or [])


def _serialize_record(record):
	result = dict(record)
	result["record_id"] = result.pop("name")
	result["processed_value"] = _loads(result.pop("processed_value_json", ""), {})
	result["original_value"] = _loads(result.pop("original_value_json", ""), {})
	result["exception_codes"] = _loads(result["exception_codes"], [])
	result["proposed_value"] = _loads(result.pop("proposed_value_json", ""), {})
	result["confirmed_value"] = _loads(result.pop("confirmed_value_json", ""), None)
	result["department"] = _display_department(result.get("department"))
	for values in (result["processed_value"], result["proposed_value"], result["confirmed_value"]):
		if isinstance(values, dict):
			for fieldname in ("department", "部门"):
				if values.get(fieldname) not in (None, ""):
					values[fieldname] = _display_department(values[fieldname])
	result["result_summary"] = _json(result["confirmed_value"] or result["proposed_value"] or result["processed_value"])
	result["source_label"] = SOURCE_LABELS.get(result.get("source_type"), result.get("source_type") or "--")
	result["exception_labels"] = [EXCEPTION_LABELS.get(code, "待人工确认") for code in result["exception_codes"]]
	result["department_comparison"] = _department_comparison(result)
	if result["department_comparison"]:
		comparison = result["department_comparison"]
		result["exception_detail"] = "钉钉考勤表部门：{0}；花名册部门：{1}。".format(comparison["source_department"], comparison["roster_department"])
	else:
		result["exception_detail"] = result.get("exception_message") or ""
	result["review_guidance"] = _review_guidance(result["exception_codes"], result.get("source_type") or "")
	result["review_options"] = _review_options(result["exception_codes"], result.get("source_type") or "")
	if result.get("source_type") == "attendance_draft":
		values = _effective_result_values(result)
		# New batches persist exception_lines.  Rebuild them from the retained
		# daily facts for old batches.  If the historic projection itself lacks a
		# marker (for example, 旷工), replay its retained original rows read-only.
		result["daily_exception_lines"] = values.get("exception_lines") or exception_lines_from_attendance_details(
			values.get("attendance_details") or [], result["exception_codes"]
		) or _restore_daily_exception_lines_from_source(result)
		result["daily_attendance_details"] = values.get("attendance_details") or []
	return result


def _effective_result_values(row: dict[str, Any]) -> dict[str, Any]:
	"""Build the current display/export row without changing its audit trail."""
	values = dict(row.get("processed_value") or {})
	values.update(row.get("proposed_value") or {})
	if row.get("confirmed_value") is not None:
		values.update(row["confirmed_value"])
	for fieldname in ("employee_code", "employee_name", "department"):
		if row.get(fieldname) not in (None, ""):
			values[fieldname] = row[fieldname]
	# Proposed values are intentionally retained for review, but they cannot be
	# presented as final red-apple amounts while a row is pending or rejected.
	if "included" in values:
		values["included"] = bool(row.get("eligible_for_downstream"))
		if not values["included"]:
			values["red_apples"] = 0
			values["amount"] = 0
	return values


def _daily_row_overrides(values: dict[str, Any]) -> dict[str, dict[str, Any]]:
	overrides = values.get("_daily_row_overrides", {}) if isinstance(values, dict) else {}
	return overrides if isinstance(overrides, dict) else {}


def _effective_daily_source_rows(row: dict[str, Any]) -> list[dict[str, Any]]:
	"""Return the source rows with approved in-system corrections overlaid."""
	original = row.get("original_value") or {}
	source_rows = original.get("rows") if isinstance(original, dict) else []
	if not isinstance(source_rows, list):
		return []
	values = _effective_result_values(row)
	overrides = _daily_row_overrides(values)
	effective_rows = []
	for source in source_rows:
		if not isinstance(source, dict):
			continue
		daily = dict(source)
		override = overrides.get(str(daily.get("source_row") or daily.get("_source_row") or ""), {})
		if isinstance(override, dict):
			daily.update(override)
		effective_rows.append(daily)
	return effective_rows


def _daily_row_editor_payload(row: dict[str, Any]) -> list[dict[str, Any]]:
	items = []
	for source in _effective_daily_source_rows(row):
		source_row = source.get("source_row") or source.get("_source_row")
		try:
			source_row = int(source_row)
		except (TypeError, ValueError):
			continue
		fields = []
		for default_name, label, aliases in ATTENDANCE_DAILY_EDIT_FIELDS:
			fieldname = next((alias for alias in aliases if alias in source), default_name)
			fields.append({"fieldname": fieldname, "label": label, "value": source.get(fieldname, "")})
		items.append({
			"source_row": source_row,
			"attendance_date": _daily_attendance_date(source.get("日期") or source.get("考勤日期")),
			"source_values": source,
			"editable_fields": fields,
		})
	return items


def _missed_punch_summary(rows: list[dict[str, Any]]) -> dict[str, int | float]:
	"""Calculate the one result sheet's effective, downstream-eligible totals."""
	def number(value: Any) -> int | float:
		try:
			parsed = float(value or 0)
		except (TypeError, ValueError):
			return 0
		return int(parsed) if parsed.is_integer() else parsed

	summary: dict[str, int | float] = {
		"included_rows": 0,
		"green_apples": 0,
		"red_apples": 0,
		"amount": 0,
	}
	for row in rows:
		if not row.get("eligible_for_downstream"):
			continue
		values = _effective_result_values(row)
		if not values.get("included"):
			continue
		summary["included_rows"] += 1
		summary["red_apples"] += number(values.get("red_apples"))
		summary["amount"] += number(values.get("amount"))
	return summary


def _signoff_datetime(value: Any) -> Any:
	"""Keep sign-off dates sortable in Excel without inventing invalid values."""
	if isinstance(value, datetime) or value in (None, ""):
		return value or ""
	try:
		parsed = datetime.fromisoformat(str(value).strip())
		return parsed.replace(tzinfo=None) if parsed.tzinfo else parsed
	except (TypeError, ValueError):
		return value


def _missed_punch_signoff_values(row: dict[str, Any], fallback_sequence: int) -> list[Any]:
	"""Project one auditable record into the printable nine-column list."""
	values = _effective_result_values(row)
	original = row.get("original_value") if isinstance(row.get("original_value"), dict) else {}
	sequence = _row_value(original, "序号")
	return [
		sequence if sequence not in (None, "") else fallback_sequence,
		values.get("department") or _row_value(original, "创建人部门", "部门") or "",
		_signoff_datetime(values.get("created_at") or _row_value(original, "创建时间")),
		_signoff_datetime(values.get("punch_time") or _row_value(original, "补卡时间")),
		values.get("punch_type") or _row_value(original, "补卡类型") or "",
		values.get("reason") or _row_value(original, "补卡理由") or "",
		values.get("employee_name") or _row_value(original, "创建人", "姓名") or "",
		"",
		"",
	]


def _apple_tree_summary(rows: list[dict[str, Any]]) -> dict[str, int | float]:
	"""Total downstream-eligible green and red apples independently."""
	def number(value: Any) -> int | float:
		try:
			parsed = float(value or 0)
		except (TypeError, ValueError):
			return 0
		return int(parsed) if parsed.is_integer() else parsed

	summary: dict[str, int | float] = {"green_apples": 0, "red_apples": 0}
	for row in rows:
		if not row.get("eligible_for_downstream"):
			continue
		values = _effective_result_values(row)
		amount = number(values.get("有效苹果数"))
		if values.get("苹果类型") == "绿苹果":
			summary["green_apples"] += amount
		elif values.get("苹果类型") == "红苹果":
			summary["red_apples"] += amount
	return summary


def _apple_tree_result_values(row: dict[str, Any], raw: dict[str, Any] | None = None) -> dict[str, Any]:
	"""Project current and legacy Apple-tree records into one visible layout."""
	raw = raw or row.get("original_value") or {}
	current = _effective_result_values(row)
	value = lambda *keys: _row_value(raw, *keys) or _row_value(current, *keys)
	return {
		"创建时间": value("创建时间", "created_at"),
		"奖惩日期": value("奖/惩日期", "奖惩日期", "award_date"),
		"部门": _display_department(value("受奖/惩人部门", "部门", "department")),
		"姓名": value("受奖/惩人", "姓名", "employee_name"),
		"绿苹果": value("绿苹果", "原始绿苹果"),
		"红苹果": value("红苹果", "原始红苹果"),
		"项目": value("奖/惩项目", "项目", "project"),
		"备注": value("备注", "remark"),
		"创建人": value("创建人", "creator"),
		"工号": _row_value(current, "工号", "employee_code") or row.get("employee_code") or "",
		"苹果类型": _row_value(current, "苹果类型") or "",
		"有效苹果数": _row_value(current, "有效苹果数") or "",
		"审批编号": value("审批编号", "approval_no") or row.get("approval_no") or "",
		"审批结果": value("审批结果", "approval_result"),
		"审批状态": value("审批状态", "approval_status"),
	}


def _signoff_number(value: Any) -> Any:
	"""Use numeric apple counts when the retained source value is numeric text."""
	if value in (None, "") or isinstance(value, (int, float)):
		return value or ""
	try:
		parsed = float(str(value).strip())
		return int(parsed) if parsed.is_integer() else parsed
	except (TypeError, ValueError):
		return value


def _apple_tree_signoff_values(row: dict[str, Any], fallback_sequence: int) -> list[Any]:
	"""Project one auditable Apple-tree record into the printable form."""
	original = row.get("original_value") if isinstance(row.get("original_value"), dict) else {}
	values = row.get("processed_value") or _apple_tree_result_values(row)
	sequence = _row_value(original, "序号")
	return [
		sequence if sequence not in (None, "") else fallback_sequence,
		_signoff_datetime(values.get("创建时间")),
		_signoff_datetime(values.get("奖惩日期")),
		values.get("部门", ""),
		values.get("姓名", ""),
		_signoff_number(values.get("绿苹果")),
		_signoff_number(values.get("红苹果")),
		values.get("项目", ""),
		values.get("备注", ""),
		values.get("创建人", ""),
		"",
		"",
	]


def _apple_tree_signoff_title(attendance_month: Any) -> str:
	try:
		return f"{int(str(attendance_month).split('-')[1])}月苹果树"
	except (IndexError, TypeError, ValueError):
		return "苹果树"


def _hydrate_apple_tree_result_rows(batch, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
	"""Backfill old simplified records from their retained source rows.

	Older batches persisted only editable fields.  Reading the immutable uploaded
	file by source-row makes both their grid and their new download usable without
	changing the audit record.
	"""
	raw_by_source_row = {}
	try:
		source_rows, _sheet_name = _read_source_rows(batch)
		raw_by_source_row = {str(raw.get("source_row")): raw for raw in source_rows}
	except Exception:
		# A missing historical attachment must not stop users viewing/exporting the
		# fields that were already persisted.
		pass
	for row in rows:
		raw = raw_by_source_row.get(str(row.get("source_row")))
		row["processed_value"] = _apple_tree_result_values(row, raw)
	return rows


def _apple_tree_trace(row: dict[str, Any]) -> str:
	return " · ".join(
		str(value) for value in (row.get("source_file"), row.get("source_sheet"), row.get("source_row"), row.get("source_id"), row.get("approval_no"))
		if value not in (None, "")
	) or "--"


def _export_processed_result(batch) -> dict[str, str]:
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter
	from frappe.utils.file_manager import save_file

	rows = _result_rows(batch)
	book = Workbook()
	sheet = book.active
	sheet.title = "加工结果"
	if batch.source_type == "attendance_draft":
		# Source file, sheet and row stay in the audit record, but do not belong in
		# the HR-facing processing workbook.
		headers = ["序号"] + [label for _field, label in ATTENDANCE_DRAFT_RESULT_COLUMNS] + ["异常说明", "处理状态", "是否计入下游"]
	elif batch.source_type == "apple_tree":
		headers = list(APPLE_TREE_SIGNOFF_COLUMNS)
		sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
		sheet.cell(row=1, column=1, value=_apple_tree_signoff_title(batch.attendance_month))
	elif batch.source_type == "missing_card":
		headers = list(MISSED_PUNCH_SIGNOFF_COLUMNS)
	else:
		headers = ["工号", "姓名", "部门", "加工结果", "异常类型", "异常说明", "处理状态", "建议值", "确认值", "可进入下游"]
	sheet.append(headers)
	for index, row in enumerate(rows, start=1):
		if batch.source_type == "attendance_draft":
			values = _effective_result_values(row)
			sheet.append([
				index,
				*[values.get(field, "") for field, _label in ATTENDANCE_DRAFT_RESULT_COLUMNS],
				"；".join(row.get("exception_labels") or []) or "无",
				row.get("review_status"),
				"是" if row.get("eligible_for_downstream") else "否",
			])
		elif batch.source_type == "apple_tree":
			sheet.append(_apple_tree_signoff_values(row, index))
		elif batch.source_type == "missing_card":
			sheet.append(_missed_punch_signoff_values(row, index))
		else:
			sheet.append([
				row.get("employee_code"), row.get("employee_name"), row.get("department"), _json(row.get("processed_value")),
				"、".join(row.get("exception_labels") or []), row.get("exception_message"), row.get("review_status"),
				_json(row.get("proposed_value")), _json(row.get("confirmed_value")) if row.get("confirmed_value") is not None else "",
				"是" if row.get("eligible_for_downstream") else "否",
			])
	if batch.source_type == "missing_card":
		thin = Side(style="thin", color="000000")
		border = Border(left=thin, right=thin, top=thin, bottom=thin)
		for cell in sheet[1]:
			cell.fill = PatternFill("solid", fgColor="D9D9D9")
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.border = border
		sheet.row_dimensions[1].height = 28
		for row_cells in sheet.iter_rows(min_row=2, min_col=3, max_col=4):
			for cell in row_cells:
				if isinstance(cell.value, datetime):
					cell.number_format = "yyyy-mm-dd hh:mm:ss"
		for row_cells in sheet.iter_rows(min_row=2, max_col=len(MISSED_PUNCH_SIGNOFF_COLUMNS)):
			for cell in row_cells:
				cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
				cell.border = border
			sheet.row_dimensions[row_cells[0].row].height = 36
		sheet.freeze_panes = "A2"
		sheet.auto_filter.ref = sheet.dimensions
		sheet.sheet_view.showGridLines = False
		sheet.page_setup.orientation = "landscape"
		sheet.page_setup.fitToWidth = 1
		sheet.print_title_rows = "1:1"
	elif batch.source_type == "apple_tree":
		thin = Side(style="thin", color="000000")
		border = Border(left=thin, right=thin, top=thin, bottom=thin)
		title = sheet["A1"]
		title.font = Font(bold=True, size=22)
		title.alignment = Alignment(horizontal="center", vertical="center")
		sheet.row_dimensions[1].height = 34
		for cell in sheet[2]:
			cell.fill = PatternFill("solid", fgColor="BDD7EE")
			cell.font = Font(bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.border = border
		sheet.row_dimensions[2].height = 34
		for row_cells in sheet.iter_rows(min_row=3, max_col=len(APPLE_TREE_SIGNOFF_COLUMNS)):
			for cell in row_cells:
				cell.alignment = Alignment(
					horizontal="left" if cell.column in (8, 9) else "center",
					vertical="center",
					wrap_text=True,
				)
				cell.border = border
			sheet.row_dimensions[row_cells[0].row].height = 48
		for cell in sheet["B"][2:]:
			if isinstance(cell.value, datetime):
				cell.number_format = "yyyy-mm-dd hh:mm:ss"
		for cell in sheet["C"][2:]:
			if isinstance(cell.value, datetime):
				cell.number_format = "yyyy-mm-dd"
		sheet.freeze_panes = "A3"
		sheet.auto_filter.ref = f"A2:L{sheet.max_row}"
		sheet.sheet_view.showGridLines = False
		sheet.page_setup.orientation = "landscape"
		sheet.page_setup.fitToWidth = 1
		sheet.print_title_rows = "1:2"
	for column_index, column in enumerate(sheet.iter_cols(), start=1):
		sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 48)
	if batch.source_type == "missing_card":
		for column_letter, width in {"A": 8, "B": 12, "C": 21, "D": 21, "E": 14, "F": 18, "G": 12, "H": 14, "I": 14}.items():
			sheet.column_dimensions[column_letter].width = width
	elif batch.source_type == "apple_tree":
		for column_letter, width in {"A": 8, "B": 20, "C": 14, "D": 15, "E": 12, "F": 8, "G": 8, "H": 50, "I": 28, "J": 12, "K": 14, "L": 16}.items():
			sheet.column_dimensions[column_letter].width = width
	output = BytesIO()
	book.save(output)
	file = save_file(f"{batch.attendance_month}_{SOURCE_LABELS[batch.source_type]}_加工结果.xlsx", output.getvalue(), None, None, is_private=1)
	return {"file_url": file.file_url, "file_name": file.file_name}


def _slot_payload(batch):
	if not batch:
		return None
	meta = _processing_meta(batch)
	metrics = meta.get("metrics", {})
	processed_rows = frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": batch.name})
	pending_exception_count = frappe.db.count(
		PROCESSING_RECORD_DOCTYPE,
		{"import_batch": batch.name, "review_status": "待审核"},
	)
	# A confirmed source may deliberately retain unresolved exception records for
	# audit.  They are excluded from downstream calculation at confirmation time,
	# so a live queue count must not make the source card appear unconfirmed.
	# For every other batch, retain the live queue-derived status so review actions
	# are reflected immediately.
	effective_status = (
		batch.status
		if batch.status == "已确认"
		else "待处理异常"
		if pending_exception_count
		else batch.status
	)
	return {
		"source_type": batch.source_type,
		"source_file": batch.source_file,
		"source_file_name": Path(batch.source_file or "").name,
		"attendance_month": batch.attendance_month,
		"row_count": metrics.get("source_rows", batch.daily_sheet_rows or 0),
		"eligible_employee_source_rows": metrics.get("eligible_employee_source_rows", 0),
		"employee_summary_count": metrics.get("processed_rows", processed_rows),
		# This is a live pending count, rather than the historic number detected
		# at import. Resolved rows stay auditable but disappear from the work queue.
		"exception_count": pending_exception_count,
		"historic_exception_count": metrics.get("exception_rows", 0),
		"exception_event_count": metrics.get("exception_events", metrics.get("exception_rows", 0)),
		"status": effective_status,
		"stored_status": batch.status,
		"can_precheck": bool(batch.source_file) and batch.status in {"待加工", "预览", "已导入"},
		"can_process": bool(batch.source_file) and batch.status in {"待加工", "预览", "已导入"},
		"can_edit": batch.status in {"待处理异常", "待确认", "已确认"},
		"can_confirm": bool(processed_rows) and batch.status in {"待处理异常", "待确认"},
		"precheck": meta.get("precheck"),
		"data_quality": meta.get("data_quality", {}),
		"processed_result": meta.get("processed_result"),
	}


def _refresh_batch_review_status(batch):
	"""Keep a confirmed source usable while unresolved rows remain excluded."""
	if batch.status == "已确认":
		return batch.status
	pending_rows = frappe.db.count(
		PROCESSING_RECORD_DOCTYPE,
		{"import_batch": batch.name, "review_status": "待审核"},
	)
	batch.status = "待处理异常" if pending_rows else "待确认"
	batch.save(ignore_permissions=True)
	return batch.status


def _monthly_final_employee_recognition(company: str, attendance_month: str) -> dict[str, int]:
	"""Summarise the attendance-draft/roster match without counting daily rows.

	The monthly final is one row per employee, whereas the DingTalk source is one
	row per employee per day.  These figures therefore use the persisted
	employee-centred attendance-draft records and only use a roster employee when
	it has a business employee code that the source can actually match.
	"""
	batch = _latest_batch(company, attendance_month, "attendance_draft")
	records = _result_rows(batch, 5000) if batch else []

	def employee_key(record: dict[str, Any]) -> str:
		code = str(record.get("employee_code") or "").strip()
		name = str(record.get("employee_name") or "").strip()
		return code or (f"name:{name}" if name else "")

	draft_people = {key for record in records if (key := employee_key(record))}
	successful_people = {
		key
		for record in records
		if record.get("eligible_for_downstream") and (key := employee_key(record))
	}
	roster_people = {
		str(employee.get("employee_code") or "").strip()
		for employee in _employee_directory(company)
		if str(employee.get("employee_code") or "").strip()
	}
	return {
		"draft_recognized_employee_count": len(draft_people),
		"roster_employee_count": len(roster_people),
		"successful_employee_count": len(successful_people),
	}


@frappe.whitelist()
def get_processing_batch(company: str, attendance_month: str):
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	slots = []
	for source_type in SOURCE_TYPES:
		batch = _latest_batch(company, attendance_month, source_type)
		slot = _slot_payload(batch)
		if slot:
			slots.append(slot)
	batch_statuses = [slot["status"] for slot in slots]
	status = "未上传" if not slots else "已确认" if len(slots) == 3 and all(value == "已确认" for value in batch_statuses) else "待处理异常" if "待处理异常" in batch_statuses else "待确认" if "待确认" in batch_statuses else "待加工"
	finalization_inputs = _finalization_inputs(company, attendance_month, slots)
	anchor_batch = _latest_batch(company, attendance_month, "attendance_draft")
	anchor_meta = _processing_meta(anchor_batch) if anchor_batch else {}
	final_outputs = anchor_meta.get("monthly_final_outputs", {})
	return {
		"batch_id": f"{company}:{attendance_month}",
		"company": company,
		"attendance_month": attendance_month,
		"status": status,
		"slots": slots,
		"finalization_inputs": finalization_inputs,
		"final_outputs": final_outputs,
		"signed_final_reconciliation": anchor_meta.get("signed_final_reconciliation", {}),
		"employee_recognition": _monthly_final_employee_recognition(company, attendance_month),
		"locked_snapshot_version": final_outputs.get("locked_snapshot_version", ""),
		"snapshot_ready": bool(finalization_inputs) and all(item["ready"] for item in finalization_inputs),
	}


@frappe.whitelist()
def register_source_file(company: str, attendance_month: str, source_type: str, file_url: str):
	_require_processing_manager()
	company, attendance_month, source_type = _require_company(company), _require_month(attendance_month), _require_source_type(source_type)
	if not file_url:
		frappe.throw(_("请先上传来源文件。"))
	if not file_url.lower().split("?", 1)[0].endswith(".xlsx"):
		frappe.throw(_("当前考勤处理仅接受 .xlsx 文件，以保证结构预检和加工结果可追溯。"))
	checksum = _file_checksum(file_url)
	batch = frappe.get_doc({
		"doctype": IMPORT_BATCH_DOCTYPE,
		"company": company,
		"attendance_month": attendance_month,
		"source_file": file_url,
		"source_type": source_type,
		"source_checksum": checksum,
		"status": "待加工",
		"imported_by": frappe.session.user,
		"imported_on": now_datetime(),
		"notes": _json({"attendance_processing_center": {"source_version": now_datetime().isoformat(), "registered_by": frappe.session.user}}),
	}).insert(ignore_permissions=True)
	return {"batch": batch.name, "source_type": source_type, "status": batch.status, "file_url": file_url}


@frappe.whitelist()
def register_monthly_support_file(company: str, attendance_month: str, source_type: str, file_url: str):
	"""Register one allowance/award workbook for the monthly-final gate.

	It remains a separate source batch so an upload never overwrites the prior
	confirmed version and the eventual monthly snapshot can be audited.
	"""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	source_type = _require_monthly_support_source_type(source_type)
	if not file_url:
		frappe.throw(_("请先上传来源文件。"))
	if not file_url.lower().split("?", 1)[0].endswith(".xlsx"):
		frappe.throw(_("月度补充来源仅接受 .xlsx 文件。"))
	batch = frappe.get_doc({
		"doctype": IMPORT_BATCH_DOCTYPE,
		"company": company,
		"attendance_month": attendance_month,
		"source_file": file_url,
		"source_type": source_type,
		"source_checksum": _file_checksum(file_url),
		"status": "待加工",
		"imported_by": frappe.session.user,
		"imported_on": now_datetime(),
		"notes": _json({"attendance_processing_center": {
			"source_version": now_datetime().isoformat(),
			"registered_by": frappe.session.user,
			"monthly_support": True,
		}}),
	}).insert(ignore_permissions=True)
	return {"batch": batch.name, "source_type": source_type, "status": batch.status, "file_url": file_url}


def _detect_bulk_source_type(file_url: str, file_name: str = "") -> str:
	"""Identify an attendance source by workbook structure.

	The workbook is the source of truth. File names vary when HR downloads or
	renames a DingTalk export, so a name is used only as a fallback when its
	structure cannot be read.
	"""
	name = f"{file_name} {Path(file_url).name}".replace(" ", "")
	name_hints = (
		("住房", "housing_allowance"), ("全勤", "full_attendance"), ("特殊工时", "special_hours"),
		("苹果", "apple_tree"), ("忘打卡", "missing_card"), ("补卡", "missing_card"), ("考勤初稿", "attendance_draft"),
	)
	try:
		workbook = _load_workbook(file_url)
		# File and worksheet names vary by DingTalk report version. Identify the
		# daily attendance source from its required headers, not only its name.
		if find_dingtalk_daily_sheet(workbook):
			return "attendance_draft"
		for sheet in workbook.worksheets:
			# Some monthly templates have a title row before their headers.
			values = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True))
			text = "|".join(str(value or "") for row in values for value in row)
			if "每日明细（钉钉导出）" in sheet.title or "每日明细（钉钉导出）" in text:
				return "attendance_draft"
			if "受奖/惩人" in text and ("绿苹果" in text or "红苹果" in text):
				return "apple_tree"
			if "补卡时间" in text and "补卡类型" in text:
				return "missing_card"
			# The full-attendance workbook may be derived from the housing template
			# and retain a copied right-hand ``住房补贴`` header.  Prefer its specific
			# marker so the two independent attendance sources are not confused.
			if "全勤奖" in text and "工号" in text and "姓名" in text:
				return "full_attendance"
			if "住房补贴" in text and "工号" in text and "姓名" in text:
				return "housing_allowance"
			if "特殊工时" in text and "工号" in text and "姓名" in text:
				return "special_hours"
	except Exception:
		# The normal upload/precheck path reports the precise workbook problem.
		# Keep filename fallback for older, established HR templates.
		pass
	for marker, source_type in name_hints:
		if marker in name:
			return source_type
	return ""


@frappe.whitelist()
def bulk_import_and_process_sources(company: str, attendance_month: str, files: str | list[dict[str, Any]]):
	"""Register and apply all six attendance sources in one step.

	The files are fully classified before any batch is written.  That gives HR
	the convenience of selecting six files together without silently putting a
	file in the wrong source slot. Invalid rows stay traceable and are excluded
	from calculation; they do not require a separate approval step.
	"""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	if isinstance(files, str):
		files = _loads(files, [])
	if not isinstance(files, list):
		frappe.throw(_("请选择六个 .xlsx 考勤来源文件。"))
	classified: dict[str, dict[str, Any]] = {}
	unmatched = []
	duplicates = []
	for item in files:
		if not isinstance(item, dict):
			unmatched.append(str(item))
			continue
		file_url = str(item.get("file_url") or "").strip()
		file_name = str(item.get("file_name") or "").strip()
		if not file_url.lower().split("?", 1)[0].endswith(".xlsx"):
			unmatched.append(file_name or file_url)
			continue
		source_type = _detect_bulk_source_type(file_url, file_name)
		if not source_type:
			unmatched.append(file_name or file_url)
			continue
		if source_type in classified:
			duplicates.append(SOURCE_LABELS[source_type])
			continue
		classified[source_type] = {"file_url": file_url, "file_name": file_name or Path(file_url).name}
	missing = [SOURCE_LABELS[source_type] for source_type in SOURCE_TYPES + MONTHLY_SUPPORT_SOURCE_TYPES if source_type not in classified]
	if unmatched or duplicates or missing:
		issues = []
		if unmatched:
			issues.append(_("无法识别：{0}").format("、".join(unmatched)))
		if duplicates:
			issues.append(_("重复来源：{0}").format("、".join(duplicates)))
		if missing:
			issues.append(_("缺少来源：{0}").format("、".join(missing)))
		frappe.throw(_("批量导入未开始。请调整文件后重试：{0}").format("；".join(issues)))
	items = []
	for source_type in SOURCE_TYPES + MONTHLY_SUPPORT_SOURCE_TYPES:
		item = classified[source_type]
		if source_type in SOURCE_TYPES:
			registered = register_source_file(company, attendance_month, source_type, item["file_url"])
			processed = process_source_slot(company, attendance_month, source_type)
		else:
			registered = register_monthly_support_file(company, attendance_month, source_type, item["file_url"])
			precheck_monthly_support_file(company, attendance_month, source_type)
			processed = process_monthly_support_file(company, attendance_month, source_type)
		metrics = processed.get("metrics") or {}
		items.append({
			"source_type": source_type,
			"label": SOURCE_LABELS[source_type],
			"file_name": item["file_name"],
			"batch": registered.get("batch"),
			"status": processed.get("status"),
			"processed_rows": metrics.get("processed_rows", 0),
			"exception_rows": metrics.get("exception_rows", 0),
		})
	return {
		"items": items,
		"notice": _("六类考勤来源已完成匹配、导入并生效。无效记录已自动排除，不需要逐类审批。"),
	}


@frappe.whitelist()
def precheck_monthly_support_file(company: str, attendance_month: str, source_type: str):
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	source_type = _require_monthly_support_source_type(source_type)
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		frappe.throw(_("请先上传该月度补充来源文件。"))
	if batch.status == "已确认" and frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": batch.name}):
		frappe.throw(_("该来源已经确认；如需更正，请上传新的来源版本。"))
	precheck = _monthly_support_precheck(batch)
	if precheck["is_valid"]:
		batch.status = "待加工"
		batch.daily_sheet_rows = cint(precheck.get("record_count"))
	else:
		batch.status = "结构异常"
	batch.save(ignore_permissions=True)
	_save_batch_notes(batch, {"monthly_support_precheck": precheck, "metrics": {"source_rows": cint(precheck.get("record_count"))}})
	return {"batch": batch.name, "source_type": source_type, "status": batch.status, "precheck": precheck}


@frappe.whitelist()
def process_monthly_support_file(company: str, attendance_month: str, source_type: str):
	"""Perform the one-time import and retain its validation result."""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	source_type = _require_monthly_support_source_type(source_type)
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		frappe.throw(_("请先上传该月度补充来源文件。"))
	if batch.status in {"已确认", "导入异常"} and frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": batch.name}):
		return {
			"batch": batch.name,
			"source_type": source_type,
			"status": batch.status,
			"metrics": _processing_meta(batch).get("metrics", {}),
			"processed_result": _processing_meta(batch).get("processed_result"),
			"message": _("该文件已经完成导入校验；如需更正，请重新上传新的来源版本。"),
		}
	# 前台不再暴露“预检”这一步；点击加工时自动执行结构校验，
	# 只有失败才留在卡片上提示用户重新上传。
	precheck = _monthly_support_precheck(batch)
	_save_batch_notes(batch, {"monthly_support_precheck": precheck})
	if not precheck.get("is_valid"):
		batch.status = "结构异常"
		batch.save(ignore_permissions=True)
		frappe.throw(precheck.get("message") or _("来源文件结构不符合要求。"))
	result = _process_monthly_support_rows(batch)
	_persist_processed_rows(batch, result)
	batch.status = result["status"]
	batch.daily_sheet_rows = cint(result["metrics"]["source_rows"])
	batch.save(ignore_permissions=True)
	processed_result = _export_processed_result(batch)
	_save_batch_notes(batch, {
		"metrics": result["metrics"],
		"processed_result": processed_result,
		"processed_on": now_datetime().isoformat(),
		"monthly_support_processing_version": 2,
		"monthly_support_import_mode": "one_time_import",
	})
	return {"batch": batch.name, "source_type": source_type, "status": batch.status, "metrics": result["metrics"], "processed_result": processed_result}


@frappe.whitelist()
def confirm_monthly_support_file(company: str, attendance_month: str, source_type: str):
	"""Compatibility endpoint for the retired monthly-support confirmation step."""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	source_type = _require_monthly_support_source_type(source_type)
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		frappe.throw(_("尚未上传该月度补充来源文件。"))
	if batch.status == "导入异常":
		frappe.throw(_("该文件存在导入校验错误，请更正后重新上传；此类来源不支持异常处理或人工确认。"))
	if batch.status != "已确认":
		frappe.throw(_("请先完成一次性导入校验。"))
	return {
		"batch": batch.name,
		"source_type": source_type,
		"status": batch.status,
		"record_count": cint((_processing_meta(batch).get("monthly_support_precheck") or {}).get("record_count")),
		"message": _("该来源已随导入自动生效，无需再次确认。"),
	}


@frappe.whitelist()
def precheck_source_slot(company: str, attendance_month: str, source_type: str):
	_require_processing_manager()
	company, attendance_month, source_type = _require_company(company), _require_month(attendance_month), _require_source_type(source_type)
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		frappe.throw(_("请先上传该来源文件。"))
	precheck = _precheck(batch)
	row_count = precheck["row_count"]
	if source_type == "attendance_draft":
		batch.daily_sheet_rows = row_count
	else:
		batch.apple_sheet_rows = row_count
	_save_batch_notes(batch, {"precheck": precheck, "metrics": {"source_rows": row_count, "processed_rows": 0, "exception_rows": 0}})
	return {"batch": batch.name, "source_type": source_type, "status": batch.status, "precheck": precheck}


@frappe.whitelist()
def process_source_slot(company: str, attendance_month: str, source_type: str):
	_require_processing_manager()
	company, attendance_month, source_type = _require_company(company), _require_month(attendance_month), _require_source_type(source_type)
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		frappe.throw(_("请先上传该来源文件。"))
	if frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": batch.name}):
		return {"batch": batch.name, "status": batch.status, "processed_result": _processing_meta(batch).get("processed_result"), "message": _("该来源版本已加工；如需重新处理，请重新上传形成新版本。")}
	result = _process_batch(batch)
	_persist_processed_rows(batch, result)
	structure_precheck = result.get("structure_precheck") or {}
	if structure_precheck and not structure_precheck.get("is_valid", True):
		batch.status = "结构异常"
	else:
		# Valid records take effect immediately. Rows needing attention remain in
		# the audit result and are excluded from downstream calculation; do not
		# wait for a second source-confirmation or approval action.
		batch.status = "已确认"
	result["status"] = batch.status
	processed_result = _export_processed_result(batch)
	_save_batch_notes(batch, {
		"precheck": result.get("structure_precheck"),
		"metrics": result.get("metrics", {}),
		"data_quality": result.get("data_quality", {}),
		"processed_result": processed_result,
		"processed_on": now_datetime().isoformat(),
	})
	return {"batch": batch.name, "source_type": source_type, "status": batch.status, "processed_result": processed_result, "metrics": result.get("metrics", {})}


@frappe.whitelist()
def list_processing_results(company: str, attendance_month: str, source_type: str, exception_only: int = 0, page_length: int = 500):
	_require_processing_manager()
	company, attendance_month, source_type = _require_company(company), _require_month(attendance_month), _require_processing_source_type(source_type)
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		return {"processed_rows": [], "can_confirm": False}
	meta = _processing_meta(batch)
	rows = _result_rows(batch, min(max(cint(page_length), 1), 5000))
	if cint(exception_only):
		rows = [row for row in rows if row["exception_codes"] and row["review_status"] == "待审核"]
	return {
		"batch": batch.name,
		"status": batch.status,
		"processed_rows": rows,
		"processed_result": meta.get("processed_result"),
		"import_validation": meta.get("monthly_support_precheck") if source_type in MONTHLY_SUPPORT_SOURCE_TYPES else {},
		"result_summary": _missed_punch_summary(rows) if source_type == "missing_card" else _apple_tree_summary(rows) if source_type == "apple_tree" else {},
		"can_confirm": bool(rows) or batch.status in {"待处理异常", "待确认"},
	}


def _attendance_draft_data_quality_value(row: dict[str, Any], *fieldnames: str) -> str:
	"""Read an attendance-draft identity field without guessing from names."""
	return str(_row_value(row, *fieldnames) or "").strip()


def _attendance_draft_data_quality_item(row: dict[str, Any]) -> dict[str, Any]:
	"""Expose a source row read-only, including its original Excel row number."""
	metadata_fields = {"source_file", "source_sheet", "source_row"}
	return {
		"source_row": cint(row.get("source_row")),
		"values": {
			fieldname: "" if value is None else str(value)
			for fieldname, value in row.items()
			if fieldname not in metadata_fields
		},
	}


@frappe.whitelist()
def get_attendance_data_quality_details(
	company: str,
	attendance_month: str,
	source_type: str,
	quality_type: str,
	page_start: int = 0,
	page_length: int = 500,
):
	"""Return the source rows behind the attendance-draft quality notices.

	These rows deliberately remain outside the employee processing records: a
	missing employee code must not become an employee exception.  Read the
	immutable uploaded workbook again so the detail view is complete and remains
	traceable even after the source batch has been confirmed.
	"""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	if _require_source_type(source_type) != "attendance_draft":
		frappe.throw(_("仅考勤初稿支持查看这类数据质量明细。"))
	quality_type = str(quality_type or "").strip()
	quality_labels = {
		"missing_employee_code": "无工号来源行",
		"blank_shift": "入离职期间空班次",
	}
	if quality_type not in quality_labels:
		frappe.throw(_("不支持的数据质量明细类型。"))
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch or not batch.source_file:
		frappe.throw(_("尚未找到当前月份的考勤初稿来源文件。"))
	workbook = _load_workbook(batch.source_file)
	sheet = find_dingtalk_daily_sheet(workbook)
	if not sheet:
		frappe.throw(_("来源文件中未找到可读取的钉钉每日明细。"))
	rows = rows_from_dingtalk_daily_sheet(sheet, source_file=batch.source_file)
	if quality_type == "missing_employee_code":
		matched_rows = [
			row for row in rows
			if not _attendance_draft_data_quality_value(row, "工号", "员工工号", "employee_code")
		]
	else:
		# This is intentionally the same rule used by the processor: every
		# employee-code row with a blank shift is retained as a quality event.
		matched_rows = [
			row for row in rows
			if _attendance_draft_data_quality_value(row, "工号", "员工工号", "employee_code")
			and not _attendance_draft_data_quality_value(row, "班次", "shift")
		]
	columns = list(dict.fromkeys(
		fieldname
		for row in matched_rows
		for fieldname in row
		if fieldname not in {"source_file", "source_sheet", "source_row"}
	))
	page_start = max(cint(page_start), 0)
	page_length = min(max(cint(page_length), 1), 500)
	return {
		"title": quality_labels[quality_type],
		"description": (
			"以下为工号为空、未进入员工异常的原始来源行。"
			if quality_type == "missing_employee_code"
			else "以下为有工号但班次为空、仅作为数据质量证据保留的原始来源行。"
		),
		"source_file_name": Path(batch.source_file).name,
		"source_sheet": sheet.title,
		"total_count": len(matched_rows),
		"columns": columns,
		"items": [_attendance_draft_data_quality_item(row) for row in matched_rows[page_start : page_start + page_length]],
	}


@frappe.whitelist()
def export_processing_result(company: str, attendance_month: str, source_type: str):
	"""Generate the one current result file from persisted review values."""
	_require_processing_manager()
	company, attendance_month, source_type = _require_company(company), _require_month(attendance_month), _require_processing_source_type(source_type)
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		frappe.throw(_("尚未上传该来源文件。"))
	if not frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": batch.name}):
		frappe.throw(_("该来源尚未完成加工，不能下载加工结果。"))
	processed_result = _export_processed_result(batch)
	_save_batch_notes(batch, {
		"processed_result": processed_result,
		"processed_result_refreshed_on": now_datetime().isoformat(),
		"processed_result_refresh_reason": "manual_download",
	})
	return {"batch": batch.name, "source_type": source_type, "processed_result": processed_result}


@frappe.whitelist()
def get_processing_record(company: str, attendance_month: str, source_type: str, record_id: str):
	_require_processing_manager()
	_require_company(company)
	_require_month(attendance_month)
	_require_processing_source_type(source_type)
	doc = frappe.get_doc(PROCESSING_RECORD_DOCTYPE, record_id)
	if doc.company != company or doc.attendance_month != attendance_month or doc.source_type != source_type:
		frappe.throw(_("无权读取该加工记录。"))
	row = _serialize_record(doc.as_dict())
	editable = sorted(set((row["proposed_value"] or {}).keys()) | set((row["confirmed_value"] or {}).keys()))
	row["editable_fields"] = [{"fieldname": field, "label": PROCESSING_FIELD_LABELS.get(field, "人工调整字段"), "value": (row["confirmed_value"] or row["proposed_value"] or {}).get(field)} for field in editable]
	row["editable_fields"].insert(0, {"fieldname": "__review_decision__", "label": "仅记录处理决定（不改数值）", "value": ""})
	if source_type == "attendance_draft":
		row["daily_rows"] = _daily_row_editor_payload(row)
	return row


@frappe.whitelist()
def update_processing_record(company: str, attendance_month: str, source_type: str, record_id: str, field_name: str, original_value: str = "", new_value: str = "", review_status: str = "待审核", reason: str = ""):
	_require_processing_manager()
	company, attendance_month, source_type = _require_company(company), _require_month(attendance_month), _require_processing_source_type(source_type)
	if review_status not in {"待审核", "已通过", "已驳回"}:
		frappe.throw(_("处理结果无效。"))
	if not (reason or "").strip():
		frappe.throw(_("人工调整必须填写原因。"))
	doc = frappe.get_doc(PROCESSING_RECORD_DOCTYPE, record_id)
	if doc.company != company or doc.attendance_month != attendance_month or doc.source_type != source_type:
		frappe.throw(_("无权修改该加工记录。"))
	proposed = _loads(doc.proposed_value_json, {})
	confirmed = _loads(doc.confirmed_value_json, None) or dict(proposed)
	decision_only = field_name == "__review_decision__"
	if decision_only and review_status == doc.review_status:
		frappe.throw(_("该记录已经完成相同处理；如需更正，请选择具体字段后提交新的调整。"))
	if not decision_only and field_name not in proposed and field_name not in confirmed:
		frappe.throw(_("不能修改未声明的加工字段。"))
	old_confirmed = deepcopy_json(confirmed)
	if not decision_only:
		confirmed[field_name] = _loads(new_value, new_value)
	history = _loads(doc.review_history_json, [])
	history.append({
		"old_value": old_confirmed,
		"new_value": deepcopy_json(confirmed),
		"field_name": field_name,
		"original_value": _loads(original_value, original_value) if not decision_only else deepcopy_json(old_confirmed),
		"reason": reason,
		"review_status": review_status,
		"reviewer": frappe.session.user,
		"reviewed_on": now_datetime().isoformat(),
	})
	doc.confirmed_value_json = _json(confirmed)
	# Keep source-specific display fields (for example Apple-tree date, project
	# and approval columns) while applying the latest reviewed editable value.
	# Earlier code replaced the whole projection with five editable fields.
	processed_value = _loads(doc.processed_value_json, {})
	if not isinstance(processed_value, dict):
		processed_value = {}
	processed_value.update(confirmed if review_status == "已通过" else proposed)
	doc.processed_value_json = _json(processed_value)
	# DingTalk forgot-punch exports normally omit 工号.  If an administrator
	# resolves the employee in review, persist that confirmed identity so the
	# current grid, export and sort key reflect the same business code.
	for identity_field, apple_field in (("employee_code", "工号"), ("employee_name", "姓名"), ("department", "部门")):
		identity_value = _row_value(confirmed, identity_field, apple_field)
		if identity_value not in (None, ""):
			setattr(doc, identity_field, identity_value)
	doc.review_status = review_status
	doc.reviewer = frappe.session.user
	doc.reviewed_on = now_datetime()
	doc.review_note = reason
	doc.review_history_json = _json(history)
	doc.eligible_for_downstream = 1 if review_status == "已通过" and _confirmed_downstream_eligible(confirmed) else 0
	doc.save(ignore_permissions=True)
	batch = frappe.get_doc(IMPORT_BATCH_DOCTYPE, doc.import_batch)
	batch_status = _refresh_batch_review_status(batch)
	# The stored file URL must always reflect the current persistent record, not
	# the original processing-time snapshot. Historical files remain untouched;
	# the batch points only at the newest auditable export.
	processed_result = _export_processed_result(batch)
	_save_batch_notes(batch, {
		"processed_result": processed_result,
		"processed_result_refreshed_on": now_datetime().isoformat(),
		"processed_result_refresh_reason": "manual_review_update",
	})
	result = _serialize_record(doc.as_dict())
	result["batch_status"] = batch_status
	result["processed_result"] = processed_result
	frappe.db.commit()
	return result


@frappe.whitelist()
def update_attendance_draft_daily_row(
	company: str,
	attendance_month: str,
	record_id: str,
	source_row: int | str,
	changes: str | dict,
	review_status: str = "已通过",
	reason: str = "",
):
	"""Correct one DingTalk daily row, then rebuild only that employee's totals.

	The uploaded workbook is immutable.  The approved overlay and both values are
	kept in review history, so the regenerated result and every downstream total
	remain traceable to a specific employee/date/source row.
	"""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	if review_status not in {"待审核", "已通过", "已驳回"}:
		frappe.throw(_("处理结果无效。"))
	if not (reason or "").strip():
		frappe.throw(_("修改钉钉每日数据必须填写原因。"))
	if isinstance(changes, str):
		changes = _loads(changes, {})
	if not isinstance(changes, dict) or not changes:
		frappe.throw(_("请至少修改一个每日考勤字段。"))
	try:
		source_row = int(source_row)
	except (TypeError, ValueError):
		frappe.throw(_("来源行号无效。"))
	doc = frappe.get_doc(PROCESSING_RECORD_DOCTYPE, record_id)
	if doc.company != company or doc.attendance_month != attendance_month or doc.source_type != "attendance_draft":
		frappe.throw(_("无权修改该钉钉考勤记录。"))
	serialized = _serialize_record(doc.as_dict())
	daily_rows = _daily_row_editor_payload(serialized)
	target = next((row for row in daily_rows if row["source_row"] == source_row), None)
	if not target:
		frappe.throw(_("未找到该员工的钉钉来源行。"))
	allowed_fields = {field["fieldname"] for field in target["editable_fields"]}
	invalid_fields = sorted(set(changes) - allowed_fields)
	if invalid_fields:
		frappe.throw(_("不能修改非钉钉每日统计字段：{0}").format("、".join(invalid_fields)))
	old_effective_row = dict(target["source_values"])
	confirmed_before = _loads(doc.confirmed_value_json, None) or {}
	overrides = _daily_row_overrides(confirmed_before)
	overrides[str(source_row)] = {**overrides.get(str(source_row), {}), **changes}

	# Rebuild this employee from its original rows plus the approved overlay. The
	# processor only reads DingTalk's explicit flags/counts; it does not derive
	# missing punches, lateness or night shifts from local schedule rules.
	effective_rows = _effective_daily_source_rows({
		"original_value": serialized.get("original_value"),
		"processed_value": serialized.get("processed_value"),
		"proposed_value": serialized.get("proposed_value"),
		"confirmed_value": {**confirmed_before, "_daily_row_overrides": overrides},
		"eligible_for_downstream": serialized.get("eligible_for_downstream"),
	})
	batch = frappe.get_doc(IMPORT_BATCH_DOCTYPE, doc.import_batch)
	rebuilt = process_attendance_draft_rows(
		effective_rows,
		attendance_month=attendance_month,
		source_file=doc.source_file or batch.source_file,
		source_sheet=doc.source_sheet or "每日统计",
		employee_directory=_employee_directory(company) or None,
		exception_policy=_attendance_draft_exception_policy(),
	)
	replacement = next((row for row in rebuilt["processed_rows"] if str(row.get("employee_code") or "") == str(doc.employee_code or "")), None)
	if not replacement:
		frappe.throw(_("更正后无法重新生成该员工的考勤汇总。"))
	confirmed = dict(replacement["proposed_value"])
	confirmed["_daily_row_overrides"] = overrides
	history = _loads(doc.review_history_json, [])
	history.append({
		"old_value": old_effective_row,
		"new_value": {**old_effective_row, **changes},
		"field_name": f"__daily_row__:{source_row}",
		"original_value": old_effective_row,
		"reason": (reason or "").strip(),
		"review_status": review_status,
		"reviewer": frappe.session.user,
		"reviewed_on": now_datetime().isoformat(),
		"source_row": source_row,
	})
	doc.processed_value_json = _json(confirmed)
	doc.confirmed_value_json = _json(confirmed)
	doc.exception_codes = _json(replacement["exception_codes"])
	doc.exception_message = replacement["exception_message"]
	doc.review_status = review_status
	doc.reviewer = frappe.session.user
	doc.reviewed_on = now_datetime()
	doc.review_note = (reason or "").strip()
	doc.review_history_json = _json(history)
	doc.eligible_for_downstream = 1 if review_status == "已通过" and _confirmed_downstream_eligible(confirmed) else 0
	doc.save(ignore_permissions=True)
	batch_status = _refresh_batch_review_status(batch)
	processed_result = _export_processed_result(batch)
	_save_batch_notes(batch, {
		"processed_result": processed_result,
		"processed_result_refreshed_on": now_datetime().isoformat(),
		"processed_result_refresh_reason": "daily_source_row_manual_update",
	})
	result = _serialize_record(doc.as_dict())
	result["batch_status"] = batch_status
	result["processed_result"] = processed_result
	frappe.db.commit()
	return result


@frappe.whitelist()
def bulk_update_processing_records(
	company: str,
	attendance_month: str,
	source_type: str,
	record_ids: str | list[str],
	select_all_pending: int = 0,
	page_start: int = 0,
	page_length: int = 20,
	review_status: str = "待审核",
	reason: str = "",
):
	"""Apply one reviewed decision to explicitly selected exception records.

	This is deliberately a *decision-only* batch action: it does not invent or
	overwrite attendance figures.  Each selected record receives its own audit
	history entry, exactly as it would through the single-record review dialog.
	"""
	_require_processing_manager()
	company, attendance_month, source_type = _require_company(company), _require_month(attendance_month), _require_processing_source_type(source_type)
	if review_status not in {"待审核", "已通过", "已驳回"}:
		frappe.throw(_("处理结果无效。"))
	if not (reason or "").strip():
		frappe.throw(_("批量处理必须填写原因。"))
	select_all_pending = cint(select_all_pending)
	if isinstance(record_ids, str):
		record_ids = _loads(record_ids, [])
	if not isinstance(record_ids, list):
		frappe.throw(_("请选择要批量处理的记录。"))
	record_ids = list(dict.fromkeys(str(record_id).strip() for record_id in record_ids if str(record_id).strip()))
	if not select_all_pending and not record_ids:
		frappe.throw(_("请选择至少一条异常记录。"))
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		frappe.throw(_("尚未上传该来源文件。"))
	if select_all_pending:
		# A source-filtered all-selection is resolved on the server at submit time,
		# so every pending record is included even when the browser only displays
		# one page.  The source boundary remains mandatory: different source types
		# can carry different review semantics.
		record_ids = [row.name for row in frappe.get_all(
			PROCESSING_RECORD_DOCTYPE,
			filters={"import_batch": batch.name, "exception_codes": ["!=", "[]"], "review_status": "待审核"},
			fields=["name"],
			order_by="modified desc",
			limit_page_length=501,
		)]
		if not record_ids:
			frappe.throw(_("当前筛选来源没有待处理异常，请刷新页面后重试。"))
	elif len(record_ids) > 500:
		frappe.throw(_("一次最多批量处理 500 条记录。"))
	if len(record_ids) > 500:
		frappe.throw(_("当前筛选待处理异常超过 500 条，请分批处理。"))
	rows = frappe.get_all(
		PROCESSING_RECORD_DOCTYPE,
		filters={"name": ["in", record_ids]},
		fields=["name", "import_batch", "exception_codes", "review_status"],
		limit_page_length=len(record_ids),
	)
	if len(rows) != len(record_ids) or any(row.import_batch != batch.name for row in rows):
		frappe.throw(_("所选记录不属于当前来源的最新加工版本，请刷新后重试。"))
	if any(not _loads(row.exception_codes, []) for row in rows):
		frappe.throw(_("批量处理仅适用于异常记录；正常记录无需审核。"))
	if any(row.review_status != "待审核" for row in rows):
		frappe.throw(_("所选记录已经处理。若需更正，请逐条使用“查看/更正记录”。"))

	processed_at = now_datetime()
	for record_id in record_ids:
		doc = frappe.get_doc(PROCESSING_RECORD_DOCTYPE, record_id)
		proposed = _loads(doc.proposed_value_json, {})
		confirmed = _loads(doc.confirmed_value_json, None) or dict(proposed)
		old_confirmed = deepcopy_json(confirmed)
		history = _loads(doc.review_history_json, [])
		history.append({
			"old_value": old_confirmed,
			"new_value": deepcopy_json(confirmed),
			"field_name": "__review_decision__",
			"original_value": deepcopy_json(old_confirmed),
			"reason": reason.strip(),
			"review_status": review_status,
			"reviewer": frappe.session.user,
			"reviewed_on": processed_at.isoformat(),
		})
		doc.confirmed_value_json = _json(confirmed)
		processed_value = _loads(doc.processed_value_json, {})
		if not isinstance(processed_value, dict):
			processed_value = {}
		processed_value.update(confirmed if review_status == "已通过" else proposed)
		doc.processed_value_json = _json(processed_value)
		doc.review_status = review_status
		doc.reviewer = frappe.session.user
		doc.reviewed_on = processed_at
		doc.review_note = reason.strip()
		doc.review_history_json = _json(history)
		doc.eligible_for_downstream = 1 if review_status == "已通过" and _confirmed_downstream_eligible(confirmed) else 0
		doc.save(ignore_permissions=True)

	batch_status = _refresh_batch_review_status(batch)
	processed_result = _export_processed_result(batch)
	_save_batch_notes(batch, {
		"processed_result": processed_result,
		"processed_result_refreshed_on": processed_at.isoformat(),
		"processed_result_refresh_reason": "bulk_manual_review_update",
	})
	all_rows = frappe.get_all(
		PROCESSING_RECORD_DOCTYPE,
		filters={"import_batch": batch.name},
		fields=["review_status", "eligible_for_downstream"],
		limit_page_length=0,
	)
	return {
		"batch": batch.name,
		"source_type": source_type,
		"updated_rows": len(record_ids),
		"review_status": review_status,
		"batch_status": batch_status,
		"included_rows": sum(1 for row in all_rows if cint(row.eligible_for_downstream)),
		"rejected_rows": sum(1 for row in all_rows if row.review_status == "已驳回"),
		"processed_result": processed_result,
	}


def _signed_final_header_text(sheet, row: int, column: int) -> str:
	return re.sub(r"\s+", "", str(sheet.cell(row, column).value or ""))


def _signed_final_column(sheet, label: str, *, header_row: int, max_column: int = 64) -> int:
	needle = re.sub(r"\s+", "", label)
	for column in range(1, min(sheet.max_column, max_column) + 1):
		if needle in _signed_final_header_text(sheet, header_row, column):
			return column
	frappe.throw(_("已签考勤终稿缺少字段：{0}").format(label))


def _signed_final_rows(file_url: str) -> list[dict[str, Any]]:
	"""Read the supplied HR sign-off form without rewriting the workbook."""
	workbook = _load_workbook(file_url)
	sheet = next((item for item in workbook.worksheets if "考勤终稿" in str(item.title or "")), workbook.worksheets[0])
	header_values = {
		row_number: next(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True), ())
		for row_number in (2, 3)
	}

	def header_column(field_key: str, label: str, row_number: int, max_column: int = 64) -> int:
		# A configured label has priority over the historical wording.  This keeps
		# the sign-off import stable when HR changes the visible Excel heading.
		candidates = _attendance_final_header_candidates(field_key, label, row_number)
		for candidate in candidates:
			needle = re.sub(r"\s+", "", candidate)
			for column, value in enumerate(header_values[row_number][:max_column], start=1):
				if needle and needle in re.sub(r"\s+", "", str(value or "")):
					return column
		frappe.throw(_("已签考勤终稿缺少字段：{0}").format(label))

	columns = {
		"department": header_column("department", "部门", 2),
		"employee_code": header_column("employee_code", "工号", 2),
		"employee_name": header_column("employee_name", "姓名", 2),
		"standard_hours": header_column("standard_hours", "标准工时", 2),
		"actual_attendance_hours": header_column("actual_attendance_hours", "钉钉导出实际出勤", 2),
		"workday_overtime_hours": header_column("workday_overtime_hours", "工作日加班", 3, 25),
		"restday_overtime_hours": header_column("restday_overtime_hours", "休息日加班", 3, 25),
		"holiday_overtime_hours": header_column("holiday_overtime_hours", "节假日加班", 3, 25),
		"personal_leave_hours": header_column("personal_leave_hours", "事假", 3, 25),
		"sick_leave_hours": header_column("sick_leave_hours", "病假", 3, 25),
		"annual_leave_hours": header_column("annual_leave_hours", "特休", 3, 25),
		"work_injury_hours": header_column("work_injury_hours", "工伤", 3, 25),
		"rest_arrangement_hours": header_column("rest_arrangement_hours", "排休", 3, 25),
		"absence_hours": header_column("absence_hours", "旷工", 3, 25),
		"large_night_shifts": header_column("large_night_shifts", "大夜班", 2),
		"small_night_shifts": header_column("small_night_shifts", "小夜班", 2),
		# These are HR-reviewed money values. Apple source rows contain counts,
		# so treating those counts as payroll amounts produces the wrong salary.
		"green_apple_amount": header_column("green_apple_amount", "绿苹果", 2),
		"red_apple_amount": header_column("red_apple_amount", "红苹果", 2),
		"housing_allowance": header_column("housing_allowance", "住房", 2),
		"full_attendance_award": header_column("full_attendance_award", "全勤", 2),
	}
	rows = []
	for source_row, source_values in enumerate(sheet.iter_rows(min_row=5, values_only=True), start=5):
		code = str(source_values[columns["employee_code"] - 1] or "").strip()
		name = str(source_values[columns["employee_name"] - 1] or "").strip()
		if not code and not name:
			continue
		if not code or not name:
			continue
		row = {"source_row": source_row}
		for fieldname, column in columns.items():
			value = source_values[column - 1] if column <= len(source_values) else None
			if fieldname in {"department", "employee_code", "employee_name"}:
				row[fieldname] = str(value or "").strip()
			else:
				row[fieldname] = _as_number(value)
		rows.append(row)
	return rows


@frappe.whitelist()
def preview_signed_attendance_final(company: str, attendance_month: str, file_url: str):
	"""Return an HR-readable reconciliation preview without changing records."""
	_require_processing_manager()
	_require_company(company), _require_month(attendance_month)
	rows = _signed_final_rows(file_url)
	amount_fields = ("green_apple_amount", "red_apple_amount", "housing_allowance", "full_attendance_award")
	return {
		"employee_rows": len(rows),
		"amount_totals": {field: round(sum(_as_number(row.get(field)) for row in rows), 2) for field in amount_fields},
		"samples": [
			{
				"employee_code": row.get("employee_code"),
				"employee_name": row.get("employee_name"),
				**{field: row.get(field) for field in amount_fields},
			}
			for row in rows[:5]
		],
	}


@frappe.whitelist()
def reconcile_attendance_draft_with_signed_final(company: str, attendance_month: str, file_url: str):
	"""Use HR's signed final as an audited correction layer over the draft.

	The raw DingTalk draft remains immutable.  Only the employee-level processed
	projection is corrected, with a review-history event that cites the supplied
	sign-off workbook.  Employees absent from the signed final are explicitly
	excluded instead of silently leaking into payroll.
	"""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	if not file_url or not file_url.lower().split("?", 1)[0].endswith(".xlsx"):
		frappe.throw(_("请上传 .xlsx 格式的已签考勤终稿。"))
	batch = _latest_batch(company, attendance_month, "attendance_draft")
	if not batch or not frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": batch.name}):
		frappe.throw(_("请先导入并加工本月考勤初稿。"))
	final_rows = _signed_final_rows(file_url)
	if not final_rows:
		frappe.throw(_("已签考勤终稿没有可识别的员工记录。"))
	final_by_code = {}
	for row in final_rows:
		code = row["employee_code"]
		if code in final_by_code:
			frappe.throw(_("已签考勤终稿存在重复工号：{0}").format(code))
		final_by_code[code] = row
	records = frappe.get_all(
		PROCESSING_RECORD_DOCTYPE,
		filters={"import_batch": batch.name},
		fields=["name", "employee_code", "employee_name", "proposed_value_json", "confirmed_value_json", "processed_value_json", "review_status", "eligible_for_downstream", "review_history_json"],
		limit_page_length=5000,
	)
	processed_at = now_datetime()
	matched_codes = set()
	changed_records = 0
	attendance_changed_records = 0
	amount_changed_records = 0
	excluded_records = 0
	approved_exceptions = 0
	fields = tuple(
		field for field, _label in ATTENDANCE_DRAFT_RESULT_COLUMNS
		if field not in {"employee_code", "employee_name", "department", "attendance_details", "source_row_count", "night_shift_matching", "clock_in_missing_count", "clock_out_missing_count"}
	) + ("green_apple_amount", "red_apple_amount", "housing_allowance", "full_attendance_award")
	amount_fields = {"green_apple_amount", "red_apple_amount", "housing_allowance", "full_attendance_award"}
	attendance_fields = set(fields) - amount_fields
	for row in records:
		doc = frappe.get_doc(PROCESSING_RECORD_DOCTYPE, row.name)
		final = final_by_code.get(str(row.employee_code or "").strip())
		proposed = _loads(doc.proposed_value_json, {})
		confirmed = _loads(doc.confirmed_value_json, None) or dict(proposed)
		old_confirmed = deepcopy_json(confirmed)
		previous_status = doc.review_status
		if final:
			matched_codes.add(final["employee_code"])
			for fieldname in fields:
				if fieldname in final:
					confirmed[fieldname] = final[fieldname]
			confirmed.update({"employee_code": final["employee_code"], "employee_name": final["employee_name"], "department": final["department"]})
			confirmed["signed_final_override"] = 1
			doc.review_status = "已通过" if previous_status == "待审核" or confirmed != proposed else previous_status
			doc.eligible_for_downstream = 1
			if previous_status == "待审核":
				approved_exceptions += 1
		else:
			doc.review_status = "已驳回"
			doc.eligible_for_downstream = 0
			excluded_records += 1
		reason = (
			_("已与人资签字考勤终稿逐项核对并采用终稿数值；来源：{0}").format(Path(file_url).name)
			if final else _("该员工不在本月人资签字考勤终稿中，明确不计入本月终稿；来源：{0}").format(Path(file_url).name)
		)
		if final:
			attendance_changed = any(_as_number(old_confirmed.get(field)) != _as_number(final.get(field)) for field in attendance_fields)
			amount_changed = any(_as_number(old_confirmed.get(field)) != _as_number(final.get(field)) for field in amount_fields)
			if attendance_changed or amount_changed:
				changed_records += 1
			attendance_changed_records += int(attendance_changed)
			amount_changed_records += int(amount_changed)
		history = _loads(doc.review_history_json, [])
		history.append({
			"old_value": old_confirmed,
			"new_value": deepcopy_json(confirmed) if final else None,
			"field_name": "__signed_attendance_final__",
			"original_value": old_confirmed,
			"reason": reason,
			"review_status": doc.review_status,
			"reviewer": frappe.session.user,
			"reviewed_on": processed_at.isoformat(),
			"source_file": file_url,
		})
		doc.confirmed_value_json = _json(confirmed) if final else ""
		processed = _loads(doc.processed_value_json, {})
		if final:
			processed.update(confirmed)
		doc.processed_value_json = _json(processed)
		doc.reviewer = frappe.session.user
		doc.reviewed_on = processed_at
		doc.review_note = reason
		doc.review_history_json = _json(history)
		doc.save(ignore_permissions=True)
	unmatched_final = sorted(set(final_by_code) - matched_codes)
	if unmatched_final:
		frappe.throw(_("已签考勤终稿中有 {0} 个工号不在考勤初稿：{1}").format(len(unmatched_final), "、".join(unmatched_final[:20])))
	_refresh_batch_review_status(batch)
	processed_result = _export_processed_result(batch)
	reconciliation = {
		"source_file": file_url,
		"source_file_name": Path(file_url).name,
		"source_checksum": _file_checksum(file_url),
		"employee_rows": len(final_rows),
		"matched_records": len(matched_codes),
		"changed_records": changed_records,
		"attendance_changed_records": attendance_changed_records,
		"amount_changed_records": amount_changed_records,
		"approved_exception_records": approved_exceptions,
		"excluded_records": excluded_records,
		"reconciled_by": frappe.session.user,
		"reconciled_on": processed_at.isoformat(),
	}
	_save_batch_notes(batch, {"signed_final_reconciliation": reconciliation, "processed_result": processed_result})
	# The reconciliation is an explicit HR confirmation action spanning many
	# records.  Commit it as one auditable unit so it also persists when invoked
	# from the local acceptance runner (web requests would otherwise be the only
	# place where Frappe commits automatically).
	frappe.db.commit()
	return {"batch": batch.name, "status": batch.status, "reconciliation": reconciliation, "processed_result": processed_result}


def deepcopy_json(value):
	return _loads(_json(value), value)


@frappe.whitelist()
def confirm_source_result(company: str, attendance_month: str, source_type: str):
	_require_processing_manager()
	company, attendance_month, source_type = _require_company(company), _require_month(attendance_month), _require_source_type(source_type)
	batch = _latest_batch(company, attendance_month, source_type)
	if not batch:
		frappe.throw(_("尚未上传该来源文件。"))
	if batch.status == "已确认":
		frappe.throw(_("该来源已经确认；如需更正，请使用“查看/更正记录”。"))
	rows = frappe.get_all(PROCESSING_RECORD_DOCTYPE, filters={"import_batch": batch.name}, fields=["name", "review_status", "eligible_for_downstream"], limit_page_length=0)
	if not rows:
		frappe.throw(_("该来源尚未生成任何加工记录，不能确认。请先检查来源文件和结构预检结果。"))
	# Pending records are deliberately excluded from downstream calculations,
	# not used to block every matching employee in this source from proceeding.
	pending_review_rows = [row for row in rows if row.review_status == "待审核"]
	batch.status = "已确认"
	batch.save(ignore_permissions=True)
	_save_batch_notes(batch, {
		"confirmed_with_pending_reviews": bool(pending_review_rows),
		"pending_review_rows": len(pending_review_rows),
		"confirmation_note": "待审核记录不会阻塞来源确认，且不计入下游计算。",
	})
	frappe.db.commit()
	return {
		"batch": batch.name,
		"status": batch.status,
		"confirmed_rows": len(rows),
		"included_rows": sum(1 for row in rows if cint(row.eligible_for_downstream)),
		"pending_review_rows": len(pending_review_rows),
		"confirmed_with_pending_reviews": bool(pending_review_rows),
		"rejected_rows": sum(1 for row in rows if row.review_status == "已驳回"),
	}


@frappe.whitelist()
def list_processing_exceptions(company: str, attendance_month: str, source_type: str = "", page_length: int = 20, page_start: int = 0):
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	if source_type:
		_require_processing_source_type(source_type)
	page_length = min(max(cint(page_length), 1), 100)
	page_start = max(cint(page_start), 0)
	batches = [_latest_batch(company, attendance_month, source) for source in SOURCE_TYPES + MONTHLY_SUPPORT_SOURCE_TYPES]
	all_batch_names = [batch.name for batch in batches if batch]
	batch_names = [batch.name for batch in batches if batch and (not source_type or batch.source_type == source_type)]
	# Keep the count and page query identical; otherwise the page indicator could
	# promise records that cannot appear in the exception work queue.
	total_pending_count = frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": ["in", all_batch_names], "exception_codes": ["!=", "[]"], "review_status": "待审核"}) if all_batch_names else 0
	filtered_pending_count = frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": ["in", batch_names], "exception_codes": ["!=", "[]"], "review_status": "待审核"}) if batch_names else 0
	if not batch_names:
		return {"review_rows": [], "total_pending_count": total_pending_count, "filtered_pending_count": 0, "source_type": source_type, "page_start": page_start, "page_length": page_length}
	records = frappe.get_all(
		PROCESSING_RECORD_DOCTYPE,
		# The work queue contains only unresolved records.  Resolved exceptions
		# remain in the source result and manual-adjustment ledger for traceability.
		filters={"import_batch": ["in", batch_names], "exception_codes": ["!=", "[]"], "review_status": "待审核"},
		fields=["name", "attendance_month", "employee_code", "employee_name", "department", "source_type", "original_value_json", "exception_codes", "exception_message", "review_status", "proposed_value_json", "confirmed_value_json", "reviewer", "reviewed_on", "review_note", "source_file", "source_sheet", "source_row", "source_id", "approval_no"],
		order_by="modified desc",
		limit_start=page_start,
		limit_page_length=page_length,
	)
	return {"review_rows": [_serialize_record(row) for row in records], "total_pending_count": total_pending_count, "filtered_pending_count": filtered_pending_count, "source_type": source_type, "page_start": page_start, "page_length": page_length}


@frappe.whitelist()
def list_processing_batches(company: str, attendance_month: str, page_length: int = 100):
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	latest = {}
	for source_type in SOURCE_TYPES:
		batch = _latest_batch(company, attendance_month, source_type)
		if batch:
			latest[source_type] = batch
	if not latest:
		return {"items": []}
	created_at = min((batch.creation for batch in latest.values()), default="")
	return {
		"items": [{
			"batch_id": f"{company}:{attendance_month}",
			"attendance_month": attendance_month,
			"attendance_draft_status": latest.get("attendance_draft").status if latest.get("attendance_draft") else "未上传",
			"apple_tree_status": latest.get("apple_tree").status if latest.get("apple_tree") else "未上传",
			"missing_card_status": latest.get("missing_card").status if latest.get("missing_card") else "未上传",
			"created_at": created_at,
		}][: min(max(cint(page_length), 1), 200)]
	}


@frappe.whitelist()
def list_daily_attendance_records(company: str, attendance_month: str, attendance_date: str = "", page_length: int = 500):
	"""Show the raw DingTalk daily-detail rows behind the monthly draft source.

	This is intentionally a view of the *same* attendance-draft upload, rather
	than a second competing daily dataset.  HR can inspect any day immediately
	after uploading, then process the same source into the monthly draft.
	"""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	if attendance_date and not re.fullmatch(r"\d{4}-\d{2}-\d{2}", attendance_date):
		frappe.throw(_("查看日期格式应为 YYYY-MM-DD。"))
	batch = _latest_batch(company, attendance_month, "attendance_draft")
	if not batch:
		return {"items": [], "available_dates": [], "source_file_name": "", "notice": _("请先导入该月钉钉日考勤明细。")}
	workbook = _load_workbook(batch.source_file)
	sheet = _attendance_draft_sheet(workbook)
	if not sheet:
		return {"items": [], "available_dates": [], "source_file_name": batch.source_file.rsplit("/", 1)[-1], "notice": _("考勤初稿中未找到可识别的钉钉每日数据表。")}
	rows = rows_from_dingtalk_daily_sheet(sheet, source_file=batch.source_file)
	available_dates = sorted({value for row in rows if (value := _daily_attendance_date(row.get("日期"))).startswith(attendance_month)})
	items = []
	excluded_missing_employee_code_rows = 0
	for row in rows:
		day = _daily_attendance_date(row.get("日期"))
		if not day.startswith(attendance_month) or (attendance_date and day != attendance_date):
			continue
		if not str(row.get("工号") or "").strip():
			excluded_missing_employee_code_rows += 1
			continue
		items.append({
			"attendance_date": day,
			"employee_name": row.get("姓名") or "",
			"employee_code": row.get("工号") or "",
			"department": _display_department(row.get("实际部门") or row.get("部门")),
			"date_type": row.get("日期类型") or "",
			"shift": row.get("班次") or "",
			"clock_in": row.get("上班时间") or "",
			"clock_out": row.get("下班时间") or "",
			"standard_hours": row.get("标准工时") or 0,
			"actual_attendance_hours": row.get("实际出勤（小时）") or 0,
			"workday_overtime_hours": row.get("工作日加班（小时）") or 0,
			"restday_overtime_hours": row.get("休息日加班（小时）") or 0,
			"holiday_overtime_hours": row.get("节假日加班（小时）") or 0,
			"clock_in_missing": row.get("上班缺卡") or row.get("上班未打卡次数") or 0,
			"clock_out_missing": row.get("下班缺卡") or row.get("下班未打卡次数") or 0,
		})
	items.sort(key=lambda item: (item["attendance_date"], str(item["department"]), str(item["employee_code"])))
	limit = min(max(cint(page_length), 1), 5000)
	return {
		"items": items[:limit],
		"total_count": len(items),
		"excluded_missing_employee_code_rows": excluded_missing_employee_code_rows,
		"available_dates": available_dates,
		"source_file_name": batch.source_file.rsplit("/", 1)[-1],
		"source_file_url": batch.source_file,
		"batch_status": batch.status,
		"notice": "",
	}


@frappe.whitelist()
def reset_attendance_month(company: str, attendance_month: str, confirm_month: str):
	"""Delete all processing state for one company/month after typed confirmation.

	The uploaded File records are deliberately not deleted: they remain the
	auditable originals, while batches, derived records, reviews and final-output
	references are removed so the month can be imported afresh.
	"""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	if confirm_month != attendance_month:
		frappe.throw(_("请输入当前处理月份 {0} 以确认清空。").format(attendance_month))
	batch_count = frappe.db.count(IMPORT_BATCH_DOCTYPE, {"company": company, "attendance_month": attendance_month})
	record_count = frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"company": company, "attendance_month": attendance_month})
	# Processing records contain review history, exception state and references
	# to final outputs, so they must be removed before their parent batches.
	frappe.db.delete(PROCESSING_RECORD_DOCTYPE, {"company": company, "attendance_month": attendance_month})
	frappe.db.delete(IMPORT_BATCH_DOCTYPE, {"company": company, "attendance_month": attendance_month})
	return {
		"attendance_month": attendance_month,
		"deleted_batch_count": batch_count,
		"deleted_record_count": record_count,
		"notice": _("已清空 {0} 的考勤处理数据；原始上传文件未删除，可重新导入。").format(attendance_month),
	}


@frappe.whitelist()
def list_manual_adjustments(company: str, attendance_month: str, page_length: int = 500):
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	records = frappe.get_all(
		PROCESSING_RECORD_DOCTYPE,
		filters={"company": company, "attendance_month": attendance_month, "review_history_json": ["!=", "[]"]},
		fields=["name", "employee_code", "employee_name", "source_type", "review_history_json"],
		order_by="modified desc",
		limit_page_length=min(max(cint(page_length), 1), 5000),
	)
	items = []
	for record in records:
		for event in _loads(record.review_history_json, []):
			items.append({"record_id": record.name, "employee_code": record.employee_code, "employee_name": record.employee_name, "source_type": record.source_type, "field_name": event.get("field_name", ""), "original_value": event.get("old_value"), "new_value": event.get("new_value"), "reason": event.get("reason", ""), "modified_by": event.get("reviewer", ""), "modified_at": event.get("reviewed_on", "")})
	return {"items": items[: min(max(cint(page_length), 1), 5000)]}


@frappe.whitelist()
def get_processing_configuration(company: str, configuration_type: str):
	_require_processing_manager()
	_require_company(company)
	if configuration_type == "field-mapping":
		return {"items": [{"name": "考勤初稿", "source_value": "钉钉每日明细明确数值列", "target_value": "标准工时、实际出勤、加班、夜班、请假、排休、旷工", "status": "已发布"}]}
	if configuration_type == "department-mapping":
		return {"items": [{"name": "员工匹配", "source_value": "工号为主键；姓名与部门辅助", "target_value": "冲突进入统一待审核", "status": "已发布"}]}
	return {"items": [{"name": "自动化边界", "source_value": "特殊工时与样例人工差异", "target_value": "不猜测；进入其他来源或统一审核", "status": "已发布"}]}


@frappe.whitelist()
def list_department_mappings(company: str, source_type: str = "missing_card"):
	"""List company-scoped, auditable aliases for historical DingTalk departments."""
	_require_processing_manager()
	company = _require_company(company)
	if source_type != "missing_card":
		frappe.throw(_("当前仅支持维护“忘打卡”来源的部门映射。"))
	items = frappe.get_all(
		DEPARTMENT_MAPPING_DOCTYPE,
		filters={"company": company, "source_type": source_type},
		fields=["name", "source_department", "target_department", "enabled", "modified", "modified_by"],
		order_by="modified desc",
		limit_page_length=500,
	)
	return {
		"items": [{
			"name": item.name,
			"source_department": item.source_department,
			"target_department": item.target_department,
			"enabled": cint(item.enabled),
			"status": "已启用" if cint(item.enabled) else "已停用",
			"modified": item.modified,
			"modified_by": item.modified_by,
		} for item in items],
	}


@frappe.whitelist()
def upsert_department_mapping(
	company: str,
	source_department: str,
	target_department: str,
	enabled: int = 1,
	source_type: str = "missing_card",
):
	"""Save a reviewed alias without rewriting existing batch history."""
	_require_processing_manager()
	company = _require_company(company)
	if source_type != "missing_card":
		frappe.throw(_("当前仅支持维护“忘打卡”来源的部门映射。"))
	source_department, target_department = (source_department or "").strip(), (target_department or "").strip()
	if not source_department or not target_department:
		frappe.throw(_("钉钉原部门和花名册目标部门均不能为空。"))
	if not frappe.db.exists("Department", target_department):
		frappe.throw(_("花名册目标部门不存在，请从有效部门中选择。"))
	existing = frappe.db.get_value(
		DEPARTMENT_MAPPING_DOCTYPE,
		{"company": company, "source_type": source_type, "source_department": source_department},
		"name",
	)
	if existing:
		doc = frappe.get_doc(DEPARTMENT_MAPPING_DOCTYPE, existing)
		doc.target_department = target_department
		doc.enabled = cint(enabled)
		doc.save(ignore_permissions=True)
	else:
		doc = frappe.get_doc({
			"doctype": DEPARTMENT_MAPPING_DOCTYPE,
			"company": company,
			"source_type": source_type,
			"source_department": source_department,
			"target_department": target_department,
			"enabled": cint(enabled),
		}).insert(ignore_permissions=True)
	return {
		"item": {
			"name": doc.name,
			"source_department": doc.source_department,
			"target_department": doc.target_department,
			"enabled": cint(doc.enabled),
			"status": "已启用" if cint(doc.enabled) else "已停用",
		},
		"notice": _("映射将用于后续重新上传并加工的忘打卡批次；已生成批次保留原始记录与审核历史。"),
	}


def _finalization_inputs(company, attendance_month, slots):
	by_source = {slot["source_type"]: slot for slot in slots}
	inputs = []
	for source_type in SOURCE_TYPES:
		slot = by_source.get(source_type)
		# Confirmation is the final decision for a main source.  Rows that remain
		# in the audit queue after confirmation have already been excluded from
		# downstream calculation, so they must not reopen the month-final gate.
		# Otherwise a card can truthfully display “已确认” while the lock button
		# remains disabled forever because of its historic pending-row count.
		ready = bool(slot and slot["status"] == "已确认")
		inputs.append({"key": source_type, "source_type": source_type, "label": SOURCE_LABELS[source_type], "status": "已就绪" if ready else slot["status"] if slot else "未就绪", "ready": ready, "snapshot_version": ""})
	for source_type in MONTHLY_SUPPORT_SOURCE_TYPES:
		batch = _latest_batch(company, attendance_month, source_type)
		meta = _processing_meta(batch) if batch else {}
		precheck = meta.get("monthly_support_precheck") or {}
		processed_rows = frappe.db.count(PROCESSING_RECORD_DOCTYPE, {"import_batch": batch.name}) if batch else 0
		needs_processing = bool(batch and batch.status == "已确认" and not processed_rows)
		inputs.append({
			"key": source_type,
			"source_type": source_type,
			"label": SOURCE_LABELS[source_type],
			"kind": "月度补充来源",
			"status": "需补做导入校验" if needs_processing else "已就绪" if batch and batch.status == "已确认" else batch.status if batch else "未就绪",
			"ready": bool(batch and batch.status == "已确认" and processed_rows and not needs_processing),
			"snapshot_version": "",
			"source_file": batch.source_file if batch else "",
			"source_file_name": Path(batch.source_file).name if batch and batch.source_file else "",
			"record_count": cint(precheck.get("record_count")),
			"processed_rows": processed_rows,
			# Kept for compatibility with existing clients.  Import validation errors
			# are shown on the source result page, never in the exception queue.
			"pending_exception_count": 0,
			"import_error_count": sum(1 for row in _result_rows(batch, 5000) if row.get("exception_codes")) if batch else 0,
			"can_precheck": bool(batch and batch.source_file and (batch.status in {"待加工", "预览", "已导入"} or needs_processing)),
			"can_process": bool(batch and precheck.get("is_valid") and not processed_rows and batch.status == "待加工"),
			"can_confirm": False,
			"description": MONTHLY_SUPPORT_SOURCE_CONFIG[source_type]["description"],
		})
	return inputs


FINAL_SIGNED_COLUMNS = (
	("employee_code", "工号"), ("employee_name", "姓名"), ("department", "部门"),
	("standard_hours", "标准工时"), ("actual_attendance_hours", "实际出勤"),
	("special_workday_hours", "平日特殊工时"), ("workday_overtime_hours", "工作日加班"),
	("special_restday_hours", "周末特殊工时"), ("restday_overtime_hours", "休息日加班"),
	("special_holiday_hours", "节假日特殊工时"), ("holiday_overtime_hours", "节假日加班"),
	("large_night_shifts", "大夜班"), ("small_night_shifts", "小夜班"),
	("personal_leave_hours", "事假"), ("sick_leave_hours", "病假"), ("annual_leave_hours", "特休"),
	("work_injury_hours", "工伤"), ("rest_arrangement_hours", "排休"), ("absence_hours", "旷工"),
	("clock_in_missing_count", "上班漏打卡"), ("clock_out_missing_count", "下班漏打卡"),
	("green_apple_amount", "绿苹果金额"), ("red_apple_amount", "红苹果金额"),
	("housing_allowance", "住房补贴"), ("full_attendance_award", "全勤奖"),
	("employee_signature", "员工签字"), ("review_note", "备注"),
)
FINAL_FINANCE_COLUMNS = (
	("employee_code", "工号"), ("employee_name", "姓名"), ("department", "部门"),
	("actual_attendance_hours", "实际出勤"), ("workday_overtime_hours", "工作日加班（含特殊工时）"),
	("restday_overtime_hours", "休息日加班（含特殊工时）"), ("holiday_overtime_hours", "节假日加班（含特殊工时）"),
	("large_night_shifts", "大夜班"), ("small_night_shifts", "小夜班"),
	("absence_hours", "旷工"), ("green_apple_amount", "绿苹果金额"), ("red_apple_amount", "红苹果金额"), ("housing_allowance", "住房补贴"),
	("full_attendance_award", "全勤奖"),
)

# The settings table stores visible Excel headings, while field keys remain
# stable implementation identifiers.  A user may therefore rename or regroup
# the workbook without changing the calculation chain or import contract.
SIGNED_FINAL_FIELD_LAYOUT = (
	("sequence", "序号", "", "人工填写"),
	("department", "部门", "", "来源字段"),
	("employee_code", "工号", "", "来源字段"),
	("employee_name", "姓名", "", "来源字段"),
	("date_of_joining", "入职时间", "", "来源字段"),
	("standard_hours", "标准工时\n（小时）", "", "来源字段"),
	("actual_attendance_hours", "钉钉导出\n实际出勤\n（小时）", "", "来源字段"),
	("actual_checked", "实际打卡出勤A(验算）", "", "计算字段"),
	("special_workday_hours", "1.5倍加班工时C", "平特", "来源字段"),
	("workday_overtime_hours", "1.5倍加班工时C", "工作日加班（小时）", "来源字段"),
	("special_restday_hours", "2倍加班工时D", "周特", "来源字段"),
	("restday_overtime_hours", "2倍加班工时D", "休息日加班(小时)", "来源字段"),
	("special_holiday_hours", "3倍加班工时", "节假日特", "来源字段"),
	("holiday_overtime_hours", "3倍加班工时", "节假日加班(小时)", "来源字段"),
	("personal_leave_hours", "请假钉钉条", "事假(小时)", "来源字段"),
	("sick_leave_hours", "请假钉钉条", "病假(小时)", "来源字段"),
	("annual_leave_hours", "请假钉钉条", "特休(小时)", "来源字段"),
	("work_injury_hours", "请假钉钉条", "工伤(小时)", "来源字段"),
	("rest_arrangement_hours", "请假钉钉条", "排休(小时)", "来源字段"),
	("absence_hours", "请假钉钉条", "旷工(小时)", "来源字段"),
	("bereavement_leave_hours", "请假钉钉条", "丧假\n(小时)", "来源字段"),
	("marriage_leave_half_days", "请假钉钉条", "婚假\n(半天)", "来源字段"),
	("public_leave_half_days", "请假钉钉条", "公假\n(半天)", "人工填写"),
	("maternity_leave_days", "请假钉钉条", "产假（天）", "人工填写"),
	("sick_leave_credit", "应补1倍工时B", "病假补工时50%", "计算字段"),
	("annual_leave_credit", "应补1倍工时B", "特休\n(小时)", "计算字段"),
	("work_injury_credit", "应补1倍工时B", "工伤\n(小时)", "计算字段"),
	("bereavement_credit", "应补1倍工时B", "丧假\n(小时)", "计算字段"),
	("marriage_credit", "应补1倍工时B", "婚假\n(半天)", "计算字段"),
	("sick_leave_deduction", "应扣2倍工时F", "病假扣工时50%", "计算字段"),
	("personal_leave_deduction", "应扣2倍工时F", "事假\n工时", "计算字段"),
	("rest_arrangement_deduction", "工作日排休应扣1.5倍工时E", "", "计算字段"),
	("settlement_one_pre", "调整前工时", "1倍结算工时=A+B", "计算字段"),
	("settlement_15_pre", "调整前工时", "1.5倍结算工时=C", "计算字段"),
	("settlement_20_pre", "调整前工时", "2倍结算工时=D", "计算字段"),
	("settlement_30", "调整前工时", "3倍结算工时", "计算字段"),
	("absence_15_pre", "调整前缺勤工时", "1.5倍缺勤工时=E", "计算字段"),
	("absence_20_pre", "调整前缺勤工时", "2倍缺勤工时=F", "计算字段"),
	("absence_30_pre", "调整前缺勤工时", "3倍缺勤工时=F", "计算字段"),
	("adjusted_absence_15", "调整后缺勤工时（验算）", "调整后1.5倍缺勤工时", "计算字段"),
	("adjusted_absence_20", "调整后缺勤工时（验算）", "调整后2倍缺勤工时", "计算字段"),
	("adjusted_one", "调整后工时", "1倍结算工时", "计算字段"),
	("adjusted_one_absence", "调整后工时", "调整后1倍缺勤工时", "计算字段"),
	("settlement_15", "调整后工时", "1.5倍结算工时=平特+工作日加班", "计算字段"),
	("settlement_20", "调整后工时", "2倍结算工时=周特+休息日加班", "计算字段"),
	("standard_hours_check", "调整后工时", "（验算用）标准工时=1倍结算工时+1.5倍缺勤工时+2倍缺勤工时", "计算字段"),
	("settlement_30_check", "调整后工时", "3倍节假日加班\n工时", "计算字段"),
	("large_night_shifts", "大\n夜\n班", "", "来源字段"),
	("small_night_shifts", "小\n夜\n班", "", "来源字段"),
	("absence_deduction", "旷工(小时)工时扣3倍", "", "计算字段"),
	("proposal_bonus", "提案改善奖金", "", "人工填写"),
	("cross_department_support", "跨部门支援奖", "", "人工填写"),
	("maintenance_bonus", "保养奖励", "", "人工填写"),
	("green_apple_amount", "绿\n苹\n果", "", "来源字段"),
	("red_apple_amount", "红苹果\n（包含忘打卡）", "", "来源字段"),
	("housing_allowance", "住房\n补贴", "", "来源字段"),
	("full_attendance_award", "全勤\n（含迟到）", "", "来源字段"),
	("employee_signature", "签名", "", "人工填写"),
	("review_note", "备注", "", "人工填写"),
)


def _attendance_final_excel_fields() -> list[dict[str, Any]]:
	"""Return the published-in-settings heading configuration with safe defaults."""
	configured = {}
	if hasattr(frappe, "get_single"):
		try:
			settings = frappe.get_single("HR Settings")
			configured = {
				str(item.field_key or "").strip(): item
				for item in settings.get("attendance_final_excel_fields") or []
				if item.enabled and str(item.field_key or "").strip()
			}
		except getattr(frappe, "DoesNotExistError", Exception):
			pass

	fields = []
	for key, main_header, sub_header, field_type in SIGNED_FINAL_FIELD_LAYOUT:
		item = configured.get(key)
		fields.append({
			"key": key,
			"main_header": str(item.excel_main_header or "").strip() if item and item.excel_main_header else main_header,
			"sub_header": str(item.excel_sub_header or "").strip() if item and item.excel_sub_header else sub_header,
			"field_type": str(item.field_type or field_type) if item else field_type,
			"comparison_policy": str(item.comparison_policy or "必须一致") if item else "必须一致",
			"numeric_tolerance": _as_number(item.numeric_tolerance) if item else 0.0,
		})
	return fields


def _attendance_final_header_candidates(field_key: str, default_label: str, row_number: int) -> list[str]:
	for field in _attendance_final_excel_fields():
		if field["key"] == field_key:
			configured_label = field["sub_header"] if row_number == 3 and field["sub_header"] else field["main_header"]
			return list(dict.fromkeys([configured_label, default_label]))
	return [default_label]


@frappe.whitelist()
def seed_attendance_final_excel_fields():
	"""Populate HR Settings with the current sign-off layout once, without overwriting edits."""
	_require_processing_manager()
	settings = frappe.get_single("HR Settings")
	existing = {str(item.field_key or "").strip() for item in settings.get("attendance_final_excel_fields") or []}
	added = 0
	for key, main_header, sub_header, field_type in SIGNED_FINAL_FIELD_LAYOUT:
		if key in existing:
			continue
		settings.append("attendance_final_excel_fields", {
			"enabled": 1,
			"field_key": key,
			"field_label": sub_header or main_header,
			"excel_main_header": main_header,
			"excel_sub_header": sub_header,
			"field_type": field_type,
			"comparison_policy": "必须一致" if field_type != "人工填写" else "仅提醒",
		})
		added += 1
	settings.save(ignore_permissions=True)
	frappe.db.commit()
	return {"added": added, "total": len(settings.get("attendance_final_excel_fields") or [])}


def _attendance_final_excel_config_hash() -> str:
	payload = [{key: item[key] for key in ("key", "main_header", "sub_header", "field_type", "comparison_policy", "numeric_tolerance")} for item in _attendance_final_excel_fields()]
	return hashlib.sha256(_json(payload).encode()).hexdigest()


# Version nine invalidates previous files and also fingerprints heading changes,
# so an edited Settings table cannot accidentally keep an old export.
MONTHLY_FINAL_LAYOUT_VERSION = 9


# The employee-facing file deliberately follows the paper confirmation form
# used by HR: two-tier headings, a coded colour band and dedicated signature
# columns.  It is different from the compact finance export by design.
SIGNED_CONFIRMATION_COLUMN_WIDTHS = (7, 12, 11, 11, 13, 14, 13, 13, 12, 8, 8, 8, 11, 8, 12, 12, 12, 16, 16)
SIGNED_CONFIRMATION_FIELD_CODES = (1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 12, 13, 14, 17, 18, 19, 20, "", "")


def _as_number(value: Any) -> float:
	try:
		return 0.0 if value in (None, "") else float(value)
	except (TypeError, ValueError):
		return 0.0


def _company_statutory_holidays(company: str, attendance_month: str) -> set[date]:
	"""Return the company's non-weekly-off holidays for the target month."""
	try:
		year, month = (int(part) for part in attendance_month.split("-", 1))
		holiday_list = frappe.db.get_value("Company", company, "default_holiday_list") if company else ""
		if not holiday_list:
			return set()
		values = frappe.get_all(
			"Holiday",
			filters={
				"parent": holiday_list,
				"holiday_date": ["between", [date(year, month, 1), date(year, month, monthrange(year, month)[1])]],
				"weekly_off": 0,
			},
			pluck="holiday_date",
		)
		return {
			value if isinstance(value, date) else date.fromisoformat(str(value)[:10])
			for value in values
		}
	except (AttributeError, TypeError, ValueError):
		return set()


def _special_hours_breakdown(entries: list[dict[str, Any]] | None, attendance_month: str, company: str = "") -> dict[str, float]:
	"""Split special-hour entries into the three overtime-rate inputs.

	The special-hours worksheet records one value for each calendar date.  The
	monthly final needs those values at their correct rate: weekday, weekend or
	holiday.  A dated row is therefore never treated as one undifferentiated
	``特殊工时`` total again.  Statutory-holiday dates come from the company's
	configured Holiday List and take precedence over the weekend split.
	"""
	try:
		year, month = (int(part) for part in attendance_month.split("-", 1))
	except (TypeError, ValueError):
		return {"special_workday_hours": 0.0, "special_restday_hours": 0.0, "special_holiday_hours": 0.0}
	result = {"special_workday_hours": 0.0, "special_restday_hours": 0.0, "special_holiday_hours": 0.0}
	statutory_holidays = _company_statutory_holidays(company, attendance_month)
	for entry in entries or []:
		day = cint(entry.get("day"))
		if not day:
			continue
		try:
			day_date = date(year, month, day)
		except ValueError:
			continue
		hours = _as_number(entry.get("hours"))
		key = (
			"special_holiday_hours"
			if day_date in statutory_holidays
			else "special_restday_hours"
			if day_date.weekday() >= 5
			else "special_workday_hours"
		)
		result[key] += hours
	return result


def _final_calculation(row: dict[str, Any]) -> dict[str, float]:
	"""Mirror the HR paper-form calculation chain for one monthly employee row."""
	actual = _as_number(row.get("actual_attendance_hours"))
	standard = _as_number(row.get("standard_hours"))
	personal = _as_number(row.get("personal_leave_hours"))
	sick = _as_number(row.get("sick_leave_hours"))
	annual = _as_number(row.get("annual_leave_hours"))
	injury = _as_number(row.get("work_injury_hours"))
	rest = _as_number(row.get("rest_arrangement_hours"))
	absence = _as_number(row.get("absence_hours"))
	bereavement = _as_number(row.get("bereavement_leave_hours"))
	marriage = _as_number(row.get("marriage_leave_half_days"))
	regular_special = _as_number(row.get("special_workday_hours"))
	rest_special = _as_number(row.get("special_restday_hours"))
	holiday_special = _as_number(row.get("special_holiday_hours"))
	settlement_15_pre = regular_special + _as_number(row.get("workday_overtime_hours"))
	settlement_20_pre = rest_special + _as_number(row.get("restday_overtime_hours"))
	settlement_30 = holiday_special + _as_number(row.get("holiday_overtime_hours"))
	actual_checked = actual - sick / 2 - annual - injury - bereavement - marriage
	credit_one = sick / 2 + annual + injury + bereavement + marriage
	absence_15_pre = rest
	absence_20_pre = sick / 2 + personal
	adjusted_absence_15 = max(absence_15_pre - settlement_15_pre, 0)
	adjusted_absence_20 = max(absence_20_pre - settlement_20_pre, 0)
	adjusted_one = actual_checked + credit_one + absence_15_pre + absence_20_pre - adjusted_absence_15 - adjusted_absence_20
	return {
		"actual_checked": actual_checked,
		"special_workday_hours": regular_special,
		"special_restday_hours": rest_special,
		"special_holiday_hours": holiday_special,
		"settlement_15_pre": settlement_15_pre,
		"settlement_20_pre": settlement_20_pre,
		"settlement_30": settlement_30,
		"adjusted_one": adjusted_one,
		"adjusted_one_absence": standard - adjusted_one,
		"settlement_15": max(settlement_15_pre - absence_15_pre, 0),
		"settlement_20": max(settlement_20_pre - absence_20_pre, 0),
	}


def _final_snapshot_batches(company: str, attendance_month: str):
	return {source_type: _latest_batch(company, attendance_month, source_type) for source_type in SOURCE_TYPES + MONTHLY_SUPPORT_SOURCE_TYPES}


def _legacy_final_snapshot_batches(company: str, attendance_month: str):
	return {source_type: _latest_batch(company, attendance_month, source_type) for source_type in SOURCE_TYPES + LEGACY_MONTHLY_SUPPORT_SOURCE_TYPES}


def _processing_state_hash(batch) -> str:
	"""Fingerprint the reviewed values that are allowed into the monthly final."""
	rows = frappe.get_all(
		PROCESSING_RECORD_DOCTYPE,
		filters={"import_batch": batch.name},
		fields=["name", "review_status", "eligible_for_downstream", "confirmed_value_json", "modified"],
		order_by="name asc",
		limit_page_length=5000,
	)
	return hashlib.sha256(_json(rows).encode()).hexdigest()


def _monthly_snapshot_version(batches: dict[str, Any]) -> str:
	"""Create one version from both source files and reviewed processing values."""
	material = [
		{
			"source_type": source_type,
			"batch": batch.name,
			"checksum": batch.source_checksum,
			"processing_state": _processing_state_hash(batch),
		}
		for source_type, batch in sorted(batches.items())
	]
	return hashlib.sha256(_json(material).encode()).hexdigest()[:16]


def _monthly_final_rows(batches: dict[str, Any]):
	"""Aggregate confirmed processing rows without recalculating their source facts."""
	rows_by_employee = defaultdict(dict)
	attendance_population = set()
	for source_type, batch in sorted(batches.items(), key=lambda item: item[0] != "attendance_draft"):
		if not batch:
			continue
		for record in _result_rows(batch, 5000):
			if not record.get("eligible_for_downstream"):
				continue
			values = _effective_result_values(record)
			code = str(values.get("employee_code") or record.get("employee_code") or "").strip()
			name = str(values.get("employee_name") or record.get("employee_name") or "").strip()
			if not code and not name:
				continue
			key = code or f"name:{name}"
			if source_type == "attendance_draft":
				attendance_population.add(key)
			elif key not in attendance_population:
				# Monthly additions and reward files can contain template carry-over or
				# historical employees. Attendance defines the payroll population; a
				# supplemental source must never create an extra salary recipient.
				continue
			output = rows_by_employee[key]
			output.setdefault("employee_code", code)
			output.setdefault("employee_name", name)
			output.setdefault("department", values.get("department") or record.get("department") or "")
			if source_type == "attendance_draft":
				for field, _label in ATTENDANCE_DRAFT_RESULT_COLUMNS:
					output[field] = values.get(field, output.get(field, 0))
				for field in ("green_apple_amount", "red_apple_amount", "housing_allowance", "full_attendance_award", "signed_final_override"):
					if field in values:
						output[field] = values.get(field)
			elif source_type == "missing_card":
				output["red_apples"] = _as_number(output.get("red_apples")) + _as_number(values.get("red_apples"))
				if not output.get("signed_final_override"):
					output["red_apple_amount"] = _as_number(output.get("red_apple_amount")) + _as_number(values.get("amount"))
			elif source_type == "apple_tree":
				apple_count = _as_number(values.get("有效苹果数"))
				if "绿" in str(values.get("苹果类型") or ""):
					output["green_apples"] = _as_number(output.get("green_apples")) + apple_count
				elif "红" in str(values.get("苹果类型") or ""):
					output["red_apples"] = _as_number(output.get("red_apples")) + apple_count
			elif source_type == "special_hours":
				# The grid supplies a dated value for every entry.  Retain the total
				# for audit, but pass the rate-specific values to the final calculator.
				breakdown = _special_hours_breakdown(values.get("special_hours_days"), batch.attendance_month, batch.company)
				output["special_hours"] = _as_number(output.get("special_hours")) + _as_number(values.get("special_hours"))
				for field, amount in breakdown.items():
					output[field] = _as_number(output.get(field)) + amount
			elif source_type in {"housing_allowance", "full_attendance"}:
				if output.get("signed_final_override"):
					continue
				field = MONTHLY_SUPPORT_SOURCE_CONFIG[source_type]["value_field"]
				output[field] = _as_number(output.get(field)) + _as_number(values.get(field))
	return sorted(rows_by_employee.values(), key=lambda row: (str(row.get("department") or ""), str(row.get("employee_code") or ""), str(row.get("employee_name") or "")))


def _finance_final_preview_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
	"""Present special hours inside the finance overtime totals.

	The finance workbook already uses the three rate-specific settlement totals.
	Keep the webpage view on that same contract, rather than exposing a separate
	"特殊工时" total that payroll could accidentally omit.
	"""
	preview_rows = []
	for source_row in rows:
		row = dict(source_row)
		calculation = _final_calculation(row)
		row["workday_overtime_hours"] = calculation["settlement_15_pre"]
		row["restday_overtime_hours"] = calculation["settlement_20_pre"]
		row["holiday_overtime_hours"] = calculation["settlement_30"]
		preview_rows.append(row)
	return preview_rows


def _save_monthly_final_file(attendance_month: str, title: str, columns, rows):
	from openpyxl import Workbook
	from frappe.utils.file_manager import save_file

	book = Workbook()
	sheet = book.active
	sheet.title = title
	sheet.append([label for _field, label in columns])
	for row in rows:
		sheet.append([row.get(field, 0 if field not in {"employee_code", "employee_name", "department", "employee_signature", "review_note"} else "") for field, _label in columns])
	for column in sheet.columns:
		sheet.column_dimensions[column[0].column_letter].width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 24)
	sheet.freeze_panes = "A2"
	output = BytesIO()
	book.save(output)
	file = save_file(f"{attendance_month}_{title}.xlsx", output.getvalue(), None, None, is_private=1)
	return {"file_url": file.file_url, "file_name": file.file_name}


def _save_monthly_finance_confirmation_file(attendance_month: str, rows):
	"""Create the compact finance-facing confirmation workbook."""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter
	from frappe.utils.file_manager import save_file

	book = Workbook()
	sheet = book.active
	sheet.title = "财务版"
	last_column = len(SIGNED_CONFIRMATION_COLUMN_WIDTHS)
	last_column_letter = get_column_letter(last_column)
	month_label = f"{int(attendance_month.split('-')[1])}月" if "-" in attendance_month else attendance_month
	sheet.merge_cells(f"A1:{last_column_letter}1")
	sheet["A1"] = f"{month_label}工时奖惩确认表"

	# Two-tier headings mirror the supplied HR form.  Single-purpose columns
	# span both heading rows; overtime columns sit beneath the grouped heading.
	vertical_headers = {
		"A": "序号", "B": "部门", "C": "姓名", "D": "标准工时\n（小时）", "E": "钉钉导出\n实际出勤\n（小时）",
		"J": "大夜\n班（55\n元）", "K": "大夜\n班（45\n元）", "L": "小\n夜\n班", "M": "旷工\n（小时）\n工时扣3\n倍",
		"N": "绿\n苹\n果", "O": "红苹果\n（包含\n忘打卡）", "P": "住房\n补贴", "Q": "全勤\n（含迟\n到）", "R": "签名", "S": "备注",
	}
	for column, label in vertical_headers.items():
		sheet[f"{column}2"] = label
		sheet.merge_cells(f"{column}2:{column}3")
	sheet.merge_cells("F2:I2")
	sheet["F2"] = "调整后工时"
	for column, label in {
		"F": "调整后\n1倍结算\n工时",
		"G": "1.5倍结算\n工时=平常\n工作日加班",
		"H": "2倍结算工\n时=周末+\n休息日加班",
		"I": "3倍节\n假日加\n班工时",
	}.items():
		sheet[f"{column}3"] = label

	for column, code in enumerate(SIGNED_CONFIRMATION_FIELD_CODES, start=1):
		sheet.cell(row=4, column=column, value=code)

	thin_black = Side(style="thin", color="000000")
	border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
	pale_yellow = PatternFill("solid", fgColor="FFFFCC")
	peach = PatternFill("solid", fgColor="F8CBAD")
	blue = PatternFill("solid", fgColor="D9E2F3")
	bright_yellow = PatternFill("solid", fgColor="FFFF00")
	code_gold = PatternFill("solid", fgColor="FFD966")
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	for row in sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=last_column):
		for cell in row:
			cell.font = Font(name="宋体", size=12, bold=True)
			cell.alignment = center
			cell.border = border
			cell.fill = pale_yellow
	for coordinate in ("F3",):
		sheet[coordinate].fill = peach
	for coordinate in ("G3",):
		sheet[coordinate].fill = blue
	for coordinate in ("H3", "M2"):
		sheet[coordinate].fill = bright_yellow
	for column in range(1, 6):
		sheet.cell(row=4, column=column).fill = code_gold
	for column in range(6, last_column + 1):
		sheet.cell(row=4, column=column).fill = PatternFill("solid", fgColor="FFFFFF")

	sheet["A1"].font = Font(name="宋体", size=24, bold=False)
	sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
	sheet.row_dimensions[1].height = 48
	sheet.row_dimensions[2].height = 40
	sheet.row_dimensions[3].height = 92
	sheet.row_dimensions[4].height = 34
	for index, width in enumerate(SIGNED_CONFIRMATION_COLUMN_WIDTHS, start=1):
		sheet.column_dimensions[get_column_letter(index)].width = width

	for index, row in enumerate(rows, start=5):
		calculation = _final_calculation(row)
		actual_hours = _as_number(row.get("actual_attendance_hours"))
		values = [
			index - 4,
			_display_department(row.get("department")),
			row.get("employee_name") or "",
			_as_number(row.get("standard_hours")),
			actual_hours,
			calculation["adjusted_one"],
			calculation["settlement_15"],
			calculation["settlement_20"],
			calculation["settlement_30"],
			_as_number(row.get("large_night_shifts")),
			"",  # The current source does not distinguish the 45-yuan large-night rate.
			_as_number(row.get("small_night_shifts")),
			_as_number(row.get("absence_hours")),
			_as_number(row.get("green_apple_amount")),
			_as_number(row.get("red_apple_amount")),
			_as_number(row.get("housing_allowance")),
			_as_number(row.get("full_attendance_award")),
			row.get("employee_signature") or "",
			row.get("review_note") or "",
		]
		for column, value in enumerate(values, start=1):
			cell = sheet.cell(row=index, column=column, value=value)
			cell.border = border
			cell.alignment = Alignment(horizontal="center" if column not in {2, 3, 18, 19} else "left", vertical="center", wrap_text=True)
			cell.font = Font(name="宋体", size=11)
			if column >= 4 and column <= 17:
				cell.number_format = "0.0"
		sheet.row_dimensions[index].height = 28

	sheet.freeze_panes = "A5"
	sheet.sheet_view.showGridLines = False
	output = BytesIO()
	book.save(output)
	file = save_file(f"{attendance_month}_财务版.xlsx", output.getvalue(), None, None, is_private=1)
	return {"file_url": file.file_url, "file_name": file.file_name}


def _save_monthly_signed_confirmation_file(attendance_month: str, rows):
	"""Create the complete multi-band employee sign-off form supplied by HR."""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter
	from frappe.utils.file_manager import save_file

	book = Workbook()
	sheet = book.active
	sheet.title = "员工签字版"
	# Labels and two-level groups come from HR Settings.  Formula positions stay
	# stable, while visible Excel wording can evolve with HR's monthly form.
	excel_fields = _attendance_final_excel_fields()
	# Preserve column A as the same blank margin used by HR's original file.
	# The visible form is B:BH, which also keeps every formula letter identical
	# to the audited manual confirmation workbook.
	form_start, form_end = 2, 60
	for offset, field in enumerate(excel_fields):
		column = form_start + offset
		sheet.cell(row=2, column=column, value=field["main_header"])
		if field["sub_header"]:
			sheet.cell(row=3, column=column, value=field["sub_header"])
	# Contiguous equal first-level headings form an Excel group; a lone field
	# spans both heading rows.  This supports renaming and regrouping in Settings.
	start = 0
	while start < len(excel_fields):
		main_header = excel_fields[start]["main_header"]
		end = start + 1
		while end < len(excel_fields) and excel_fields[end]["main_header"] == main_header and excel_fields[end]["sub_header"]:
			end += 1
		if end - start > 1:
			sheet.merge_cells(start_row=2, start_column=form_start + start, end_row=2, end_column=form_start + end - 1)
		else:
			sheet.merge_cells(start_row=2, start_column=form_start + start, end_row=3, end_column=form_start + start)
		start = end

	field_codes = [1, 2, "", 4, "", 5, 6, "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", 7, 8, 9, 10, "", 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, "", ""]
	for column, code in enumerate(field_codes, start=form_start):
		sheet.cell(row=4, column=column, value=code)

	thin = Side(style="thin", color="000000")
	blue_side = Side(style="medium", color="0000FF")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	pale = PatternFill("solid", fgColor="FFFFCC")
	peach = PatternFill("solid", fgColor="F8CBAD")
	blue = PatternFill("solid", fgColor="D9E2F3")
	yellow = PatternFill("solid", fgColor="FFFF00")
	gold = PatternFill("solid", fgColor="FFD966")
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	for row in sheet.iter_rows(min_row=2, max_row=4, min_col=form_start, max_col=form_end):
		for cell in row:
			cell.font = Font(name="宋体", size=10, bold=True)
			cell.alignment = center
			cell.border = border
			cell.fill = pale
	for column in range(26, 33):
		for row in (2, 3): sheet.cell(row=row, column=column).fill = peach
	for column in range(34, 43):
		for row in (2, 3): sheet.cell(row=row, column=column).fill = blue
	for column in (12, 13, 14, 15, 31, 32, 46, 47, 51):
		for row in (2, 3): sheet.cell(row=row, column=column).fill = yellow
	for column in range(2, 10): sheet.cell(row=4, column=column).fill = gold

	sheet.merge_cells("D1:BH1")
	month_label = f"{int(attendance_month.split('-')[1])}月" if "-" in attendance_month else attendance_month
	sheet["D1"] = f"{month_label}工时奖惩确认表"
	sheet["D1"].font = Font(name="宋体", size=20)
	sheet["D1"].alignment = Alignment(horizontal="center", vertical="center")
	sheet.row_dimensions[1].height = 42
	sheet.row_dimensions[2].height = 32
	sheet.row_dimensions[3].height = 72
	sheet.row_dimensions[4].height = 30
	sheet.column_dimensions["A"].width = 2
	for column in range(form_start, form_end + 1): sheet.column_dimensions[get_column_letter(column)].width = 9
	for column in (3, 5, 59, 60): sheet.column_dimensions[get_column_letter(column)].width = 13
	for column in (4, 7, 8, 9, 47): sheet.column_dimensions[get_column_letter(column)].width = 11

	for excel_row, row in enumerate(rows, start=5):
		special = _final_calculation(row)
		actual = _as_number(row.get("actual_attendance_hours"))
		standard = _as_number(row.get("standard_hours"))
		workday, restday, holiday = (_as_number(row.get("workday_overtime_hours")), _as_number(row.get("restday_overtime_hours")), _as_number(row.get("holiday_overtime_hours")))
		personal, sick, annual, injury, rest, absence = (_as_number(row.get("personal_leave_hours")), _as_number(row.get("sick_leave_hours")), _as_number(row.get("annual_leave_hours")), _as_number(row.get("work_injury_hours")), _as_number(row.get("rest_arrangement_hours")), _as_number(row.get("absence_hours")))
		values = [excel_row - 4, _display_department(row.get("department")), row.get("employee_code") or "", row.get("employee_name") or "", "", standard, actual,
			f"=H{excel_row}-Q{excel_row}/2-R{excel_row}-S{excel_row}-V{excel_row}-W{excel_row}", special["special_workday_hours"], workday, special["special_restday_hours"], restday, special["special_holiday_hours"], holiday,
			personal, sick, annual, injury, rest, absence, 0, 0, 0, 0,
			f"=Q{excel_row}*0.5", f"=R{excel_row}", f"=S{excel_row}", f"=V{excel_row}", f"=W{excel_row}", f"=Q{excel_row}*0.5", f"=P{excel_row}", f"=T{excel_row}",
			f"=I{excel_row}+Z{excel_row}+AA{excel_row}+AB{excel_row}+AC{excel_row}+AD{excel_row}", f"=J{excel_row}+K{excel_row}", f"=L{excel_row}+M{excel_row}", f"=N{excel_row}+O{excel_row}", f"=AG{excel_row}", f"=AE{excel_row}+AF{excel_row}", 0,
			f"=IF(AL{excel_row}-AI{excel_row}>0,AL{excel_row}-AI{excel_row},0)", f"=IF(AM{excel_row}-AJ{excel_row}>0,AM{excel_row}-AJ{excel_row},0)", f"=AH{excel_row}+AL{excel_row}+AM{excel_row}-AO{excel_row}-AP{excel_row}", f"=G{excel_row}-AQ{excel_row}", f"=IF(AI{excel_row}-AL{excel_row}>0,AI{excel_row}-AL{excel_row},0)", f"=IF(AJ{excel_row}-AM{excel_row}>0,AJ{excel_row}-AM{excel_row},0)", f"=AQ{excel_row}+AN{excel_row}+AO{excel_row}", f"=AK{excel_row}",
			_as_number(row.get("large_night_shifts")), _as_number(row.get("small_night_shifts")), absence, 0, 0, 0, _as_number(row.get("green_apple_amount")), _as_number(row.get("red_apple_amount")), _as_number(row.get("housing_allowance")), _as_number(row.get("full_attendance_award")), row.get("employee_signature") or "", row.get("review_note") or ""]
		for column, value in enumerate(values, start=form_start):
			cell = sheet.cell(row=excel_row, column=column, value=value)
			cell.border = border
			cell.alignment = Alignment(horizontal="center" if column not in {3, 5, 59, 60} else "left", vertical="center", wrap_text=True)
			cell.font = Font(name="宋体", size=10)
			if 7 <= column <= 58: cell.number_format = "0.0"
		sheet.row_dimensions[excel_row].height = 26

	for row in sheet.iter_rows(min_row=1, max_row=max(4, len(rows) + 4), min_col=form_start, max_col=form_end):
		row[0].border = Border(left=blue_side, top=row[0].border.top, right=row[0].border.right, bottom=row[0].border.bottom)
		row[-1].border = Border(left=row[-1].border.left, top=row[-1].border.top, right=blue_side, bottom=row[-1].border.bottom)
	sheet.freeze_panes = "B5"
	sheet.sheet_view.showGridLines = False
	output = BytesIO()
	book.save(output)
	file = save_file(f"{attendance_month}_员工签字版.xlsx", output.getvalue(), None, None, is_private=1)
	return {"file_url": file.file_url, "file_name": file.file_name}


@frappe.whitelist()
def generate_monthly_final_files(company: str, attendance_month: str, snapshot_version: str = ""):
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	state = get_processing_batch(company, attendance_month)
	readiness = state["finalization_inputs"]
	blocked = [item for item in readiness if not item["ready"]]
	if blocked:
		return {"blocked": True, "reason": _("月度终稿来源未完备，尚未生成任何文件。"), "readiness": readiness, "missing_sources": [item["label"] for item in blocked]}
	batches = _final_snapshot_batches(company, attendance_month)
	if any(not batch for batch in batches.values()):
		return {"blocked": True, "reason": _("终稿来源不完整，尚未生成文件。"), "readiness": readiness}
	locked_snapshot_version = _monthly_snapshot_version(batches)
	config_hash = _attendance_final_excel_config_hash()
	anchor_batch = batches["attendance_draft"]
	existing_outputs = _processing_meta(anchor_batch).get("monthly_final_outputs", {})
	if existing_outputs.get("locked_snapshot_version") == locked_snapshot_version and existing_outputs.get("layout_version") == MONTHLY_FINAL_LAYOUT_VERSION and existing_outputs.get("attendance_final_excel_config_hash") == config_hash and existing_outputs.get("signed_file_url") and existing_outputs.get("finance_file_url"):
		return {"blocked": False, "readiness": readiness, "final_outputs": existing_outputs, "snapshot_version": locked_snapshot_version}
	rows = _monthly_final_rows(batches)
	if not rows:
		return {"blocked": True, "reason": _("没有可进入下游的员工数据，尚未生成终稿。"), "readiness": readiness}
	signed = _save_monthly_signed_confirmation_file(attendance_month, rows)
	finance = _save_monthly_finance_confirmation_file(attendance_month, rows)
	final_outputs = {
		"locked_version": locked_snapshot_version,
		"locked_snapshot_version": locked_snapshot_version,
		"signed_file_url": signed["file_url"],
		"signed_file_name": signed["file_name"],
		"finance_file_url": finance["file_url"],
		"finance_file_name": finance["file_name"],
		"generated_on": now_datetime().isoformat(),
		"employee_count": len(rows),
		"layout_version": MONTHLY_FINAL_LAYOUT_VERSION,
		"attendance_final_excel_config_hash": config_hash,
	}
	_save_batch_notes(anchor_batch, {"monthly_final_outputs": final_outputs})
	frappe.db.commit()
	return {"blocked": False, "readiness": readiness, "final_outputs": final_outputs, "snapshot_version": locked_snapshot_version}


@frappe.whitelist()
def get_monthly_final_preview(company: str, attendance_month: str, kind: str = "signed"):
	"""Return the exact current locked-final table for in-system review."""
	_require_processing_manager()
	company, attendance_month = _require_company(company), _require_month(attendance_month)
	if kind not in {"signed", "finance"}:
		frappe.throw(_("终稿预览类型不正确。"))
	state = get_processing_batch(company, attendance_month)
	outputs = state.get("final_outputs") or {}
	if not outputs.get("locked_snapshot_version"):
		return {"available": False, "reason": _("请先锁定并生成月度终稿。")}
	batches = _final_snapshot_batches(company, attendance_month)
	if any(not batch for batch in batches.values()):
		return {"available": False, "reason": _("终稿来源不完整，无法提供预览。")}
	locked_version = str(outputs.get("locked_snapshot_version") or "")
	current_version = _monthly_snapshot_version(batches)
	preview_batches = batches
	if current_version != locked_version:
		legacy_batches = _legacy_final_snapshot_batches(company, attendance_month)
		legacy_matches = bool(all(legacy_batches.values())) and _monthly_snapshot_version(legacy_batches) == locked_version
		if not legacy_matches:
			return {"available": False, "stale": True, "reason": _("来源或人工处理已变化，请重新锁定并生成终稿后再查看。")}
		preview_batches = legacy_batches
	rows = _monthly_final_rows(preview_batches)
	columns = FINAL_SIGNED_COLUMNS if kind == "signed" else FINAL_FINANCE_COLUMNS
	return {
		"available": True,
		"kind": kind,
		"title": _("员工签字版") if kind == "signed" else _("财务版"),
		"locked_snapshot_version": locked_version,
		"columns": [{"field": field, "label": label} for field, label in columns],
		"rows": rows if kind == "signed" else _finance_final_preview_rows(rows),
	}
