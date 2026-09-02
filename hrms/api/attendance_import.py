import hashlib
import json
import re
from collections import defaultdict
from datetime import date, datetime, time
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt, getdate, now_datetime


REQUIRED_ATTENDANCE_SHEETS = ["1.1每日统计", "1.2请假单", "1.3苹果树"]
DINGTALK_EXPORT_V1_SHEETS = ["每日统计", "打卡时间", "原始记录", "月度汇总"]
COMPANY_ATTENDANCE_WORKBOOK_SOURCES = {
	"dingtalk_raw": {
		"sheet_name": "每日统计（钉钉导出）",
		"header_rows": (3, 4),
		"data_start_row": 5,
	},
	"manual_adjustment": {
		"sheet_name": "每日统计（修改后）",
		"header_rows": (1, 2),
		"data_start_row": 3,
	},
}
COMPANY_ATTENDANCE_REGISTER_V1_SHEETS = {
	"daily_statistics": {"sheet_name": "每日统计", "header_rows": (1, 2), "data_start_row": 3, "writes_daily_checks": True},
	"attendance_detail": {"sheet_name": "出勤明细", "header_rows": (1, 2), "data_start_row": 3, "writes_daily_checks": False},
	"attendance_exception": {"sheet_name": "出勤异常", "header_rows": (2,), "data_start_row": 3, "writes_daily_checks": False},
	"apple_tree": {"sheet_name": "苹果树", "header_rows": (3,), "data_start_row": 4, "writes_daily_checks": False},
}
COMPANY_DAILY_STATISTICS_COLUMNS = [
	("姓名", ""), ("工号", ""), ("日期", ""), ("实际部门", ""), ("班次", ""), ("上班时间", ""), ("下班时间", ""),
	("上班缺卡", ""), ("下班缺卡", ""), ("旷工", ""), ("标准工时", ""), ("实际出勤（小时）", ""), ("关联审批单", ""),
	("工作日加班（小时）", ""), ("休息日加班（小时）", ""), ("节假日加班（小时）", ""), ("大夜班", ""), ("小夜班", ""),
	("请假", "事假(小时)"), ("", "病假(小时)"), ("", "婚假(天)"), ("", "特休(小时)"), ("", "丧假(小时)"),
	("", "工伤(小时)"), ("", "公假(天)"), ("", "产假(天)"), ("", "团圆假(天)"), ("", "排休(小时)"),
	("", "旷工(小时)"),
	# The company source sheet retains a second leave summary block without units.
	# It is kept as source evidence and is not added to leave hours.
	("请假", "婚假"), ("", "丧假"), ("", "公假"), ("", "产假"), ("", "团圆假"),
	("旷工", ""), ("上班未打卡次数", ""), ("下班未打卡次数", ""), ("迟到次数", ""), ("早退次数", ""),
]
ATTENDANCE_IMPORT_TEMPLATES = [
	{
		"key": "company_attendance_register_v1",
		"label": "公司考勤工作簿（推荐）",
		"description": "包含每日统计、出勤明细、出勤异常、苹果树及填写说明；这是公司人工汇总后的标准上传格式。",
		"upload_mode": "whole_workbook",
		"sheet_names": [source["sheet_name"] for source in COMPANY_ATTENDANCE_REGISTER_V1_SHEETS.values()],
		"writes": "每日统计写入每日考勤核对；苹果树作为奖惩来源留存；出勤明细和出勤异常用于核对。",
	},
	{
		"key": "company_daily_statistics_v1",
		"label": "每日统计（单表）",
		"description": "按员工、日期填写工时、班次、加班、请假和缺卡数据。下载后请替换到整套工作簿的“每日统计”页签。",
		"upload_mode": "replace_sheet",
		"sheet_names": ["每日统计"],
		"writes": "必须放入“公司考勤工作簿（推荐）”后一起上传。",
	},
	{
		"key": "attendance_detail_v1",
		"label": "出勤明细（部门日报）",
		"description": "按部门记录现有人数、出勤人数、请假人员及当日人员变动。",
		"upload_mode": "replace_sheet",
		"sheet_names": ["出勤明细"],
		"writes": "用于部门核对，不直接覆盖每日考勤结果。",
	},
	{
		"key": "attendance_exception_v1",
		"label": "出勤异常处理表",
		"description": "记录迟到、早退、旷工、未打卡和后补假卡的处理及签字确认。",
		"upload_mode": "replace_sheet",
		"sheet_names": ["出勤异常"],
		"writes": "用于异常核对；最终处理仍在系统“考勤确认”中完成。",
	},
	{
		"key": "apple_reward_v1",
		"label": "苹果树奖惩表",
		"description": "记录奖惩日期、人员、绿苹果、红苹果、奖惩项目和备注。",
		"upload_mode": "replace_sheet",
		"sheet_names": ["苹果树"],
		"writes": "放入整套工作簿后作为奖惩来源留存，是否计薪取决于审批/确认状态。",
	},
	{
		"key": "legacy_workbook_v1",
		"label": "旧版三表兼容模板",
		"description": "仅用于历史数据兼容，包含 1.1每日统计、1.2请假单、1.3苹果树。新月份优先使用公司考勤工作簿或钉钉原始导出。",
		"upload_mode": "whole_workbook",
		"sheet_names": REQUIRED_ATTENDANCE_SHEETS,
		"writes": "兼容导入历史数据。",
	},
]
# Export profiles follow the company documents in ``5.2人资考勤.xlsx``:
# operational detail is kept separate from the three payroll-facing monthly forms.
ATTENDANCE_EXPORT_PROFILES = {
	"company_attendance_workbook": {
		"label": "公司考勤工作簿",
		"sheet_keys": ["daily_statistics", "attendance_detail", "leave_evidence", "attendance_exception", "missing_card", "apple_reward", "monthly_draft", "monthly_signed", "monthly_finance"],
	},
	"daily_statistics": {"label": "每日统计", "sheet_keys": ["daily_statistics"]},
	"attendance_detail": {"label": "出勤明细", "sheet_keys": ["attendance_detail"]},
	"leave_evidence": {"label": "请假单", "sheet_keys": ["leave_evidence"]},
	"attendance_exception": {"label": "出勤异常", "sheet_keys": ["attendance_exception"]},
	"missing_card": {"label": "忘打卡", "sheet_keys": ["missing_card"]},
	"apple_reward": {"label": "苹果树", "sheet_keys": ["apple_reward"]},
	"monthly_draft": {"label": "考勤初稿", "sheet_keys": ["monthly_draft"]},
	"monthly_signed": {"label": "考勤终稿（签字版）", "sheet_keys": ["monthly_signed"]},
	"monthly_finance": {"label": "考勤终稿（财务版）", "sheet_keys": ["monthly_finance"]},
}
DINGTALK_EXPORT_V1_SCHEMA = {
	"source_type": "dingtalk_export_v1",
	"daily_header_rows": (3, 4),
	"daily_data_start_row": 5,
	"raw_header_row": 3,
	"raw_data_start_row": 4,
	"monthly_header_rows": (3, 4),
	"monthly_data_start_row": 5,
}
DINGTALK_DAILY_FIELD_MAPPING = {
	"姓名": "employee_name",
	"工号": "employee_code",
	"UserId": "dingtalk_user_id",
	"日期": "attendance_date",
	"workDate": "source_work_date",
	"日期类型": "date_type",
	"考勤组": "attendance_group",
	"部门": "source_department",
	"实际部门": "actual_department",
	"班次": "shift_name",
	"上班时间": "actual_in_time",
	"下班时间": "actual_out_time",
	"上班缺卡": "missing_in",
	"下班缺卡": "missing_out",
	"标准工时": "standard_hours",
	"实际出勤(小时)": "actual_attendance_hours",
	"实际出勤（小时）": "actual_attendance_hours",
	"关联的审批单": "approval_reference",
	"关联审批单": "approval_reference",
	"工作日加班(小时)": "workday_overtime_hours",
	"工作日加班（小时）": "workday_overtime_hours",
	"休息日加班(小时)": "restday_overtime_hours",
	"休息日加班（小时）": "restday_overtime_hours",
	"节假日加班(小时)": "holiday_overtime_hours",
	"节假日加班（小时）": "holiday_overtime_hours",
	"请假/事假(小时)": "personal_leave_hours",
	"请假/病假(小时)": "sick_leave_hours",
	"请假/婚假(天)": "marriage_leave_days",
	"请假/特休(小时)": "annual_leave_hours",
	"请假/丧假(小时)": "bereavement_leave_hours",
	"请假/工伤(小时)": "work_injury_leave_hours",
	"请假/公假(天)": "public_leave_days",
	"请假/产假(天)": "maternity_leave_days",
	"请假/团圆假(天)": "reunion_leave_days",
	"请假/排休(小时)": "rest_leave_hours",
	"请假/旷工(小时)": "absent_hours",
	"旷工": "absence_summary",
	"旷工_2": "absent_hours",
}
ATTENDANCE_IMPORT_REQUIRED_FIELDS = ("employee_name", "employee_code", "attendance_date")
ATTENDANCE_FIELD_CATALOG = [
	{"fieldname": "employee_name", "label": "姓名", "required": 1},
	{"fieldname": "employee_code", "label": "工号", "required": 1},
	{"fieldname": "attendance_date", "label": "日期", "required": 1},
	{"fieldname": "dingtalk_user_id", "label": "UserId", "required": 0},
	{"fieldname": "attendance_group", "label": "考勤组", "required": 0},
	{"fieldname": "actual_department", "label": "实际部门", "required": 0},
	{"fieldname": "shift_name", "label": "班次", "required": 0},
	{"fieldname": "actual_in_time", "label": "上班时间", "required": 0},
	{"fieldname": "actual_out_time", "label": "下班时间", "required": 0},
	{"fieldname": "standard_hours", "label": "标准工时", "required": 0},
	{"fieldname": "actual_attendance_hours", "label": "实际出勤", "required": 0},
	{"fieldname": "approval_reference", "label": "关联审批单", "required": 0},
	{"fieldname": "workday_overtime_hours", "label": "工作日加班", "required": 0},
	{"fieldname": "restday_overtime_hours", "label": "休息日加班", "required": 0},
	{"fieldname": "holiday_overtime_hours", "label": "节假日加班", "required": 0},
	{"fieldname": "personal_leave_hours", "label": "事假", "required": 0},
	{"fieldname": "sick_leave_hours", "label": "病假", "required": 0},
	{"fieldname": "annual_leave_hours", "label": "特休", "required": 0},
	{"fieldname": "absent_hours", "label": "旷工", "required": 0},
]
ATTENDANCE_RULE_APPLICATION_MODES = ("仅展示", "导入校验", "异常提示")
SUPPORTED_ATTENDANCE_HINT_RULE_CODES = ("ATT-LATE-30", "ATT-MISSING-CARD", "ATT-ABSENT-NO-LEAVE")
ATTENDANCE_DRAFT_IMPORT_RULE_CODES = (
	"ATT-DRAFT-MISSING-PUNCH",
	"ATT-DRAFT-LATE",
	"ATT-DRAFT-EARLY",
	"ATT-DRAFT-ABSENCE-MARKER",
	"ATT-DRAFT-RESTDAY-CLOCK-WITHOUT-OVERTIME",
	"ATT-DRAFT-SHIFT-MISSING",
)
ATTENDANCE_BATCH_DOCTYPE = "HRMS Attendance Import Batch"
DAY_CHECK_DOCTYPE = "HRMS Attendance Day Check"
LEAVE_EVIDENCE_DOCTYPE = "HRMS Attendance Leave Evidence"
EXCEPTION_DOCTYPE = "HRMS Attendance Exception"
APPLE_RECORD_DOCTYPE = "HRMS Apple Reward Record"
MONTHLY_SUMMARY_DOCTYPE = "HRMS Monthly Attendance Summary"
MONTH_LOCK_DOCTYPE = "HRMS Attendance Month Lock"
LOCK_AUDIT_DOCTYPE = "HRMS Attendance Lock Audit"
DEPARTMENT_CONFIRMATION_DOCTYPE = "HRMS Attendance Department Confirmation"
CUSTOM_RULE_DOCTYPE = "HRMS Attendance Custom Rule"
APPLE_UNIT_AMOUNT = 5
STANDARD_DAY_HOURS = 8
TEST_ATTENDANCE_DEMO_COMPANY = "TEST-HRMS"
TEST_ATTENDANCE_DEMO_MONTH = "2099-02"
TEST_ATTENDANCE_DEMO_CHECKSUM = hashlib.sha256(b"TEST-HRMS attendance demo v1").hexdigest()
LARGE_NIGHT_SHIFT_ALLOWANCE = 45
SMALL_NIGHT_SHIFT_ALLOWANCE = 24


def _rule_execution_notice():
	return "规则不会自动修改导入数据、月度终稿或薪资。"


def _rule_execution_state(rule):
	"""Keep configured formula text separate from the small audited rule executor.

	Custom expressions are deliberately never evaluated.  A rule is executable only
	when its code is backed by reviewed server-side logic below.
	"""
	mode = getattr(rule, "application_mode", "") or "仅展示"
	if not getattr(rule, "enabled", 0):
		return {"status": "已停用", "description": "规则已停用，不参与任何检查。"}
	if mode == "仅展示":
		return {"status": "说明规则", "description": "展示制度来源和处理建议，不读取考勤数据。"}
	if mode == "导入校验":
		return {"status": "导入合同校验", "description": "导入时使用固定字段映射和必填项校验；不会执行自定义公式。"}
	if getattr(rule, "rule_code", "") in ATTENDANCE_DRAFT_IMPORT_RULE_CODES:
		return {"status": "导入时运行", "description": "在考勤初稿加工时按受控内置逻辑识别；可在此启用或停用，不会执行自定义公式。"}
	if getattr(rule, "rule_code", "") in SUPPORTED_ATTENDANCE_HINT_RULE_CODES:
		return {"status": "可运行", "description": "可对当前公司和月份的有效日核对数据运行只读提示。"}
	return {"status": "待接入执行器", "description": "已保存为规则说明；自定义公式不会被系统直接执行。"}


def _import_validation(field_mapping):
	mapped_targets = {target for target in (field_mapping or {}).values() if target and target != "source_only"}
	missing = [fieldname for fieldname in ATTENDANCE_IMPORT_REQUIRED_FIELDS if fieldname not in mapped_targets]
	return {
		"status": "可导入" if not missing else "需核对",
		"required_target_fields": list(ATTENDANCE_IMPORT_REQUIRED_FIELDS),
		"missing_required_fields": missing,
		"matched_field_count": len(mapped_targets),
		"source_only_field_count": sum(1 for target in (field_mapping or {}).values() if target == "source_only"),
		"notice": "字段映射只决定文件如何写入每日考勤核对；规则不会自动修改导入数据、月度终稿或薪资。",
	}


def _with_import_validation(preview):
	preview["import_validation"] = _import_validation(preview.get("field_mapping") or {})
	return preview


@frappe.whitelist()
def get_attendance_field_mapping_catalog():
	"""Expose the fixed, auditable import contract used by all workbook previews."""
	return {
		"write_policy": "只读说明，不写入考勤数据",
		"notice": "导入文件先按该映射预览。确认导入后保留原始行；需要改正时请在每日考勤核对创建人工更正，不要直接覆盖原始文件记录。",
		"profiles": [
			{
				"source_type": "company_attendance_register_v1",
				"label": "公司考勤工作簿（推荐）",
				"required_target_fields": list(ATTENDANCE_IMPORT_REQUIRED_FIELDS),
				"source_sheets": ["每日统计", "出勤明细", "出勤异常", "苹果树"],
			},
			{
				"source_type": "dingtalk_export_v1",
				"label": "钉钉四表原始导出",
				"required_target_fields": list(ATTENDANCE_IMPORT_REQUIRED_FIELDS),
				"source_sheets": DINGTALK_EXPORT_V1_SHEETS,
			},
			{
				"source_type": "company_attendance_workbook_v1",
				"label": "公司人工修正工作簿",
				"required_target_fields": list(ATTENDANCE_IMPORT_REQUIRED_FIELDS),
				"source_sheets": [source["sheet_name"] for source in COMPANY_ATTENDANCE_WORKBOOK_SOURCES.values()],
			},
			{
				"source_type": "legacy_workbook",
				"label": "旧版三表兼容文件",
				"required_target_fields": list(ATTENDANCE_IMPORT_REQUIRED_FIELDS),
				"source_sheets": REQUIRED_ATTENDANCE_SHEETS,
			},
		],
		"fields": [dict(field) for field in ATTENDANCE_FIELD_CATALOG],
		"default_mapping": dict(DINGTALK_DAILY_FIELD_MAPPING),
	}


DEFAULT_ATTENDANCE_CUSTOM_RULES = [
	{
		"rule_code": "ATT-DRAFT-MISSING-PUNCH",
		"rule_name": "考勤初稿：缺卡识别",
		"rule_group": "考勤",
		"rule_type": "导入异常识别",
		"source_module": "钉钉每日统计",
		"source_document": "每日统计 / 上班未打卡次数、下班未打卡次数",
		"trigger_condition": "来源明确给出上班未打卡次数或下班未打卡次数大于 0。",
		"formula": "clock_in_missing_count > 0 || clock_out_missing_count > 0",
		"action_result": "按员工＋日期生成缺卡待核验事件；不自动扣款。",
		"priority": 5,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "ATT-DRAFT-LATE",
		"rule_name": "考勤初稿：迟到识别",
		"rule_group": "考勤",
		"rule_type": "导入异常识别",
		"source_module": "钉钉每日统计",
		"source_document": "每日统计 / 应上班时间、上班时间、迟到次数、请假",
		"trigger_condition": "工作日实际上班时间晚于应上班时间，且无请假证据；不存在迟到宽限分钟。",
		"formula": "late_count > 0 || (actual_in_time > scheduled_in_time && !leave_evidence)",
		"action_result": "按员工＋日期生成迟到待核验事件；不直接扣款，需结合请假或主管说明后处理。",
		"priority": 6,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "ATT-DRAFT-EARLY",
		"rule_name": "考勤初稿：早退识别",
		"rule_group": "考勤",
		"rule_type": "导入异常识别",
		"source_module": "钉钉每日统计",
		"source_document": "每日统计 / 早退次数",
		"trigger_condition": "来源明确给出早退次数大于 0。",
		"formula": "early_count > 0",
		"action_result": "按员工＋日期生成早退待核验事件；不自动形成缺勤工时。",
		"priority": 7,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "ATT-DRAFT-ABSENCE-MARKER",
		"rule_name": "考勤初稿：旷工标记核验",
		"rule_group": "考勤",
		"rule_type": "导入异常识别",
		"source_module": "钉钉每日统计",
		"source_document": "每日统计 / 旷工",
		"trigger_condition": "来源的无单位“旷工”标记大于 0。",
		"formula": "absence_marker_count > 0",
		"action_result": "按员工＋日期生成旷工待核验事件；来源无小时单位，不能直接用于薪资扣减。",
		"priority": 8,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "ATT-DRAFT-RESTDAY-CLOCK-WITHOUT-OVERTIME",
		"rule_name": "考勤初稿：休息日打卡未计加班",
		"rule_group": "考勤",
		"rule_type": "导入异常识别",
		"source_module": "钉钉每日统计",
		"source_document": "每日统计 / 日期类型、上班时间、下班时间、关联审批单、休息日加班（小时）",
		"trigger_condition": "日期类型为休息日，存在上班或下班打卡时间，未匹配加班申请，且休息日加班工时为 0。",
		"formula": "is_rest_day && (clock_in || clock_out) && !overtime_approval && restday_overtime_hours <= 0",
		"action_result": "置顶进入考勤初稿异常；由人事人工填写实际休息日加班工时或确认本次打卡不计加班。",
		"priority": 1,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "ATT-DRAFT-SHIFT-MISSING",
		"rule_name": "考勤初稿：班次缺失核验",
		"rule_group": "考勤",
		"rule_type": "数据质量",
		"source_module": "钉钉每日统计",
		"source_document": "每日统计 / 班次、员工入职日期、离职日期",
		"trigger_condition": "班次为空，且日期不在花名册已知的入职前或离职后期间。",
		"formula": "!shift && attendance_date >= date_of_joining && (!relieving_date || attendance_date <= relieving_date)",
		"action_result": "生成排班数据待核验；入职前、离职后空班次仅留为数据质量记录，不进员工异常。",
		"priority": 9,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "ATT-LATE-30",
		"rule_name": "迟到（无宽限）",
		"rule_group": "考勤",
		"rule_type": "异常判定",
		"source_module": "人资考勤",
		"source_document": "5.2人资考勤.xlsx / 人资考勤制度作业规范",
		"trigger_condition": "工作日实际上班时间晚于应上班时间，且迟到时长大于 0 分钟。",
		"formula": "late_minutes > 0",
		"action_result": "生成迟到待核验；无迟到宽限，需核对请假或主管说明后再处理。",
		"priority": 10,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "ATT-MISSING-CARD",
		"rule_name": "忘打卡",
		"rule_group": "考勤",
		"rule_type": "异常判定",
		"source_module": "人资考勤",
		"source_document": "5.2人资考勤.xlsx / 1.10忘打卡",
		"trigger_condition": "上班缺卡或下班缺卡，且无不可抗力或HR主管确认。",
		"formula": "missing_in || missing_out",
		"action_result": "生成异常，月底统计每次2个红苹果；员工需提交钉钉补卡。",
		"priority": 20,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "ATT-ABSENT-NO-LEAVE",
		"rule_name": "无有效请假旷工",
		"rule_group": "考勤",
		"rule_type": "异常判定",
		"source_module": "人资考勤",
		"source_document": "5.2人资考勤.xlsx / 1.6出勤异常",
		"trigger_condition": "工作日无打卡或有旷工小时，且没有审批通过、已结束的请假证据。",
		"formula": "standard_hours > 0 && !actual_in_time && valid_leave_hours <= 0",
		"action_result": "生成旷工异常，薪资侧按3倍旷工工时扣除。",
		"priority": 30,
		"application_mode": "异常提示",
	},
	{
		"rule_code": "APPLE-ATTENDANCE",
		"rule_name": "苹果树考勤奖惩",
		"rule_group": "苹果树",
		"rule_type": "奖惩汇总",
		"source_module": "苹果树",
		"source_document": "4.2苹果树.xlsx / 钉钉苹果树导出",
		"trigger_condition": "钉钉苹果树记录审批结果为审批通过且审批状态未终止。",
		"formula": "(green_apples - red_apples) * 5",
		"action_result": "并入月度考勤终稿的苹果树金额，作为薪资前置数据。",
		"priority": 40,
		"application_mode": "仅展示",
	},
	{
		"rule_code": "KPI-LATE-SUBMIT",
		"rule_name": "KPI资料延迟/错误红苹果",
		"rule_group": "KPI",
		"rule_type": "奖惩来源",
		"source_module": "绩效管理",
		"source_document": "4.4 KPI绩效管理.xlsx / KPI绩效管理办法",
		"trigger_condition": "每月15日前未准时、正确提交月会资料，或资料错误被退回。",
		"formula": "late_submit || data_error",
		"action_result": "罚红苹果1颗/项，作为苹果树来源记录进入月度汇总。",
		"priority": 50,
		"application_mode": "仅展示",
	},
	{
		"rule_code": "SEVENS-RECTIFY",
		"rule_name": "7S整改闭环",
		"rule_group": "7S",
		"rule_type": "稽核来源",
		"source_module": "绩效管理",
		"source_document": "4.3 7S.xlsx / 7S作业规范",
		"trigger_condition": "每月20日稽核后，部门需在次月5日前提交整改后图片并完成会签。",
		"formula": "audit_issue && !rectified_before_next_month_5",
		"action_result": "作为7S缺失项/未改善项记录，可进入苹果树或绩效扣分流程。",
		"priority": 60,
		"application_mode": "仅展示",
	},
]


def _get_file_content(file_url):
	name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not name:
		frappe.throw(_("未找到上传文件"))
	file_doc = frappe.get_doc("File", name)
	content = file_doc.get_content()
	if isinstance(content, str):
		content = content.encode()
	return content


def _load_workbook(file_url):
	from openpyxl import load_workbook

	return load_workbook(BytesIO(_get_file_content(file_url)), data_only=True, read_only=True)


def _sheet_by_required_name(workbook, required_name):
	for sheet_name in workbook.sheetnames:
		if sheet_name.strip() == required_name:
			return workbook[sheet_name]
	return None


def _cell_text(value):
	if value is None:
		return ""
	if isinstance(value, datetime):
		return value.strftime("%Y-%m-%d %H:%M:%S")
	if isinstance(value, date):
		return value.isoformat()
	return str(value).strip()


def _normalise_header(value):
	return re.sub(r"\s+", "", _cell_text(value).replace("\n", ""))


def _normalise_dingtalk_header(value):
	return _normalise_header(value).replace("（", "(").replace("）", ")").replace("／", "/")


def _read_sheet_rows(sheet, max_rows=None):
	rows = []
	for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
		values = [_cell_text(value) for value in row]
		if any(values):
			rows.append(values)
		if max_rows and index >= max_rows:
			break
	return rows


def _find_header_index(rows, required_headers):
	required = {_normalise_header(header) for header in required_headers}
	best_index = 0
	best_score = -1
	for index, row in enumerate(rows[:20]):
		headers = {_normalise_header(value) for value in row if value}
		score = len(required & headers)
		if score > best_score:
			best_index = index
			best_score = score
	return best_index


def _rows_as_dicts(sheet, required_headers):
	rows = _read_sheet_rows(sheet)
	if not rows:
		return []
	header_index = _find_header_index(rows, required_headers)
	headers = []
	seen = defaultdict(int)
	for value in rows[header_index]:
		header = _normalise_header(value)
		if not header:
			headers.append("")
			continue
		seen[header] += 1
		headers.append(f"{header}_{seen[header]}" if seen[header] > 1 else header)

	items = []
	for row in rows[header_index + 1 :]:
		item = {}
		for index, header in enumerate(headers):
			if header:
				item[header] = row[index] if index < len(row) else ""
		if any(item.values()):
			items.append(item)
	return items


def _first_value(row, *headers):
	for header in headers:
		value = row.get(_normalise_header(header))
		if value not in (None, ""):
			return value
	return ""


def _float_value(row, *headers):
	return flt(_first_value(row, *headers))


def _int_value(row, *headers):
	return int(flt(_first_value(row, *headers)))


def _duration_hours(value):
	text = _cell_text(value)
	if not text:
		return 0
	match = re.search(r"(-?\d+(?:\.\d+)?)\s*(小时|H|h|天|半天)?", text)
	if not match:
		return flt(text)
	number = flt(match.group(1))
	unit = match.group(2) or "小时"
	if unit == "天":
		return number * STANDARD_DAY_HOURS
	if unit == "半天":
		return number * 4
	return number


def _parse_date(value):
	if not value:
		return None
	if isinstance(value, (date, datetime)):
		return getdate(value)
	text = _cell_text(value)
	match = re.search(r"(\d{2,4})[-/年.](\d{1,2})[-/月.](\d{1,2})", text)
	if match:
		year = int(match.group(1))
		if year < 100:
			year += 2000
		return date(year, int(match.group(2)), int(match.group(3)))
	try:
		return getdate(text)
	except Exception:
		return None


def _parse_datetime(value):
	if not value:
		return None
	if isinstance(value, datetime):
		return value
	if isinstance(value, date):
		return datetime.combine(value, time.min)
	text = _cell_text(value)
	for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
		try:
			return datetime.strptime(text, fmt)
		except ValueError:
			continue
	try:
		parsed_date = _parse_date(text)
		return datetime.combine(parsed_date, time.min) if parsed_date else None
	except Exception:
		return None


def _parse_shift_start_time(shift_name):
	text = _cell_text(shift_name)
	match = re.search(r"(\d{1,2}):(\d{2})\s*[-~至]", text)
	if match:
		return f"{int(match.group(1)):02d}:{match.group(2)}"
	match = re.search(r"(\d{1,2}):(\d{2})", text)
	if match:
		return f"{int(match.group(1)):02d}:{match.group(2)}"
	return ""


def _time_text_to_minutes(value):
	text = _cell_text(value)
	match = re.search(r"(\d{1,2}):(\d{2})", text)
	if not match:
		return None
	return int(match.group(1)) * 60 + int(match.group(2))


def _is_valid_approval(row):
	result = _first_value(row, "审批结果")
	status = _first_value(row, "审批状态")
	if not result and not status:
		return 0
	return 1 if result == "审批通过" and status == "已结束" else 0


def _month_bounds(attendance_month):
	month = (attendance_month or "").strip()
	if not re.match(r"^\d{4}-\d{2}$", month):
		frappe.throw(_("考勤月份格式应为 YYYY-MM"))
	start = getdate(f"{month}-01")
	if start.month == 12:
		end = date(start.year + 1, 1, 1)
	else:
		end = date(start.year, start.month + 1, 1)
	return start, end


def _require_company(company):
	company = (company or "").strip()
	if not company:
		frappe.throw(_("请先选择公司。"))
	if not frappe.db.exists("Company", company):
		frappe.throw(_("公司 {0} 不存在").format(company))
	return company


def _attendance_scope_filters(company, attendance_month, attendance_lock_version):
	company = (company or "").strip()
	attendance_month = (attendance_month or "").strip()
	attendance_lock_version = str(attendance_lock_version or "").strip()
	if not company or not attendance_month or not attendance_lock_version:
		frappe.throw(_("考勤范围必须包含公司、月份和锁定版本。"))
	return {
		"company": company,
		"attendance_month": attendance_month,
		"attendance_lock_version": attendance_lock_version,
	}


def _source_file_checksum(file_url):
	return hashlib.sha256(_get_file_content(file_url)).hexdigest()


def _employee_lookup(employee_code=None, employee_name=None):
	if employee_code:
		for fieldname in ("custom_employee_code",):
			try:
				name = frappe.db.get_value("Employee", {fieldname: employee_code}, "name")
			except Exception:
				name = None
			if name:
				return name
	if employee_name:
		return frappe.db.get_value("Employee", {"employee_name": employee_name}, "name")
	return None


def _employee_matches_company(employee, company):
	return bool(employee and frappe.db.get_value("Employee", employee, "company") == company)


def _department_lookup(department):
	department = (department or "").strip()
	if not department:
		return None
	if frappe.db.exists("Department", department):
		return department
	return frappe.db.get_value("Department", {"department_name": department}, "name")


def _preview_sheet(workbook, sheet_name):
	sheet = _sheet_by_required_name(workbook, sheet_name)
	if not sheet:
		return {"sheet_name": sheet_name, "found": False, "row_count": 0, "headers": []}
	rows = _read_sheet_rows(sheet, max_rows=12)
	header_index = _find_header_index(rows, ["姓名", "工号", "日期"] if sheet_name != "1.3苹果树" else ["受奖/惩人", "奖/惩日期"])
	return {
		"sheet_name": sheet_name,
		"found": True,
		"row_count": max(sheet.max_row - header_index - 1, 0),
		"headers": rows[header_index][:30] if rows else [],
	}


def _is_dingtalk_export_v1(workbook):
	return all(_sheet_by_required_name(workbook, sheet_name) for sheet_name in DINGTALK_EXPORT_V1_SHEETS)


def _flatten_dingtalk_headers(sheet, header_rows=(3, 4)):
	start_row, end_row = header_rows
	rows = list(sheet.iter_rows(min_row=start_row, max_row=end_row, values_only=True))
	if len(rows) != 2:
		return []

	parents, children = rows
	headers = []
	active_parent = ""
	seen = defaultdict(int)
	for parent, child in zip(parents, children):
		parent_text = _normalise_dingtalk_header(parent)
		child_text = _normalise_dingtalk_header(child)
		if parent_text:
			active_parent = parent_text
		header = f"{active_parent}/{child_text}" if child_text and active_parent else (child_text or active_parent)
		seen[header] += 1
		headers.append(f"{header}_{seen[header]}" if header and seen[header] > 1 else header)
	return headers


def _flatten_dingtalk_daily_headers(sheet):
	return _flatten_dingtalk_headers(sheet)


def _daily_rows_from_header_rows(sheet, header_rows=(3, 4), data_start_row=5):
	headers = _flatten_dingtalk_headers(sheet, header_rows)
	items = []
	for row_index, values in enumerate(sheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
		item = {
			header: _cell_text(values[index]) if index < len(values) else ""
			for index, header in enumerate(headers)
			if header
		}
		if any(item.values()):
			item["_source_row"] = row_index
			items.append(item)
	return items


def _dingtalk_daily_rows(sheet):
	return _daily_rows_from_header_rows(sheet)


def _dingtalk_export_import_source_kind():
	return "钉钉原始导出"


def _count_nonempty_rows(sheet, start_row):
	return sum(1 for row in sheet.iter_rows(min_row=start_row, values_only=True) if any(_cell_text(value) for value in row))


def _dingtalk_sheet_preview(sheet_name, sheet, row_count, headers):
	return {
		"sheet_name": sheet_name,
		"found": bool(sheet),
		"row_count": row_count,
		"headers": headers[:40],
	}


def _preview_dingtalk_export_v1(workbook):
	daily_sheet = _sheet_by_required_name(workbook, "每日统计")
	raw_sheet = _sheet_by_required_name(workbook, "原始记录")
	monthly_sheet = _sheet_by_required_name(workbook, "月度汇总")
	clock_sheet = _sheet_by_required_name(workbook, "打卡时间")
	daily_rows = _dingtalk_daily_rows(daily_sheet)
	daily_headers = _flatten_dingtalk_daily_headers(daily_sheet)

	quality_counts = {
		"missing_employee_code": 0,
		"missing_attendance_group": 0,
		"planned_hours_without_actual": 0,
		"duplicate_userid_workdate": 0,
	}

	seen_user_dates = set()
	for row in daily_rows:
		if not row.get("工号"):
			quality_counts["missing_employee_code"] += 1
		if row.get("考勤组") in ("", "未加入考勤组"):
			quality_counts["missing_attendance_group"] += 1
		if row.get("标准工时") and not row.get("实际出勤(小时)"):
			quality_counts["planned_hours_without_actual"] += 1
		user_date = (row.get("UserId", ""), row.get("workDate") or row.get("日期", ""))
		if all(user_date):
			if user_date in seen_user_dates:
				quality_counts["duplicate_userid_workdate"] += 1
			else:
				seen_user_dates.add(user_date)

	raw_headers = _read_sheet_rows(raw_sheet, max_rows=DINGTALK_EXPORT_V1_SCHEMA["raw_header_row"])[-1]
	monthly_headers = _flatten_dingtalk_daily_headers(monthly_sheet)
	clock_headers = _flatten_dingtalk_daily_headers(clock_sheet)
	sheets = [
		_dingtalk_sheet_preview("每日统计", daily_sheet, len(daily_rows), daily_headers),
		_dingtalk_sheet_preview(
			"打卡时间",
			clock_sheet,
			_count_nonempty_rows(clock_sheet, DINGTALK_EXPORT_V1_SCHEMA["daily_data_start_row"]),
			clock_headers,
		),
		_dingtalk_sheet_preview(
			"原始记录",
			raw_sheet,
			_count_nonempty_rows(raw_sheet, DINGTALK_EXPORT_V1_SCHEMA["raw_data_start_row"]),
			raw_headers,
		),
		_dingtalk_sheet_preview(
			"月度汇总",
			monthly_sheet,
			_count_nonempty_rows(monthly_sheet, DINGTALK_EXPORT_V1_SCHEMA["monthly_data_start_row"]),
			monthly_headers,
		),
	]
	return _with_import_validation({
		"source_type": DINGTALK_EXPORT_V1_SCHEMA["source_type"],
		"required_sheets": DINGTALK_EXPORT_V1_SHEETS,
		"sheets": sheets,
		"missing_sheets": [sheet["sheet_name"] for sheet in sheets if not sheet["found"]],
		"record_counts": {
			"daily_statistics": len(daily_rows),
			"raw_records": sheets[2]["row_count"],
			"monthly_people": sheets[3]["row_count"],
		},
		"field_mapping": {
			header: DINGTALK_DAILY_FIELD_MAPPING.get(header, "source_only") for header in daily_headers if header
		},
		"quality_warnings": [
			{"code": "missing_employee_code", "label": "缺工号", "count": quality_counts["missing_employee_code"]},
			{"code": "missing_attendance_group", "label": "未加入考勤组", "count": quality_counts["missing_attendance_group"]},
			{
				"code": "planned_hours_without_actual",
				"label": "标准工时有值但实际出勤为空",
				"count": quality_counts["planned_hours_without_actual"],
			},
			{
				"code": "duplicate_userid_workdate",
				"label": "重复 UserId + workDate",
				"count": quality_counts["duplicate_userid_workdate"],
			},
		],
		"parsing": {
			"daily_header_rows": list(DINGTALK_EXPORT_V1_SCHEMA["daily_header_rows"]),
			"daily_data_start_row": DINGTALK_EXPORT_V1_SCHEMA["daily_data_start_row"],
		},
		"database_writes": 0,
	})


def _is_company_attendance_workbook(workbook):
	return all(_sheet_by_required_name(workbook, source["sheet_name"]) for source in COMPANY_ATTENDANCE_WORKBOOK_SOURCES.values())


def _is_company_attendance_register_v1(workbook):
	return all(_sheet_by_required_name(workbook, source["sheet_name"]) for source in COMPANY_ATTENDANCE_REGISTER_V1_SHEETS.values())


def _company_register_sheet_preview(workbook, source_key, source):
	sheet = _sheet_by_required_name(workbook, source["sheet_name"])
	if not sheet:
		return {"source_key": source_key, "sheet_name": source["sheet_name"], "found": False, "row_count": 0, "headers": []}
	if source_key == "daily_statistics":
		headers = _flatten_dingtalk_headers(sheet, source["header_rows"])
		row_count = len(_daily_rows_from_header_rows(sheet, source["header_rows"], source["data_start_row"]))
	else:
		rows = _read_sheet_rows(sheet, max_rows=max(source["header_rows"]))
		headers = rows[-1] if rows else []
		row_count = _count_nonempty_rows(sheet, source["data_start_row"])
	return {
		"source_key": source_key,
		"sheet_name": source["sheet_name"],
		"found": True,
		"row_count": row_count,
		"headers": headers[:40],
		"writes_daily_checks": source["writes_daily_checks"],
		"import_behavior": {
			"daily_statistics": "写入每日考勤核对",
			"attendance_detail": "保留为部门日报核对来源，不写入个人日考勤",
			"attendance_exception": "保留为异常核对来源，不自动生成处理结论",
			"apple_tree": "写入苹果树奖惩来源，待月度核对后使用",
		}.get(source_key, "仅留存来源说明"),
	}


def _preview_company_attendance_register_v1(workbook):
	sheets = [
		_company_register_sheet_preview(workbook, source_key, source)
		for source_key, source in COMPANY_ATTENDANCE_REGISTER_V1_SHEETS.items()
	]
	daily_sheet = _sheet_by_required_name(workbook, "每日统计")
	daily_headers = _flatten_dingtalk_headers(daily_sheet, COMPANY_ATTENDANCE_REGISTER_V1_SHEETS["daily_statistics"]["header_rows"])
	return _with_import_validation({
		"source_type": "company_attendance_register_v1",
		"required_sheets": [source["sheet_name"] for source in COMPANY_ATTENDANCE_REGISTER_V1_SHEETS.values()],
		"sheets": sheets,
		"missing_sheets": [sheet["sheet_name"] for sheet in sheets if not sheet["found"]],
		"field_mapping": {header: DINGTALK_DAILY_FIELD_MAPPING.get(header, "source_only") for header in daily_headers if header},
		"template_key": "company_attendance_register_v1",
		"database_writes": 0,
	})


def _preview_company_daily_source(workbook, source_kind, source):
	sheet = _sheet_by_required_name(workbook, source["sheet_name"])
	headers = _flatten_dingtalk_headers(sheet, source["header_rows"])
	rows = _daily_rows_from_header_rows(sheet, source["header_rows"], source["data_start_row"])
	return {
		"source_kind": source_kind,
		"sheet_name": source["sheet_name"],
		"found": bool(sheet),
		"header_rows": list(source["header_rows"]),
		"data_start_row": source["data_start_row"],
		"row_count": len(rows),
		"headers": headers,
		"field_mapping": {header: DINGTALK_DAILY_FIELD_MAPPING.get(header, "source_only") for header in headers if header},
	}


def _preview_company_attendance_workbook(workbook):
	daily_sources = {
		source_kind: _preview_company_daily_source(workbook, source_kind, source)
		for source_kind, source in COMPANY_ATTENDANCE_WORKBOOK_SOURCES.items()
	}
	reference_sheet_names = ["请假单（钉钉导出）", "出勤异常", "苹果树（钉钉导出）", "苹果树（修改后）", "考勤初稿", "考勤终稿（签字版）", "考勤终稿（财务版）"]
	reference_sheets = [
		{
			"sheet_name": sheet_name,
			"found": bool(_sheet_by_required_name(workbook, sheet_name)),
		}
		for sheet_name in reference_sheet_names
	]
	preview = {
		"source_type": "company_attendance_workbook_v1",
		"daily_sources": daily_sources,
		"reference_sheets": reference_sheets,
		"database_writes": 0,
	}
	# Both daily sheets use the same company field contract. The raw export is
	# preferred for validation because it retains UserId and source evidence.
	preview["field_mapping"] = daily_sources.get("dingtalk_raw", {}).get("field_mapping", {})
	return _with_import_validation(preview)


@frappe.whitelist()
def preview_attendance_workbook(file_url: str, company: str = ""):
	# Company is accepted from the workbench so preview and import share one API contract.
	# Preview stays read-only and therefore does not require or write a company value.
	workbook = _load_workbook(file_url)
	if _is_company_attendance_register_v1(workbook):
		return _preview_company_attendance_register_v1(workbook)
	if _is_company_attendance_workbook(workbook):
		return _preview_company_attendance_workbook(workbook)
	if _is_dingtalk_export_v1(workbook):
		return _preview_dingtalk_export_v1(workbook)
	sheets = [_preview_sheet(workbook, sheet_name) for sheet_name in REQUIRED_ATTENDANCE_SHEETS]
	legacy_headers = sheets[0].get("headers", []) if sheets else []
	return _with_import_validation({
		"required_sheets": REQUIRED_ATTENDANCE_SHEETS,
		"sheets": sheets,
		"missing_sheets": [sheet["sheet_name"] for sheet in sheets if not sheet["found"]],
		"field_mapping": {header: DINGTALK_DAILY_FIELD_MAPPING.get(header, "source_only") for header in legacy_headers if header},
		"database_writes": 0,
	})


def _attendance_template_or_throw(template_key):
	for template in ATTENDANCE_IMPORT_TEMPLATES:
		if template["key"] == template_key:
			return template
	frappe.throw(_("未找到考勤导入模板：{0}").format(template_key))


@frappe.whitelist()
def list_attendance_import_templates():
	return [dict(template) for template in ATTENDANCE_IMPORT_TEMPLATES]


def _style_template_headers(sheet, header_rows, widths=None):
	from openpyxl.styles import Alignment, Font, PatternFill

	fill = PatternFill("solid", fgColor="D9EAD3")
	font = Font(bold=True, color="1F2937")
	for row in header_rows:
		for cell in sheet[row]:
			if cell.value not in (None, ""):
				cell.fill = fill
				cell.font = font
				cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	for index, width in enumerate(widths or [], start=1):
		sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width


def _add_daily_statistics_template_sheet(workbook, sheet_name="每日统计"):
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	sheet = workbook.create_sheet(sheet_name)
	sheet.append([parent for parent, _child in COMPANY_DAILY_STATISTICS_COLUMNS])
	sheet.append([child for _parent, child in COMPANY_DAILY_STATISTICS_COLUMNS])

	# Match the company register: fixed fields are vertically merged, while the
	# two leave sections retain their original grouped headers for direct upload.
	for column, (_parent, child) in enumerate(COMPANY_DAILY_STATISTICS_COLUMNS, start=1):
		if not child:
			sheet.merge_cells(start_row=1, start_column=column, end_row=2, end_column=column)
	sheet.merge_cells("S1:AC1")
	sheet.merge_cells("AD1:AH1")

	yellow = PatternFill("solid", fgColor="FFF2CC")
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(COMPANY_DAILY_STATISTICS_COLUMNS)):
		for cell in row:
			cell.fill = yellow
			cell.font = Font(name="新宋体", size=12, bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.border = border

	for index, width in enumerate([15] + [13] * 38, start=1):
		sheet.column_dimensions[get_column_letter(index)].width = width
	sheet.row_dimensions[1].height = 51.2
	sheet.row_dimensions[2].height = 51.2
	sheet.freeze_panes = "A3"
	sheet.auto_filter.ref = f"A2:{get_column_letter(len(COMPANY_DAILY_STATISTICS_COLUMNS))}1000"
	return sheet


def _add_attendance_detail_template_sheet(workbook):
	sheet = workbook.create_sheet("出勤明细")
	sheet.append(["部门", "统计日期", "", "", "", "", "备注", "对比日期", ""])
	sheet.append(["", "现有人数", "出勤人数", "请假人数", "请假人员", "请假说明", "", "现有人数", "人员变动说明"])
	_style_template_headers(sheet, [1, 2], [18, 14, 14, 14, 16, 34, 34, 14, 34])
	sheet.freeze_panes = "A3"
	return sheet


def _add_attendance_exception_template_sheet(workbook):
	sheet = workbook.create_sheet("出勤异常")
	sheet.append(["出勤异常：迟到、早退、旷工、未打卡、后补假卡"])
	sheet.append(["序号", "姓名", "工号", "出勤日期", "单位", "应上班时间", "实际上班时间", "实际下班时间", "异常类型", "处理方式", "备注", "签字确认"])
	_style_template_headers(sheet, [1, 2], [10, 14, 14, 15, 18, 30, 16, 16, 16, 28, 34, 18])
	sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
	sheet.freeze_panes = "A3"
	sheet.auto_filter.ref = "A2:L3"
	return sheet


def _add_apple_reward_template_sheet(workbook):
	sheet = workbook.create_sheet("苹果树")
	# Preserve the source workbook's third-row header convention so users can
	# replace this sheet in the company workbook without shifting data rows.
	sheet.append([])
	sheet.append([])
	sheet.append(["序号", "创建时间", "奖/惩日期", "部门", "受奖/惩人", "工号", "绿苹果", "红苹果", "奖/惩项目", "备注", "创建人", "审批编号"])
	_style_template_headers(sheet, [3], [10, 20, 16, 18, 16, 14, 12, 12, 48, 36, 16, 20])
	sheet.freeze_panes = "A4"
	sheet.auto_filter.ref = "A3:L4"
	return sheet


def _add_legacy_template_sheets(workbook):
	daily = workbook.create_sheet("1.1每日统计")
	daily.append(["姓名", "工号", "日期", "部门", "班次", "上班时间", "下班时间", "标准工时", "实际出勤(小时)", "工作日加班(小时)", "休息日加班(小时)", "节假日加班(小时)"])
	_style_template_headers(daily, [1], [16] * 12)
	daily.freeze_panes = "A2"
	leave = workbook.create_sheet("1.2请假单")
	leave.append(["姓名", "工号", "部门", "请假类型", "开始时间", "结束时间", "时长", "请假事由", "审批编号", "审批结果", "审批状态"])
	_style_template_headers(leave, [1], [16, 14, 18, 16, 20, 20, 12, 36, 20, 16, 16])
	leave.freeze_panes = "A2"
	apple = workbook.create_sheet("1.3苹果树")
	apple.append(["奖/惩日期", "部门", "受奖/惩人", "工号", "绿苹果", "红苹果", "奖/惩项目", "备注", "创建人", "审批编号", "审批结果", "审批状态"])
	_style_template_headers(apple, [1], [16, 18, 16, 14, 12, 12, 42, 34, 16, 20, 16, 16])
	apple.freeze_panes = "A2"


def _add_template_instructions(workbook, template):
	from openpyxl.styles import Alignment, Font, PatternFill

	sheet = workbook.create_sheet("填写说明", 0)
	sheet.append(["模板名称", template["label"]])
	sheet.append(["上传方式", "可直接上传整套工作簿" if template["upload_mode"] == "whole_workbook" else "此为单表模板，请替换进“公司考勤工作簿（推荐）”后再上传"])
	sheet.append(["工作表", "、".join(template["sheet_names"])])
	sheet.append(["系统处理", template["writes"]])
	sheet.append(["填写规则", "工号、姓名、部门须与当前公司员工及组织主数据一致；日期使用 YYYY-MM-DD；工时使用数字。"])
	sheet.append(["数据安全", "上传先预览、再确认导入；月度锁定后只能通过更正批次修改。"])
	for cell in sheet[1]:
		cell.fill = PatternFill("solid", fgColor="D9EAD3")
		cell.font = Font(bold=True)
	for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
		for cell in row:
			cell.alignment = Alignment(vertical="top", wrap_text=True)
	sheet.column_dimensions["A"].width = 18
	sheet.column_dimensions["B"].width = 100


@frappe.whitelist()
def create_attendance_import_template_file(template_key: str):
	from openpyxl import Workbook
	from hrms.utils.export_watermark import save_workbook_with_logo_watermark

	template = _attendance_template_or_throw(template_key)
	workbook = Workbook()
	workbook.remove(workbook.active)
	_add_template_instructions(workbook, template)

	if template_key == "company_attendance_register_v1":
		_add_daily_statistics_template_sheet(workbook)
		_add_attendance_detail_template_sheet(workbook)
		_add_attendance_exception_template_sheet(workbook)
		_add_apple_reward_template_sheet(workbook)
	elif template_key == "company_daily_statistics_v1":
		_add_daily_statistics_template_sheet(workbook)
	elif template_key == "attendance_detail_v1":
		_add_attendance_detail_template_sheet(workbook)
	elif template_key == "attendance_exception_v1":
		_add_attendance_exception_template_sheet(workbook)
	elif template_key == "apple_reward_v1":
		_add_apple_reward_template_sheet(workbook)
	elif template_key == "legacy_workbook_v1":
		_add_legacy_template_sheets(workbook)

	output = BytesIO()
	save_workbook_with_logo_watermark(workbook, output)
	filename = f"{template['label']}导入模板.xlsx"
	file_doc = frappe.get_doc({"doctype": "File", "file_name": filename, "content": output.getvalue(), "is_private": 0}).insert(ignore_permissions=True)
	return {"file_url": file_doc.file_url, "file_name": filename, "template_key": template_key}


def _attendance_export_profile_or_throw(export_profile):
	profile = ATTENDANCE_EXPORT_PROFILES.get(export_profile)
	if not profile:
		frappe.throw(_("未找到考勤导出表单：{0}").format(export_profile))
	return profile


def _export_number(value):
	value = flt(value)
	return "" if value == 0 else value


def _export_leave_days(value):
	value = flt(value)
	return "" if value == 0 else round(value / STANDARD_DAY_HOURS, 2)


def _export_attendance_date(value):
	if not value:
		return ""
	attendance_date = getdate(value)
	return f"{attendance_date.strftime('%y-%m-%d')} 星期{'一二三四五六日'[attendance_date.weekday()]}"


def _apply_export_header_style(sheet, header_rows, widths, fill_color="FFF2CC"):
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	fill = PatternFill("solid", fgColor=fill_color)
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for row in sheet.iter_rows(min_row=min(header_rows), max_row=max(header_rows), min_col=1, max_col=len(widths)):
		for cell in row:
			cell.fill = fill
			cell.font = Font(name="新宋体", size=11, bold=True)
			cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
			cell.border = border
	for index, width in enumerate(widths, start=1):
		sheet.column_dimensions[get_column_letter(index)].width = width


def _append_daily_statistics_export_rows(sheet, rows):
	for row in rows:
		sheet.append(
			[
				getattr(row, "employee_name", ""), getattr(row, "employee_code", ""), _export_attendance_date(getattr(row, "attendance_date", "")),
				getattr(row, "department", ""), getattr(row, "shift_name", ""), getattr(row, "actual_in_time", ""), getattr(row, "actual_out_time", ""),
				"缺卡" if getattr(row, "missing_in", 0) else "", "缺卡" if getattr(row, "missing_out", 0) else "", _export_number(getattr(row, "absent_hours", 0)),
				_export_number(getattr(row, "standard_hours", 0)), _export_number(getattr(row, "actual_attendance_hours", 0)), getattr(row, "leave_summary", ""),
				_export_number(getattr(row, "workday_overtime_hours", 0)), _export_number(getattr(row, "restday_overtime_hours", 0)), _export_number(getattr(row, "holiday_overtime_hours", 0)),
				_export_number(getattr(row, "large_night_shift_count", 0)), _export_number(getattr(row, "small_night_shift_count", 0)),
				_export_number(getattr(row, "personal_leave_hours", 0)), _export_number(getattr(row, "sick_leave_hours", 0)), _export_leave_days(getattr(row, "marriage_leave_hours", 0)),
				_export_number(getattr(row, "annual_leave_hours", 0)), _export_number(getattr(row, "bereavement_leave_hours", 0)), _export_number(getattr(row, "work_injury_leave_hours", 0)),
				_export_leave_days(getattr(row, "public_leave_hours", 0)), _export_leave_days(getattr(row, "maternity_leave_hours", 0)), _export_leave_days(getattr(row, "reunion_leave_hours", 0)),
				_export_number(getattr(row, "rest_leave_hours", 0)), _export_number(getattr(row, "absent_hours", 0)),
				_export_leave_days(getattr(row, "marriage_leave_hours", 0)), _export_leave_days(getattr(row, "bereavement_leave_hours", 0)), _export_leave_days(getattr(row, "public_leave_hours", 0)),
				_export_leave_days(getattr(row, "maternity_leave_hours", 0)), _export_leave_days(getattr(row, "reunion_leave_hours", 0)), _export_number(getattr(row, "absent_hours", 0)),
				_export_number(getattr(row, "missing_in", 0)), _export_number(getattr(row, "missing_out", 0)), _export_number(getattr(row, "late_count", 0)), _export_number(getattr(row, "early_count", 0)),
			]
		)


def _add_attendance_detail_export_sheet(workbook, rows, attendance_month):
	from openpyxl.styles import Alignment, Border, Side

	sheet = workbook.create_sheet("出勤明细")
	month_label = attendance_month.replace("-", "")
	sheet.append(["部门", month_label, "", "", "", "", "备注", "对比日期", "", "", "", ""])
	sheet.append(["", "现有人数", "出勤人数", "请假人数", "请假人员", "请假说明", "", "现有人数", "出勤人数", "请假人数", "请假人员", "请假说明"])
	sheet.merge_cells("A1:A2")
	sheet.merge_cells("B1:G1")
	sheet.merge_cells("H1:L1")
	_apply_export_header_style(sheet, [1, 2], [18, 13, 13, 13, 20, 38, 36, 13, 13, 13, 20, 38])
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	by_department = defaultdict(list)
	for row in rows:
		by_department[getattr(row, "department", "") or _("未分配部门")].append(row)
	for department, members in sorted(by_department.items()):
		attending = sum(1 for row in members if flt(getattr(row, "actual_attendance_hours", 0)) > 0)
		leaving = [row for row in members if flt(getattr(row, "leave_hours", 0)) > 0]
		sheet.append([department, len(members), attending, len(leaving), "、".join(getattr(row, "employee_name", "") for row in leaving), "；".join(getattr(row, "leave_summary", "") for row in leaving if getattr(row, "leave_summary", "")), "", len(members), attending, len(leaving), "", ""])
		for cell in sheet[sheet.max_row]:
			cell.border = border
			cell.alignment = Alignment(vertical="top", wrap_text=True)
	sheet.freeze_panes = "A3"
	return sheet


def _add_leave_evidence_export_sheet(workbook, rows):
	from openpyxl.styles import Alignment, Border, Side

	headers = [
		"序号", "数据id", "请假类型（实际）", "请假说明确认", "请假类型", "开始时间", "结束时间", "时长", "逝者关系", "死亡证明", "结婚证明",
		"结婚证、出生证、准生证", "亲属关系证明", "病假证明", "请假事由", "图片", "审批编号", "创建时间", "创建人", "当前负责人", "审批结果", "审批状态",
		"更新时间", "完成时间", "创建人部门", "审批单标题", "历史审批人", "耗时(时:分:秒)", "审批记录",
	]
	sheet = workbook.create_sheet("请假单")
	sheet.append(headers)
	_apply_export_header_style(sheet, [1], [10, 20, 16, 18, 16, 20, 20, 12, 16, 16, 16, 22, 20, 24, 36, 20, 20, 20, 16, 16, 16, 16, 20, 20, 18, 32, 24, 18, 50])
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for index, row in enumerate(rows, start=1):
		duration = _export_number(getattr(row, "leave_hours", 0))
		sheet.append([
			index, "", getattr(row, "leave_type", ""), "", getattr(row, "leave_type", ""), getattr(row, "leave_start", ""), getattr(row, "leave_end", ""),
			f"{duration:g}小时" if isinstance(duration, (int, float)) else "", "", "", "", "", "", "", getattr(row, "leave_reason", ""), "", getattr(row, "approval_no", ""),
			getattr(row, "creation", ""), getattr(row, "employee_name", ""), "", getattr(row, "approval_result", ""), getattr(row, "approval_status", ""), "", "", getattr(row, "department", ""), "", "", "", "",
		])
		for cell in sheet[sheet.max_row]:
			cell.border = border
			cell.alignment = Alignment(vertical="top", wrap_text=True)
	sheet.freeze_panes = "A2"
	return sheet


def _add_attendance_exception_export_sheet(workbook, rows):
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

	sheet = workbook.create_sheet("出勤异常")
	sheet.append(["", "出勤异常：迟到、早退、旷工、未打卡、后补假卡"])
	sheet.append(["", "序号", "姓名", "工号", "出勤日期", "单位", "应上班时间", "实际上班时间", "实际下班时间", "异常类型", "处理方式", "备注", "签字确认"])
	sheet.merge_cells("B1:M1")
	_apply_export_header_style(sheet, [1, 2], [4, 10, 14, 14, 17, 18, 30, 17, 17, 18, 28, 36, 18])
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for index, row in enumerate(rows, start=1):
		sheet.append(["", index, getattr(row, "employee_name", ""), getattr(row, "employee_code", ""), _export_attendance_date(getattr(row, "attendance_date", "")), getattr(row, "department", ""), getattr(row, "expected_shift", ""), getattr(row, "actual_in_time", ""), getattr(row, "actual_out_time", ""), getattr(row, "exception_type", ""), getattr(row, "handling_method", ""), getattr(row, "remarks", ""), getattr(row, "confirmed_by", "")])
		for cell in sheet[sheet.max_row]:
			cell.border = border
			cell.alignment = Alignment(vertical="top", wrap_text=True)
	sheet.freeze_panes = "B3"
	return sheet


def _add_missing_card_export_sheet(workbook, rows, attendance_month):
	from openpyxl.styles import Alignment, Border, Font, Side

	sheet = workbook.create_sheet("忘打卡")
	sheet.append([])
	sheet.append([f"{attendance_month.replace('-', '')}忘打卡名单"])
	sheet.merge_cells("B2:J2")
	sheet["B2"].font = Font(name="新宋体", size=14, bold=True)
	sheet["B2"].alignment = Alignment(horizontal="center")
	sheet.append(["", "序号", "部门", "创建时间", "补卡时间", "补卡类型", "补卡理由", "创建人", "签名", "备注"])
	_apply_export_header_style(sheet, [3], [4, 10, 18, 20, 20, 18, 30, 16, 18, 36])
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	missing_rows = [row for row in rows if getattr(row, "exception_type", "") == "忘打卡"]
	for index, row in enumerate(missing_rows, start=1):
		sheet.append(["", index, getattr(row, "department", ""), getattr(row, "creation", ""), _export_attendance_date(getattr(row, "attendance_date", "")), "忘刷卡补卡", getattr(row, "remarks", "") or "忘打卡", getattr(row, "employee_name", ""), getattr(row, "confirmed_by", ""), getattr(row, "handling_method", "")])
		for cell in sheet[sheet.max_row]:
			cell.border = border
			cell.alignment = Alignment(vertical="top", wrap_text=True)
	sheet.freeze_panes = "B4"
	return sheet


def _add_apple_reward_export_sheet(workbook, rows):
	from openpyxl.styles import Alignment, Border, Side

	sheet = workbook.create_sheet("苹果树")
	sheet.append([])
	sheet.append([])
	sheet.append(["", "序号", "创建时间", "奖/惩日期", "部门", "受奖/惩人", "绿苹果", "红苹果", "奖/惩项目", "备注", "创建人", "审批编号"])
	_apply_export_header_style(sheet, [3], [4, 10, 20, 16, 18, 16, 12, 12, 48, 36, 16, 20])
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for index, row in enumerate(rows, start=1):
		sheet.append(["", index, getattr(row, "creation", ""), getattr(row, "reward_date", ""), getattr(row, "department", ""), getattr(row, "employee_name", ""), _export_number(getattr(row, "green_apples", 0)), _export_number(getattr(row, "red_apples", 0)), getattr(row, "reward_item", ""), "", getattr(row, "created_by_name", ""), getattr(row, "approval_no", "")])
		for cell in sheet[sheet.max_row]:
			cell.border = border
			cell.alignment = Alignment(vertical="top", wrap_text=True)
	sheet.freeze_panes = "B4"
	return sheet


def _monthly_export_columns(profile_key):
	base = [
		("序号", "sequence"), ("部门", "department"), ("姓名", "employee_name"), ("工号", "employee_code"), ("入职时间", "date_of_joining"),
		("标准工时（小时）", "standard_hours"), ("钉钉导出实际出勤（小时）", "actual_attendance_hours"), ("1.5倍加班（小时）", "overtime_1_5_hours"),
		("2倍加班（小时）", "overtime_2_hours"), ("3倍加班（小时）", "overtime_3_hours"), ("大夜班", "large_night_shift_count"), ("小夜班", "small_night_shift_count"),
		("事假（小时）", "personal_leave_hours"), ("病假（小时）", "sick_leave_hours"), ("特休（小时）", "annual_leave_hours"), ("工伤（小时）", "work_injury_leave_hours"),
		("排休（小时）", "rest_leave_hours"), ("旷工（小时）", "absent_hours"),
	]
	if profile_key == "monthly_draft":
		return base + [("签名", ""), ("备注", "")]
	return base + [
		("实际打卡出勤A（验算）", "actual_clock_attendance_hours"), ("应补1倍工时B", "paid_leave_makeup_hours"), ("应扣2倍工时F", "leave_deductible_hours"),
		("工作日排休应扣1.5倍工时E", "workday_rest_leave_hours"), ("调整后缺勤工时", "adjusted_absence_hours"), ("调整后工时", "adjusted_working_hours"),
		("1.5倍结算工时", "overtime_1_5_settlement_hours"), ("2倍结算工时", "overtime_2_settlement_hours"), ("3倍结算工时", "overtime_3_settlement_hours"),
		("绿苹果", "green_apples"), ("红苹果", "red_apples"), ("苹果树金额", "apple_reward_amount"), ("红苹果扣款", "red_apple_penalty"),
		("夜班津贴", "night_shift_allowance"), ("全勤（含迟到）", "full_attendance_deduction"), ("签名", ""), ("备注", ""),
	]


def _add_monthly_attendance_export_sheet(workbook, profile_key, attendance_month, rows):
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	name_map = {"monthly_draft": "考勤初稿", "monthly_signed": "考勤终稿（签字版）", "monthly_finance": "考勤终稿（财务版）"}
	columns = _monthly_export_columns(profile_key)
	sheet = workbook.create_sheet(name_map[profile_key])
	sheet.append([f"{attendance_month.replace('-', '年')}月工时奖惩确认表"])
	sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
	sheet["A1"].font = Font(name="新宋体", size=16, bold=True)
	sheet["A1"].alignment = Alignment(horizontal="center")
	sheet.append([label for label, _fieldname in columns])
	_apply_export_header_style(sheet, [2], [12, 16, 14, 14, 16] + [16] * (len(columns) - 5))
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for index, row in enumerate(rows, start=1):
		values = []
		for _label, fieldname in columns:
			if fieldname == "sequence":
				values.append(index)
			elif fieldname == "date_of_joining":
				values.append(str(getattr(row, fieldname, "") or ""))
			elif not fieldname:
				values.append("")
			elif fieldname in {"department", "employee_name", "employee_code"}:
				values.append(getattr(row, fieldname, ""))
			else:
				values.append(_export_number(getattr(row, fieldname, 0)))
		sheet.append(values)
		for cell in sheet[sheet.max_row]:
			cell.border = border
			cell.alignment = Alignment(vertical="top", wrap_text=True)
	sheet.freeze_panes = "A3"
	sheet.auto_filter.ref = f"A2:{get_column_letter(len(columns))}{max(sheet.max_row, 3)}"
	return sheet


@frappe.whitelist()
def download_attendance_export(company: str, attendance_month: str, export_profile: str = "company_attendance_workbook"):
	"""Generate a read-only company/month export in the source workbook formats."""
	from openpyxl import Workbook
	from hrms.utils.export_watermark import save_workbook_with_logo_watermark

	_require_attendance_reviewer()
	company = _require_company(company)
	attendance_month = attendance_month or datetime.today().strftime("%Y-%m")
	_month_bounds(attendance_month)
	profile = _attendance_export_profile_or_throw(export_profile)
	daily_rows = _prefer_manual_daily_rows(_get_month_records(DAY_CHECK_DOCTYPE, "attendance_date", attendance_month, company))
	batch_ids = sorted({_cell_text(getattr(row, "import_batch", "")) for row in daily_rows if getattr(row, "import_batch", "")})
	leave_rows = frappe.get_all(LEAVE_EVIDENCE_DOCTYPE, filters={"import_batch": ["in", batch_ids or ["__none__"]]}, fields=["*"], order_by="leave_start asc, employee_name asc", limit_page_length=0)
	exceptions = frappe.get_all(EXCEPTION_DOCTYPE, filters={"import_batch": ["in", batch_ids or ["__none__"]]}, fields=["*"], order_by="attendance_date asc, employee_name asc", limit_page_length=0)
	apple_rows = _company_apple_records(attendance_month, company)
	monthly_rows = frappe.get_all(MONTHLY_SUMMARY_DOCTYPE, filters={"company": company, "attendance_month": attendance_month}, fields=["*"], order_by="department asc, employee_name asc", limit_page_length=0)

	workbook = Workbook()
	workbook.remove(workbook.active)
	for sheet_key in profile["sheet_keys"]:
		if sheet_key == "daily_statistics":
			sheet = _add_daily_statistics_template_sheet(workbook)
			_append_daily_statistics_export_rows(sheet, daily_rows)
		elif sheet_key == "attendance_detail":
			_add_attendance_detail_export_sheet(workbook, daily_rows, attendance_month)
		elif sheet_key == "leave_evidence":
			_add_leave_evidence_export_sheet(workbook, leave_rows)
		elif sheet_key == "attendance_exception":
			_add_attendance_exception_export_sheet(workbook, exceptions)
		elif sheet_key == "missing_card":
			_add_missing_card_export_sheet(workbook, exceptions, attendance_month)
		elif sheet_key == "apple_reward":
			_add_apple_reward_export_sheet(workbook, apple_rows)
		elif sheet_key in {"monthly_draft", "monthly_signed", "monthly_finance"}:
			_add_monthly_attendance_export_sheet(workbook, sheet_key, attendance_month, monthly_rows)

	output = BytesIO()
	save_workbook_with_logo_watermark(workbook, output)
	filename = f"{attendance_month}_{profile['label']}.xlsx"
	file_doc = frappe.get_doc({"doctype": "File", "file_name": filename, "content": output.getvalue(), "is_private": 0}).insert(ignore_permissions=True)
	return {
		"file_url": file_doc.file_url,
		"file_name": filename,
		"export_profile": export_profile,
		"row_counts": {"daily_statistics": len(daily_rows), "leave_evidence": len(leave_rows), "attendance_exception": len(exceptions), "apple_reward": len(apple_rows), "monthly_summary": len(monthly_rows)},
	}


def _insert_day_check(batch_name, row, company, source_kind="旧模板", source_sheet="", correction_version=1, allow_unmatched=False):
	employee_code = _first_value(row, "工号")
	employee_name = _first_value(row, "姓名")
	attendance_date = _parse_date(_first_value(row, "日期", "workDate"))
	if not employee_name or not attendance_date:
		return None
	employee = _employee_lookup(employee_code, employee_name)
	if not _employee_matches_company(employee, company) and not allow_unmatched:
		return None
	if employee and not _employee_matches_company(employee, company):
		employee = None
	shift_name = _first_value(row, "班次")
	actual_in_time = _first_value(row, "上班时间")
	actual_out_time = _first_value(row, "下班时间")
	missing_in = 1 if _first_value(row, "上班缺卡") or _float_value(row, "上班未打卡次数") else 0
	missing_out = 1 if _first_value(row, "下班缺卡") or _float_value(row, "下班未打卡次数") else 0
	absent_hours = _float_value(row, "请假/旷工(小时)", "旷工_2", "旷工(小时)", "旷工")
	standard_hours = _float_value(row, "标准工时")
	actual_attendance_hours = _float_value(row, "实际出勤（小时）", "实际出勤(小时)", "实际出勤")
	personal_leave_hours = _float_value(row, "请假/事假(小时)", "事假(小时)")
	sick_leave_hours = _float_value(row, "请假/病假(小时)", "病假(小时)")
	annual_leave_hours = _float_value(row, "请假/特休(小时)", "特休(小时)")
	work_injury_leave_hours = _float_value(row, "请假/工伤(小时)", "工伤(小时)")
	rest_leave_hours = _float_value(row, "请假/排休(小时)", "排休(小时)")
	bereavement_leave_hours = _float_value(row, "请假/丧假(小时)", "丧假(小时)")
	marriage_leave_hours = _float_value(row, "请假/婚假(小时)", "婚假(小时)") or _float_value(row, "请假/婚假(天)", "婚假(天)") * STANDARD_DAY_HOURS
	public_leave_hours = _float_value(row, "请假/公假(小时)", "公假(小时)") or _float_value(row, "请假/公假(天)", "公假(天)") * STANDARD_DAY_HOURS
	maternity_leave_hours = _float_value(row, "请假/产假(小时)", "产假(小时)") or _float_value(row, "请假/产假(天)", "产假(天)") * STANDARD_DAY_HOURS
	reunion_leave_hours = _float_value(row, "请假/团圆假(小时)", "团圆假(小时)") or _float_value(row, "请假/团圆假(天)", "团圆假(天)") * STANDARD_DAY_HOURS
	leave_hours = sum(
		[
			personal_leave_hours,
			sick_leave_hours,
			annual_leave_hours,
			work_injury_leave_hours,
			rest_leave_hours,
			bereavement_leave_hours,
			marriage_leave_hours,
			public_leave_hours,
			maternity_leave_hours,
			reunion_leave_hours,
		]
	)
	approval_summary = _first_value(row, "关联审批单", "关联的审批单")
	has_overtime = flt(_first_value(row, "工作日加班（小时）", "工作日加班(小时)")) or flt(_first_value(row, "休息日加班（小时）", "休息日加班(小时)")) or flt(_first_value(row, "节假日加班（小时）", "节假日加班(小时)"))
	overtime_without_approval = 1 if has_overtime and "加班" not in approval_summary else 0
	attendance_result = "异常" if missing_in or missing_out or absent_hours or _int_value(row, "迟到次数") or _int_value(row, "早退次数") else "正常"

	doc = frappe.get_doc(
		{
			"doctype": DAY_CHECK_DOCTYPE,
			"import_batch": batch_name,
			"company": company,
			"source_kind": source_kind,
			"source_sheet": source_sheet,
			"source_row_number": row.get("_source_row", 0),
			"correction_version": correction_version,
			"attendance_date": attendance_date,
			"employee": employee,
			"employee_code": employee_code,
			"employee_name": employee_name,
			"attendance_group": _first_value(row, "考勤组"),
			"department": _department_lookup(_first_value(row, "实际部门", "部门")),
			"position": _first_value(row, "职位"),
			"user_id": _first_value(row, "UserId"),
			"date_type": _first_value(row, "日期类型"),
			"shift_name": shift_name,
			"scheduled_in_time": _first_value(row, "应上班时间") or _parse_shift_start_time(shift_name),
			"scheduled_out_time": _first_value(row, "应下班时间"),
			"actual_in_time": actual_in_time,
			"actual_out_time": actual_out_time,
			"missing_in": missing_in,
			"missing_out": missing_out,
			"attendance_result": attendance_result,
			"attendance_duration_hours": actual_attendance_hours,
			"absent_hours": absent_hours,
			"standard_hours": standard_hours,
			"actual_attendance_hours": actual_attendance_hours,
			"workday_overtime_hours": _float_value(row, "工作日加班（小时）", "工作日加班(小时)"),
			"restday_overtime_hours": _float_value(row, "休息日加班（小时）", "休息日加班(小时)"),
			"holiday_overtime_hours": _float_value(row, "节假日加班（小时）", "节假日加班(小时)"),
			"large_night_shift_count": _float_value(row, "大夜班"),
			"small_night_shift_count": _float_value(row, "小夜班"),
			"leave_summary": _first_value(row, "请假") or approval_summary,
			"leave_hours": leave_hours,
			"personal_leave_hours": personal_leave_hours,
			"sick_leave_hours": sick_leave_hours,
			"annual_leave_hours": annual_leave_hours,
			"work_injury_leave_hours": work_injury_leave_hours,
			"rest_leave_hours": rest_leave_hours,
			"bereavement_leave_hours": bereavement_leave_hours,
			"marriage_leave_hours": marriage_leave_hours,
			"public_leave_hours": public_leave_hours,
			"maternity_leave_hours": maternity_leave_hours,
			"reunion_leave_hours": reunion_leave_hours,
			"valid_leave_hours": 0,
			"invalid_leave_hours": 0,
			"overtime_without_approval": overtime_without_approval,
			"late_count": _int_value(row, "迟到次数"),
			"early_count": _int_value(row, "早退次数"),
			"raw_row_json": json.dumps(row, ensure_ascii=False, default=str),
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def _insert_leave_evidence(batch_name, row):
	employee_name = _first_value(row, "创建人", "姓名")
	leave_start = _parse_datetime(_first_value(row, "开始时间"))
	leave_end = _parse_datetime(_first_value(row, "结束时间"))
	if not employee_name or not leave_start:
		return None
	employee_code = _first_value(row, "工号")
	valid = _is_valid_approval(row)
	doc = frappe.get_doc(
		{
			"doctype": LEAVE_EVIDENCE_DOCTYPE,
			"import_batch": batch_name,
			"employee": _employee_lookup(employee_code, employee_name),
			"employee_code": employee_code,
			"employee_name": employee_name,
			"department": _department_lookup(_first_value(row, "创建人部门", "部门")),
			"leave_type": _first_value(row, "请假类型（实际）", "请假类型"),
			"leave_start": leave_start,
			"leave_end": leave_end,
			"leave_hours": _duration_hours(_first_value(row, "时长")),
			"leave_reason": _first_value(row, "请假事由"),
			"approval_no": _first_value(row, "审批编号"),
			"approval_result": _first_value(row, "审批结果"),
			"approval_status": _first_value(row, "审批状态"),
			"is_valid_approval": valid,
			"raw_row_json": json.dumps(row, ensure_ascii=False, default=str),
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def _person_keys(row):
	keys = []
	for fieldname in ("employee", "employee_code", "employee_name"):
		value = _cell_text(getattr(row, fieldname, ""))
		if value and value not in keys:
			keys.append(value)
	return keys


def _primary_person_key(row):
	keys = _person_keys(row)
	return keys[0] if keys else ""


def _index_records_by_person(records):
	index = defaultdict(list)
	for record in records:
		for key in _person_keys(record):
			index[key].append(record)
	return index


def _records_for_same_person(index, row):
	records = []
	seen = set()
	for key in _person_keys(row):
		for record in index.get(key, []):
			name = getattr(record, "name", None) or id(record)
			if name in seen:
				continue
			seen.add(name)
			records.append(record)
	return records


def _apply_leave_evidence_to_day_checks(batch_name):
	leaves = frappe.get_all(
		LEAVE_EVIDENCE_DOCTYPE,
		filters={"import_batch": batch_name},
		fields=["employee", "employee_code", "employee_name", "leave_type", "leave_start", "leave_end", "leave_hours", "is_valid_approval", "approval_no"],
	)
	evidence_by_person = _index_records_by_person(leaves)

	day_checks = frappe.get_all(DAY_CHECK_DOCTYPE, filters={"import_batch": batch_name}, fields=["*"])
	for day_check in day_checks:
		matched_valid = []
		matched_invalid = []
		attendance_date = getdate(day_check.attendance_date)
		for leave in _records_for_same_person(evidence_by_person, day_check):
			start = getdate(leave.leave_start)
			end = getdate(leave.leave_end) if leave.leave_end else start
			if start <= attendance_date <= end:
				if leave.is_valid_approval:
					matched_valid.append(leave)
				else:
					matched_invalid.append(leave)
		valid_hours = sum(flt(leave.leave_hours) for leave in matched_valid)
		invalid_hours = sum(flt(leave.leave_hours) for leave in matched_invalid)
		updates = {
			"valid_leave_hours": min(flt(day_check.leave_hours) or valid_hours, valid_hours) if valid_hours else 0,
			"invalid_leave_hours": invalid_hours,
			"valid_leave_summary": "；".join(f"{leave.leave_type}{flt(leave.leave_hours):g}H" for leave in matched_valid[:4]),
		}
		if matched_valid and day_check.attendance_result == "异常" and not (day_check.missing_in or day_check.missing_out or day_check.late_count or day_check.early_count or day_check.absent_hours):
			updates["attendance_result"] = "请假"
		frappe.db.set_value(DAY_CHECK_DOCTYPE, day_check.name, updates)


def _insert_apple_record(batch_name, row):
	employee_name = _first_value(row, "受奖/惩人")
	reward_date = _parse_date(_first_value(row, "奖/惩日期"))
	if not employee_name or not reward_date:
		return None
	green = _float_value(row, "绿苹果")
	red = _float_value(row, "红苹果")
	employee_code = _first_value(row, "工号")
	valid = _is_valid_approval(row)
	doc = frappe.get_doc(
		{
			"doctype": APPLE_RECORD_DOCTYPE,
			"import_batch": batch_name,
			"reward_date": reward_date,
			"employee": _employee_lookup(employee_code, employee_name),
			"employee_code": employee_code,
			"employee_name": employee_name,
			"department": _department_lookup(_first_value(row, "受奖/惩人部门", "部门")),
			"reward_item": _first_value(row, "奖/惩项目"),
			"green_apples": green,
			"red_apples": red,
			"reward_amount": (green - red) * APPLE_UNIT_AMOUNT,
			"approval_no": _first_value(row, "审批编号"),
			"approval_result": _first_value(row, "审批结果"),
			"approval_status": _first_value(row, "审批状态"),
			"is_valid_approval": valid,
			"created_by_name": _first_value(row, "创建人"),
			"raw_row_json": json.dumps(row, ensure_ascii=False, default=str),
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def _create_attendance_batch(file_url, attendance_month, company, source_type):
	source_checksum = _source_file_checksum(file_url)
	# A revoked batch has no derived rows left. It is retained only as audit
	# evidence and must not prevent the same source file from being imported again.
	existing = frappe.db.get_value(
		ATTENDANCE_BATCH_DOCTYPE,
		{"company": company, "source_checksum": source_checksum, "status": ["!=", "已撤销"]},
		"name",
	)
	if existing:
		return frappe.get_doc(ATTENDANCE_BATCH_DOCTYPE, existing), True

	batch = frappe.get_doc(
		{
			"doctype": ATTENDANCE_BATCH_DOCTYPE,
			"company": company,
			"attendance_month": attendance_month,
			"source_file": file_url,
			"source_type": source_type,
			"source_checksum": source_checksum,
			"status": "已导入",
			"imported_by": frappe.session.user,
			"imported_on": now_datetime(),
		}
	)
	batch.insert(ignore_permissions=True)
	return batch, False


def _duplicate_batch_result(batch):
	"""Describe an active identical import without presenting it as a new write."""
	return {
		"batch": batch.name,
		"duplicate": 1,
		"status": batch.status,
		"daily_sheet_rows": batch.daily_sheet_rows,
		"inserted_day_checks": frappe.db.count(DAY_CHECK_DOCTYPE, {"import_batch": batch.name}),
		"rejected_company_or_employee_rows": 0,
	}


def _import_company_attendance_workbook(workbook, file_url, attendance_month, company):
	correction_version = _correction_version_for_import(company, attendance_month)
	batch, duplicate = _create_attendance_batch(file_url, attendance_month, company, "company_attendance_workbook_v1")
	if duplicate:
		return _duplicate_batch_result(batch)

	row_counts = {}
	rejected_rows = 0
	inserted_rows = 0
	for source_kind, source in COMPANY_ATTENDANCE_WORKBOOK_SOURCES.items():
		rows = _daily_rows_from_header_rows(workbook[source["sheet_name"]], source["header_rows"], source["data_start_row"])
		for row in rows:
			if _insert_day_check(
				batch.name,
				row,
				company,
				"钉钉原始导出" if source_kind == "dingtalk_raw" else "人工调整",
				source["sheet_name"],
				correction_version,
			):
				inserted_rows += 1
			else:
				rejected_rows += 1
		row_counts[source_kind] = len(rows)

	batch.daily_sheet_rows = sum(row_counts.values())
	batch.notes = json.dumps(
		{
			"daily_sources": row_counts,
			"daily_statistics_imported": inserted_rows,
			"rejected_company_or_employee_rows": rejected_rows,
		},
		ensure_ascii=False,
	)
	batch.save(ignore_permissions=True)
	frappe.db.commit()
	return {
		"batch": batch.name,
		"daily_sheet_rows": batch.daily_sheet_rows,
		"daily_sources": row_counts,
		"inserted_day_checks": inserted_rows,
		"rejected_company_or_employee_rows": rejected_rows,
	}


def _import_company_attendance_register_v1(workbook, file_url, attendance_month, company):
	correction_version = _correction_version_for_import(company, attendance_month)
	batch, duplicate = _create_attendance_batch(file_url, attendance_month, company, "company_attendance_register_v1")
	if duplicate:
		return _duplicate_batch_result(batch)

	daily_source = COMPANY_ATTENDANCE_REGISTER_V1_SHEETS["daily_statistics"]
	daily_sheet = _sheet_by_required_name(workbook, daily_source["sheet_name"])
	daily_rows = _daily_rows_from_header_rows(daily_sheet, daily_source["header_rows"], daily_source["data_start_row"])
	rejected_rows = 0
	inserted_rows = 0
	for row in daily_rows:
		if _insert_day_check(batch.name, row, company, "人工调整", daily_source["sheet_name"], correction_version):
			inserted_rows += 1
		else:
			rejected_rows += 1

	apple_source = COMPANY_ATTENDANCE_REGISTER_V1_SHEETS["apple_tree"]
	apple_sheet = _sheet_by_required_name(workbook, apple_source["sheet_name"])
	apple_rows = _rows_as_dicts(apple_sheet, ["奖/惩日期", "受奖/惩人", "绿苹果", "红苹果"])
	inserted_apples = sum(1 for row in apple_rows if _insert_apple_record(batch.name, row))
	reference_counts = {
		"出勤明细": _count_nonempty_rows(
			_sheet_by_required_name(workbook, "出勤明细"),
			COMPANY_ATTENDANCE_REGISTER_V1_SHEETS["attendance_detail"]["data_start_row"],
		),
		"出勤异常": _count_nonempty_rows(
			_sheet_by_required_name(workbook, "出勤异常"),
			COMPANY_ATTENDANCE_REGISTER_V1_SHEETS["attendance_exception"]["data_start_row"],
		),
	}
	batch.daily_sheet_rows = len(daily_rows)
	batch.apple_sheet_rows = len(apple_rows)
	batch.notes = json.dumps(
		{
			"daily_statistics_imported": inserted_rows,
			"apple_records_imported": inserted_apples,
			"rejected_company_or_employee_rows": rejected_rows,
			"reference_only_sheet_rows": reference_counts,
		},
		ensure_ascii=False,
	)
	batch.save(ignore_permissions=True)
	frappe.db.commit()
	return {
		"batch": batch.name,
		"daily_sheet_rows": len(daily_rows),
		"inserted_day_checks": inserted_rows,
		"apple_sheet_rows": len(apple_rows),
		"inserted_apple_records": inserted_apples,
		"rejected_company_or_employee_rows": rejected_rows,
		"reference_only_sheet_rows": reference_counts,
	}


def _import_dingtalk_export_v1(workbook, file_url, attendance_month, company):
	correction_version = _correction_version_for_import(company, attendance_month)
	batch, duplicate = _create_attendance_batch(file_url, attendance_month, company, "dingtalk_export_v1")
	if duplicate:
		return _duplicate_batch_result(batch)

	daily_rows = _dingtalk_daily_rows(_sheet_by_required_name(workbook, "每日统计"))
	inserted_rows = 0
	rejected_rows = 0
	for row in daily_rows:
		if _insert_day_check(
			batch.name,
			row,
			company,
			_dingtalk_export_import_source_kind(),
			"每日统计",
			correction_version,
		):
			inserted_rows += 1
		else:
			rejected_rows += 1

	preview = _preview_dingtalk_export_v1(workbook)
	batch.daily_sheet_rows = len(daily_rows)
	batch.notes = json.dumps(
		{
			"daily_statistics_imported": inserted_rows,
			"rejected_company_or_employee_rows": rejected_rows,
			"source_only_record_counts": preview["record_counts"],
		},
		ensure_ascii=False,
	)
	batch.save(ignore_permissions=True)
	frappe.db.commit()
	return {
		"batch": batch.name,
		"daily_sheet_rows": len(daily_rows),
		"inserted_day_checks": inserted_rows,
		"rejected_company_or_employee_rows": rejected_rows,
		"source_only_record_counts": preview["record_counts"],
	}


def _record_import_preview_metadata(batch_name, preview):
	"""Persist the read-only mapping verdict alongside the batch for audit.

	The uploaded file remains the source evidence. This stores only the mapping
	and validation explanation shown before the user clicked import.
	"""
	if not batch_name:
		return
	batch = frappe.get_doc(ATTENDANCE_BATCH_DOCTYPE, batch_name)
	try:
		notes = json.loads(batch.notes or "{}")
	except (TypeError, ValueError):
		notes = {"previous_notes": batch.notes or ""}
	notes["import_preview"] = {
		"source_type": preview.get("source_type", "legacy_workbook"),
		"field_mapping": preview.get("field_mapping", {}),
		"import_validation": preview.get("import_validation", {}),
		"recorded_at": str(now_datetime()),
		"recorded_by": frappe.session.user,
	}
	batch.notes = json.dumps(notes, ensure_ascii=False, default=str)
	batch.save(ignore_permissions=True)
	frappe.db.commit()


@frappe.whitelist()
def import_attendance_workbook(file_url: str, attendance_month: str = "", company: str = ""):
	company = _require_company(company)
	workbook = _load_workbook(file_url)
	preview = preview_attendance_workbook(file_url)
	if not attendance_month:
		attendance_month = datetime.today().strftime("%Y-%m")

	if preview.get("source_type") == "company_attendance_register_v1":
		result = _import_company_attendance_register_v1(workbook, file_url, attendance_month, company)
		_record_import_preview_metadata(result.get("batch"), preview)
		return {**result, "import_validation": preview.get("import_validation", {})}
	if preview.get("source_type") == "company_attendance_workbook_v1":
		result = _import_company_attendance_workbook(workbook, file_url, attendance_month, company)
		_record_import_preview_metadata(result.get("batch"), preview)
		return {**result, "import_validation": preview.get("import_validation", {})}
	if preview.get("source_type") == "dingtalk_export_v1":
		result = _import_dingtalk_export_v1(workbook, file_url, attendance_month, company)
		_record_import_preview_metadata(result.get("batch"), preview)
		return {**result, "import_validation": preview.get("import_validation", {})}
	if preview["missing_sheets"]:
		frappe.throw(_("缺少必要工作表：{0}").format("、".join(preview["missing_sheets"])))

	correction_version = _correction_version_for_import(company, attendance_month)
	batch, duplicate = _create_attendance_batch(file_url, attendance_month, company, "legacy_workbook")
	if duplicate:
		return _duplicate_batch_result(batch)

	daily_rows = _rows_as_dicts(_sheet_by_required_name(workbook, "1.1每日统计"), ["姓名", "工号", "日期", "班次"])
	leave_rows = _rows_as_dicts(_sheet_by_required_name(workbook, "1.2请假单"), ["请假类型", "开始时间", "结束时间"])
	apple_rows = _rows_as_dicts(_sheet_by_required_name(workbook, "1.3苹果树"), ["奖/惩日期", "受奖/惩人", "绿苹果", "红苹果"])

	rejected_rows = 0
	inserted_rows = 0
	for row in daily_rows:
		if _insert_day_check(batch.name, row, company, "旧模板", "1.1每日统计", correction_version):
			inserted_rows += 1
		else:
			rejected_rows += 1
	for row in leave_rows:
		_insert_leave_evidence(batch.name, row)
	for row in apple_rows:
		_insert_apple_record(batch.name, row)
	_apply_leave_evidence_to_day_checks(batch.name)

	batch.daily_sheet_rows = len(daily_rows)
	batch.leave_sheet_rows = len(leave_rows)
	batch.apple_sheet_rows = len(apple_rows)
	batch.save(ignore_permissions=True)
	frappe.db.commit()
	result = {
		"batch": batch.name,
		"daily_sheet_rows": len(daily_rows),
		"inserted_day_checks": inserted_rows,
		"leave_sheet_rows": len(leave_rows),
		"apple_sheet_rows": len(apple_rows),
		"rejected_company_or_employee_rows": rejected_rows,
	}
	_record_import_preview_metadata(result.get("batch"), preview)
	return {**result, "import_validation": preview.get("import_validation", {})}


def _make_exception(day_check, exception_type, handling_method="", deduct_absence_hours=0, full_attendance_deduction=0, red_apple_penalty=0, remarks=""):
	doc = frappe.get_doc(
		{
			"doctype": EXCEPTION_DOCTYPE,
			"import_batch": day_check.import_batch,
			"day_check": day_check.name,
			"attendance_date": day_check.attendance_date,
			"employee": day_check.employee,
			"employee_code": day_check.employee_code,
			"employee_name": day_check.employee_name,
			"department": day_check.department,
			"exception_type": exception_type,
			"expected_shift": day_check.shift_name,
			"actual_in_time": day_check.actual_in_time,
			"actual_out_time": day_check.actual_out_time,
			"handling_method": handling_method,
			"deduct_absence_hours": deduct_absence_hours,
			"full_attendance_deduction": full_attendance_deduction,
			"red_apple_penalty": red_apple_penalty,
			"confirmation_status": "待确认",
			"remarks": remarks,
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def _build_exception_candidates(day_check):
	candidates = []
	if not day_check.employee:
		candidates.append(
			{
				"exception_type": "员工未匹配",
				"handling_method": "请在人事侧确认钉钉 UserId、工号与员工档案的映射；确认前不得进入月度锁定或薪资。",
				"remarks": "钉钉原始数据已保留，等待人事匹配员工。",
			}
		)
	if day_check.missing_in or day_check.missing_out:
		candidates.append(
			{
				"exception_type": "忘打卡",
				"handling_method": "钉钉补卡；月底统计2个红苹果/次。不可抗因素需HR主管确认后视为已打卡。",
				"red_apple_penalty": 2,
			}
		)

	shift_start = day_check.scheduled_in_time or _parse_shift_start_time(day_check.shift_name)
	shift_start_minutes = _time_text_to_minutes(shift_start)
	actual_in_minutes = _time_text_to_minutes(day_check.actual_in_time)
	late_minutes = 0
	if shift_start_minutes is not None and actual_in_minutes is not None and actual_in_minutes > shift_start_minutes:
		late_minutes = actual_in_minutes - shift_start_minutes
	if flt(day_check.late_count) or late_minutes:
		deduct_hours = 0.5 if 0 < late_minutes <= 30 else round(late_minutes / 60, 2) if late_minutes else 0
		candidates.append(
			{
				"exception_type": "迟到",
				"handling_method": "需提交钉钉事假单；0-0.5H按0.5H缺勤并扣全勤10元，超过0.5H按实际迟到时长计缺勤。",
				"deduct_absence_hours": deduct_hours,
				"full_attendance_deduction": 10 if 0 < late_minutes <= 30 else 0,
			}
		)
	if flt(day_check.early_count):
		candidates.append(
			{
				"exception_type": "早退",
				"handling_method": "未请假或未提前告知主管离岗，早退缺勤时数按旷工处理。",
				"remarks": "请核对请假或主管说明。",
			}
		)
	# A blank clock-time cell in a historical export is not proof of absence.
	# Only a source-provided absence value or an explicitly abnormal/no-show result
	# may create a salary-relevant absence review item.
	result_text = _cell_text(day_check.attendance_result).strip().lower()
	explicit_absence = any(token in result_text for token in ("旷工", "缺勤", "未出勤", "absence", "absent", "no show"))
	if flt(day_check.absent_hours) or (explicit_absence and not day_check.valid_leave_hours):
		candidates.append(
			{
				"exception_type": "旷工",
				"handling_method": "工作日未打卡且无有效请假；薪资按3倍旷工工时扣除。",
				"deduct_absence_hours": flt(day_check.absent_hours) or flt(day_check.standard_hours),
			}
		)
	if flt(day_check.overtime_without_approval):
		candidates.append(
			{
				"exception_type": "未申请加班",
				"handling_method": "请核对钉钉加班审批；无审批时在每日表标黄提醒。",
				"remarks": "工作日/周末出勤存在加班时数但未匹配加班审批。",
			}
		)
	return candidates


@frappe.whitelist()
def generate_attendance_exceptions(batch: str):
	existing = frappe.get_all(EXCEPTION_DOCTYPE, filters={"import_batch": batch}, pluck="name")
	for name in existing:
		frappe.delete_doc(EXCEPTION_DOCTYPE, name, ignore_permissions=True, force=True)

	created = []
	day_checks = frappe.get_all(
		DAY_CHECK_DOCTYPE,
		filters={"import_batch": batch},
		fields=["*"],
		order_by="attendance_date asc, employee_name asc",
	)
	for row in day_checks:
		day_check = frappe._dict(row)
		for candidate in _build_exception_candidates(day_check):
			created.append(_make_exception(day_check, **candidate))

	frappe.db.set_value(ATTENDANCE_BATCH_DOCTYPE, batch, "status", "已生成异常")
	frappe.db.commit()
	return {"created": len(created), "sample_exception_names": created[:20]}


def _get_month_records(doctype, date_field, attendance_month, company):
	start, end = _month_bounds(attendance_month)
	return frappe.get_all(
		doctype,
		filters=[["company", "=", company], [date_field, ">=", start], [date_field, "<", end]],
		fields=["*"],
		order_by=f"{date_field} asc",
	)


def _get_or_create_month_lock(company, attendance_month):
	name = frappe.db.get_value(MONTH_LOCK_DOCTYPE, {"company": company, "attendance_month": attendance_month}, "name")
	if name:
		return frappe.get_doc(MONTH_LOCK_DOCTYPE, name)
	lock = frappe.get_doc(
		{
			"doctype": MONTH_LOCK_DOCTYPE,
			"company": company,
			"attendance_month": attendance_month,
			"status": "草稿",
			"active_version": 1,
		}
	)
	lock.insert(ignore_permissions=True)
	return lock


def _append_lock_audit(lock, action, reason=""):
	doc = frappe.get_doc(
		{
			"doctype": LOCK_AUDIT_DOCTYPE,
			"month_lock": lock.name,
			"company": lock.company,
			"attendance_month": lock.attendance_month,
			"action": action,
			"lock_version": lock.active_version,
			"reason": reason,
			"operator": frappe.session.user,
			"occurred_on": now_datetime(),
			"source_checksum": lock.source_checksum,
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def _prepare_month_lock_for_generation(company, attendance_month):
	lock = _get_or_create_month_lock(company, attendance_month)
	if lock.status == "已锁定":
		frappe.throw(_("该公司考勤月份已锁定，不能重新生成。"))
	if lock.status == "已重开":
		lock.active_version = int(lock.active_version or 1) + 1
		lock.status = "草稿"
		lock.save(ignore_permissions=True)
		_append_lock_audit(lock, "创建更正版本", _("解锁后的更正版本"))
	return lock


def _correction_version_for_import(company, attendance_month):
	lock = _get_or_create_month_lock(company, attendance_month)
	if lock.status == "已锁定":
		frappe.throw(_("该公司考勤月份已锁定，不能导入更正数据。"))
	return int(lock.active_version or 1) + (1 if lock.status == "已重开" else 0)


def _department_confirmation_metrics(company, attendance_month):
	"""Build the department-level confirmation figures from effective daily records.

	The department report is a confirmation view, not a second calculation source.
	Every number remains traceable to the selected employee/day records.
	"""
	rows = _prefer_manual_daily_rows(_get_month_records(DAY_CHECK_DOCTYPE, "attendance_date", attendance_month, company))
	metrics = defaultdict(lambda: {"employees": set(), "attendance": set(), "leave": set(), "exceptions": set()})
	day_check_department = {}
	for row in rows:
		department = _cell_text(getattr(row, "department", "")).strip()
		if not department:
			continue
		person_key = _primary_person_key(row)
		if not person_key:
			continue
		metrics[department]["employees"].add(person_key)
		if flt(getattr(row, "actual_attendance_hours", 0)) > 0:
			metrics[department]["attendance"].add(person_key)
		if flt(getattr(row, "valid_leave_hours", 0)) > 0 or flt(getattr(row, "leave_hours", 0)) > 0:
			metrics[department]["leave"].add(person_key)
		day_check_department[getattr(row, "name", "")] = (department, person_key)

	batch_ids = sorted({_cell_text(getattr(row, "import_batch", "")) for row in rows if getattr(row, "import_batch", "")})
	if batch_ids:
		for exception in frappe.get_all(
			EXCEPTION_DOCTYPE,
			filters={"import_batch": ["in", batch_ids], "confirmation_status": ["!=", "已驳回"]},
			fields=["day_check", "department", "employee", "employee_code", "employee_name"],
			limit_page_length=0,
		):
			department, person_key = day_check_department.get(
				getattr(exception, "day_check", ""),
				(_cell_text(getattr(exception, "department", "")).strip(), _primary_person_key(exception)),
			)
			if department and person_key:
				metrics[department]["exceptions"].add(person_key)

	return rows, {
		department: {
			"current_headcount": len(values["employees"]),
			"attendance_count": len(values["attendance"]),
			"leave_count": len(values["leave"]),
			"exception_count": len(values["exceptions"]),
		}
		for department, values in metrics.items()
	}


def _ensure_monthly_department_confirmations(company, attendance_month, source_checksum=""):
	"""Create or refresh pending department confirmations for the active month version."""
	if not frappe.db.exists("DocType", DEPARTMENT_CONFIRMATION_DOCTYPE):
		return {"created": 0, "updated": 0, "departments": 0, "pending": 0}
	lock = _get_or_create_month_lock(company, attendance_month)
	if lock.status == "已锁定":
		return {"created": 0, "updated": 0, "departments": 0, "pending": 0}
	rows, metrics_by_department = _department_confirmation_metrics(company, attendance_month)
	if not source_checksum:
		_, source_checksum = _source_summary_metadata(rows)
	version = int(lock.active_version or 1)
	existing = frappe.get_all(
		DEPARTMENT_CONFIRMATION_DOCTYPE,
		filters={
			"company": company,
			"attendance_month": attendance_month,
			"confirmation_scope": "月度部门工时",
			"attendance_lock_version": version,
		},
		fields=["name", "department", "confirmation_status", "source_checksum"],
		limit_page_length=0,
	)
	existing_by_department = {row.department: row for row in existing}
	created = 0
	updated = 0
	for department, metrics in metrics_by_department.items():
		current = existing_by_department.get(department)
		values = {
			"company": company,
			"attendance_month": attendance_month,
			"confirmation_scope": "月度部门工时",
			"attendance_lock_version": version,
			"department": department,
			"source_checksum": source_checksum,
			**metrics,
		}
		if current:
			doc = frappe.get_doc(DEPARTMENT_CONFIRMATION_DOCTYPE, current.name)
			# Recalculate only in a draft version. A changed source invalidates an
			# earlier confirmation and requires the department to reconfirm.
			if current.source_checksum != source_checksum and current.confirmation_status == "已确认":
				values.update({"confirmation_status": "待部门确认", "confirmed_by": "", "confirmed_on": None, "signoff_attachment": ""})
			doc.update(values)
			doc.save(ignore_permissions=True)
			updated += 1
		else:
			frappe.get_doc({"doctype": DEPARTMENT_CONFIRMATION_DOCTYPE, "confirmation_status": "待部门确认", **values}).insert(ignore_permissions=True)
			created += 1
	pending = frappe.db.count(
		DEPARTMENT_CONFIRMATION_DOCTYPE,
		{
			"company": company,
			"attendance_month": attendance_month,
			"confirmation_scope": "月度部门工时",
			"attendance_lock_version": version,
			"confirmation_status": ["!=", "已确认"],
		},
	)
	return {"created": created, "updated": updated, "departments": len(metrics_by_department), "pending": pending}


@frappe.whitelist()
def list_attendance_department_confirmations(company: str, attendance_month: str, page_length: int = 100):
	company = _require_company(company)
	_month_bounds(attendance_month)
	lock = _get_or_create_month_lock(company, attendance_month)
	_ensure_monthly_department_confirmations(company, attendance_month)
	return frappe.get_all(
		DEPARTMENT_CONFIRMATION_DOCTYPE,
		filters={
			"company": company,
			"attendance_month": attendance_month,
			"confirmation_scope": "月度部门工时",
			"attendance_lock_version": int(lock.active_version or 1),
		},
		fields=["name", "department", "current_headcount", "attendance_count", "leave_count", "exception_count", "confirmation_status", "department_contact", "confirmed_by", "confirmed_on", "signoff_attachment", "return_reason", "remarks", "source_checksum"],
		order_by="department asc",
		limit_page_length=int(page_length or 100),
	)


@frappe.whitelist()
def review_attendance_department_confirmation(name: str, decision: str, remarks: str = "", signoff_attachment: str = ""):
	"""Record HR's receipt of the department confirmation or a department return."""
	_require_attendance_reviewer()
	if decision not in ("confirm", "return"):
		frappe.throw(_("确认操作仅支持 confirm 或 return。"))
	doc = frappe.get_doc(DEPARTMENT_CONFIRMATION_DOCTYPE, name)
	lock = _get_or_create_month_lock(doc.company, doc.attendance_month)
	if lock.status == "已锁定":
		frappe.throw(_("该月份已经锁定，不能修改部门确认。"))
	if int(doc.attendance_lock_version or 0) != int(lock.active_version or 1):
		frappe.throw(_("该部门确认属于旧版本，请确认当前考勤版本。"))
	remarks = (remarks or "").strip()
	if decision == "return" and not remarks:
		frappe.throw(_("退回部门确认必须填写原因。"))
	if decision == "confirm":
		doc.confirmation_status = "已确认"
		doc.confirmed_by = frappe.session.user
		doc.confirmed_on = now_datetime()
		doc.return_reason = ""
		if signoff_attachment:
			doc.signoff_attachment = signoff_attachment
	else:
		doc.confirmation_status = "已退回"
		doc.return_reason = remarks
		doc.confirmed_by = ""
		doc.confirmed_on = None
	if remarks:
		doc.remarks = remarks
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "confirmation_status": doc.confirmation_status}


def _assert_month_ready_for_lock(company, attendance_month):
	day_checks = _get_month_records(DAY_CHECK_DOCTYPE, "attendance_date", attendance_month, company)
	unmatched = [row for row in day_checks if not getattr(row, "employee", "")]
	missing_department = [row for row in day_checks if not getattr(row, "department", "")]
	batch_ids = sorted({_cell_text(getattr(row, "import_batch", "")) for row in day_checks if getattr(row, "import_batch", "")})
	start, end = _month_bounds(attendance_month)
	pending = frappe.get_all(
		EXCEPTION_DOCTYPE,
		filters=[
			["import_batch", "in", batch_ids or ["__none__"]],
			["attendance_date", ">=", start],
			["attendance_date", "<", end],
			["confirmation_status", "=", "待确认"],
		],
		fields=["name"],
	)
	if unmatched:
		frappe.throw(_("存在 {0} 条未匹配员工的日考勤，不能锁定。").format(len(unmatched)))
	if missing_department:
		frappe.throw(_("存在 {0} 条未匹配部门的日考勤，不能锁定。").format(len(missing_department)))
	if pending:
		frappe.throw(_("存在 {0} 条待确认考勤异常，不能锁定。").format(len(pending)))
	if frappe.db.exists("DocType", DEPARTMENT_CONFIRMATION_DOCTYPE):
		lock = _get_or_create_month_lock(company, attendance_month)
		confirmation = _ensure_monthly_department_confirmations(company, attendance_month)
		if confirmation["pending"]:
			frappe.throw(_("存在 {0} 个待部门确认或已退回的月度考勤汇总，不能锁定。").format(confirmation["pending"]))


def _daily_identity_key(row):
	return (
		getattr(row, "employee", "") or getattr(row, "employee_code", "") or getattr(row, "employee_name", ""),
		_cell_text(getattr(row, "attendance_date", "")),
	)


def _row_correction_version(row):
	try:
		return int(getattr(row, "correction_version", 1) or 1)
	except (TypeError, ValueError):
		return 1


def _prefer_manual_daily_rows(rows):
	"""Choose a deterministic effective record for an employee/day.

	A live DingTalk API pull is newer evidence than an imported historical
	export, but neither may silently overwrite a deliberate HR adjustment.
	"""
	source_priority = {"旧模板": 1, "钉钉原始导出": 2, "钉钉API同步": 3, "人工调整": 4}
	selected = {}
	for row in rows:
		key = _daily_identity_key(row)
		if not key[0]:
			continue
		current = selected.get(key)
		if not current or _row_correction_version(row) > _row_correction_version(current):
			selected[key] = row
		elif _row_correction_version(row) == _row_correction_version(current):
			current_priority = source_priority.get(getattr(current, "source_kind", ""), 0)
			row_priority = source_priority.get(getattr(row, "source_kind", ""), 0)
			if row_priority > current_priority or (
				row_priority == current_priority and str(getattr(row, "modified", "")) > str(getattr(current, "modified", ""))
			):
				selected[key] = row
	return list(selected.values())


def _require_attendance_reviewer():
	frappe.only_for(("System Manager", "HR Manager"))


def _month_batch_ids(company, attendance_month):
	start, end = _month_bounds(attendance_month)
	return frappe.get_all(
		DAY_CHECK_DOCTYPE,
		filters=[["company", "=", company], ["attendance_date", ">=", start], ["attendance_date", "<", end]],
		pluck="import_batch",
		limit_page_length=0,
	)


@frappe.whitelist()
def get_attendance_review_dashboard(company: str, attendance_month: str = "", batch: str = "", attendance_date: str = ""):
	"""Return the small, auditable set of metrics used by the attendance workbench."""
	company = _require_company(company)
	attendance_month = attendance_month or datetime.today().strftime("%Y-%m")
	start, end = _month_bounds(attendance_month)
	filters = [["company", "=", company], ["attendance_date", ">=", start], ["attendance_date", "<", end]]
	if attendance_date:
		filters.append(["attendance_date", "=", getdate(attendance_date)])
	if batch:
		filters.append(["import_batch", "=", batch])
	rows = frappe.get_all(DAY_CHECK_DOCTYPE, filters=filters, fields=["name", "import_batch", "employee", "employee_code", "attendance_date", "source_kind", "attendance_result", "missing_in", "missing_out", "actual_attendance_hours", "valid_leave_hours"], limit_page_length=0)
	effective = _prefer_manual_daily_rows(rows) if not batch else rows
	batch_ids = sorted({_cell_text(getattr(row, "import_batch", "")) for row in rows if getattr(row, "import_batch", "")})
	exception_filters = [["import_batch", "in", batch_ids or ["__none__"]]]
	exceptions = frappe.get_all(EXCEPTION_DOCTYPE, filters=exception_filters, fields=["name", "exception_type", "confirmation_status"], limit_page_length=0)
	source_counts = defaultdict(int)
	for row in rows:
		source_counts[getattr(row, "source_kind", "未标注") or "未标注"] += 1
	exception_types = defaultdict(int)
	for row in exceptions:
		exception_types[row.exception_type or "未分类"] += 1
	return {
		"attendance_month": attendance_month,
		"attendance_date": str(getdate(attendance_date)) if attendance_date else "",
		"total_rows": len(effective),
		"attendance_people": len({getattr(row, "employee", "") or getattr(row, "employee_code", "") for row in effective if getattr(row, "employee", "") or getattr(row, "employee_code", "")}),
		"normal_rows": sum(1 for row in effective if getattr(row, "attendance_result", "") == "正常"),
		"anomaly_rows": sum(1 for row in effective if getattr(row, "attendance_result", "") == "异常"),
		"missing_rows": sum(1 for row in effective if getattr(row, "missing_in", 0) or getattr(row, "missing_out", 0)),
		"unmatched_rows": sum(1 for row in effective if not getattr(row, "employee", "")),
		"source_counts": dict(source_counts),
		"exceptions": {
			"total": len(exceptions),
			"pending": sum(1 for row in exceptions if row.confirmation_status == "待确认"),
			"confirmed": sum(1 for row in exceptions if row.confirmation_status == "已确认"),
			"rejected": sum(1 for row in exceptions if row.confirmation_status == "已驳回"),
			"by_type": dict(exception_types),
		},
	}


def _company_apple_records(attendance_month, company):
	start, end = _month_bounds(attendance_month)
	batches = frappe.get_all(ATTENDANCE_BATCH_DOCTYPE, filters={"company": company}, pluck="name")
	if not batches:
		return []
	return frappe.get_all(
		APPLE_RECORD_DOCTYPE,
		filters={"import_batch": ["in", batches], "reward_date": ["between", [start, end]]},
		fields=["*"],
		order_by="reward_date asc",
	)


def _source_summary_metadata(rows):
	batch_ids = sorted({_cell_text(getattr(row, "import_batch", "")) for row in rows if getattr(row, "import_batch", "")})
	batch_records = frappe.get_all(
		ATTENDANCE_BATCH_DOCTYPE,
		filters={"name": ["in", batch_ids]},
		fields=["name", "source_checksum"],
	)
	checksums = sorted({_cell_text(getattr(batch, "source_checksum", "")) for batch in batch_records if getattr(batch, "source_checksum", "")})
	payload = json.dumps({"batches": batch_ids, "source_checksums": checksums}, ensure_ascii=False, sort_keys=True)
	return ",".join(batch_ids), hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _calculate_monthly_values(values):
	sick_half_hours = values["sick_leave_hours"] * 0.5
	paid_leave_makeup_hours = (
		sick_half_hours
		+ values["annual_leave_hours"]
		+ values["work_injury_leave_hours"]
		+ values["reunion_leave_hours"]
		+ values["bereavement_leave_hours"]
		+ values["marriage_leave_hours"]
	)
	actual_clock_attendance_hours = max(
		values["actual_attendance_hours"]
		- sick_half_hours
		- values["annual_leave_hours"]
		- values["work_injury_leave_hours"]
		- values["reunion_leave_hours"]
		- values["bereavement_leave_hours"]
		- values["marriage_leave_hours"],
		0,
	)
	leave_deductible_hours = values["personal_leave_hours"] + sick_half_hours
	workday_rest_leave_hours = values["rest_leave_hours"]
	adjusted_1_5_absence_hours = max(workday_rest_leave_hours - values["overtime_1_5_hours"], 0)
	adjusted_2_absence_hours = max(leave_deductible_hours - values["overtime_2_hours"], 0)
	overtime_1_5_settlement_hours = max(values["overtime_1_5_hours"] - workday_rest_leave_hours, 0)
	overtime_2_settlement_hours = max(values["overtime_2_hours"] - leave_deductible_hours, 0)
	adjusted_working_hours = (
		actual_clock_attendance_hours
		+ paid_leave_makeup_hours
		+ workday_rest_leave_hours
		+ leave_deductible_hours
		- adjusted_1_5_absence_hours
		- adjusted_2_absence_hours
	)
	adjusted_absence_hours = max(values["standard_hours"] - adjusted_working_hours, 0)
	full_attendance_basis = max(values["standard_hours"] - values["actual_attendance_hours"] - values["rest_leave_hours"] - values["reunion_leave_hours"] + sick_half_hours, 0)
	if full_attendance_basis > 48:
		full_attendance_deduction = 200
	elif full_attendance_basis > 32:
		full_attendance_deduction = 150
	elif full_attendance_basis > 16:
		full_attendance_deduction = 100
	elif full_attendance_basis > 0.5:
		full_attendance_deduction = 50
	elif full_attendance_basis > 0:
		full_attendance_deduction = 10
	else:
		full_attendance_deduction = 0
	return {
		"actual_clock_attendance_hours": actual_clock_attendance_hours,
		"paid_leave_makeup_hours": paid_leave_makeup_hours,
		"leave_deductible_hours": leave_deductible_hours,
		"workday_rest_leave_hours": workday_rest_leave_hours,
		"adjusted_1_5_absence_hours": adjusted_1_5_absence_hours,
		"adjusted_2_absence_hours": adjusted_2_absence_hours,
		"adjusted_absence_hours": adjusted_absence_hours,
		"adjusted_working_hours": adjusted_working_hours,
		"overtime_1_5_settlement_hours": overtime_1_5_settlement_hours,
		"overtime_2_settlement_hours": overtime_2_settlement_hours,
		"overtime_3_settlement_hours": values["overtime_3_hours"],
		"night_shift_allowance": values["large_night_shift_count"] * LARGE_NIGHT_SHIFT_ALLOWANCE + values["small_night_shift_count"] * SMALL_NIGHT_SHIFT_ALLOWANCE,
		"full_attendance_deduction": full_attendance_deduction,
		"absence_deduction_hours": values["absent_hours"] * 3,
		"red_apple_penalty": values["red_apples"] * APPLE_UNIT_AMOUNT,
	}


@frappe.whitelist()
def generate_monthly_attendance_summary(company: str, attendance_month: str):
	company = _require_company(company)
	_month_bounds(attendance_month)
	lock = _prepare_month_lock_for_generation(company, attendance_month)
	attendance_lock_version = str(lock.active_version)
	daily_rows = _prefer_manual_daily_rows(_get_month_records(DAY_CHECK_DOCTYPE, "attendance_date", attendance_month, company))
	source_batch_ids, source_checksum = _source_summary_metadata(daily_rows)
	summaries = defaultdict(lambda: defaultdict(float))
	identity = {}
	person_aliases = {}
	for row in daily_rows:
		key = _primary_person_key(row)
		if not key:
			continue
		identity[key] = row
		for alias in _person_keys(row):
			person_aliases[alias] = key
		summaries[key]["standard_hours"] += flt(row.standard_hours)
		summaries[key]["actual_attendance_hours"] += flt(row.actual_attendance_hours)
		summaries[key]["overtime_1_5_hours"] += flt(row.workday_overtime_hours)
		summaries[key]["overtime_2_hours"] += flt(row.restday_overtime_hours)
		summaries[key]["overtime_3_hours"] += flt(row.holiday_overtime_hours)
		summaries[key]["leave_hours"] += flt(row.leave_hours)
		summaries[key]["absent_hours"] += flt(row.absent_hours)
		summaries[key]["personal_leave_hours"] += flt(row.personal_leave_hours)
		summaries[key]["sick_leave_hours"] += flt(row.sick_leave_hours)
		summaries[key]["annual_leave_hours"] += flt(row.annual_leave_hours)
		summaries[key]["work_injury_leave_hours"] += flt(row.work_injury_leave_hours)
		summaries[key]["rest_leave_hours"] += flt(row.rest_leave_hours)
		summaries[key]["reunion_leave_hours"] += flt(row.reunion_leave_hours)
		summaries[key]["bereavement_leave_hours"] += flt(row.bereavement_leave_hours)
		summaries[key]["marriage_leave_hours"] += flt(row.marriage_leave_hours)
		summaries[key]["large_night_shift_count"] += flt(row.large_night_shift_count)
		summaries[key]["small_night_shift_count"] += flt(row.small_night_shift_count)

	for row in _company_apple_records(attendance_month, company):
		if not row.is_valid_approval:
			continue
		key = next((person_aliases[alias] for alias in _person_keys(row) if alias in person_aliases), _primary_person_key(row))
		if not key:
			continue
		identity.setdefault(key, row)
		for alias in _person_keys(row):
			person_aliases.setdefault(alias, key)
		summaries[key]["green_apples"] += flt(row.green_apples)
		summaries[key]["red_apples"] += flt(row.red_apples)
		summaries[key]["apple_reward_amount"] += flt(row.reward_amount)

	existing_summaries = frappe.get_all(
		MONTHLY_SUMMARY_DOCTYPE,
		filters=_attendance_scope_filters(company, attendance_month, attendance_lock_version),
		fields=["name", "employee", "employee_code", "employee_name"],
	)
	existing_by_person = {_primary_person_key(row): row.name for row in existing_summaries if _primary_person_key(row)}
	created = []
	for key, values in summaries.items():
		source = identity[key]
		employee = getattr(source, "employee", None)
		date_of_joining = frappe.db.get_value("Employee", employee, "date_of_joining") if employee else None
		calculated = _calculate_monthly_values(values)
		summary_values = {
				"doctype": MONTHLY_SUMMARY_DOCTYPE,
				"company": company,
				"attendance_month": attendance_month,
				"attendance_lock_version": attendance_lock_version,
				"lock_status": "草稿",
				"source_batch_ids": source_batch_ids,
				"source_checksum": source_checksum,
				"employee": employee,
				"employee_code": getattr(source, "employee_code", ""),
				"employee_name": getattr(source, "employee_name", ""),
				"department": getattr(source, "department", ""),
				"date_of_joining": date_of_joining,
				"standard_hours": values["standard_hours"],
				"actual_attendance_hours": values["actual_attendance_hours"],
				"overtime_1_5_hours": values["overtime_1_5_hours"],
				"overtime_2_hours": values["overtime_2_hours"],
				"overtime_3_hours": values["overtime_3_hours"],
				"leave_hours": values["leave_hours"],
				"absent_hours": values["absent_hours"],
				"personal_leave_hours": values["personal_leave_hours"],
				"sick_leave_hours": values["sick_leave_hours"],
				"annual_leave_hours": values["annual_leave_hours"],
				"work_injury_leave_hours": values["work_injury_leave_hours"],
				"rest_leave_hours": values["rest_leave_hours"],
				"reunion_leave_hours": values["reunion_leave_hours"],
				"large_night_shift_count": values["large_night_shift_count"],
				"small_night_shift_count": values["small_night_shift_count"],
				"actual_clock_attendance_hours": calculated["actual_clock_attendance_hours"],
				"paid_leave_makeup_hours": calculated["paid_leave_makeup_hours"],
				"leave_deductible_hours": calculated["leave_deductible_hours"],
				"workday_rest_leave_hours": calculated["workday_rest_leave_hours"],
				"adjusted_1_5_absence_hours": calculated["adjusted_1_5_absence_hours"],
				"adjusted_2_absence_hours": calculated["adjusted_2_absence_hours"],
				"adjusted_absence_hours": calculated["adjusted_absence_hours"],
				"adjusted_working_hours": calculated["adjusted_working_hours"],
				"overtime_1_5_settlement_hours": calculated["overtime_1_5_settlement_hours"],
				"overtime_2_settlement_hours": calculated["overtime_2_settlement_hours"],
				"overtime_3_settlement_hours": calculated["overtime_3_settlement_hours"],
				"night_shift_allowance": calculated["night_shift_allowance"],
				"full_attendance_deduction": calculated["full_attendance_deduction"],
				"absence_deduction_hours": calculated["absence_deduction_hours"],
				"green_apples": values["green_apples"],
				"red_apples": values["red_apples"],
				"apple_reward_amount": values["apple_reward_amount"],
				"red_apple_penalty": calculated["red_apple_penalty"],
				"status": "草稿",
		}
		existing_name = existing_by_person.get(key)
		if existing_name:
			doc = frappe.get_doc(MONTHLY_SUMMARY_DOCTYPE, existing_name)
			doc.update(summary_values)
			doc.save(ignore_permissions=True)
		else:
			doc = frappe.get_doc(summary_values)
			doc.insert(ignore_permissions=True)
		created.append(doc.name)

	lock.source_batch_ids = source_batch_ids
	lock.source_checksum = source_checksum
	lock.save(ignore_permissions=True)
	frappe.db.commit()
	return {"created": len(created), "summaries": created, "company": company, "attendance_lock_version": attendance_lock_version}


@frappe.whitelist()
def lock_attendance_month(company: str, attendance_month: str, reason: str = ""):
	company = _require_company(company)
	_month_bounds(attendance_month)
	lock = _get_or_create_month_lock(company, attendance_month)
	if lock.status == "已锁定":
		frappe.throw(_("该公司考勤月份已经锁定。"))
	_assert_month_ready_for_lock(company, attendance_month)
	scope = _attendance_scope_filters(company, attendance_month, str(lock.active_version))
	summaries = frappe.get_all(MONTHLY_SUMMARY_DOCTYPE, filters=scope, fields=["name"])
	if not summaries:
		frappe.throw(_("没有可锁定的月度考勤草稿。"))
	locked_on = now_datetime()
	lock.status = "已锁定"
	lock.locked_by = frappe.session.user
	lock.locked_on = locked_on
	lock.save(ignore_permissions=True)
	for summary in summaries:
		frappe.db.set_value(
			MONTHLY_SUMMARY_DOCTYPE,
			summary.name,
			{"lock_status": "已锁定", "locked_by": frappe.session.user, "locked_on": locked_on},
		)
	audit_name = _append_lock_audit(lock, "锁定", reason)
	frappe.db.commit()
	return {
		"month_lock": lock.name,
		"company": company,
		"attendance_month": attendance_month,
		"attendance_lock_version": str(lock.active_version),
		"audit": audit_name,
	}


@frappe.whitelist()
def unlock_attendance_month(company: str, attendance_month: str, reason: str):
	company = _require_company(company)
	_month_bounds(attendance_month)
	reason = (reason or "").strip()
	if not reason:
		frappe.throw(_("解锁考勤月份必须填写原因。"))
	lock = _get_or_create_month_lock(company, attendance_month)
	if lock.status != "已锁定":
		frappe.throw(_("只有已锁定的考勤月份可以解锁。"))
	lock.status = "已重开"
	lock.reopened_by = frappe.session.user
	lock.reopened_on = now_datetime()
	lock.save(ignore_permissions=True)
	audit_name = _append_lock_audit(lock, "解锁", reason)
	frappe.db.commit()
	return {
		"month_lock": lock.name,
		"company": company,
		"attendance_month": attendance_month,
		"attendance_lock_version": str(lock.active_version),
		"audit": audit_name,
	}


def _attendance_demo_employee(employee_code):
	return frappe.db.get_value(
		"Employee",
		{"custom_employee_code": employee_code, "company": TEST_ATTENDANCE_DEMO_COMPANY},
		["name", "employee_name", "department"],
		as_dict=True,
	)


def _get_or_create_attendance_demo_batch():
	name = frappe.db.get_value(
		ATTENDANCE_BATCH_DOCTYPE,
		{"company": TEST_ATTENDANCE_DEMO_COMPANY, "source_checksum": TEST_ATTENDANCE_DEMO_CHECKSUM},
		"name",
	)
	if name:
		return frappe.get_doc(ATTENDANCE_BATCH_DOCTYPE, name), True
	batch = frappe.get_doc(
		{
			"doctype": ATTENDANCE_BATCH_DOCTYPE,
			"company": TEST_ATTENDANCE_DEMO_COMPANY,
			"attendance_month": TEST_ATTENDANCE_DEMO_MONTH,
			"source_type": "test_attendance_demo",
			"source_checksum": TEST_ATTENDANCE_DEMO_CHECKSUM,
			"status": "已导入",
			"notes": "TEST-HRMS attendance demo: raw DingTalk source plus manual adjustment.",
		}
	)
	batch.insert(ignore_permissions=True)
	return batch, False


def _seed_attendance_demo_day_check(batch, row, source_kind, source_row):
	employee_code = row["工号"]
	if frappe.db.exists(
		DAY_CHECK_DOCTYPE,
		{
			"import_batch": batch.name,
			"employee_code": employee_code,
			"attendance_date": _parse_date(row["日期"]),
			"source_kind": source_kind,
		},
	):
		return False
	return bool(
		_insert_day_check(
			batch.name,
			{**row, "_source_row": source_row},
			TEST_ATTENDANCE_DEMO_COMPANY,
			source_kind,
			"TEST-HRMS attendance demo",
			1,
		)
	)


@frappe.whitelist()
def seed_test_attendance_demo(dry_run: int | str = 0):
	"""Create an idempotent attendance-only closure in TEST-HRMS / 2099-02."""
	dry_run = bool(int(dry_run or 0))
	if not frappe.db.exists("Company", TEST_ATTENDANCE_DEMO_COMPANY):
		frappe.throw(_("请先创建 TEST-HRMS 演示公司及员工种子。"))
	people = {code: _attendance_demo_employee(code) for code in ("TEST-REG-003", "TEST-MOV-007")}
	missing_people = [code for code, employee in people.items() if not employee]
	if missing_people:
		frappe.throw(_("缺少 TEST-HRMS 演示员工：{0}").format("、".join(missing_people)))
	if dry_run:
		return {
			"company": TEST_ATTENDANCE_DEMO_COMPANY,
			"attendance_month": TEST_ATTENDANCE_DEMO_MONTH,
			"dry_run": True,
			"would_create": ["导入批次", "钉钉原始日统计", "人工调整日统计", "考勤异常", "月度终稿", "月度锁定", "锁定审计"],
		}

	lock = _get_or_create_month_lock(TEST_ATTENDANCE_DEMO_COMPANY, TEST_ATTENDANCE_DEMO_MONTH)
	if lock.status == "已锁定":
		return get_test_attendance_demo_status()

	batch, batch_exists = _get_or_create_attendance_demo_batch()
	regular = people["TEST-REG-003"]
	moved = people["TEST-MOV-007"]
	rows = [
		(
			"钉钉原始导出",
			2,
			{
				"姓名": regular.employee_name,
				"工号": "TEST-REG-003",
				"日期": f"{TEST_ATTENDANCE_DEMO_MONTH}-03",
				"班次": "08:00-17:00",
				"上班时间": "08:00",
				"下班时间": "17:00",
				"标准工时": 8,
				"实际出勤(小时)": 8,
				"实际部门": regular.department,
			},
		),
		(
			"人工调整",
			3,
			{
				"姓名": regular.employee_name,
				"工号": "TEST-REG-003",
				"日期": f"{TEST_ATTENDANCE_DEMO_MONTH}-03",
				"班次": "08:00-17:00",
				"上班时间": "08:00",
				"下班时间": "16:30",
				"标准工时": 8,
				"实际出勤(小时)": 7.5,
				"请假/事假(小时)": 0.5,
				"关联审批单": "DEMO-LEAVE-2099-02-03",
				"实际部门": regular.department,
			},
		),
		(
			"钉钉原始导出",
			4,
			{
				"姓名": moved.employee_name,
				"工号": "TEST-MOV-007",
				"日期": f"{TEST_ATTENDANCE_DEMO_MONTH}-04",
				"班次": "08:00-17:00",
				"上班时间": "08:00",
				"下班时间": "",
				"下班缺卡": "是",
				"标准工时": 8,
				"实际出勤(小时)": 8,
				"实际部门": moved.department,
			},
		),
	]
	created_days = sum(1 for source_kind, source_row, row in rows if _seed_attendance_demo_day_check(batch, row, source_kind, source_row))
	batch.daily_sheet_rows = len(rows)
	batch.save(ignore_permissions=True)
	exception_result = generate_attendance_exceptions(batch.name)
	for name in frappe.get_all(EXCEPTION_DOCTYPE, filters={"import_batch": batch.name}, pluck="name"):
		frappe.db.set_value(
			EXCEPTION_DOCTYPE,
			name,
			{"confirmation_status": "已确认", "confirmed_by": frappe.session.user, "confirmed_on": now_datetime(), "remarks": "TEST-HRMS 演示：日核对已确认。"},
		)
	monthly_result = generate_monthly_attendance_summary(TEST_ATTENDANCE_DEMO_COMPANY, TEST_ATTENDANCE_DEMO_MONTH)
	department_confirmation = _ensure_monthly_department_confirmations(TEST_ATTENDANCE_DEMO_COMPANY, TEST_ATTENDANCE_DEMO_MONTH)
	for name in frappe.get_all(
		DEPARTMENT_CONFIRMATION_DOCTYPE,
		filters={
			"company": TEST_ATTENDANCE_DEMO_COMPANY,
			"attendance_month": TEST_ATTENDANCE_DEMO_MONTH,
			"confirmation_scope": "月度部门工时",
			"attendance_lock_version": monthly_result["attendance_lock_version"],
		},
		pluck="name",
	):
		frappe.db.set_value(
			DEPARTMENT_CONFIRMATION_DOCTYPE,
			name,
			{"confirmation_status": "已确认", "confirmed_by": frappe.session.user, "confirmed_on": now_datetime(), "remarks": "TEST-HRMS 演示：部门月度工时已确认。"},
		)
	lock_result = lock_attendance_month(TEST_ATTENDANCE_DEMO_COMPANY, TEST_ATTENDANCE_DEMO_MONTH, "TEST-HRMS 演示月度确认")
	frappe.db.commit()
	return {
		"company": TEST_ATTENDANCE_DEMO_COMPANY,
		"attendance_month": TEST_ATTENDANCE_DEMO_MONTH,
		"batch": batch.name,
		"batch_existing": batch_exists,
		"created_day_checks": created_days,
		"exceptions": exception_result["created"],
		"department_confirmations": department_confirmation,
		"monthly": monthly_result,
		"lock": lock_result,
	}


@frappe.whitelist()
def get_test_attendance_demo_status():
	"""Read-only inventory for the TEST-HRMS attendance closure."""
	company = TEST_ATTENDANCE_DEMO_COMPANY
	month = TEST_ATTENDANCE_DEMO_MONTH
	batches = frappe.get_all(ATTENDANCE_BATCH_DOCTYPE, filters={"company": company, "attendance_month": month}, fields=["name", "status", "daily_sheet_rows", "source_checksum"])
	batch_ids = [batch.name for batch in batches]
	return {
		"company": company,
		"attendance_month": month,
		"batches": batches,
		"day_checks": frappe.get_all(DAY_CHECK_DOCTYPE, filters=[["company", "=", company], ["attendance_date", ">=", f"{month}-01"], ["attendance_date", "<", "2099-03-01"]], fields=["name", "employee_code", "attendance_date", "source_kind", "actual_attendance_hours", "leave_hours"]),
		"exceptions": frappe.get_all(EXCEPTION_DOCTYPE, filters={"import_batch": ["in", batch_ids or ["__none__"]]}, fields=["name", "employee_code", "exception_type", "confirmation_status"]),
		"month_lock": frappe.db.get_value(MONTH_LOCK_DOCTYPE, {"company": company, "attendance_month": month}, ["name", "status", "active_version", "source_checksum"], as_dict=True),
		"department_confirmations": frappe.get_all(DEPARTMENT_CONFIRMATION_DOCTYPE, filters={"company": company, "attendance_month": month}, fields=["name", "department", "confirmation_status", "attendance_lock_version"]),
		"summaries": frappe.get_all(MONTHLY_SUMMARY_DOCTYPE, filters={"company": company, "attendance_month": month}, fields=["name", "employee_code", "attendance_lock_version", "lock_status", "actual_attendance_hours", "leave_hours"]),
	}


def _list_records(doctype, filters=None, fields=None, page_length=50):
	return frappe.get_all(
		doctype,
		filters=filters or {},
		fields=fields or ["*"],
		order_by="modified desc",
		limit_page_length=int(page_length or 50),
	)


@frappe.whitelist()
def list_attendance_day_checks(
	company: str,
	batch: str = "",
	attendance_month: str = "",
	attendance_date: str = "",
	source_kind: str = "",
	effective_only: int = 1,
	page_length: int = 50,
):
	company = _require_company(company)
	if attendance_month:
		rows = _get_month_records(DAY_CHECK_DOCTYPE, "attendance_date", attendance_month, company)
		if batch:
			rows = [row for row in rows if row.import_batch == batch]
	else:
		filters = {"company": company}
		if batch:
			filters["import_batch"] = batch
		rows = _list_records(DAY_CHECK_DOCTYPE, filters=filters, page_length=page_length)
	if attendance_date:
		selected_date = getdate(attendance_date)
		rows = [row for row in rows if getdate(getattr(row, "attendance_date", selected_date)) == selected_date]
	if source_kind:
		rows = [row for row in rows if getattr(row, "source_kind", "") == source_kind]
	if int(effective_only or 0):
		rows = _prefer_manual_daily_rows(rows)
	return sorted(rows, key=lambda row: (str(getattr(row, "attendance_date", "")), str(getattr(row, "employee_name", ""))), reverse=True)[: int(page_length or 50)]


@frappe.whitelist()
def get_attendance_day_check_review_context(name: str):
	"""Return source and effective values for a human review without opening a raw DocType form."""
	_require_attendance_reviewer()
	day_check = frappe.get_doc(DAY_CHECK_DOCTYPE, name)
	batch = frappe.get_doc(ATTENDANCE_BATCH_DOCTYPE, day_check.import_batch) if day_check.import_batch else None
	try:
		raw_row = json.loads(day_check.raw_row_json or "{}")
	except (TypeError, ValueError):
		raw_row = {"raw_value": day_check.raw_row_json or ""}
	return {
		"day_check": day_check.as_dict(),
		"batch": batch.as_dict() if batch else {},
		"raw_row": raw_row,
		"notice": "人工更正会新增一个版本，不会覆盖或删除原始导入记录。",
	}


def _manual_adjustment_fields():
	return (
		"attendance_date", "employee", "employee_code", "employee_name", "attendance_group", "department", "position", "user_id",
		"date_type", "shift_name", "scheduled_in_time", "scheduled_out_time", "actual_in_time", "actual_out_time",
		"missing_in", "missing_out", "attendance_result", "attendance_duration_hours", "absent_hours", "standard_hours",
		"actual_attendance_hours", "workday_overtime_hours", "restday_overtime_hours", "holiday_overtime_hours",
		"large_night_shift_count", "small_night_shift_count", "leave_summary", "leave_hours", "personal_leave_hours",
		"sick_leave_hours", "annual_leave_hours", "work_injury_leave_hours", "rest_leave_hours", "bereavement_leave_hours",
		"marriage_leave_hours", "public_leave_hours", "maternity_leave_hours", "reunion_leave_hours", "valid_leave_hours",
		"invalid_leave_hours", "valid_leave_summary", "overtime_without_approval", "late_count", "early_count",
	)


@frappe.whitelist()
def create_attendance_manual_adjustment(name: str, changes: str | dict, reason: str):
	"""Create an auditable correction version; raw imports are never overwritten."""
	_require_attendance_reviewer()
	if not (reason or "").strip():
		frappe.throw(_("人工更正必须填写原因。"))
	if isinstance(changes, str):
		changes = json.loads(changes or "{}")
	changes = frappe._dict(changes or {})
	original = frappe.get_doc(DAY_CHECK_DOCTYPE, name)
	attendance_month = getdate(original.attendance_date).strftime("%Y-%m")
	_correction_version_for_import(original.company, attendance_month)

	matching_versions = frappe.get_all(
		DAY_CHECK_DOCTYPE,
		filters={"company": original.company, "employee_code": original.employee_code, "attendance_date": original.attendance_date},
		fields=["correction_version"],
		limit_page_length=0,
	)
	next_version = max([int(flt(getattr(row, "correction_version", 0))) for row in matching_versions] or [0]) + 1
	payload = {fieldname: getattr(original, fieldname, None) for fieldname in _manual_adjustment_fields()}
	for fieldname in _manual_adjustment_fields():
		if fieldname in changes:
			payload[fieldname] = changes[fieldname]
	try:
		original_raw_row = json.loads(original.raw_row_json or "{}")
	except (TypeError, ValueError):
		original_raw_row = {"raw_value": original.raw_row_json or ""}
	doc = frappe.get_doc(
		{
			"doctype": DAY_CHECK_DOCTYPE,
			"import_batch": original.import_batch,
			"company": original.company,
			"source_kind": "人工调整",
			"source_sheet": original.source_sheet,
			"source_row_number": original.source_row_number,
			"correction_version": next_version,
			"adjusted_from": original.name,
			"manual_adjustment_reason": (reason or "").strip(),
			"adjusted_by": frappe.session.user,
			"adjusted_on": now_datetime(),
			"raw_row_json": json.dumps(
				{
					"adjusted_from": original.name,
					"adjustment_reason": (reason or "").strip(),
					"adjusted_by": frappe.session.user,
					"adjusted_on": str(now_datetime()),
					"original_row": original_raw_row,
				},
				ensure_ascii=False,
				default=str,
			),
			**payload,
		}
	)
	doc.insert(ignore_permissions=True)
	frappe.db.commit()
	return {
		"name": doc.name,
		"adjusted_from": original.name,
		"correction_version": doc.correction_version,
		"source_kind": doc.source_kind,
		"notice": "已创建人工更正版本；原始导入记录仍完整保留。",
	}


@frappe.whitelist()
def list_attendance_leave_evidence(batch: str = "", page_length: int = 50):
	filters = {"import_batch": batch} if batch else {}
	return _list_records(LEAVE_EVIDENCE_DOCTYPE, filters=filters, page_length=page_length)


@frappe.whitelist()
def list_attendance_exceptions(company: str = "", attendance_month: str = "", attendance_date: str = "", batch: str = "", confirmation_status: str = "", page_length: int = 50):
	"""List exceptions in the active company/month rather than an unscoped global queue."""
	filters = {}
	if batch:
		filters["import_batch"] = batch
	elif company and attendance_month:
		company = _require_company(company)
		batch_ids = sorted(set(_month_batch_ids(company, attendance_month)))
		filters["import_batch"] = ["in", batch_ids or ["__none__"]]
	if confirmation_status:
		filters["confirmation_status"] = confirmation_status
	rows = _list_records(EXCEPTION_DOCTYPE, filters=filters, page_length=page_length)
	if attendance_date:
		selected_date = getdate(attendance_date)
		rows = [row for row in rows if getdate(row.attendance_date) == selected_date]
	return rows


def _attendance_import_batch_impact(batch_name):
	"""Count only rows derived from one import batch; source files stay untouched."""
	return {
		"day_checks": frappe.db.count(DAY_CHECK_DOCTYPE, {"import_batch": batch_name}),
		"exceptions": frappe.db.count(EXCEPTION_DOCTYPE, {"import_batch": batch_name}),
		"leave_evidence": frappe.db.count(LEAVE_EVIDENCE_DOCTYPE, {"import_batch": batch_name}),
		"apple_records": frappe.db.count(APPLE_RECORD_DOCTYPE, {"import_batch": batch_name}),
	}


def _attendance_import_revoke_blocker(batch_doc):
	if batch_doc.status == "已撤销":
		return "该批次已撤销。"
	if frappe.db.get_value(MONTH_LOCK_DOCTYPE, {"company": batch_doc.company, "attendance_month": batch_doc.attendance_month}, "status") == "已锁定":
		return "该月份已锁定，不能撤销导入。请按解锁与重算流程处理。"
	if frappe.db.exists(MONTHLY_SUMMARY_DOCTYPE, {"company": batch_doc.company, "attendance_month": batch_doc.attendance_month}):
		return "该月份已生成月度考勤终稿。请先按终稿重算流程处理，避免数据前后不一致。"
	return ""


@frappe.whitelist()
def list_attendance_import_batches(company: str, attendance_month: str = "", include_revoked: int = 0, page_length: int = 100):
	"""List import batches with their removable impact before a user can clear test data."""
	_require_attendance_reviewer()
	company = _require_company(company)
	filters = {"company": company}
	if attendance_month:
		_month_bounds(attendance_month)
		filters["attendance_month"] = attendance_month
	if not int(include_revoked or 0):
		filters["status"] = ["!=", "已撤销"]
	batches = frappe.get_all(
		ATTENDANCE_BATCH_DOCTYPE,
		filters=filters,
		fields=["name", "company", "attendance_month", "source_file", "source_type", "status", "daily_sheet_rows", "leave_sheet_rows", "apple_sheet_rows", "imported_by", "imported_on", "creation", "modified"],
		order_by="imported_on desc, creation desc",
		limit_page_length=min(max(int(page_length or 100), 1), 200),
	)
	result = []
	for batch in batches:
		batch_doc = frappe.get_doc(ATTENDANCE_BATCH_DOCTYPE, batch.name)
		blocker = _attendance_import_revoke_blocker(batch_doc)
		batch.update(
			{
				"impact": _attendance_import_batch_impact(batch.name),
				"can_revoke": 0 if blocker else 1,
				"revoke_blocker": blocker,
			}
		)
		result.append(batch)
	return result


@frappe.whitelist()
def revoke_attendance_import_batch(batch: str, reason: str = "", enforce_role: bool = True):
	"""Withdraw a non-locked import batch without deleting its source evidence.

	API raw payloads and uploaded Excel files remain attached to the batch so HR can
	audit what happened.  Only replaceable day drafts, exceptions and derivative
	evidence are removed.
	"""
	if enforce_role:
		_require_attendance_reviewer()
	batch_doc = frappe.get_doc(ATTENDANCE_BATCH_DOCTYPE, batch)
	blocker = _attendance_import_revoke_blocker(batch_doc)
	if batch_doc.status == "已撤销":
		return {"batch": batch_doc.name, "already_revoked": True}
	if blocker:
		frappe.throw(_(blocker))

	deleted = {}
	for doctype in (EXCEPTION_DOCTYPE, DAY_CHECK_DOCTYPE, LEAVE_EVIDENCE_DOCTYPE, APPLE_RECORD_DOCTYPE):
		names = frappe.get_all(doctype, filters={"import_batch": batch_doc.name}, pluck="name", limit_page_length=0)
		for name in names:
			# Generated import rows can be numerous. The regular delete flow queues one
			# ``delete_dynamic_links`` job per row, which can fill the default RQ queue
			# during a rollback and leave the user with a 550 queue-overload error.
			# These rows are generated drafts, so remove their dynamic references inline
			# and retain the import batch + source file as the audit record.
			_delete_generated_import_row(doctype, name)
		deleted[doctype] = len(names)

	try:
		notes = json.loads(batch_doc.notes or "{}")
	except (TypeError, ValueError):
		notes = {"previous_notes": batch_doc.notes or ""}
	notes["revoked_at"] = str(now_datetime())
	notes["revoked_by"] = frappe.session.user
	notes["revoke_reason"] = reason or "人事撤销本次导入"
	batch_doc.status = "已撤销"
	batch_doc.notes = json.dumps(notes, ensure_ascii=False)
	batch_doc.save(ignore_permissions=True)
	return {"batch": batch_doc.name, "status": batch_doc.status, "deleted": deleted}


def _delete_generated_import_row(doctype: str, name: str):
	"""Delete one replaceable import draft without creating an RQ cleanup job."""
	from frappe.model.delete_doc import delete_dynamic_links

	frappe.delete_doc(
		doctype,
		name,
		ignore_permissions=True,
		force=True,
		for_reload=True,
		delete_permanently=True,
	)
	# ``for_reload`` deliberately bypasses Frappe's asynchronous dynamic-link
	# cleanup. Execute the same cleanup in this transaction instead of adding
	# thousands of small default-queue jobs for one rollback.
	delete_dynamic_links(doctype, name)


@frappe.whitelist()
def revoke_latest_attendance_import_batch(company: str, attendance_month: str, reason: str = ""):
	"""Undo the latest active import in the chosen company/month, never globally."""
	_require_attendance_reviewer()
	company = _require_company(company)
	_month_bounds(attendance_month)
	batch_name = frappe.db.get_value(
		ATTENDANCE_BATCH_DOCTYPE,
		{"company": company, "attendance_month": attendance_month, "status": ["!=", "已撤销"]},
		"name",
		order_by="imported_on desc, creation desc",
	)
	if not batch_name:
		frappe.throw(_("当前公司和月份没有可撤回的导入批次。"))
	result = revoke_attendance_import_batch(batch_name, reason=reason or "撤回最近一次导入", enforce_role=False)
	frappe.db.commit()
	return {**result, "message": "已撤回最近一次导入；原始文件和批次记录已保留，派生考勤数据已清除。"}


@frappe.whitelist()
def bulk_revoke_attendance_import_batches(company: str, batches_json: str | list, reason: str = ""):
	"""Clear selected import data atomically after validating every selected batch."""
	_require_attendance_reviewer()
	company = _require_company(company)
	batch_names = json.loads(batches_json) if isinstance(batches_json, str) else batches_json
	if not isinstance(batch_names, list) or not batch_names:
		frappe.throw(_("请至少选择一个导入批次。"))
	batch_names = list(dict.fromkeys(str(name).strip() for name in batch_names if str(name).strip()))
	if len(batch_names) > 100:
		frappe.throw(_("一次最多清除 100 个导入批次。"))
	batch_docs = []
	for batch_name in batch_names:
		batch_doc = frappe.get_doc(ATTENDANCE_BATCH_DOCTYPE, batch_name)
		if batch_doc.company != company:
			frappe.throw(_("不能清除其他公司的导入批次。"))
		blocker = _attendance_import_revoke_blocker(batch_doc)
		if blocker:
			frappe.throw(_("批次 {0} 无法清除：{1}").format(batch_doc.name, blocker))
		batch_docs.append(batch_doc)

	results = [revoke_attendance_import_batch(batch_doc.name, reason=reason or "批量清除测试导入数据", enforce_role=False) for batch_doc in batch_docs]
	frappe.db.commit()
	deleted = defaultdict(int)
	for result in results:
		for doctype, count in (result.get("deleted") or {}).items():
			deleted[doctype] += count
	return {
		"processed": len(results),
		"batches": [result["batch"] for result in results],
		"deleted": dict(deleted),
		"message": "已清除所选批次派生的考勤数据；导入文件和批次审计记录仍保留。",
	}


@frappe.whitelist()
def get_attendance_exception_review_context(name: str):
	_require_attendance_reviewer()
	exception = frappe.get_doc(EXCEPTION_DOCTYPE, name)
	day_check = frappe.get_doc(DAY_CHECK_DOCTYPE, exception.day_check) if exception.day_check else None
	batch = frappe.get_doc(ATTENDANCE_BATCH_DOCTYPE, exception.import_batch) if exception.import_batch else None
	return {
		"exception": exception.as_dict(),
		"day_check": day_check.as_dict() if day_check else {},
		"batch": batch.as_dict() if batch else {},
	}


@frappe.whitelist()
def review_attendance_exception(name: str, decision: str, remarks: str = ""):
	"""Confirm or reject an exception before it can affect a locked month/payroll."""
	_require_attendance_reviewer()
	if decision not in ("confirm", "reject"):
		frappe.throw(_("处理决定必须为确认或驳回。"))
	if decision == "reject" and not (remarks or "").strip():
		frappe.throw(_("驳回异常时必须填写原因，便于审计追溯。"))
	doc = frappe.get_doc(EXCEPTION_DOCTYPE, name)
	doc.confirmation_status = "已确认" if decision == "confirm" else "已驳回"
	doc.confirmed_by = frappe.session.user
	doc.confirmed_on = now_datetime()
	doc.remarks = (remarks or "").strip()
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name, "confirmation_status": doc.confirmation_status, "confirmed_by": doc.confirmed_by, "confirmed_on": doc.confirmed_on}


@frappe.whitelist()
def review_attendance_exceptions(names_json: str | list, decision: str, remarks: str = ""):
	"""Batch review is intentionally limited to a single decision and audit note."""
	_require_attendance_reviewer()
	names = json.loads(names_json) if isinstance(names_json, str) else names_json
	if not isinstance(names, list) or not names:
		frappe.throw(_("请至少选择一条考勤异常。"))
	if len(names) > 200:
		frappe.throw(_("一次最多处理 200 条考勤异常。"))
	results = [review_attendance_exception(name, decision, remarks) for name in names]
	return {"processed": len(results), "results": results}


@frappe.whitelist()
def list_monthly_attendance_summary(company: str, attendance_month: str = "", page_length: int = 50):
	filters = {"company": _require_company(company)}
	if attendance_month:
		filters["attendance_month"] = attendance_month
	return _list_records(MONTHLY_SUMMARY_DOCTYPE, filters=filters, page_length=page_length)


@frappe.whitelist()
def seed_attendance_custom_rules():
	created = []
	for rule in DEFAULT_ATTENDANCE_CUSTOM_RULES:
		existing = frappe.db.get_value(CUSTOM_RULE_DOCTYPE, {"rule_code": rule["rule_code"]}, "name")
		if existing:
			continue
		doc = frappe.get_doc(
			{
				"doctype": CUSTOM_RULE_DOCTYPE,
				"enabled": 1,
				**rule,
			}
		)
		doc.insert(ignore_permissions=True)
		created.append(doc.name)
	frappe.db.commit()
	return {"created": len(created), "rules": created}


@frappe.whitelist()
def list_attendance_custom_rules(rule_group: str = "", enabled_only: int = 0, page_length: int = 100):
	filters = {}
	if rule_group:
		filters["rule_group"] = rule_group
	if int(enabled_only or 0):
		filters["enabled"] = 1
	return _list_records(
		CUSTOM_RULE_DOCTYPE,
		filters=filters,
		fields=[
			"name",
			"rule_code",
			"rule_name",
			"rule_group",
			"rule_type",
			"source_module",
			"source_document",
			"trigger_condition",
			"formula",
			"action_result",
			"priority",
			"enabled",
			"application_mode",
			"last_evaluated_on",
			"last_hit_count",
			"last_evaluation_summary",
		],
		page_length=page_length,
	)


@frappe.whitelist()
def upsert_attendance_custom_rule(rule: str | dict):
	if isinstance(rule, str):
		rule = json.loads(rule)
	rule = frappe._dict(rule or {})
	if not rule.rule_code or not rule.rule_name:
		frappe.throw(_("规则编码和规则名称不能为空"))

	values = {
		"rule_code": rule.rule_code,
		"rule_name": rule.rule_name,
		"rule_group": rule.rule_group or "考勤",
		"rule_type": rule.rule_type or "自定义",
		"source_module": rule.source_module,
		"source_document": rule.source_document,
		"trigger_condition": rule.trigger_condition,
		"formula": rule.formula,
		"action_result": rule.action_result,
		"priority": int(flt(rule.priority)),
		"enabled": 1 if str(rule.enabled) in ("1", "true", "True", "on", "是") else 0,
		"application_mode": rule.application_mode if rule.application_mode in ATTENDANCE_RULE_APPLICATION_MODES else "仅展示",
		"remarks": rule.remarks,
	}
	existing = frappe.db.get_value(CUSTOM_RULE_DOCTYPE, {"rule_code": rule.rule_code}, "name")
	if existing:
		doc = frappe.get_doc(CUSTOM_RULE_DOCTYPE, existing)
		doc.update(values)
	else:
		doc = frappe.get_doc({"doctype": CUSTOM_RULE_DOCTYPE, **values})
	doc.save(ignore_permissions=True)
	frappe.db.commit()
	return {"name": doc.name}


def _late_minutes_for_rule(day_check):
	"""Return a verifiable late duration, or ``None`` when source data is insufficient."""
	shift_start = getattr(day_check, "scheduled_in_time", "") or _parse_shift_start_time(getattr(day_check, "shift_name", ""))
	shift_start_minutes = _time_text_to_minutes(shift_start)
	actual_in_minutes = _time_text_to_minutes(getattr(day_check, "actual_in_time", ""))
	if shift_start_minutes is None or actual_in_minutes is None:
		return None
	return max(actual_in_minutes - shift_start_minutes, 0)


def _rule_hit_rows(rule_code, day_checks):
	"""Return auditable hit rows for reviewed built-in hint rules only.

	The returned reason is displayed to HR so a count is never presented without
	the supporting date, employee and source values.  User-entered formula text is
	never interpreted here.
	"""
	hits = []
	for row in day_checks:
		reason = ""
		if rule_code == "ATT-MISSING-CARD":
			parts = []
			if getattr(row, "missing_in", 0):
				parts.append("上班缺卡")
			if getattr(row, "missing_out", 0):
				parts.append("下班缺卡")
			reason = "、".join(parts)
		elif rule_code == "ATT-ABSENT-NO-LEAVE":
			if (
				flt(getattr(row, "standard_hours", 0)) > 0
				and not getattr(row, "actual_in_time", "")
				and flt(getattr(row, "valid_leave_hours", 0)) <= 0
			):
				reason = "应出勤且无上班打卡、无有效请假"
		elif rule_code == "ATT-LATE-30":
			late_minutes = _late_minutes_for_rule(row)
			if late_minutes is not None and late_minutes > 0 and flt(getattr(row, "valid_leave_hours", 0)) <= 0:
				reason = "实际上班晚于排班开始 %s 分钟" % late_minutes
		if reason:
			hits.append({"row": row, "reason": reason})
	return hits if rule_code in SUPPORTED_ATTENDANCE_HINT_RULE_CODES else None


def _rule_hit_payload(hit):
	row = hit["row"]
	return {
		"day_check": row.name,
		"attendance_date": str(getattr(row, "attendance_date", "") or ""),
		"employee_name": getattr(row, "employee_name", "") or "",
		"employee_code": getattr(row, "employee_code", "") or "",
		"department": getattr(row, "department", "") or "",
		"shift_name": getattr(row, "shift_name", "") or "",
		"scheduled_in_time": getattr(row, "scheduled_in_time", "") or _parse_shift_start_time(getattr(row, "shift_name", "")),
		"actual_in_time": getattr(row, "actual_in_time", "") or "",
		"actual_out_time": getattr(row, "actual_out_time", "") or "",
		"source_kind": getattr(row, "source_kind", "") or "",
		"reason": hit["reason"],
	}


def _rule_evaluation_snapshot(company, attendance_month, attendance_date, day_checks, hit_count):
	return json.dumps(
		{
			"company": company,
			"attendance_month": attendance_month,
			"attendance_date": attendance_date or "整月",
			"effective_day_checks": len(day_checks),
			"hit_count": hit_count,
			"checked_at": str(now_datetime()),
		},
		ensure_ascii=False,
	)


@frappe.whitelist()
def get_attendance_rule_usage_summary(company: str = "", attendance_month: str = ""):
	"""Show each rule's real execution boundary and current data scope."""
	if company:
		_require_company(company)
	rules = list_attendance_custom_rules(page_length=200)
	mode_counts = defaultdict(int)
	for rule in rules:
		mode_counts[getattr(rule, "application_mode", "") or "仅展示"] += 1
		state = _rule_execution_state(rule)
		rule.execution_status = state["status"]
		rule.execution_description = state["description"]
	day_checks = []
	if company and attendance_month:
		day_checks = _prefer_manual_daily_rows(_get_month_records(DAY_CHECK_DOCTYPE, "attendance_date", attendance_month, company))
	return {
		"execution_notice": _rule_execution_notice(),
		"company": company,
		"attendance_month": attendance_month,
		"enabled_rules": sum(1 for rule in rules if getattr(rule, "enabled", 0)),
		"effective_day_check_count": len(day_checks),
		"executable_rule_count": sum(1 for rule in rules if getattr(rule, "execution_status", "") == "可运行"),
		"mode_counts": dict(mode_counts),
		"rules": rules,
		"execution_policy": {
			"仅展示": "仅展示公司制度、7S、KPI、苹果树等规则来源。",
			"导入校验": "仅提示字段或数据质量问题，不修改导入行。",
			"异常提示": "需由人事手工运行提示检查；结果不自动生成扣款、不自动进入薪资。",
		},
	}


@frappe.whitelist()
def evaluate_attendance_rules(company: str, attendance_month: str, attendance_date: str = ""):
	"""Run safe hint checks and retain the exact checked range, never a formula result."""
	_require_attendance_reviewer()
	company = _require_company(company)
	rules = list_attendance_custom_rules(enabled_only=1, page_length=200)
	day_checks = list_attendance_day_checks(
		company=company,
		attendance_month=attendance_month,
		attendance_date=attendance_date,
		effective_only=1,
		page_length=5000,
	)
	results = []
	for rule in rules:
		mode = getattr(rule, "application_mode", "") or "仅展示"
		state = _rule_execution_state(rule)
		if state["status"] != "可运行":
			results.append({"rule_code": rule.rule_code, "rule_name": rule.rule_name, "status": "未执行", "reason": state["description"], "hit_count": 0, "sample_hits": []})
			continue
		hits = _rule_hit_rows(rule.rule_code, day_checks)
		if hits is None:
			results.append({"rule_code": rule.rule_code, "rule_name": rule.rule_name, "status": "未执行", "reason": "该规则尚未接入受控执行器", "hit_count": 0, "sample_hits": []})
			continue
		doc = frappe.get_doc(CUSTOM_RULE_DOCTYPE, rule.name)
		doc.last_evaluated_on = now_datetime()
		doc.last_hit_count = len(hits)
		doc.last_evaluation_summary = _rule_evaluation_snapshot(company, attendance_month, attendance_date, day_checks, len(hits))
		doc.save(ignore_permissions=True)
		results.append({"rule_code": rule.rule_code, "rule_name": rule.rule_name, "status": "已提示", "reason": "只读核对，不改写考勤数据", "hit_count": len(hits), "sample_hits": [_rule_hit_payload(hit) for hit in hits[:5]]})
	frappe.db.commit()
	return {
		"company": company,
		"attendance_month": attendance_month,
		"attendance_date": attendance_date,
		"rules_evaluated": sum(1 for item in results if item["status"] == "已提示"),
		"results": results,
		"execution_notice": _rule_execution_notice(),
	}


@frappe.whitelist()
def get_attendance_rule_hits(company: str, attendance_month: str, rule_code: str, attendance_date: str = "", page_length: int = 200):
	"""Show the current evidence behind a supported rule's prompt count."""
	_require_attendance_reviewer()
	company = _require_company(company)
	rule = frappe.db.get_value(CUSTOM_RULE_DOCTYPE, {"rule_code": rule_code}, ["name", "rule_name", "enabled", "application_mode"], as_dict=True)
	if not rule:
		frappe.throw(_("未找到考勤规则。"))
	if rule_code not in SUPPORTED_ATTENDANCE_HINT_RULE_CODES or not rule.enabled or rule.application_mode != "异常提示":
		return {
			"rule": rule,
			"hits": [],
			"notice": "该规则当前不支持命中明细。只有已启用的内置异常提示规则会读取日核对数据。",
		}
	day_checks = list_attendance_day_checks(
		company=company,
		attendance_month=attendance_month,
		attendance_date=attendance_date,
		effective_only=1,
		page_length=5000,
	)
	hits = _rule_hit_rows(rule_code, day_checks) or []
	return {
		"rule": rule,
		"company": company,
		"attendance_month": attendance_month,
		"attendance_date": attendance_date or "整月",
		"effective_day_check_count": len(day_checks),
		"hits": [_rule_hit_payload(hit) for hit in hits[: int(page_length or 200)]],
		"notice": "以下为当前有效日核对记录的只读命中明细；请进入每日核对或异常处理完成处理。",
	}
