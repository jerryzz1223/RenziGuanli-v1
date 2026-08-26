from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
import re

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import cstr, getdate, nowdate


IMPORT_HEADER_ALIASES = {
	"部门": "source_department",
	"原部门": "source_department",
	"所属部门": "source_department",
	"姓名": "employee_name",
	"员工姓名": "employee_name",
	"可支援部门": "support_department",
	"支援部门": "support_department",
	"可支援岗位": "support_designation",
	"支援岗位": "support_designation",
	"岗位": "support_designation",
	"资格状态": "qualification_status",
	"考核通过日期": "qualified_on",
	"考核日期": "qualified_on",
	"有效开始日期": "valid_from",
	"有效截止日期": "valid_until",
	"备注": "remarks",
}
IMPORT_REQUIRED_FIELDS = {"employee_name", "support_department", "support_designation"}
IMPORT_STATUS_VALUES = {"有效", "待复核", "暂停", "失效"}
CAPABILITY_REQUIRED_FIELD_LABELS = {
	"employee": _("员工"),
	"support_department": _("可支援部门"),
	"support_designation": _("可支援岗位"),
}


class CrossDepartmentSupportCapability(Document):
	"""A single, auditable cross-department qualification for an employee."""

	def validate(self):
		if self.valid_from and self.valid_until and getdate(self.valid_from) > getdate(self.valid_until):
			frappe.throw(_("有效开始日期不能晚于有效截止日期。"))

		missing = [fieldname for fieldname in CAPABILITY_REQUIRED_FIELD_LABELS if not self.get(fieldname)]
		if missing:
			# Import exceptions are deliberately saved as inactive review records so
			# an HR user can complete them in the ledger rather than re-uploading a
			# whole spreadsheet. They must never be made dispatchable prematurely.
			if self.qualification_status == "有效" or self.is_active:
				frappe.throw(_("启用支援能力前请补齐：{0}。").format("、".join(CAPABILITY_REQUIRED_FIELD_LABELS[field] for field in missing)))
			self.is_active = 0
		else:
			# Once an imported exception has been completed, remove the transient
			# import warning. The HR user still explicitly decides whether to set it
			# to 有效 and 可派.
			if not self.is_new() and self.import_validation_note:
				self.import_validation_note = ""

		if self.qualification_status != "有效":
			self.is_active = 0


def _can_read_support_capabilities() -> bool:
	return frappe.has_permission("Cross Department Support Capability", "read")


def _can_create_support_capabilities() -> bool:
	return frappe.has_permission("Cross Department Support Capability", "create")


def _normalise_import_text(value) -> str:
	return re.sub(r"\s+", "", cstr(value or "").replace("（", "(").replace("）", ")"))


def _format_import_value(value) -> str:
	if isinstance(value, datetime):
		return value.date().isoformat()
	if isinstance(value, date):
		return value.isoformat()
	return cstr(value).strip()


def _safe_import_date(value, field_label: str, errors: list[str]):
	"""Keep malformed spreadsheet dates from stopping a staging import."""
	if not value:
		return None
	try:
		return getdate(value)
	except Exception:
		errors.append(_("{0}格式无效：{1}").format(field_label, _format_import_value(value)))
		return None


def _get_import_file_content(file_url: str) -> bytes:
	file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
	if not file_name:
		frappe.throw(_("未找到上传的 Excel 文件。"))
	file_doc = frappe.get_doc("File", file_name)
	file_doc.check_permission("read")
	content = file_doc.get_content()
	return content.encode() if isinstance(content, str) else content


def _find_import_sheet_and_header(workbook):
	candidates = []
	for sheet in workbook.worksheets:
		for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 15), values_only=True), start=1):
			mapping = {}
			for column, value in enumerate(row, start=1):
				fieldname = IMPORT_HEADER_ALIASES.get(_normalise_import_text(value))
				if fieldname and fieldname not in mapping:
					mapping[fieldname] = column
			if IMPORT_REQUIRED_FIELDS.issubset(mapping):
				candidates.append((len(mapping), sheet, row_number, mapping))
	if not candidates:
		frappe.throw(_("未识别到表头。请使用“部门、姓名、可支援部门、可支援岗位”四列的格式。"))
	return max(candidates, key=lambda item: item[0])


def _resolve_link(doctype: str, value: str) -> str:
	"""Resolve an imported display value to a current master-data record name."""
	if frappe.db.exists(doctype, value):
		return value

	meta = frappe.get_meta(doctype)
	title_field = meta.title_field
	if not title_field:
		return ""
	matches = frappe.get_all(doctype, filters={title_field: value}, pluck="name", limit_page_length=2)
	return matches[0] if len(matches) == 1 else ""


def _resolve_employee(employee_name: str, source_department: str) -> tuple[str, str]:
	filters = {"employee_name": employee_name}
	if source_department:
		filters["department"] = source_department
	rows = frappe.get_all("Employee", filters=filters, fields=["name", "employee_name"], limit_page_length=2)
	# The source department in the spreadsheet is a human-readable hint rather
	# than a dependency. When this independent directory uses a department name
	# that has not been created in HRMS, still accept a uniquely named employee.
	if not rows and source_department:
		rows = frappe.get_all("Employee", filters={"employee_name": employee_name}, fields=["name", "employee_name"], limit_page_length=2)
	if len(rows) == 1:
		return rows[0].name, ""
	if len(rows) > 1:
		return "", _("员工“{0}”存在同名记录，请补齐正确的原部门后再导入。").format(employee_name)
	return "", _("未找到员工“{0}”，请先在员工花名册中确认姓名和部门。").format(employee_name)


def _support_import_plan(file_url: str):
	from openpyxl import load_workbook

	workbook = load_workbook(BytesIO(_get_import_file_content(file_url)), data_only=True, read_only=True)
	_score, sheet, header_row, mapping = _find_import_sheet_and_header(workbook)
	rows = []
	last_source_department = ""
	last_employee_name = ""
	last_support_department = ""
	for row_number, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
		data = {
			fieldname: _format_import_value(values[column - 1] if len(values) >= column else "")
			for fieldname, column in mapping.items()
		}
		if not any(data.values()):
			continue

		# The supplied table merges/omits values for a person's second
		# qualification. Carry them down just as an HR user reads it. In
		# particular, an empty support department means the prior row's support
		# department when the employee name is also omitted.
		continued_employee = not data.get("employee_name")
		if data.get("source_department"):
			last_source_department = data["source_department"]
		else:
			data["source_department"] = last_source_department
		if data.get("employee_name"):
			last_employee_name = data["employee_name"]
		else:
			data["employee_name"] = last_employee_name
		if data.get("support_department"):
			last_support_department = data["support_department"]
		elif continued_employee:
			data["support_department"] = last_support_department

		row = {"row_number": row_number, **data, "errors": [], "action": "新增"}
		missing = [field for field in IMPORT_REQUIRED_FIELDS if not data.get(field)]
		if missing:
			row["errors"].append(_("缺少：{0}").format("、".join({
				"employee_name": _("姓名"), "support_department": _("可支援部门"), "support_designation": _("可支援岗位")
			}[field] for field in missing)))
			rows.append(row)
			continue

		source_department = data.get("source_department") or ""
		support_department = data["support_department"]
		support_designation = data["support_designation"]
		employee, employee_error = _resolve_employee(data["employee_name"], source_department)
		if employee_error:
			row["errors"].append(employee_error)

		status = data.get("qualification_status") or "有效"
		if status not in IMPORT_STATUS_VALUES:
			row["errors"].append(_("资格状态应为：{0}。").format("、".join(IMPORT_STATUS_VALUES)))
		row.update({
			"employee": employee,
			"source_department_name": source_department,
			"support_department_name": support_department,
			"support_designation_name": support_designation,
			"qualification_status": status,
		})
		if not row["errors"] and frappe.db.exists(
			"Cross Department Support Capability",
			{"employee": employee, "support_department": support_department, "support_designation": support_designation},
		):
			row["action"] = "跳过（已存在）"
		rows.append(row)

	failed = sum(bool(row["errors"]) for row in rows)
	for row in rows:
		if row["errors"] and not row["action"].startswith("跳过"):
			row["action"] = _("待复核（可导入）")
	return {
		"sheet_name": sheet.title,
		"header_row": header_row,
		"rows": rows,
		"total": len(rows),
		"failed": failed,
		"can_import": bool(rows),
	}


@frappe.whitelist()
def download_cross_department_support_template():
	"""Create a spreadsheet that deliberately mirrors the supplied four-column list."""
	if not _can_create_support_capabilities():
		frappe.throw(_("你没有维护跨部门支援名单的权限。"), frappe.PermissionError)

	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill

	workbook = Workbook()
	sheet = workbook.active
	sheet.title = "跨部门支援"
	sheet.merge_cells("A1:D1")
	sheet["A1"] = "支援人员名单"
	sheet["A1"].font = Font(bold=True, size=16)
	sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
	sheet.row_dimensions[1].height = 28
	headers = ["部门", "姓名", "可支援部门", "可支援岗位", "资格状态", "考核通过日期", "有效开始日期", "有效截止日期", "备注"]
	for column, header in enumerate(headers, start=1):
		cell = sheet.cell(row=2, column=column, value=header)
		cell.font = Font(bold=True, color="FFFFFF")
		cell.fill = PatternFill("solid", fgColor="1F4E78")
		cell.alignment = Alignment(horizontal="center", vertical="center")
		sheet.column_dimensions[cell.column_letter].width = 18
	sheet.freeze_panes = "A3"
	sheet.auto_filter.ref = f"A2:I2"
	output = BytesIO()
	workbook.save(output)
	filename = "跨部门支援名单导入模板.xlsx"
	file_doc = frappe.get_doc({"doctype": "File", "file_name": filename, "content": output.getvalue(), "is_private": 0}).insert(ignore_permissions=True)
	return {"file_url": file_doc.file_url, "file_name": filename}


@frappe.whitelist()
def preview_cross_department_support_import(file_url: str):
	if not _can_create_support_capabilities():
		frappe.throw(_("你没有维护跨部门支援名单的权限。"), frappe.PermissionError)
	return _support_import_plan(file_url)


@frappe.whitelist()
def import_cross_department_support_capabilities(file_url: str):
	if not _can_create_support_capabilities():
		frappe.throw(_("你没有维护跨部门支援名单的权限。"), frappe.PermissionError)

	plan = _support_import_plan(file_url)
	if not plan["can_import"]:
		frappe.throw(_("未读取到可导入的名单行。"))

	inserted = 0
	skipped = 0
	pending_review = 0
	for row in plan["rows"]:
		if row["action"].startswith("跳过"):
			skipped += 1
			continue
		review_errors = list(row["errors"])
		qualified_on = _safe_import_date(row.get("qualified_on"), _("考核通过日期"), review_errors)
		valid_from = _safe_import_date(row.get("valid_from"), _("有效开始日期"), review_errors)
		valid_until = _safe_import_date(row.get("valid_until"), _("有效截止日期"), review_errors)
		if valid_from and valid_until and valid_from > valid_until:
			review_errors.append(_("有效开始日期不能晚于有效截止日期。"))
			valid_until = None
		needs_review = bool(review_errors)
		raw_source_department = row.get("source_department_name") or ""
		# source_department is a Link field fetched from Employee. Never write a
		# display-only Excel department into it: legacy department names (such as
		# 品保课) may not exist in the current Department master and would make
		# Frappe reject the whole import.
		source_department = frappe.db.get_value("Employee", row.get("employee"), "department") if row.get("employee") else ""
		import_note_parts = review_errors
		if raw_source_department and not _resolve_link("Department", raw_source_department):
			import_note_parts.append(_("Excel 原部门：{0}").format(raw_source_department))
		doc = frappe.get_doc({
			"doctype": "Cross Department Support Capability",
			"employee": row.get("employee") or None,
			# Keep the spreadsheet values visible when the employee could not be
			# matched. HR can then select the correct employee directly in the
			# imported record.
			"employee_name": row.get("employee_name") or None,
			"source_department": source_department or None,
			"support_department": row.get("support_department_name") or row.get("support_department") or None,
			"support_designation": row.get("support_designation_name") or row.get("support_designation") or None,
			"qualification_status": "待复核" if needs_review else row["qualification_status"],
			"is_active": 0 if needs_review else 1 if row["qualification_status"] == "有效" else 0,
			"qualified_on": qualified_on,
			"valid_from": valid_from,
			"valid_until": valid_until,
			"remarks": row.get("remarks") or None,
			"import_validation_note": _("Excel 第 {0} 行待修正：{1}").format(row["row_number"], "；".join(import_note_parts)) if needs_review else None,
		})
		# This directory is a query-oriented staging ledger. A record with an
		# unresolved employee or blank target fields must still be retained for
		# later correction in the maintenance ledger.
		doc.insert(ignore_mandatory=True, ignore_links=True)
		inserted += 1
		pending_review += needs_review

	return {
		"inserted": inserted,
		"skipped": skipped,
		"pending_review": pending_review,
		"total": len(plan["rows"]),
		"imported_on": nowdate(),
	}


@frappe.whitelist()
def get_available_support_candidates(
	source_department: str | None = None,
	source_designation: str | None = None,
	support_department: str | None = None,
	support_designation: str | None = None,
	employee_keyword: str | None = None,
	employee: str | None = None,
	include_unavailable: int | str = 0,
	page: int | str = 1,
	page_length: int | str = 10,
):
	"""Return supportable employees for the search page.

	The status and validity period are enforced server-side. This keeps the page,
	list view, exports, and future integration entry points aligned.
	"""
	if not _can_read_support_capabilities():
		frappe.throw(_("你没有查看跨部门支援名单的权限。"), frappe.PermissionError)

	filters: dict[str, object] = {}
	if source_department:
		filters["source_department"] = source_department
	if source_designation:
		filters["source_designation"] = source_designation
	if support_department:
		filters["support_department"] = support_department
	if support_designation:
		filters["support_designation"] = support_designation
	if not frappe.utils.cint(include_unavailable):
		filters.update({"qualification_status": "有效", "is_active": 1})

	rows = frappe.get_list(
		"Cross Department Support Capability",
		filters=filters,
		fields=[
			"name",
			"employee",
			"employee_name",
			"employee_code",
			"source_department",
			"source_designation",
			"support_department",
			"support_designation",
			"qualification_status",
			"is_active",
			"qualified_on",
			"valid_from",
			"valid_until",
			"remarks",
		],
		order_by="support_department asc, support_designation asc, employee_name asc",
		limit_page_length=500,
	)

	today = getdate()
	result = []
	keyword = frappe.utils.cstr(employee_keyword or employee).strip().lower()
	keyword_parts = [part.strip() for part in keyword.split("/") if part.strip()] or [keyword]
	for row in rows:
		employee_values = [frappe.utils.cstr(row.get(field)).lower() for field in ("employee", "employee_name", "employee_code")]
		if keyword and not any(part in value for part in keyword_parts for value in employee_values):
			continue
		is_current = not row.valid_from or getdate(row.valid_from) <= today
		is_current = is_current and (not row.valid_until or getdate(row.valid_until) >= today)
		row["availability"] = "可派" if row.is_active and row.qualification_status == "有效" and is_current else "不可派"
		if frappe.utils.cint(include_unavailable) or row["availability"] == "可派":
			result.append(row)

	total_count = len(result)
	page = max(frappe.utils.cint(page), 1)
	page_length = min(max(frappe.utils.cint(page_length), 1), 100)
	start = (page - 1) * page_length
	return {
		"rows": result[start : start + page_length],
		"count": total_count,
		"page": page,
		"page_length": page_length,
		"total_pages": max((total_count + page_length - 1) // page_length, 1),
	}


@frappe.whitelist()
def get_support_filter_options():
	"""Return the existing values used by the type-ahead query controls."""
	if not _can_read_support_capabilities():
		frappe.throw(_("你没有查看跨部门支援名单的权限。"), frappe.PermissionError)

	rows = frappe.get_all(
		"Cross Department Support Capability",
		fields=["source_department", "source_designation", "support_department", "support_designation", "employee_name", "employee_code"],
		limit_page_length=500,
	)

	def values(fieldname):
		return sorted({frappe.utils.cstr(row.get(fieldname)).strip() for row in rows if row.get(fieldname)})

	employees = sorted(
		{
			"{0} / {1}".format(row.employee_name, row.employee_code)
			if row.employee_code
			else frappe.utils.cstr(row.employee_name)
			for row in rows
			if row.employee_name
		}
	)
	return {
		"source_departments": values("source_department"),
		"source_designations": values("source_designation"),
		"support_departments": values("support_department"),
		"support_designations": values("support_designation"),
		"employees": employees,
	}
