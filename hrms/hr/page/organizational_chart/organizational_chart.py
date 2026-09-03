import json
import re
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path

import frappe
from frappe import _
from frappe.query_builder.functions import Count
from frappe.utils import cint, cstr


HYBRID_MANAGER_KEYWORDS = ("课长", "组长", "主管", "总监", "副总", "经理", "班长", "代理")
TEMPLATE_DEPARTMENT_NODE_TYPES = {"division", "department", "team"}
YONGXIN_Q2_ORG_TEMPLATE = Path(__file__).with_name("yongxin_q2_org_structure.json")
YONGXIN_Q3_DEPARTMENT_HIERARCHY = Path(__file__).with_name("yongxin_q3_department_hierarchy.json")
YONGXIN_Q3_BASELINE_WORKBOOK = Path("/Users/lrj/Documents/SAD/YOngxin/人资/副本人资系统沟通表260713.xlsx")
YONGXIN_ORG_EXPORT_TEMPLATE = Path(
	"/Users/lrj/Documents/SAD/YOngxin/人资/人资二/人资系统资料/人资流程模块/1.人力资源规划/1.2组织架构.xlsx"
)
YONGXIN_Q3_ORG_WORKBOOK_CANDIDATES = (
	Path("/Users/lrj/Documents/SAD/YOngxin/模版/组织架构图260626.xlsx"),
	Path("/home/frappe/frappe-bench/sites/hrms.localhost/private/files/组织架构图260626.xlsx"),
	YONGXIN_ORG_EXPORT_TEMPLATE,
)
YONGXIN_ORG_SHEET_ALIASES = ("26Q3组织架构图", "组织架构图")
WORKBOOK_SNAPSHOT_NODE_ROWS = {
	4: "company_leadership",
	6: "director",
	10: "division",
	13: "department",
	18: "team",
	21: "work_level",
	23: "position_group",
}
WORKBOOK_SNAPSHOT_EMPLOYEE_ROWS = range(24, 29)
YONGXIN_Q3_ORG_SHEET = "26Q3组织架构图"
YONGXIN_Q3_SNAPSHOT_VERSION = "2026Q3"
YONGXIN_Q3_HIERARCHY_CONFIRMATION = "同步2026Q3组织架构"
YONGXIN_COMPANY_NAME = "永新"
YONGXIN_LEGACY_COMPANY_KEYS = ("1",)
ORGANIZATION_TECHNICAL_SUFFIX_RE = re.compile(r"\s*[-－]\s*(?:\d+[A-Za-z]?|1D|11)\s*$", re.IGNORECASE)
GRADE_TAG_CANDIDATES = {"直线级", "间师级", "文师级", "文组级"}

MANUAL_ORGANIZATION_MODE_MESSAGE = _(
	"组织架构现为手动维护模式；请在人事界面逐项创建部门、岗位和任职关系，系统不会再依据原表或员工档案自动生成。"
)

HYBRID_ROSTER_FIELD_MAP = {
	"工号": "custom_employee_code",
	"员工编号": "custom_employee_code",
	"姓名": "employee_name",
	"公司": "company",
	"分支机构": "branch",
	"分公司": "branch",
	"部门": "department",
	"现职务": "designation",
	"职位": "designation",
	"职务": "designation",
	"岗位": "designation",
	"职级": "grade",
	"员工等级": "grade",
	"上级主管": "reports_to",
	"直接上级": "reports_to",
	"汇报对象": "reports_to",
	"在职状态": "status",
	"员工状态": "status",
	"联系电话": "cell_number",
}

DEPARTMENT_QUICK_EDIT_FIELDS = {
	"department_name",
	"company",
	"parent_department",
	"is_group",
	"disabled",
	"leave_block_list",
	"payroll_cost_center",
	"hrms_org_level",
	"hrms_org_role",
	"hrms_org_manager",
	"hrms_org_proxy",
	"hrms_org_card_content",
	"hrms_planned_headcount",
	"hrms_actual_headcount",
	"hrms_vacancy_count",
	"hrms_recruitment_plan",
	"hrms_roster_assignable",
}


def _company_business_weight(company: str):
	"""Measure whether a Yongxin company key actually carries organization data."""
	if not company:
		return 0
	return frappe.db.count("Department", {"company": company}) + frappe.db.count(
		"Employee", {"company": company, "status": "Active"}
	)


@frappe.whitelist()
def preview_multiple_position_organization_import(company: str | None = None, version_code: str | None = None):
	"""Return a no-write staging payload for the Q3 visual organization chart.

	The source workbook is spatial and deliberately contains ambiguous labels.  This
	preview creates only candidates: organization nodes, position nodes, known
	grade-tag candidates, and rows that still need HR confirmation.  It never
	creates Employee, Department, Designation, or assignment records.
	"""
	frappe.throw(MANUAL_ORGANIZATION_MODE_MESSAGE)

	if not frappe.has_permission("Employee", "read"):
		frappe.throw(_("您无权预览包含员工映射信息的组织架构。"))

	source = _resolve_yongxin_org_workbook()
	if not source:
		frappe.throw(_("未找到组织架构原表。"))
	workbook = load_workbook(source, read_only=False, data_only=True)
	sheet = next((workbook[name] for name in YONGXIN_ORG_SHEET_ALIASES if name in workbook.sheetnames), None)
	if not sheet:
		frappe.throw(_("组织架构原表中未找到组织图工作表。"))

	company = _normalize_yongxin_company(company or YONGXIN_COMPANY_NAME) or YONGXIN_COMPANY_NAME
	version_code = cstr(version_code).strip() or YONGXIN_Q3_SNAPSHOT_VERSION
	employee_lookup = _get_employee_lookup(_get_active_employees(company))
	nodes = _parse_workbook_snapshot_nodes(sheet, employee_lookup)
	root = next((node for node in nodes if node.get("node_type") == "company_leadership"), None)
	if not root:
		frappe.throw(_("组织架构原表中未找到总经理节点。"))
	_build_workbook_snapshot_relationships(nodes, root)
	return _build_multiple_position_import_preview(
		nodes=nodes,
		company=company,
		version_code=version_code,
		source_document=source.name,
		source_sheet=sheet.title,
	)


@frappe.whitelist()
def create_multiple_position_organization_draft(company: str | None = None, version_code: str | None = None):
	"""Persist only source-derived draft candidates; never assign or update employees."""
	frappe.throw(MANUAL_ORGANIZATION_MODE_MESSAGE)
	if not frappe.has_permission("Employee", "read"):
		frappe.throw(_("您无权创建包含员工映射信息的组织草稿。"))

	preview = preview_multiple_position_organization_import(company=company, version_code=version_code)
	company = preview["company"]
	version_code = preview["version_code"]
	existing = frappe.db.exists(
		"Organization Structure Version", {"company": company, "version_code": version_code}
	)
	if existing:
		return {
			"version": existing,
			"created": False,
			"message": _("多岗位组织草稿已存在；未重复创建。"),
			"summary": preview["summary"],
		}

	version = frappe.get_doc(
		{
			"doctype": "Organization Structure Version",
			"version_code": version_code,
			"company": company,
			"status": "草稿",
			"source_file_name": preview["source_document"],
			"source_sheet": preview["source_sheet"],
			"source_reference": f"{preview['source_document']}::{preview['source_sheet']}",
		}
	).insert(ignore_permissions=False)

	grade_tags = {}
	for row in preview["grade_tag_candidates"]:
		grade_tag = frappe.get_doc(
			{
				"doctype": "Grade Tag",
				"tag_code": row["tag_code"],
				"tag_name": row["tag_name"],
				"source_version": version.name,
			}
		).insert(ignore_permissions=False)
		grade_tags[row["tag_name"]] = grade_tag.name

	organization_nodes = {}
	for row in preview["organization_nodes"]:
		parent = organization_nodes.get(row.get("parent_node_code"))
		node = frappe.get_doc(
			{
				"doctype": "Organization Node",
				"node_code": row["node_code"],
				"structure_version": version.name,
				"node_type": row["node_type"],
				"display_name": row["display_name"],
				"parent_node": parent,
				"source_sheet": row["source_sheet"],
				"source_cell": row["source_cell"],
				"source_text": row["source_text"],
				"confirmation_status": row["confirmation_status"],
				"planned_headcount": row["planned_headcount"],
				"current_headcount": row["current_headcount"],
				"vacancy_count": row["vacancy_count"],
			}
		).insert(ignore_permissions=False)
		organization_nodes[row["node_code"]] = node.name

	for row in preview["positions"]:
		position = frappe.get_doc(
			{
				"doctype": "Organization Position",
				"position_code": row["position_code"],
				"structure_version": version.name,
				"organization_node": organization_nodes.get(row["organization_node_code"]),
				"position_name": row["position_name"],
				"source_sheet": row["source_sheet"],
				"source_cell": row["source_cell"],
				"source_text": row["source_text"],
				"confirmation_status": row["confirmation_status"],
				"suggested_grade_tags": [
					{"grade_tag": grade_tags[tag]}
					for tag in row["suggested_grade_tags"]
					if tag in grade_tags
				],
			}
		).insert(ignore_permissions=False)

	return {
		"version": version.name,
		"created": True,
		"message": _("已创建多岗位组织草稿；尚未改动任何员工主职或附职。"),
		"summary": preview["summary"],
	}


@frappe.whitelist()
def get_multiple_position_draft_status(company: str | None = None, version_code: str | None = None):
	"""Expose draft progress without mixing it into the published live tree."""
	company = _normalize_yongxin_company(company or YONGXIN_COMPANY_NAME) or YONGXIN_COMPANY_NAME
	version_code = cstr(version_code).strip() or YONGXIN_Q3_SNAPSHOT_VERSION
	version = frappe.db.get_value(
		"Organization Structure Version",
		{"company": company, "version_code": version_code},
		["name", "status", "modified"],
		as_dict=True,
	)
	if not version:
		return {
			"exists": False,
			"company": company,
			"version_code": version_code,
			"message": _("尚未建立多岗位组织草稿；实时架构仍按当前已发布的部门与员工资料显示。"),
		}

	return {
		"exists": True,
		"version": version.name,
		"version_code": version_code,
		"status": version.status,
		"organization_node_count": frappe.db.count("Organization Node", {"structure_version": version.name}),
		"position_count": frappe.db.count("Organization Position", {"structure_version": version.name}),
		"message": _("多岗位组织草稿已建立，仍待人工确认主职、附职和职级标签后发布。"),
	}


def _build_multiple_position_import_preview(nodes, company, version_code, source_document, source_sheet):
	"""Translate parsed visual cards into explicit candidates without guessing data."""
	nodes_by_id = {node["node_id"]: node for node in nodes}
	organization_types = {"company_leadership", "director", "division", "department", "team"}
	organization_type_labels = {
		"company_leadership": "公司",
		"director": "分管",
		"division": "分管",
		"department": "课",
		"team": "组",
	}

	def ancestor(node, allowed_types):
		current = nodes_by_id.get(node.get("parent_node_id"))
		while current:
			if current.get("node_type") in allowed_types:
				return current
			current = nodes_by_id.get(current.get("parent_node_id"))
		return None

	def node_code(node):
		return f"ORG-{version_code}-{node['source_cell']}"

	def position_code(node):
		return f"POS-{version_code}-{node['source_cell']}"

	organization_nodes = []
	positions = []
	grade_tags = set()
	unclassified_level_cards = []
	manager_cards_requiring_confirmation = []
	employee_candidates = []

	for node in nodes:
		node_type = node.get("node_type")
		if node_type in organization_types:
			parent = ancestor(node, organization_types)
			organization_nodes.append(
				{
					"node_code": node_code(node),
					"node_type": organization_type_labels[node_type],
					"display_name": node.get("name"),
					"parent_node_code": node_code(parent) if parent else None,
					"source_sheet": source_sheet,
					"source_cell": node.get("source_cell"),
					"source_text": "\n".join(node.get("lines") or []),
					"planned_headcount": node.get("planned_headcount", 0),
					"current_headcount": node.get("current_headcount", 0),
					"vacancy_count": node.get("vacancy_count", 0),
					"confirmation_status": "待确认",
				}
			)
			if node.get("manager_names") or node.get("proxy_names"):
				manager_cards_requiring_confirmation.append(
					{
						"organization_node_code": node_code(node),
						"source_cell": node.get("source_cell"),
						"manager_names": node.get("manager_names") or [],
						"proxy_names": node.get("proxy_names") or [],
						"reason": _("原表负责人/代理人需人工选择实际岗位节点和任职类型。"),
					}
				)
			continue

		if node_type == "work_level":
			label = cstr(node.get("name")).strip()
			if label in GRADE_TAG_CANDIDATES:
				grade_tags.add(label)
			else:
				unclassified_level_cards.append(
					{
						"source_cell": node.get("source_cell"),
						"label": label,
						"source_text": "\n".join(node.get("lines") or []),
						"reason": _("该行位于职级位置，但名称不属于已确认的非互斥职级标签。"),
					}
				)
			continue

		if node_type == "position_group":
			organization_parent = ancestor(node, organization_types)
			work_level = ancestor(node, {"work_level"})
			level_label = cstr(work_level.get("name") if work_level else "").strip()
			if level_label in GRADE_TAG_CANDIDATES:
				grade_tags.add(level_label)
			positions.append(
				{
					"position_code": position_code(node),
					"organization_node_code": node_code(organization_parent) if organization_parent else None,
					"position_name": node.get("name"),
					"suggested_grade_tags": [level_label] if level_label in GRADE_TAG_CANDIDATES else [],
					"source_sheet": source_sheet,
					"source_cell": node.get("source_cell"),
					"source_text": "\n".join(node.get("lines") or []),
					"confirmation_status": "待确认",
				}
			)
			continue

		if node_type == "employee_group":
			for person in node.get("people") or []:
				if person.get("role") != _("员工"):
					continue
				employee_candidates.append(
					{
						"employee_name": person.get("employee_name"),
						"matched_employee": person.get("matched_employee"),
						"source_cell": node.get("source_cell"),
						"reason": _("人员只生成待映射清单；初始预览不创建主职或附职。"),
					}
				)

	return {
		"write_mode": "preview_only",
		"company": company,
		"version_code": version_code,
		"source_document": source_document,
		"source_sheet": source_sheet,
		"organization_nodes": organization_nodes,
		"positions": positions,
		"grade_tag_candidates": [
			{"tag_code": f"GRADE-{version_code}-{label}", "tag_name": label, "source_version": version_code}
			for label in sorted(grade_tags)
		],
		"unclassified_level_cards": unclassified_level_cards,
		"manager_cards_requiring_confirmation": manager_cards_requiring_confirmation,
		"employee_candidates": employee_candidates,
		"summary": {
			"organization_node_count": len(organization_nodes),
			"position_count": len(positions),
			"grade_tag_candidate_count": len(grade_tags),
			"unclassified_level_count": len(unclassified_level_cards),
			"manager_confirmation_count": len(manager_cards_requiring_confirmation),
			"employee_mapping_count": len(employee_candidates),
		},
	}


def _resolve_yongxin_company_candidates(company: str | None = None):
	"""Return only Yongxin's current and known legacy company keys."""
	candidates = []

	def add(candidate):
		candidate = cstr(candidate).strip()
		if candidate and candidate not in candidates and frappe.db.exists("Company", candidate):
			candidates.append(candidate)

	if company and company != "All Companies" and _is_yongxin_company(company):
		add(company)
	add(YONGXIN_COMPANY_NAME)
	for candidate in frappe.get_all(
		"Company", filters={"company_name": YONGXIN_COMPANY_NAME}, pluck="name", order_by="name asc"
	):
		add(candidate)
	for candidate in YONGXIN_LEGACY_COMPANY_KEYS:
		add(candidate)
	default_company = frappe.defaults.get_user_default("Company")
	if default_company in YONGXIN_LEGACY_COMPANY_KEYS or _is_yongxin_company(default_company):
		add(default_company)
	return candidates


def _normalize_yongxin_company(company: str | None = None):
	candidates = _resolve_yongxin_company_candidates(company)
	if not candidates:
		return None
	# Prefer the Yongxin key that owns the live hierarchy. This repairs sites
	# where an empty "永新" record coexists with legacy data stored under "1".
	return max(candidates, key=lambda candidate: (_company_business_weight(candidate), -candidates.index(candidate)))


def _employee_business_number(employee):
	return cstr(employee.get("custom_employee_code")).strip()


@frappe.whitelist()
def resolve_employee_code(employee_code: str, company: str | None = None):
	"""Resolve the company work number to the Employee link key."""
	lookup_value = cstr(employee_code).strip()
	if not lookup_value or len(lookup_value) > 140 or re.search(r"[/\?#\x00-\x1f]", lookup_value):
		frappe.throw(_("员工编号无效"))

	company = _normalize_yongxin_company(company)
	fields = _get_meta_fields("Employee", ["name", "employee_name", "custom_employee_code", "company"])
	filters = {}
	if company and company != "All Companies":
		filters["company"] = company
	matches = frappe.get_list(
		"Employee",
		fields=fields,
		filters={**filters, "custom_employee_code": lookup_value},
		limit_page_length=3,
	)
	if not matches:
		frappe.throw(_("未找到员工编号 {0} 对应的员工档案").format(lookup_value))
	if len(matches) > 1:
		frappe.throw(_("员工编号 {0} 对应多份档案，请先清理重复编号").format(lookup_value))

	employee = matches[0]
	frappe.get_doc("Employee", employee.name).check_permission("read")
	return {
		"name": employee.name,
		"employee_name": employee.get("employee_name"),
		"employee_code": _employee_business_number(employee),
	}


@frappe.whitelist()
def get_children(parent: str | None = None, company: str | None = None, exclude_node: str | None = None):
	filters = [["status", "=", "Active"]]
	if company and company != "All Companies":
		filters.append(["company", "=", company])

	if parent and company and parent != company:
		filters.append(["reports_to", "=", parent])
	else:
		filters.append(["reports_to", "=", ""])

	if exclude_node:
		filters.append(["name", "!=", exclude_node])

	employees = frappe.get_all(
		"Employee",
		fields=[
			"employee_name as name",
			"name as id",
			"lft",
			"rgt",
			"reports_to",
			"image",
			"designation as title",
		],
		filters=filters,
		order_by="name",
	)

	for employee in employees:
		employee.connections = get_connections(employee.id, employee.lft, employee.rgt)
		employee.expandable = bool(employee.connections)

	return employees


def get_connections(employee: str, lft: int, rgt: int) -> int:
	Employee = frappe.qb.DocType("Employee")
	query = (
		frappe.qb.from_(Employee)
		.select(Count(Employee.name))
		.where((Employee.lft > lft) & (Employee.rgt < rgt) & (Employee.status == "Active"))
	).run()

	return query[0][0]


@frappe.whitelist()
def get_employee_roster_field_map():
	return HYBRID_ROSTER_FIELD_MAP


@frappe.whitelist()
def get_yongxin_q2_org_template_preview():
	seed = _load_yongxin_q2_org_template()
	return {
		"title": seed.get("title"),
		"source_document": seed.get("source_document"),
		"source_sheet": seed.get("source_sheet"),
		"department_count": _count_seed_nodes(seed.get("department_tree") or []),
		"chart_node_count": _count_seed_nodes([seed.get("chart_tree")]) if seed.get("chart_tree") else 0,
		"position_count": len(seed.get("position_templates") or []),
		"staffing_summary_count": len(seed.get("staffing_summary") or []),
	}


@frappe.whitelist()
def preview_yongxin_department_hierarchy(company: str | None = None):
	"""Return a no-write comparison between the Q2 chart and Department records.

	A Company is the business root.  ``All Departments`` stays an ERPNext
	technical root and is never proposed as a business department in this result.
	"""
	company = _normalize_yongxin_company(company)
	departments = _get_departments(company)
	departments_by_label = defaultdict(list)
	for department in departments:
		label = _organization_business_name(department.get("department_name") or department.name)
		if label:
			departments_by_label[label].append(department)

	template_units = _collect_template_department_units(_load_yongxin_q2_org_template().get("chart_tree"))
	team_occurrences = Counter(
		(unit.get("parent_department"), unit.get("department_name"))
		for unit in template_units
		if unit.get("node_type") == "team"
	)
	team_sequence = defaultdict(int)
	matched_departments = set()
	rows = []
	for unit in template_units:
		candidates = departments_by_label.get(unit["department_name"], [])
		expected_parent = unit.get("parent_department") or _get_company_label(company)
		team_key = (unit.get("parent_department"), unit.get("department_name"))
		if unit.get("node_type") == "team":
			team_sequence[team_key] += 1
		team_index = team_sequence[team_key]
		# The source chart uses the same label for the 总办室 course and its
		# internal team.  A Department cannot be its own parent, so the team
		# must be resolved as a position/work-level grouping, never auto-linked.
		if unit.get("node_type") == "team" and unit.get("parent_department") == unit.get("department_name"):
			rows.append(
				{
					**unit,
					"expected_parent": expected_parent,
					"status": "ambiguous",
					"status_label": _("需确认"),
					"message": _("课与组同名，不能将部门设为自己的上级；将作为该课内岗位分组处理。"),
				}
			)
			continue
		if unit.get("node_type") == "team" and team_occurrences[team_key] > 1:
			suggested_name = _suggest_template_team_name(unit, team_index)
			rows.append(
				{
					**unit,
					"expected_parent": expected_parent,
					"suggested_department_name": suggested_name,
					"status": "ambiguous",
					"status_label": _("需确认"),
					"message": _("同一课下存在同名组，需确认正式名称后才能创建。建议：{0}").format(
						suggested_name
					),
				}
			)
			continue
		if not candidates:
			rows.append(
				{
					**unit,
					"expected_parent": expected_parent,
					"suggested_department_name": _suggest_template_team_name(unit),
					"status": "needs_create",
					"status_label": _("待创建"),
					"message": _("架构表中存在该{0}，但当前未建立对应部门记录。").format(
						_("组") if unit.get("node_type") == "team" else _("部门")
					),
				}
			)
			continue
		if len(candidates) > 1:
			rows.append(
				{
					**unit,
					"expected_parent": expected_parent,
					"suggested_department_name": _suggest_template_team_name(unit),
					"status": "ambiguous",
					"status_label": _("需确认"),
					"message": _("当前有 {0} 个同名部门，不能自动确定关系。").format(len(candidates)),
				}
			)
			continue

		department = candidates[0]
		matched_departments.add(department.name)
		actual_parent = cstr(department.get("parent_department")).strip()
		actual_parent_label = (
			_get_company_label(company)
			if _is_all_departments_name(actual_parent)
			else _organization_business_name(actual_parent)
		)
		parent_matches = actual_parent_label == expected_parent
		rows.append(
			{
				**unit,
				"department": department.name,
				"current_parent": actual_parent_label,
				"expected_parent": expected_parent,
				"status": "aligned" if parent_matches else "needs_update",
				"status_label": _("已匹配") if parent_matches else _("待调整"),
				"message": _("当前上级关系正确。") if parent_matches else _("可在确认后调整为建议上级。"),
			}
		)

	for department in departments:
		if department.name in matched_departments:
			continue
		rows.append(
			{
				"node_type": "department",
				"department_name": _organization_business_name(department.get("department_name") or department.name),
				"department": department.name,
				"current_parent": _get_company_label(company)
				if _is_all_departments_name(department.get("parent_department"))
				else _organization_business_name(department.get("parent_department")),
				"expected_parent": "",
				"status": "not_in_template",
				"status_label": _("未纳入架构表"),
				"message": _("当前部门未在 Q2 架构表中找到唯一节点，保持现状。"),
			}
		)

	counts = Counter(row["status"] for row in rows)
	return {
		"company": company,
		"company_label": _get_company_label(company),
		"root_label": _("{0}（公司根节点）").format(_get_company_label(company)),
		"write_mode": False,
		"source_document": _load_yongxin_q2_org_template().get("source_document"),
		"source_sheet": _load_yongxin_q2_org_template().get("source_sheet"),
		"rows": rows,
		"summary": {
			"aligned_count": counts["aligned"],
			"needs_update_count": counts["needs_update"],
			"needs_create_count": counts["needs_create"],
			"ambiguous_count": counts["ambiguous"],
			"unmapped_count": counts["not_in_template"],
		},
	}


@frappe.whitelist()
def preview_yongxin_position_hierarchy(company: str | None = None):
	"""Compare the Q2 role ladder with the current HR Department/Grade/Designation data."""
	company = _normalize_yongxin_company(company)
	employees = _get_active_employees(company)
	employees_by_department = defaultdict(list)
	for employee in employees:
		if employee.get("department"):
			employees_by_department[employee.get("department")].append(employee)

	departments_by_label = defaultdict(list)
	for department in _get_departments(company):
		label = _organization_business_name(department.get("department_name") or department.name)
		if label:
			departments_by_label[label].append(department)

	rows = []
	for position in _load_yongxin_q2_org_template().get("position_templates") or []:
		department_label = _organization_business_name(position.get("department"))
		expected_designation = cstr(position.get("designation")).strip()
		expected_level = _position_hierarchy_level(expected_designation)
		parent_level = _position_hierarchy_level(position.get("parent_designation"))
		department_matches = departments_by_label.get(department_label, [])
		if len(department_matches) != 1:
			rows.append(
				{
					"department": department_label,
					"expected_designation": expected_designation,
					"expected_level": expected_level,
					"parent_designation": position.get("parent_designation"),
					"status": "missing_department" if not department_matches else "ambiguous_department",
					"status_label": _("缺少部门") if not department_matches else _("部门需确认"),
					"message": _("无法唯一匹配当前部门。"),
					"source_cell": position.get("source_cell"),
				}
			)
			continue

		department = department_matches[0]
		current_employees = employees_by_department.get(department.name, [])
		level_employees = [
			employee
			for employee in current_employees
			if _position_hierarchy_level(employee.get("designation")) == expected_level
		]
		matching_employees = [
			employee for employee in level_employees if _matches_template_designation(expected_designation, employee.get("designation"))
		]
		parent_employees = [
			employee
			for employee in current_employees
			if parent_level and _position_hierarchy_level(employee.get("designation")) == parent_level
		]
		if matching_employees and (not parent_level or parent_employees):
			status = "aligned"
		elif level_employees and (not parent_level or parent_employees):
			status = "level_only"
		elif level_employees:
			status = "missing_parent_position"
		else:
			status = "missing_position"
		rows.append(
			{
				"department": department_label,
				"expected_designation": expected_designation,
				"expected_level": expected_level,
				"parent_designation": position.get("parent_designation"),
				"matched_employees": [
					{
						"name": employee.get("employee_name") or employee.name,
						"designation": employee.get("designation"),
						"grade": employee.get("grade") or _("未设置职级"),
					}
					for employee in (matching_employees or level_employees)
				],
				"status": status,
				"status_label": {
					"aligned": _("岗位名称已匹配"),
					"level_only": _("仅职位层级匹配"),
					"missing_parent_position": _("缺少上级职位层级"),
				}.get(status, _("缺少对应职位层级")),
				"message": _position_hierarchy_message(
					expected_designation,
					expected_level,
					parent_level,
					matching_employees,
					level_employees,
					parent_employees,
				),
				"source_cell": position.get("source_cell"),
			}
		)

	counts = Counter(row["status"] for row in rows)
	no_grade_count = sum(1 for employee in employees if not cstr(employee.get("grade")).strip())
	return {
		"company": company,
		"company_label": _get_company_label(company),
		"source_document": _load_yongxin_q2_org_template().get("source_document"),
		"source_sheet": _load_yongxin_q2_org_template().get("source_sheet"),
		"write_mode": False,
		"rows": rows,
		"summary": {
			"aligned_count": counts["aligned"],
			"level_only_count": counts["level_only"],
			"missing_position_count": counts["missing_position"] + counts["missing_parent_position"],
			"missing_department_count": counts["missing_department"] + counts["ambiguous_department"],
			"missing_grade_count": no_grade_count,
		},
	}


def _position_hierarchy_level(designation):
	"""Reduce live and template job titles to the organization-chart role ladder."""
	title = cstr(designation).strip()
	if not title:
		return ""

	normalized = (
		title.replace("（代）", "")
		.replace("(代)", "")
		.replace("代理", "")
		.strip()
	)
	if "总经理" in normalized:
		return "副总经理" if "副总" in normalized else "总经理"
	if "总监" in normalized:
		return "总监"
	if "课长" in normalized:
		return "课长"
	if "主管" in normalized:
		return "主管"
	if "班长" in normalized:
		return "班长"
	if "组长" in normalized:
		return "组长"
	if "直线级" in normalized:
		return "直线级"
	if "间师级" in normalized:
		return "间师级"
	if "文组级" in normalized or "文师级" in normalized:
		return "文师级"
	if normalized in {"QE", "QC", "IQC", "IPQC", "FQC"}:
		return normalized
	if any(keyword in normalized for keyword in ("经理", "主任")):
		return "经理/主任"
	if any(keyword in normalized for keyword in ("助理", "文员", "专员", "工程师", "技术员", "检验员", "作业员", "员工")):
		return "员工"
	return normalized


def _normalize_template_designation(designation):
	return re.sub(r"[（(].*?[）)]", "", cstr(designation)).replace("代理", "").replace(" ", "").strip()


def _matches_template_designation(expected_designation, current_designation):
	"""Require a role-name match when the Q2 template names a specific team position."""
	expected = _normalize_template_designation(expected_designation)
	current = _normalize_template_designation(current_designation)
	if not expected or not current:
		return False

	if expected in {"总经理", "副总", "总监", "课长", "主管", "班长", "组长"}:
		return expected in current or current in expected
	if expected == current:
		return True

	for alternative in re.split(r"、|,|，|/", expected):
		team_name = alternative.replace("组长", "").replace("班长", "").strip()
		if team_name and team_name in current:
			return True
	return False


def _position_hierarchy_message(
	expected_designation,
	expected_level,
	parent_level,
	matching_employees,
	level_employees,
	parent_employees,
):
	if matching_employees and (not parent_level or parent_employees):
		return _("已按“{0}”精确匹配 {1} 名员工。").format(expected_designation, len(matching_employees))
	if level_employees and parent_level and not parent_employees:
		return _("已找到同级职位，但当前部门未找到上级“{0}”层级岗位。").format(parent_level)
	if level_employees:
		return _("已找到“{0}”层级员工，但岗位名称尚未精确对应“{1}”。").format(expected_level, expected_designation)
	return _("当前部门未找到“{0}”层级岗位。").format(expected_level or _("未定义"))
	if parent_level and not parent_employees:
		return _("已匹配 {0} 名员工，但当前部门未找到上级“{1}”层级岗位。").format(
			len(matching_employees), parent_level
		)
	if parent_level:
		return _("已匹配 {0} 名员工，上级“{1}”层级已存在。").format(len(matching_employees), parent_level)
	return _("已匹配 {0} 名员工。").format(len(matching_employees))


@frappe.whitelist()
def preview_yongxin_q3_organization_snapshot(
	source_path: str | None = None,
	company: str | None = None,
	snapshot_version: str | None = None,
):
	"""Read the Q3 source workbook and return a versioned organization precheck without writes."""

	source = Path(source_path or YONGXIN_Q3_BASELINE_WORKBOOK)
	if not source.exists():
		frappe.throw(_("未找到组织基线文件：{0}").format(source))

	workbook = _load_xlsx_workbook(source)
	snapshot_version = snapshot_version or YONGXIN_Q3_SNAPSHOT_VERSION
	company = _normalize_yongxin_company(company or YONGXIN_COMPANY_NAME) or YONGXIN_COMPANY_NAME
	org_sheet = workbook[YONGXIN_Q3_ORG_SHEET] if YONGXIN_Q3_ORG_SHEET in workbook.sheetnames else None

	return {
		"company": company,
		"company_context": {
			"root_company": YONGXIN_COMPANY_NAME,
			"isolation_doctype": "Company",
			"default_locked": _is_yongxin_company(company),
		},
		"snapshot_version": snapshot_version,
		"source_document": source.name,
		"source_path": str(source),
		"organization_sheet": YONGXIN_Q3_ORG_SHEET,
		"write_mode": "preview_only",
		"raw_name_policy": {
			"display_name_uses_business_name": True,
			"raw_name_retained": True,
			"technical_suffix_examples": ["-11", "1D", "- 1D"],
			"dingtalk_path_example": "林俊松-陈文萍-品保课",
		},
		"sheets": [
			{"sheet_name": sheet.title, "row_count": sheet.max_row or 0, "column_count": sheet.max_column or 0}
			for sheet in workbook.worksheets
		],
		"organization_cells": _preview_q3_org_cells(org_sheet),
		"roster_precheck": _preview_organization_sheet(
			workbook,
			"花名册",
			{
				"employee_code": ("工号", "员工编号"),
				"employee_name": ("姓名", "员工姓名"),
				"raw_department_name": ("部门",),
				"raw_designation_name": ("岗位", "职位", "职务", "现职务"),
				"grade": ("职级", "员工等级"),
				"reports_to": ("上级主管", "直接上级", "汇报对象"),
			},
		),
		"dingtalk_precheck": _preview_organization_sheet(
			workbook,
			"每日统计（钉钉导出）",
			{
				"employee_code": ("工号", "员工编号"),
				"employee_name": ("姓名", "员工姓名"),
				"raw_department_name": ("部门",),
				"department_name": ("实际部门",),
				"raw_designation_name": ("职位", "岗位", "职务", "现职务"),
			},
		),
	}


def _load_yongxin_q3_department_hierarchy():
	"""Load the reviewed Q3 folder tree, independent of a server-local workbook path."""
	if not YONGXIN_Q3_DEPARTMENT_HIERARCHY.exists():
		frappe.throw(_("未找到 2026Q3 部门层级配置。"))
	payload = json.loads(YONGXIN_Q3_DEPARTMENT_HIERARCHY.read_text(encoding="utf-8"))
	if not payload.get("nodes"):
		frappe.throw(_("2026Q3 部门层级配置为空。"))
	return payload


def _q3_department_hierarchy_fields():
	return [
		"name",
		"department_name",
		"parent_department",
		"company",
		"is_group",
		"disabled",
		"hrms_org_level",
		"hrms_org_role",
		"hrms_org_manager",
		"hrms_org_proxy",
		"hrms_planned_headcount",
		"hrms_actual_headcount",
		"hrms_vacancy_count",
		"hrms_org_source_cell",
		"hrms_roster_assignable",
	]


def _q3_department_hierarchy_records(company):
	meta = frappe.get_meta("Department")
	fields = [field for field in _q3_department_hierarchy_fields() if field in {item.fieldname for item in meta.fields} or field in {"name", "department_name", "parent_department", "company", "is_group", "disabled"}]
	return frappe.get_all("Department", filters={"company": company}, fields=fields, limit_page_length=0)


def _q3_department_hierarchy_preview(company):
	payload = _load_yongxin_q3_department_hierarchy()
	nodes = payload.get("nodes") or []
	records = _q3_department_hierarchy_records(company)
	by_source = {cstr(row.get("hrms_org_source_cell")).strip(): row for row in records if cstr(row.get("hrms_org_source_cell")).strip()}
	by_name = {cstr(row.get("department_name") or row.get("name")).strip(): row for row in records}
	target_names = {node["name"] for node in nodes}
	target_sources = {node["source_cell"] for node in nodes}
	rows, source_to_target_name = [], {node["source_cell"]: node["name"] for node in nodes}
	source_to_existing_name = {
		node["source_cell"]: (by_source.get(node["source_cell"]) or by_name.get(node["name"]) or {}).get("name")
		for node in nodes
	}
	for node in nodes:
		existing = by_source.get(node["source_cell"]) or by_name.get(node["name"])
		expected_parent = source_to_existing_name.get(node.get("parent_source_cell")) or source_to_target_name.get(node.get("parent_source_cell"), "")
		current_parent = cstr(existing.get("parent_department")).strip() if existing else ""
		parent_matches = bool(existing) and current_parent == expected_parent
		shape_matches = bool(existing) and cint(existing.get("is_group")) == cint(node.get("is_group"))
		roster_matches = not meta_has_field("Department", "hrms_roster_assignable") or bool(existing) and cint(existing.get("hrms_roster_assignable")) == cint(node.get("roster_assignable"))
		rows.append(
			{
				"source_cell": node["source_cell"],
				"source_label": node.get("source_label") or node["name"],
				"department_name": node["name"],
				"node_type": node.get("node_type"),
				"expected_parent": expected_parent,
				"current_parent": current_parent,
				"is_group": cint(node.get("is_group")),
				"roster_assignable": cint(node.get("roster_assignable")),
				"status": "create" if not existing else "aligned" if parent_matches and shape_matches and roster_matches else "update",
			}
		)

	existing_names = {row.name for row in records}
	group_employee_rows = frappe.get_all(
		"Employee",
		filters={"company": company, "status": "Active", "department": ["in", list(existing_names)]} if existing_names else {"name": ["=", ""]},
		fields=["name", "employee_name", "custom_employee_code", "department"],
		limit_page_length=0,
	)
	group_department_names = {
		source_to_existing_name.get(node["source_cell"]) or node["name"]
		for node in nodes
		if cint(node.get("is_group"))
	}
	legacy_employee_assignments = [
		row
		for row in group_employee_rows
		if cstr(row.get("department")).strip() in group_department_names
	]
	return {
		"company": company,
		"version": payload.get("version"),
		"source_document": payload.get("source_document"),
		"source_sheet": payload.get("source_sheet"),
		"confirmation_text": payload.get("confirmation_text") or YONGXIN_Q3_HIERARCHY_CONFIRMATION,
		"write_mode": "preview_only",
		"summary": {
			"node_count": len(nodes),
			"folder_count": len([node for node in nodes if cint(node.get("is_group"))]),
			"roster_leaf_count": len([node for node in nodes if cint(node.get("roster_assignable"))]),
			"create_count": len([row for row in rows if row["status"] == "create"]),
			"update_count": len([row for row in rows if row["status"] == "update"]),
			"aligned_count": len([row for row in rows if row["status"] == "aligned"]),
			"legacy_employee_assignment_count": len(legacy_employee_assignments),
		},
		"nodes": rows,
		"legacy_employee_assignments": legacy_employee_assignments,
		"unmapped_departments": [
			{
				"department": row.name,
				"parent_department": row.get("parent_department"),
				"source_cell": row.get("hrms_org_source_cell"),
			}
			for row in records
			if row.name not in target_names and cstr(row.get("hrms_org_source_cell")).strip() not in target_sources
		],
	}


def meta_has_field(doctype, fieldname):
	return bool(frappe.get_meta(doctype).get_field(fieldname))


@frappe.whitelist()
def preview_yongxin_q3_department_hierarchy(company: str | None = None):
	"""Return the exact folder/leaf plan before it changes any Department records."""
	company = _normalize_yongxin_company(company)
	if not company or not frappe.db.exists("Company", company):
		frappe.throw(_("请选择有效的公司。"))
	return _q3_department_hierarchy_preview(company)


def _q3_department_apply_node(node, company, source_to_department):
	department = frappe.db.get_value(
		"Department", {"company": company, "hrms_org_source_cell": node["source_cell"]}, "name"
	) if meta_has_field("Department", "hrms_org_source_cell") else None
	department = department or frappe.db.get_value("Department", {"company": company, "department_name": node["name"]}, "name")
	doc = frappe.get_doc("Department", department) if department else frappe.new_doc("Department")
	doc.department_name = node["name"]
	doc.company = company
	# Department links store the document name, not the visible label.  The
	# source-to-document mapping is populated parent-first during this import,
	# so a renamed/autonamed Department cannot break the folder relationship.
	doc.parent_department = source_to_department.get(node.get("parent_source_cell"), "")
	doc.is_group = cint(node.get("is_group"))
	doc.disabled = 0
	numeric_fields = {
		"hrms_planned_headcount",
		"hrms_actual_headcount",
		"hrms_vacancy_count",
		"hrms_roster_assignable",
	}
	for fieldname, key in [
		("hrms_org_level", "level"),
		("hrms_org_role", "role"),
		("hrms_org_manager", "manager"),
		("hrms_org_proxy", "proxy"),
		("hrms_planned_headcount", "planned"),
		("hrms_actual_headcount", "current"),
		("hrms_vacancy_count", "vacancy"),
		("hrms_org_source_cell", "source_cell"),
		("hrms_roster_assignable", "roster_assignable"),
	]:
		if meta_has_field("Department", fieldname):
			value = node.get(key)
			doc.set(fieldname, cint(value) if fieldname in numeric_fields else value or "")
	doc.save(ignore_permissions=False)
	return doc.name


@frappe.whitelist()
def import_yongxin_q3_department_hierarchy(
	company: str | None = None, confirmation: str = "", dry_run: int | str = 0
):
	"""Apply the reviewed Q3 folder tree; roster assignments are never guessed."""
	if not frappe.has_permission("Department", "create") and not frappe.has_permission("Department", "write"):
		frappe.throw(_("没有权限同步组织架构。"))
	company = _normalize_yongxin_company(company)
	if not company or not frappe.db.exists("Company", company):
		frappe.throw(_("请选择有效的公司。"))
	preview = _q3_department_hierarchy_preview(company)
	if cint(dry_run):
		return {**preview, "dry_run": True}
	if confirmation != preview["confirmation_text"]:
		frappe.throw(_("请准确输入确认文字“{0}”后再同步。 ").format(preview["confirmation_text"]))

	payload = _load_yongxin_q3_department_hierarchy()
	# Parents are listed before their children in the reviewed Q3 source map.
	# Retain real document names as we save each node, rather than assuming that
	# a Department document name always equals its department_name.
	source_to_department = {}
	savepoint = "yongxin_q3_department_hierarchy"
	frappe.db.savepoint(savepoint)
	try:
		created, updated = [], []
		for node in payload["nodes"]:
			existed = frappe.db.get_value(
				"Department", {"company": company, "hrms_org_source_cell": node["source_cell"]}, "name"
			) if meta_has_field("Department", "hrms_org_source_cell") else None
			existed = existed or frappe.db.exists("Department", {"company": company, "department_name": node["name"]})
			name = _q3_department_apply_node(node, company, source_to_department)
			source_to_department[node["source_cell"]] = name
			(created if not existed else updated).append(name)
		frappe.db.commit()
		frappe.clear_cache(doctype="Department")
		return {
			**_q3_department_hierarchy_preview(company),
			"write_mode": "applied",
			"created_departments": created,
			"updated_departments": updated,
		}
	except Exception:
		frappe.db.rollback(save_point=savepoint)
		raise


@frappe.whitelist()
def import_yongxin_q2_org_structure(company: str | None = None, dry_run: int | str = 0):
	"""Seed Department parent hierarchy and Designation reporting logic from 1.2组织架构.xlsx."""
	frappe.throw(MANUAL_ORGANIZATION_MODE_MESSAGE)

	if not frappe.has_permission("Department", "create") and not frappe.has_permission("Department", "write"):
		frappe.throw(_("没有权限导入组织架构。"))
	company = _normalize_yongxin_company(company)
	if not company:
		frappe.throw(_("请先创建或选择公司。"))
	if not frappe.db.exists("Company", company):
		frappe.throw(_("公司 {0} 不存在。").format(company))

	seed = _load_yongxin_q2_org_template()
	preview = get_yongxin_q2_org_template_preview()
	if cint(dry_run):
		preview.update({"company": company, "dry_run": True})
		return preview

	result = {
		"company": company,
		"title": seed.get("title"),
		"created_departments": [],
		"updated_departments": [],
		"created_designations": [],
		"updated_designations": [],
	}
	source_to_department = {}
	seen_department_labels = defaultdict(int)

	for node in seed.get("department_tree") or []:
		_import_department_node(
			node=node,
			company=company,
			parent_department=None,
			result=result,
			source_to_department=source_to_department,
			seen_department_labels=seen_department_labels,
		)

	_apply_staffing_summary(seed.get("staffing_summary") or [], company, result)
	_import_position_templates(seed.get("position_templates") or [], company, source_to_department, result)
	frappe.db.commit()
	return result


@frappe.whitelist()
def get_hybrid_tree(company: str | None = None, source_mode: str | None = None):
	company = _normalize_yongxin_company(company)
	source_mode = cstr(source_mode).strip() or "live"
	if source_mode == "workbook_snapshot":
		return _get_yongxin_workbook_snapshot_tree(company)
	if source_mode == "quarterly_template":
		return _get_yongxin_template_tree(company)

	# The published view reads the same Department folder links shown in the
	# Department list.  The workbook is only used by the explicit, confirmed Q3
	# sync action; it never rewrites this hierarchy during ordinary page loads.
	departments = _get_departments(company)
	employees = _get_active_employees(company)
	staffing = _get_department_staffing(company)

	employees_by_department = defaultdict(list)
	missing_department_count = 0
	for employee in employees:
		department = employee.get("department")
		if department:
			employees_by_department[department].append(employee)
		else:
			missing_department_count += 1

	department_children = defaultdict(list)
	root_departments = []
	department_by_name = {department.name: department for department in departments}
	for department in departments:
		parent = department.get("parent_department")
		if parent and parent in department_by_name:
			department_children[parent].append(department)
		else:
			root_departments.append(department)

	missing_manager_count = sum(
		1
		for department in departments
		if not _get_department_managers(employees_by_department.get(department.name, []))
		and not department.get("hrms_org_manager")
	)
	staffing_summary = _summarize_staffing(root_departments, staffing)
	reported_current_headcount = staffing_summary["current_headcount"]
	current_headcount = len(employees) or reported_current_headcount
	department_nodes = [
		_build_department_node(department, employees_by_department, department_children, staffing)
		for department in root_departments
	]
	root_node = _build_company_root_node(
		company=company,
		department_nodes=department_nodes,
		staffing_summary={
			"planned_headcount": staffing_summary["planned_headcount"],
			"current_headcount": current_headcount,
			"vacancy_count": staffing_summary["vacancy_count"],
		},
		missing_department_count=missing_department_count,
		missing_manager_count=missing_manager_count,
	)
	root_node["connections"] = len(root_node["children"])
	root_node["expandable"] = bool(root_node["children"])

	return {
		"company": company,
		"source_mode": "live",
		"source_label": _("部门管理文件夹树（与部门管理页面使用同一上下级关系）"),
		"root": root_node,
		"summary": {
			"planned_headcount": root_node["planned_headcount"],
			"current_headcount": root_node["current_headcount"],
			"vacancy_count": root_node["vacancy_count"],
			"department_count": len(departments),
			"missing_department_count": missing_department_count,
			"missing_manager_count": missing_manager_count,
		},
		"field_map": HYBRID_ROSTER_FIELD_MAP,
	}


def _get_manual_organization_records(company=None):
	"""Read a standalone chart only; no HR master data is part of this view."""
	if not frappe.db.exists("DocType", "Organization Node"):
		return {"nodes": []}

	versions = frappe.get_all(
		"Organization Structure Version",
		filters={"company": company, "status": ["!=", "已归档"]} if company else {"status": ["!=", "已归档"]},
		pluck="name",
	)
	if not versions:
		return {"nodes": []}

	nodes = frappe.get_all(
		"Organization Node",
		filters={"structure_version": ["in", versions], "confirmation_status": "已确认"},
		fields=["name", "node_code", "node_type", "display_name", "parent_node", "planned_headcount", "current_headcount", "vacancy_count"],
		order_by="creation asc",
	)
	return {"nodes": nodes}


def _build_manual_organization_tree(company, manual):
	nodes_by_name = {node.name: node for node in manual["nodes"]}
	children_by_parent = defaultdict(list)
	for node in manual["nodes"]:
		(children_by_parent[node.parent_node] if node.parent_node in nodes_by_name else children_by_parent[None]).append(node)
	def organization_node(node):
		child_nodes = [organization_node(child) for child in children_by_parent[node.name]]
		return {
			"node_id": f"organization_node:{node.name}", "id": f"organization_node:{node.name}",
			"node_type": "organization_node", "name": node.display_name, "title": node.node_type,
			"people": [], "planned_headcount": cint(node.planned_headcount),
			"current_headcount": cint(node.current_headcount), "vacancy_count": cint(node.vacancy_count),
			"children": child_nodes, "connections": len(child_nodes), "expandable": bool(child_nodes),
		}

	children = [organization_node(node) for node in children_by_parent[None]]
	return {
		"node_id": f"company:{company or 'all'}", "id": f"company:{company or 'all'}", "node_type": "company",
		"name": _("{0}（总公司）").format(_get_company_label(company)), "title": _("人工维护组织根节点"),
		"people": [], "lines": [], "planned_headcount": sum(cint(node.planned_headcount) for node in manual["nodes"] if not node.parent_node),
		"current_headcount": sum(cint(node.current_headcount) for node in manual["nodes"] if not node.parent_node),
		"vacancy_count": sum(cint(node.vacancy_count) for node in manual["nodes"] if not node.parent_node), "children": children,
		"connections": len(children), "expandable": bool(children),
	}


@frappe.whitelist()
def export_organization_chart_excel(company: str | None = None):
	"""Refresh the supplied organization-chart workbook without changing its layout."""
	from copy import copy

	from openpyxl import load_workbook
	from openpyxl.styles import Alignment, Border, Font, Side
	from frappe.utils.file_manager import save_file
	from hrms.utils.export_watermark import save_workbook_with_logo_watermark

	if not YONGXIN_ORG_EXPORT_TEMPLATE.exists():
		frappe.throw(_("未找到组织架构 Excel 模板。"))
	payload = get_hybrid_tree(company=company)
	root = payload.get("root") or {}
	if not root:
		frappe.throw(_("暂无可导出的组织架构。"))
	book = load_workbook(YONGXIN_ORG_EXPORT_TEMPLATE)
	sheet = book["组织架构图"]
	for worksheet in list(book.worksheets):
		if worksheet != sheet:
			book.remove(worksheet)
	_refresh_organization_export_title(sheet, payload)
	_refresh_organization_export_nodes(sheet, root)
	_refresh_organization_export_summary(sheet, get_organization_report(payload.get("company")), copy, Alignment, Border, Font, Side)

	output = BytesIO()
	save_workbook_with_logo_watermark(book, output)
	company_label = _get_company_label(payload.get("company")) or YONGXIN_COMPANY_NAME
	file = save_file(f"{company_label}_组织架构图.xlsx", output.getvalue(), None, None, is_private=1)
	return {"file_url": file.file_url, "file_name": file.file_name}


def _refresh_organization_export_title(sheet, payload):
	for row in sheet.iter_rows():
		for cell in row:
			if "组织架构图" in cstr(cell.value):
				cell.value = f"{_get_company_label(payload.get('company'))}组织架构图"
				return


def _refresh_organization_export_nodes(sheet, root):
	nodes_by_name = {}
	summary_header_row = next(
		(
			cell.row
			for row in sheet.iter_rows()
			for cell in row
			if cstr(cell.value).strip() == "部门"
		),
		sheet.max_row + 1,
	)
	for node in _flatten_organization_export_nodes(root):
		if node.get("node_type") != "department":
			continue
		name = _organization_business_name(node.get("name"))
		if name:
			nodes_by_name[name] = node
	for row in sheet.iter_rows():
		for cell in row:
			if cell.row >= summary_header_row:
				continue
			lines = [line.strip() for line in cstr(cell.value).splitlines() if line.strip()]
			if not lines:
				continue
			node = nodes_by_name.get(_organization_business_name(lines[0]))
			if not node:
				continue
			cell.value = "\n".join(part for part in [node.get("name"), node.get("title"), _organization_export_staffing(node)] if part)
			below = sheet.cell(row=cell.row + 1, column=cell.column)
			if "编制" in cstr(below.value) or "空缺" in cstr(below.value):
				below.value = _organization_export_staffing(node)


def _refresh_organization_export_summary(sheet, report, copy, Alignment, Border, Font, Side):
	department_header = None
	for row in sheet.iter_rows():
		for cell in row:
			if cstr(cell.value).strip() == "部门":
				department_header = cell
				break
		if department_header:
			break
	if not department_header:
		return
	header_row = department_header.row
	columns = {
		"department": department_header.column,
		"planned_headcount": _find_organization_export_header_column(sheet, header_row, "编制人数"),
		"current_headcount": _find_organization_export_header_column(sheet, header_row, "现有人数"),
		"vacancy_count": _find_organization_export_header_column(sheet, header_row, "空缺人数"),
		"fulfillment_rate": _find_organization_export_header_column(sheet, header_row, "岗位满足率"),
		"vacancy_notes": _find_organization_export_header_column(sheet, header_row, "Q2招聘岗位及人数"),
	}
	if not all(columns.values()):
		return
	total_row = next((row for row in range(header_row + 1, sheet.max_row + 1) if cstr(sheet.cell(row=row, column=department_header.column).value).strip() == "汇总"), sheet.max_row + 1)
	rows = report.get("rows") or []
	required_rows = len(rows) + 1
	available_rows = total_row - header_row - 1
	if required_rows > available_rows:
		sheet.insert_rows(total_row, required_rows - available_rows)
		total_row += required_rows - available_rows
	thin = Side(style="thin", color="000000")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	for offset, data in enumerate(rows + [{"department": "汇总", **(report.get("total") or {}), "vacancy_notes": "-"}], start=1):
		row = header_row + offset
		for key, column in columns.items():
			cell = sheet.cell(row=row, column=column, value=data.get(key) if key != "vacancy_notes" else data.get(key) or "-")
			cell.font = copy(sheet.cell(row=header_row + 1, column=column).font) if header_row + 1 < total_row else Font(name="宋体", size=9)
			cell.alignment = Alignment(horizontal="left" if key in {"department", "vacancy_notes"} else "center", vertical="center", wrap_text=True)
			cell.border = border
			if key == "fulfillment_rate" and data.get(key) is not None:
				cell.number_format = "0%"
		sheet.row_dimensions[row].height = 24


def _find_organization_export_header_column(sheet, row, label):
	for cell in sheet[row]:
		if cstr(cell.value).strip() == label:
			return cell.column
	return None


def _flatten_organization_export_nodes(node):
	if not node:
		return []
	result = [node]
	for child in node.get("children") or []:
		result.extend(_flatten_organization_export_nodes(child))
	return result


def _organization_export_staffing(node):
	return _("编制/实际:{0}/{1}\n空缺:{2}").format(
		cint(node.get("planned_headcount")), cint(node.get("current_headcount")), cint(node.get("vacancy_count"))
	)


@frappe.whitelist()
def get_organization_report(company: str | None = None):
	company = _normalize_yongxin_company(company)
	departments = _get_departments(company)
	employees = _get_active_employees(company)
	staffing = _get_department_staffing(company)
	notes = _get_department_vacancy_notes(company)
	employee_counts = defaultdict(int)
	for employee in employees:
		if employee.get("department"):
			employee_counts[employee.get("department")] += 1

	department_by_name = {department.name: department for department in departments}
	rows = []
	for department in departments:
		staffing_row = staffing.get(department.name, {})
		planned = cint(staffing_row.get("planned_headcount"))
		current = employee_counts.get(department.name, 0)
		vacancy = cint(staffing_row.get("vacancy_count")) or max(planned - current, 0)
		rows.append(
			{
				"department": _organization_business_name(department.get("department_name") or department.name),
				"parent_department": _organization_business_name(
					department_by_name.get(department.get("parent_department"), {}).get("department_name")
					if department.get("parent_department") in department_by_name
					else ""
				),
				"level": cint(department.get("hrms_org_level")) or _department_tree_level(department, department_by_name),
				"planned_headcount": planned,
				"current_headcount": current,
				"vacancy_count": vacancy,
				"fulfillment_rate": (current / planned) if planned else None,
				"vacancy_notes": "、".join(notes.get(department.name, []))
				or cstr(staffing_row.get("recruitment_plan")).strip(),
			}
		)

	root_departments = [
		department
		for department in departments
		if not department.get("parent_department") or department.get("parent_department") not in department_by_name
	]
	staffing_summary = _summarize_staffing(root_departments, staffing)
	total_planned = cint(staffing_summary.get("planned_headcount"))
	total_current = len(employees)
	total_vacancy = cint(staffing_summary.get("vacancy_count")) or max(total_planned - total_current, 0)
	return {
		"company": company,
		"title": _("{0}组织报表").format(_get_company_label(company)),
		"columns": ["部门/课别", "编制人数", "现有人数", "空缺人数", "岗位满足率", "备注"],
		"rows": rows,
		"total": {
			"planned_headcount": total_planned,
			"current_headcount": total_current,
			"vacancy_count": total_vacancy,
			"fulfillment_rate": (total_current / total_planned) if total_planned else None,
		},
	}


def _department_tree_level(department, department_by_name):
	level = 1
	seen = {department.name}
	parent = department.get("parent_department")
	while parent and parent in department_by_name and parent not in seen:
		seen.add(parent)
		level += 1
		parent = department_by_name[parent].get("parent_department")
	return level


def _get_department_vacancy_notes(company=None):
	notes = defaultdict(list)
	if not frappe.db.exists("DocType", "Staffing Plan"):
		return notes
	filters = {"company": company} if company else {}
	for plan in frappe.get_all("Staffing Plan", filters=filters, fields=["name", "department"]):
		try:
			doc = frappe.get_doc("Staffing Plan", plan.name)
		except Exception:
			continue
		for row in doc.get("staffing_plan_details") or []:
			vacancies = cint(row.get("vacancies")) or max(
				cint(row.get("number_of_positions")) - cint(row.get("current_count")),
				0,
			)
			if vacancies:
				notes[plan.department].append(_("{0}：{1}人").format(row.get("designation") or _("未设置岗位"), vacancies))
	return notes


@frappe.whitelist()
def get_hybrid_node_detail(
	node_id: str | None = None,
	node_type: str | None = None,
	company: str | None = None,
	search: str | None = None,
	source_mode: str | None = None,
):
	requested_company = cstr(company).strip()
	company = _normalize_yongxin_company(company)
	node_id = node_id or ""
	node_type = node_type or _node_type_from_id(node_id)
	search = (search or "").strip()
	if cstr(source_mode).strip() == "workbook_snapshot":
		return _get_workbook_snapshot_node_detail(node_id, node_type, company, search)
	if not node_id.startswith("department:"):
		template_detail = _get_template_node_detail(
			node_id,
			node_type,
			company,
			search,
			allow_unimported=requested_company == YONGXIN_COMPANY_NAME or _is_yongxin_company(company),
		)
		if template_detail:
			return template_detail

	if node_type == "company":
		employees = _get_node_employees(company=company, search=search)
		return {
			"node_type": "company",
			"node_id": node_id,
			"title": _get_company_label(company),
			"subtitle": _("公司组织总览"),
			"metrics": _company_metrics(company),
			"employees": employees[:100],
			"actions": {"can_add_department": frappe.has_permission("Department", "create")},
		}

	if node_type in {"work_level", "position_group"}:
		return _get_live_group_node_detail(node_id, company, search)

	department = _node_value(node_id)
	if node_type in {"department", "employee_group"}:
		department_doc = frappe._dict()
		if department and frappe.db.exists("Department", department):
			department_doc = frappe.get_cached_doc("Department", department)
		employees = _get_node_employees(department=department, company=company, search=search)
		people = _get_template_people_for_department(department_doc, company)
		staffing = _get_department_staffing(company).get(department, {})
		current_headcount = len(employees) or staffing.get("current_headcount", 0)
		return {
			"node_type": node_type,
			"node_id": node_id,
			"department": department,
			"title": getattr(department_doc, "department_name", None) or department or "未分配部门",
			"subtitle": _department_subtitle(department_doc),
			"metrics": {
				"planned_headcount": staffing.get("planned_headcount", 0),
				"current_headcount": current_headcount,
				"vacancy_count": staffing.get("vacancy_count", max((staffing.get("planned_headcount", 0) or 0) - current_headcount, 0)),
				"employee_count": current_headcount,
			},
			"employees": employees[:100],
			"people": people,
			"relationships": _get_department_relationships(department_doc, company),
			"actions": {
				"can_add_department": frappe.has_permission("Department", "create"),
				"can_edit_department": bool(department) and frappe.has_permission("Department", "write", department),
				"can_delete_department": bool(department) and frappe.has_permission("Department", "delete", department),
			},
		}

	employee = _node_value(node_id)
	if node_type == "manager" and employee and frappe.db.exists("Employee", employee):
		doc = frappe.get_cached_doc("Employee", employee)
		reports = _get_node_employees(manager=employee, company=company, search=search)
		return {
			"node_type": "manager",
			"node_id": node_id,
			"employee": employee,
			"title": doc.employee_name or employee,
			"subtitle": doc.designation or "管理人员",
			"metrics": {
				"direct_report_count": len(reports),
				"employee_count": len(reports) + 1,
				"planned_headcount": 0,
				"current_headcount": len(reports) + 1,
				"vacancy_count": 0,
			},
			"employees": [_employee_row(doc)] + reports[:99],
			"actions": {"can_open_employee": frappe.has_permission("Employee", "read", employee)},
		}

	if node_type == "employee" and employee and frappe.db.exists("Employee", employee):
		doc = frappe.get_cached_doc("Employee", employee)
		manager_name = ""
		if getattr(doc, "reports_to", None):
			manager_name = frappe.db.get_value("Employee", doc.reports_to, "employee_name") or doc.reports_to
		return {
			"node_type": "employee",
			"node_id": node_id,
			"employee": employee,
			"title": doc.employee_name or employee,
			"subtitle": doc.designation or _("员工"),
			"metrics": {"employee_count": 1, "current_headcount": 1},
			"employees": [_employee_row(doc)],
			"relationships": {"reports_to": {"name": doc.reports_to, "label": manager_name} if manager_name else None},
			"actions": {"can_open_employee": frappe.has_permission("Employee", "read", employee)},
		}

	return {
		"node_type": node_type,
		"node_id": node_id,
		"title": "未找到节点",
		"subtitle": "请刷新组织架构图",
		"metrics": {},
		"employees": [],
		"actions": {},
	}


def _get_manual_organization_node_detail(node_id, node_type, company, search):
	"""Details for the standalone chart; it never loads Employee or Department."""
	if node_type != "organization_node":
		return None
	manual = _get_manual_organization_records(company)
	node_name = _node_value(node_id)
	node = next((row for row in manual["nodes"] if row.name == node_name), None)
	if not node:
		return None
	return {
		"node_type": node_type, "node_id": node_id, "title": node.display_name,
		"subtitle": _("{0}；仅用于组织图展示。 ").format(node.node_type),
		"metrics": {"planned_headcount": cint(node.planned_headcount), "current_headcount": cint(node.current_headcount), "vacancy_count": cint(node.vacancy_count)},
		"employees": [], "people": [],
		"actions": {},
	}


def _get_template_people_for_department(department_doc, company=None):
	source_cell = getattr(department_doc, "hrms_org_source_cell", None)
	if not source_cell:
		return []
	seed = _load_yongxin_q2_org_template()
	node = _find_template_node(seed.get("chart_tree"), f"department:{source_cell}")
	if not node:
		return []
	department = frappe._dict(department_doc.as_dict() if hasattr(department_doc, "as_dict") else department_doc)
	return _build_template_people(node, _get_employee_lookup(_get_active_employees(company)), department)


@frappe.whitelist()
def update_department_fields(department: str, values: str | dict):
	if not department or not frappe.db.exists("Department", department):
		frappe.throw(_("部门不存在。"))

	doc = frappe.get_doc("Department", department)
	doc.check_permission("write")
	values = frappe.parse_json(values) if isinstance(values, str) else (values or {})
	if "hrms_org_manager" in values:
		values["hrms_org_manager"] = "、".join(_split_organization_people(values["hrms_org_manager"]))
	if "hrms_org_proxy" in values:
		values["hrms_org_proxy"] = "、".join(_split_organization_people(values["hrms_org_proxy"]))
	if "hrms_org_card_content" in values:
		values["hrms_org_card_content"] = cstr(values["hrms_org_card_content"]).strip()
	meta = frappe.get_meta("Department")
	updated = {}
	target_name = cstr(values.get("department_name") or doc.department_name).strip()
	target_parent = cstr(values.get("parent_department") if "parent_department" in values else doc.parent_department).strip()
	target_is_group = cint(values.get("is_group") if "is_group" in values else doc.is_group)
	target_roster_assignable = cint(
		values.get("hrms_roster_assignable") if "hrms_roster_assignable" in values else doc.get("hrms_roster_assignable")
	)
	if target_name and target_name != doc.name:
		from hrms.overrides.department_identity import validate_department_name_available

		validate_department_name_available(target_name, doc.company, doc.name)
	if target_parent:
		if target_parent == doc.name or _would_create_department_loop(doc.name, target_parent):
			frappe.throw(_("上级部门不能选择当前部门或其下级部门。"))
		parent_doc = frappe.get_doc("Department", target_parent)
		if parent_doc.company != doc.company:
			frappe.throw(_("不能跨公司调整部门层级。"))
		if not cint(parent_doc.is_group):
			frappe.throw(_("上级部门“{0}”必须先设置为文件夹部门。 ").format(parent_doc.department_name))
	if target_is_group and target_roster_assignable:
		frappe.throw(_("文件夹部门不能用于花名册归属；请只在末级部门启用。"))

	for fieldname, value in values.items():
		if fieldname not in DEPARTMENT_QUICK_EDIT_FIELDS or not meta.has_field(fieldname):
			continue
		doc.set(fieldname, value)
		updated[fieldname] = value

	if updated:
		doc.save(ignore_permissions=False)
		if "department_name" in updated and doc.name != target_name:
			from hrms.api.department_identity import rename_department_document

			department = rename_department_document(doc.name, target_name)
			return {
				"name": department,
				"department_name": target_name,
				"updated": updated,
				"message": _("部门名称及正式关联名称已同步更新。"),
			}

	return {
		"name": doc.name,
		"department_name": doc.get("department_name"),
		"updated": updated,
	}


@frappe.whitelist()
def update_employee_group(
	node_id: str,
	fieldname: str,
	new_value: str,
	company: str | None = None,
):
	if fieldname not in {"grade", "designation"}:
		frappe.throw(_("仅支持调整员工职级或岗位。"))
	company = _normalize_yongxin_company(company)
	payload = _node_value(node_id)
	department = payload.split("::", 1)[0]
	employees = [employee for employee in _get_active_employees(company) if employee.get("department") == department]
	managers = _get_department_managers(employees)
	department_row = frappe._dict(name=department, department_name=department)
	levels = _build_work_level_nodes(department_row, employees, managers)
	candidates = levels + [child for level in levels for child in level.get("children", [])]
	group = next((row for row in candidates if row.get("node_id") == node_id), None)
	if not group:
		frappe.throw(_("当前职级或岗位分组已变化，请刷新后重试。"))
	expected_type = "work_level" if fieldname == "grade" else "position_group"
	if group.get("node_type") != expected_type:
		frappe.throw(_("分组类型与调整字段不一致。"))

	new_value = cstr(new_value).strip()
	link_doctype = "Employee Grade" if fieldname == "grade" else "Designation"
	if new_value and not frappe.db.exists(link_doctype, new_value):
		frappe.throw(_("{0} {1} 不存在。").format(_(link_doctype), new_value))

	employee_names = [person.get("employee") for person in _collect_node_people(group) if person.get("employee")]
	updated = []
	for employee_name in employee_names:
		doc = frappe.get_doc("Employee", employee_name)
		doc.check_permission("write")
		if cstr(doc.get(fieldname)).strip() == new_value:
			continue
		doc.set(fieldname, new_value or None)
		doc.save(ignore_permissions=False)
		updated.append(employee_name)

	return {"node_id": node_id, "fieldname": fieldname, "new_value": new_value, "updated": updated}


@frappe.whitelist()
def move_organization_node(node_id: str, target_node_id: str, company: str | None = None):
	"""Persist a tree-builder move in the same records used by Department and Employee.

	There is intentionally no visual-only canvas state: a department move writes
	``parent_department``; a person move writes department/职级/岗位 or 直接上级;
	a position move writes the position's parent.  Therefore the Department list,
	employee archive, and chart always read the same hierarchy.
	"""
	company = _normalize_yongxin_company(company)
	node_type = _node_type_from_id(node_id)
	target_type = _node_type_from_id(target_node_id)
	if not node_id or not target_node_id or node_id == target_node_id:
		frappe.throw(_("请选择不同的有效节点。"))

	if node_type == "department":
		department = _node_value(node_id)
		if target_type not in {"department", "company", "company_leadership"}:
			frappe.throw(_("部门只能拼接到公司根节点或另一个部门下。"))
		if not frappe.db.exists("Department", department):
			frappe.throw(_("部门不存在。"))
		doc = frappe.get_doc("Department", department)
		doc.check_permission("write")
		parent = _node_value(target_node_id) if target_type == "department" else ""
		if parent and not frappe.db.exists("Department", parent):
			frappe.throw(_("请先将该分管节点导入为部门后再拼接。"))
		if parent == doc.name or _would_create_department_loop(doc.name, parent):
			frappe.throw(_("不能把部门移动到自己或自己的下级部门。"))
		if parent:
			parent_doc = frappe.get_doc("Department", parent)
			if parent_doc.company != doc.company:
				frappe.throw(_("不能跨公司调整部门层级。"))
		doc.parent_department = parent
		doc.save(ignore_permissions=False)
		return {"message": _("部门层级已同步到部门管理。"), "updated": [doc.name]}

	if node_type == "employee":
		employee = _node_value(node_id)
		if not frappe.db.exists("Employee", employee):
			frappe.throw(_("员工不存在。"))
		doc = frappe.get_doc("Employee", employee)
		doc.check_permission("write")
		if company and doc.company != company:
			frappe.throw(_("不能跨公司调整员工归属。"))
		if target_type == "employee":
			manager = _node_value(target_node_id)
			if manager == doc.name or not frappe.db.exists("Employee", manager):
				frappe.throw(_("直接上级无效。"))
			if _would_create_employee_reporting_loop(doc.name, manager):
				frappe.throw(_("不能把员工设为自己的下级。"))
			doc.reports_to = manager
		elif target_type in {"department", "work_level", "position_group"}:
			context = _get_live_node_context(target_node_id, company)
			if not context.get("department"):
				frappe.throw(_("目标节点未关联有效部门。"))
			doc.department = context["department"]
			if target_type in {"work_level", "position_group"} and context.get("work_level"):
				doc.grade = context["work_level"] if context["work_level"] != _("直线级") else ""
			if target_type == "position_group" and context.get("designation"):
				doc.designation = context["designation"] if context["designation"] != _("员工") else ""
		else:
			frappe.throw(_("人员可移动到部门、职级、岗位或另一名员工下。"))
		doc.save(ignore_permissions=False)
		return {"message": _("员工归属已同步到花名册与组织树。"), "updated": [doc.name]}

	if node_type == "position_group" and target_type == "position_group":
		source = _get_live_node_context(node_id, company)
		target = _get_live_node_context(target_node_id, company)
		designation = source.get("designation")
		parent_designation = target.get("designation")
		if not designation or not parent_designation or designation == parent_designation:
			frappe.throw(_("岗位上下级关系无效。"))
		if not frappe.db.exists("Designation", designation) or not frappe.db.exists("Designation", parent_designation):
			frappe.throw(_("请先维护岗位主数据后再调整岗位层级。"))
		if _would_create_designation_loop(designation, parent_designation):
			frappe.throw(_("不能把岗位移动到自己的下级岗位。"))
		doc = frappe.get_doc("Designation", designation)
		doc.check_permission("write")
		if frappe.get_meta("Designation").has_field("hrms_parent_designation"):
			doc.hrms_parent_designation = parent_designation
			doc.save(ignore_permissions=False)
			return {"message": _("岗位上下级已同步到组织树。"), "updated": [doc.name]}
		frappe.throw(_("当前站点尚未安装岗位上级字段，请先执行系统迁移。"))

	frappe.throw(_("当前节点类型暂不支持拖动。"))


@frappe.whitelist()
def delete_departments(departments: str | list):
	departments = frappe.parse_json(departments) if isinstance(departments, str) else (departments or [])
	departments = [department for department in departments if department]
	result = {"deleted": [], "failed": []}

	for department in departments:
		try:
			_validate_department_delete(department)
			frappe.delete_doc("Department", department, ignore_permissions=False)
			result["deleted"].append(department)
		except Exception as exc:
			result["failed"].append({"name": department, "message": frappe.utils.cstr(exc)})

	result["deleted_count"] = len(result["deleted"])
	result["failed_count"] = len(result["failed"])
	return result


def _load_yongxin_q2_org_template():
	if not YONGXIN_Q2_ORG_TEMPLATE.exists():
		frappe.throw(_("未找到组织架构模板文件。"))
	return json.loads(YONGXIN_Q2_ORG_TEMPLATE.read_text(encoding="utf-8"))


def _resolve_yongxin_org_workbook():
	"""Use the current Q3 source first, with the in-repository form as a deployable fallback."""
	for source in YONGXIN_Q3_ORG_WORKBOOK_CANDIDATES:
		if source.exists():
			return source
	return None


def _get_yongxin_workbook_snapshot_tree(company=None):
	"""Build a read-only, source-faithful tree from the merged-cell organization chart.

	The workbook is a spatial diagram rather than a normalized list: each branch is
	represented by a merged cell.  We retain every role/person cell and connect it
	to the closest eligible node on the level above, so a page user can inspect the
	exact source hierarchy without first changing live Department or Employee data.
	"""
	from openpyxl import load_workbook

	source = _resolve_yongxin_org_workbook()
	if not source:
		# The Excel file is intentionally optional in deployed sites.  The checked-in
		# snapshot keeps the chart, people and hierarchy readable until the next
		# source workbook is uploaded to the server.
		payload = _get_yongxin_template_tree(company)
		payload.update(
			{
				"source_mode": "quarterly_template",
				"snapshot_fallback": True,
				"source_document": YONGXIN_Q2_ORG_TEMPLATE.name,
				"source_sheet": "内置组织快照",
				"source_label": _("内置组织快照 · 原始 Excel 未部署，当前页面仍可正常查看"),
			}
		)
		return payload
	workbook = load_workbook(source, read_only=False, data_only=True)
	sheet = next((workbook[name] for name in YONGXIN_ORG_SHEET_ALIASES if name in workbook.sheetnames), None)
	if not sheet:
		frappe.throw(_("组织架构原表中未找到组织图工作表。"))

	employee_lookup = _get_employee_lookup(_get_active_employees(company))
	nodes = _parse_workbook_snapshot_nodes(sheet, employee_lookup)
	_apply_workbook_snapshot_card_overrides(nodes, company, employee_lookup)
	root = next((node for node in nodes if node.get("node_type") == "company_leadership"), None)
	if not root:
		frappe.throw(_("组织架构原表中未找到总经理节点。"))
	_build_workbook_snapshot_relationships(nodes, root)
	_summary = _workbook_snapshot_summary(root, nodes)
	return {
		"company": company,
		"root": root,
		"summary": _summary,
		"field_map": HYBRID_ROSTER_FIELD_MAP,
		"source_mode": "workbook_snapshot",
		"source_document": source.name,
		"source_sheet": sheet.title,
		"source_label": _("原表视图 · {0} / {1}").format(source.name, sheet.title),
	}


def _split_organization_people(value):
	"""Accept the department editor's common multi-person separators."""
	if isinstance(value, (list, tuple)):
		values = value
	else:
		values = re.split(r"[、,，;；\n]+", cstr(value))
	return _deduplicate_names(cstr(name).strip() for name in values if cstr(name).strip())


def _apply_workbook_snapshot_card_overrides(nodes, company, employee_lookup):
	"""Layer editable Department card fields over the source-faithful Excel tree.

	The workbook still owns node identity and parent/child layout. A Department
	mapped to the same source cell may only replace card presentation fields; it
	never changes the Excel hierarchy or writes employee assignments.
	"""
	departments_by_source = {
		department.get("hrms_org_source_cell"): department
		for department in _get_departments(company)
		if department.get("hrms_org_source_cell")
	}
	for node in nodes:
		department = departments_by_source.get(node.get("source_cell"))
		if not department:
			continue
		node["department"] = department.name
		node["department_label"] = department.get("department_name") or department.name
		role = cstr(department.get("hrms_org_role")).strip() or node.get("role") or _("负责人")
		manager_names = _split_organization_people(department.get("hrms_org_manager")) or node.get("manager_names") or []
		proxy_names = _split_organization_people(department.get("hrms_org_proxy")) or node.get("proxy_names") or []
		node["role"] = role
		node["manager_names"] = manager_names
		node["proxy_names"] = proxy_names
		node["card_content"] = cstr(department.get("hrms_org_card_content")).strip()
		node["title"] = _workbook_snapshot_title(node.get("lines") or [], node.get("node_type"), manager_names, proxy_names)
		if node["card_content"]:
			node["title"] = " · ".join(part for part in [node["title"], node["card_content"]] if part)
		node["people"] = _build_workbook_snapshot_people(node, employee_lookup)


def _parse_workbook_snapshot_nodes(sheet, employee_lookup):
	nodes = []
	seen_cells = set()
	for merged in sheet.merged_cells.ranges:
		row = merged.min_row
		node_type = WORKBOOK_SNAPSHOT_NODE_ROWS.get(row)
		if row in WORKBOOK_SNAPSHOT_EMPLOYEE_ROWS:
			node_type = "employee_group"
		if not node_type:
			continue
		cell = sheet.cell(row=row, column=merged.min_col)
		value = _organization_source_name(cell.value)
		if not value or cell.coordinate in seen_cells:
			continue
		seen_cells.add(cell.coordinate)
		node = _parse_workbook_snapshot_card(value, node_type, cell.coordinate)
		if not node:
			continue
		node.update(
			{
				"row": row,
				"column_start": merged.min_col,
				"column_end": merged.max_col,
				"column_center": (merged.min_col + merged.max_col) / 2,
				"children": [],
			}
		)
		node["people"] = _build_workbook_snapshot_people(node, employee_lookup)
		nodes.append(node)
	return sorted(nodes, key=lambda node: (node["row"], node["column_center"]))


def _parse_workbook_snapshot_card(value, node_type, source_cell):
	lines = [re.sub(r"\s+", " ", line).strip() for line in value.splitlines() if line.strip()]
	if not lines:
		return None
	name = lines[0]
	planned, current, vacancy = _workbook_snapshot_staffing(value)
	manager_names, proxy_names = _workbook_snapshot_manager_names(lines)
	employee_names = _workbook_snapshot_employee_names(lines, node_type)
	return {
		"node_id": f"snapshot:{source_cell}",
		"id": f"snapshot:{source_cell}",
		"node_type": node_type,
		"name": name,
		"title": _workbook_snapshot_title(lines, node_type, manager_names, proxy_names),
		"source_cell": source_cell,
		"lines": lines,
		"manager_names": manager_names,
		"proxy_names": proxy_names,
		"employee_names": employee_names,
		"planned_headcount": planned,
		"current_headcount": current,
		"vacancy_count": vacancy,
	}


def _workbook_snapshot_staffing(value):
	compact = re.sub(r"\s+", "", value)
	staffing = re.search(r"编制(?:/实际)?[:：]?(\d+)/(\d+)", compact)
	vacancy = re.search(r"空缺[:：]?(\d+)", compact)
	planned = cint(staffing.group(1)) if staffing else 0
	current = cint(staffing.group(2)) if staffing else 0
	return planned, current, cint(vacancy.group(1)) if vacancy else max(planned - current, 0)


def _workbook_snapshot_manager_names(lines):
	managers, proxies = [], []
	for index, line in enumerate(lines):
		compact = re.sub(r"\s+", "", line)
		role_match = re.match(r"(?:总经理|副总经理|技术总监|管理总监|总监|副总|分管|课长|组长|班长|副组长)[:：](.*)", compact)
		if role_match:
			managers.extend(_workbook_snapshot_names(role_match.group(1)))
			continue
		if compact.startswith("代理人[:：]"):
			proxies.extend(_workbook_snapshot_names(compact.split("：", 1)[-1].split(":", 1)[-1]))
			continue
		if compact in {"组长", "班长"} and index + 1 < len(lines):
			managers.extend(_workbook_snapshot_names(lines[index + 1]))
		# A few source cards contain only a leader name, without a role label.
		if index == 1 and not managers and (node_like_name := _workbook_snapshot_names(line)):
			managers.extend(node_like_name)
	return _deduplicate_names(managers), _deduplicate_names(proxies)


def _workbook_snapshot_employee_names(lines, node_type):
	if node_type not in {"work_level", "position_group", "employee_group"}:
		return []
	source_lines = lines if node_type == "employee_group" else lines[1:]
	return _deduplicate_names(name for line in source_lines for name in _workbook_snapshot_names(line))


def _workbook_snapshot_names(value):
	value = re.sub(r"TBA\s*[*×xX]?\s*\d*", "", cstr(value), flags=re.IGNORECASE)
	return [
		name
		for name in re.findall(r"[\u4e00-\u9fff]{2,4}", value)
		if name
		not in {"总经理", "副总经理", "技术总监", "管理总监", "代理人", "直线级", "间师级", "文组级", "文师级", "副组长", "组长", "班长", "员工", "模具员", "备品员", "助理", "物料员", "资讯助理", "膜厚室", "清洁", "厂区"}
	]


def _deduplicate_names(names):
	result = []
	for name in names:
		name = cstr(name).strip()
		if name and name not in result:
			result.append(name)
	return result


def _workbook_snapshot_title(lines, node_type, manager_names, proxy_names):
	parts = []
	if manager_names:
		parts.append(_("负责人：{0}").format("、".join(manager_names)))
	if proxy_names:
		parts.append(_("代理人：{0}").format("、".join(proxy_names)))
	if node_type in {"work_level", "position_group", "employee_group"} and len(lines) > 1:
		parts.append(_("原表人员：{0}人").format(len(_workbook_snapshot_employee_names(lines, node_type))))
	return " · ".join(parts)


def _build_workbook_snapshot_people(node, employee_lookup):
	people = []
	for role, names in [
		(node.get("role") or _("负责人"), node.get("manager_names") or []),
		(_("代理人"), node.get("proxy_names") or []),
		(_("员工"), node.get("employee_names") or []),
	]:
		for name in names:
			person = _build_person_token(name, role, employee_lookup)
			if person:
				people.append(person)
	return people


def _build_workbook_snapshot_relationships(nodes, root):
	by_type = defaultdict(list)
	for node in nodes:
		by_type[node["node_type"]].append(node)

	for node in nodes:
		node["parent_node_id"] = None
		node["children"] = []
		if node is root:
			continue
		parent = _find_workbook_snapshot_parent(node, by_type, root)
		node["parent_node_id"] = parent.get("node_id")
		parent["children"].append(node)

	for node in nodes:
		node["connections"] = len(node["children"])
		node["expandable"] = bool(node["children"])


def _find_workbook_snapshot_parent(node, by_type, root):
	node_type = node["node_type"]
	if node_type == "director":
		return root
	if node_type == "division":
		manager_names = set(node.get("manager_names") or [])
		directors = [
			candidate
			for candidate in by_type["director"]
			if manager_names.intersection(candidate.get("manager_names") or [])
		]
		return _closest_workbook_snapshot_node(node, directors) if directors else root
	parent_types = {
		"department": ("division",),
		"team": ("department",),
		"work_level": ("team",),
		"position_group": ("work_level", "team"),
		"employee_group": ("position_group", "work_level", "team"),
	}.get(node_type, ())
	candidates = [
		candidate
		for parent_type in parent_types
		for candidate in by_type[parent_type]
		if candidate["row"] < node["row"]
	]
	return _closest_workbook_snapshot_node(node, candidates) if candidates else root


def _closest_workbook_snapshot_node(node, candidates):
	return min(
		candidates,
		key=lambda candidate: (
			abs(candidate["column_center"] - node["column_center"]),
			-node["row"],
			candidate["column_start"],
		),
	)


def _workbook_snapshot_summary(root, nodes):
	people = {
		person.get("lookup_name") or person.get("employee_name")
		for node in nodes
		for person in node.get("people") or []
		if person.get("role") == _("员工") and (person.get("lookup_name") or person.get("employee_name"))
	}
	matched_people = {
		person.get("lookup_name") or person.get("employee_name")
		for node in nodes
		for person in node.get("people") or []
		if person.get("role") == _("员工") and person.get("matched_employee")
	}
	return {
		"planned_headcount": root.get("planned_headcount", 0),
		"current_headcount": root.get("current_headcount", 0),
		"vacancy_count": root.get("vacancy_count", 0),
		"department_count": len([node for node in nodes if node["node_type"] == "department"]),
		"source_employee_count": len(people),
		"matched_employee_count": len(matched_people),
		"missing_department_count": 0,
		"missing_manager_count": 0,
	}


def _get_workbook_snapshot_node_detail(node_id, node_type, company=None, search=None):
	payload = _get_yongxin_workbook_snapshot_tree(company)
	root = payload.get("root") or {}
	nodes = _flatten_organization_export_nodes(root)
	node = next((item for item in nodes if item.get("node_id") == node_id), None)
	if not node:
		return {"node_type": node_type, "node_id": node_id, "title": _("未找到原表节点"), "employees": [], "actions": {}}
	search = cstr(search).strip().lower()
	people = []
	seen_people = set()
	for descendant in _flatten_organization_export_nodes(node):
		for person in descendant.get("people") or []:
			key = (person.get("role"), person.get("lookup_name") or person.get("employee_name") or person.get("name"))
			if key in seen_people:
				continue
			seen_people.add(key)
			people.append(person)
	if search:
		people = [
			person
			for person in people
			if search in " ".join(cstr(person.get(key)) for key in ("name", "employee_name", "designation", "role")).lower()
		]
	parent = next((item for item in nodes if item.get("node_id") == node.get("parent_node_id")), None)
	return {
		"node_type": node.get("node_type"),
		"node_id": node_id,
		"title": node.get("name"),
		"subtitle": " · ".join(
			part
			for part in [
				_("原表单元格：{0}").format(node.get("source_cell")),
				node.get("title"),
			]
			if part
		),
		"card_content": node.get("card_content"),
		"metrics": {
			"planned_headcount": node.get("planned_headcount", 0),
			"current_headcount": node.get("current_headcount", 0),
			"vacancy_count": node.get("vacancy_count", 0),
			"employee_count": len([person for person in people if person.get("role") == _("员工")]),
		},
		"employees": people,
		"people": people,
		"relationships": {
			"parent": _workbook_snapshot_relation(parent),
			"children": [_workbook_snapshot_relation(child) for child in node.get("children") or []],
		},
		"actions": {
			"can_edit_department": bool(node.get("department"))
			and frappe.has_permission("Department", "write", node.get("department")),
		},
	}


def _workbook_snapshot_relation(node):
	if not node:
		return None
	return {"name": node.get("node_id"), "label": node.get("name"), "node_type": node.get("node_type")}


def _collect_template_department_units(chart_tree):
	"""Flatten department/team nodes while retaining the nearest actual department parent."""
	units = []

	def walk(node, parent_department="", management_path=None):
		node = frappe._dict(node or {})
		management_path = list(management_path or [])
		node_type = node.get("node_type")
		label = _organization_business_name(node.get("name"))
		if node_type in {"division", "director"} and label:
			management_path.append(label)
		if node_type in {"department", "team"} and label:
			units.append(
				{
					"node_type": node_type,
					"department_name": label,
					"parent_department": parent_department,
					"management_path": " / ".join(management_path),
					"source_cell": node.get("source_cell"),
					"manager_names": node.get("manager_names") or [],
				}
			)
			next_parent = label
		else:
			next_parent = parent_department
		for child in node.get("children") or []:
			walk(child, next_parent, management_path)

	walk(chart_tree)
	return units


def _suggest_template_team_name(unit, sibling_index=0):
	if unit.get("node_type") != "team":
		return unit.get("department_name") or ""
	managers = [
		_normalize_person_lookup(manager)
		for manager in unit.get("manager_names") or []
		if _normalize_person_lookup(manager) and not _normalize_person_lookup(manager).startswith("TBA")
	]
	suffix = "、".join(managers) or (_("第{0}组").format(sibling_index) if sibling_index else "") or _("待确认")
	return _("{0}（{1}）").format(unit.get("department_name") or _("小组"), suffix)


def _load_xlsx_workbook(source: Path):
	try:
		from openpyxl import load_workbook
	except ImportError:
		frappe.throw(_("缺少 openpyxl，无法读取组织基线 Excel。"))
	return load_workbook(source, read_only=True, data_only=True)


def _preview_q3_org_cells(sheet, limit=80):
	if not sheet:
		return []

	cells = []
	for row in sheet.iter_rows():
		for cell in row:
			value = _organization_source_name(cell.value)
			if not value:
				continue
			first_line = value.splitlines()[0].strip()
			cells.append(
				{
					"source_cell": cell.coordinate,
					"raw_department_name": first_line,
					"department_name": _organization_business_name(first_line),
					"raw_value": value,
				}
			)
			if len(cells) >= limit:
				return cells
	return cells


def _preview_organization_sheet(workbook, sheet_name, field_aliases, limit=25):
	if sheet_name not in workbook.sheetnames:
		return {"sheet_name": sheet_name, "found": False, "rows": [], "errors": [_("工作表不存在。")]}

	sheet = workbook[sheet_name]
	header_row, header_indexes = _find_organization_header_indexes(sheet, field_aliases)
	if not header_indexes:
		return {
			"sheet_name": sheet_name,
			"found": True,
			"header_row": None,
			"rows": [],
			"errors": [_("未找到姓名、部门等组织字段表头。")],
		}

	rows = []
	for row_index, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
		raw = {
			fieldname: values[index] if index < len(values) else None
			for fieldname, index in header_indexes.items()
		}
		normalized = _normalize_organization_preview_row(row_index, raw)
		if not normalized:
			continue
		rows.append(normalized)
		if len(rows) >= limit:
			break

	return {
		"sheet_name": sheet_name,
		"found": True,
		"header_row": header_row,
		"matched_fields": sorted(header_indexes),
		"rows": rows,
		"errors": [],
	}


def _find_organization_header_indexes(sheet, field_aliases, max_scan_rows=12):
	for row_index, values in enumerate(sheet.iter_rows(max_row=max_scan_rows, values_only=True), start=1):
		headers = {
			_clean_organization_header(value): index
			for index, value in enumerate(values)
			if _clean_organization_header(value)
		}
		matches = {}
		for fieldname, aliases in field_aliases.items():
			for alias in aliases:
				index = headers.get(_clean_organization_header(alias))
				if index is not None:
					matches[fieldname] = index
					break
		if matches.get("employee_name") and (matches.get("raw_department_name") or matches.get("department_name")):
			return row_index, matches
	return None, {}


def _normalize_organization_preview_row(row_index, raw):
	raw_department_name = _organization_source_name(
		raw.get("raw_department_name") or raw.get("department_name")
	)
	department_name = _organization_business_name(raw.get("department_name")) or _organization_business_name(raw_department_name)
	raw_designation_name = _organization_source_name(raw.get("raw_designation_name"))
	designation_name = _organization_business_name(raw_designation_name)
	employee_name = _organization_source_name(raw.get("employee_name"))

	if not employee_name and not raw_department_name and not raw_designation_name:
		return None

	return {
		"row_number": row_index,
		"employee_code": _organization_source_name(raw.get("employee_code")),
		"employee_name": employee_name,
		"raw_department_name": raw_department_name,
		"department_name": department_name,
		"raw_designation_name": raw_designation_name,
		"designation_name": designation_name,
		"grade": _organization_source_name(raw.get("grade")),
		"reports_to": _organization_source_name(raw.get("reports_to")),
	}


def _clean_organization_header(value):
	return re.sub(r"\s+", "", _organization_source_name(value).replace("/", ""))


def _organization_source_name(value):
	return cstr(value).strip()


def _organization_business_name(value):
	text = _organization_source_name(value)
	if not text:
		return ""
	text = ORGANIZATION_TECHNICAL_SUFFIX_RE.sub("", text).strip()
	for separator in ("／", "\\", ">", "｜", "|"):
		text = text.replace(separator, "/")
	parts = [part.strip() for part in text.split("/") if part.strip()]
	if parts:
		text = parts[-1]
	dash_parts = [part.strip() for part in re.split(r"\s*[-－]\s*", text) if part.strip()]
	if len(dash_parts) > 1:
		text = dash_parts[-1]
	return ORGANIZATION_TECHNICAL_SUFFIX_RE.sub("", text).strip()


def _count_seed_nodes(nodes):
	return sum(1 + _count_seed_nodes(node.get("children") or []) for node in nodes if node)


def _company_has_imported_org_template(company=None):
	if not company or not frappe.get_meta("Department").has_field("hrms_org_source_cell"):
		return False
	return bool(
		frappe.db.exists(
			"Department",
			{
				"company": company,
				"hrms_org_source_cell": ["in", _template_department_source_cells()],
			},
		)
	)


def _is_yongxin_company(company=None):
	if not company:
		return False
	if cstr(company).strip() == YONGXIN_COMPANY_NAME or cstr(company).strip() in YONGXIN_LEGACY_COMPANY_KEYS:
		return True
	return cstr(frappe.db.get_value("Company", company, "company_name")).strip() == YONGXIN_COMPANY_NAME


def _template_department_source_cells():
	seed = _load_yongxin_q2_org_template()
	cells = []

	def collect(node):
		if not node:
			return
		if node.get("node_type") in {"division", "department", "team"} and node.get("source_cell"):
			cells.append(node.get("source_cell"))
		for child in node.get("children") or []:
			collect(child)

	collect(seed.get("chart_tree"))
	try:
		cells.extend(node["source_cell"] for node in _load_yongxin_q3_department_hierarchy().get("nodes") or [])
	except Exception:
		# A missing optional Q3 deployment file must not hide ordinary departments.
		pass
	return list(dict.fromkeys(cells))


def _get_yongxin_template_tree(company=None):
	seed = _load_yongxin_q2_org_template()
	departments = _get_departments(company)
	departments_by_source = {
		department.get("hrms_org_source_cell"): department
		for department in departments
		if department.get("hrms_org_source_cell")
	}
	employees = _get_active_employees(company)
	employees_by_department = defaultdict(list)
	for employee in employees:
		if employee.get("department"):
			employees_by_department[employee.get("department")].append(employee)
	employee_lookup = _get_employee_lookup(employees)

	root = _build_template_tree_node(
		seed.get("chart_tree"),
		departments_by_source,
		employees_by_department,
		employee_lookup,
	)
	summary = {
		"planned_headcount": root.get("planned_headcount", 0),
		"current_headcount": root.get("current_headcount", 0),
		"vacancy_count": root.get("vacancy_count", 0),
		"department_count": _count_template_nodes_by_type(root, {"division", "department", "team"}),
		"missing_department_count": 0,
		"missing_manager_count": _count_template_missing_managers(root),
	}
	return {
		"root": root,
		"summary": summary,
		"field_map": HYBRID_ROSTER_FIELD_MAP,
		"source_document": seed.get("source_document"),
		"source_sheet": seed.get("source_sheet"),
	}


def _build_template_tree_node(
	node,
	departments_by_source,
	employees_by_department,
	employee_lookup,
	parent_department=None,
):
	node = frappe._dict(node or {})
	source_cell = node.get("source_cell")
	department = departments_by_source.get(source_cell)
	effective_department = department or parent_department
	children = [
		_build_template_tree_node(
			child,
			departments_by_source,
			employees_by_department,
			employee_lookup,
			effective_department,
		)
		for child in node.get("children") or []
	]
	result = {
		"node_id": node.get("node_id") or f"{node.get('node_type')}:{source_cell}",
		"id": node.get("id") or f"{node.get('node_type')}:{source_cell}",
		"node_type": node.get("node_type") or "template",
		"template_node_type": node.get("node_type"),
		"name": node.get("name"),
		"title": node.get("title") or _template_manager_label(node),
		"role": node.get("role"),
		"manager_names": "、".join(node.get("manager_names") or []),
		"proxy_names": "、".join(node.get("proxy_names") or []),
		"card_content": "",
		"lines": node.get("lines") or [],
		"employee_names": node.get("employee_names") or [],
		"people": _build_template_people(node, employee_lookup, effective_department),
		"source_cell": source_cell,
		"planned_headcount": node.get("planned_headcount") or 0,
		"current_headcount": node.get("current_headcount") or 0,
		"vacancy_count": node.get("vacancy_count") or 0,
		"children": children,
	}

	if department:
		result.update(
			{
				"node_id": f"department:{department.name}",
				"id": f"department:{department.name}",
				"node_type": "department",
				"template_node_type": node.get("node_type"),
				"name": department.get("department_name") or node.get("name"),
				"title": _department_manager_label(department) or node.get("title") or _template_manager_label(node),
				"department": department.name,
				"role": department.get("hrms_org_role") or node.get("role"),
				"manager_names": department.get("hrms_org_manager") or "、".join(node.get("manager_names") or []),
				"proxy_names": department.get("hrms_org_proxy") or "、".join(node.get("proxy_names") or []),
				"card_content": cstr(department.get("hrms_org_card_content")).strip(),
				"planned_headcount": cint(department.get("hrms_planned_headcount")) or node.get("planned_headcount") or 0,
				"current_headcount": cint(department.get("hrms_actual_headcount")) or node.get("current_headcount") or 0,
				"vacancy_count": cint(department.get("hrms_vacancy_count")) or node.get("vacancy_count") or 0,
				"recruitment_plan": department.get("hrms_recruitment_plan"),
			}
		)

	result["connections"] = len(children)
	result["expandable"] = bool(children)
	return result


def _get_employee_lookup(employees):
	lookup = {}
	for employee in employees or []:
		row = _employee_row(employee)
		for key in {row.get("employee_name"), _normalize_person_lookup(row.get("employee_name"))}:
			if key and key not in lookup:
				lookup[key] = row
	return lookup


def _build_template_people(node, employee_lookup, department=None):
	people = []
	seen = set()
	for role, names in [
		(node.get("role") or _("负责人"), node.get("manager_names") or []),
		(_("代理人"), node.get("proxy_names") or []),
		(_("员工"), node.get("employee_names") or []),
	]:
		for name in names:
			person = _build_person_token(
				name=name,
				role=role,
				employee_lookup=employee_lookup,
				department=department,
				expected_designation=role,
			)
			if not person:
				continue
			key = (person.get("lookup_name"), person.get("role"))
			if key in seen:
				continue
			seen.add(key)
			people.append(person)
	return people


def _build_person_token(name, role, employee_lookup, department=None, expected_designation=None):
	display_name = cstr(name).strip()
	lookup_name = _normalize_person_lookup(display_name)
	if not lookup_name or lookup_name in {"(代)", "(兼)", "（代）", "（兼）"}:
		return None
	matched = employee_lookup.get(display_name) or employee_lookup.get(lookup_name)
	department_name = department.get("name") if department else ""
	department_label = department.get("department_name") if department else ""
	if matched:
		return {
			"name": display_name,
			"lookup_name": lookup_name,
			"role": role or "",
			"employee": matched.get("name"),
			"employee_name": matched.get("employee_name") or display_name,
			"employee_code": matched.get("employee_code"),
			"department": matched.get("department") or department_name,
			"department_label": department_label,
			"designation": matched.get("designation") or expected_designation or role or "",
			"grade": matched.get("grade"),
			"reports_to": matched.get("reports_to"),
			"branch": matched.get("branch"),
			"cell_number": matched.get("cell_number"),
			"image": matched.get("image"),
			"matched_employee": True,
			"match_status": _("已匹配员工档案"),
		}
	return {
		"name": display_name,
		"lookup_name": lookup_name,
		"role": role or "",
		"employee": "",
		"employee_name": display_name,
		"employee_code": "",
		"department": department_name,
		"department_label": department_label,
		"designation": expected_designation or role or "",
		"grade": "",
		"reports_to": "",
		"branch": "",
		"cell_number": "",
		"image": "",
		"matched_employee": False,
		"match_status": _("待匹配员工档案"),
	}


def _normalize_person_lookup(value):
	value = cstr(value).strip()
	if not value:
		return ""
	value = value.replace("（", "(").replace("）", ")")
	value = re.sub(r"\s*\([^)]*\)\s*$", "", value).strip()
	return value


def _template_manager_label(node):
	role = node.get("role") or ""
	manager = "、".join(node.get("manager_names") or [])
	return "：".join(part for part in [role, manager] if part)


def _count_template_nodes_by_type(node, node_types):
	return (1 if node.get("template_node_type") in node_types else 0) + sum(
		_count_template_nodes_by_type(child, node_types) for child in node.get("children") or []
	)


def _count_template_missing_managers(node):
	missing = 0
	if node.get("template_node_type") in {"division", "department", "team"} and not node.get("manager_names"):
		missing = 1
	return missing + sum(_count_template_missing_managers(child) for child in node.get("children") or [])


def _get_template_node_detail(node_id, node_type, company=None, search=None, allow_unimported=False):
	if not allow_unimported and not _company_has_imported_org_template(company):
		return None
	seed = _load_yongxin_q2_org_template()
	node = _find_template_node(seed.get("chart_tree"), node_id, node_type)
	if not node:
		return None
	departments_by_source = {
		department.get("hrms_org_source_cell"): department
		for department in _get_departments(company)
		if department.get("hrms_org_source_cell")
	}
	department = departments_by_source.get(node.get("source_cell"))
	employee_lookup = _get_employee_lookup(_get_active_employees(company))
	employee_names = _template_employee_names(node)
	employees = _get_employees_by_names(employee_names, company=company, search=search, employee_lookup=employee_lookup)
	people = _build_template_people(node, employee_lookup, department)
	return {
		"node_type": node.get("node_type"),
		"node_id": node_id,
		"title": node.get("name"),
		"subtitle": " · ".join(
			part
			for part in [
				node.get("source_cell") and _("来源单元格：{0}").format(node.get("source_cell")),
				_template_manager_label(node),
				node.get("proxy_names") and _("代理人：{0}").format("、".join(node.get("proxy_names") or [])),
			]
			if part
		),
		"metrics": {
			"planned_headcount": node.get("planned_headcount") or 0,
			"current_headcount": node.get("current_headcount") or len(employee_names),
			"vacancy_count": node.get("vacancy_count") or 0,
			"employee_count": len(employee_names),
		},
		"employees": employees,
		"people": people,
		"template_lines": node.get("lines") or node.get("employee_names") or [],
		"actions": {"can_add_department": frappe.has_permission("Department", "create")},
	}


def _find_template_node(node, node_id, node_type=None):
	if not node:
		return None
	if node.get("node_id") == node_id or node.get("id") == node_id:
		return node
	if ":" in node_id and node.get("source_cell") == node_id.split(":", 1)[1]:
		return node
	for child in node.get("children") or []:
		found = _find_template_node(child, node_id, node_type)
		if found:
			return found
	return None


def _template_employee_names(node):
	names = []
	for value in (node.get("employee_names") or []):
		value = cstr(value).strip()
		if value:
			names.append(value)
	for child in node.get("children") or []:
		names.extend(_template_employee_names(child))
	return names


def _get_employees_by_names(names, company=None, search=None, employee_lookup=None):
	clean_names = [name for name in names if name and not name.startswith("TBA")]
	if search:
		clean_names = [name for name in clean_names if search in name]
	if not clean_names:
		return []
	filters = {"employee_name": ["in", clean_names]}
	if company:
		filters["company"] = company
	fields = _get_meta_fields(
		"Employee",
		[
			"name",
			"employee_name",
			"custom_employee_code",
			"department",
			"designation",
			"grade",
			"reports_to",
			"branch",
			"cell_number",
			"image",
		],
	)
	rows = frappe.get_all("Employee", fields=fields, filters=filters, limit_page_length=500)
	by_name = employee_lookup or _get_employee_lookup(rows)
	return [
		dict(by_name[name] if name in by_name else by_name.get(_normalize_person_lookup(name)), matched_employee=True, match_status=_("已匹配员工档案"))
		if name in by_name or _normalize_person_lookup(name) in by_name
		else {
			"name": "",
			"employee_name": name,
			"employee_code": "",
			"department": "",
			"designation": "组织图人员",
			"grade": "",
			"reports_to": "",
			"branch": "",
			"cell_number": "",
			"image": "",
			"matched_employee": False,
			"match_status": _("待匹配员工档案"),
		}
		for name in clean_names
	]


def _import_department_node(
	node,
	company,
	parent_department,
	result,
	source_to_department,
	seen_department_labels,
):
	node_type = cstr(node.get("node_type")).strip()
	if node_type not in TEMPLATE_DEPARTMENT_NODE_TYPES:
		for child in node.get("children") or []:
			_import_department_node(
				node=child,
				company=company,
				parent_department=parent_department,
				result=result,
				source_to_department=source_to_department,
				seen_department_labels=seen_department_labels,
			)
		return

	department = _upsert_department_from_node(node, company, parent_department, result, seen_department_labels)
	source_to_department[node.get("source_cell")] = department
	for child in node.get("children") or []:
		_import_department_node(
			node=child,
			company=company,
			parent_department=department,
			result=result,
			source_to_department=source_to_department,
			seen_department_labels=seen_department_labels,
		)


def _upsert_department_from_node(node, company, parent_department, result, seen_department_labels):
	source_cell = node.get("source_cell")
	department = _find_department_by_source(company, source_cell)
	department_name = _department_display_name(node, parent_department, seen_department_labels)
	created = False

	if department:
		doc = frappe.get_doc("Department", department)
	else:
		department = frappe.db.get_value("Department", {"department_name": department_name, "company": company}, "name")
		if department and _would_create_department_loop(department, parent_department):
			department_name = _source_scoped_department_name(department_name, node)
			department = frappe.db.get_value("Department", {"department_name": department_name, "company": company}, "name")
		if department:
			doc = frappe.get_doc("Department", department)
		else:
			doc = frappe.new_doc("Department")
			created = True

	doc.department_name = department_name
	doc.company = company
	if frappe.get_meta("Department").has_field("parent_department"):
		doc.parent_department = parent_department
	if frappe.get_meta("Department").has_field("is_group"):
		doc.is_group = 1 if node.get("children") else 0
	_set_if_exists(doc, "hrms_org_level", node.get("level"))
	_set_if_exists(doc, "hrms_org_role", node.get("role"))
	_set_if_exists(doc, "hrms_org_manager", "、".join(node.get("manager_names") or []))
	_set_if_exists(doc, "hrms_org_proxy", "、".join(node.get("proxy_names") or []))
	_set_if_exists(doc, "hrms_planned_headcount", cint(node.get("planned_headcount")))
	_set_if_exists(doc, "hrms_actual_headcount", cint(node.get("actual_headcount")))
	_set_if_exists(doc, "hrms_vacancy_count", cint(node.get("vacancy_count")))
	_set_if_exists(doc, "hrms_org_source_cell", source_cell)
	doc.save(ignore_permissions=False)

	result["created_departments" if created else "updated_departments"].append(doc.name)
	return doc.name


def _source_scoped_department_name(department_name, node):
	source_cell = cstr(node.get("source_cell")).strip()
	manager = "、".join(name for name in (node.get("manager_names") or []) if name and not name.startswith("TBA"))
	suffix = manager or source_cell
	return f"{department_name}（{suffix}）" if suffix else department_name


def _would_create_department_loop(department, parent_department):
	if not department or not parent_department or department == parent_department:
		return bool(department and parent_department and department == parent_department)
	bounds = frappe.db.get_value("Department", department, ["lft", "rgt"], as_dict=True)
	parent_bounds = frappe.db.get_value("Department", parent_department, ["lft", "rgt"], as_dict=True)
	if not bounds or not parent_bounds:
		return False
	return parent_bounds.lft > bounds.lft and parent_bounds.rgt < bounds.rgt


def _department_display_name(node, parent_department, seen_department_labels):
	label = cstr(node.get("label")).strip()
	source_cell = cstr(node.get("source_cell")).strip()
	key = (parent_department or "", label)
	seen_department_labels[key] += 1
	if seen_department_labels[key] <= 1:
		return label

	manager = "、".join(name for name in (node.get("manager_names") or []) if name and not name.startswith("TBA"))
	if manager:
		return f"{label}（{manager}）"
	return f"{label}（{source_cell}）" if source_cell else f"{label}-{seen_department_labels[key]}"


def _find_department_by_source(company, source_cell):
	if not source_cell or not frappe.get_meta("Department").has_field("hrms_org_source_cell"):
		return None
	return frappe.db.get_value(
		"Department",
		{"company": company, "hrms_org_source_cell": source_cell},
		"name",
	)


def _apply_staffing_summary(rows, company, result):
	meta = frappe.get_meta("Department")
	if not meta.has_field("hrms_recruitment_plan"):
		return
	for row in rows:
		department_name = cstr(row.get("department")).strip()
		if not department_name:
			continue
		departments = frappe.get_all(
			"Department",
			filters={"company": company, "department_name": ["like", f"{department_name}%"]},
			fields=["name"],
			limit_page_length=20,
		)
		for department in departments:
			doc = frappe.get_doc("Department", department.name)
			_set_if_exists(doc, "hrms_planned_headcount", cint(row.get("planned_headcount")))
			_set_if_exists(doc, "hrms_actual_headcount", cint(row.get("actual_headcount")))
			_set_if_exists(doc, "hrms_vacancy_count", cint(row.get("vacancy_count")))
			_set_if_exists(doc, "hrms_recruitment_plan", row.get("recruitment_plan"))
			doc.save(ignore_permissions=False)
			if department.name not in result["updated_departments"]:
				result["updated_departments"].append(department.name)


def _import_position_templates(rows, company, source_to_department, result):
	for row in rows:
		department = _find_department_by_label(company, row.get("department"))
		parent_designation = _ensure_designation(row.get("parent_designation"), company, department, None, result)
		_ensure_designation(
			row.get("designation"),
			company,
			department,
			parent_designation,
			result,
			source_cell=row.get("source_cell"),
		)


def _find_department_by_label(company, label):
	label = cstr(label).strip()
	if not label:
		return None
	return frappe.db.get_value("Department", {"company": company, "department_name": label}, "name") or frappe.db.get_value(
		"Department",
		{"company": company, "department_name": ["like", f"{label}%"]},
		"name",
	)


def _ensure_designation(designation, company, department, parent_designation, result, source_cell=None):
	designation = cstr(designation).strip()
	if not designation:
		return None
	existing = frappe.db.exists("Designation", designation) or frappe.db.get_value(
		"Designation", {"designation_name": designation}, "name"
	)
	created = False
	if existing:
		doc = frappe.get_doc("Designation", existing)
	else:
		doc = frappe.new_doc("Designation")
		doc.designation_name = designation
		created = True
	_set_if_exists(doc, "hrms_parent_designation", parent_designation)
	_set_if_exists(doc, "hrms_source_department", department)
	_set_if_exists(doc, "hrms_org_source_cell", source_cell)
	doc.save(ignore_permissions=False)
	result["created_designations" if created else "updated_designations"].append(doc.name)
	return doc.name


def _set_if_exists(doc, fieldname, value):
	if frappe.get_meta(doc.doctype).has_field(fieldname):
		doc.set(fieldname, value)


def _get_meta_fields(doctype, requested):
	meta = frappe.get_meta(doctype)
	return [field for field in requested if field == "name" or meta.has_field(field)]


def _validate_department_delete(department):
	if not frappe.db.exists("Department", department):
		frappe.throw(_("部门 {0} 不存在。").format(department))
	if not frappe.has_permission("Department", "delete", department):
		frappe.throw(_("没有权限删除部门 {0}。").format(department))
	child = frappe.db.get_value("Department", {"parent_department": department}, "name")
	if child:
		frappe.throw(_("部门 {0} 下仍有子部门 {1}，请先调整层级。").format(department, child))
	employee = frappe.db.get_value("Employee", {"department": department, "status": "Active"}, "name")
	if employee:
		frappe.throw(_("部门 {0} 下仍有关联在职员工 {1}，请先转移员工。").format(department, employee))


def _get_departments(company=None):
	filters = {}
	if company:
		filters["company"] = company
	fields = _get_meta_fields(
		"Department",
		[
			"name",
			"department_name",
			"parent_department",
			"company",
			"is_group",
			"lft",
			"rgt",
			"hrms_org_level",
			"hrms_org_role",
			"hrms_org_manager",
			"hrms_org_proxy",
			"hrms_org_card_content",
			"hrms_planned_headcount",
			"hrms_actual_headcount",
			"hrms_vacancy_count",
			"hrms_recruitment_plan",
			"hrms_org_source_cell",
			"hrms_roster_assignable",
		],
	)
	departments = frappe.get_all("Department", fields=fields, filters=filters, order_by="lft asc, department_name asc")
	if "hrms_org_source_cell" not in fields:
		return departments
	allowed_source_cells = set(_template_department_source_cells())
	return [
		department
		for department in departments
		if not department.get("hrms_org_source_cell")
		or department.get("hrms_org_source_cell") in allowed_source_cells
	]


def _get_active_employees(company=None):
	filters = {"status": "Active"}
	if company:
		filters["company"] = company
	fields = _get_meta_fields(
		"Employee",
		[
			"name",
			"employee_name",
			"custom_employee_code",
			"company",
			"branch",
			"department",
			"designation",
			"grade",
			"reports_to",
			"image",
			"cell_number",
			"status",
		],
	)
	return frappe.get_all("Employee", fields=fields, filters=filters, order_by="department asc, designation asc, employee_name asc")


def _get_company_label(company=None):
	if not company:
		return "全部公司"
	return frappe.db.get_value("Company", company, "company_name") or company


def _company_metrics(company=None):
	departments = _get_departments(company)
	employees = _get_active_employees(company)
	staffing = _get_department_staffing(company)
	department_by_name = {department.name: department for department in departments}
	root_departments = [
		department
		for department in departments
		if not department.get("parent_department") or department.get("parent_department") not in department_by_name
	]
	staffing_summary = _summarize_staffing(root_departments, staffing)
	return {
		"planned_headcount": staffing_summary["planned_headcount"],
		"current_headcount": len(employees) or staffing_summary["current_headcount"],
		"vacancy_count": staffing_summary["vacancy_count"],
		"department_count": len(departments),
	}


def _summarize_staffing(root_departments, staffing):
	roots_with_counts = [department for department in root_departments if staffing.get(department.name, {}).get("planned_headcount")]
	source_departments = roots_with_counts or [frappe._dict({"name": name}) for name in staffing]
	return {
		"planned_headcount": sum(staffing.get(department.name, {}).get("planned_headcount", 0) for department in source_departments),
		"current_headcount": sum(staffing.get(department.name, {}).get("current_headcount", 0) for department in source_departments),
		"vacancy_count": sum(staffing.get(department.name, {}).get("vacancy_count", 0) for department in source_departments),
	}


def _is_management_designation(designation):
	return bool(_classify_management_role(designation)) or any(
		keyword in cstr(designation) for keyword in HYBRID_MANAGER_KEYWORDS
	)


def _get_department_managers(employees):
	direct_manager_names = {cstr(employee.get("reports_to")).strip() for employee in employees if employee.get("reports_to")}
	return [
		employee
		for employee in employees
		if _is_management_designation(employee.get("designation")) or employee.name in direct_manager_names
	]


def _get_department_staffing(company=None):
	staffing = defaultdict(lambda: {"planned_headcount": 0, "current_headcount": 0, "vacancy_count": 0})
	for department in _get_departments(company):
		planned = cint(department.get("hrms_planned_headcount"))
		current = cint(department.get("hrms_actual_headcount"))
		vacancy = cint(department.get("hrms_vacancy_count"))
		if planned or current or vacancy:
			staffing[department.name] = {
				"planned_headcount": planned,
				"current_headcount": current,
				"vacancy_count": vacancy if vacancy else max(planned - current, 0),
				"recruitment_plan": department.get("hrms_recruitment_plan"),
			}
	if not frappe.db.exists("DocType", "Staffing Plan"):
		return staffing

	filters = {}
	if company:
		filters["company"] = company
	for plan in frappe.get_all("Staffing Plan", filters=filters, fields=["name", "department"]):
		try:
			doc = frappe.get_doc("Staffing Plan", plan.name)
		except Exception:
			continue
		department = doc.get("department")
		for row in doc.get("staffing_plan_details") or []:
			positions = frappe.utils.cint(row.get("number_of_positions"))
			current = frappe.utils.cint(row.get("current_count"))
			vacancies = frappe.utils.cint(row.get("vacancies"))
			staffing[department]["planned_headcount"] += positions
			staffing[department]["current_headcount"] += current
			staffing[department]["vacancy_count"] += vacancies if vacancies else max(positions - current, 0)
	return staffing


def _build_department_node(department, employees_by_department, department_children, staffing):
	employees = employees_by_department.get(department.name, [])
	managers = _get_department_managers(employees)
	staffing_row = staffing.get(department.name, {})
	current_headcount = len(employees) or staffing_row.get("current_headcount", 0)
	children = [
		_build_department_node(child, employees_by_department, department_children, staffing)
		for child in department_children.get(department.name, [])
	]
	children.extend(_build_work_level_nodes(department, employees, managers))
	node = {
		"node_id": f"department:{department.name}",
		"id": f"department:{department.name}",
		"node_type": "department",
		"name": department.get("department_name") or department.name,
		"title": _manager_label(managers) or _department_manager_label(department) or "负责人未设置",
		"department": department.name,
		"role": department.get("hrms_org_role"),
		"manager_names": department.get("hrms_org_manager"),
		"proxy_names": department.get("hrms_org_proxy"),
		"card_content": department.get("hrms_org_card_content"),
		# Employees are rendered once as leaf nodes under 职级 → 岗位.  Keeping
		# them as department chips as well was the source of the apparent
		# "unclassified people" in the old hybrid chart.
		"people": [],
		"recruitment_plan": department.get("hrms_recruitment_plan"),
		"planned_headcount": staffing_row.get("planned_headcount", 0),
		"current_headcount": current_headcount,
		"vacancy_count": staffing_row.get(
			"vacancy_count",
			max((staffing_row.get("planned_headcount", 0) or 0) - current_headcount, 0),
		),
		"children": children,
	}
	node["connections"] = len(children)
	node["expandable"] = bool(children)
	return node


def _build_company_root_node(
	company,
	department_nodes,
	staffing_summary,
	missing_department_count,
	missing_manager_count,
):
	"""Expose Company as the only visible root of the editable folder tree."""
	children = sorted(
		department_nodes,
		key=lambda node: (
			cint(node.get("org_level")) or 999,
			cstr(node.get("name")).strip(),
		),
	)
	return {
		"node_id": f"company:{company or 'all'}",
		"id": f"company:{company or 'all'}",
		"node_type": "company",
		"name": _("{0}（总公司）").format(_get_company_label(company)),
		"title": _("公司根节点"),
		"role": _("总公司"),
		"people": [],
		"lines": [],
		"planned_headcount": staffing_summary.get("planned_headcount", 0),
		"current_headcount": staffing_summary.get("current_headcount", 0),
		"vacancy_count": staffing_summary.get("vacancy_count", 0),
		"missing_department_count": missing_department_count,
		"missing_manager_count": missing_manager_count,
		"children": children,
		"connections": len(children),
		"expandable": bool(children),
	}


def _build_live_management_hierarchy(
	company,
	employees,
	department_nodes,
	staffing_summary,
	missing_department_count,
	missing_manager_count,
):
	"""Keep the Excel management line while Department remains the editable source of truth."""
	template_root = frappe._dict(_load_yongxin_q2_org_template().get("chart_tree") or {})
	department_nodes_by_label = {
		_organization_business_name(node.get("name")): node
		for node in department_nodes
		if _organization_business_name(node.get("name"))
	}
	used_departments = set()
	employees_by_name = {
		_normalize_person_lookup(employee.get("employee_name") or employee.name): employee
		for employee in employees
		if _normalize_person_lookup(employee.get("employee_name") or employee.name)
	}

	def build_template_node(template_node):
		template_node = frappe._dict(template_node or {})
		node_type = template_node.get("node_type")
		business_name = _organization_business_name(template_node.get("name"))
		if node_type in TEMPLATE_DEPARTMENT_NODE_TYPES and business_name in department_nodes_by_label:
			live_department = department_nodes_by_label[business_name]
			used_departments.add(live_department.get("department"))
			return live_department

		children = (
			_build_live_division_nodes(template_node, build_template_node)
			if node_type == "company_leadership"
			else [
				child
				for child in (build_template_node(child) for child in template_node.get("children") or [])
				if child
			]
		)
		if node_type in {"department", "team"}:
			return children[0] if len(children) == 1 else None
		if node_type not in {"company_leadership", "director", "division"}:
			return None
		if node_type != "company_leadership" and not children:
			return None

		role, fallback_names = _template_role_and_names(template_node)
		manager_names = _resolve_management_names(role, fallback_names, employees, employees_by_name)
		people = _resolve_management_people(manager_names, employees_by_name)
		child_summary = _summarize_live_nodes(children)
		if node_type == "company_leadership":
			deputy_role, deputy_fallback = _template_role_and_names(
				frappe._dict({"name": template_node.get("title")})
			)
			deputy_names = _resolve_management_names(deputy_role, deputy_fallback, employees, employees_by_name)
			children = _wrap_deputy_divisions(children, deputy_names, company)
			director_lines = []
			for director in template_node.get("children") or []:
				if director.get("node_type") != "director":
					continue
				director_role, director_fallback = _template_role_and_names(director)
				director_names = _resolve_management_names(director_role, director_fallback, employees, employees_by_name)
				director_lines.append(_("{0}：{1}").format(director_role, "、".join(director_names)))
			return {
				"node_id": template_node.get("node_id") or f"company_leadership:{company}",
				"id": template_node.get("id") or f"company_leadership:{company}",
				"node_type": "company_leadership",
				"name": _("总经理：{0}").format("、".join(manager_names)),
				"title": _("副总经理：{0}").format("、".join(deputy_names)),
				"role": "总经理",
				"people": people + _resolve_management_people(deputy_names, employees_by_name),
				"lines": director_lines
				+ [
					_("编制/实际：{0}/{1}").format(
						staffing_summary.get("planned_headcount", 0), staffing_summary.get("current_headcount", 0)
					),
					_("空缺：{0}").format(staffing_summary.get("vacancy_count", 0)),
				],
				"planned_headcount": staffing_summary.get("planned_headcount", 0),
				"current_headcount": staffing_summary.get("current_headcount", 0),
				"vacancy_count": staffing_summary.get("vacancy_count", 0),
				"missing_department_count": missing_department_count,
				"missing_manager_count": missing_manager_count,
				"children": children,
				"connections": len(children),
				"expandable": bool(children),
			}

		if node_type == "director":
			name = _("{0}：{1}").format(role, "、".join(manager_names))
			title = _("直属分管 {0} 个").format(len(children))
		else:
			name = template_node.get("name") or _("分管")
			title = _("分管：{0}").format("、".join(manager_names))
		return {
			"node_id": template_node.get("node_id"),
			"id": template_node.get("id"),
			"node_type": node_type,
			"name": name,
			"title": title,
			"role": role,
			"people": people,
			"source_cell": template_node.get("source_cell"),
			"planned_headcount": child_summary["planned_headcount"],
			"current_headcount": child_summary["current_headcount"],
			"vacancy_count": child_summary["vacancy_count"],
			"children": children,
			"connections": len(children),
			"expandable": bool(children),
		}

	root = build_template_node(template_root)
	if not root:
		root = {
			"node_id": f"company_leadership:{company or 'all'}",
			"id": f"company_leadership:{company or 'all'}",
			"node_type": "company_leadership",
			"name": _("管理层级"),
			"title": "",
			"people": [],
			"lines": [],
			"children": [],
		}

	remaining_departments = [node for node in department_nodes if node.get("department") not in used_departments]
	if remaining_departments:
		root["children"].append(
			{
				"node_id": "division:unassigned",
				"id": "division:unassigned",
				"node_type": "division",
				"name": _("未归属分管"),
				"title": _("请维护部门上级关系"),
				"people": [],
				"planned_headcount": sum(node.get("planned_headcount", 0) for node in remaining_departments),
				"current_headcount": sum(node.get("current_headcount", 0) for node in remaining_departments),
				"vacancy_count": sum(node.get("vacancy_count", 0) for node in remaining_departments),
				"children": remaining_departments,
				"connections": len(remaining_departments),
				"expandable": True,
			}
		)
	root["connections"] = len(root.get("children") or [])
	root["expandable"] = bool(root.get("children"))
	return root


def _wrap_deputy_divisions(children, deputy_names, company):
	"""Make 总经理 → 副总经理 → 分管 a real visible branch, not card text."""
	deputies = {name for name in deputy_names if name and name != _("未设置")}
	if not deputies:
		return children
	grouped = defaultdict(list)
	remaining = []
	for child in children:
		# Imported divisions are Department records (and therefore carry the
		# ``department`` node type); the source label is the stable indicator.
		if child.get("node_type") != "division" and not cstr(child.get("name")).endswith("分管"):
			remaining.append(child)
			continue
		manager_names = [person.get("employee_name") or person.get("name") for person in child.get("people") or []]
		manager_names.extend(re.findall(r"(.+?)分管$", cstr(child.get("name"))))
		deputy = next((name for name in manager_names if name in deputies), None)
		if deputy:
			grouped[deputy].append(child)
		else:
			remaining.append(child)
	wrapped = []
	for deputy in deputy_names:
		branches = grouped.get(deputy)
		if not branches:
			continue
		summary = _summarize_live_nodes(branches)
		node_id = f"director:deputy:{company}:{deputy}"
		wrapped.append(
			{
				"node_id": node_id,
				"id": node_id,
				"node_type": "director",
				"name": _("副总经理：{0}").format(deputy),
				"title": _("分管 {0} 个组织单元").format(len(branches)),
				"role": "副总经理",
				"people": [],
				"planned_headcount": summary["planned_headcount"],
				"current_headcount": summary["current_headcount"],
				"vacancy_count": summary["vacancy_count"],
				"children": branches,
				"connections": len(branches),
				"expandable": True,
			}
		)
	return wrapped + remaining


def _build_live_division_nodes(template_root, build_node):
	"""Build only the management/division branch nodes defined by the quarterly chart."""
	return [child for child in (build_node(child) for child in template_root.get("children") or []) if child]


def _template_role_and_names(node):
	name = cstr(node.get("name")).strip()
	if "：" in name:
		role, names = name.split("：", 1)
		return role.strip(), [part.strip() for part in re.split(r"[、,，]", names) if part.strip()]
	manager_names = [cstr(value).strip() for value in node.get("manager_names") or [] if cstr(value).strip()]
	return cstr(node.get("role")).strip() or _("负责人"), manager_names


def _classify_management_role(designation):
	designation = cstr(designation).strip()
	if "副总经理" in designation:
		return "副总经理"
	if "总经理" in designation:
		return "总经理"
	if "技术总监" in designation:
		return "技术总监"
	if "管理总监" in designation:
		return "管理总监"
	if "总监" in designation:
		return "总监"
	if "课长" in designation:
		return "课长"
	if "主管" in designation:
		return "主管"
	if "组长" in designation or "班长" in designation:
		return "组长"
	if "经理" in designation:
		return "经理"
	return ""


def _resolve_management_names(role, fallback_names, employees, employees_by_name):
	fallback_names = list(fallback_names or [])
	role_matches = [
		employee.get("employee_name") or employee.name
		for employee in employees
		if _classify_management_role(employee.get("designation")) == role
	]
	if role_matches:
		return role_matches
	matched_fallbacks = [
		name for name in fallback_names if _normalize_person_lookup(name) in employees_by_name
	]
	return matched_fallbacks or fallback_names or [_('未设置')]


def _resolve_management_people(names, employees_by_name):
	people = []
	seen = set()
	for name in names or []:
		employee = employees_by_name.get(_normalize_person_lookup(name))
		if employee and employee.name not in seen:
			seen.add(employee.name)
			people.append(_employee_row(employee))
	return people


def _summarize_live_nodes(nodes):
	return {
		"planned_headcount": sum(cint(node.get("planned_headcount")) for node in nodes),
		"current_headcount": sum(cint(node.get("current_headcount")) for node in nodes),
		"vacancy_count": sum(cint(node.get("vacancy_count")) for node in nodes),
	}


def _build_management_node(employee):
	return {
		"node_id": f"manager:{employee.name}",
		"id": f"manager:{employee.name}",
		"node_type": "manager",
		"name": employee.get("employee_name") or employee.name,
		"title": employee.get("designation") or "管理人员",
		"employee": employee.name,
		"department": employee.get("department"),
		"image": employee.get("image"),
		"planned_headcount": 0,
		"current_headcount": 1,
		"vacancy_count": 0,
		"children": [],
		"connections": 0,
		"expandable": False,
	}


def _build_employee_group_node(department, employees, managers):
	manager_names = {employee.name for employee in managers}
	group_employees = [employee for employee in employees if employee.name not in manager_names]
	group_people = [_employee_row(employee) for employee in group_employees]
	group_count = len(group_people)
	return {
		"node_id": f"employee_group:{department.name}",
		"id": f"employee_group:{department.name}",
		"node_type": "employee_group",
		"name": department.get("department_name") or department.name,
		"title": f"员工 {group_count} 人",
		"department": department.name,
		"planned_headcount": 0,
		"current_headcount": group_count,
		"vacancy_count": 0,
		"people": group_people,
		"children": [],
		"connections": group_count,
		"expandable": False,
	}


def _build_grade_group_nodes(department, employees, managers):
	manager_names = {employee.name for employee in managers}
	groups = defaultdict(list)
	for employee in employees:
		if employee.name in manager_names:
			continue
		grade = cstr(employee.get("grade")).strip()
		designation = cstr(employee.get("designation")).strip()
		if grade:
			group_key = ("grade", grade)
		elif designation:
			group_key = ("designation", designation)
		else:
			group_key = ("grade", _("未设置职级"))
		groups[group_key].append(employee)

	nodes = []
	for index, ((group_kind, label), group_employees) in enumerate(
		sorted(groups.items(), key=lambda item: (item[0][0] != "grade", item[0][1]))
	):
		people = [_employee_row(employee) for employee in group_employees]
		node_id = f"grade_group:{department.name}::{index}"
		nodes.append(
			{
				"node_id": node_id,
				"id": node_id,
				"node_type": "grade_group",
				"name": label,
				"title": _("职级 · {0} 人").format(len(people))
				if group_kind == "grade"
				else _("岗位 · {0} 人").format(len(people)),
				"department": department.name,
				"group_kind": group_kind,
				"group_label": label,
				"planned_headcount": 0,
				"current_headcount": len(people),
				"vacancy_count": 0,
				"people": people,
				"children": [],
				"connections": len(people),
				"expandable": False,
			}
		)
	return nodes


def _build_work_level_nodes(department, employees, managers):
	groups = defaultdict(list)
	for employee in employees:
		work_level = cstr(employee.get("grade")).strip() or _("直线级")
		groups[work_level].append(employee)

	nodes = []
	for level_index, (work_level, level_employees) in enumerate(sorted(groups.items())):
		children = _build_position_group_nodes(department, work_level, level_index, level_employees)
		group_people = [_employee_row(employee) for employee in level_employees]
		node_id = f"work_level:{department.name}::{level_index}"
		nodes.append(
			{
				"node_id": node_id,
				"id": node_id,
				"node_type": "work_level",
				"name": work_level,
				"title": _("职级 · {0} 人").format(len(group_people)),
				"department": department.name,
				"work_level": work_level,
				"planned_headcount": 0,
				"current_headcount": len(group_people),
				"vacancy_count": 0,
				"people": [],
				"children": children,
				"connections": len(children),
				"expandable": bool(children),
			}
		)
	return nodes


def _build_position_group_nodes(department, work_level, level_index, employees):
	groups = defaultdict(list)
	for employee in employees:
		position = cstr(employee.get("designation")).strip() or _("员工")
		groups[position].append(employee)

	nodes_by_designation = {}
	for position_index, (position, position_employees) in enumerate(sorted(groups.items())):
		group_people = [_employee_row(employee) for employee in position_employees]
		node_id = f"position_group:{department.name}::{level_index}::{position_index}"
		employee_nodes = [_build_employee_node(employee) for employee in position_employees]
		nodes_by_designation[position] = {
			"node_id": node_id,
			"id": node_id,
			"node_type": "position_group",
			"name": position,
			"title": _("岗位 · {0} 人").format(len(group_people)),
			"department": department.name,
			"work_level": work_level,
			"designation": position,
			"planned_headcount": 0,
			"current_headcount": len(group_people),
			"vacancy_count": 0,
			"people": [],
			"children": employee_nodes,
			"connections": len(employee_nodes),
			"expandable": bool(employee_nodes),
		}

	parents = _get_designation_parent_map(nodes_by_designation)
	roots = []
	for designation, node in nodes_by_designation.items():
		parent = parents.get(designation)
		if parent and parent in nodes_by_designation and parent != designation:
			nodes_by_designation[parent]["children"].insert(0, node)
			nodes_by_designation[parent]["connections"] = len(nodes_by_designation[parent]["children"])
			nodes_by_designation[parent]["expandable"] = True
		else:
			roots.append(node)
	return roots


def _build_employee_node(employee):
	name = employee.get("employee_name") or employee.name
	return {
		"node_id": f"employee:{employee.name}",
		"id": f"employee:{employee.name}",
		"node_type": "employee",
		"name": name,
		"title": employee.get("designation") or _("员工"),
		"employee": employee.name,
		"employee_route": employee.name,
		"employee_code": _employee_business_number(employee),
		"department": employee.get("department"),
		"work_level": cstr(employee.get("grade")).strip() or _("直线级"),
		"designation": employee.get("designation") or _("员工"),
		"people": [],
		"planned_headcount": 0,
		"current_headcount": 1,
		"vacancy_count": 0,
		"children": [],
		"connections": 0,
		"expandable": False,
	}


def _get_designation_parent_map(designations):
	if not designations or not frappe.get_meta("Designation").has_field("hrms_parent_designation"):
		return {}
	rows = frappe.get_all(
		"Designation",
		filters={"name": ["in", list(designations)]},
		fields=["name", "hrms_parent_designation"],
	)
	return {row.name: row.hrms_parent_designation for row in rows if row.get("hrms_parent_designation")}


def _get_live_group_node_detail(node_id, company=None, search=None):
	context = _get_live_node_context(node_id, company)
	department = context.get("department")
	department_row = context.get("department_row") or frappe._dict(name=department, department_name=department)
	group = context.get("node")
	if not group:
		return {
			"node_type": _node_type_from_id(node_id),
			"node_id": node_id,
			"title": _("职级或岗位"),
			"subtitle": department,
			"metrics": {},
			"employees": [],
			"actions": {},
		}

	people = _collect_node_people(group)
	if search:
		keyword = cstr(search).lower()
		people = [
			person
			for person in people
			if keyword
			in " ".join(
				cstr(person.get(field))
				for field in ("employee_name", "employee_code", "designation", "grade", "department")
			).lower()
		]
	return {
		"node_type": group.get("node_type"),
		"node_id": node_id,
		"department": department,
		"title": group.get("name"),
		"subtitle": _("{0} · {1}").format(group.get("title"), department),
		"metrics": {"current_headcount": len(people), "employee_count": len(people)},
		"employees": people[:100],
		"people": people,
		"relationships": _get_department_relationships(department_row, company),
		"actions": {},
	}


def _flatten_tree_nodes(nodes):
	result = []
	for node in nodes or []:
		result.append(node)
		result.extend(_flatten_tree_nodes(node.get("children") or []))
	return result


def _collect_node_people(node):
	people = list(node.get("people") or [])
	if node.get("node_type") == "employee" and node.get("employee"):
		people.append({
			"employee": node.get("employee"),
			"name": node.get("employee"),
			"employee_name": node.get("name"),
			"employee_code": node.get("employee_code"),
			"department": node.get("department"),
			"designation": node.get("designation"),
			"grade": node.get("work_level"),
		})
	for child in node.get("children") or []:
		people.extend(_collect_node_people(child))
	seen = set()
	return [person for person in people if person.get("employee") and not (person.get("employee") in seen or seen.add(person.get("employee")))]


def _get_live_node_context(node_id, company=None):
	payload = _node_value(node_id)
	department = payload.split("::", 1)[0]
	if not department:
		return {}
	employees = [employee for employee in _get_active_employees(company) if employee.get("department") == department]
	department_row = (
		frappe.get_cached_doc("Department", department)
		if frappe.db.exists("Department", department)
		else frappe._dict(name=department, department_name=department)
	)
	work_levels = _build_work_level_nodes(department_row, employees, _get_department_managers(employees))
	node = next((item for item in _flatten_tree_nodes(work_levels) if item.get("node_id") == node_id), None)
	if not node:
		return {}
	return {"node": node, "department": department, "department_row": department_row, "work_level": node.get("work_level"), "designation": node.get("designation")}


def _get_grade_group_node_detail(node_id, company=None, search=None):
	payload = _node_value(node_id)
	department = payload.rsplit("::", 1)[0] if "::" in payload else payload
	employees = [employee for employee in _get_active_employees(company) if employee.get("department") == department]
	managers = _get_department_managers(employees)
	department_row = frappe._dict(name=department, department_name=department)
	group = next(
		(row for row in _build_grade_group_nodes(department_row, employees, managers) if row["node_id"] == node_id),
		None,
	)
	if not group:
		return {
			"node_type": "grade_group",
			"node_id": node_id,
			"title": _("职级分组"),
			"subtitle": department,
			"metrics": {},
			"employees": [],
			"actions": {"can_add_department": frappe.has_permission("Department", "create")},
		}

	people = group.get("people") or []
	if search:
		keyword = cstr(search).lower()
		people = [
			person
			for person in people
			if keyword
			in " ".join(
				cstr(person.get(field))
				for field in ("employee_name", "employee_code", "designation", "grade", "department")
			).lower()
		]
	return {
		"node_type": "grade_group",
		"node_id": node_id,
		"department": department,
		"title": group.get("name"),
		"subtitle": _("{0} · {1}").format(group.get("title"), department),
		"metrics": {"current_headcount": len(people), "employee_count": len(people)},
		"employees": people[:100],
		"people": people,
		"actions": {"can_add_department": frappe.has_permission("Department", "create")},
	}


def _manager_label(managers):
	if not managers:
		return ""
	return "、".join((employee.get("employee_name") or employee.name) for employee in managers[:3])


def _department_manager_label(department):
	role = department.get("hrms_org_role") or ""
	manager = department.get("hrms_org_manager") or ""
	return "：".join(part for part in [role, manager] if part)


def _department_subtitle(department_doc):
	if not department_doc:
		return ""
	return ""


def _get_department_relationships(department_doc, company=None):
	if not department_doc or not department_doc.get("name"):
		return {"parent": None, "children": []}

	department_name = department_doc.get("name")
	parent_name = cstr(department_doc.get("parent_department")).strip()
	parent = None
	parent_label = frappe.db.get_value("Department", parent_name, "department_name") if parent_name else ""
	if (
		parent_name
		and not _is_all_departments_name(parent_name)
		and not _is_all_departments_name(parent_label)
		and frappe.db.exists("Department", parent_name)
	):
		parent_label = parent_label or parent_name
		parent = {"name": parent_name, "label": parent_label}

	filters = {"parent_department": department_name}
	if company and company != "All Companies":
		filters["company"] = company
	children = frappe.get_list(
		"Department",
		filters=filters,
		fields=["name", "department_name"],
		order_by="department_name asc",
		limit_page_length=0,
	)
	return {
		"parent": parent,
		"children": [
			{"name": child.name, "label": child.get("department_name") or child.name}
			for child in children
		],
	}


def _is_all_departments_name(value):
	value = re.sub(r"\s+", "", cstr(value)).lower()
	return value in {"alldepartments", "all部门s", "所有部门"}


def _node_type_from_id(node_id):
	if ":" in node_id:
		return node_id.split(":", 1)[0]
	return ""


def _node_value(node_id):
	if ":" in node_id:
		return node_id.split(":", 1)[1]
	return node_id


def _would_create_employee_reporting_loop(employee, manager):
	"""Return true when assigning manager would create an Employee.reports_to cycle."""
	seen = {employee}
	current = manager
	while current:
		if current in seen:
			return True
		seen.add(current)
		current = frappe.db.get_value("Employee", current, "reports_to")
	return False


def _would_create_designation_loop(designation, parent_designation):
	seen = {designation}
	current = parent_designation
	while current:
		if current in seen:
			return True
		seen.add(current)
		current = frappe.db.get_value("Designation", current, "hrms_parent_designation")
	return False


def _get_node_employees(department=None, manager=None, company=None, search=None):
	filters = {"status": "Active"}
	if company:
		filters["company"] = company
	if department:
		filters["department"] = department
	if manager:
		filters["reports_to"] = manager
	or_filters = None
	if search:
		or_filters = {
			"employee_name": ["like", f"%{search}%"],
			"name": ["like", f"%{search}%"],
			"custom_employee_code": ["like", f"%{search}%"],
		}
	fields = _get_meta_fields(
		"Employee",
		[
			"name",
			"employee_name",
			"custom_employee_code",
			"department",
			"designation",
			"grade",
			"reports_to",
			"branch",
			"cell_number",
			"image",
		],
	)
	return [
		_employee_row(employee)
		for employee in frappe.get_all(
			"Employee",
			fields=fields,
			filters=filters,
			or_filters=or_filters,
			order_by="department asc, designation asc, employee_name asc",
			limit_page_length=500,
		)
	]


def _employee_row(employee):
	return {
		"name": employee.get("name"),
		"employee": employee.get("name"),
		"employee_route": employee.get("name"),
		"employee_name": employee.get("employee_name") or employee.get("name"),
		"employee_code": _employee_business_number(employee),
		"department": employee.get("department"),
		"designation": employee.get("designation"),
		"grade": employee.get("grade"),
		"reports_to": employee.get("reports_to"),
		"branch": employee.get("branch"),
		"cell_number": employee.get("cell_number"),
		"image": employee.get("image"),
		"matched_employee": True,
		"match_status": _("已匹配员工档案"),
	}
