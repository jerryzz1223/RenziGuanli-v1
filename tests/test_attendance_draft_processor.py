"""Contracts for the single-result DingTalk attendance-draft processor."""

from __future__ import annotations

import importlib.util
import sys
import unittest
from pathlib import Path

try:
	from openpyxl import load_workbook
except ModuleNotFoundError:
	load_workbook = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PROCESSOR_PATH = PROJECT_ROOT / "hrms" / "api" / "attendance_processors" / "attendance_draft.py"
REAL_WORKBOOK = Path("/Users/lrj/Desktop/薪酬计算设计表单/考勤数据/1.考勤初稿.xlsx")


def _processor():
	spec = importlib.util.spec_from_file_location("attendance_draft_processor_contract", PROCESSOR_PATH)
	module = importlib.util.module_from_spec(spec)
	sys.modules[spec.name] = module
	spec.loader.exec_module(module)
	return module


processor = _processor()


class _FakeWorksheet:
	def __init__(self, title, rows):
		self.title = title
		self._rows = rows

	def iter_rows(self, min_row=1, max_row=None, values_only=False):
		end = max_row or len(self._rows)
		yield from self._rows[min_row - 1 : end]


class _FakeWorkbook:
	def __init__(self, sheets):
		self.worksheets = sheets


class AttendanceDraftProcessorContractTest(unittest.TestCase):
	def test_daily_statistics_export_is_detected_by_headers_not_worksheet_name(self):
		sheet = _FakeWorksheet("每日统计", [
			("每日统计配置版 统计日期：2026-07-01 至 2026-07-31",),
			("报表生成时间：2026-08-20 09:56",),
			("姓名", "工号", "日期", "实际部门", "班次", "标准工时", "实际出勤（小时）", "请假"),
			("", "", "", "", "", "", "", "事假(小时)"),
			("张三", "E-001", "26-07-01 星期三", "工程课", "白班", 8, 8, 0),
		])

		location = processor.dingtalk_daily_header_location(sheet)
		rows = processor.rows_from_dingtalk_daily_sheet(sheet, source_file="daily.xlsx")

		self.assertEqual(location["header_row"], 3)
		self.assertIs(processor.find_dingtalk_daily_sheet(_FakeWorkbook([sheet])), sheet)
		self.assertEqual(rows[0]["source_row"], 5)
		self.assertEqual(rows[0]["工号"], "E-001")
		self.assertEqual(rows[0]["请假/事假(小时)"], 0)

	def test_structure_and_single_employee_result_contract(self):
		rows = [
			{
				"姓名": "张三",
				"工号": "E-001",
				"日期": "26-06-01 星期一",
				"实际部门": "工程课",
				"班次": "白班",
				"标准工时": "8",
				"实际出勤（小时）": "7.5",
				"工作日加班（小时）": "1",
				"source_file": "sample.xlsx",
				"source_sheet": "每日明细（钉钉导出）",
				"source_row": 3,
			},
			{
				"姓名": "张三",
				"工号": "E-001",
				"日期": "26-06-02 星期二",
				"实际部门": "工程课",
				"班次": "白班",
				"标准工时": 8,
				"实际出勤（小时）": 8,
				"工作日加班（小时）": 0.5,
				"source_file": "sample.xlsx",
				"source_sheet": "每日明细（钉钉导出）",
				"source_row": 4,
			},
		]
		result = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")

		self.assertEqual(result["status"], "待确认")
		self.assertEqual(result["metrics"]["source_rows"], 2)
		self.assertEqual(result["metrics"]["processed_rows"], 1)
		row = result["processed_rows"][0]
		self.assertEqual(row["source_type"], "attendance_draft")
		self.assertEqual(row["employee_code"], "E-001")
		self.assertEqual(row["processed_value"]["standard_hours"], 16)
		self.assertEqual(row["processed_value"]["actual_attendance_hours"], 15.5)
		self.assertEqual(row["processed_value"]["workday_overtime_hours"], 1.5)
		self.assertEqual(row["review_status"], "无需审核")
		self.assertTrue(row["eligible_for_downstream"])
		self.assertEqual(row["source_row"], 3)
		self.assertEqual(len(row["original_value"]["source_rows"]), 2)
		self.assertEqual([item["attendance_date"] for item in row["processed_value"]["attendance_details"]], ["2026-06-01", "2026-06-02"])

	def test_missing_employee_code_source_accounts_do_not_become_employee_exceptions(self):
		rows = [
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课", "班次": "白班", "标准工时": 8, "source_row": 3},
			{"姓名": "车间二楼放料", "工号": "", "UserId": "station-001", "日期": "26-06-01", "实际部门": "连续课", "班次": "未排班", "source_row": 4},
			{"姓名": "车间二楼放料", "工号": "", "UserId": "station-001", "日期": "26-06-02", "实际部门": "连续课", "班次": "未排班", "source_row": 5},
		]

		result = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")

		self.assertEqual(result["metrics"]["source_rows"], 3)
		self.assertEqual(result["metrics"]["eligible_employee_source_rows"], 1)
		self.assertEqual(result["metrics"]["excluded_missing_employee_code_rows"], 2)
		self.assertEqual(result["metrics"]["excluded_missing_employee_code_accounts"], 1)
		self.assertEqual(result["metrics"]["processed_rows"], 1)
		self.assertEqual(result["processed_rows"][0]["employee_name"], "张三")
		self.assertEqual(result["data_quality"]["excluded_missing_employee_code_accounts"][0]["source_account_name"], "车间二楼放料")

	def test_duplicate_dates_and_identity_conflicts_enter_review_without_loss(self):
		rows = [
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课", "班次": "白班", "标准工时": 8, "source_file": "a.xlsx", "source_sheet": "每日明细（钉钉导出）", "source_row": 3},
			{"姓名": "李四", "工号": "E-001", "日期": "26-06-01", "实际部门": "品质课", "班次": "", "标准工时": "bad", "source_file": "a.xlsx", "source_sheet": "每日明细（钉钉导出）", "source_row": 4},
		]
		result = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")

		self.assertEqual(result["metrics"]["processed_rows"], 1)
		self.assertEqual(result["metrics"]["source_rows"], 2)
		row = result["processed_rows"][0]
		self.assertEqual(row["review_status"], "待审核")
		self.assertFalse(row["eligible_for_downstream"])
		self.assertTrue({"ATTENDANCE_DATE_DUPLICATE", "EMPLOYEE_CODE_NAME_CONFLICT", "EMPLOYEE_DEPARTMENT_CONFLICT", "INVALID_NUMERIC_VALUE"}.issubset(row["exception_codes"]))

	def test_explicit_dingtalk_missing_punch_counts_remain_downstream_attendance_facts(self):
		rows = [{
			"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课", "班次": "白班",
			"标准工时": 8, "上班时间": "08:01", "下班时间": "17:30", "上班未打卡次数": 1, "下班未打卡次数": 2,
			"source_file": "sample.xlsx", "source_sheet": "每日明细（钉钉导出）", "source_row": 3,
		}]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertEqual(row["processed_value"]["clock_in_missing_count"], 1)
		self.assertEqual(row["processed_value"]["clock_out_missing_count"], 2)
		self.assertTrue({"CLOCK_IN_MISSING", "CLOCK_OUT_MISSING"}.issubset(row["exception_codes"]))
		self.assertEqual(row["review_status"], "无需审核")
		self.assertTrue(row["eligible_for_downstream"])
		self.assertEqual(row["processed_value"]["attendance_details"], [{
			"attendance_date": "2026-06-01", "shift": "白班", "clock_in": "08:01", "clock_out": "17:30",
			"clock_in_missing": 1, "clock_out_missing": 2, "late_count": 0, "early_count": 0,
			"absence_marker_count": 0, "absence_hours": 0, "source_row": 3,
		}])

	def test_daily_statistics_missing_card_markers_are_source_facts(self):
		rows = [{
			"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课", "班次": "夜班",
			"上班缺卡": "是", "下班缺卡": 1, "source_row": 5,
		}]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertEqual(row["processed_value"]["clock_in_missing_count"], 1)
		self.assertEqual(row["processed_value"]["clock_out_missing_count"], 1)
		self.assertEqual(row["processed_value"]["exception_lines"], [{
			"attendance_date": "2026-06-01", "shift": "夜班", "clock_in": "", "clock_out": "",
			"clock_in_missing": 1, "clock_out_missing": 1, "late_count": 0, "early_count": 0,
			"absence_marker_count": 0, "absence_hours": 0, "source_row": 5,
			"exception_codes": ["CLOCK_IN_MISSING", "CLOCK_OUT_MISSING"],
		}])

	def test_exactly_one_clock_time_creates_the_missing_side_review_event(self):
		rows = [
			{
				"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课", "班次": "白班",
				"上班时间": "08:00", "下班时间": "", "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 3,
			},
			{
				"姓名": "张三", "工号": "E-001", "日期": "26-06-02", "实际部门": "工程课", "班次": "白班",
				"上班时间": "", "下班时间": "17:00", "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 4,
			},
		]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertEqual(row["processed_value"]["clock_in_missing_count"], 1)
		self.assertEqual(row["processed_value"]["clock_out_missing_count"], 1)
		self.assertTrue({"CLOCK_IN_MISSING", "CLOCK_OUT_MISSING"}.issubset(row["exception_codes"]))
		self.assertEqual([(item["attendance_date"], item["exception_codes"]) for item in row["processed_value"]["exception_lines"]], [
			("2026-06-01", ["CLOCK_OUT_MISSING"]),
			("2026-06-02", ["CLOCK_IN_MISSING"]),
		])

	def test_no_clock_times_is_not_mistaken_for_a_single_punch(self):
		rows = [{
			"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课", "班次": "白班",
			"上班时间": "", "下班时间": "", "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 3,
		}]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertEqual(row["processed_value"]["clock_in_missing_count"], 0)
		self.assertEqual(row["processed_value"]["clock_out_missing_count"], 0)
		self.assertNotIn("CLOCK_IN_MISSING", row["exception_codes"])
		self.assertNotIn("CLOCK_OUT_MISSING", row["exception_codes"])

	def test_daily_attendance_facts_stay_downstream_and_workday_absence_becomes_payroll_hours(self):
		rows = [
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-07", "日期类型": "周末休息日", "实际部门": "工程课", "班次": "白班", "旷工": 1, "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 3},
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-08", "日期类型": "工作日", "实际部门": "工程课", "班次": "白班 08:00-17:00", "标准工时": 8, "旷工": 1, "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 4},
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-09", "日期类型": "工作日", "实际部门": "工程课", "班次": "白班 08:00-17:00", "标准工时": 8, "实际出勤（小时）": 6, "下班时间": "15:00", "早退次数": 1, "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 5},
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-10", "日期类型": "工作日", "实际部门": "工程课", "班次": "白班 08:00-17:00", "标准工时": 8, "实际出勤（小时）": 7.5, "迟到次数": 1, "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 6},
		]
		result = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")
		row = result["processed_rows"][0]

		self.assertEqual(row["processed_value"]["late_count"], 1)
		self.assertEqual(row["processed_value"]["early_count"], 1)
		self.assertEqual(row["processed_value"]["absence_marker_count"], 2)
		self.assertEqual(row["processed_value"]["absence_hours"], 10)
		self.assertTrue({"CLOCK_IN_MISSING", "LATE_MARKED", "EARLY_MARKED", "ABSENCE_MARKED"}.issubset(row["exception_codes"]))
		self.assertEqual(result["metrics"]["exception_events"], 4)
		self.assertEqual(row["review_status"], "无需审核")
		self.assertTrue(row["eligible_for_downstream"])
		self.assertEqual([(event["attendance_date"], event["code"]) for event in row["exception_events"]], [
			("2026-06-08", "ABSENCE_MARKED"),
			("2026-06-09", "CLOCK_IN_MISSING"),
			("2026-06-09", "EARLY_MARKED"),
			("2026-06-10", "LATE_MARKED"),
		])

	def test_rest_day_clock_without_overtime_application_requires_manual_hours_confirmation(self):
		rows = [{
			"姓名": "张三", "工号": "E-001", "日期": "26-06-07", "日期类型": "周末休息日", "实际部门": "工程课",
			"班次": "休息", "上班时间": "09:02", "下班时间": "17:41", "实际出勤（小时）": 0,
			"休息日加班（小时）": 0, "关联审批单": "", "source_row": 3,
		}]

		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertIn("RESTDAY_CLOCKED_WITHOUT_OVERTIME", row["exception_codes"])
		self.assertEqual(row["review_status"], "待审核")
		self.assertFalse(row["eligible_for_downstream"])
		self.assertEqual(row["exception_events"], [{
			"code": "RESTDAY_CLOCKED_WITHOUT_OVERTIME", "attendance_date": "2026-06-07", "source_row": 3,
		}])
		self.assertEqual(row["processed_value"]["exception_lines"][0]["restday_overtime_hours"], 0)
		self.assertTrue(row["processed_value"]["exception_lines"][0]["restday_clocked_without_overtime"])

	def test_rest_day_clock_with_overtime_application_does_not_require_manual_confirmation(self):
		rows = [{
			"姓名": "张三", "工号": "E-001", "日期": "26-06-07", "日期类型": "周末休息日", "实际部门": "工程课",
			"班次": "休息", "上班时间": "09:02", "下班时间": "17:41", "休息日加班（小时）": 0,
			"关联审批单": "加班申请 OT-001", "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 3,
		}]

		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertNotIn("RESTDAY_CLOCKED_WITHOUT_OVERTIME", row["exception_codes"])
		self.assertEqual(row["review_status"], "无需审核")

	def test_early_departure_with_leave_evidence_does_not_create_absence_hours(self):
		rows = [{
			"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "日期类型": "工作日", "实际部门": "工程课",
			"班次": "生产夜班 20:00-次日04:30", "标准工时": 8, "实际出勤（小时）": 3.5,
			"下班时间": "次日 00:04", "早退次数": 1, "关联审批单": "事假 4.5小时",
			"source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 3,
		}]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertEqual(row["processed_value"]["absence_hours"], 0)
		self.assertEqual(row["review_status"], "无需审核")
		self.assertTrue(row["eligible_for_downstream"])

	def test_historic_attendance_details_restore_only_actual_exception_dates(self):
		lines = processor.exception_lines_from_attendance_details([
			{"attendance_date": "2026-06-01", "early_count": 0, "clock_in_missing": 0, "clock_out_missing": 0, "late_count": 0, "absence_marker_count": 0, "source_row": 3},
			{"attendance_date": "2026-06-02", "early_count": 1, "clock_in_missing": 0, "clock_out_missing": 0, "late_count": 0, "absence_marker_count": 0, "source_row": 4},
			{"attendance_date": "2026-06-03", "early_count": 0, "clock_in_missing": 0, "clock_out_missing": 1, "late_count": 0, "absence_marker_count": 0, "source_row": 5},
		], ["EARLY_MARKED", "CLOCK_OUT_MISSING"])

		self.assertEqual([(line["attendance_date"], line["exception_codes"]) for line in lines], [
			("2026-06-02", ["EARLY_MARKED"]),
			("2026-06-03", ["CLOCK_OUT_MISSING"]),
		])

	def test_one_minute_late_without_leave_is_a_review_event(self):
		rows = [{
			"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课",
			"班次": "白班 08:00-17:00", "标准工时": 8, "上班时间": "08:01", "下班时间": "17:00",
			"source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 3,
		}]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertEqual(row["processed_value"]["late_count"], 1)
		self.assertIn("LATE_MARKED", row["exception_codes"])
		self.assertEqual(row["processed_value"]["exception_lines"][0]["attendance_date"], "2026-06-01")

	def test_leave_evidence_prevents_time_only_late_alert(self):
		rows = [{
			"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课",
			"班次": "白班 08:00-17:00", "标准工时": 8, "上班时间": "08:01", "下班时间": "17:00",
			"请假/事假(小时)": 1, "迟到次数": 1, "source_file": "sample.xlsx", "source_sheet": "每日统计", "source_row": 3,
		}]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		self.assertEqual(row["processed_value"]["late_count"], 0)
		self.assertNotIn("LATE_MARKED", row["exception_codes"])

	def test_blank_shift_outside_known_employment_period_is_data_quality_only(self):
		rows = [
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课", "班次": "", "source_row": 3},
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-11", "实际部门": "工程课", "班次": "", "source_row": 4},
			{"姓名": "张三", "工号": "E-001", "日期": "26-06-21", "实际部门": "工程课", "班次": "", "source_row": 5},
		]
		roster = [{"employee_code": "E-001", "employee_name": "张三", "department": "工程课", "date_of_joining": "2026-06-10", "relieving_date": "2026-06-20"}]
		result = processor.process_attendance_draft_rows(rows, attendance_month="2026-06", employee_directory=roster, source_file="sample.xlsx", source_sheet="每日统计")
		row = result["processed_rows"][0]

		self.assertEqual(row["exception_codes"], [])
		self.assertEqual(row["exception_events"], [])
		self.assertEqual([event["attendance_date"] for event in row["data_quality_events"]], ["2026-06-01", "2026-06-11", "2026-06-21"])
		self.assertEqual(result["data_quality"]["lifecycle_excluded_blank_shift_rows"], 3)

	def test_duplicate_dingtalk_headers_do_not_overwrite_the_first_column(self):
		headers = processor.flatten_dingtalk_headers(("姓名", "旷工", "旷工"), ("", "", ""))
		self.assertEqual(headers, ["姓名", "旷工", "旷工_2"])

	def test_large_night_shift_uses_only_the_dingtalk_export_value(self):
		rows = [
			{
				"姓名": "张三", "工号": "E-001", "日期": "26-06-01", "实际部门": "工程课", "班次": "夜班", "标准工时": 8,
				"上班时间": "20:05", "下班时间": "07:58", "大夜班": 9,
				"source_file": "sample.xlsx", "source_sheet": "每日明细（钉钉导出）", "source_row": 3,
			},
			{
				"姓名": "张三", "工号": "E-001", "日期": "26-06-02", "实际部门": "工程课", "班次": "夜班", "标准工时": 8,
				"上班时间": "20:05", "大夜班": 1,
				"source_file": "sample.xlsx", "source_sheet": "每日明细（钉钉导出）", "source_row": 4,
			},
		]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06")["processed_rows"][0]

		# HRMS does not use clock times or a local rule to replace DingTalk's count.
		self.assertEqual(row["processed_value"]["large_night_shifts"], 10)
		self.assertEqual(row["processed_value"]["night_shift_matching"]["mode"], "source_only")

	def test_department_group_and_section_suffixes_are_the_same_department(self):
		rows = [{
			"姓名": "朱耀辉", "工号": "164", "日期": "26-06-01", "实际部门": "设备组", "班次": "白班", "标准工时": 8,
			"source_file": "sample.xlsx", "source_sheet": "每日明细（钉钉导出）", "source_row": 3,
		}]
		roster = [{"employee_code": "164", "employee_name": "朱耀辉", "department": "设备课"}]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06", employee_directory=roster)["processed_rows"][0]

		self.assertNotIn("EMPLOYEE_DEPARTMENT_MISMATCH", row["exception_codes"])
		self.assertEqual(row["review_status"], "无需审核")
		self.assertTrue(row["eligible_for_downstream"])

	def test_dingtalk_department_identifier_is_removed_before_matching_and_display(self):
		rows = [{
			"姓名": "朱耀辉", "工号": "164", "日期": "26-06-01", "实际部门": "工程课 - 11", "班次": "白班", "标准工时": 8,
			"source_file": "sample.xlsx", "source_sheet": "每日明细（钉钉导出）", "source_row": 3,
		}]
		roster = [{"employee_code": "164", "employee_name": "朱耀辉", "department": "工程课"}]
		row = processor.process_attendance_draft_rows(rows, attendance_month="2026-06", employee_directory=roster)["processed_rows"][0]

		self.assertNotIn("EMPLOYEE_DEPARTMENT_MISMATCH", row["exception_codes"])
		self.assertEqual(row["department"], "工程课")
		self.assertEqual(row["processed_value"]["department"], "工程课")

	@unittest.skipUnless(load_workbook is not None, "openpyxl is unavailable")
	@unittest.skipUnless(REAL_WORKBOOK.exists(), "Local real DingTalk sample is unavailable")
	def test_real_dingtalk_sample_has_5820_source_rows_and_194_employee_results(self):
		book = load_workbook(REAL_WORKBOOK, data_only=True, read_only=True)
		rows = processor.rows_from_dingtalk_daily_sheet(book["每日明细（钉钉导出）"], source_file=str(REAL_WORKBOOK))
		result = processor.process_attendance_draft_rows(rows, attendance_month="2026-06", source_file=str(REAL_WORKBOOK))

		self.assertEqual(result["metrics"]["source_rows"], 5820)
		self.assertEqual(result["metrics"]["processed_rows"], 194)
		yang_bo = next(row for row in result["processed_rows"] if row["employee_code"] == "946")
		self.assertEqual(yang_bo["processed_value"]["standard_hours"], 168)
		self.assertEqual(yang_bo["processed_value"]["actual_attendance_hours"], 156.5)
		self.assertEqual(yang_bo["processed_value"]["workday_overtime_hours"], 28)
		self.assertEqual(yang_bo["processed_value"]["restday_overtime_hours"], 9)
		self.assertEqual(yang_bo["processed_value"]["annual_leave_hours"], 8)
		# Daily late/early/missing-card facts remain visible as events, but they
		# no longer remove employees with valid source rows from the monthly final.
		self.assertGreater(result["metrics"]["exception_events"], 0)
		self.assertEqual(result["metrics"]["eligible_rows"], 194)


if __name__ == "__main__":
	unittest.main()
