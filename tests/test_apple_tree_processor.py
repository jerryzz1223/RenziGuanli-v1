from __future__ import annotations

import importlib.util
import sys
import unittest
from datetime import datetime
from io import BytesIO
from pathlib import Path
from types import ModuleType, SimpleNamespace
from unittest.mock import patch

try:
	from openpyxl import load_workbook
except ModuleNotFoundError:
	load_workbook = None


MODULE_PATH = (
	Path(__file__).parents[1] / "hrms" / "api" / "attendance_processors" / "apple_tree.py"
)
REAL_WORKBOOK = Path("/Users/lrj/Desktop/薪酬计算设计表单/考勤数据/2.苹果树合计.xlsx")
HOUSING_ALLOWANCE_WORKBOOK = Path("/Users/lrj/Desktop/薪酬计算设计表单/考勤数据/4.住房补贴（月）.xlsx")
FULL_ATTENDANCE_WORKBOOK = Path("/Users/lrj/Desktop/薪酬计算设计表单/考勤数据/5.全勤奖（月）.xlsx")
SPECIAL_HOURS_WORKBOOK = Path("/Users/lrj/Desktop/薪酬计算设计表单/考勤数据/7.特殊工时（月）.xlsx")
ATTENDANCE_SOURCE_DIR = Path("/Users/lrj/Desktop/薪酬计算设计表单/考勤数据")
ATTENDANCE_PAGE = (
	Path(__file__).parents[1]
	/ "hrms"
	/ "hr"
	/ "page"
	/ "attendance_import_center"
	/ "attendance_import_center.js"
)
PROCESSING_CENTER = Path(__file__).parents[1] / "hrms" / "api" / "attendance_processing_center.py"
SPEC = importlib.util.spec_from_file_location("apple_tree_processor", MODULE_PATH)
apple_tree = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = apple_tree
SPEC.loader.exec_module(apple_tree)

AppleTreeRules = apple_tree.AppleTreeRules
apply_reviews = apple_tree.apply_reviews
build_employee_summary = apple_tree.build_employee_summary
normalize_apple_tree_rows = apple_tree.normalize_apple_tree_rows
preflight_apple_tree_rows = apple_tree.preflight_apple_tree_rows
process_apple_tree_rows = apple_tree.process_apple_tree_rows


def processing_center_module():
	"""Load the export helpers with a minimal Frappe facade for workbook tests."""
	frappe = ModuleType("frappe")
	frappe._ = lambda text: text
	def whitelist(function=None, **_kwargs):
		return (lambda decorated: decorated) if function is None else function

	frappe.whitelist = whitelist
	frappe.PermissionError = PermissionError
	frappe.session = SimpleNamespace(user="test@example.com")
	frappe.get_all = lambda *args, **kwargs: []
	frappe.db = SimpleNamespace(
		count=lambda *args, **kwargs: 0,
		get_value=lambda *args, **kwargs: None,
	)
	frappe_utils = ModuleType("frappe.utils")
	frappe_utils.cint = lambda value: int(value or 0)
	frappe_utils.flt = lambda value, *_args, **_kwargs: float(value or 0)
	frappe_utils.now_datetime = lambda: None
	file_manager = ModuleType("frappe.utils.file_manager")
	frappe_utils.file_manager = file_manager
	hrms = ModuleType("hrms")
	hrms.__path__ = [str(PROCESSING_CENTER.parents[1])]
	hrms_api = ModuleType("hrms.api")
	hrms_api.__path__ = [str(PROCESSING_CENTER.parent)]
	processors = ModuleType("hrms.api.attendance_processors")
	processors.__path__ = [str(PROCESSING_CENTER.parent / "attendance_processors")]
	module_names = (
		"frappe", "frappe.utils", "frappe.utils.file_manager", "hrms", "hrms.api", "hrms.api.attendance_processors"
	)
	old_modules = {name: sys.modules.get(name) for name in module_names}
	fake_modules = {
		"frappe": frappe,
		"frappe.utils": frappe_utils,
		"frappe.utils.file_manager": file_manager,
		"hrms": hrms,
		"hrms.api": hrms_api,
		"hrms.api.attendance_processors": processors,
	}
	sys.modules.update(fake_modules)
	try:
		spec = importlib.util.spec_from_file_location("attendance_processing_center_export_contract", PROCESSING_CENTER)
		module = importlib.util.module_from_spec(spec)
		sys.modules[spec.name] = module
		spec.loader.exec_module(module)
		return module, file_manager, fake_modules
	finally:
		for name, old_module in old_modules.items():
			if old_module is None:
				sys.modules.pop(name, None)
			else:
				sys.modules[name] = old_module


EMPLOYEES = [
	{"employee_code": "E-001", "employee_name": "张三", "department": "连续课", "employment_status": "在职"},
	{"employee_code": "E-002", "employee_name": "李四", "department": "品管课", "employment_status": "在职"},
]


def source_row(**overrides):
	row = {
		"数据id": "DATA-001",
		"审批编号": "APPROVAL-001",
		"奖/惩日期": "2026-06-15",
		"创建时间": "2026-06-16 08:30:00",
		"受奖/惩人部门": "连续课",
		"受奖/惩人": "张三",
		"绿苹果": "2",
		"红苹果": "15",
		"奖/惩项目": "连续课/绿苹果/保养/作业员完成保养。2颗",
		"备注": "保养A线",
		"创建人": "主管甲",
		"审批结果": "审批通过",
		"审批状态": "已结束",
	}
	row.update(overrides)
	return row


class AppleTreeProcessorContractTest(unittest.TestCase):
	def confirmed_rules(self):
		return AppleTreeRules(target_month="2026-06")

	def test_green_project_uses_only_green_value_and_keeps_provenance(self):
		processed = normalize_apple_tree_rows(
			[source_row()],
			rules=self.confirmed_rules(),
			employees=EMPLOYEES,
			source_file="苹果树.xlsx",
			source_sheet="钉钉导出数据",
			start_row=2,
		)

		self.assertEqual(len(processed), 1)
		row = processed[0]
		self.assertEqual(row["苹果类型"], "绿苹果")
		self.assertEqual(row["原始有效苹果数"], 2)
		self.assertEqual(row["有效苹果数"], 2)
		self.assertEqual(row["工号"], "E-001")
		self.assertTrue(row["include_in_downstream"])
		self.assertEqual(row["exception_codes"], [])
		self.assertEqual(row["exception_message"], "")
		self.assertEqual(row["review_status"], "无需审核")
		self.assertEqual(row["proposed_value"]["有效苹果数"], 2)
		self.assertEqual(row["processed_value"]["奖惩日期"], "2026-06-15")
		self.assertEqual(row["processed_value"]["项目"], row["项目"])
		self.assertEqual(row["processed_value"]["有效苹果数"], 2)
		self.assertIsNone(row["confirmed_value"])
		self.assertEqual(row["source_file"], "苹果树.xlsx")
		self.assertEqual(row["source_type"], "apple_tree")
		self.assertEqual(row["source_kind"], "dingtalk_export")
		self.assertEqual(row["source_sheet"], "钉钉导出数据")
		self.assertEqual(row["source_row"], 2)
		self.assertEqual(row["source_id"], "DATA-001")
		self.assertEqual(row["approval_no"], "APPROVAL-001")
		self.assertEqual(row["original_data"]["红苹果"], "15")

	def test_dingtalk_department_identifier_does_not_create_a_false_department_mismatch(self):
		row = normalize_apple_tree_rows(
			[source_row(**{"受奖/惩人部门": "连续课 - 11"})],
			rules=self.confirmed_rules(),
			employees=[{"employee_code": "E-001", "employee_name": "张三", "department": "连续课", "employment_status": "在职"}],
			source_file="苹果树.xlsx",
		)[0]

		self.assertNotIn("EMPLOYEE_DEPARTMENT_MISMATCH", row["exception_codes"])
		self.assertEqual(row["部门"], "连续课")
		self.assertEqual(row["processed_value"]["部门"], "连续课")

	def test_structure_preflight_reports_missing_columns_without_consuming_rows(self):
		valid = preflight_apple_tree_rows([source_row()])
		invalid_row = source_row()
		invalid_row.pop("数据id")
		invalid_row.pop("审批状态")
		invalid = preflight_apple_tree_rows([invalid_row])

		self.assertTrue(valid["可加工"])
		self.assertEqual(valid["状态"], "通过")
		self.assertFalse(invalid["可加工"])
		self.assertEqual(invalid["状态"], "不通过")
		self.assertEqual(set(invalid["缺失字段"]), {"数据ID", "审批状态"})

	def test_hr_monthly_register_does_not_require_nonexistent_dingtalk_audit_columns(self):
		row = source_row(**{"source_kind": "monthly_summary", "绿苹果": 8, "奖/惩项目": "人资组/绿苹果/新员工/每天2颗"})
		for field in ("数据id", "审批编号", "审批结果", "审批状态"):
			row.pop(field, None)
		preflight = preflight_apple_tree_rows([row])
		processed = normalize_apple_tree_rows(
			[row],
			rules=self.confirmed_rules(),
			employees=EMPLOYEES,
			source_file="7月苹果树.xlsx",
			source_sheet="苹果树合计",
			start_row=4,
		)

		self.assertTrue(preflight["可加工"])
		self.assertEqual(preflight["来源口径"], "人资月度汇总表")
		self.assertEqual(processed[0]["review_status"], "无需审核")
		self.assertEqual(processed[0]["有效苹果数"], 8)
		self.assertTrue(processed[0]["source_id"].startswith("monthly-summary:苹果树合计:"))
		self.assertNotIn("AMOUNT_TEXT_CONFLICT", processed[0]["exception_codes"])

	def test_historical_event_before_relieving_date_is_not_a_former_employee_exception(self):
		row = source_row(**{"source_kind": "monthly_summary"})
		for field in ("数据id", "审批编号", "审批结果", "审批状态"):
			row.pop(field, None)
		processed = normalize_apple_tree_rows(
			[row],
			rules=self.confirmed_rules(),
			employees=[{
				"employee_code": "E-001", "employee_name": "张三", "department": "连续课",
				"employment_status": "Left", "date_of_joining": "2026-01-01", "relieving_date": "2026-08-14",
			}],
			source_file="7月苹果树.xlsx",
		)[0]
		self.assertNotIn("FORMER_EMPLOYEE_REQUIRES_CONFIRMATION", processed["exception_codes"])

	def test_page_does_not_classify_apple_tree_as_special_hours(self):
		source = ATTENDANCE_PAGE.read_text(encoding="utf-8")

		self.assertIn("独立苹果树奖惩来源；不包含特殊工时。", source)
		self.assertNotIn("独立奖惩与特殊工时来源。", source)

	def test_full_apple_tree_result_has_browser_and_download_contract(self):
		"""Apple-tree rows must stay columnar after persistence, not become JSON blobs."""
		center_source = PROCESSING_CENTER.read_text(encoding="utf-8")
		page_source = ATTENDANCE_PAGE.read_text(encoding="utf-8")

		self.assertIn('limit_page_length=5000', center_source)
		self.assertIn('_employee_directory(batch.company)', center_source)
		self.assertIn('row.get("original_data") or row.get("original_value")', center_source)
		self.assertIn('"奖惩日期"', center_source)
		self.assertIn('"有效苹果数"', center_source)
		self.assertIn('(\"奖惩日期\", \"奖/惩日期\")', center_source)
		self.assertIn('(\"部门\", \"受奖/惩人部门\")', center_source)
		self.assertIn('(\"项目\", \"奖/惩项目\")', center_source)
		self.assertIn('def _hydrate_apple_tree_result_rows', center_source)
		self.assertIn('"是否计入下游", "来源追溯"', center_source)
		self.assertIn('processed_value.update(confirmed if review_status == "已通过" else proposed)', center_source)
		self.assertIn('def bulk_update_processing_records(', center_source)
		self.assertIn('show_bulk_processing_dialog', page_source)
		self.assertIn('bulk_update_processing_records', page_source)
		self.assertIn('批量确认当前数据（通过）', page_source)
		self.assertIn('"review_status": "待审核"', center_source)
		self.assertNotIn('bulk_resolve_apple_tree_employees', center_source)
		self.assertNotIn('批量重新匹配工号', page_source)
		self.assertIn('isAppleTree', page_source)
		self.assertIn('"奖惩日期"', page_source)
		self.assertIn('"奖/惩日期"', page_source)
		self.assertIn('download_processing_file', page_source)

	def test_apple_tree_export_is_only_the_printable_signoff_list(self):
		if load_workbook is None:
			self.skipTest("openpyxl is unavailable")
		center, file_manager, frappe_modules = processing_center_module()
		saved = {}
		file_manager.save_file = lambda file_name, content, *_args, **_kwargs: (
			saved.update({"file_name": file_name, "content": content})
			or SimpleNamespace(file_name=file_name, file_url="/private/files/test.xlsx")
		)
		row = {
			"employee_code": "E-001",
			"employee_name": "张三",
			"department": "连续课",
			"processed_value": {"工号": "E-001", "苹果类型": "绿苹果", "有效苹果数": 2},
			"proposed_value": {"工号": "E-001", "姓名": "张三", "部门": "连续课", "苹果类型": "绿苹果", "有效苹果数": 2},
			"confirmed_value": None,
			"original_value": source_row(),
			"exception_labels": ["花名册未找到员工"],
			"exception_message": "未匹配到员工工号。",
			"review_status": "待审核",
			"eligible_for_downstream": False,
			"source_file": "苹果树.xlsx",
			"source_sheet": "钉钉导出数据",
			"source_row": 2,
			"source_id": "DATA-001",
			"approval_no": "APPROVAL-001",
		}
		row["processed_value"] = center._apple_tree_result_values(row)
		batch = SimpleNamespace(source_type="apple_tree", attendance_month="2026-06")
		with patch.dict(sys.modules, frappe_modules), patch.object(center, "_result_rows", return_value=[row]):
			result = center._export_processed_result(batch)

		book = load_workbook(BytesIO(saved["content"]), read_only=True, data_only=True)
		sheet = book["加工结果"]
		export_rows = sheet.iter_rows(values_only=False)
		title = [cell.value for cell in next(export_rows)]
		headers = [cell.value for cell in next(export_rows)]
		values = [cell.value for cell in next(export_rows)]
		expected_headers = [
			"序号", "创建时间", "奖/惩日期", "受奖/惩人部门", "受奖/惩人", "绿苹果", "红苹果", "奖/惩项目", "备注", "创建人", "签名", "备注",
		]
		self.assertEqual(title[0], "6月苹果树")
		self.assertEqual(headers, expected_headers)
		self.assertEqual(values[0], 1)
		self.assertIsInstance(values[1], datetime)
		self.assertIsInstance(values[2], datetime)
		self.assertEqual(values[3:5], ["连续课", "张三"])
		self.assertEqual(values[5:10], [2, 15, "连续课/绿苹果/保养/作业员完成保养。2颗", "保养A线", "主管甲"])
		self.assertEqual(values[10:12], [None, None])
		self.assertEqual(sheet.max_column, 12)
		self.assertEqual(result["file_url"], "/private/files/test.xlsx")

	def test_missed_punch_export_is_only_the_plain_signoff_list(self):
		if load_workbook is None:
			self.skipTest("openpyxl is unavailable")
		center, file_manager, frappe_modules = processing_center_module()
		saved = {}
		file_manager.save_file = lambda file_name, content, *_args, **_kwargs: (
			saved.update({"file_name": file_name, "content": content})
			or SimpleNamespace(file_name=file_name, file_url="/private/files/test.xlsx")
		)
		row = {
			"employee_code": "E-001",
			"employee_name": "张三",
			"department": "工程课",
			"processed_value": {
				"employee_code": "E-001", "employee_name": "张三", "department": "工程课",
				"created_at": "2026-06-09 09:00:00", "punch_time": "2026-06-10 08:00:00",
				"punch_type": "忘刷卡补卡", "reason": "忘打卡", "included": True,
			},
			"proposed_value": {},
			"confirmed_value": None,
			"original_value": {"序号": 7, "创建人": "张三", "创建人部门": "工程课"},
			"eligible_for_downstream": True,
			"source_file": "忘打卡.xlsx",
			"source_sheet": "钉钉导出数据",
			"source_row": 2,
			"source_id": "DATA-001",
			"approval_no": "APPROVAL-001",
		}
		batch = SimpleNamespace(source_type="missing_card", attendance_month="2026-06")
		with patch.dict(sys.modules, frappe_modules), patch.object(center, "_result_rows", return_value=[row]):
			center._export_processed_result(batch)

		book = load_workbook(BytesIO(saved["content"]), read_only=True, data_only=True)
		sheet = book["加工结果"]
		values = list(sheet.iter_rows(values_only=True))
		self.assertEqual(values[0], ("序号", "部门", "创建时间", "补卡时间", "补卡类型", "补卡理由", "创建人", "签名", "备注"))
		self.assertEqual(values[1][0:2], (7, "工程课"))
		self.assertIsInstance(values[1][2], datetime)
		self.assertIsInstance(values[1][3], datetime)
		self.assertEqual(values[1][4:7], ("忘刷卡补卡", "忘打卡", "张三"))
		self.assertEqual(values[1][7:9], (None, None))
		self.assertEqual(sheet.max_column, 9)

	def test_monthly_signed_file_uses_the_full_hr_confirmation_layout(self):
		if load_workbook is None:
			self.skipTest("openpyxl is unavailable")
		center, file_manager, frappe_modules = processing_center_module()
		saved = {}
		file_manager.save_file = lambda file_name, content, *_args, **_kwargs: (
			saved.update({"file_name": file_name, "content": content})
			or SimpleNamespace(file_name=file_name, file_url="/private/files/test.xlsx")
		)
		with patch.dict(sys.modules, frappe_modules):
			result = center._save_monthly_signed_confirmation_file("2026-06", [{
				"department": "工程课 - 11", "employee_code": "103", "employee_name": "张三", "standard_hours": 168,
				"actual_attendance_hours": 164.5, "workday_overtime_hours": 35, "restday_overtime_hours": 5,
				"holiday_overtime_hours": 0, "special_workday_hours": 10.5, "special_restday_hours": 1,
				"large_night_shifts": 1, "small_night_shifts": 2,
				"absence_hours": 0, "green_apples": 3, "red_apples": 10, "housing_allowance": 200,
				"full_attendance_award": 150,
			}])

		book = load_workbook(BytesIO(saved["content"]), read_only=False, data_only=False)
		sheet = book["员工签字版"]
		self.assertEqual(sheet["D1"].value, "6月工时奖惩确认表")
		self.assertIn("AQ2:AV2", {str(item) for item in sheet.merged_cells.ranges})
		# This is the paper-form order.  It specifically prevents the historic
		# regression where the employee code was written under the department
		# heading in the signed confirmation workbook.
		self.assertEqual(sheet["B5"].value, 1)
		self.assertEqual(sheet["C5"].value, "工程课")
		self.assertEqual(sheet["D5"].value, "103")
		self.assertEqual(sheet["E5"].value, "张三")
		self.assertEqual(sheet["G5"].value, 168)
		self.assertEqual(sheet["H5"].value, 164.5)
		self.assertEqual(sheet["J5"].value, 10.5)
		self.assertEqual(sheet["K5"].value, 35)
		self.assertEqual(sheet["L5"].value, 1)
		self.assertEqual(sheet["M5"].value, 5)
		self.assertEqual(sheet["AI5"].value, "=J5+K5")
		self.assertEqual(sheet["AR5"].value, "=G5-AQ5")
		self.assertEqual(sheet.max_column, 60)
		self.assertEqual(sheet["B4"].fill.fgColor.rgb, "00FFD966")
		self.assertEqual(sheet["Z3"].fill.fgColor.rgb, "00F8CBAD")
		self.assertEqual(sheet["AH3"].fill.fgColor.rgb, "00D9E2F3")
		self.assertEqual(sheet["L3"].fill.fgColor.rgb, "00FFFF00")
		self.assertEqual(result["file_url"], "/private/files/test.xlsx")

	def test_monthly_signed_file_reads_custom_excel_headers_from_hr_settings(self):
		if load_workbook is None:
			self.skipTest("openpyxl is unavailable")
		center, file_manager, frappe_modules = processing_center_module()
		saved = {}
		file_manager.save_file = lambda file_name, content, *_args, **_kwargs: (
			saved.update({"content": content})
			or SimpleNamespace(file_name=file_name, file_url="/private/files/test.xlsx")
		)
		configured_rows = [
			SimpleNamespace(
				enabled=1,
				field_key="special_workday_hours",
				excel_main_header="平日加班",
				excel_sub_header="平日特",
				field_type="来源字段",
				comparison_policy="必须一致",
				numeric_tolerance=0,
			),
			SimpleNamespace(
				enabled=1,
				field_key="workday_overtime_hours",
				excel_main_header="平日加班",
				excel_sub_header="工作日加班",
				field_type="来源字段",
				comparison_policy="必须一致",
				numeric_tolerance=0,
			),
		]
		settings = SimpleNamespace(get=lambda fieldname: configured_rows if fieldname == "attendance_final_excel_fields" else [])
		center.frappe.get_single = lambda _doctype: settings
		with patch.dict(sys.modules, frappe_modules):
			center._save_monthly_signed_confirmation_file("2026-06", [{
				"department": "工程课", "employee_code": "103", "employee_name": "张三", "standard_hours": 168,
				"actual_attendance_hours": 168, "special_workday_hours": 4, "workday_overtime_hours": 8,
			}])

		book = load_workbook(BytesIO(saved["content"]), read_only=False, data_only=False)
		sheet = book["员工签字版"]
		self.assertEqual(sheet["J2"].value, "平日加班")
		self.assertEqual(sheet["J3"].value, "平日特")
		self.assertEqual(sheet["K3"].value, "工作日加班")
		self.assertIn("J2:K2", {str(item) for item in sheet.merged_cells.ranges})
		self.assertEqual(sheet["AU5"].value, "=AQ5+AN5+AO5")

	def test_special_hours_are_split_by_date_before_final_calculation(self):
		center, _file_manager, _frappe_modules = processing_center_module()
		self.assertEqual(
			center._special_hours_breakdown(
				[{"day": 1, "hours": 10.5}, {"day": 6, "hours": 1}], "2026-06"
			),
			{"special_workday_hours": 10.5, "special_restday_hours": 1.0, "special_holiday_hours": 0.0},
		)
		calculation = center._final_calculation({
			"standard_hours": 168, "actual_attendance_hours": 168,
			"special_workday_hours": 10.5, "workday_overtime_hours": 12.5,
			"special_restday_hours": 1, "restday_overtime_hours": 22,
		})
		self.assertEqual(calculation["settlement_15"], 23)
		self.assertEqual(calculation["settlement_20"], 23)
		self.assertEqual(calculation["adjusted_one_absence"], 0)

	def test_special_hours_use_company_holiday_list_before_weekend_split(self):
		center, _file_manager, _frappe_modules = processing_center_module()
		center.frappe.db.get_value = lambda *_args, **_kwargs: "2026年节假日"
		center.frappe.get_all = lambda *_args, **_kwargs: ["2026-06-01"]
		self.assertEqual(
			center._special_hours_breakdown(
				[{"day": 1, "hours": 8}, {"day": 6, "hours": 4}, {"day": 8, "hours": 2}],
				"2026-06",
				"永新公司",
			),
			{"special_workday_hours": 2.0, "special_restday_hours": 4.0, "special_holiday_hours": 8.0},
		)

	def test_finance_preview_merges_rate_matched_special_hours(self):
		center, _file_manager, _frappe_modules = processing_center_module()
		rows = center._finance_final_preview_rows([{
			"special_workday_hours": 10.5, "workday_overtime_hours": 12.5,
			"special_restday_hours": 1, "restday_overtime_hours": 22,
			"special_holiday_hours": 3, "holiday_overtime_hours": 5,
		}])
		self.assertEqual(rows[0]["workday_overtime_hours"], 23)
		self.assertEqual(rows[0]["restday_overtime_hours"], 23)
		self.assertEqual(rows[0]["holiday_overtime_hours"], 8)
		self.assertNotIn(("special_hours", "特殊工时"), center.FINAL_FINANCE_COLUMNS)
		self.assertNotIn(("special_hours", "特殊工时"), center.FINAL_SIGNED_COLUMNS)
		self.assertIn(("special_holiday_hours", "节假日特殊工时"), center.FINAL_SIGNED_COLUMNS)

	def test_daily_attendance_date_normalizes_dingtalk_display_dates(self):
		center, _file_manager, _frappe_modules = processing_center_module()
		self.assertEqual(center._daily_attendance_date("26-06-01 星期一"), "2026-06-01")
		self.assertEqual(center._daily_attendance_date("2026/06/30"), "2026-06-30")
		self.assertEqual(center._daily_attendance_date("未填写"), "")

	def test_bulk_import_classifies_the_six_standard_file_names_before_writing_batches(self):
		center, _file_manager, _frappe_modules = processing_center_module()
		self.assertEqual(center._detect_bulk_source_type("", "1.考勤初稿.xlsx"), "attendance_draft")
		self.assertEqual(center._detect_bulk_source_type("", "2.苹果树合计.xlsx"), "apple_tree")
		self.assertEqual(center._detect_bulk_source_type("", "3.忘打卡合计.xlsx"), "missing_card")
		self.assertEqual(center._detect_bulk_source_type("", "4.住房补贴（月）.xlsx"), "housing_allowance")
		self.assertEqual(center._detect_bulk_source_type("", "5.全勤奖（月）.xlsx"), "full_attendance")
		self.assertEqual(center._detect_bulk_source_type("", "7.特殊工时（月）.xlsx"), "special_hours")

	def test_bulk_import_identifies_real_workbooks_by_structure_not_filename(self):
		"""A renamed source must still be sent to its correct monthly slot."""
		if load_workbook is None:
			self.skipTest("openpyxl is unavailable")
		fixtures = (
			("attendance_draft", "1.考勤初稿.xlsx"),
			("apple_tree", "2.苹果树合计.xlsx"),
			("missing_card", "3.忘打卡合计.xlsx"),
			("housing_allowance", "4.住房补贴（月）.xlsx"),
			("full_attendance", "5.全勤奖（月）.xlsx"),
			("special_hours", "7.特殊工时（月）.xlsx"),
		)
		missing = [name for _source_type, name in fixtures if not (ATTENDANCE_SOURCE_DIR / name).exists()]
		if missing:
			self.skipTest(f"real source fixtures are unavailable: {', '.join(missing)}")
		center, _file_manager, frappe_modules = processing_center_module()
		with patch.dict(sys.modules, frappe_modules):
			for expected, filename in fixtures:
				workbook = load_workbook(ATTENDANCE_SOURCE_DIR / filename, read_only=True, data_only=True)
				with patch.object(center, "_load_workbook", return_value=workbook):
					self.assertEqual(center._detect_bulk_source_type("/private/files/upload.xlsx", "upload.xlsx"), expected)

	def test_wang_chuanrui_real_special_hours_reconcile_to_23_hours_at_15_rate(self):
		if load_workbook is None or not SPECIAL_HOURS_WORKBOOK.exists():
			self.skipTest("Special-hours sample workbook is unavailable")
		center, _file_manager, _frappe_modules = processing_center_module()
		sheet = load_workbook(SPECIAL_HOURS_WORKBOOK, data_only=True)["特殊工时（人为登记）"]
		day_row = [sheet.cell(3, column).value for column in range(1, sheet.max_column + 1)]
		wang_row = next(row for row in range(4, sheet.max_row + 1) if str(sheet.cell(row, 3).value or "").strip() == "王传瑞")
		entries = [
			{"day": day, "hours": sheet.cell(wang_row, column).value}
			for column, day in enumerate(day_row, start=1)
			if isinstance(day, int) and sheet.cell(wang_row, column).value not in (None, "")
		]
		breakdown = center._special_hours_breakdown(entries, "2026-06")
		self.assertEqual(breakdown, {"special_workday_hours": 10.5, "special_restday_hours": 1.0, "special_holiday_hours": 0.0})
		self.assertEqual(
			center._final_calculation({"special_workday_hours": breakdown["special_workday_hours"], "workday_overtime_hours": 12.5})["settlement_15"],
			23,
		)

	def test_monthly_amount_parser_ignores_new_hire_and_leaver_reference_sections(self):
		"""Only the numbered master lists become payroll records.

		The real housing and full-attendance exports append two unnumbered
		reference tables.  Their columns shift left and otherwise make names look
		like employee codes and amounts look blank.
		"""
		if load_workbook is None:
			self.skipTest("openpyxl is unavailable")
		fixtures = (
			("housing_allowance", HOUSING_ALLOWANCE_WORKBOOK),
			("full_attendance", FULL_ATTENDANCE_WORKBOOK),
		)
		missing = [str(path) for _source_type, path in fixtures if not path.exists()]
		if missing:
			self.skipTest(f"monthly amount fixtures are unavailable: {', '.join(missing)}")
		center, _file_manager, _frappe_modules = processing_center_module()
		for source_type, path in fixtures:
			with self.subTest(source_type=source_type):
				workbook = load_workbook(path, data_only=True, read_only=True)
				sheet = workbook[workbook.sheetnames[0]]
				batch = SimpleNamespace(source_type=source_type, attendance_month="2026-06", source_file="/private/files/monthly.xlsx")
				rows = center._monthly_amount_rows(sheet, batch, center.MONTHLY_SUPPORT_SOURCE_CONFIG[source_type])
				self.assertGreater(len(rows), 150)
				self.assertEqual(len(rows), len({row["employee_code"] for row in rows}))
				self.assertTrue(all(str(row["employee_code"]).strip().isdigit() for row in rows))
				self.assertTrue(all(isinstance(row[center.MONTHLY_SUPPORT_SOURCE_CONFIG[source_type]["value_field"]], (int, float)) for row in rows))

	def test_special_hours_does_not_block_on_a_historical_department_label(self):
		if load_workbook is None or not SPECIAL_HOURS_WORKBOOK.exists():
			self.skipTest("Special-hours sample workbook is unavailable")
		center, _file_manager, _frappe_modules = processing_center_module()
		workbook = load_workbook(SPECIAL_HOURS_WORKBOOK, data_only=True, read_only=True)
		batch = SimpleNamespace(source_type="special_hours", attendance_month="2026-06", source_file="/private/files/special.xlsx", company="永新")
		with patch.object(center, "_load_workbook", return_value=workbook):
			raw_rows, _sheets = center._read_monthly_support_rows(batch)
		self.assertTrue(all(row["department"] for row in raw_rows))
		roster = [
			{"employee_code": row["employee_code"], "employee_name": row["employee_name"], "department": "当前部门"}
			for row in raw_rows
		]
		with patch.object(center, "_load_workbook", return_value=load_workbook(SPECIAL_HOURS_WORKBOOK, data_only=True, read_only=True)), patch.object(center, "_employee_directory", return_value=roster):
			result = center._process_monthly_support_rows(batch)
		self.assertFalse(any("EMPLOYEE_DEPARTMENT_MISMATCH" in row["exception_codes"] for row in result["processed_rows"]))

	def test_monthly_finance_file_uses_the_compact_confirmation_layout(self):
		if load_workbook is None:
			self.skipTest("openpyxl is unavailable")
		center, file_manager, frappe_modules = processing_center_module()
		saved = {}
		file_manager.save_file = lambda file_name, content, *_args, **_kwargs: (
			saved.update({"file_name": file_name, "content": content})
			or SimpleNamespace(file_name=file_name, file_url="/private/files/test.xlsx")
		)
		with patch.dict(sys.modules, frappe_modules):
			center._save_monthly_finance_confirmation_file("2026-06", [{
				"department": "工程课", "employee_name": "张三", "standard_hours": 168, "actual_attendance_hours": 168,
				"special_workday_hours": 10.5, "workday_overtime_hours": 12.5,
				"special_restday_hours": 1, "restday_overtime_hours": 22,
			}])
		book = load_workbook(BytesIO(saved["content"]), read_only=False, data_only=False)
		sheet = book["财务版"]
		self.assertEqual(sheet["A1"].value, "6月工时奖惩确认表")
		self.assertIn("F2:I2", {str(item) for item in sheet.merged_cells.ranges})
		self.assertEqual(sheet["A4"].fill.fgColor.rgb, "00FFD966")
		self.assertEqual(sheet["F5"].value, 168)
		self.assertEqual(sheet["G5"].value, 23)
		self.assertEqual(sheet["H5"].value, 23)

	def test_result_summary_keeps_green_and_red_apple_totals_separate(self):
		center, _file_manager, _frappe_modules = processing_center_module()
		rows = [
			{
				"eligible_for_downstream": True,
				"processed_value": {"苹果类型": "绿苹果", "有效苹果数": 5},
				"proposed_value": {},
				"confirmed_value": None,
			},
			{
				"eligible_for_downstream": True,
				"processed_value": {"苹果类型": "红苹果", "有效苹果数": 2},
				"proposed_value": {},
				"confirmed_value": None,
			},
			{
				"eligible_for_downstream": False,
				"processed_value": {"苹果类型": "红苹果", "有效苹果数": 99},
				"proposed_value": {},
				"confirmed_value": None,
			},
		]

		self.assertEqual(center._apple_tree_summary(rows), {"green_apples": 5, "red_apples": 2})

	def test_download_endpoint_generates_current_processed_result(self):
		"""A completed source must take the export path, not the error branch."""
		center, _file_manager, _frappe_modules = processing_center_module()
		center.frappe.get_roles = lambda _user: ["HR Manager"]
		center.frappe.db.count = lambda *_args, **_kwargs: 1
		batch = SimpleNamespace(name="BATCH-001")
		generated = {"file_url": "/private/files/apple-tree-current.xlsx", "file_name": "苹果树加工结果.xlsx"}
		with patch.object(center, "_latest_batch", return_value=batch), patch.object(center, "_export_processed_result", return_value=generated) as export_result, patch.object(center, "_save_batch_notes") as save_notes, patch.object(center, "now_datetime", return_value=SimpleNamespace(isoformat=lambda: "2026-06-30T12:00:00")):
			result = center.export_processing_result("测试公司", "2026-06", "apple_tree")

		export_result.assert_called_once_with(batch)
		save_notes.assert_called_once()
		self.assertEqual(result["processed_result"], generated)

	def test_red_project_accepts_blank_inactive_value(self):
		row = source_row(
			**{
				"数据id": "DATA-002",
				"审批编号": "APPROVAL-002",
				"受奖/惩人": "李四",
				"受奖/惩人部门": "品管课",
				"绿苹果": "",
				"红苹果": "3",
				"奖/惩项目": "品管课/红苹果/异常/操作不当。3颗",
			}
		)
		processed = normalize_apple_tree_rows(
			[row], rules=self.confirmed_rules(), employees=EMPLOYEES, source_file="sample.xlsx"
		)[0]

		self.assertEqual(processed["苹果类型"], "红苹果")
		self.assertEqual(processed["有效苹果数"], 3)
		self.assertEqual(processed["review_status"], "无需审核")
		self.assertTrue(processed["include_in_downstream"])

	def test_anomalies_are_retained_and_never_silently_dropped(self):
		rows = [
			source_row(**{"数据id": "DUP", "审批编号": "DUP-APP"}),
			source_row(
				**{
					"数据id": "DUP",
					"审批编号": "DUP-APP",
					"受奖/惩人": "李四",
					"受奖/惩人部门": "品管课",
					"奖/惩项目": "品管课/绿苹果/异常/发现异常。5颗",
				}
			),
			source_row(
				**{
					"数据id": "REJECTED",
					"审批编号": "REJECTED-APP",
					"审批结果": "审批未通过",
					"审批状态": "已结束",
				}
			),
			source_row(
				**{
					"数据id": "WRONG-MONTH",
					"审批编号": "WRONG-MONTH-APP",
					"奖/惩日期": "2026-05-31",
				}
			),
			source_row(
				**{
					"数据id": "NAME-MISMATCH",
					"审批编号": "NAME-MISMATCH-APP",
					"工号": "E-001",
					"受奖/惩人": "李四",
				}
			),
		]

		processed = normalize_apple_tree_rows(rows, rules=self.confirmed_rules(), employees=EMPLOYEES)

		# 审批未通过的苹果树记录不进入人工审批队列。
		self.assertEqual(len(processed), len(rows) - 1)
		self.assertIn("DUPLICATE_SOURCE_ID", processed[0]["exception_codes"])
		self.assertIn("DUPLICATE_APPROVAL_NO", processed[1]["exception_codes"])
		self.assertIn("AMOUNT_TEXT_CONFLICT", processed[1]["exception_codes"])
		for row in processed:
			self.assertEqual(row["review_status"], "待审核")
			self.assertFalse(row["include_in_downstream"])
		self.assertIn("MONTH_MISMATCH", processed[2]["exception_codes"])
		self.assertIn("EMPLOYEE_NAME_MISMATCH", processed[3]["exception_codes"])

	def test_name_department_disambiguation_and_employee_conflicts_enter_review(self):
		employees = [
			{"employee_code": "E-101", "employee_name": "同名员工", "department": "连续课", "employment_status": "在职"},
			{"employee_code": "E-102", "employee_name": "同名员工", "department": "品管课", "employment_status": "在职"},
			{"employee_code": "E-103", "employee_name": "离职员工", "department": "连续课", "employment_status": "已离职"},
		]
		rows = [
			source_row(**{"受奖/惩人": "同名员工", "受奖/惩人部门": "品管课"}),
			source_row(
				**{
					"数据id": "DATA-002",
					"审批编号": "APPROVAL-002",
					"工号": "E-101",
					"受奖/惩人": "错误姓名",
					"受奖/惩人部门": "错误部门",
				}
			),
			source_row(
				**{
					"数据id": "DATA-003",
					"审批编号": "APPROVAL-003",
					"受奖/惩人": "离职员工",
				}
			),
		]
		processed = normalize_apple_tree_rows(
			rows, rules=self.confirmed_rules(), employees=employees, source_file="sample.xlsx"
		)

		self.assertEqual(processed[0]["工号"], "E-102")
		self.assertEqual(processed[0]["review_status"], "无需审核")
		self.assertIn("EMPLOYEE_NAME_MISMATCH", processed[1]["exception_codes"])
		self.assertIn("EMPLOYEE_DEPARTMENT_MISMATCH", processed[1]["exception_codes"])
		self.assertIn("FORMER_EMPLOYEE_REQUIRES_CONFIRMATION", processed[2]["exception_codes"])

	def test_in_progress_apple_approval_remains_reviewable_while_closed_rows_disappear(self):
		rows = [
			source_row(**{"数据id": "PENDING", "审批编号": "PENDING-APP", "审批结果": "审批中", "审批状态": "审批中"}),
			source_row(**{"数据id": "TERMINATED", "审批编号": "TERMINATED-APP", "审批状态": "终止"}),
			source_row(**{"数据id": "BUSINESS", "审批编号": "BUSINESS-APP", "审批类型": "因公打卡"}),
		]
		processed = normalize_apple_tree_rows(rows, rules=self.confirmed_rules(), employees=EMPLOYEES)

		self.assertEqual(len(processed), 1)
		self.assertEqual(processed[0]["数据ID"], "PENDING")
		self.assertEqual(processed[0]["review_status"], "待审核")

	def test_missing_source_offline_rows_and_same_time_approvals_enter_review(self):
		rows = [
			source_row(**{"数据id": "DATA-101", "审批编号": "APPROVAL-101"}),
			source_row(**{"数据id": "DATA-102", "审批编号": "APPROVAL-102"}),
			source_row(
				**{
					"数据id": "DATA-103",
					"审批编号": "APPROVAL-103",
					"source_kind": "offline",
					"创建时间": "2026-06-17 09:00:00",
				}
			),
		]
		processed = normalize_apple_tree_rows(
			rows, rules=self.confirmed_rules(), employees=EMPLOYEES, source_file="", source_sheet=""
		)

		self.assertIn("MULTIPLE_APPROVALS_SAME_TIME", processed[0]["exception_codes"])
		self.assertIn("MULTIPLE_APPROVALS_SAME_TIME", processed[1]["exception_codes"])
		self.assertIn("SOURCE_FILE_MISSING", processed[0]["exception_codes"])
		self.assertIn("SOURCE_SHEET_MISSING", processed[0]["exception_codes"])
		self.assertIn("OFFLINE_ENTRY_REQUIRES_CONFIRMATION", processed[2]["exception_codes"])
		self.assertTrue(all(row["review_status"] == "待审核" for row in processed))

	def test_target_month_is_required_but_award_date_is_the_confirmed_default_basis(self):
		processed = normalize_apple_tree_rows(
			[source_row()], rules=AppleTreeRules(), employees=EMPLOYEES
		)[0]

		self.assertEqual(AppleTreeRules().month_basis, "award_date")
		self.assertEqual(processed["review_status"], "待审核")
		self.assertIn("TARGET_MONTH_REQUIRED", processed["exception_codes"])

	def test_review_keeps_original_value_and_requires_audited_confirmation(self):
		processed = normalize_apple_tree_rows(
			[source_row(**{"奖/惩项目": "连续课/绿苹果/保养/作业员完成保养。3颗"})],
			rules=self.confirmed_rules(),
			employees=EMPLOYEES,
			source_file="sample.xlsx",
		)
		reviewed = apply_reviews(
			processed,
			[
				{
					"source_id": "DATA-001",
					"review_status": "已通过",
					"proposed_value": {"有效苹果数": 3},
					"confirmed_value": {"有效苹果数": 3},
					"reviewer": "HR-001",
					"reviewed_on": "2026-07-08 10:00:00",
					"review_note": "员工签字确认后调整",
				}
			],
		)[0]

		self.assertEqual(reviewed["original_value"]["有效苹果数"], 2)
		self.assertEqual(reviewed["有效苹果数"], 3)
		self.assertEqual(reviewed["processed_value"]["有效苹果数"], 3)
		self.assertEqual(reviewed["original_data"]["绿苹果"], "2")
		self.assertEqual(reviewed["proposed_value"]["有效苹果数"], 3)
		self.assertEqual(reviewed["confirmed_value"]["有效苹果数"], 3)
		self.assertEqual(reviewed["review_status"], "已通过")
		self.assertEqual(reviewed["reviewer"], "HR-001")
		self.assertTrue(reviewed["include_in_downstream"])
		self.assertEqual(reviewed["review_history"][0]["original_value"]["有效苹果数"], 2)

	def test_row_level_missing_values_are_reviewed_instead_of_dropped(self):
		row = source_row(
			**{
				"奖/惩日期": "",
				"创建时间": "",
				"创建人": "",
				"审批结果": "",
				"审批状态": "",
			}
		)
		processed = normalize_apple_tree_rows(
			[row], rules=self.confirmed_rules(), employees=EMPLOYEES, source_file="sample.xlsx"
		)[0]

		self.assertEqual(processed["review_status"], "待审核")
		self.assertFalse(processed["include_in_downstream"])
		self.assertTrue(
			{
				"MISSING_AWARD_DATE",
				"MISSING_CREATED_AT",
				"MISSING_CREATOR",
				"MISSING_APPROVAL_RESULT",
				"MISSING_APPROVAL_STATUS",
			}.issubset(processed["exception_codes"])
		)

	def test_processor_returns_one_dataset_and_summary_is_only_an_optional_statistic(self):
		rows = [
			source_row(),
			source_row(
				**{
					"数据id": "DATA-002",
					"审批编号": "APPROVAL-002",
					"创建时间": "2026-06-17 08:30:00",
					"绿苹果": "3",
					"奖/惩项目": "连续课/绿苹果/支援/额外支援。3颗",
				}
			),
			source_row(
				**{
					"数据id": "DATA-003",
					"审批编号": "APPROVAL-003",
					"创建时间": "2026-06-18 08:30:00",
					"审批状态": "审批中",
				}
			),
		]
		processed = process_apple_tree_rows(
			rows, rules=self.confirmed_rules(), employees=EMPLOYEES, source_file="苹果树.xlsx"
		)

		summary = build_employee_summary(processed)

		self.assertIsInstance(processed, list)
		self.assertEqual(len(processed), 3)
		self.assertEqual(len(summary), 1)
		self.assertEqual(summary[0]["工号"], "E-001")
		self.assertEqual(summary[0]["绿苹果合计"], 5)
		self.assertEqual(summary[0]["红苹果合计"], 0)
		self.assertEqual(summary[0]["计入记录数"], 2)
		self.assertEqual(processed[0]["source_file"], "苹果树.xlsx")
		self.assertIn("APPROVAL_NOT_FINISHED", processed[2]["exception_codes"])

	def test_real_workbook_contract_retains_all_reviewable_source_rows(self):
		if load_workbook is None or not REAL_WORKBOOK.exists():
			return
		workbook = load_workbook(REAL_WORKBOOK, read_only=True, data_only=True)
		sheet = workbook["钉钉导出数据"]
		values = sheet.iter_rows(values_only=True)
		headers = [str(value or "").strip() for value in next(values)]
		raw_rows = [dict(zip(headers, row, strict=True)) for row in values]

		processed = normalize_apple_tree_rows(
			raw_rows,
			rules=self.confirmed_rules(),
			source_file=REAL_WORKBOOK.name,
			source_sheet=sheet.title,
		)

		reviewable_rows = [row for row in raw_rows if not apple_tree.is_auto_excluded_apple_tree_row(row)]
		self.assertEqual(len(raw_rows), 761)
		self.assertEqual(len(reviewable_rows), 750)
		self.assertEqual(len(processed), len(reviewable_rows))
		self.assertEqual(sum(row["苹果类型"] == "绿苹果" for row in processed), 718)
		self.assertEqual(sum(row["苹果类型"] == "红苹果" for row in processed), 32)
		self.assertEqual(sum("MONTH_MISMATCH" in row["exception_codes"] for row in processed), 104)
		self.assertEqual(
			sum("INACTIVE_APPLE_VALUE_CONFLICT" in row["exception_codes"] for row in processed), 1
		)
		self.assertEqual(processed[0]["source_row"], 2)
		self.assertEqual(processed[-1]["source_row"], 762)


if __name__ == "__main__":
	unittest.main()
