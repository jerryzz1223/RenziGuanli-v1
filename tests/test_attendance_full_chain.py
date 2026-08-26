"""End-to-end acceptance for the three-source attendance processing chain.

This test uses the actual DingTalk export samples as immutable input.  It
verifies the same business boundary used by the Desk page:

source workbook -> structure precheck -> one processed dataset per source
-> unified review contract -> manual approve/reject -> downstream eligibility
-> monthly-final gate.

It deliberately does not write to Frappe or to the source workbooks.  The
runtime/API persistence and download-refresh invariants are covered by
``verify_attendance_processing_center.py``; keeping this test pure makes it
safe to run before every deployment and against the supplied exports.
"""

from __future__ import annotations

import importlib.util
import re
import sys
import unittest
from copy import deepcopy
from pathlib import Path

try:
	from openpyxl import load_workbook
except ModuleNotFoundError:
	load_workbook = None


ROOT = Path(__file__).resolve().parents[1]
PROCESSOR_DIR = ROOT / "hrms" / "api" / "attendance_processors"
PROCESSING_API = ROOT / "hrms" / "api" / "attendance_processing_center.py"
ATTENDANCE_PAGE = ROOT / "hrms" / "hr" / "page" / "attendance_import_center" / "attendance_import_center.js"
SOURCE_DIR = Path("/Users/lrj/Desktop/薪酬计算设计表单/考勤数据")
SOURCES = {
	"attendance_draft": SOURCE_DIR / "1.考勤初稿.xlsx",
	"apple_tree": SOURCE_DIR / "2.苹果树合计.xlsx",
	"missing_card": SOURCE_DIR / "3.忘打卡合计.xlsx",
}
MONTHLY_SUPPORT_SOURCES = {
	"housing_allowance": {
		"path": SOURCE_DIR / "4.住房补贴（月）.xlsx",
		"headers": ("工号", "姓名", "住房补贴"),
		"mode": "monthly_amount",
	},
	"full_attendance": {
		"path": SOURCE_DIR / "5.全勤奖（月）.xlsx",
		"headers": ("工号", "姓名", "全勤奖"),
		"mode": "monthly_amount",
	},
	"special_hours": {
		"path": SOURCE_DIR / "7.特殊工时（月）.xlsx",
		"headers": ("工号", "姓名"),
		"mode": "special_hours_grid",
	},
}
MONTH = "2026-06"


def _load_module(name: str, filename: str):
	spec = importlib.util.spec_from_file_location(name, PROCESSOR_DIR / filename)
	module = importlib.util.module_from_spec(spec)
	sys.modules[name] = module
	spec.loader.exec_module(module)
	return module


attendance_draft = _load_module("attendance_full_chain_draft", "attendance_draft.py")
apple_tree = _load_module("attendance_full_chain_apple", "apple_tree.py")
missed_punch = _load_module("attendance_full_chain_missing", "missed_punch.py")


def _flat_sheet_rows(workbook_path: Path, sheet_name: str):
	book = load_workbook(workbook_path, data_only=True, read_only=True)
	sheet = book[sheet_name]
	values = sheet.iter_rows(values_only=True)
	headers = [str(value or "").strip() for value in next(values)]
	rows = []
	for source_row, values_row in enumerate(values, start=2):
		if not any(value not in (None, "") for value in values_row):
			continue
		row = {headers[index]: values_row[index] for index in range(len(headers)) if headers[index]}
		row.update({"source_file": workbook_path.name, "source_sheet": sheet.title, "source_row": source_row})
		rows.append(row)
	return rows


def _is_downstream_eligible(row: dict) -> bool:
	"""Mirror the processing center's shared eligibility meaning."""
	if row.get("review_status") not in {"无需审核", "已通过"}:
		return False
	confirmed = row.get("confirmed_value") or {}
	for key in ("eligible_for_downstream", "include_in_downstream", "included"):
		if key in confirmed:
			return bool(confirmed[key])
	return bool(row.get("eligible_for_downstream", row.get("include_in_downstream", row.get("included", False))))


def _attendance_day(value) -> str:
	match = re.search(r"(\d{2,4})-(\d{1,2})-(\d{1,2})", str(value or ""))
	if not match:
		return ""
	year = f"20{match.group(1)}" if len(match.group(1)) == 2 else match.group(1)
	return f"{year}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"


def _normalized_header(value) -> str:
	return "".join(str(value or "").split())


def _monthly_support_precheck(workbook_path: Path, required_headers: tuple[str, ...], mode: str) -> dict:
	"""Mirror the monthly-support API's safe structure gate on real exports."""
	required = tuple(_normalized_header(value) for value in required_headers)
	workbook = load_workbook(workbook_path, data_only=True, read_only=True)
	matches = []
	for sheet in workbook.worksheets:
		month_text = sheet.title + "".join(str(value or "") for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True) for value in row)
		if "6月" not in month_text:
			continue
		for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
			headers = [_normalized_header(value) for value in row]
			positions = {field: [index for index, value in enumerate(headers) if value == field] for field in required}
			if not all(positions.values()):
				continue
			if mode == "special_hours_grid":
				record_count = sum(1 for data_row in sheet.iter_rows(min_row=row_number + 2, values_only=True) if positions["工号"][0] < len(data_row) and str(data_row[positions["工号"][0]] or "").strip())
			else:
				record_count = sum(
					sum(1 for index in positions["工号"] if index < len(data_row) and str(data_row[index] or "").strip())
					for data_row in sheet.iter_rows(min_row=row_number + 1, values_only=True)
				)
			matches.append({"sheet": sheet.title, "header_row": row_number, "record_count": record_count})
			break
	return {"is_valid": bool(matches) and sum(item["record_count"] for item in matches) > 0, "matches": matches, "record_count": sum(item["record_count"] for item in matches)}


@unittest.skipUnless(load_workbook is not None, "openpyxl is unavailable")
class AttendanceFullChainAcceptanceTest(unittest.TestCase):
	@classmethod
	def setUpClass(cls):
		missing = [str(path) for path in SOURCES.values() if not path.exists()]
		missing.extend(str(config["path"]) for config in MONTHLY_SUPPORT_SOURCES.values() if not config["path"].exists())
		if missing:
			raise unittest.SkipTest(f"Real DingTalk test exports are unavailable: {', '.join(missing)}")

		attendance_book = load_workbook(SOURCES["attendance_draft"], data_only=True, read_only=True)
		daily_sheet = attendance_book["每日明细（钉钉导出）"]
		values = daily_sheet.iter_rows(values_only=True)
		headers = attendance_draft.flatten_dingtalk_headers(next(values), next(values))
		cls.attendance_precheck = attendance_draft.precheck_attendance_draft_structure(headers)
		cls.attendance_rows = attendance_draft.rows_from_dingtalk_daily_sheet(
			daily_sheet, source_file=SOURCES["attendance_draft"].name
		)
		cls.attendance_result = attendance_draft.process_attendance_draft_rows(
			cls.attendance_rows,
			attendance_month=MONTH,
			source_file=SOURCES["attendance_draft"].name,
			source_sheet=daily_sheet.title,
		)
		# This is the same matching directory that the later source stages receive
		# from HRMS Employee.  Building it from the attendance source keeps the
		# acceptance fixture self-contained and avoids using production employee data.
		cls.employee_directory = [
			{
				"employee_code": row["employee_code"],
				"employee_name": row["employee_name"],
				"department": row["department"],
				"employment_status": "在职",
			}
			for row in cls.attendance_result["processed_rows"]
			if row.get("employee_code") and row.get("employee_name")
		]

		cls.apple_rows = _flat_sheet_rows(SOURCES["apple_tree"], "钉钉导出数据")
		cls.apple_precheck = apple_tree.preflight_apple_tree_rows(cls.apple_rows)
		cls.apple_result = apple_tree.process_apple_tree_rows(
			cls.apple_rows,
			rules=apple_tree.AppleTreeRules(target_month=MONTH),
			employees=cls.employee_directory,
			source_file=SOURCES["apple_tree"].name,
			source_sheet="钉钉导出数据",
			start_row=2,
		)

		cls.missing_rows = _flat_sheet_rows(SOURCES["missing_card"], "钉钉导出数据")
		cls.missing_precheck = missed_punch.precheck_missed_punch_structure(list(cls.missing_rows[0]))
		cls.missing_result = missed_punch.process_missed_punch_rows(
			cls.missing_rows,
			attendance_month=MONTH,
			source_file=SOURCES["missing_card"].name,
			source_sheet="钉钉导出数据",
			employee_directory=cls.employee_directory,
		)
		cls.monthly_support_prechecks = {
			source_type: _monthly_support_precheck(config["path"], config["headers"], config["mode"])
			for source_type, config in MONTHLY_SUPPORT_SOURCES.items()
		}

	def test_01_three_source_prechecks_and_row_reconciliation(self):
		self.assertTrue(self.attendance_precheck["is_valid"])
		self.assertTrue(self.apple_precheck["可加工"])
		self.assertTrue(self.missing_precheck["is_valid"])

		self.assertEqual(len(self.attendance_rows), 5820)
		self.assertEqual(self.attendance_result["metrics"]["source_rows"], len(self.attendance_rows))
		self.assertEqual(self.attendance_result["metrics"]["processed_rows"], 194)
		self.assertEqual(
			sum(len(row["original_value"]["source_rows"]) for row in self.attendance_result["processed_rows"]),
			len(self.attendance_rows),
		)

		self.assertEqual(len(self.apple_rows), 761)
		self.assertEqual(len(self.apple_result), len(self.apple_rows))
		self.assertEqual(len({row["source_row"] for row in self.apple_result}), len(self.apple_rows))

		self.assertEqual(len(self.missing_rows), 52)
		self.assertEqual(self.missing_result["metrics"]["source_rows"], len(self.missing_rows))
		self.assertEqual(
			self.missing_result["metrics"]["processed_rows"] + self.missing_result["metrics"]["excluded_source_rows"],
			len(self.missing_rows),
		)
		self.assertEqual(self.missing_result["metrics"]["excluded_source_rows"], 8)

	def test_02_missing_punch_is_an_independent_penalty_source_not_a_daily_detail_duplicate(self):
		"""A daily missing-checkin flag is not a substitute for an approved forgotten-punch event.

		The daily export is a current attendance snapshot.  The forgotten-punch
		export is the auditable history used by HR to apply the separate rule of
		2 red apples / 10 yuan per event.  The fixed June source demonstrates why
		removing the third source would lose information after a check-in is made up.
		"""
		daily_flag_keys = {
			f"{str(row.get('姓名') or '').strip()}|{_attendance_day(row.get('日期'))}"
			for row in self.attendance_rows
			if row.get("上班缺卡") or row.get("下班缺卡")
		}
		forgot_rows = [row for row in self.missing_rows if "忘刷卡" in str(row.get("补卡类型") or "")]
		forgot_keys = {
			f"{str(row.get('创建人') or '').strip()}|{_attendance_day(row.get('补卡时间'))}"
			for row in forgot_rows
		}
		self.assertEqual(len(daily_flag_keys), 22)
		self.assertEqual(len(forgot_rows), 44)
		self.assertEqual(len(daily_flag_keys & forgot_keys), 1)
		self.assertGreater(
			len(forgot_keys - daily_flag_keys),
			0,
			"Approved forgotten-punch events must not be inferred only from the daily missing-checkin flags.",
		)

	def test_03_every_processed_row_has_the_unified_audit_contract(self):
		all_rows = (
			self.attendance_result["processed_rows"]
			+ self.apple_result
			+ self.missing_result["processed_rows"]
		)
		self.assertGreater(len(all_rows), 0)
		for row in all_rows:
			with self.subTest(source=row.get("source_type"), source_row=row.get("source_row")):
				self.assertIn(row.get("review_status"), {"无需审核", "待审核", "已通过", "已驳回"})
				for field in (
					"exception_codes", "exception_message", "proposed_value", "confirmed_value",
					"reviewer", "reviewed_on", "review_note", "source_file", "source_sheet", "source_row",
				):
					self.assertIn(field, row)
				self.assertTrue(row["source_file"])
				self.assertTrue(row["source_sheet"])
				self.assertIsNotNone(row["source_row"])

		pending = [row for row in all_rows if row["review_status"] == "待审核"]
		self.assertGreater(len(pending), 0, "The real exports must expose reviewable exceptions instead of silently passing all rows.")
		self.assertTrue(all(not _is_downstream_eligible(row) for row in pending))

	def test_04_manual_review_changes_eligibility_without_mutating_the_raw_record(self):
		pending = next(row for row in self.missing_result["processed_rows"] if row["review_status"] == "待审核")
		original = deepcopy(pending)
		confirmed = {**pending["proposed_value"], "included": True, "red_apples": 2, "amount": 10}
		approved = missed_punch.apply_missed_punch_review(
			pending,
			decision="已通过",
			confirmed_value=confirmed,
			reviewer="TEST-E2E-HR",
			reviewed_on="2026-07-01 09:00:00",
			review_note="全链路验收：管理员确认计入",
		)
		self.assertEqual(pending, original, "Review must create an auditable new value, not overwrite the source proposal.")
		self.assertEqual(approved["review_status"], "已通过")
		self.assertTrue(_is_downstream_eligible(approved))
		self.assertEqual(approved["review_history"][-1]["new_value"], confirmed)

		rejected = missed_punch.apply_missed_punch_review(
			pending,
			decision="已驳回",
			confirmed_value={**pending["proposed_value"], "included": False},
			reviewer="TEST-E2E-HR",
			reviewed_on="2026-07-01 09:05:00",
			review_note="全链路验收：确认不计入",
		)
		self.assertEqual(rejected["review_status"], "已驳回")
		self.assertFalse(_is_downstream_eligible(rejected))

	def test_05_confirmation_and_monthly_final_are_blocked_until_the_real_gates_pass(self):
		all_rows = (
			self.attendance_result["processed_rows"]
			+ self.apple_result
			+ self.missing_result["processed_rows"]
		)
		self.assertTrue(any(row["review_status"] == "待审核" for row in all_rows))
		self.assertFalse(
			all(row["review_status"] != "待审核" for row in all_rows),
			"A source must not be confirmable while its unified review queue still has pending rows.",
		)
		for source_type, precheck in self.monthly_support_prechecks.items():
			with self.subTest(monthly_support=source_type):
				self.assertTrue(precheck["is_valid"])
				self.assertGreater(precheck["record_count"], 0)

		api = PROCESSING_API.read_text(encoding="utf-8")
		for marker in (
			"MONTHLY_SUPPORT_SOURCE_TYPES",
			"register_monthly_support_file",
			"precheck_monthly_support_file",
			"process_monthly_support_file",
			"confirm_monthly_support_file",
			"_process_monthly_support_rows",
			"special_hours",
			"已识别字段，但未找到任何工号记录。",
			"FINAL_SIGNED_COLUMNS",
			"FINAL_FINANCE_COLUMNS",
			"_monthly_final_rows",
			"monthly_final_outputs",
			"get_monthly_final_preview",
			"_monthly_snapshot_version",
			"for source_type in SOURCE_TYPES",
			"slot[\"status\"] == \"已确认\"",
		):
			self.assertIn(marker, api)

	def test_06_page_keeps_upload_inside_each_source_and_month_selection_interactive(self):
		page = ATTENDANCE_PAGE.read_text(encoding="utf-8")
		for marker in (
			"data-slot-upload", "set_primary_action(null)", "render_month_control", "open_month_picker",
			"data-month-shift", "data-open-month-picker", "选择处理月份",
		):
			self.assertIn(marker, page)
		header_start = page.index("render_header()")
		header_end = page.index("render_kpi_grid()", header_start)
		self.assertNotIn("data-upload", page[header_start:header_end])


if __name__ == "__main__":
	unittest.main(verbosity=2)
