#!/usr/bin/env python3
"""Contract tests for the standalone missed-punch attendance processor."""

from __future__ import annotations

import importlib.util
import sys
from pathlib import Path


PROJECT_ROOT = Path(__file__).resolve().parents[1]
PROCESSOR_PATH = PROJECT_ROOT / "hrms" / "api" / "attendance_processors" / "missed_punch.py"


def _load_processor_module():
	spec = importlib.util.spec_from_file_location("missed_punch_processor_contract", PROCESSOR_PATH)
	module = importlib.util.module_from_spec(spec)
	sys.modules[spec.name] = module
	spec.loader.exec_module(module)
	return module


processor = _load_processor_module()
MissedPunchRules = processor.MissedPunchRules
apply_missed_punch_review = processor.apply_missed_punch_review
precheck_missed_punch_structure = processor.precheck_missed_punch_structure
process_missed_punch_rows = processor.process_missed_punch_rows
summarize_missed_punch_rows = processor.summarize_missed_punch_rows


SAMPLE_WORKBOOK = Path("/Users/lrj/Desktop/薪酬计算设计表单/考勤数据/3.忘打卡合计.xlsx")


def _employee(code, name, department, status="在职"):
	return {
		"employee_code": code,
		"employee_name": name,
		"department": department,
		"employment_status": status,
	}


def _row(
	approval_no,
	punch_time="2026-06-10 08:00",
	name="张三",
	department="工程课",
	punch_type="忘刷卡补卡",
	approval_result="审批通过",
	approval_status="已结束",
	data_id=None,
):
	return {
		"数据id": data_id or f"DATA-{approval_no}",
		"审批编号": approval_no,
		"创建时间": "2026-06-09 09:00",
		"补卡时间": punch_time,
		"创建人": name,
		"创建人部门": department,
		"补卡类型": punch_type,
		"补卡理由": "忘打卡",
		"审批结果": approval_result,
		"审批状态": approval_status,
	}


def test_structure_precheck_recognises_the_dingtalk_contract():
	headers = list(_row("APP-001"))
	result = precheck_missed_punch_structure(headers)

	assert result["is_valid"] is True
	assert result["missing_required_fields"] == []
	assert result["field_mapping"]["source_id"] == "数据id"
	assert result["field_mapping"]["employee_name"] == "创建人"
	assert result["field_mapping"]["punch_time"] == "补卡时间"

	invalid = precheck_missed_punch_structure(["创建人", "补卡理由"])
	assert invalid["is_valid"] is False
	assert "punch_time" in invalid["missing_required_fields"]
	assert "approval_no" in invalid["missing_required_fields"]


def test_default_rules_return_one_processed_dataset_and_keep_reviewable_rows():
	rows = [
		_row("APP-001"),
		_row("APP-002", punch_time="2026-06-11 08:00", punch_type="因公补卡"),
		_row("APP-003", punch_time="2026-05-31 20:00"),
		_row("APP-004", punch_time="2026-06-12 08:00", approval_result="--", approval_status="终止"),
		_row("APP-001", punch_time="2026-06-13 08:00", data_id="DATA-DUPLICATE"),
	]
	result = process_missed_punch_rows(
		rows,
		attendance_month="2026-06",
		source_file="3.忘打卡合计.xlsx",
		source_sheet="钉钉导出数据",
		employee_directory=[_employee("E-001", "张三", "工程课")],
	)

	assert result["status"] == "待处理异常"
	assert set(result) == {"status", "structure_precheck", "processed_rows", "metrics"}
	assert len(result["processed_rows"]) == len(rows) - 1
	assert result["metrics"]["source_rows"] == len(rows)
	assert result["metrics"]["excluded_source_rows"] == 1
	assert result["processed_rows"][0]["included"] is True
	assert result["processed_rows"][0]["red_apples"] == 2
	assert result["processed_rows"][0]["amount"] == 10
	assert result["processed_rows"][0]["proposed_value"]["created_at"] == "2026-06-09 09:00"
	assert result["processed_rows"][0]["proposed_value"]["punch_type"] == "忘刷卡补卡"
	assert result["processed_rows"][0]["review_status"] == "无需审核"
	assert "OUTSIDE_ATTENDANCE_MONTH" in result["processed_rows"][1]["exception_codes"]
	assert "APPROVAL_NOT_APPROVED" in result["processed_rows"][2]["exception_codes"]
	assert "APPROVAL_NOT_ENDED" in result["processed_rows"][2]["exception_codes"]
	assert "DUPLICATE_APPROVAL_NO" in result["processed_rows"][3]["exception_codes"]
	assert all(
		field in result["processed_rows"][1]
		for field in (
			"exception_codes",
			"exception_message",
			"review_status",
			"proposed_value",
			"confirmed_value",
			"reviewer",
			"reviewed_on",
			"review_note",
		)
	)
	assert summarize_missed_punch_rows(result["processed_rows"]) == [
		{
			"employee_code": "E-001",
			"employee_name": "张三",
			"department": "工程课",
			"missed_punch_count": 1,
			"red_apples": 2,
			"amount": 10,
			"source_rows": [2],
			"approval_nos": ["APP-001"],
		}
	]
	assert {
		item["source_row"] for item in result["processed_rows"] if item["review_status"] == "待审核"
	} == {4, 5, 6}


def test_identity_department_departure_and_same_time_conflicts_are_not_silent():
	rows = [
		{**_row("APP-101", name="错误姓名", department="旧部门"), "工号": "E-101"},
		_row("APP-102", name="同名员工", punch_time="2026-06-11 08:00"),
		_row("APP-103", name="离职员工", punch_time="2026-06-12 08:00"),
		_row("APP-104", name="张三", punch_time="2026-06-13 08:00"),
		_row("APP-105", name="张三", punch_time="2026-06-13 08:00"),
	]
	employees = [
		_employee("E-101", "正确姓名", "新部门"),
		_employee("E-201", "同名员工", "工程课"),
		_employee("E-202", "同名员工", "工程课"),
		_employee("E-301", "离职员工", "工程课", "已离职"),
		_employee("E-001", "张三", "工程课"),
	]
	result = process_missed_punch_rows(
		rows,
		attendance_month="2026-06",
		source_file="sample.xlsx",
		source_sheet="钉钉导出数据",
		employee_directory=employees,
	)

	assert {"EMPLOYEE_NAME_CONFLICT", "DEPARTMENT_CONFLICT"}.issubset(result["processed_rows"][0]["exception_codes"])
	assert "EMPLOYEE_AMBIGUOUS" in result["processed_rows"][1]["exception_codes"]
	assert "FORMER_EMPLOYEE_REQUIRES_CONFIRMATION" in result["processed_rows"][2]["exception_codes"]
	assert "MULTIPLE_APPROVALS_SAME_PUNCH_TIME" in result["processed_rows"][3]["exception_codes"]
	assert "MULTIPLE_APPROVALS_SAME_PUNCH_TIME" in result["processed_rows"][4]["exception_codes"]
	assert all(record["included"] is False for record in result["processed_rows"])
	assert all(record["review_status"] == "待审核" for record in result["processed_rows"])


def test_name_and_department_can_disambiguate_but_missing_source_still_requires_review():
	resolved = process_missed_punch_rows(
		[_row("APP-301", name="同名员工", department="财务课")],
		attendance_month="2026-06",
		source_file="sample.xlsx",
		source_sheet="钉钉导出数据",
		employee_directory=[
			_employee("E-301", "同名员工", "工程课"),
			_employee("E-302", "同名员工", "财务课"),
		],
	)
	assert resolved["processed_rows"][0]["employee_code"] == "E-302"
	assert resolved["processed_rows"][0]["review_status"] == "无需审核"

	missing_source = process_missed_punch_rows(
		[_row("APP-302")],
		attendance_month="2026-06",
		source_file="",
		source_sheet="",
		employee_directory=[_employee("E-001", "张三", "工程课")],
	)
	assert "SOURCE_FILE_MISSING" in missing_source["processed_rows"][0]["exception_codes"]
	assert "SOURCE_SHEET_MISSING" in missing_source["processed_rows"][0]["exception_codes"]


def test_name_only_source_rows_can_use_business_employee_code_aliases():
	"""DingTalk forgot-punch exports contain no 工号; roster aliases must still match."""
	result = process_missed_punch_rows(
		[_row("APP-303", name="张三", department="工程课")],
		attendance_month="2026-06",
		source_file="sample.xlsx",
		source_sheet="钉钉导出数据",
		employee_directory=[{
			"custom_employee_code": "E-303",
			"employee_name": "张三",
			"department": "工程课",
			"employment_status": "在职",
		}],
	)
	record = result["processed_rows"][0]
	assert record["employee_code"] == "E-303"
	assert record["review_status"] == "无需审核"


def test_confirmed_department_mapping_prevents_a_repeat_department_conflict():
	result = process_missed_punch_rows(
		[_row("APP-304", name="时杰", department="CCD+")],
		attendance_month="2026-06",
		source_file="sample.xlsx",
		source_sheet="钉钉导出数据",
		employee_directory=[_employee("E-304", "时杰", "品管课")],
		department_mapping={"CCD+": "品管课"},
	)
	record = result["processed_rows"][0]
	assert record["source_department"] == "CCD+"
	assert record["department"] == "品管课"
	assert "DEPARTMENT_CONFLICT" not in record["exception_codes"]
	assert record["review_status"] == "无需审核"


def test_offline_entries_are_explicit_and_rule_controlled():
	base = {
		**_row("", name="张三"),
		"source_kind": "offline",
		"数据id": "",
		"审批编号": "",
	}
	unconfirmed = process_missed_punch_rows(
		[base],
		attendance_month="2026-06",
		source_file="人工补录.xlsx",
		source_sheet="人工补录",
		employee_directory=[_employee("E-001", "张三", "工程课")],
	)
	record = unconfirmed["processed_rows"][0]
	assert record["source_kind"] == "offline"
	assert "OFFLINE_REASON_REQUIRED" in record["exception_codes"]
	assert "OFFLINE_CONFIRMER_REQUIRED" in record["exception_codes"]
	assert record["included"] is False

	confirmed_row = {
		**base,
		"manual_reason": "纸质签卡单补录",
		"confirmed_by": "HR-001",
		"manual_confirmed": True,
	}
	confirmed = process_missed_punch_rows(
		[confirmed_row],
		attendance_month="2026-06",
		source_file="人工补录.xlsx",
		source_sheet="人工补录",
		employee_directory=[_employee("E-001", "张三", "工程课")],
	)
	confirmed_record = confirmed["processed_rows"][0]
	assert confirmed_record["included"] is False
	assert confirmed_record["review_status"] == "待审核"
	assert confirmed_record["exception_codes"] == ["OFFLINE_ENTRY_REQUIRES_CONFIRMATION"]
	assert confirmed_record["approval_no"] == ""
	assert confirmed_record["source_id"] == ""


def test_former_employee_policy_can_be_explicitly_overridden():
	result = process_missed_punch_rows(
		[_row("APP-501", name="离职员工")],
		attendance_month="2026-06",
		source_file="sample.xlsx",
		source_sheet="钉钉导出数据",
		employee_directory=[_employee("E-501", "离职员工", "工程课", "已离职")],
		rules=MissedPunchRules(former_employee_policy="include"),
	)
	assert result["processed_rows"][0]["included"] is True
	assert result["processed_rows"][0]["exception_codes"] == []


def test_review_application_is_audited_and_does_not_mutate_the_proposal():
	result = process_missed_punch_rows(
		[_row("APP-601", name="待匹配员工")],
		attendance_month="2026-06",
		source_file="sample.xlsx",
		source_sheet="钉钉导出数据",
		employee_directory=[],
	)
	original = result["processed_rows"][0]
	confirmed_value = {
		**original["proposed_value"],
		"employee_code": "E-601",
		"employee_name": "待匹配员工",
		"department": "工程课",
		"included": True,
		"red_apples": 2,
		"amount": 10,
	}
	reviewed = apply_missed_punch_review(
		original,
		decision="已通过",
		confirmed_value=confirmed_value,
		reviewer="HR-001",
		reviewed_on="2026-07-02 09:30:00",
		review_note="已核对员工花名册",
	)

	assert original["review_status"] == "待审核"
	assert original["confirmed_value"] is None
	assert reviewed["review_status"] == "已通过"
	assert reviewed["included"] is True
	assert reviewed["employee_code"] == "E-601"
	assert reviewed["review_history"][-1]["old_value"] is None
	assert reviewed["review_history"][-1]["new_value"] == confirmed_value
	assert reviewed["review_history"][-1]["reason"] == "已核对员工花名册"


def test_known_sample_contract_when_the_user_workbook_is_available():
	if not SAMPLE_WORKBOOK.exists():
		return
	try:
		from openpyxl import load_workbook
	except ModuleNotFoundError:
		return

	workbook = load_workbook(SAMPLE_WORKBOOK, data_only=True, read_only=True)
	sheet = workbook["钉钉导出数据"]
	headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
	rows = [dict(zip(headers, (cell.value for cell in row), strict=True)) for row in sheet.iter_rows(min_row=2)]
	people = {}
	for row in rows:
		name = row.get("创建人")
		people.setdefault(name, _employee(f"SAMPLE-{len(people) + 1:03d}", name, row.get("创建人部门")))

	result = process_missed_punch_rows(
		rows,
		attendance_month="2026-06",
		source_file=SAMPLE_WORKBOOK.name,
		source_sheet="钉钉导出数据",
		employee_directory=list(people.values()),
	)

	assert result["metrics"]["source_rows"] == 52
	assert result["metrics"]["excluded_source_rows"] == 8
	assert len(result["processed_rows"]) == 44
	assert sum("OUTSIDE_ATTENDANCE_MONTH" in row["exception_codes"] for row in result["processed_rows"]) == 6
	assert sum(row["included"] for row in result["processed_rows"]) == 37


if __name__ == "__main__":
	test_structure_precheck_recognises_the_dingtalk_contract()
	test_default_rules_return_one_processed_dataset_and_keep_reviewable_rows()
	test_identity_department_departure_and_same_time_conflicts_are_not_silent()
	test_name_and_department_can_disambiguate_but_missing_source_still_requires_review()
	test_name_only_source_rows_can_use_business_employee_code_aliases()
	test_confirmed_department_mapping_prevents_a_repeat_department_conflict()
	test_offline_entries_are_explicit_and_rule_controlled()
	test_former_employee_policy_can_be_explicitly_overridden()
	test_review_application_is_audited_and_does_not_mutate_the_proposal()
	test_known_sample_contract_when_the_user_workbook_is_available()
	print("Missed-punch processor contract passed.")
