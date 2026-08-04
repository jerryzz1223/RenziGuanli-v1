#!/usr/bin/env python3
"""Phase 1 contract for previewing the real DingTalk attendance export.

The test intentionally loads the user-provided workbook read-only. It stubs
only Frappe's import-time helpers, so exercising the parser cannot write data.
"""

from __future__ import annotations

import importlib.util
import sys
import types
from datetime import datetime
from pathlib import Path

try:
	from openpyxl import load_workbook
except ModuleNotFoundError:
	load_workbook = None


PROJECT_ROOT = Path(__file__).resolve().parents[1]
API_PATH = PROJECT_ROOT / "hrms" / "api" / "attendance_import.py"
REAL_EXPORT = Path("/Users/lrj/Desktop/考勤表.xlsx")
COMPANY_WORKBOOK = Path("/Users/lrj/Documents/SAD/YOngxin/人资/副本人资系统沟通表260713.xlsx")
COMPANY_ATTENDANCE_REGISTER = Path("/Users/lrj/Documents/SAD/YOngxin/人资/各种表单/考勤.xlsx")
DAILY_ATTENDANCE_REGISTER = Path("/Users/lrj/Documents/SAD/YOngxin/日考勤.xlsx")


def _skip_reason() -> str | None:
	if load_workbook is None:
		return "Skipping DingTalk export preview contract: install openpyxl to load workbook fixtures."
	return None


if "pytest" in sys.modules and (reason := _skip_reason()):
	import pytest

	pytest.skip(reason, allow_module_level=True)


def load_attendance_module():
	frappe = types.ModuleType("frappe")
	frappe._ = lambda value: value
	def whitelist(function=None):
		return function if function is not None else lambda decorated: decorated

	frappe.whitelist = whitelist
	frappe.throw = lambda message: (_ for _ in ()).throw(RuntimeError(message))
	frappe.db = types.SimpleNamespace()
	frappe.get_doc = lambda *_args, **_kwargs: None

	utils = types.ModuleType("frappe.utils")
	utils.flt = lambda value: float(value or 0)
	utils.getdate = lambda value: datetime.fromisoformat(str(value)).date()
	utils.now_datetime = datetime.now

	sys.modules["frappe"] = frappe
	sys.modules["frappe.utils"] = utils
	spec = importlib.util.spec_from_file_location("attendance_import_preview_contract", API_PATH)
	module = importlib.util.module_from_spec(spec)
	sys.modules[spec.name] = module
	spec.loader.exec_module(module)
	return module


def test_preview_real_dingtalk_export():
	if not REAL_EXPORT.exists():
		return
	module = load_attendance_module()
	workbook = load_workbook(REAL_EXPORT, data_only=True, read_only=True)
	preview = module._preview_dingtalk_export_v1(workbook)

	assert preview["source_type"] == "dingtalk_export_v1"
	assert preview["missing_sheets"] == []
	assert preview["record_counts"] == {
		"daily_statistics": 3029,
		"raw_records": 4058,
		"monthly_people": 233,
	}
	assert "请假/事假(小时)" in preview["field_mapping"]
	assert preview["database_writes"] == 0
	assert preview["import_validation"]["status"] == "可导入"
	assert preview["import_validation"]["missing_required_fields"] == []
	assert preview["import_validation"]["matched_field_count"] >= 20
	warnings = {warning["code"]: warning["count"] for warning in preview["quality_warnings"]}
	assert warnings["missing_employee_code"] == 299
	assert warnings["missing_attendance_group"] == 299
	assert warnings["planned_hours_without_actual"] == 544
	assert warnings["duplicate_userid_workdate"] == 0


def test_preview_company_workbook_preserves_raw_and_manual_daily_sources():
	if not COMPANY_WORKBOOK.exists():
		return
	module = load_attendance_module()
	workbook = load_workbook(COMPANY_WORKBOOK, data_only=True, read_only=True)
	preview = module._preview_company_attendance_workbook(workbook)

	assert preview["source_type"] == "company_attendance_workbook_v1"
	assert preview["database_writes"] == 0
	assert preview["daily_sources"]["dingtalk_raw"]["sheet_name"] == "每日统计（钉钉导出）"
	assert preview["daily_sources"]["dingtalk_raw"]["header_rows"] == [3, 4]
	assert preview["daily_sources"]["dingtalk_raw"]["data_start_row"] == 5
	assert preview["daily_sources"]["manual_adjustment"]["sheet_name"] == "每日统计（修改后）"
	assert preview["daily_sources"]["manual_adjustment"]["header_rows"] == [1, 2]
	assert preview["daily_sources"]["manual_adjustment"]["data_start_row"] == 3
	assert "UserId" in preview["daily_sources"]["dingtalk_raw"]["headers"]
	assert "UserId" not in preview["daily_sources"]["manual_adjustment"]["headers"]
	assert "请假/事假(小时)" in preview["daily_sources"]["dingtalk_raw"]["field_mapping"]
	assert "日期类型" in preview["daily_sources"]["dingtalk_raw"]["field_mapping"]
	assert "上班缺卡" in preview["daily_sources"]["dingtalk_raw"]["field_mapping"]
	for field in ("请假/旷工(小时)", "请假/公假(天)", "请假/产假(天)", "请假/团圆假(天)"):
		assert field in preview["daily_sources"]["dingtalk_raw"]["field_mapping"]


def test_company_attendance_register_has_a_dedicated_preview_contract():
	if not COMPANY_ATTENDANCE_REGISTER.exists():
		return
	module = load_attendance_module()
	workbook = load_workbook(COMPANY_ATTENDANCE_REGISTER, data_only=True, read_only=True)
	preview = module._preview_company_attendance_register_v1(workbook)

	assert module._is_company_attendance_register_v1(workbook)
	assert preview["source_type"] == "company_attendance_register_v1"
	assert preview["missing_sheets"] == []
	assert preview["sheets"][0]["sheet_name"] == "每日统计"
	assert preview["sheets"][0]["row_count"] == 198
	assert "请假/事假(小时)" in preview["field_mapping"]
	assert {sheet["sheet_name"] for sheet in preview["sheets"]} == {"每日统计", "出勤明细", "出勤异常", "苹果树"}
	assert preview["import_validation"]["status"] == "可导入"
	assert preview["sheets"][0]["import_behavior"] == "写入每日考勤核对"
	assert next(sheet for sheet in preview["sheets"] if sheet["sheet_name"] == "出勤异常")["import_behavior"] == "保留为异常核对来源，不自动生成处理结论"


def test_daily_attendance_register_uses_the_company_39_column_layout():
	if not DAILY_ATTENDANCE_REGISTER.exists():
		return
	module = load_attendance_module()
	workbook = load_workbook(DAILY_ATTENDANCE_REGISTER, data_only=True, read_only=True)
	preview = module._preview_company_attendance_register_v1(workbook)
	daily_sheet = next(sheet for sheet in preview["sheets"] if sheet["sheet_name"] == "每日统计")

	assert preview["source_type"] == "company_attendance_register_v1"
	assert daily_sheet["row_count"] == 198
	assert len(daily_sheet["headers"]) == 39
	assert "请假/婚假" in preview["field_mapping"]
	assert "请假/团圆假" in preview["field_mapping"]
	assert "旷工" in preview["field_mapping"]


def test_daily_statistics_template_matches_the_company_39_column_layout():
	from openpyxl import Workbook

	module = load_attendance_module()
	workbook = Workbook()
	workbook.remove(workbook.active)
	sheet = module._add_daily_statistics_template_sheet(workbook)

	assert sheet.max_column == 39
	assert sheet["L1"].value == "实际出勤（小时）"
	assert sheet["S1"].value == "请假"
	assert sheet["S2"].value == "事假(小时)"
	assert sheet["AD1"].value == "请假"
	assert sheet["AD2"].value == "婚假"
	assert sheet["AH2"].value == "团圆假"
	assert sheet["AI1"].value == "旷工"
	assert "S1:AC1" in {str(cell_range) for cell_range in sheet.merged_cells.ranges}
	assert "AD1:AH1" in {str(cell_range) for cell_range in sheet.merged_cells.ranges}
	assert sheet.freeze_panes == "A3"

	row = [""] * 39
	row[0:3] = ["模板测试员工", "E-001", "2026-07-01"]
	row[11] = "8"
	row[13] = "3"
	row[34] = "2"
	sheet.append(row)
	parsed = module._daily_rows_from_header_rows(sheet, (1, 2), 3)
	assert len(parsed) == 1
	assert parsed[0]["姓名"] == "模板测试员工"
	assert parsed[0]["工号"] == "E-001"
	assert parsed[0]["日期"] == "2026-07-01"
	assert parsed[0]["实际出勤(小时)"] == "8"
	assert parsed[0]["工作日加班(小时)"] == "3"
	assert parsed[0]["旷工_2"] == "2"
	assert parsed[0]["_source_row"] == 3


def test_export_profiles_cover_the_company_detail_and_monthly_forms():
	from openpyxl import Workbook

	module = load_attendance_module()
	assert set(module.ATTENDANCE_EXPORT_PROFILES) == {
		"company_attendance_workbook",
		"daily_statistics",
		"attendance_detail",
		"leave_evidence",
		"attendance_exception",
		"missing_card",
		"apple_reward",
		"monthly_draft",
		"monthly_signed",
		"monthly_finance",
	}

	workbook = Workbook()
	workbook.remove(workbook.active)
	rows = [
		types.SimpleNamespace(
			department="工程课", employee_name="模板测试员工", employee_code="E-001", date_of_joining="2026-01-01",
			standard_hours=8, actual_attendance_hours=8, overtime_1_5_hours=3, overtime_2_hours=0, overtime_3_hours=0,
			large_night_shift_count=0, small_night_shift_count=0, personal_leave_hours=0, sick_leave_hours=0,
			annual_leave_hours=0, work_injury_leave_hours=0, rest_leave_hours=0, absent_hours=0,
			actual_clock_attendance_hours=8, paid_leave_makeup_hours=0, leave_deductible_hours=0, workday_rest_leave_hours=0,
			adjusted_absence_hours=0, adjusted_working_hours=8, overtime_1_5_settlement_hours=3,
			overtime_2_settlement_hours=0, overtime_3_settlement_hours=0, green_apples=2, red_apples=0,
			apple_reward_amount=20, red_apple_penalty=0, night_shift_allowance=0, full_attendance_deduction=0,
		)
	]
	module._add_monthly_attendance_export_sheet(workbook, "monthly_draft", "2026-07", rows)
	module._add_monthly_attendance_export_sheet(workbook, "monthly_signed", "2026-07", rows)
	module._add_monthly_attendance_export_sheet(workbook, "monthly_finance", "2026-07", rows)

	assert workbook.sheetnames == ["考勤初稿", "考勤终稿（签字版）", "考勤终稿（财务版）"]
	assert workbook["考勤初稿"]["A1"].value == "2026年07月工时奖惩确认表"
	assert workbook["考勤终稿（签字版）"].max_column > workbook["考勤初稿"].max_column
	assert workbook["考勤终稿（财务版）"].max_column > workbook["考勤初稿"].max_column
	assert workbook["考勤终稿（财务版）"]["C3"].value == "模板测试员工"


def test_revoked_batch_does_not_block_reimport_of_the_same_source_file():
	module = load_attendance_module()
	captured = {}
	active_batch = types.SimpleNamespace(name="ACTIVE-BATCH", status="已导入", daily_sheet_rows=12)
	module._source_file_checksum = lambda _file_url: "same-source-checksum"

	def get_value(doctype, filters, fieldname):
		captured.update({"doctype": doctype, "filters": filters, "fieldname": fieldname})
		return active_batch.name

	module.frappe.db = types.SimpleNamespace(get_value=get_value)
	module.frappe.get_doc = lambda doctype, name: active_batch
	batch, duplicate = module._create_attendance_batch("/private/files/attendance.xlsx", "2026-07", "永新", "company_attendance_register_v1")

	assert duplicate is True
	assert batch is active_batch
	assert captured["filters"]["status"] == ["!=", "已撤销"]


def test_field_mapping_catalog_explains_the_import_contract_without_writing_data():
	module = load_attendance_module()
	catalog = module.get_attendance_field_mapping_catalog()

	profile = next(item for item in catalog["profiles"] if item["source_type"] == "dingtalk_export_v1")
	assert "employee_name" in profile["required_target_fields"]
	assert "attendance_date" in profile["required_target_fields"]
	assert catalog["write_policy"] == "只读说明，不写入考勤数据"


def test_rule_evaluation_is_explicit_and_never_mutates_imported_attendance():
	module = load_attendance_module()
	assert module.ATTENDANCE_RULE_APPLICATION_MODES == ("仅展示", "导入校验", "异常提示")
	assert module._rule_execution_notice() == "规则不会自动修改导入数据、月度终稿或薪资。"


def test_attendance_template_catalog_explains_uploadable_company_workbook():
	module = load_attendance_module()
	templates = module.list_attendance_import_templates()
	keys = {template["key"] for template in templates}

	assert "company_attendance_register_v1" in keys
	assert "company_daily_statistics_v1" in keys
	assert "attendance_exception_v1" in keys
	assert "apple_reward_v1" in keys
	company_workbook = next(template for template in templates if template["key"] == "company_attendance_register_v1")
	assert company_workbook["upload_mode"] == "whole_workbook"
	assert "每日统计" in company_workbook["sheet_names"]


def test_monthly_scope_uses_company_month_and_lock_version():
	module = load_attendance_module()
	assert module._attendance_scope_filters("Company A", "2026-07", "2") == {
		"company": "Company A",
		"attendance_month": "2026-07",
		"attendance_lock_version": "2",
	}


def test_real_dingtalk_export_has_a_daily_statistics_import_path():
	module = load_attendance_module()
	assert module._dingtalk_export_import_source_kind() == "钉钉原始导出"


def test_manual_adjustment_takes_precedence_without_deleting_raw_source():
	module = load_attendance_module()
	raw = types.SimpleNamespace(employee_code="E-001", employee_name="张三", attendance_date="2026-07-01", source_kind="钉钉原始导出")
	manual = types.SimpleNamespace(employee_code="E-001", employee_name="张三", attendance_date="2026-07-01", source_kind="人工调整")
	other = types.SimpleNamespace(employee_code="E-002", employee_name="李四", attendance_date="2026-07-01", source_kind="钉钉原始导出")

	effective = module._prefer_manual_daily_rows([raw, manual, other])
	assert manual in effective
	assert raw not in effective
	assert other in effective


def test_newer_correction_version_takes_precedence_over_older_manual_adjustment():
	module = load_attendance_module()
	manual_v1 = types.SimpleNamespace(employee_code="E-001", employee_name="张三", attendance_date="2026-07-01", source_kind="人工调整", correction_version=1)
	correction_v2 = types.SimpleNamespace(employee_code="E-001", employee_name="张三", attendance_date="2026-07-01", source_kind="钉钉原始导出", correction_version=2)

	effective = module._prefer_manual_daily_rows([manual_v1, correction_v2])
	assert effective == [correction_v2]


if __name__ == "__main__":
	if reason := _skip_reason():
		print(reason)
		raise SystemExit(0)
	test_preview_real_dingtalk_export()
	test_preview_company_workbook_preserves_raw_and_manual_daily_sources()
	test_company_attendance_register_has_a_dedicated_preview_contract()
	test_daily_attendance_register_uses_the_company_39_column_layout()
	test_daily_statistics_template_matches_the_company_39_column_layout()
	test_revoked_batch_does_not_block_reimport_of_the_same_source_file()
	test_field_mapping_catalog_explains_the_import_contract_without_writing_data()
	test_rule_evaluation_is_explicit_and_never_mutates_imported_attendance()
	test_attendance_template_catalog_explains_uploadable_company_workbook()
	test_monthly_scope_uses_company_month_and_lock_version()
	test_real_dingtalk_export_has_a_daily_statistics_import_path()
	test_manual_adjustment_takes_precedence_without_deleting_raw_source()
	test_newer_correction_version_takes_precedence_over_older_manual_adjustment()
	print("DingTalk export preview contract passed.")
